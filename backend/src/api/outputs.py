import csv
import json
import re
from datetime import date as dt_date
from html import escape
from pathlib import Path

import pandas as pd
from fastapi import APIRouter, HTTPException, Depends
from fastapi.responses import FileResponse, HTMLResponse
from loguru import logger
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from src.services.output_builder import (
    build_output_preview,
    build_delivery_preview,
    build_daily_output_bundle,
    build_weekly_weight_summary_workbook,
)
from src.db import session_scope
from src.models.order import Order
from src.services import order_form_service, facility_service, order_service
from src.api.auth import require_role

router = APIRouter()

_PREVIEW_LIMIT_DEFAULT = 10
_FINAL_DIET_LABELS = {
    "regular": "常食",
    "regular_bag": "常食袋分け",
    "soft": "軟菜",
    "mixer": "ミキサー",
    "daycare": "通所",
    "staff": "職員",
    "diabetes": "糖尿",
    "no_meat": "禁食肉禁",
    "no_fish": "禁食魚禁",
    "no_fried": "揚げ物禁",
    "forbidden_other": "禁食その他",
    "forbidden": "禁食",
    "禁食": "禁食",
}


def _output_file_for_type(order_id: str, output_type: str) -> tuple[Path, str]:
    if output_type == "labels":
        outputs = build_output_preview(order_id, "labels")
        return Path(outputs["labels"]), "ラベルCSV"
    if output_type == "delivery":
        outputs = build_output_preview(order_id, "delivery")
        return Path(outputs["delivery_note"]), "納品書Excel"
    if output_type == "aggregate":
        outputs = build_output_preview(order_id, "aggregate")
        return Path(outputs["aggregate"]), "総量CSV"
    if output_type == "order_form_saved_sheet":
        return order_form_service.build_saved_sheet_order_form_excel(order_id=order_id), "FAX読取シートExcel"
    raise ValueError("invalid output type")


def _cell_color(fill: PatternFill) -> str:
    color = getattr(fill, "fgColor", None)
    rgb = str(getattr(color, "rgb", "") or "")
    if len(rgb) == 8:
        return f"#{rgb[-6:]}"
    return ""


def _cell_css(cell) -> str:
    styles: list[str] = []
    if cell.font and cell.font.bold:
        styles.append("font-weight:700")
    if cell.alignment:
        if cell.alignment.horizontal:
            styles.append(f"text-align:{cell.alignment.horizontal}")
        if cell.alignment.vertical:
            styles.append(f"vertical-align:{cell.alignment.vertical}")
        if cell.alignment.wrap_text:
            styles.append("white-space:pre-wrap")
    color = _cell_color(cell.fill)
    if color and color != "#000000":
        styles.append(f"background:{color}")
    border_styles = []
    for side_name in ("top", "right", "bottom", "left"):
        side = getattr(cell.border, side_name)
        if getattr(side, "style", None):
            border_styles.append(side_name)
    if border_styles:
        styles.append("border:1px solid #1f2933")
    return ";".join(styles)


def _render_xlsx_preview(path: Path, title: str) -> str:
    workbook = load_workbook(path, data_only=True)
    sheet_html: list[str] = []
    for worksheet in workbook.worksheets:
        if worksheet.sheet_state != "visible":
            continue
        merged_lookup: dict[tuple[int, int], tuple[int, int]] = {}
        covered: set[tuple[int, int]] = set()
        for merged in worksheet.merged_cells.ranges:
            rowspan = merged.max_row - merged.min_row + 1
            colspan = merged.max_col - merged.min_col + 1
            merged_lookup[(merged.min_row, merged.min_col)] = (rowspan, colspan)
            for row_idx in range(merged.min_row, merged.max_row + 1):
                for col_idx in range(merged.min_col, merged.max_col + 1):
                    if (row_idx, col_idx) != (merged.min_row, merged.min_col):
                        covered.add((row_idx, col_idx))
        rows_html: list[str] = []
        for row_idx in range(1, min(worksheet.max_row, 90) + 1):
            cells_html: list[str] = []
            for col_idx in range(1, min(worksheet.max_column, 20) + 1):
                if (row_idx, col_idx) in covered:
                    continue
                cell = worksheet.cell(row=row_idx, column=col_idx)
                attrs = []
                rowspan, colspan = merged_lookup.get((row_idx, col_idx), (1, 1))
                if rowspan > 1:
                    attrs.append(f'rowspan="{rowspan}"')
                if colspan > 1:
                    attrs.append(f'colspan="{colspan}"')
                style = _cell_css(cell)
                if style:
                    attrs.append(f'style="{escape(style)}"')
                value = cell.value
                text = value.strftime("%Y-%m-%d") if hasattr(value, "strftime") else str(value or "")
                cells_html.append(f"<td {' '.join(attrs)}>{escape(text)}</td>")
            rows_html.append(f"<tr>{''.join(cells_html)}</tr>")
        sheet_html.append(
            f"<section><h2>{escape(worksheet.title)}</h2><div class=\"sheet-wrap\"><table>{''.join(rows_html)}</table></div></section>"
        )
    return _wrap_preview_html(title, "".join(sheet_html))


def _render_csv_preview(path: Path, title: str, encoding: str) -> str:
    with open(path, newline="", encoding=encoding, errors="replace") as f:
        reader = csv.reader(f)
        rows = list(reader)
    body = "".join(
        "<tr>" + "".join(f"<td>{escape(str(cell))}</td>" for cell in row) + "</tr>"
        for row in rows[:500]
    )
    return _wrap_preview_html(title, f"<section><h2>{escape(title)}</h2><div class=\"sheet-wrap\"><table>{body}</table></div></section>")


def _wrap_preview_html(title: str, body: str) -> str:
    return f"""<!doctype html>
<html lang="ja">
<head>
  <meta charset="utf-8" />
  <title>{escape(title)} プレビュー</title>
  <style>
    body {{ background:#f3f1ea; color:#1f2a2a; font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",sans-serif; margin:0; padding:20px; }}
    h1 {{ font-size:18px; margin:0 0 16px; }}
    h2 {{ font-size:14px; margin:18px 0 8px; }}
    .sheet-wrap {{ background:white; border:1px solid #d8d2c4; overflow:auto; max-height:80vh; }}
    table {{ border-collapse:collapse; background:white; }}
    td {{ border:1px solid #d7d7d7; font-size:12px; min-width:56px; padding:4px 6px; white-space:pre; }}
  </style>
</head>
<body>
  <h1>{escape(title)} プレビュー</h1>
  {body}
</body>
</html>"""


def _normalize_delivery_diet_for_display(value: object) -> str:
    text = str(value or "").strip()
    compact = re.sub(r"[\s　]+", "", text)
    lowered = text.lower()
    if lowered in {"change_1", "change_2", "change1", "change2"} or compact in {"変更1", "変更2", "変更１", "変更２"}:
        return "regular"
    if lowered in {"regular", "standard"} or "常食" in text:
        return "regular"
    if lowered in {"regular_bag"} or "袋" in text:
        return "regular_bag"
    if lowered == "soft" or "軟菜" in text:
        return "soft"
    if lowered == "mixer" or "ミキサ" in text or "ﾐｷｻ" in text:
        return "mixer"
    if lowered == "daycare" or "通所" in text:
        return "daycare"
    if lowered == "staff" or "職員" in text:
        return "staff"
    if lowered in {"diabetes", "diabetic"} or "糖尿" in text:
        return "diabetes"
    if lowered == "no_meat" or "肉禁" in text:
        return "no_meat"
    if lowered == "no_fish" or "魚禁" in text:
        return "no_fish"
    if lowered == "no_fried" or "揚げ物" in text or "揚物" in text:
        return "no_fried"
    if lowered == "forbidden_other" or "その他" in text:
        return "forbidden_other"
    if lowered in {"forbidden", "禁食"} or "禁食" in text:
        return "forbidden"
    return lowered or text


def _normalize_delivery_area_for_display(value: object) -> str:
    text = str(value or "").strip()
    if not text or text.upper() == "X":
        return ""
    text = text.replace("階", "F").replace("ｆ", "F").replace("Ｆ", "F")
    if not re.search(r"\dF|月|花", text, re.IGNORECASE):
        return ""
    if "花" in text:
        return "2F"
    if "月" in text:
        return "3F"
    return text.upper()


def _delivery_final_header_base(column: dict) -> str:
    diet = _normalize_delivery_diet_for_display(column.get("diet_type"))
    if diet in _FINAL_DIET_LABELS:
        return _FINAL_DIET_LABELS[diet]
    raw = str(column.get("header") or column.get("name") or "").strip()
    raw = re.sub(r"\d+\s*回目", "", raw).strip()
    if raw in {"1回目", "2回目", "3回目", "変更1", "変更2", "変更１", "変更２"}:
        return "常食"
    return raw or str(column.get("name") or "").strip()


def _build_delivery_render_columns(columns: list | None) -> list[dict]:
    render_columns: list[dict] = [
        {"kind": "field", "source": "date", "header": "日付"},
        {"kind": "field", "source": "daypart", "header": "区分"},
        {"kind": "field", "source": "menu_category", "header": "献立区分"},
        {"kind": "field", "source": "menu_name", "header": "メニュー名"},
    ]
    quantity_groups: dict[tuple[str, str], dict] = {}
    for column in columns or []:
        if not isinstance(column, dict) or column.get("source") != "quantity" or not column.get("name"):
            continue
        diet = _normalize_delivery_diet_for_display(column.get("diet_type") or column.get("name"))
        if diet in {"change_1", "change_2", "change1", "change2", "変更1", "変更2", "変更１", "変更２"}:
            diet = "regular"
        area = _normalize_delivery_area_for_display(column.get("area_id") or column.get("name"))
        base = _delivery_final_header_base(column)
        key = (diet or base, area)
        group = quantity_groups.setdefault(
            key,
            {
                "kind": "quantity",
                "header_base": base,
                "area": area,
                "source_names": [],
            },
        )
        group["source_names"].append(str(column.get("name")))
    base_counts: dict[str, int] = {}
    for group in quantity_groups.values():
        base_counts[str(group.get("header_base") or "")] = base_counts.get(str(group.get("header_base") or ""), 0) + 1
    for group in quantity_groups.values():
        base = str(group.get("header_base") or "")
        area = str(group.get("area") or "")
        group["header"] = f"{base}\n{area}" if area and base_counts.get(base, 0) > 1 else base
        render_columns.append(group)
    render_columns.append({"kind": "field", "source": "note", "header": "備考欄"})
    return render_columns


def _format_delivery_html_cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return str(value)


def _delivery_render_cell_value(row: dict, column: dict) -> object:
    if column.get("kind") == "quantity":
        total = 0.0
        has_value = False
        for name in column.get("source_names") or []:
            value = row.get(str(name))
            if value is None or value == "":
                continue
            try:
                total += float(value)
                has_value = True
            except (TypeError, ValueError):
                return value
        if not has_value:
            return ""
        return int(total) if total.is_integer() else total
    source = column.get("source")
    if source == "menu_name":
        return row.get("menu_name")
    if source == "daypart":
        value = str(row.get("daypart") or "").strip()
        return value[:1] if value in {"朝食", "昼食", "夕食"} else value
    return row.get(str(source))


def _render_editable_delivery_note_html(
    order_id: str,
    title: str,
    headers: list,
    rows: list,
    facility_name: str | None = None,
    columns: list | None = None,
    raw_rows: list | None = None,
) -> str:
    render_columns = _build_delivery_render_columns(columns)
    render_rows = [row for row in (raw_rows or []) if isinstance(row, dict)]
    if not render_rows:
        render_columns = [{"kind": "field", "source": f"col-{idx}", "header": str(header or "")} for idx, header in enumerate(headers)]
        render_rows = [
            {f"col-{idx}": row[idx] if isinstance(row, list) and idx < len(row) else "" for idx in range(len(render_columns))}
            for row in rows
        ]
    header_html = "".join(
        f'<th class="col-{idx}" contenteditable="true" data-edit="h-{idx}">{escape(str(column.get("header") or ""))}</th>'
        for idx, column in enumerate(render_columns)
    )
    row_html: list[str] = []
    for row_idx, row in enumerate(render_rows):
        daypart = str(row.get("daypart") or "")
        row_class = "daypart-start" if row_idx == 0 or daypart != str(render_rows[row_idx - 1].get("daypart") or "") else ""
        cell_html = []
        for col_idx, column in enumerate(render_columns):
            value = _delivery_render_cell_value(row, column)
            class_name = "menu-cell" if column.get("source") == "menu_name" else ""
            cell_html.append(
                f'<td class="col-{col_idx} {class_name}" contenteditable="true" data-edit="r-{row_idx}-c-{col_idx}">'
                f"{escape(_format_delivery_html_cell(value))}</td>"
            )
        row_html.append(f'<tr class="{row_class}">{"".join(cell_html)}</tr>')
    body_html = "".join(row_html)
    safe_title = escape(title)
    safe_order_id = escape(order_id)
    safe_facility_name = escape(facility_name or "")
    return f"""<!doctype html>
<html lang="ja">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>{safe_title}</title>
  <style>
    @page {{ size: A4 landscape; margin: 8mm; }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      background: #e9ecef;
      color: #111827;
      font-family: -apple-system, BlinkMacSystemFont, "Hiragino Kaku Gothic ProN", "Yu Gothic", Meiryo, sans-serif;
      font-size: 12px;
    }}
    .toolbar {{
      position: sticky;
      top: 0;
      z-index: 10;
      display: flex;
      align-items: center;
      gap: 8px;
      padding: 10px 14px;
      background: #ffffff;
      border-bottom: 1px solid #cbd5e1;
    }}
    .toolbar-title {{ font-weight: 700; margin-right: auto; }}
    button {{
      border: 1px solid #64748b;
      background: #ffffff;
      color: #0f172a;
      border-radius: 6px;
      padding: 7px 12px;
      font: inherit;
      cursor: pointer;
    }}
    button.primary {{ background: #0f172a; color: #ffffff; border-color: #0f172a; }}
    .sheet {{
      width: 297mm;
      min-height: 210mm;
      margin: 12px auto;
      padding: 9mm;
      background: #ffffff;
      box-shadow: 0 2px 10px rgba(15, 23, 42, 0.18);
    }}
    .company {{
      display: grid;
      grid-template-columns: 1fr 1.4fr;
      gap: 8mm;
      margin-bottom: 8mm;
    }}
    .doc-title {{ font-size: 24px; font-weight: 800; margin: 0 0 5mm; }}
    .facility-label {{ font-weight: 700; border-bottom: 2px solid #111827; display: inline-block; margin-bottom: 3mm; }}
    .facility-box {{
      border: 2px solid #111827;
      min-height: 18mm;
      display: flex;
      align-items: center;
      justify-content: center;
      font-size: 24px;
      font-weight: 700;
      padding: 4mm;
    }}
    .company-name {{ font-size: 18px; font-weight: 800; margin-bottom: 5mm; }}
    .company-line {{ font-size: 10px; font-weight: 700; margin: 0 0 3mm; }}
    table {{
      width: 100%;
      border-collapse: collapse;
      table-layout: auto;
      border: 2px solid #111827;
      background: #ffffff;
    }}
    th, td {{
      border: 1px solid #111827;
      min-height: 7mm;
      height: 7mm;
      padding: 2px 4px;
      text-align: center;
      vertical-align: middle;
      white-space: pre-wrap;
      word-break: break-word;
    }}
    th {{ font-weight: 400; font-size: 14px; height: 12mm; min-height: 12mm; }}
    thead th {{ border-top-width: 2px; border-bottom-width: 2px; }}
    tr.daypart-start td {{ border-top-width: 2px; }}
    tbody tr:last-child td {{ border-bottom-width: 2px; }}
    th:first-child, td:first-child {{ border-left-width: 2px; }}
    th:last-child, td:last-child {{ border-right-width: 2px; }}
    td.menu-cell {{ text-align: left; font-weight: 700; }}
    .col-0 {{ width: 9%; }}
    .col-1 {{ width: 6%; }}
    .col-2 {{ width: 9%; }}
    .col-3 {{ width: 24%; }}
    th[class*="col-"]:not(.col-0):not(.col-1):not(.col-2):not(.col-3),
    td[class*="col-"]:not(.col-0):not(.col-1):not(.col-2):not(.col-3) {{ width: 7%; }}
    [contenteditable="true"] {{ cursor: text; }}
    [contenteditable="true"]:focus {{
      outline: 2px solid #2563eb;
      outline-offset: -2px;
      background: #eff6ff;
    }}
    @media print {{
      body {{ background: #ffffff; }}
      .toolbar {{ display: none; }}
      .sheet {{
        width: auto;
        min-height: auto;
        margin: 0;
        padding: 0;
        box-shadow: none;
      }}
    }}
  </style>
</head>
<body>
  <div class="toolbar">
    <div class="toolbar-title">{safe_title}</div>
    <button type="button" id="reset-edits">編集をリセット</button>
    <button type="button" class="primary" id="print-note">印刷/PDF保存</button>
  </div>
  <main class="sheet">
    <section class="company">
      <div>
        <h1 class="doc-title">【納品書】</h1>
        <div class="facility-label">施設名</div>
        <div class="facility-box" contenteditable="true" data-edit="facility-name">{safe_facility_name}</div>
      </div>
      <div>
        <div class="company-name" contenteditable="true" data-edit="company-name">株式会社アドオンミール</div>
        <p class="company-line" contenteditable="true" data-edit="company-1">福岡本社: 〒810-0001</p>
        <p class="company-line" contenteditable="true" data-edit="company-2">福岡県福岡市中央区天神２丁目３番１０号　天神パインクレスト719</p>
        <p class="company-line" contenteditable="true" data-edit="company-3">電話0120-907-056　Fax050-3092-0899</p>
      </div>
    </section>
    <table aria-label="納品書">
      <thead><tr>{header_html}</tr></thead>
      <tbody>{body_html}</tbody>
    </table>
  </main>
  <script>
    (() => {{
      const storageKey = "delivery-note-html:{safe_order_id}";
      const saved = JSON.parse(localStorage.getItem(storageKey) || "{{}}");
      const initial = {{}};
      const persist = () => localStorage.setItem(storageKey, JSON.stringify(saved));
      document.querySelectorAll("[data-edit]").forEach((node) => {{
        const key = node.getAttribute("data-edit");
        initial[key] = node.textContent || "";
        if (Object.prototype.hasOwnProperty.call(saved, key)) {{
          node.textContent = saved[key];
        }}
        node.addEventListener("input", () => {{
          saved[key] = node.textContent || "";
          persist();
        }});
      }});
      document.getElementById("print-note")?.addEventListener("click", () => window.print());
      document.getElementById("reset-edits")?.addEventListener("click", () => {{
        if (window.confirm("この画面で保存された編集内容をリセットします。")) {{
          localStorage.removeItem(storageKey);
          Object.keys(saved).forEach((key) => delete saved[key]);
          document.querySelectorAll("[data-edit]").forEach((node) => {{
            const key = node.getAttribute("data-edit");
            node.textContent = initial[key] || "";
          }});
        }}
      }});
    }})();
  </script>
</body>
</html>"""


def _extract_delivery_note_sheet(html: str) -> tuple[str, str]:
    title_start = html.find("<title>")
    title_end = html.find("</title>")
    title = html[title_start + len("<title>") : title_end] if title_start >= 0 and title_end > title_start else "納品書"
    sheet_start = html.find('<main class="sheet"')
    if sheet_start < 0:
        sheet_start = html.find('<main class="sheet">')
    sheet_end = html.find("</main>", sheet_start)
    if sheet_start < 0 or sheet_end < 0:
        raise ValueError("delivery note sheet not found")
    return title, html[sheet_start : sheet_end + len("</main>")]


def _extract_delivery_note_style(html: str) -> str:
    style_start = html.find("<style>")
    style_end = html.find("</style>", style_start)
    if style_start < 0 or style_end < 0:
        return ""
    return html[style_start + len("<style>") : style_end]


def _render_editable_daily_delivery_note_html(
    target_date: dt_date,
    *,
    status: str | None = None,
) -> tuple[str, dict]:
    orders = order_service.list_orders_by_line_date(target_date, status=status)
    order_ids = [
        str(order_summary.get("id") or "").strip()
        for order_summary in orders
        if str(order_summary.get("id") or "").strip()
    ]
    if not order_ids:
        raise ValueError("対象日の納品書出力対象がありません")

    safe_date = escape(target_date.isoformat())
    safe_status = escape(status or "全て")
    order_ids_json = json.dumps(order_ids, ensure_ascii=False)
    html = f"""<!doctype html>
<html lang="ja">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>{safe_date} 当日納品書HTML</title>
  <style>
    @page {{ size: A4 landscape; margin: 8mm; }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      background: #e9ecef;
      color: #111827;
      font-family: -apple-system, BlinkMacSystemFont, "Hiragino Kaku Gothic ProN", "Yu Gothic", Meiryo, sans-serif;
      font-size: 12px;
    }}
    .toolbar {{
      position: sticky;
      top: 0;
      z-index: 10;
      display: flex;
      align-items: center;
      gap: 8px;
      padding: 10px 14px;
      background: #ffffff;
      border-bottom: 1px solid #cbd5e1;
    }}
    .toolbar-title {{ font-weight: 700; margin-right: auto; }}
    button {{
      border: 1px solid #64748b;
      background: #ffffff;
      color: #0f172a;
      border-radius: 6px;
      padding: 7px 12px;
      font: inherit;
      cursor: pointer;
    }}
    button.primary {{ background: #0f172a; color: #ffffff; border-color: #0f172a; }}
    button:disabled {{ opacity: 0.45; cursor: wait; }}
    .status-line {{
      max-width: 297mm;
      margin: 10px auto;
      padding: 8px 12px;
      background: #ffffff;
      border: 1px solid #cbd5e1;
    }}
    .sheet {{
      width: 297mm;
      min-height: 210mm;
      margin: 12px auto;
      padding: 9mm;
      background: #ffffff;
      box-shadow: 0 2px 10px rgba(15, 23, 42, 0.18);
    }}
    .company {{ display: grid; grid-template-columns: 1fr 1.4fr; gap: 8mm; margin-bottom: 8mm; }}
    .doc-title {{ font-size: 24px; font-weight: 800; margin: 0 0 5mm; }}
    .facility-label {{ font-weight: 700; border-bottom: 2px solid #111827; display: inline-block; margin-bottom: 3mm; }}
    .facility-box {{ border: 2px solid #111827; min-height: 18mm; display: flex; align-items: center; justify-content: center; font-size: 24px; font-weight: 700; padding: 4mm; }}
    .company-name {{ font-size: 18px; font-weight: 800; margin-bottom: 5mm; }}
    .company-line {{ font-size: 10px; font-weight: 700; margin: 0 0 3mm; }}
    table {{ width: 100%; border-collapse: collapse; table-layout: auto; border: 2px solid #111827; background: #ffffff; }}
    th, td {{ border: 1px solid #111827; min-height: 7mm; height: 7mm; padding: 2px 4px; text-align: center; vertical-align: middle; white-space: pre-wrap; word-break: break-word; }}
    th {{ font-weight: 400; font-size: 14px; height: 12mm; min-height: 12mm; }}
    thead th {{ border-top-width: 2px; border-bottom-width: 2px; }}
    tr.daypart-start td {{ border-top-width: 2px; }}
    tbody tr:last-child td {{ border-bottom-width: 2px; }}
    th:first-child, td:first-child {{ border-left-width: 2px; }}
    th:last-child, td:last-child {{ border-right-width: 2px; }}
    td.menu-cell {{ text-align: left; font-weight: 700; }}
    .col-0 {{ width: 9%; }}
    .col-1 {{ width: 6%; }}
    .col-2 {{ width: 9%; }}
    .col-3 {{ width: 24%; }}
    th[class*="col-"]:not(.col-0):not(.col-1):not(.col-2):not(.col-3),
    td[class*="col-"]:not(.col-0):not(.col-1):not(.col-2):not(.col-3) {{ width: 7%; }}
    [contenteditable="true"] {{ cursor: text; }}
    [contenteditable="true"]:focus {{ outline: 2px solid #2563eb; outline-offset: -2px; background: #eff6ff; }}
    .bundle-page {{ break-after: page; page-break-after: always; }}
    .bundle-page:last-child {{ break-after: auto; page-break-after: auto; }}
    .bundle-title {{ display: none; }}
    .build-errors {{ max-width: 297mm; margin: 10px auto; padding: 8px 12px; background: #fff7ed; border: 1px solid #fed7aa; }}
    @media print {{
      body {{ background: #ffffff; }}
      .toolbar, .status-line, .build-errors {{ display: none; }}
      .sheet {{ width: auto; min-height: auto; margin: 0; padding: 0; box-shadow: none; }}
      .bundle-page {{ break-after: page; page-break-after: always; }}
      .bundle-page:last-child {{ break-after: auto; page-break-after: auto; }}
    }}
  </style>
</head>
<body>
  <div class="toolbar">
    <div class="toolbar-title">{safe_date} 当日納品書HTML / ステータス: {safe_status} / {len(order_ids)}件</div>
    <button type="button" id="print-note" class="primary" disabled>読み込み中</button>
  </div>
  <div class="status-line" id="load-status">納品書を読み込んでいます: 0 / {len(order_ids)}件</div>
  <div id="build-errors" class="build-errors" hidden></div>
  <div id="daily-sheets"></div>
  <script>
    (() => {{
      const orderIds = {order_ids_json};
      const root = document.getElementById("daily-sheets");
      const status = document.getElementById("load-status");
      const errors = document.getElementById("build-errors");
      const printButton = document.getElementById("print-note");
      let done = 0;
      let failed = 0;
      const parser = new DOMParser();
      const authHeader = window.sessionStorage.getItem("auth_header") || "";
      const updateStatus = () => {{
        status.textContent = `納品書を読み込んでいます: ${{done}} / ${{orderIds.length}}件`;
        if (done >= orderIds.length) {{
          status.textContent = failed ? `読み込み完了: ${{orderIds.length - failed}}件 / 失敗 ${{failed}}件` : `読み込み完了: ${{orderIds.length}}件`;
          printButton.textContent = "印刷/PDF保存";
          printButton.disabled = false;
        }}
      }};
      const appendError = (orderId, message) => {{
        failed += 1;
        errors.hidden = false;
        const line = document.createElement("div");
        line.textContent = `${{orderId}}: ${{message}}`;
        errors.appendChild(line);
      }};
      const loadOne = async (orderId, index) => {{
        try {{
          const response = await fetch(`/api/outputs/delivery-notes/html?order_id=${{encodeURIComponent(orderId)}}&date={safe_date}`, {{
            headers: authHeader ? {{ Authorization: authHeader }} : undefined,
          }});
          if (!response.ok) {{
            throw new Error(await response.text());
          }}
          const doc = parser.parseFromString(await response.text(), "text/html");
          const sheet = doc.querySelector(".sheet");
          if (!sheet) {{
            throw new Error("納品書本体が見つかりません");
          }}
          const section = document.createElement("section");
          section.className = "bundle-page";
          section.dataset.orderId = orderId;
          section.appendChild(sheet);
          root.appendChild(section);
        }} catch (error) {{
          appendError(orderId, error instanceof Error ? error.message : String(error));
        }} finally {{
          done += 1;
          updateStatus();
        }}
      }};
      printButton.addEventListener("click", () => window.print());
      orderIds.forEach((orderId, index) => loadOne(orderId, index));
      updateStatus();
    }})();
  </script>
</body>
</html>"""
    return html, {"total_orders": len(orders), "success_orders": len(order_ids), "error_orders": 0}


def _delivery_facility_name(order_id: str) -> str:
    with session_scope() as session:
        order = session.get(Order, order_id)
        facility_id = str(getattr(order, "facility_code", "") or "").strip() if order else ""
    if not facility_id:
        return ""
    facility = facility_service.get_facility(facility_id)
    if isinstance(facility, dict):
        return str(facility.get("name") or facility.get("facility_name") or facility_id)
    return facility_id


def _preview_csv(path: str, encoding: str, limit: int) -> dict:
    with open(path, newline="", encoding=encoding, errors="replace") as f:
        reader = csv.reader(f)
        header = next(reader, [])
        rows: list[list[str]] = []
        for idx, row in enumerate(reader):
            if idx >= limit:
                break
            rows.append([str(cell) for cell in row])
    return {"headers": header, "rows": rows}


def _normalize_cell(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and pd.isna(value):
        return ""
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return str(value)


def _preview_excel(path: str, limit: int) -> dict:
    df = pd.read_excel(path)
    if limit > 0:
        df = df.head(limit)
    headers = [str(col) for col in df.columns]
    rows: list[list[str]] = []
    for _, row in df.iterrows():
        rows.append([_normalize_cell(row.get(col)) for col in df.columns])
    return {"headers": headers, "rows": rows}


def _parse_iso_date(value: str) -> dt_date:
    try:
        return dt_date.fromisoformat(value)
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=400, detail="date must be YYYY-MM-DD") from exc


@router.get("/labels", dependencies=[Depends(require_role("operator"))])
def download_labels(order_id: str):
    outputs = build_output_preview(order_id, "labels")
    path = outputs["labels"]
    logger.info("Output download", order_id=order_id, output="labels", path=path)
    return FileResponse(path, media_type="text/csv", filename=f"{order_id}_labels.csv")


@router.get("/delivery-notes", dependencies=[Depends(require_role("operator"))])
def download_delivery(order_id: str):
    outputs = build_output_preview(order_id, "delivery")
    path = outputs["delivery_note"]
    logger.info("Output download", order_id=order_id, output="delivery", path=path)
    return FileResponse(
        path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=f"{order_id}_delivery.xlsx",
    )


@router.get("/delivery-notes/html", response_class=HTMLResponse, dependencies=[Depends(require_role("operator"))])
def view_delivery_note_html(order_id: str, date: str | None = None):
    try:
        target_date = _parse_iso_date(date) if date else None
        preview = build_delivery_preview(order_id, include_diagnostics=False, target_date=target_date)
        headers = preview.get("headers", [])
        rows = preview.get("rows", [])
        facility_name = _delivery_facility_name(order_id)
        html = _render_editable_delivery_note_html(
            order_id,
            f"{order_id} 納品書",
            headers,
            rows,
            facility_name,
            preview.get("columns", []),
            preview.get("raw_rows", []),
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=500, detail=f"delivery note html build failed: {exc}") from exc
    logger.info("Output html view", order_id=order_id, output="delivery")
    timings = preview.get("timings", {}) if isinstance(preview, dict) else {}
    timing_header = ",".join(f"{key};dur={value}" for key, value in timings.items())
    headers = {"Server-Timing": timing_header} if timing_header else None
    return HTMLResponse(html, headers=headers)


@router.get("/daily-delivery-notes/html", response_class=HTMLResponse, dependencies=[Depends(require_role("operator"))])
def view_daily_delivery_note_html(date: str, status: str | None = None):
    target_date = _parse_iso_date(date)
    try:
        html, summary = _render_editable_daily_delivery_note_html(target_date, status=status)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=500, detail=f"daily delivery note html build failed: {exc}") from exc
    logger.info(
        "Daily output html view",
        date=target_date.isoformat(),
        output="delivery",
        total_orders=summary.get("total_orders"),
        success_orders=summary.get("success_orders"),
        error_orders=summary.get("error_orders"),
    )
    return HTMLResponse(
        html,
        headers={
            "X-Daily-Delivery-Total-Orders": str(summary.get("total_orders", 0)),
            "X-Daily-Delivery-Success-Orders": str(summary.get("success_orders", 0)),
            "X-Daily-Delivery-Error-Orders": str(summary.get("error_orders", 0)),
        },
    )


@router.get("/order-form-saved-sheet", dependencies=[Depends(require_role("operator"))])
def download_order_form_saved_sheet(order_id: str):
    try:
        path = order_form_service.build_saved_sheet_order_form_excel(order_id=order_id)
    except ValueError as exc:
        detail = str(exc)
        if detail == "order not found":
            raise HTTPException(status_code=404, detail=detail) from exc
        raise HTTPException(status_code=400, detail=detail) from exc
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=500, detail=f"order form saved sheet build failed: {exc}") from exc
    logger.info("Output download", order_id=order_id, output="order_form_saved_sheet", path=path)
    return FileResponse(
        path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=path.name,
    )


@router.get("/manufacturing-aggregate", dependencies=[Depends(require_role("operator"))])
def download_aggregate(order_id: str):
    outputs = build_output_preview(order_id, "aggregate")
    path = outputs["aggregate"]
    logger.info("Output download", order_id=order_id, output="aggregate", path=path)
    return FileResponse(path, media_type="text/csv", filename=f"{order_id}_aggregate.csv")


@router.get("/preview", dependencies=[Depends(require_role("operator"))])
def preview_output(order_id: str, type: str, limit: int = _PREVIEW_LIMIT_DEFAULT):
    limit = max(1, min(limit, _PREVIEW_LIMIT_DEFAULT))
    if type == "labels":
        outputs = build_output_preview(order_id, "labels")
        payload = _preview_csv(outputs["labels"], "cp932", limit)
    elif type == "delivery":
        preview = build_delivery_preview(order_id)
        headers = preview.get("headers", [])
        rows = preview.get("rows", [])
        payload = {
            "headers": headers,
            "rows": rows[:limit],
            "ocr_entry_count": preview.get("ocr_entry_count"),
        }
    elif type == "aggregate":
        outputs = build_output_preview(order_id, "aggregate")
        payload = _preview_csv(outputs["aggregate"], "cp932", limit)
    else:
        raise HTTPException(status_code=400, detail="invalid output type")
    return {"type": type, **payload}


@router.get("/file-preview", response_class=HTMLResponse, dependencies=[Depends(require_role("operator"))])
def preview_output_file(order_id: str, type: str):
    try:
        path, title = _output_file_for_type(order_id, type)
        if type in {"labels", "aggregate"}:
            html = _render_csv_preview(path, title, "cp932")
        else:
            html = _render_xlsx_preview(path, title)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=500, detail=f"failed to render output preview: {exc}") from exc
    return HTMLResponse(html)


@router.get("/daily-bundle", dependencies=[Depends(require_role("operator"))])
def download_daily_bundle(
    date: str,
    bundle_type: str = "both",
    status: str | None = None,
):
    target_date = _parse_iso_date(date)
    try:
        bundle_path, summary = build_daily_output_bundle(
            target_date,
            bundle_type=bundle_type,
            status=status,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=500, detail=f"bundle build failed: {exc}") from exc
    headers = {
        "X-Daily-Bundle-Total-Orders": str(summary.get("total_orders", 0)),
        "X-Daily-Bundle-Success-Orders": str(summary.get("success_orders", 0)),
        "X-Daily-Bundle-Error-Orders": str(summary.get("error_orders", 0)),
        "X-Daily-Bundle-Type": str(summary.get("bundle_type", bundle_type)),
    }
    file_format = str(summary.get("file_format") or "xlsx")
    filename = f"daily_outputs_{target_date.isoformat()}_{summary.get('bundle_type', bundle_type)}.{file_format}"
    media_type = (
        "application/zip"
        if file_format == "zip"
        else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    return FileResponse(
        str(bundle_path),
        media_type=media_type,
        filename=filename,
        headers=headers,
    )


@router.get("/weekly-weight", dependencies=[Depends(require_role("operator"))])
def download_weekly_weight(date: str, status: str | None = None):
    target_date = _parse_iso_date(date)
    try:
        weight_path = build_weekly_weight_summary_workbook(target_date, status=status)
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=500, detail=f"weekly weight build failed: {exc}") from exc
    return FileResponse(
        str(weight_path),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=weight_path.name,
    )
