from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
import os
import re

from loguru import logger
from openpyxl import load_workbook

from src.services import config_service
from src.services.sagawa_tracking_service import (
    TrackingStatus,
    lookup_tracking_statuses,
    normalize_tracking_key,
)

try:
    import pdfplumber
except Exception:  # noqa: BLE001
    pdfplumber = None


DATA_DIR = Path(__file__).resolve().parents[1] / "data"
SHIPPING_TEMPLATE_PATH = Path(
    os.getenv("SHIPPING_TEMPLATE_PATH", DATA_DIR / "shipping_template.xlsx")
)
OUTPUT_DIR = Path("/tmp/shipping-outputs")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

_TRACKING_PATTERN = re.compile(r"お問い合せ送り状№：([0-9\\-]+)")
_SHIP_DATE_PATTERN = re.compile(r"出荷予定日：\s*([0-9]{4}/[0-9]{2}/[0-9]{2})")
_ARRIVAL_PATTERN = re.compile(r"【配達指定】\s*([0-9]{1,2})月([0-9]{1,2})日")
_POSTAL_PATTERN = re.compile(r"^〒\\d{3}-\\d{4}")

_PREFECTURES = {
    "北海道",
    "青森",
    "岩手",
    "宮城",
    "秋田",
    "山形",
    "福島",
    "茨城",
    "栃木",
    "群馬",
    "埼玉",
    "千葉",
    "東京",
    "神奈川",
    "新潟",
    "富山",
    "石川",
    "福井",
    "山梨",
    "長野",
    "岐阜",
    "静岡",
    "愛知",
    "三重",
    "滋賀",
    "京都",
    "大阪",
    "兵庫",
    "奈良",
    "和歌山",
    "鳥取",
    "島根",
    "岡山",
    "広島",
    "山口",
    "徳島",
    "香川",
    "愛媛",
    "高知",
    "福岡",
    "佐賀",
    "長崎",
    "熊本",
    "大分",
    "宮崎",
    "鹿児島",
    "沖縄",
}

_SKIP_PREFIXES = (
    "TEL",
    "FAX",
    "出荷予定日",
    "便種",
    "【",
    "個数",
    "№",
    "送り状有効期限",
    "お客様",
    "お問い合せ",
    "集荷",
    "配達",
    "箱類",
    "クール便",
    "冷蔵",
)

_DELIVERY_FLAG_HEADER = "配達完了フラグ"


@dataclass
class ShippingRecord:
    ship_date: date | None
    arrival_date: date | None
    tracking_number: str
    facility_name: str


def _parse_date(text: str | None) -> date | None:
    if not text:
        return None
    try:
        return datetime.strptime(text, "%Y/%m/%d").date()
    except Exception:
        return None


def _parse_arrival(text: str | None, year_hint: int | None) -> date | None:
    if not text:
        return None
    match = _ARRIVAL_PATTERN.search(text)
    if not match:
        return None
    month = int(match.group(1))
    day = int(match.group(2))
    year = year_hint or datetime.now().year
    try:
        return date(year, month, day)
    except ValueError:
        return None


def _extract_facility_from_lines(lines: list[str]) -> str | None:
    for idx, line in enumerate(lines):
        if _POSTAL_PATTERN.match(line):
            for candidate in lines[idx + 1 : idx + 40]:
                if not candidate:
                    continue
                if candidate in _PREFECTURES:
                    continue
                if any(candidate.startswith(prefix) for prefix in _SKIP_PREFIXES):
                    continue
                if any(token in candidate for token in ("県", "市", "町", "村", "区", "丁目", "番地", "番")):
                    continue
                if re.search(r"[0-9０-９]", candidate):
                    continue
                return candidate
    return None


def _resolve_facility_name(page_text: str) -> str:
    candidates = config_service.match_facility_candidates(page_text)
    if candidates:
        best = max(
            candidates,
            key=lambda item: (
                1 if item.get("auto") else 0,
                float(item.get("score") or 0),
            ),
        )
        name = best.get("facility_name")
        if isinstance(name, str) and name.strip():
            return name.strip()
    lines = [line.strip() for line in page_text.splitlines() if line.strip()]
    extracted = _extract_facility_from_lines(lines)
    return extracted or ""


def extract_shipping_records(pdf_bytes: bytes) -> list[ShippingRecord]:
    if not pdfplumber:
        raise RuntimeError("pdfplumber is required for shipping PDF parsing")
    records: list[ShippingRecord] = []
    seen_numbers: set[str] = set()
    with pdfplumber.open(BytesIO(pdf_bytes)) as pdf:
        for page in pdf.pages:
            page_text = page.extract_text() or ""
            if not page_text.strip():
                continue
            track_match = _TRACKING_PATTERN.search(page_text)
            if not track_match:
                continue
            tracking_number = track_match.group(1)
            if tracking_number in seen_numbers:
                continue
            seen_numbers.add(tracking_number)

            ship_date = _parse_date(
                next((m.group(1) for m in _SHIP_DATE_PATTERN.finditer(page_text)), None)
            )
            arrival_date = _parse_arrival(page_text, ship_date.year if ship_date else None)
            facility_name = _resolve_facility_name(page_text)
            records.append(
                ShippingRecord(
                    ship_date=ship_date,
                    arrival_date=arrival_date,
                    tracking_number=tracking_number,
                    facility_name=facility_name,
                )
            )
    return records


def _find_header_row(ws, header_tokens: list[str], max_scan: int = 30) -> int | None:
    for row in range(1, max_scan + 1):
        values = [ws.cell(row=row, column=col).value for col in range(1, ws.max_column + 1)]
        for token in header_tokens:
            if token in values:
                return row
    return None


def _resolve_template_sheet(workbook, ship_date: date | None):
    if ship_date:
        month = ship_date.month
        for name in workbook.sheetnames:
            if "佐川" in name and str(month) in name:
                return workbook[name]
    return workbook.active


def build_shipping_excel(records: list[ShippingRecord]) -> Path:
    if not SHIPPING_TEMPLATE_PATH.exists():
        raise FileNotFoundError(f"shipping template missing: {SHIPPING_TEMPLATE_PATH}")
    workbook = load_workbook(SHIPPING_TEMPLATE_PATH)
    ship_date = next((r.ship_date for r in records if r.ship_date), None)
    ws = _resolve_template_sheet(workbook, ship_date)

    header_row = _find_header_row(ws, ["お届け先", "伝票番号"])
    if not header_row:
        header_row = 2

    col_map = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=header_row, column=col).value
        if value == "お届け先":
            col_map["facility"] = col
        elif value == "段ボールサイズ":
            col_map["size"] = col
        elif value == "伝票番号":
            col_map["tracking"] = col
        elif value == "到着日時":
            col_map["arrival"] = col
    if "facility" in col_map:
        col_map.setdefault("date", max(col_map["facility"] - 1, 1))
    col_map.setdefault("date", 2)

    # Clear existing rows (keep formatting)
    for row in range(header_row + 1, ws.max_row + 1):
        for key in ("date", "facility", "size", "tracking", "arrival"):
            col_idx = col_map.get(key)
            if not col_idx:
                continue
            ws.cell(row=row, column=col_idx).value = None

    start_row = header_row + 1
    for idx, record in enumerate(records):
        row_idx = start_row + idx
        if record.ship_date:
            ws.cell(row=row_idx, column=col_map["date"]).value = record.ship_date
        ws.cell(row=row_idx, column=col_map["facility"]).value = record.facility_name
        ws.cell(row=row_idx, column=col_map["tracking"]).value = record.tracking_number
        if record.arrival_date:
            ws.cell(row=row_idx, column=col_map["arrival"]).value = record.arrival_date

    output_path = OUTPUT_DIR / f"shipping_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    workbook.save(output_path)
    logger.info("Shipping Excel created", path=str(output_path), records=len(records))
    return output_path


def _normalize_tracking_number(value: object) -> str:
    if value is None:
        return ""
    return normalize_tracking_key(str(value))


def _coerce_excel_date(value: object) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    raw = str(value).strip()
    if not raw:
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(raw, fmt).date()
        except Exception:
            continue
    return None


def _find_column_indexes(ws, header_row: int) -> dict[str, int]:
    col_map: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=header_row, column=col).value
        key = str(value).strip() if value is not None else ""
        if key == "発送日":
            col_map["ship_date"] = col
        elif key == "お届け先":
            col_map["facility"] = col
        elif key == "段ボールサイズ":
            col_map["size"] = col
        elif key == "伝票番号":
            col_map["tracking"] = col
        elif key == "到着日時":
            col_map["arrival"] = col
        elif key == _DELIVERY_FLAG_HEADER:
            col_map["delivery_flag"] = col
    if "arrival" not in col_map:
        col_map["arrival"] = ws.max_column + 1
        ws.cell(row=header_row, column=col_map["arrival"]).value = "到着日時"
    if "delivery_flag" not in col_map:
        col_map["delivery_flag"] = ws.max_column + 1
        ws.cell(row=header_row, column=col_map["delivery_flag"]).value = _DELIVERY_FLAG_HEADER
    return col_map


def _has_tracking_number(value: str) -> bool:
    digits = re.sub(r"\D", "", value)
    return len(digits) >= 10


def get_tracking_statuses(tracking_numbers: list[str]) -> list[dict]:
    statuses = get_tracking_status_records(tracking_numbers)
    return [status.serialize() for status in statuses]


def get_tracking_status_records(tracking_numbers: list[str]) -> list[TrackingStatus]:
    return lookup_tracking_statuses(tracking_numbers)


def enrich_tracking_excel(excel_bytes: bytes) -> tuple[Path, dict]:
    workbook = load_workbook(BytesIO(excel_bytes))
    row_refs: list[dict] = []
    unique_numbers: list[str] = []
    seen_numbers: set[str] = set()
    for ws in workbook.worksheets:
        header_row = _find_header_row(ws, ["伝票番号"], max_scan=40)
        if not header_row:
            continue
        col_map = _find_column_indexes(ws, header_row)
        tracking_col = col_map.get("tracking")
        if not tracking_col:
            continue
        for row_idx in range(header_row + 1, ws.max_row + 1):
            raw_tracking = ws.cell(row=row_idx, column=tracking_col).value
            tracking_key = _normalize_tracking_number(raw_tracking)
            if not tracking_key or not _has_tracking_number(tracking_key):
                continue
            ship_date = None
            ship_date_col = col_map.get("ship_date")
            if ship_date_col:
                ship_date = _coerce_excel_date(ws.cell(row=row_idx, column=ship_date_col).value)
            facility_name = ""
            facility_col = col_map.get("facility")
            if facility_col:
                raw_facility = ws.cell(row=row_idx, column=facility_col).value
                facility_name = str(raw_facility).strip() if raw_facility is not None else ""
            row_refs.append(
                {
                    "sheet": ws,
                    "row": row_idx,
                    "tracking_key": tracking_key,
                    "arrival_col": col_map["arrival"],
                    "flag_col": col_map["delivery_flag"],
                    "facility_name": facility_name,
                    "ship_date": ship_date,
                }
            )
            if tracking_key not in seen_numbers:
                seen_numbers.add(tracking_key)
                unique_numbers.append(tracking_key)

    status_map: dict[str, TrackingStatus] = {}
    status_records = lookup_tracking_statuses(unique_numbers)
    for status in status_records:
        if not status.tracking_key:
            continue
        status_map[status.tracking_key] = status
    facility_by_tracking: dict[str, str] = {}
    ship_date_by_tracking: dict[str, date] = {}
    for ref in row_refs:
        facility_name = str(ref.get("facility_name") or "").strip()
        tracking_key = str(ref.get("tracking_key") or "").strip()
        if facility_name and tracking_key and tracking_key not in facility_by_tracking:
            facility_by_tracking[tracking_key] = facility_name
        ship_date = ref.get("ship_date")
        if tracking_key and isinstance(ship_date, date) and tracking_key not in ship_date_by_tracking:
            ship_date_by_tracking[tracking_key] = ship_date

    delivered_rows = 0
    pending_rows = 0
    updated_arrival_rows = 0
    error_rows = 0
    for ref in row_refs:
        status = status_map.get(ref["tracking_key"])
        ws = ref["sheet"]
        row_idx = ref["row"]
        arrival_cell = ws.cell(row=row_idx, column=ref["arrival_col"])
        flag_cell = ws.cell(row=row_idx, column=ref["flag_col"])
        if not status:
            flag_cell.value = "照会失敗"
            pending_rows += 1
            error_rows += 1
            continue
        if status.error:
            flag_cell.value = "照会失敗"
            pending_rows += 1
            error_rows += 1
            continue
        if status.delivered:
            delivered_rows += 1
            flag_cell.value = "完了"
            has_arrival = arrival_cell.value is not None and str(arrival_cell.value).strip() != ""
            if not has_arrival and status.arrival_text:
                arrival_cell.value = status.arrival_text
                updated_arrival_rows += 1
        else:
            pending_rows += 1
            if status.status == "該当なし":
                flag_cell.value = "該当なし"
            elif status.status == "invalid":
                flag_cell.value = "形式不正"
            else:
                flag_cell.value = "未完了"

    output_path = OUTPUT_DIR / f"shipping_enriched_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    workbook.save(output_path)
    summary = {
        "total_rows": len(row_refs),
        "lookup_count": len(unique_numbers),
        "delivered_rows": delivered_rows,
        "pending_rows": pending_rows,
        "updated_arrival_rows": updated_arrival_rows,
        "error_rows": error_rows,
        "all_delivered": len(row_refs) > 0 and pending_rows == 0,
        "_status_items": [status.serialize() for status in status_records],
        "_facility_by_tracking": facility_by_tracking,
        "_ship_date_by_tracking": ship_date_by_tracking,
    }
    logger.info("Shipping Excel enriched", path=str(output_path), **summary)
    return output_path, summary
