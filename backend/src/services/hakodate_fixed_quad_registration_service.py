from __future__ import annotations

from dataclasses import asdict, dataclass
import json
from pathlib import Path
from typing import Any

import cv2
import fitz
import numpy as np
from openpyxl import load_workbook
from PIL import Image

from src.services import workbook_pdf_renderer


FORBIDDEN_DOWNSTREAM_METHODS = [
    "fax_candidate_gap_sequence_match",
    "drop_one_extra_fax_line_by_min_affine_residual",
    "ordered_affine_dp_match_with_interpolated_missing_lines",
]
ORDER_FORM_TEMPLATE_Y_EDGE_COUNT = 59
TWO_STAGE_HEADER_BOUNDARY_MIN_WIDTH_RATIO = 0.12
ORDER_FORM_TEMPLATE_HEADER_EDGE_ROWS = (7, 9)
ORDER_FORM_TEMPLATE_BODY_EDGE_ROWS = tuple(range(11, 68))


@dataclass(frozen=True)
class FixedQuadTemplateRegistrationResult:
    facility_code: str
    order_id: str
    fax_pdf: str
    template_pdf: str
    quad_source: str | None
    quad_px: list[list[float]]
    rectified_canvas_size: list[int]
    template_axes_x: list[int]
    template_axes_y_used_count: int
    template_axes_y_used_first_last: list[int]
    template_axes_y_all: list[int]
    template_outer_grid_bbox_used: list[int]
    legacy_manifest_template_bbox_not_used: list[float]
    forbidden_downstream_methods: list[str]
    outputs: dict[str, str | None]


def render_pdf_page_to_rgb_at_dpi(pdf_path: str | Path, *, dpi: int = 220) -> Image.Image:
    doc = fitz.open(str(pdf_path))
    if doc.page_count < 1:
        raise ValueError(f"pdf has no pages: {pdf_path}")
    page = doc.load_page(0)
    pix = page.get_pixmap(matrix=fitz.Matrix(dpi / 72, dpi / 72), alpha=False)
    return Image.frombytes("RGB", [pix.width, pix.height], pix.samples)


def _edge_locked_binarize(arr: np.ndarray) -> np.ndarray:
    gray = cv2.cvtColor(arr, cv2.COLOR_RGB2GRAY)
    gray = cv2.GaussianBlur(gray, (3, 3), 0)
    return cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)[1]


def _edge_locked_masks(thresholded: np.ndarray) -> tuple[np.ndarray, np.ndarray, np.ndarray]:
    height, width = thresholded.shape
    horizontal = cv2.morphologyEx(
        thresholded,
        cv2.MORPH_OPEN,
        cv2.getStructuringElement(cv2.MORPH_RECT, (max(28, width // 90), 1)),
        iterations=1,
    )
    vertical = cv2.morphologyEx(
        thresholded,
        cv2.MORPH_OPEN,
        cv2.getStructuringElement(cv2.MORPH_RECT, (1, max(28, height // 90))),
        iterations=1,
    )
    combined = cv2.bitwise_or(horizontal, vertical)
    combined = cv2.morphologyEx(
        combined,
        cv2.MORPH_CLOSE,
        cv2.getStructuringElement(cv2.MORPH_RECT, (9, 9)),
        iterations=1,
    )
    combined = cv2.dilate(
        combined,
        cv2.getStructuringElement(cv2.MORPH_RECT, (3, 3)),
        iterations=1,
    )
    return horizontal, vertical, combined


def _edge_locked_choose_component(combined: np.ndarray) -> tuple[np.ndarray, dict[str, Any], str]:
    height, width = combined.shape
    count, labels, stats, _centroids = cv2.connectedComponentsWithStats((combined > 0).astype("uint8"), 8)
    best: dict[str, Any] | None = None
    for label in range(1, count):
        x, y, w, h, area = stats[label]
        if w < width * 0.35 or h < height * 0.25 or area < 800:
            continue
        score = w * h + area * 8
        if y < height * 0.03:
            score *= 0.75
        row = {
            "label": int(label),
            "x": int(x),
            "y": int(y),
            "w": int(w),
            "h": int(h),
            "area": int(area),
            "score": float(score),
        }
        if best is None or row["score"] > best["score"]:
            best = row
    source = "large_grid_component"
    if best is None:
        for label in range(1, count):
            x, y, w, h, area = stats[label]
            row = {
                "label": int(label),
                "x": int(x),
                "y": int(y),
                "w": int(w),
                "h": int(h),
                "area": int(area),
                "score": float(area),
            }
            if best is None or row["score"] > best["score"]:
                best = row
        source = "fallback_largest_component"
    if best is None:
        raise ValueError("edge locked grid component not found")
    return (labels == best["label"]).astype("uint8") * 255, best, source


def _edge_locked_order_quad(points: np.ndarray) -> list[np.ndarray]:
    points = np.asarray(points, float)
    sums = points.sum(axis=1)
    diffs = np.diff(points, axis=1).reshape(-1)
    return [
        points[np.argmin(sums)],
        points[np.argmin(diffs)],
        points[np.argmax(sums)],
        points[np.argmax(diffs)],
    ]


def _edge_locked_line2(point1: np.ndarray, point2: np.ndarray) -> np.ndarray:
    point1 = np.asarray(point1, float)
    point2 = np.asarray(point2, float)
    vector = point2 - point1
    norm = np.linalg.norm(vector)
    return np.array([vector[0] / norm, vector[1] / norm, point1[0], point1[1]], float)


def _edge_locked_dist(points: np.ndarray, line: np.ndarray) -> np.ndarray:
    vx, vy, x0, y0 = line
    normal = np.array([-vy, vx], float)
    return (points - np.array([x0, y0], float)) @ normal


def _edge_locked_fit(points: np.ndarray, fallback: np.ndarray) -> tuple[np.ndarray, str]:
    if len(points) >= 30:
        vx, vy, x0, y0 = cv2.fitLine(
            np.asarray(points, np.float32),
            cv2.DIST_L1,
            0,
            0.01,
            0.01,
        ).flatten()
        return np.array([vx, vy, x0, y0], float), "fit"
    return fallback, "fallback"


def _edge_locked_intersect(line1: np.ndarray, line2: np.ndarray) -> np.ndarray | None:
    vx1, vy1, x1, y1 = line1
    vx2, vy2, x2, y2 = line2
    matrix = np.array([[vx1, -vx2], [vy1, -vy2]], float)
    if abs(np.linalg.det(matrix)) < 1e-8:
        return None
    t, _ = np.linalg.solve(matrix, np.array([x2 - x1, y2 - y1], float))
    return np.array([x1 + vx1 * t, y1 + vy1 * t], float)


def _edge_locked_line_points(mask: np.ndarray) -> np.ndarray:
    return np.column_stack(np.where(mask > 0)[::-1]).astype(float)


def _edge_locked_refine(
    component_mask: np.ndarray,
    horizontal_mask: np.ndarray,
    vertical_mask: np.ndarray,
) -> tuple[list[np.ndarray], list[np.ndarray], dict[str, np.ndarray], dict[str, np.ndarray], dict[str, str], dict[str, int]]:
    ys, xs = np.where(component_mask > 0)
    points = np.column_stack([xs, ys]).astype(np.float32)
    initial = _edge_locked_order_quad(cv2.boxPoints(cv2.minAreaRect(points)))
    top_left, top_right, bottom_right, bottom_left = initial
    base = {
        "top": _edge_locked_line2(top_left, top_right),
        "right": _edge_locked_line2(top_right, bottom_right),
        "bottom": _edge_locked_line2(bottom_left, bottom_right),
        "left": _edge_locked_line2(top_left, bottom_left),
    }
    horizontal_points = _edge_locked_line_points(cv2.bitwise_and(component_mask, horizontal_mask))
    vertical_points = _edge_locked_line_points(cv2.bitwise_and(component_mask, vertical_mask))
    component_points = _edge_locked_line_points(component_mask)
    refined: dict[str, np.ndarray] = {}
    sources: dict[str, str] = {}
    counts: dict[str, int] = {}
    for edge, base_line in base.items():
        if edge in ("top", "bottom") and len(horizontal_points) > 0:
            primary = horizontal_points
        elif edge in ("left", "right") and len(vertical_points) > 0:
            primary = vertical_points
        else:
            primary = component_points
        # Exact accepted v4 lock: only pixels close to the initial outer edge.
        distances = np.abs(_edge_locked_dist(primary, base_line))
        band = 22.0
        selected = primary[distances <= band]
        if len(selected) < 30:
            fallback_distances = np.abs(_edge_locked_dist(component_points, base_line))
            selected = component_points[fallback_distances <= band]
        refined[edge], sources[edge] = _edge_locked_fit(selected, base_line)
        counts[edge] = int(len(selected))
    quad = [
        _edge_locked_intersect(refined["top"], refined["left"]),
        _edge_locked_intersect(refined["top"], refined["right"]),
        _edge_locked_intersect(refined["bottom"], refined["right"]),
        _edge_locked_intersect(refined["bottom"], refined["left"]),
    ]
    if any(point is None for point in quad):
        quad = initial
    return initial, quad, base, refined, sources, counts  # type: ignore[return-value]


def _edge_locked_collect(
    mask: np.ndarray,
    line: np.ndarray,
    point1: np.ndarray,
    point2: np.ndarray,
    *,
    search: int = 14,
    samples: int = 360,
) -> dict[str, float]:
    height, width = mask.shape
    point1 = np.asarray(point1, float)
    point2 = np.asarray(point2, float)
    vx, vy, _x0, _y0 = line
    normal = np.array([-vy, vx], float)
    normal = normal / (np.linalg.norm(normal) or 1)
    hits = 0
    offsets: list[int] = []
    miss = 0
    max_miss = 0
    for index in range(samples):
        point = point1 * (1 - index / (samples - 1)) + point2 * (index / (samples - 1))
        found = None
        for delta in range(-search, search + 1):
            candidate = point + normal * delta
            x = int(round(candidate[0]))
            y = int(round(candidate[1]))
            if 0 <= x < width and 0 <= y < height and mask[y, x] > 0:
                found = delta
                break
        if found is None:
            miss += 1
            max_miss = max(max_miss, miss)
        else:
            hits += 1
            offsets.append(abs(found))
            miss = 0
    line_length = float(np.linalg.norm(point2 - point1))
    return {
        "hit_rate": round(hits / samples, 4),
        "mean_abs_offset_px": round(float(np.mean(offsets)) if offsets else 999, 3),
        "max_abs_offset_px": round(float(np.max(offsets)) if offsets else 999, 3),
        "gap_max_px_est": round(max_miss * line_length / max(1, samples - 1), 1),
    }


def _edge_locked_validate(
    horizontal_mask: np.ndarray,
    vertical_mask: np.ndarray,
    combined_mask: np.ndarray,
    quad: list[np.ndarray],
    lines: dict[str, np.ndarray],
) -> dict[str, dict[str, float]]:
    top_left, top_right, bottom_right, bottom_left = quad
    return {
        "top": _edge_locked_collect(cv2.bitwise_or(horizontal_mask, combined_mask), lines["top"], top_left, top_right),
        "right": _edge_locked_collect(cv2.bitwise_or(vertical_mask, combined_mask), lines["right"], top_right, bottom_right),
        "bottom": _edge_locked_collect(
            cv2.bitwise_or(horizontal_mask, combined_mask),
            lines["bottom"],
            bottom_left,
            bottom_right,
        ),
        "left": _edge_locked_collect(cv2.bitwise_or(vertical_mask, combined_mask), lines["left"], top_left, bottom_left),
    }


def _edge_locked_reasons(
    metrics: dict[str, dict[str, float]],
    edge_sources: dict[str, str],
    component_source: str,
) -> list[str]:
    reasons: list[str] = []
    if component_source != "large_grid_component":
        reasons.append(component_source)
    for edge, metric in metrics.items():
        if edge_sources.get(edge) != "fit":
            reasons.append(f"{edge}_line_not_refit")
        if metric["hit_rate"] < 0.78:
            reasons.append(f"{edge}_hit_rate_low:{metric['hit_rate']}")
        if metric["mean_abs_offset_px"] > 4.5:
            reasons.append(f"{edge}_offset_high:{metric['mean_abs_offset_px']}")
        if metric["gap_max_px_est"] > 140:
            reasons.append(f"{edge}_gap_large:{metric['gap_max_px_est']}")
    return reasons


def estimate_edge_locked_quad_from_pdf(
    pdf_path: str | Path,
    *,
    dpi: int = 220,
) -> dict[str, Any]:
    image = render_pdf_page_to_rgb_at_dpi(pdf_path, dpi=dpi)
    arr = np.array(image)
    thresholded = _edge_locked_binarize(arr)
    horizontal_mask, vertical_mask, combined_mask = _edge_locked_masks(thresholded)
    component_mask, component, component_source = _edge_locked_choose_component(combined_mask)
    initial, refined, _base, refined_lines, edge_sources, edge_point_counts = _edge_locked_refine(
        component_mask,
        horizontal_mask,
        vertical_mask,
    )
    metrics = _edge_locked_validate(horizontal_mask, vertical_mask, combined_mask, refined, refined_lines)
    reasons = _edge_locked_reasons(metrics, edge_sources, component_source)
    shift = {
        key: round(float(np.linalg.norm(np.asarray(refined[index]) - np.asarray(initial[index]))), 2)
        for index, key in enumerate(["TL", "TR", "BR", "BL"])
    }
    return {
        "status": "ok" if not reasons else "ng",
        "reasons": reasons,
        "component_source": component_source,
        "component": component,
        "edge_sources": edge_sources,
        "edge_point_counts": edge_point_counts,
        "initial_quad_px": [[round(float(x), 2), round(float(y), 2)] for x, y in initial],
        "refined_quad_px": [[round(float(x), 2), round(float(y), 2)] for x, y in refined],
        "corner_shift_px": shift,
        "metrics": metrics,
    }


def resolve_fixed_quad_px_for_manifest_item(item: dict[str, Any]) -> tuple[list[list[float]], str, dict[str, Any] | None]:
    quad = item.get("quad_px")
    if isinstance(quad, list) and len(quad) == 4:
        return quad, str(item.get("quad_source") or "manifest_quad_px"), None
    pdf_path = item.get("fax_pdf") or item.get("local_pdf")
    if not pdf_path:
        raise ValueError("quad_px missing and fax_pdf/local_pdf unavailable")
    estimate = estimate_edge_locked_quad_from_pdf(str(pdf_path), dpi=220)
    if estimate["status"] != "ok":
        raise ValueError(f"edge locked quad estimation failed: {estimate['reasons']}")
    return estimate["refined_quad_px"], "edge_locked_v4_estimated_from_fax_pdf", estimate


def load_fixed_quad_manifest_item(
    manifest_path: Path,
    *,
    facility_code: str,
    order_id: str,
) -> dict[str, Any]:
    with manifest_path.open() as f:
        manifest = json.load(f)
    for item in manifest.get("results") or []:
        if item.get("facility_code") == facility_code and item.get("order_id") == order_id:
            return item
    raise ValueError(f"manifest item not found: facility={facility_code} order={order_id}")


def render_pdf_page_to_bgr(pdf_path: str, *, width: int | None = None) -> np.ndarray:
    doc = fitz.open(pdf_path)
    if doc.page_count < 1:
        raise ValueError(f"pdf has no pages: {pdf_path}")
    page = doc[0]
    scale = (float(width) / page.rect.width) if width else 1.0
    pix = page.get_pixmap(matrix=fitz.Matrix(scale, scale), alpha=False)
    rgb = np.frombuffer(pix.samples, dtype=np.uint8).reshape(pix.height, pix.width, pix.n)
    if pix.n == 4:
        rgb = cv2.cvtColor(rgb, cv2.COLOR_RGBA2RGB)
    return cv2.cvtColor(rgb, cv2.COLOR_RGB2BGR)


def render_template_pdf_to_canvas(template_pdf: str, *, width: int, height: int) -> np.ndarray:
    doc = fitz.open(template_pdf)
    if doc.page_count < 1:
        raise ValueError(f"template pdf has no pages: {template_pdf}")
    page = doc[0]
    scale = float(width) / page.rect.width
    pix = page.get_pixmap(matrix=fitz.Matrix(scale, scale), alpha=False)
    rgb = np.frombuffer(pix.samples, dtype=np.uint8).reshape(pix.height, pix.width, pix.n)
    if pix.n == 4:
        rgb = cv2.cvtColor(rgb, cv2.COLOR_RGBA2RGB)
    image = cv2.cvtColor(rgb, cv2.COLOR_RGB2BGR)
    if image.shape[:2] != (height, width):
        image = cv2.resize(image, (width, height), interpolation=cv2.INTER_AREA)
    return image


def cluster_projection(values: np.ndarray, threshold: float) -> list[tuple[int, int]]:
    indexes = np.where(values > threshold)[0]
    if len(indexes) == 0:
        return []
    clusters: list[tuple[int, int]] = []
    start = prev = int(indexes[0])
    for value in indexes[1:]:
        current = int(value)
        if current > prev + 1:
            clusters.append((start, prev))
            start = current
        prev = current
    clusters.append((start, prev))
    return clusters


def _fill_missing_table_y_edges_by_gap(
    *,
    strong_ys: list[int],
    weak_ys: list[int],
    table_y0: float,
    expected_count: int,
) -> list[int]:
    table_ys = sorted(y for y in strong_ys if y >= table_y0 - 5)
    if len(table_ys) >= expected_count:
        return table_ys[:expected_count]

    missing_count = expected_count - len(table_ys)
    if missing_count <= 0:
        return table_ys

    existing = set(table_ys)
    weak_candidates = sorted(
        y
        for y in weak_ys
        if y >= table_y0 - 5 and all(abs(y - existing_y) > 3 for existing_y in existing)
    )
    if not weak_candidates:
        return table_ys

    gaps = np.diff(np.array(table_ys, dtype=np.float32)) if len(table_ys) >= 2 else np.array([], dtype=np.float32)
    positive_gaps = gaps[gaps > 0]
    if len(positive_gaps) == 0:
        return table_ys
    typical_gap = float(np.median(positive_gaps))

    scored: list[tuple[float, int]] = []
    for y in weak_candidates:
        insert_at = int(np.searchsorted(np.array(table_ys, dtype=np.int32), y))
        if insert_at <= 0 or insert_at >= len(table_ys):
            continue
        prev_y = table_ys[insert_at - 1]
        next_y = table_ys[insert_at]
        original_gap = float(next_y - prev_y)
        left_gap = float(y - prev_y)
        right_gap = float(next_y - y)
        if left_gap <= 0 or right_gap <= 0:
            continue
        # A partial header ruling should explain an unusually large gap by
        # splitting it into two normal-looking neighboring row gaps.
        if original_gap < typical_gap * 1.55:
            continue
        split_score = abs(left_gap - typical_gap) + abs(right_gap - typical_gap)
        split_score += abs((left_gap + right_gap) - original_gap) * 0.01
        scored.append((float(split_score), int(y)))

    selected = [y for _score, y in sorted(scored)[:missing_count]]
    return sorted(table_ys + selected)


def extract_template_axes_from_image(
    template_image: np.ndarray,
    *,
    manifest_template_bbox: list[float],
) -> tuple[list[int], list[int], list[int], list[int]]:
    height, width = template_image.shape[:2]
    gray = cv2.cvtColor(template_image, cv2.COLOR_BGR2GRAY)
    binary = cv2.threshold(gray, 180, 255, cv2.THRESH_BINARY_INV)[1]
    vertical = cv2.morphologyEx(
        binary,
        cv2.MORPH_OPEN,
        cv2.getStructuringElement(cv2.MORPH_RECT, (1, 80)),
    )
    horizontal = cv2.morphologyEx(
        binary,
        cv2.MORPH_OPEN,
        cv2.getStructuringElement(cv2.MORPH_RECT, (80, 1)),
    )
    vertical_clusters = cluster_projection(vertical.sum(axis=0) / 255, height * 0.25)
    horizontal_clusters = cluster_projection(horizontal.sum(axis=1) / 255, width * 0.25)
    xs = [int(round((start + end) / 2)) for start, end in vertical_clusters]
    strong_ys = [int(round((start + end) / 2)) for start, end in horizontal_clusters]
    if not xs or not strong_ys:
        raise ValueError("template axes not found")

    weak_horizontal_clusters = cluster_projection(
        horizontal.sum(axis=1) / 255,
        max(50.0, width * TWO_STAGE_HEADER_BOUNDARY_MIN_WIDTH_RATIO),
    )
    weak_ys = [int(round((start + end) / 2)) for start, end in weak_horizontal_clusters]
    ys_all = sorted(set(strong_ys + weak_ys))

    manifest_y0 = float(manifest_template_bbox[1])
    ys = _fill_missing_table_y_edges_by_gap(
        strong_ys=strong_ys,
        weak_ys=weak_ys,
        table_y0=manifest_y0,
        expected_count=ORDER_FORM_TEMPLATE_Y_EDGE_COUNT,
    )
    if not ys:
        raise ValueError("template table y axes not found")
    if len(ys) != ORDER_FORM_TEMPLATE_Y_EDGE_COUNT:
        raise ValueError(
            "template table y axes incomplete: "
            f"expected={ORDER_FORM_TEMPLATE_Y_EDGE_COUNT} actual={len(ys)}"
        )
    return xs, ys, [int(v) for v in xs], [int(v) for v in ys_all]


def _normalize_explicit_template_axes(
    template_axes_x: list[Any] | None,
    template_axes_y: list[Any] | None,
) -> tuple[list[int], list[int]]:
    if template_axes_x is None or template_axes_y is None:
        raise ValueError("canonical template axes require both x and y")
    try:
        xs = sorted({int(round(float(value))) for value in template_axes_x})
        ys = [int(round(float(value))) for value in template_axes_y]
    except Exception as exc:  # noqa: BLE001
        raise ValueError(f"canonical template axes invalid:{exc}") from exc
    if len(xs) < 2:
        raise ValueError("canonical template x axes incomplete")
    if len(ys) != ORDER_FORM_TEMPLATE_Y_EDGE_COUNT:
        raise ValueError(
            "canonical template y axes incomplete: "
            f"expected={ORDER_FORM_TEMPLATE_Y_EDGE_COUNT} actual={len(ys)}"
        )
    if any(ys[index] >= ys[index + 1] for index in range(len(ys) - 1)):
        raise ValueError("canonical template y axes are not strictly increasing")
    return xs, ys


def resolve_template_axes_from_manifest_or_image(
    *,
    item: dict[str, Any],
    template_image: np.ndarray,
    manifest_template_bbox: list[float],
) -> tuple[list[int], list[int], list[int], list[int]]:
    has_x = "template_axes_x" in item
    has_y = "template_axes_y" in item
    if has_x or has_y:
        xs, ys = _normalize_explicit_template_axes(
            item.get("template_axes_x") if isinstance(item, dict) else None,
            item.get("template_axes_y") if isinstance(item, dict) else None,
        )
        return xs, ys, list(xs), list(ys)
    return extract_template_axes_from_image(
        template_image,
        manifest_template_bbox=manifest_template_bbox,
    )


def canonical_template_axes_from_workbook(
    workbook_path: Path | str,
    *,
    sheet_name: str,
    canvas_width: int,
    canvas_height: int,
    table_bbox: list[float] | None = None,
) -> tuple[list[int], list[int]]:
    workbook = load_workbook(Path(workbook_path))
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"template axes sheet not found: {sheet_name}")
    geometry = workbook_pdf_renderer.worksheet_render_geometry(workbook[sheet_name])
    image_width = float(geometry["image_width"])
    image_height = float(geometry["image_height"])
    if image_width <= 0 or image_height <= 0:
        raise ValueError("template axes geometry invalid")
    x_scale = float(canvas_width) / image_width
    y_scale = float(canvas_height) / image_height
    raw_x_positions = sorted(int(value) for value in geometry["x_positions"].values())
    xs = sorted({int(round(float(value) * x_scale)) for value in raw_x_positions})
    if table_bbox is not None and len(table_bbox) == 4:
        x0 = float(table_bbox[0])
        x1 = float(table_bbox[2])
        tolerance = max(8.0, float(canvas_width) * 0.01)
        xs = [x for x in xs if x0 - tolerance <= float(x) <= x1 + tolerance]
    y_positions = geometry["y_positions"]
    edge_rows = list(ORDER_FORM_TEMPLATE_HEADER_EDGE_ROWS) + list(ORDER_FORM_TEMPLATE_BODY_EDGE_ROWS)
    missing_rows = [row for row in edge_rows if row not in y_positions]
    if missing_rows:
        raise ValueError(f"canonical template y edge rows missing:{missing_rows}")
    ys = [int(round(float(y_positions[row]) * y_scale)) for row in edge_rows]
    return _normalize_explicit_template_axes(xs, ys)


def draw_fixed_quad_overlay(original: np.ndarray, quad: np.ndarray) -> np.ndarray:
    image = original.copy()
    cv2.polylines(image, [quad.astype(np.int32)], True, (0, 0, 255), 6, cv2.LINE_AA)
    labels = ["TL", "TR", "BR", "BL"]
    colors = [(0, 0, 255), (255, 0, 0), (0, 160, 0), (0, 140, 255)]
    for point, label, color in zip(quad, labels, colors):
        x, y = map(int, point)
        cv2.circle(image, (x, y), 12, color, -1, cv2.LINE_AA)
        cv2.putText(
            image,
            label,
            (x + 10, y + 32),
            cv2.FONT_HERSHEY_SIMPLEX,
            0.9,
            color,
            3,
            cv2.LINE_AA,
        )
    cv2.putText(
        image,
        "STEP1 original FAX + fixed accepted 4 points",
        (30, 50),
        cv2.FONT_HERSHEY_SIMPLEX,
        1.1,
        (0, 0, 255),
        3,
        cv2.LINE_AA,
    )
    return image


def draw_rectified_fax_overlay(rectified: np.ndarray, table_bbox: list[int]) -> np.ndarray:
    x0, y0, x1, y1 = table_bbox
    image = rectified.copy()
    cv2.rectangle(image, (x0, y0), (x1, y1), (0, 0, 255), 4, cv2.LINE_AA)
    cv2.putText(
        image,
        "STEP2 rectified FAX by accepted 4 points -> actual template outer grid",
        (30, 60),
        cv2.FONT_HERSHEY_SIMPLEX,
        1.0,
        (0, 0, 255),
        3,
        cv2.LINE_AA,
    )
    return image


def draw_template_grid_overlay(rectified: np.ndarray, xs: list[int], ys: list[int]) -> np.ndarray:
    x0, y0, x1, y1 = [xs[0], ys[0], xs[-1], ys[-1]]
    image = rectified.copy()
    for x in xs:
        if x0 <= x <= x1:
            cv2.line(image, (x, y0), (x, y1), (0, 255, 0), 2, cv2.LINE_AA)
    for y in ys:
        if y0 <= y <= y1:
            cv2.line(image, (x0, y), (x1, y), (0, 255, 0), 2, cv2.LINE_AA)
    cv2.rectangle(image, (x0, y0), (x1, y1), (0, 180, 0), 3, cv2.LINE_AA)
    cv2.putText(
        image,
        "STEP3 rectified FAX + template ruling grid (uniform green)",
        (30, 60),
        cv2.FONT_HERSHEY_SIMPLEX,
        1.0,
        (0, 160, 0),
        3,
        cv2.LINE_AA,
    )
    return image


def rectify_fax_to_template_grid(
    original: np.ndarray,
    *,
    quad_px: list[list[float]],
    table_bbox: list[int],
    canvas_width: int,
    canvas_height: int,
) -> np.ndarray:
    quad = np.array(quad_px, dtype=np.float32)
    dst = np.array(
        [
            [table_bbox[0], table_bbox[1]],
            [table_bbox[2], table_bbox[1]],
            [table_bbox[2], table_bbox[3]],
            [table_bbox[0], table_bbox[3]],
        ],
        dtype=np.float32,
    )
    transform = cv2.getPerspectiveTransform(quad, dst)
    return cv2.warpPerspective(
        original,
        transform,
        (canvas_width, canvas_height),
        flags=cv2.INTER_CUBIC,
        borderMode=cv2.BORDER_CONSTANT,
        borderValue=(255, 255, 255),
    )


def write_fixed_quad_registration_outputs(
    *,
    output_dir: Path,
    facility_code: str,
    order_id: str,
    step1: np.ndarray,
    step2: np.ndarray,
    step3: np.ndarray,
    result: FixedQuadTemplateRegistrationResult,
) -> FixedQuadTemplateRegistrationResult:
    output_dir.mkdir(parents=True, exist_ok=True)
    paths = {
        "step1": output_dir / "01_original_fax_fixed_4points.png",
        "step2": output_dir / "02_rectified_fax_by_fixed_4points.png",
        "step3": output_dir / "03_rectified_fax_template_grid_overlay.png",
    }
    cv2.imwrite(str(paths["step1"]), step1)
    cv2.imwrite(str(paths["step2"]), step2)
    cv2.imwrite(str(paths["step3"]), step3)

    pdf_path = output_dir / f"{facility_code}_{order_id}_original_quad_rectify_template_overlay.pdf"
    pdf = fitz.open()
    for key in ("step1", "step2", "step3"):
        image = cv2.imread(str(paths[key]))
        height, width = image.shape[:2]
        page = pdf.new_page(width=width, height=height)
        page.insert_image(fitz.Rect(0, 0, width, height), filename=str(paths[key]))
    pdf.save(str(pdf_path))

    updated = FixedQuadTemplateRegistrationResult(
        **{
            **asdict(result),
            "outputs": {key: str(value) for key, value in paths.items()} | {"pdf": str(pdf_path)},
        }
    )
    summary_path = output_dir / "summary.json"
    with summary_path.open("w") as f:
        json.dump(asdict(updated), f, ensure_ascii=False, indent=2)
    return updated


def build_fixed_quad_template_registration(
    *,
    facility_code: str,
    order_id: str,
    fax_pdf: str,
    template_pdf: str,
    quad_px: list[list[float]] | None,
    manifest_template_bbox: list[float],
    canvas_width: int,
    canvas_height: int,
    render_width: int,
    quad_source: str | None = None,
    output_dir: Path | None = None,
    template_axes_x: list[int] | None = None,
    template_axes_y: list[int] | None = None,
) -> tuple[FixedQuadTemplateRegistrationResult, dict[str, np.ndarray]]:
    if quad_px is None:
        estimate = estimate_edge_locked_quad_from_pdf(fax_pdf, dpi=220)
        if estimate["status"] != "ok":
            raise ValueError(f"edge locked quad estimation failed: {estimate['reasons']}")
        quad_px = estimate["refined_quad_px"]
        if quad_source is None:
            quad_source = "edge_locked_v4_estimated_from_fax_pdf"
    original = render_pdf_page_to_bgr(fax_pdf, width=render_width)
    template = render_template_pdf_to_canvas(
        template_pdf,
        width=canvas_width,
        height=canvas_height,
    )
    if template_axes_x is not None or template_axes_y is not None:
        xs, ys = _normalize_explicit_template_axes(template_axes_x, template_axes_y)
        ys_all = list(ys)
    else:
        xs, ys, _xs_all, ys_all = extract_template_axes_from_image(
            template,
            manifest_template_bbox=manifest_template_bbox,
        )
    table_bbox = [xs[0], ys[0], xs[-1], ys[-1]]
    rectified = rectify_fax_to_template_grid(
        original,
        quad_px=quad_px,
        table_bbox=table_bbox,
        canvas_width=canvas_width,
        canvas_height=canvas_height,
    )

    quad = np.array(quad_px, dtype=np.float32)
    step_images = {
        "step1": draw_fixed_quad_overlay(original, quad),
        "step2": draw_rectified_fax_overlay(rectified, table_bbox),
        "step3": draw_template_grid_overlay(rectified, xs, ys),
    }
    result = FixedQuadTemplateRegistrationResult(
        facility_code=facility_code,
        order_id=order_id,
        fax_pdf=fax_pdf,
        template_pdf=template_pdf,
        quad_source=quad_source,
        quad_px=quad_px,
        rectified_canvas_size=[canvas_width, canvas_height],
        template_axes_x=xs,
        template_axes_y_used_count=len(ys),
        template_axes_y_used_first_last=[ys[0], ys[-1]],
        template_axes_y_all=ys_all,
        template_outer_grid_bbox_used=table_bbox,
        legacy_manifest_template_bbox_not_used=manifest_template_bbox,
        forbidden_downstream_methods=FORBIDDEN_DOWNSTREAM_METHODS,
        outputs={"step1": None, "step2": None, "step3": None, "pdf": None},
    )
    if output_dir is not None:
        result = write_fixed_quad_registration_outputs(
            output_dir=output_dir,
            facility_code=facility_code,
            order_id=order_id,
            step1=step_images["step1"],
            step2=step_images["step2"],
            step3=step_images["step3"],
            result=result,
        )
    return result, step_images


def replay_fixed_quad_template_overlay(
    *,
    manifest_path: Path,
    facility_code: str,
    order_id: str,
    output_dir: Path,
    render_width: int,
) -> dict[str, Any]:
    item = load_fixed_quad_manifest_item(
        manifest_path,
        facility_code=facility_code,
        order_id=order_id,
    )

    existing_step2 = cv2.imread(item["step2_png"])
    if existing_step2 is None:
        raise ValueError(f"step2 canvas not found: {item['step2_png']}")
    canvas_height, canvas_width = existing_step2.shape[:2]

    result, _images = build_fixed_quad_template_registration(
        facility_code=facility_code,
        order_id=order_id,
        fax_pdf=item["fax_pdf"],
        template_pdf=item["template_pdf"],
        quad_px=item["quad_px"],
        manifest_template_bbox=item["template_bbox"],
        canvas_width=canvas_width,
        canvas_height=canvas_height,
        render_width=render_width,
        quad_source=item.get("quad_source"),
        output_dir=output_dir,
    )
    return asdict(result)
