from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter, range_boundaries
from PIL import Image, ImageColor, ImageDraw, ImageFont


_DEFAULT_DPI = 144
_DEFAULT_MARGIN_PX = 12
_EMU_PER_PIXEL = 9525
_DEFAULT_FONT_PATHS = [
    "/System/Library/Fonts/ヒラギノ角ゴシック W3.ttc",
    "/Library/Fonts/Arial Unicode.ttf",
    "/System/Library/Fonts/Supplemental/Arial Unicode.ttf",
    "/System/Library/Fonts/Helvetica.ttc",
]


def _sheet_print_range(worksheet) -> tuple[int, int, int, int]:
    raw = str(worksheet.print_area or "").strip()
    if raw:
        area_ref = raw.split("!", 1)[-1].replace("$", "").split(",", 1)[0]
        min_col, min_row, max_col, max_row = range_boundaries(area_ref)
        return min_col, min_row, max_col, max_row
    return 1, 1, max(worksheet.max_column, 1), max(worksheet.max_row, 1)


def _column_width_to_pixels(width: float | None, *, dpi: int) -> int:
    excel_width = float(width if width is not None else 8.43)
    base_pixels = max(8.0, excel_width * 7.0 + 5.0)
    return max(8, int(round(base_pixels * float(dpi) / 96.0)))


def _row_height_to_pixels(height_points: float | None, *, dpi: int) -> int:
    points = float(height_points if height_points is not None else 15.0)
    return max(12, int(round(points * float(dpi) / 72.0)))


def _line_width(side_style: str | None) -> int:
    style = str(side_style or "").strip().lower()
    if not style:
        return 0
    if style in {"hair"}:
        return 1
    if style in {"medium", "mediumdashed", "mediumdashdot", "mediumdashdotdot"}:
        return 2
    if style in {"thick", "double"}:
        return 3
    return 1


def _line_style_rank(side_style: str | None) -> int:
    style = str(side_style or "").strip().lower()
    if not style:
        return 0
    if style in {"hair"}:
        return 1
    if style in {"dotted", "dashdotdot", "slantdashdot"}:
        return 2
    if style in {"dashed", "dashdot"}:
        return 3
    if style in {"thin"}:
        return 4
    if style in {"medium", "mediumdashed", "mediumdashdot", "mediumdashdotdot"}:
        return 5
    if style in {"thick", "double"}:
        return 6
    return 4


def _color_to_rgb(color, *, default: tuple[int, int, int]) -> tuple[int, int, int]:
    value = getattr(color, "rgb", None) or getattr(color, "value", None)
    if isinstance(value, str):
        token = value.strip()
        if len(token) == 8:
            token = token[2:]
        if len(token) == 6:
            try:
                return ImageColor.getrgb(f"#{token}")
            except ValueError:
                pass
    return default


def _border_priority(side) -> tuple[int, int]:
    style = str(getattr(side, "style", None) or "").strip().lower()
    return _line_style_rank(style), _line_width(style)


def _upsert_border_segment(
    segments: dict[tuple[str, int, int, int], dict],
    *,
    key: tuple[str, int, int, int],
    side,
) -> None:
    style = str(getattr(side, "style", None) or "").strip().lower()
    width = _line_width(style)
    if width <= 0:
        return
    candidate = {
        "style": style,
        "width": width,
        "color": _color_to_rgb(getattr(side, "color", None), default=(0, 0, 0)),
        "priority": _border_priority(side),
    }
    current = segments.get(key)
    if current is None or candidate["priority"] > current["priority"]:
        segments[key] = candidate


def _collect_border_segments(
    worksheet,
    *,
    min_col: int,
    min_row: int,
    max_col: int,
    max_row: int,
    x_positions: dict[int, int],
    y_positions: dict[int, int],
) -> dict[tuple[str, int, int, int], dict]:
    segments: dict[tuple[str, int, int, int], dict] = {}
    merged = _merged_rectangles(
        worksheet,
        min_col=min_col,
        min_row=min_row,
        max_col=max_col,
        max_row=max_row,
    )
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            cell = worksheet.cell(row=row, column=col)
            border = cell.border
            x0 = x_positions[col]
            x1 = x_positions[col + 1]
            y0 = y_positions[row]
            y1 = y_positions[row + 1]
            merge_rect = merged.get((row, col))
            if not merge_rect or col == merge_rect["start_col"]:
                _upsert_border_segment(
                    segments,
                    key=("v", x0, y0, y1),
                    side=border.left,
                )
            if not merge_rect or col == merge_rect["end_col"]:
                _upsert_border_segment(
                    segments,
                    key=("v", x1, y0, y1),
                    side=border.right,
                )
            if not merge_rect or row == merge_rect["start_row"]:
                _upsert_border_segment(
                    segments,
                    key=("h", y0, x0, x1),
                    side=border.top,
                )
            if not merge_rect or row == merge_rect["end_row"]:
                _upsert_border_segment(
                    segments,
                    key=("h", y1, x0, x1),
                    side=border.bottom,
                )
    return segments


def _draw_styled_line(
    draw: ImageDraw.ImageDraw,
    *,
    start: tuple[int, int],
    end: tuple[int, int],
    style: str,
    fill: tuple[int, int, int],
    width: int,
) -> None:
    normalized = str(style or "").strip().lower()
    if normalized == "dotted":
        x0, y0 = start
        x1, y1 = end
        if x0 == x1:
            step = max(width * 3, 4)
            dot = max(width, 1)
            for y in range(min(y0, y1), max(y0, y1), step):
                draw.line([(x0, y), (x0, min(y + dot, max(y0, y1)))], fill=fill, width=width)
            return
        if y0 == y1:
            step = max(width * 3, 4)
            dot = max(width, 1)
            for x in range(min(x0, x1), max(x0, x1), step):
                draw.line([(x, y0), (min(x + dot, max(x0, x1)), y0)], fill=fill, width=width)
            return
    draw.line([start, end], fill=fill, width=width)


def _draw_border_segments(
    draw: ImageDraw.ImageDraw,
    segments: dict[tuple[str, int, int, int], dict],
) -> None:
    for key, segment in segments.items():
        orientation, fixed, start, end = key
        width = int(segment["width"])
        style = str(segment["style"])
        color = segment["color"]
        if orientation == "v":
            _draw_styled_line(
                draw,
                start=(fixed, start),
                end=(fixed, end),
                style=style,
                fill=color,
                width=width,
            )
        else:
            _draw_styled_line(
                draw,
                start=(start, fixed),
                end=(end, fixed),
                style=style,
                fill=color,
                width=width,
            )


def _load_font(size_px: int, *, bold: bool = False) -> ImageFont.FreeTypeFont | ImageFont.ImageFont:
    requested = max(10, int(size_px))
    for path in _DEFAULT_FONT_PATHS:
        if Path(path).exists():
            try:
                return ImageFont.truetype(path, requested)
            except OSError:
                continue
    return ImageFont.load_default()


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return f"{value.month}/{value.day}"
    if isinstance(value, date):
        return f"{value.month}/{value.day}"
    return str(value)


def _text_bbox(draw: ImageDraw.ImageDraw, text: str, font: ImageFont.ImageFont) -> tuple[int, int]:
    if not text:
        return 0, 0
    bbox = draw.multiline_textbbox((0, 0), text, font=font, spacing=2)
    return int(bbox[2] - bbox[0]), int(bbox[3] - bbox[1])


def _draw_cell_text(
    draw: ImageDraw.ImageDraw,
    *,
    rect: tuple[int, int, int, int],
    text: str,
    font: ImageFont.ImageFont,
    font_color: tuple[int, int, int],
    horizontal: str,
    vertical: str,
) -> None:
    if not text:
        return
    x0, y0, x1, y1 = rect
    text_w, text_h = _text_bbox(draw, text, font)
    if horizontal == "center":
        text_x = x0 + max(4, ((x1 - x0) - text_w) // 2)
    elif horizontal == "right":
        text_x = max(x0 + 4, x1 - text_w - 4)
    else:
        text_x = x0 + 4
    if vertical == "top":
        text_y = y0 + 3
    elif vertical == "bottom":
        text_y = max(y0 + 3, y1 - text_h - 3)
    else:
        text_y = y0 + max(3, ((y1 - y0) - text_h) // 2)
    draw.multiline_text((text_x, text_y), text, fill=font_color, font=font, spacing=2)


def _merged_rectangles(
    worksheet,
    *,
    min_col: int,
    min_row: int,
    max_col: int,
    max_row: int,
) -> dict[tuple[int, int], dict[str, int]]:
    merged: dict[tuple[int, int], dict[str, int]] = {}
    for merged_range in worksheet.merged_cells.ranges:
        start_col, start_row, end_col, end_row = range_boundaries(str(merged_range))
        if end_col < min_col or start_col > max_col or end_row < min_row or start_row > max_row:
            continue
        clipped_start_col = max(start_col, min_col)
        clipped_start_row = max(start_row, min_row)
        clipped_end_col = min(end_col, max_col)
        clipped_end_row = min(end_row, max_row)
        for row in range(clipped_start_row, clipped_end_row + 1):
            for col in range(clipped_start_col, clipped_end_col + 1):
                merged[(row, col)] = {
                    "anchor_row": start_row,
                    "anchor_col": start_col,
                    "start_row": clipped_start_row,
                    "start_col": clipped_start_col,
                    "end_row": clipped_end_row,
                    "end_col": clipped_end_col,
                }
    return merged


def worksheet_render_geometry(
    worksheet,
    *,
    dpi: int = _DEFAULT_DPI,
    margin_px: int = _DEFAULT_MARGIN_PX,
) -> dict:
    min_col, min_row, max_col, max_row = _sheet_print_range(worksheet)
    col_widths = {}
    row_heights = {}
    for col in range(min_col, max_col + 1):
        dim = worksheet.column_dimensions[get_column_letter(col)]
        col_widths[col] = _column_width_to_pixels(dim.width, dpi=dpi)
    default_row_height = worksheet.sheet_format.defaultRowHeight
    for row in range(min_row, max_row + 1):
        dim = worksheet.row_dimensions[row]
        row_heights[row] = _row_height_to_pixels(dim.height or default_row_height, dpi=dpi)

    x_positions = {min_col: margin_px}
    current_x = margin_px
    for col in range(min_col, max_col + 1):
        current_x += col_widths[col]
        x_positions[col + 1] = current_x
    y_positions = {min_row: margin_px}
    current_y = margin_px
    for row in range(min_row, max_row + 1):
        current_y += row_heights[row]
        y_positions[row + 1] = current_y

    return {
        "print_range": [min_col, min_row, max_col, max_row],
        "x_positions": x_positions,
        "y_positions": y_positions,
        "image_width": current_x + margin_px,
        "image_height": current_y + margin_px,
    }


def render_worksheet_to_image(
    worksheet,
    *,
    dpi: int = _DEFAULT_DPI,
    margin_px: int = _DEFAULT_MARGIN_PX,
) -> Image.Image:
    geometry = worksheet_render_geometry(worksheet, dpi=dpi, margin_px=margin_px)
    min_col, min_row, max_col, max_row = [int(value) for value in geometry["print_range"]]
    x_positions = geometry["x_positions"]
    y_positions = geometry["y_positions"]
    image_width = int(geometry["image_width"])
    image_height = int(geometry["image_height"])
    canvas = Image.new("RGB", (image_width, image_height), "white")
    draw = ImageDraw.Draw(canvas)

    merged = _merged_rectangles(
        worksheet,
        min_col=min_col,
        min_row=min_row,
        max_col=max_col,
        max_row=max_row,
    )

    text_jobs: list[tuple[tuple[int, int, int, int], object]] = []
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            merge_rect = merged.get((row, col))
            if merge_rect:
                if (row, col) != (merge_rect["start_row"], merge_rect["start_col"]):
                    continue
                cell = worksheet.cell(row=merge_rect["anchor_row"], column=merge_rect["anchor_col"])
                end_row = merge_rect["end_row"]
                end_col = merge_rect["end_col"]
            else:
                cell = worksheet.cell(row=row, column=col)
                if isinstance(cell, MergedCell):
                    continue
                end_row = row
                end_col = col
            rect = (
                x_positions[col],
                y_positions[row],
                x_positions[end_col + 1],
                y_positions[end_row + 1],
            )
            fill = cell.fill
            if getattr(fill, "fill_type", None) == "solid":
                fill_rgb = _color_to_rgb(fill.fgColor, default=(255, 255, 255))
                draw.rectangle(rect, fill=fill_rgb)
            text_jobs.append((rect, cell))

    _draw_border_segments(
        draw,
        _collect_border_segments(
            worksheet,
            min_col=min_col,
            min_row=min_row,
            max_col=max_col,
            max_row=max_row,
            x_positions=x_positions,
            y_positions=y_positions,
        ),
    )

    for rect, cell in text_jobs:
        text = _cell_text(cell.value).strip()
        if not text:
            continue
        font_size_px = int(round(float(cell.font.sz or 11.0) * float(dpi) / 72.0))
        font = _load_font(font_size_px, bold=bool(getattr(cell.font, "bold", False)))
        font_color = _color_to_rgb(getattr(cell.font, "color", None), default=(0, 0, 0))
        alignment = cell.alignment
        horizontal = str(getattr(alignment, "horizontal", "") or "left").strip().lower()
        vertical = str(getattr(alignment, "vertical", "") or "center").strip().lower()
        _draw_cell_text(
            draw,
            rect=rect,
            text=text,
            font=font,
            font_color=font_color,
            horizontal=horizontal,
            vertical=vertical,
        )

    for image in getattr(worksheet, "_images", []):
        anchor = getattr(image, "anchor", None)
        start = getattr(anchor, "_from", None)
        if start is None:
            continue
        start_col = int(getattr(start, "col", 0)) + 1
        start_row = int(getattr(start, "row", 0)) + 1
        if start_col < min_col or start_row < min_row or start_col > (max_col + 1) or start_row > (max_row + 1):
            continue
        try:
            blob = image._data()
        except Exception:
            continue
        try:
            pasted = Image.open(BytesIO(blob)).convert("RGBA")
        except Exception:
            continue
        offset_x = int(round(float(getattr(start, "colOff", 0) or 0) / float(_EMU_PER_PIXEL) * float(dpi) / 96.0))
        offset_y = int(round(float(getattr(start, "rowOff", 0) or 0) / float(_EMU_PER_PIXEL) * float(dpi) / 96.0))
        x = x_positions.get(start_col, margin_px) + offset_x
        y = y_positions.get(start_row, margin_px) + offset_y
        target_w = max(1, int(round(float(getattr(image, "width", pasted.width)) * float(dpi) / 96.0)))
        target_h = max(1, int(round(float(getattr(image, "height", pasted.height)) * float(dpi) / 96.0)))
        end = getattr(anchor, "to", None)
        if end is not None:
            end_col = int(getattr(end, "col", 0)) + 1
            end_row = int(getattr(end, "row", 0)) + 1
            if end_col in x_positions and end_row in y_positions:
                end_offset_x = int(round(float(getattr(end, "colOff", 0) or 0) / float(_EMU_PER_PIXEL) * float(dpi) / 96.0))
                end_offset_y = int(round(float(getattr(end, "rowOff", 0) or 0) / float(_EMU_PER_PIXEL) * float(dpi) / 96.0))
                target_w = max(1, x_positions[end_col] + end_offset_x - x)
                target_h = max(1, y_positions[end_row] + end_offset_y - y)
        if (target_w, target_h) != pasted.size:
            pasted = pasted.resize((target_w, target_h))
        canvas.paste(pasted, (x, y), pasted)

    return canvas


def render_workbook_to_pdf_bytes(
    workbook,
    *,
    sheet_name: str | None = None,
    dpi: int = _DEFAULT_DPI,
) -> bytes:
    target_sheet = sheet_name or workbook.sheetnames[0]
    image = render_worksheet_to_image(workbook[target_sheet], dpi=dpi)
    buffer = BytesIO()
    image.save(buffer, format="PDF", resolution=float(dpi))
    return buffer.getvalue()


def render_workbook_path_to_pdf(
    workbook_path: Path | str,
    *,
    output_path: Path | str,
    sheet_name: str | None = None,
    dpi: int = _DEFAULT_DPI,
) -> Path:
    workbook = load_workbook(Path(workbook_path))
    pdf_bytes = render_workbook_to_pdf_bytes(workbook, sheet_name=sheet_name, dpi=dpi)
    target = Path(output_path)
    target.parent.mkdir(parents=True, exist_ok=True)
    target.write_bytes(pdf_bytes)
    return target


def render_workbooks_to_pdf_bytes(
    workbooks: Iterable,
    *,
    sheet_name: str | None = None,
    dpi: int = _DEFAULT_DPI,
) -> bytes:
    images = [
        render_worksheet_to_image(workbook[sheet_name or workbook.sheetnames[0]], dpi=dpi).convert("RGB")
        for workbook in workbooks
    ]
    if not images:
        return b""
    buffer = BytesIO()
    images[0].save(buffer, format="PDF", save_all=True, append_images=images[1:], resolution=float(dpi))
    return buffer.getvalue()
