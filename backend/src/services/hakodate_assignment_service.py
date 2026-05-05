from __future__ import annotations

from dataclasses import asdict, dataclass
from copy import deepcopy
from datetime import date, datetime
import hashlib
from io import BytesIO
import json
import os
import re
import unicodedata
from typing import Any

from loguru import logger
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils import range_boundaries

from src.services.grid_detector import GridDetectionResult, detect_table_grid
from src.services import order_form_service
from src.services.template_field_schema_service import (
    canonical_field_name_from_template_column,
    derive_row_fields_from_template,
)
from src.services import workbook_pdf_renderer
from src.services.workbook_pdf_renderer import render_workbook_path_to_pdf


_VALID_STRATEGIES = {"legacy", "hakodate", "both"}
_NUMERIC_RE = re.compile(r"^\d+(?:\.\d+)?$")
_SIGNATURE_KEYS = ("hakodate_template_signature", "template_signature")
_SIGNATURE_COMPONENT_KEYS = (
    "hakodate_template_signature_components",
    "template_signature_components",
)
_STRUCTURE_BODY_START_ROW = 11
_STRUCTURE_BODY_END_ROW = 67
_STRUCTURE_HEADER_ROWS = (7, 8, 9)


class HakodateOcrUnavailable(RuntimeError):
    pass


@dataclass(frozen=True)
class HakodateToken:
    text: str
    x: float
    y: float
    bbox: list[float]
    confidence: float | None = None


@dataclass(frozen=True)
class HakodateCellOcrResult:
    text: str
    normalized: str
    confidence: float | None = None
    crop_bbox_px: list[int] | None = None
    crop_image_path: str | None = None


def normalize_quantity_assignment_strategy(value: object) -> str:
    normalized = str(value or "").strip().lower()
    if normalized in _VALID_STRATEGIES:
        return normalized
    return "legacy"


def resolve_quantity_assignment_strategy(template: dict[str, Any] | None = None) -> str:
    if isinstance(template, dict):
        configured = normalize_quantity_assignment_strategy(
            template.get("quantity_assignment_strategy")
        )
        if configured != "legacy" or "quantity_assignment_strategy" in template:
            return configured
    return normalize_quantity_assignment_strategy(
        os.getenv("OCR_QUANTITY_ASSIGNMENT_STRATEGY", "legacy")
    )


def _normalize_digits(value: object) -> str:
    text = unicodedata.normalize("NFKC", str(value or "")).strip()
    text = text.replace(",", "")
    text = re.sub(r"\s+", "", text)
    return text


def _is_numeric_text(value: object) -> bool:
    return bool(_NUMERIC_RE.fullmatch(_normalize_digits(value)))


def _normalize_slot_text(value: object) -> str:
    if isinstance(value, datetime):
        return f"{value.month}/{value.day}"
    if isinstance(value, date):
        return f"{value.month}/{value.day}"
    return unicodedata.normalize("NFKC", str(value or "")).strip()


def _normalize_menu_key(value: object) -> str:
    text = _normalize_slot_text(value)
    text = re.sub(r"\s+", "", text)
    text = text.replace("（", "(").replace("）", ")")
    text = text.replace("①", "1").replace("②", "2")
    return text


def _is_weekday_only(value: object) -> bool:
    text = _normalize_slot_text(value)
    return bool(text and re.fullmatch(r"[()（）月火水木金土日]+", text))


def _quantity_field_indexes(fields: list[str]) -> list[int]:
    return [
        idx
        for idx, field in enumerate(fields)
        if str(field or "").strip().startswith("qty.")
    ]


def _quantity_fields(fields: list[str]) -> list[str]:
    return [field for field in fields if str(field or "").strip().startswith("qty.")]


def _normalized_template_columns(template: dict[str, Any]) -> list[dict[str, Any]]:
    columns = template.get("columns")
    if not isinstance(columns, list):
        return []
    normalized: list[dict[str, Any]] = []
    for fallback_index, column in enumerate(columns):
        if not isinstance(column, dict):
            continue
        try:
            index = int(column.get("index"))
        except Exception:
            index = fallback_index
        normalized.append(
            {
                "index": index,
                "role": str(column.get("role") or "").strip().lower(),
                "field": str(
                    column.get("name")
                    or column.get("field")
                    or column.get("header")
                    or ""
                ).strip(),
                "diet_type": str(column.get("diet_type") or "").strip().lower(),
                "area_id": str(column.get("area_id") or "").strip().lower(),
            }
        )
    return sorted(normalized, key=lambda item: item["index"])


def _template_data_row_count(
    template: dict[str, Any],
    skeleton_rows: list[dict[str, Any]] | None = None,
) -> int | None:
    for key in ("hakodate_data_row_count", "data_row_count", "grid_data_row_count"):
        try:
            value = int(template.get(key))
        except Exception:
            value = 0
        if value > 0:
            return value
    if isinstance(skeleton_rows, list) and skeleton_rows:
        return len(skeleton_rows)
    row_edges = template.get("grid_row_edges") or template.get("row_edges")
    if isinstance(row_edges, list) and len(row_edges) >= 2:
        return max(0, len(row_edges) - 1 - _header_row_count(template))
    return None


def _stored_template_signature(template: dict[str, Any]) -> str:
    for key in _SIGNATURE_KEYS:
        value = template.get(key)
        if isinstance(value, str) and value.strip():
            return value.strip()
    return ""


def _stored_template_signature_components(template: dict[str, Any]) -> dict[str, Any]:
    for key in _SIGNATURE_COMPONENT_KEYS:
        value = template.get(key)
        if isinstance(value, dict):
            return dict(value)
    return {}


def build_facility_template_signature(
    template: dict[str, Any],
    *,
    facility_id: str | None = None,
    skeleton_rows: list[dict[str, Any]] | None = None,
) -> dict[str, Any]:
    fields = derive_row_fields_from_template(template)
    components = {
        "version": "1",
        "facility_id": str(facility_id or template.get("facility_id") or "").strip(),
        "fields": fields,
        "quantity_fields": _quantity_fields(fields),
        "columns": _normalized_template_columns(template),
        "header_rows": _header_row_count(template),
        "data_row_count": _template_data_row_count(template, skeleton_rows),
        "row_roles": template.get("hakodate_row_roles") or template.get("row_roles") or [],
    }
    encoded = json.dumps(components, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return {
        "version": "1",
        "hash": hashlib.sha256(encoded.encode("utf-8")).hexdigest(),
        "components": components,
    }


def validate_facility_template_signature(
    template: dict[str, Any],
    *,
    facility_id: str | None = None,
    skeleton_rows: list[dict[str, Any]] | None = None,
    require_signature: bool = True,
) -> dict[str, Any]:
    current = build_facility_template_signature(
        template,
        facility_id=facility_id,
        skeleton_rows=skeleton_rows,
    )
    stored = _stored_template_signature(template)
    blockers: list[str] = []
    warnings: list[str] = []
    if not stored:
        if require_signature:
            blockers.append("hakodate_template_signature_missing")
        else:
            warnings.append("hakodate_template_signature_missing")
    elif stored != current["hash"]:
        blockers.append("template_stale_due_to_facility_category_change")
    return {
        "status": "ok" if not blockers else "blocked",
        "blockers": blockers,
        "warnings": warnings,
        "stored_signature": stored or None,
        "current_signature": current["hash"],
        "current_components": current["components"],
        "stored_components": _stored_template_signature_components(template),
    }


def build_facility_template_regeneration_candidate(
    template: dict[str, Any],
    *,
    facility_id: str | None = None,
    skeleton_rows: list[dict[str, Any]] | None = None,
) -> dict[str, Any]:
    validation = validate_facility_template_signature(
        template,
        facility_id=facility_id,
        skeleton_rows=skeleton_rows,
        require_signature=False,
    )
    current_components = validation.get("current_components") or {}
    stored_components = validation.get("stored_components") or {}
    current_quantity = list(current_components.get("quantity_fields") or [])
    stored_quantity = list(stored_components.get("quantity_fields") or [])
    candidate_template = deepcopy(template)
    candidate_template["hakodate_template_signature"] = validation.get("current_signature")
    candidate_template["hakodate_template_signature_components"] = current_components
    added = [field for field in current_quantity if field not in stored_quantity]
    removed = [field for field in stored_quantity if field not in current_quantity]
    reordered = (
        bool(stored_quantity)
        and not added
        and not removed
        and stored_quantity != current_quantity
    )
    return {
        "version": "1",
        "status": "candidate_ready",
        "reason": (
            "template_stale_due_to_facility_category_change"
            if validation.get("stored_signature") and validation.get("blockers")
            else "template_signature_missing"
            if not validation.get("stored_signature")
            else "template_signature_current"
        ),
        "stored_signature": validation.get("stored_signature"),
        "candidate_signature": validation.get("current_signature"),
        "diff": {
            "quantity_fields_before": stored_quantity,
            "quantity_fields_after": current_quantity,
            "quantity_fields_added": added,
            "quantity_fields_removed": removed,
            "quantity_fields_reordered": reordered,
            "affected_data_row_count": current_components.get("data_row_count"),
        },
        "candidate_template": candidate_template,
    }


def _field_label(field: str) -> str:
    return str(field or "").strip().replace("qty.", "")


def _find_band_index(value: float, edges: list[float]) -> int | None:
    if len(edges) < 2:
        return None
    for idx in range(len(edges) - 1):
        left = float(edges[idx])
        right = float(edges[idx + 1])
        if left <= value < right or (idx == len(edges) - 2 and left <= value <= right):
            return idx
    return None


def _cell_bbox(
    *,
    row_index: int,
    col_index: int,
    row_edges: list[float],
    column_edges: list[float],
) -> list[float]:
    return [
        float(column_edges[col_index]),
        float(row_edges[row_index]),
        float(column_edges[col_index + 1]),
        float(row_edges[row_index + 1]),
    ]


def _worksheet_row_to_structure_grid_index(worksheet_row: int) -> int | None:
    if worksheet_row in (7, 8):
        return worksheet_row - 7
    if _STRUCTURE_BODY_START_ROW <= worksheet_row <= _STRUCTURE_BODY_END_ROW:
        return 2 + (worksheet_row - _STRUCTURE_BODY_START_ROW)
    return None


def _worksheet_merged_cell_map(worksheet: Any) -> dict[tuple[int, int], dict[str, Any]]:
    merged: dict[tuple[int, int], dict[str, Any]] = {}
    for merged_range in worksheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
        payload = {
            "range": str(merged_range),
            "min_col": int(min_col),
            "min_row": int(min_row),
            "max_col": int(max_col),
            "max_row": int(max_row),
            "row_span": int(max_row - min_row + 1),
            "col_span": int(max_col - min_col + 1),
        }
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                merged[(row, col)] = payload
    return merged


def _merged_or_single_cell_bbox(
    *,
    row_index: int,
    col_index: int,
    worksheet_row: int,
    worksheet_col: int,
    row_edges: list[float],
    column_edges: list[float],
    merged_cells: dict[tuple[int, int], dict[str, Any]],
) -> tuple[list[float], dict[str, Any] | None]:
    merged = merged_cells.get((worksheet_row, worksheet_col))
    if not merged:
        return _cell_bbox(
            row_index=row_index,
            col_index=col_index,
            row_edges=row_edges,
            column_edges=column_edges,
        ), None
    start_row_index = _worksheet_row_to_structure_grid_index(int(merged["min_row"]))
    end_row_index = _worksheet_row_to_structure_grid_index(int(merged["max_row"]))
    if start_row_index is None or end_row_index is None:
        return _cell_bbox(
            row_index=row_index,
            col_index=col_index,
            row_edges=row_edges,
            column_edges=column_edges,
        ), None
    start_col_index = max(0, int(merged["min_col"]) - 1)
    end_col_index = max(0, int(merged["max_col"]) - 1)
    if (
        start_col_index >= len(column_edges) - 1
        or end_col_index >= len(column_edges) - 1
        or start_row_index >= len(row_edges) - 1
        or end_row_index >= len(row_edges) - 1
        or end_col_index < start_col_index
        or end_row_index < start_row_index
    ):
        return _cell_bbox(
            row_index=row_index,
            col_index=col_index,
            row_edges=row_edges,
            column_edges=column_edges,
        ), None
    return [
        float(column_edges[start_col_index]),
        float(row_edges[start_row_index]),
        float(column_edges[end_col_index + 1]),
        float(row_edges[end_row_index + 1]),
    ], dict(merged)


def _worksheet_merge_regions_for_grid(
    worksheet: Any,
    *,
    row_edges: list[float],
    column_edges: list[float],
    quantity_columns: set[int] | None = None,
) -> list[dict[str, Any]]:
    quantity_columns = set(quantity_columns or set())
    regions: list[dict[str, Any]] = []
    for merged_range in worksheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
        start_row_index = _worksheet_row_to_structure_grid_index(int(min_row))
        end_row_index = _worksheet_row_to_structure_grid_index(int(max_row))
        if start_row_index is None or end_row_index is None:
            continue
        start_col_index = max(0, int(min_col) - 1)
        end_col_index = max(0, int(max_col) - 1)
        if (
            start_col_index >= len(column_edges) - 1
            or end_col_index >= len(column_edges) - 1
            or start_row_index >= len(row_edges) - 1
            or end_row_index >= len(row_edges) - 1
            or end_col_index < start_col_index
            or end_row_index < start_row_index
        ):
            continue
        regions.append(
            {
                "range": str(merged_range),
                "min_col": int(min_col),
                "min_row": int(min_row),
                "max_col": int(max_col),
                "max_row": int(max_row),
                "row_span": int(max_row - min_row + 1),
                "col_span": int(max_col - min_col + 1),
                "start_col_index": start_col_index,
                "end_col_index": end_col_index,
                "start_row_index": start_row_index,
                "end_row_index": end_row_index,
                "bbox": [
                    float(column_edges[start_col_index]),
                    float(row_edges[start_row_index]),
                    float(column_edges[end_col_index + 1]),
                    float(row_edges[end_row_index + 1]),
                ],
                "is_quantity": any(
                    col_idx in quantity_columns
                    for col_idx in range(int(min_col), int(max_col) + 1)
                ),
            }
        )
    return regions


def _subtract_grid_line_intervals(
    start: float,
    end: float,
    intervals: list[tuple[float, float]],
) -> list[tuple[float, float]]:
    segments = [(float(start), float(end))]
    for raw_a, raw_b in intervals:
        a = max(float(start), float(raw_a))
        b = min(float(end), float(raw_b))
        if b <= a:
            continue
        next_segments: list[tuple[float, float]] = []
        for seg_a, seg_b in segments:
            if b <= seg_a or a >= seg_b:
                next_segments.append((seg_a, seg_b))
                continue
            if seg_a < a:
                next_segments.append((seg_a, a))
            if b < seg_b:
                next_segments.append((b, seg_b))
        segments = next_segments
    return segments


def _merge_aware_grid_line_segments(
    *,
    row_edges: list[float],
    column_edges: list[float],
    merge_regions: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    if len(row_edges) < 2 or len(column_edges) < 2:
        return []
    horizontal_skips: dict[int, list[tuple[float, float]]] = {
        idx: [] for idx in range(len(row_edges))
    }
    vertical_skips: dict[int, list[tuple[float, float]]] = {
        idx: [] for idx in range(len(column_edges))
    }
    for region in merge_regions:
        try:
            start_row = int(region["start_row_index"])
            end_row = int(region["end_row_index"])
            start_col = int(region["start_col_index"])
            end_col = int(region["end_col_index"])
        except Exception:
            continue
        if (
            start_row < 0
            or end_row >= len(row_edges) - 1
            or start_col < 0
            or end_col >= len(column_edges) - 1
        ):
            continue
        x0 = float(column_edges[start_col])
        x1 = float(column_edges[end_col + 1])
        y0 = float(row_edges[start_row])
        y1 = float(row_edges[end_row + 1])
        for edge_idx in range(start_row + 1, end_row + 1):
            horizontal_skips.setdefault(edge_idx, []).append((x0, x1))
        for edge_idx in range(start_col + 1, end_col + 1):
            vertical_skips.setdefault(edge_idx, []).append((y0, y1))

    segments: list[dict[str, Any]] = []
    y_start = float(row_edges[0])
    y_end = float(row_edges[-1])
    for idx, x in enumerate(column_edges):
        for seg_start, seg_end in _subtract_grid_line_intervals(
            y_start,
            y_end,
            vertical_skips.get(idx, []),
        ):
            if seg_end > seg_start:
                segments.append(
                    {
                        "orientation": "vertical",
                        "start": [float(x), seg_start],
                        "end": [float(x), seg_end],
                    }
                )
    x_start = float(column_edges[0])
    x_end = float(column_edges[-1])
    for idx, y in enumerate(row_edges):
        for seg_start, seg_end in _subtract_grid_line_intervals(
            x_start,
            x_end,
            horizontal_skips.get(idx, []),
        ):
            if seg_end > seg_start:
                segments.append(
                    {
                        "orientation": "horizontal",
                        "start": [seg_start, float(y)],
                        "end": [seg_end, float(y)],
                    }
                )
    return segments


def _distance_to_center(token: HakodateToken, bbox: list[float]) -> float:
    cx = (bbox[0] + bbox[2]) / 2.0
    cy = (bbox[1] + bbox[3]) / 2.0
    return ((float(token.x) - cx) ** 2 + (float(token.y) - cy) ** 2) ** 0.5


def _edge_margin_ratio(token: HakodateToken, bbox: list[float]) -> float:
    width = max(float(bbox[2]) - float(bbox[0]), 1e-9)
    height = max(float(bbox[3]) - float(bbox[1]), 1e-9)
    x_margin = min(float(token.x) - float(bbox[0]), float(bbox[2]) - float(token.x)) / width
    y_margin = min(float(token.y) - float(bbox[1]), float(bbox[3]) - float(token.y)) / height
    return max(0.0, min(x_margin, y_margin))


def _token_payload(token: HakodateToken) -> dict[str, Any]:
    payload = asdict(token)
    payload["digits_normalized"] = _normalize_digits(token.text)
    return payload


def _grid_payload(grid: GridDetectionResult | dict[str, Any]) -> dict[str, Any]:
    if isinstance(grid, GridDetectionResult):
        return {
            "table_box": list(grid.table_box),
            "column_edges": list(grid.column_edges),
            "row_edges": list(grid.row_edges),
            "confidence": float(grid.confidence),
        }
    payload = {
        "table_box": list(grid.get("table_box") or []),
        "column_edges": list(grid.get("column_edges") or grid.get("grid_column_edges") or []),
        "row_edges": list(grid.get("row_edges") or grid.get("grid_row_edges") or []),
        "confidence": float(grid.get("confidence") or 0.0),
    }
    for key in (
        "column_edges_source",
        "row_edges_source",
        "detected_column_count",
        "detected_row_count",
    ):
        if key in grid:
            payload[key] = grid.get(key)
    return payload


def _uses_synthetic_grid_edges(grid_data: dict[str, Any]) -> bool:
    synthetic_sources = {
        "synthesized_from_template",
        "template_fallback",
    }
    return (
        str(grid_data.get("column_edges_source") or "").strip() in synthetic_sources
        or str(grid_data.get("row_edges_source") or "").strip() in synthetic_sources
    )


def _header_row_count(template: dict[str, Any]) -> int:
    header_rows = int(
        template.get("hakodate_header_rows")
        if template.get("hakodate_header_rows") is not None
        else template.get("grid_header_rows")
        if template.get("grid_header_rows") is not None
        else template.get("header_rows")
        if template.get("header_rows") is not None
        else 2
    )
    if header_rows <= 0:
        return 2
    return header_rows


def _synthesize_edges(start: float, end: float, count: int) -> list[float]:
    span = float(end) - float(start)
    if count < 1 or span <= 0:
        return []
    return [float(start) + span * idx / count for idx in range(count + 1)]


def _match_detected_edges_to_expected(
    *,
    detected_edges: list[float],
    expected_cell_count: int,
    table_start: float,
    table_end: float,
    tolerance: float,
) -> tuple[list[float], str | None]:
    """Match detected FAX grid lines to the fixed template line count.

    This intentionally refuses to synthesize evenly-spaced lines. Hakodate
    assignment depends on observed grid evidence, not a guessed structure.
    """
    if expected_cell_count < 1 or table_end <= table_start:
        return [], "invalid_expected_grid"
    normalized = sorted(
        {
            max(table_start, min(table_end, float(edge)))
            for edge in detected_edges
            if table_start <= float(edge) <= table_end
        }
    )
    expected_edge_count = expected_cell_count + 1
    if len(normalized) == expected_edge_count:
        return normalized, None
    if len(normalized) < expected_edge_count:
        return [], "detected_grid_line_count_insufficient"

    targets = [
        table_start + (table_end - table_start) * idx / expected_cell_count
        for idx in range(expected_edge_count)
    ]
    matched: list[float] = []
    cursor = 0
    for target in targets:
        best_index: int | None = None
        best_distance: float | None = None
        for edge_index in range(cursor, len(normalized)):
            remaining_edges = len(normalized) - edge_index
            remaining_targets = expected_edge_count - len(matched)
            if remaining_edges < remaining_targets:
                break
            distance = abs(normalized[edge_index] - target)
            if best_distance is None or distance < best_distance:
                best_index = edge_index
                best_distance = distance
        if best_index is None or best_distance is None or best_distance > tolerance:
            return [], "detected_grid_line_match_failed"
        matched.append(normalized[best_index])
        cursor = best_index + 1
    return matched, None


def _project_structure_edges_to_detected_table_box(
    *,
    structure_edges: list[float],
    structure_start: float,
    structure_end: float,
    detected_start: float,
    detected_end: float,
    expected_edge_count: int,
) -> tuple[list[float], dict[str, Any], str | None]:
    if expected_edge_count < 2 or len(structure_edges) != expected_edge_count:
        return [], {}, "structure_projection_edge_count_mismatch"
    if structure_end <= structure_start or detected_end <= detected_start:
        return [], {}, "structure_projection_box_invalid"
    structure_span = structure_end - structure_start
    detected_span = detected_end - detected_start
    edges = [
        detected_start + ((float(edge) - structure_start) / structure_span) * detected_span
        for edge in structure_edges
    ]
    if len(edges) != expected_edge_count or any(right <= left for left, right in zip(edges, edges[1:])):
        return [], {}, "structure_projection_edges_invalid"
    return edges, {
        "method": "structure_edges_projected_to_detected_table_box",
        "expected_edge_count": expected_edge_count,
        "structure_start": structure_start,
        "structure_end": structure_end,
        "detected_start": detected_start,
        "detected_end": detected_end,
    }, None


def _detect_dense_horizontal_edges_from_table_image(
    page_image: Any,
    *,
    table_box: list[float],
    expected_edge_count: int,
    row_band: tuple[float, float] = (0.36, 0.98),
) -> tuple[list[float], dict[str, Any], str | None]:
    try:
        import cv2
        import numpy as np
    except Exception as exc:  # noqa: BLE001
        return [], {"error": str(exc)}, "dense_horizontal_detector_unavailable"
    if expected_edge_count < 2:
        return [], {}, "invalid_expected_dense_row_count"
    image = np.array(page_image.convert("RGB"))
    gray = cv2.cvtColor(image, cv2.COLOR_RGB2GRAY)
    height, width = gray.shape[:2]
    px_box = _normalized_bbox_to_px(
        table_box,
        image_width=width,
        image_height=height,
        inner_margin_ratio=0.0,
    )
    if not px_box:
        return [], {}, "dense_horizontal_table_box_invalid"
    x0, y0, x1, y1 = px_box
    roi = gray[y0:y1, x0:x1]
    if roi.size == 0:
        return [], {}, "dense_horizontal_roi_empty"
    _, thresholded = cv2.threshold(roi, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)
    band_x0 = max(0, min(thresholded.shape[1], int(thresholded.shape[1] * row_band[0])))
    band_x1 = max(0, min(thresholded.shape[1], int(thresholded.shape[1] * row_band[1])))
    if band_x1 <= band_x0:
        band_x0, band_x1 = 0, thresholded.shape[1]
    band = thresholded[:, band_x0:band_x1]
    best: dict[str, Any] | None = None
    for divisor in (80, 100, 120, 140, 180):
        kernel_width = max(10, band.shape[1] // divisor)
        kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (kernel_width, 1))
        lines = cv2.morphologyEx(band, cv2.MORPH_OPEN, kernel, iterations=1)
        projection = lines.sum(axis=1) / 255.0
        min_length = max(20.0, band.shape[1] * 0.20)
        indexes = np.where(projection >= min_length)[0].tolist()
        groups: list[list[int]] = []
        for idx in indexes:
            if not groups or idx - groups[-1][-1] > 3:
                groups.append([idx])
            else:
                groups[-1].append(idx)
        centers = [int(sum(group) / len(group)) for group in groups if group]
        if not centers:
            continue
        count_error = abs(len(centers) - expected_edge_count)
        candidate = {
            "centers": centers,
            "count_error": count_error,
            "kernel_width": kernel_width,
            "divisor": divisor,
            "line_count": len(centers),
            "row_band": list(row_band),
        }
        if best is None or (count_error, -len(centers)) < (int(best["count_error"]), -int(best["line_count"])):
            best = candidate
    if not best:
        for band_start, band_end in (row_band, (0.45, 0.98), (0.55, 0.98), (0.1, 0.98), (0.0, 1.0)):
            band_x0 = max(0, min(thresholded.shape[1], int(thresholded.shape[1] * band_start)))
            band_x1 = max(0, min(thresholded.shape[1], int(thresholded.shape[1] * band_end)))
            if band_x1 <= band_x0:
                continue
            band = thresholded[:, band_x0:band_x1]
            projection = band.sum(axis=1) / 255.0
            for min_ratio in (0.08, 0.05, 0.10, 0.03, 0.15):
                indexes = np.where(projection >= max(8.0, band.shape[1] * min_ratio))[0].tolist()
                groups: list[list[int]] = []
                for idx in indexes:
                    if not groups or idx - groups[-1][-1] > 3:
                        groups.append([idx])
                    else:
                        groups[-1].append(idx)
                centers = [int(sum(group) / len(group)) for group in groups if group]
                if not centers:
                    continue
                normalized_centers = list(centers)
                if len(normalized_centers) == expected_edge_count - 1 and len(normalized_centers) >= 3:
                    median_gap = float(np.median(np.diff(np.array(normalized_centers))))
                    top_gap = normalized_centers[0]
                    bottom_gap = (y1 - y0) - normalized_centers[-1]
                    if median_gap > 0 and 0 < top_gap <= median_gap * 0.5:
                        normalized_centers = [0] + normalized_centers
                    elif median_gap > 0 and 0 < bottom_gap <= median_gap * 0.5:
                        normalized_centers = normalized_centers + [y1 - y0]
                count_error = abs(len(normalized_centers) - expected_edge_count)
                candidate = {
                    "centers": normalized_centers,
                    "count_error": count_error,
                    "kernel_width": 0,
                    "divisor": 0,
                    "line_count": len(normalized_centers),
                    "row_band": [band_start, band_end],
                    "projection_min_ratio": min_ratio,
                    "method": "dense_horizontal_projection_detection",
                }
                if best is None or (count_error, -len(normalized_centers)) < (
                    int(best["count_error"]),
                    -int(best["line_count"]),
                ):
                    best = candidate
    if not best:
        return [], {}, "dense_horizontal_lines_not_found"
    centers = [int(value) for value in best["centers"]]
    count_error = abs(len(centers) - expected_edge_count)
    if count_error > 1:
        return [], {key: value for key, value in best.items() if key != "centers"}, "dense_horizontal_line_count_mismatch"
    if len(centers) == expected_edge_count - 1:
        bottom_gap = (y1 - y0) - centers[-1]
        median_gap = float(np.median(np.diff(np.array(centers)))) if len(centers) >= 3 else 0.0
        if median_gap > 0 and 0 < bottom_gap <= median_gap * 0.35:
            centers.append(y1 - y0)
        else:
            return [], {key: value for key, value in best.items() if key != "centers"}, "dense_horizontal_line_count_insufficient"
    if len(centers) > expected_edge_count:
        # Drop the weakest ambiguity by keeping the first expected table edges.
        centers = centers[:expected_edge_count]
    if len(centers) < expected_edge_count:
        return [], {key: value for key, value in best.items() if key != "centers"}, "dense_horizontal_line_count_insufficient"
    edges = [(y0 + center) / float(height) for center in centers]
    evidence = {
        "method": best.get("method") or "dense_horizontal_line_detection",
        "line_count": len(edges),
        "expected_edge_count": expected_edge_count,
        "kernel_width": best["kernel_width"],
        "divisor": best["divisor"],
        "row_band": best["row_band"],
        "table_box_px": [x0, y0, x1, y1],
    }
    return edges, evidence, None


def _detect_dense_vertical_edges_from_table_image(
    page_image: Any,
    *,
    table_box: list[float],
    expected_edge_count: int,
    structure_edges: list[float] | None = None,
    column_band: tuple[float, float] = (0.0, 0.9),
) -> tuple[list[float], dict[str, Any], str | None]:
    try:
        import cv2
        import numpy as np
    except Exception as exc:  # noqa: BLE001
        return [], {"error": str(exc)}, "dense_vertical_detector_unavailable"
    if expected_edge_count < 2:
        return [], {}, "invalid_expected_dense_column_count"
    image = np.array(page_image.convert("RGB"))
    gray = cv2.cvtColor(image, cv2.COLOR_RGB2GRAY)
    height, width = gray.shape[:2]
    px_box = _normalized_bbox_to_px(
        table_box,
        image_width=width,
        image_height=height,
        inner_margin_ratio=0.0,
    )
    if not px_box:
        return [], {}, "dense_vertical_table_box_invalid"
    x0, y0, x1, y1 = px_box
    roi = gray[y0:y1, x0:x1]
    if roi.size == 0:
        return [], {}, "dense_vertical_roi_empty"
    _, thresholded = cv2.threshold(roi, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)
    band_y0 = max(0, min(thresholded.shape[0], int(thresholded.shape[0] * column_band[0])))
    band_y1 = max(0, min(thresholded.shape[0], int(thresholded.shape[0] * column_band[1])))
    if band_y1 <= band_y0:
        band_y0, band_y1 = 0, thresholded.shape[0]
    band = thresholded[band_y0:band_y1, :]
    best: dict[str, Any] | None = None
    for divisor in (80, 100, 120, 160, 200):
        kernel_height = max(10, band.shape[0] // divisor)
        kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1, kernel_height))
        lines = cv2.morphologyEx(band, cv2.MORPH_OPEN, kernel, iterations=1)
        projection = lines.sum(axis=0) / 255.0
        min_length = max(20.0, band.shape[0] * 0.20)
        indexes = np.where(projection >= min_length)[0].tolist()
        groups: list[list[int]] = []
        for idx in indexes:
            if not groups or idx - groups[-1][-1] > 3:
                groups.append([idx])
            else:
                groups[-1].append(idx)
        centers = [int(sum(group) / len(group)) for group in groups if group]
        if not centers:
            continue
        merged = centers
        if structure_edges and len(structure_edges) == expected_edge_count:
            sx0 = float(structure_edges[0])
            sx1 = float(structure_edges[-1])
            span = sx1 - sx0
            expected_px = [
                int(round(((float(edge) - sx0) / span) * (x1 - x0)))
                for edge in structure_edges
                if span > 0
            ]
            selected: list[int] = []
            used: set[int] = set()
            tolerance_px = max(18, int((x1 - x0) * 0.04))
            for expected in expected_px:
                best_index: int | None = None
                best_distance: int | None = None
                for idx, center in enumerate(centers):
                    if idx in used:
                        continue
                    distance = abs(center - expected)
                    if best_distance is None or distance < best_distance:
                        best_index = idx
                        best_distance = distance
                if best_index is None or best_distance is None or best_distance > tolerance_px:
                    selected = []
                    break
                selected.append(centers[best_index])
                used.add(best_index)
            if len(selected) == expected_edge_count and selected == sorted(selected):
                merged = selected
        else:
            merged = []
            merge_px = max(8, int((x1 - x0) * 0.012))
            for center in centers:
                if not merged or center - merged[-1] > merge_px:
                    merged.append(center)
                else:
                    merged[-1] = int((merged[-1] + center) / 2)
        count_error = abs(len(merged) - expected_edge_count)
        candidate = {
            "centers": merged,
            "count_error": count_error,
            "kernel_height": kernel_height,
            "divisor": divisor,
            "line_count": len(merged),
            "column_band": list(column_band),
            "raw_line_count": len(centers),
        }
        if best is None or (count_error, -len(merged)) < (int(best["count_error"]), -int(best["line_count"])):
            best = candidate
    if not best:
        for band_start, band_end in (column_band, (0.0, 1.0), (0.05, 0.95), (0.1, 0.9)):
            band_y0 = max(0, min(thresholded.shape[0], int(thresholded.shape[0] * band_start)))
            band_y1 = max(0, min(thresholded.shape[0], int(thresholded.shape[0] * band_end)))
            if band_y1 <= band_y0:
                continue
            band = thresholded[band_y0:band_y1, :]
            projection = band.sum(axis=0) / 255.0
            for min_ratio in (0.08, 0.05, 0.10, 0.03, 0.15):
                indexes = np.where(projection >= max(8.0, band.shape[0] * min_ratio))[0].tolist()
                groups: list[list[int]] = []
                for idx in indexes:
                    if not groups or idx - groups[-1][-1] > 3:
                        groups.append([idx])
                    else:
                        groups[-1].append(idx)
                centers = [int(sum(group) / len(group)) for group in groups if group]
                if len(centers) != expected_edge_count:
                    continue
                candidate = {
                    "centers": centers,
                    "count_error": 0,
                    "kernel_height": 0,
                    "divisor": 0,
                    "line_count": len(centers),
                    "column_band": [band_start, band_end],
                    "raw_line_count": len(centers),
                    "projection_min_ratio": min_ratio,
                    "method": "dense_vertical_projection_detection",
                    "merge_px": 0,
                }
                best = candidate
                break
            if best:
                break
    if not best:
        return [], {}, "dense_vertical_lines_not_found"
    centers = [int(value) for value in best["centers"]]
    if len(centers) != expected_edge_count:
        return [], {key: value for key, value in best.items() if key != "centers"}, "dense_vertical_line_count_mismatch"
    edges = [(x0 + center) / float(width) for center in centers]
    evidence = {
        "method": best.get("method") or "dense_vertical_line_detection",
        "line_count": len(edges),
        "expected_edge_count": expected_edge_count,
        "kernel_height": best["kernel_height"],
        "divisor": best["divisor"],
        "column_band": best["column_band"],
        "merge_px": best.get("merge_px"),
        "table_box_px": [x0, y0, x1, y1],
    }
    return edges, evidence, None


def _align_grid_edges_to_template(
    *,
    grid: GridDetectionResult | dict[str, Any],
    template: dict[str, Any],
    skeleton_rows: list[dict[str, Any]] | None,
) -> tuple[dict[str, Any], list[str], list[str]]:
    grid_data = _grid_payload(grid)
    table_box = grid_data.get("table_box") or []
    if not isinstance(table_box, list) or len(table_box) < 4:
        return grid_data, [], ["grid_table_box_missing"]

    warnings: list[str] = []
    blockers: list[str] = []
    fields = derive_row_fields_from_template(template)
    column_edges = [float(value) for value in (grid_data.get("column_edges") or [])]
    row_edges = [float(value) for value in (grid_data.get("row_edges") or [])]
    expected_columns = len(fields)
    expected_rows = _header_row_count(template) + len(skeleton_rows or [])
    detected_column_count = max(0, len(column_edges) - 1)
    detected_row_count = max(0, len(row_edges) - 1)
    tolerance = float(template.get("hakodate_grid_line_match_tolerance") or 0.04)
    tolerance = max(0.005, min(tolerance, 0.12))

    if expected_columns >= 2 and detected_column_count != expected_columns:
        matched, reason = _match_detected_edges_to_expected(
            detected_edges=column_edges,
            expected_cell_count=expected_columns,
            table_start=float(table_box[0]),
            table_end=float(table_box[2]),
            tolerance=tolerance,
        )
        grid_data["detected_column_count"] = detected_column_count
        if matched:
            grid_data["column_edges"] = matched
            grid_data["column_edges_source"] = "detected_template_matched"
            warnings.append("grid_column_edges_template_matched")
        else:
            blockers.append(reason or "grid_column_edges_unmatched")
    elif expected_columns >= 2:
        grid_data["column_edges_source"] = grid_data.get("column_edges_source") or "detected_exact_count"
        grid_data["detected_column_count"] = detected_column_count

    if expected_rows >= 2 and detected_row_count != expected_rows:
        matched, reason = _match_detected_edges_to_expected(
            detected_edges=row_edges,
            expected_cell_count=expected_rows,
            table_start=float(table_box[1]),
            table_end=float(table_box[3]),
            tolerance=tolerance,
        )
        grid_data["detected_row_count"] = detected_row_count
        if matched:
            grid_data["row_edges"] = matched
            grid_data["row_edges_source"] = "detected_template_matched"
            warnings.append("grid_row_edges_template_matched")
        else:
            blockers.append(reason or "grid_row_edges_unmatched")
    elif expected_rows >= 2:
        grid_data["row_edges_source"] = grid_data.get("row_edges_source") or "detected_exact_count"
        grid_data["detected_row_count"] = detected_row_count

    return grid_data, warnings, blockers


def _template_table_box(template: dict[str, Any]) -> list[float]:
    for value in (
        template.get("grid_table_box"),
        template.get("table_box"),
        ((template.get("rois") or {}).get("table_box") if isinstance(template.get("rois"), dict) else None),
    ):
        if not isinstance(value, list) or len(value) < 4:
            continue
        try:
            box = [float(value[idx]) for idx in range(4)]
        except Exception:  # noqa: BLE001
            continue
        if box[2] > box[0] and box[3] > box[1]:
            return box
    return []


def _fallback_grid_from_template(
    *,
    template: dict[str, Any],
    skeleton_rows: list[dict[str, Any]] | None,
) -> dict[str, Any] | None:
    table_box = _template_table_box(template)
    fields = derive_row_fields_from_template(template)
    row_count = _header_row_count(template) + len(skeleton_rows or [])
    if not table_box or len(fields) < 2 or row_count < 2:
        return None
    column_edges = _synthesize_edges(table_box[0], table_box[2], len(fields))
    row_edges = _synthesize_edges(table_box[1], table_box[3], row_count)
    if not column_edges or not row_edges:
        return None
    return {
        "table_box": table_box,
        "column_edges": column_edges,
        "row_edges": row_edges,
        "confidence": 0.2,
        "column_edges_source": "template_fallback",
        "row_edges_source": "template_fallback",
        "detected_column_count": 0,
        "detected_row_count": 0,
    }


def _merged_value_map(worksheet: Any) -> dict[tuple[int, int], object]:
    merged: dict[tuple[int, int], object] = {}
    for merged_range in worksheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
        anchor_value = worksheet.cell(min_row, min_col).value
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                merged[(row, col)] = anchor_value
    return merged


def _effective_cell_value(
    worksheet: Any,
    merged_values: dict[tuple[int, int], object],
    *,
    row: int,
    col: int,
) -> object:
    value = worksheet.cell(row=row, column=col).value
    if value is not None:
        return value
    return merged_values.get((row, col))


def _template_slot_label(column: dict[str, Any], field: str) -> str:
    label = str(column.get("header") or column.get("label") or "").strip()
    if label:
        return label
    role = str(column.get("role") or "").strip().lower()
    if role == "date":
        return "日付"
    if role == "daypart":
        return "区分"
    if role == "menu_name":
        return "献立"
    if role == "note" or field == "remarks":
        return "備考欄"
    if field.startswith("qty."):
        return field.removeprefix("qty.")
    return field or role or "unknown"


def _canonical_slot_from_template_column(
    column: dict[str, Any],
    *,
    fallback_index: int,
) -> dict[str, Any] | None:
    field = canonical_field_name_from_template_column(column, fallback_index=fallback_index)
    role = str(column.get("role") or "").strip().lower()
    if not field:
        return None
    if role == "note":
        slot_role = "note"
        slot_name = "note"
    elif field == "date_mmdd":
        slot_role = "date"
        slot_name = "date"
    elif field == "daypart":
        slot_role = "daypart"
        slot_name = "daypart"
    elif field == "menu":
        slot_role = "menu_name"
        slot_name = "menu_name"
    elif field.startswith("qty."):
        slot_role = "quantity"
        slot_name = field
    else:
        slot_role = role or "unknown"
        slot_name = field
    source_index = column.get("source_index")
    try:
        explicit_source_index = int(source_index) if source_index is not None else None
    except Exception:
        explicit_source_index = None
    return {
        "role": slot_role,
        "slot_name": slot_name,
        "label": _template_slot_label(column, field),
        "canonical_field": field,
        "canonical_source": "facility_fax_template",
        "logical_index": fallback_index,
        "explicit_source_index": explicit_source_index,
    }


def _canonical_slot_specs_from_template(
    template: dict[str, Any] | None,
) -> list[dict[str, Any]]:
    if not isinstance(template, dict):
        return []
    raw_columns = template.get("columns")
    if not isinstance(raw_columns, list) or not raw_columns:
        raw_columns = template.get("grid_columns")
    if not isinstance(raw_columns, list) or not raw_columns:
        return []

    ordered = sorted(
        [column for column in raw_columns if isinstance(column, dict)],
        key=lambda column: int(column.get("index") or 0),
    )
    specs: list[dict[str, Any]] = []
    for fallback_index, column in enumerate(ordered):
        spec = _canonical_slot_from_template_column(column, fallback_index=fallback_index)
        if not spec:
            continue
        specs.append(spec)
    return specs


def _canonical_column_slots_from_template(
    template: dict[str, Any] | None,
    *,
    col_count: int,
) -> dict[int, dict[str, Any]]:
    slots: dict[int, dict[str, Any]] = {}
    for spec in _canonical_slot_specs_from_template(template):
        source_index = spec.get("explicit_source_index")
        if source_index is None:
            continue
        try:
            worksheet_col = int(source_index) + 1
        except Exception:
            continue
        if worksheet_col <= 0 or worksheet_col > col_count:
            continue
        slots[worksheet_col] = {
            "role": spec["role"],
            "slot_name": spec["slot_name"],
            "label": spec["label"],
            "canonical_field": spec["canonical_field"],
            "canonical_source": "facility_fax_template",
        }
    return slots


def _canonical_slot_match_score(slot: dict[str, Any], spec: dict[str, Any]) -> int:
    slot_role = str(slot.get("role") or "").strip()
    slot_name = str(slot.get("slot_name") or "").strip()
    slot_label = _normalize_slot_text(slot.get("label") or "")
    spec_role = str(spec.get("role") or "").strip()
    spec_name = str(spec.get("slot_name") or "").strip()
    spec_label = _normalize_slot_text(spec.get("label") or "")
    if not spec_name:
        return 0
    if slot_name == spec_name:
        return 100
    if spec_name == "qty.placeholder_x" and slot_role == "spacer":
        return 95
    if spec_role == "note" and slot_role == "note":
        return 90
    if spec_role in {"date", "daypart", "menu_name"} and slot_role == spec_role:
        return 90
    if spec_role == slot_role and spec_label and slot_label == spec_label:
        return 80
    return 0


def _apply_implicit_canonical_slots(
    slots: list[dict[str, Any]],
    specs: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    if not specs:
        return slots
    assigned_cols = {
        int(slot.get("worksheet_col_index") or 0)
        for slot in slots
        if str(slot.get("canonical_source") or "").strip() == "facility_fax_template"
    }
    for spec in specs:
        if spec.get("explicit_source_index") is not None:
            continue
        matches: list[tuple[int, int, dict[str, Any]]] = []
        for slot in slots:
            worksheet_col = int(slot.get("worksheet_col_index") or 0)
            if worksheet_col <= 0 or worksheet_col in assigned_cols:
                continue
            score = _canonical_slot_match_score(slot, spec)
            if score > 0:
                matches.append((score, worksheet_col, slot))
        if not matches:
            continue
        if str(spec.get("role") or "").strip() == "note":
            selected = sorted(matches, key=lambda item: item[1])
        else:
            selected = [max(matches, key=lambda item: (item[0], -item[1]))]
        for _score, worksheet_col, slot in selected:
            slot.update(
                {
                    "role": spec["role"],
                    "slot_name": spec["slot_name"],
                    "label": spec["label"],
                    "canonical_field": spec["canonical_field"],
                    "canonical_source": "facility_fax_template",
                }
            )
            assigned_cols.add(worksheet_col)
    return slots


def _column_slots_from_worksheet(
    worksheet: Any,
    *,
    col_count: int,
    template: dict[str, Any] | None = None,
) -> list[dict[str, Any]]:
    merged_values = _merged_value_map(worksheet)
    canonical_slots = _canonical_column_slots_from_template(template, col_count=col_count)
    canonical_specs = _canonical_slot_specs_from_template(template)
    slots: list[dict[str, Any]] = []
    for raw_col_index in range(col_count):
        worksheet_col = raw_col_index + 1
        header_values = [
            _normalize_slot_text(
                _effective_cell_value(worksheet, merged_values, row=row, col=worksheet_col)
            )
            for row in _STRUCTURE_HEADER_ROWS
        ]
        body_samples: list[str] = []
        for row in range(_STRUCTURE_BODY_START_ROW, _STRUCTURE_BODY_END_ROW + 1):
            text = _normalize_slot_text(worksheet.cell(row=row, column=worksheet_col).value)
            if text:
                body_samples.append(text)
            if len(body_samples) >= 8:
                break
        header_blob = "".join(value for value in header_values if value).replace(" ", "")
        sample_set = {sample for sample in body_samples if sample}
        role = "unknown"
        slot_name = f"col_{raw_col_index}"
        label = header_blob or slot_name
        if "日付" in header_blob:
            role, slot_name, label = "date", "date", "日付"
        elif "区分" in header_blob:
            if sample_set and sample_set <= {"朝", "昼", "夕", "タ", "粉", "塩", "師"}:
                role, slot_name, label = "daypart", "daypart", "区分"
            else:
                role, slot_name, label = "aux", "aux", "補助区分"
        elif "献立" in header_blob:
            role, slot_name, label = "menu_name", "menu_name", "献立"
        elif "肉禁" in header_blob:
            role, slot_name, label = "quantity", "qty.no_meat_x", "肉禁"
        elif "魚禁" in header_blob:
            role, slot_name, label = "quantity", "qty.no_fish_x", "魚禁"
        elif "常食" in header_blob:
            role, slot_name, label = "quantity", "qty.regular_x", "常食"
        elif "糖尿" in header_blob:
            role, slot_name, label = "quantity", "qty.diabetes_x", "糖尿"
        elif "職員" in header_blob:
            role, slot_name, label = "quantity", "qty.staff_x", "職員"
        elif "通所" in header_blob:
            role, slot_name, label = "quantity", "qty.daycare_x", "通所"
        elif "軟菜" in header_blob:
            role, slot_name, label = "quantity", "qty.soft_x", "軟菜"
        elif "ミキサ" in header_blob or "ﾐｷｻ" in header_blob:
            role, slot_name, label = "quantity", "qty.mixer_x", "ミキサー"
        elif "袋分" in header_blob:
            role, slot_name, label = "quantity", "qty.regular_bag_x", "袋分け"
        elif "変更1" in header_blob or "変更①" in header_blob:
            role, slot_name, label = "quantity", "qty.change_1_x", "変更1"
        elif "変更2" in header_blob or "変更②" in header_blob:
            role, slot_name, label = "quantity", "qty.change_2_x", "変更2"
        elif "備考" in header_blob:
            role, slot_name, label = "note", "note", "備考欄"
        elif not header_blob and not body_samples:
            role, slot_name, label = "spacer", "spacer", "spacer"
        slot = {
            "raw_col_index": raw_col_index,
            "worksheet_col_index": worksheet_col,
            "role": role,
            "slot_name": slot_name,
            "label": label,
            "header_values": header_values,
            "body_samples": body_samples,
            "source": "worksheet_header_inference",
        }
        canonical = canonical_slots.get(worksheet_col)
        if canonical:
            slot.update(canonical)
        slots.append(slot)
    return _apply_implicit_canonical_slots(slots, canonical_specs)


def _workbook_physical_row_map(worksheet: Any, *, row_count: int, header_rows: int = 2) -> dict[int, dict[str, Any]]:
    rows: dict[int, dict[str, Any]] = {}
    last_date = ""
    for row_index in range(header_rows, row_count):
        worksheet_row = _STRUCTURE_BODY_START_ROW + (row_index - header_rows)
        if worksheet_row > _STRUCTURE_BODY_END_ROW:
            break
        date_text = _normalize_slot_text(worksheet.cell(row=worksheet_row, column=1).value)
        if date_text and not _is_weekday_only(date_text):
            last_date = date_text
        rows[row_index] = {
            "worksheet_row": worksheet_row,
            "row_index": row_index,
            "date": date_text,
            "effective_date": last_date,
            "daypart": _normalize_slot_text(worksheet.cell(row=worksheet_row, column=2).value),
            "aux": _normalize_slot_text(worksheet.cell(row=worksheet_row, column=3).value),
            "menu_name": _normalize_slot_text(worksheet.cell(row=worksheet_row, column=4).value),
            "menu_key": _normalize_menu_key(worksheet.cell(row=worksheet_row, column=4).value),
        }
    return rows


def _week_sheet_name_from_template(template: dict[str, Any]) -> str:
    for key in ("hakodate_week_sheet_name", "week_sheet_name"):
        value = str(template.get(key) or "").strip()
        if value:
            return value
    return order_form_service._DEFAULT_WEEK_SHEET  # noqa: SLF001


def _source_worksheet_for_structure_template(*, facility_id: str, week_sheet_name: str) -> Any:
    facility = order_form_service.config_service.get_facility_config(facility_id)
    if not facility:
        raise ValueError("facility not found")
    fax_template_id = str(order_form_service._infer_fax_template_id_from_facility(facility) or "").strip()  # noqa: SLF001
    if not fax_template_id:
        raise ValueError("facility fax_template_id not found")
    source_workbook_name = order_form_service.resolve_facility_source_workbook_name_for_week_sheet(
        facility,
        week_sheet_name,
    )
    source_workbook_path = order_form_service._resolve_source_workbook_path(source_workbook_name)  # noqa: SLF001
    workbook = load_workbook(source_workbook_path, data_only=True)
    if week_sheet_name not in workbook.sheetnames:
        raise ValueError(f"week sheet not found in source workbook: {week_sheet_name}")
    return workbook[week_sheet_name]


def _structure_grid_for_facility_template(
    *,
    facility_id: str,
    week_sheet_name: str,
    dpi: int,
) -> tuple[dict[str, Any], Any]:
    import tempfile
    from pathlib import Path

    with tempfile.TemporaryDirectory(prefix="hakodate_structure_") as tmp:
        output_dir = Path(tmp)
        structure_xlsx = order_form_service.build_fax_structure_only_excel(
            facility_id=facility_id,
            week_sheet_name=week_sheet_name,
            output_dir=output_dir,
        )
        structure_pdf = output_dir / f"{structure_xlsx.stem}.pdf"
        render_workbook_path_to_pdf(
            structure_xlsx,
            output_path=structure_pdf,
            sheet_name=week_sheet_name,
            dpi=dpi,
        )
        workbook = load_workbook(structure_xlsx, data_only=True)
        worksheet = workbook[week_sheet_name]
        structure_grid = _structure_grid_from_rendered_worksheet(worksheet, dpi=dpi)
    return structure_grid, _source_worksheet_for_structure_template(
        facility_id=facility_id,
        week_sheet_name=week_sheet_name,
    )


def _structure_grid_from_rendered_worksheet(worksheet: Any, *, dpi: int) -> dict[str, Any]:
    min_col, min_row, max_col, max_row = workbook_pdf_renderer._sheet_print_range(worksheet)  # noqa: SLF001
    margin_px = workbook_pdf_renderer._DEFAULT_MARGIN_PX  # noqa: SLF001
    default_row_height = worksheet.sheet_format.defaultRowHeight
    col_widths: dict[int, int] = {}
    row_heights: dict[int, int] = {}
    for col in range(min_col, max_col + 1):
        dim = worksheet.column_dimensions[get_column_letter(col)]
        col_widths[col] = workbook_pdf_renderer._column_width_to_pixels(dim.width, dpi=dpi)  # noqa: SLF001
    for row in range(min_row, max_row + 1):
        dim = worksheet.row_dimensions[row]
        row_heights[row] = workbook_pdf_renderer._row_height_to_pixels(dim.height or default_row_height, dpi=dpi)  # noqa: SLF001

    x_positions = {min_col: margin_px}
    current_x = margin_px
    for col in range(min_col, max_col + 1):
        current_x += col_widths[col]
        x_positions[col + 1] = current_x
    y_positions = {min_row: margin_px}
    current_y = margin_px
    for row in range(min_row, max_row + 1):
        current_y += row_heights[row]
        y_positions[row + 1] = current_y
    image_width = current_x + margin_px
    image_height = current_y + margin_px

    table_cols = list(range(1, max_col + 1))
    header_rows = [7, 8]
    body_rows = list(range(_STRUCTURE_BODY_START_ROW, _STRUCTURE_BODY_END_ROW + 1))
    grid_rows = header_rows + body_rows
    x_edges_px = [x_positions[col] for col in table_cols if col in x_positions]
    if max_col + 1 in x_positions:
        x_edges_px.append(x_positions[max_col + 1])
    y_edges_px = [y_positions[row] for row in grid_rows if row in y_positions]
    if _STRUCTURE_BODY_END_ROW + 1 in y_positions:
        y_edges_px.append(y_positions[_STRUCTURE_BODY_END_ROW + 1])
    if len(x_edges_px) < 2 or len(y_edges_px) < 2:
        raise ValueError("structure_grid_not_found")
    column_edges = [float(x) / float(image_width) for x in x_edges_px]
    row_edges = [float(y) / float(image_height) for y in y_edges_px]
    return {
        "table_box": [column_edges[0], row_edges[0], column_edges[-1], row_edges[-1]],
        "column_edges": column_edges,
        "row_edges": row_edges,
        "confidence": 1.0,
        "column_edges_source": "structure_workbook_geometry",
        "row_edges_source": "structure_workbook_geometry",
        "detected_column_count": len(column_edges) - 1,
        "detected_row_count": len(row_edges) - 1,
    }


def _map_token_to_structure_space(
    token: HakodateToken,
    *,
    source_box: list[float],
    target_box: list[float],
) -> HakodateToken | None:
    try:
        sx0, sy0, sx1, sy1 = [float(value) for value in source_box]
        tx0, ty0, tx1, ty1 = [float(value) for value in target_box]
    except Exception:
        return None
    sw = sx1 - sx0
    sh = sy1 - sy0
    tw = tx1 - tx0
    th = ty1 - ty0
    if sw <= 0 or sh <= 0 or tw <= 0 or th <= 0:
        return None

    def map_x(value: float) -> float:
        return tx0 + ((value - sx0) / sw) * tw

    def map_y(value: float) -> float:
        return ty0 + ((value - sy0) / sh) * th

    return HakodateToken(
        text=token.text,
        x=map_x(float(token.x)),
        y=map_y(float(token.y)),
        bbox=[
            map_x(float(token.bbox[0])),
            map_y(float(token.bbox[1])),
            map_x(float(token.bbox[2])),
            map_y(float(token.bbox[3])),
        ],
        confidence=token.confidence,
    )


def _map_structure_bbox_to_fax_bbox(
    bbox: list[float],
    *,
    structure_box: list[float],
    fax_box: list[float],
) -> list[float] | None:
    try:
        sx0, sy0, sx1, sy1 = [float(value) for value in structure_box]
        fx0, fy0, fx1, fy1 = [float(value) for value in fax_box]
        bx0, by0, bx1, by1 = [float(value) for value in bbox]
    except Exception:
        return None
    sw = sx1 - sx0
    sh = sy1 - sy0
    fw = fx1 - fx0
    fh = fy1 - fy0
    if sw <= 0 or sh <= 0 or fw <= 0 or fh <= 0:
        return None

    def map_x(value: float) -> float:
        return fx0 + ((value - sx0) / sw) * fw

    def map_y(value: float) -> float:
        return fy0 + ((value - sy0) / sh) * fh

    return [map_x(bx0), map_y(by0), map_x(bx1), map_y(by1)]


def _render_pdf_page_image(pdf_bytes: bytes, template: dict[str, Any]):
    try:
        import pdfplumber
    except Exception as exc:  # noqa: BLE001
        logger.warning("Hakodate PDF rendering unavailable: {}", str(exc))
        raise HakodateOcrUnavailable(str(exc)) from exc

    page_index = max(int(template.get("page", 1)) - 1, 0)
    resolution = int(
        template.get("hakodate_cell_ocr_resolution")
        or template.get("hakodate_ocr_resolution")
        or template.get("grid_dpi")
        or 300
    )
    with pdfplumber.open(BytesIO(pdf_bytes)) as pdf:
        if not pdf.pages:
            return None
        page = pdf.pages[page_index] if page_index < len(pdf.pages) else pdf.pages[0]
        return page.to_image(resolution=resolution).original.convert("RGB")


def _normalized_bbox_to_px(
    bbox: list[float],
    *,
    image_width: int,
    image_height: int,
    inner_margin_ratio: float = 0.0,
) -> list[int] | None:
    try:
        x0, y0, x1, y1 = [float(value) for value in bbox]
    except Exception:
        return None
    if x1 <= x0 or y1 <= y0:
        return None
    margin_x = max(0.0, (x1 - x0) * inner_margin_ratio)
    margin_y = max(0.0, (y1 - y0) * inner_margin_ratio)
    px0 = max(0, min(image_width, int(round((x0 + margin_x) * image_width))))
    py0 = max(0, min(image_height, int(round((y0 + margin_y) * image_height))))
    px1 = max(0, min(image_width, int(round((x1 - margin_x) * image_width))))
    py1 = max(0, min(image_height, int(round((y1 - margin_y) * image_height))))
    if px1 <= px0 or py1 <= py0:
        return None
    return [px0, py0, px1, py1]


def _ocr_quantity_cell(
    *,
    page_image: Any,
    fax_cell_bbox: list[float],
    template: dict[str, Any],
    cell_id: str,
    debug_dir: str | None = None,
) -> HakodateCellOcrResult:
    _ = page_image, fax_cell_bbox, template, cell_id, debug_dir
    raise HakodateOcrUnavailable("tesseract_ocr_removed")


def _save_structure_slot_overlay(
    *,
    page_image: Any,
    cells: list[dict[str, Any]],
    actual_box: list[float],
    structure_box: list[float],
    actual_columns: list[float] | None = None,
    actual_rows: list[float] | None = None,
    worksheet: Any | None = None,
    merge_regions: list[dict[str, Any]] | None = None,
    debug_dir: str | None,
) -> str | None:
    if not debug_dir:
        return None
    try:
        from pathlib import Path
        from PIL import Image, ImageDraw

        out_dir = Path(debug_dir)
        out_dir.mkdir(parents=True, exist_ok=True)
        image = page_image.convert("RGBA")
        overlay = Image.new("RGBA", image.size, (0, 0, 0, 0))
        draw = ImageDraw.Draw(overlay)
        width, height = image.size

        def px_box(box: list[float]) -> tuple[int, int, int, int]:
            return (
                int(float(box[0]) * width),
                int(float(box[1]) * height),
                int(float(box[2]) * width),
                int(float(box[3]) * height),
            )

        def px_point(point: list[float]) -> tuple[int, int]:
            return (int(float(point[0]) * width), int(float(point[1]) * height))

        draw.rectangle(px_box(actual_box), outline=(255, 0, 0, 255), width=4)
        mapped_structure_box = _map_structure_bbox_to_fax_bbox(
            structure_box,
            structure_box=structure_box,
            fax_box=actual_box,
        )
        if mapped_structure_box:
            draw.rectangle(px_box(mapped_structure_box), outline=(0, 128, 255, 255), width=2)

        grid_drawn = False
        if actual_columns and actual_rows and worksheet is not None:
            current_merge_regions = list(merge_regions or [])
            if not current_merge_regions:
                quantity_columns = order_form_service._worksheet_quantity_column_indexes(worksheet)  # noqa: SLF001
                current_merge_regions = _worksheet_merge_regions_for_grid(
                    worksheet,
                    row_edges=actual_rows,
                    column_edges=actual_columns,
                    quantity_columns=quantity_columns,
                )
            for segment in _merge_aware_grid_line_segments(
                row_edges=actual_rows,
                column_edges=actual_columns,
                merge_regions=current_merge_regions,
            ):
                draw.line(
                    [px_point(segment["start"]), px_point(segment["end"])],
                    fill=(0, 180, 0, 255),
                    width=2,
                )
            for region in current_merge_regions:
                if not region.get("is_quantity"):
                    continue
                box = region.get("bbox")
                if not isinstance(box, list) or len(box) != 4:
                    continue
                draw.rectangle(px_box(box), fill=(255, 128, 0, 35), outline=(255, 128, 0, 255), width=4)
            grid_drawn = True

        for cell in cells:
            box = cell.get("fax_cell_bbox")
            if not grid_drawn and isinstance(box, list) and len(box) == 4:
                draw.rectangle(px_box(box), outline=(0, 180, 0, 255), width=2)
            center = cell.get("fax_cell_center") or []
            if isinstance(center, list) and len(center) == 2:
                cx = int(float(center[0]) * width)
                cy = int(float(center[1]) * height)
                draw.ellipse((cx - 4, cy - 4, cx + 4, cy + 4), fill=(255, 128, 0, 255))
        image = Image.alpha_composite(image, overlay).convert("RGB")
        path = out_dir / "structure_slot_overlay.png"
        image.save(path)
        return str(path)
    except Exception as exc:  # noqa: BLE001
        logger.warning("Failed to save Hakodate structure slot overlay: {}", str(exc))
        return None


def _blocked_structure_slot_assignment(
    *,
    strategy: str,
    template: dict[str, Any],
    blockers: list[str],
    warnings: list[str] | None = None,
    facility_id: str | None = None,
    skeleton_rows: list[dict[str, Any]] | None = None,
    signature_validation: dict[str, Any] | None = None,
    grid: dict[str, Any] | None = None,
) -> dict[str, Any]:
    return {
        "version": "1",
        "strategy": strategy,
        "status": "blocked",
        "blockers": blockers,
        "warnings": list(warnings or ["structure_slot_assignment"]),
        "fields": derive_row_fields_from_template(template),
        "quantity_fields": [],
        "template_signature": signature_validation
        if signature_validation is not None
        else validate_facility_template_signature(
            template,
            facility_id=facility_id,
            skeleton_rows=skeleton_rows,
            require_signature=True,
        ),
        "header_rows": 2,
        "grid": grid,
        "assignments": [],
        "review_candidates": [],
        "rejected_candidates": [],
        "metrics": {
            "ocr_numeric_candidate_count": 0,
            "assigned_count": 0,
            "review_count": 0,
            "rejected_count": 0,
        },
    }


def build_structure_slot_assignment_from_pdf(
    *,
    pdf_bytes: bytes,
    template: dict[str, Any],
    strategy: str,
    skeleton_rows: list[dict[str, Any]] | None,
) -> dict[str, Any] | None:
    facility_id = str(template.get("facility_id") or "").strip()
    if not facility_id:
        return _blocked_structure_slot_assignment(
            strategy=strategy,
            template=template,
            blockers=["facility_id_missing_for_structure_slot"],
            warnings=["structure_slot_assignment", "structure_slot_requires_facility_template"],
            skeleton_rows=skeleton_rows,
        )
    signature_validation = validate_facility_template_signature(
        template,
        facility_id=facility_id,
        skeleton_rows=skeleton_rows,
        require_signature=True,
    )
    if signature_validation.get("blockers"):
        return _blocked_structure_slot_assignment(
            strategy=strategy,
            template=template,
            blockers=list(signature_validation.get("blockers") or []),
            warnings=["structure_slot_assignment", "structure_template_signature_blocked"],
            facility_id=facility_id,
            skeleton_rows=skeleton_rows,
            signature_validation=signature_validation,
        )
    week_sheet_name = _week_sheet_name_from_template(template)
    dpi = int(template.get("hakodate_structure_dpi") or template.get("grid_dpi") or 200)
    try:
        structure_grid, worksheet = _structure_grid_for_facility_template(
            facility_id=facility_id,
            week_sheet_name=week_sheet_name,
            dpi=dpi,
        )
    except Exception as exc:  # noqa: BLE001
        return _blocked_structure_slot_assignment(
            strategy=strategy,
            template=template,
            blockers=[str(exc) or "structure_template_unavailable"],
            warnings=["structure_slot_assignment", "structure_template_unavailable"],
            facility_id=facility_id,
            skeleton_rows=skeleton_rows,
            signature_validation=signature_validation,
        )

    structure_columns = [float(value) for value in (structure_grid.get("column_edges") or [])]
    structure_rows = [float(value) for value in (structure_grid.get("row_edges") or [])]
    if len(structure_columns) < 2 or len(structure_rows) < 2:
        return _blocked_structure_slot_assignment(
            strategy=strategy,
            template=template,
            blockers=["structure_grid_invalid"],
            facility_id=facility_id,
            skeleton_rows=skeleton_rows,
            signature_validation=signature_validation,
            grid={"structure_grid": structure_grid},
        )
    actual_grid = detect_table_grid(
        pdf_bytes,
        {
            **template,
            "grid_dpi": dpi,
            "grid_auto_table_box": True,
            "grid_auto_use_raw_edges": True,
            "grid_table_search_region": [0.0, 0.05, 1.0, 0.98],
        },
    )
    if actual_grid is None:
        return _blocked_structure_slot_assignment(
            strategy=strategy,
            template=template,
            blockers=["actual_grid_not_found_for_structure_slot"],
            facility_id=facility_id,
            skeleton_rows=skeleton_rows,
            signature_validation=signature_validation,
            grid={"structure_grid": structure_grid, "actual_grid": None},
        )
    actual_grid_data = _grid_payload(actual_grid)
    actual_box = actual_grid_data.get("table_box") or []
    structure_box = structure_grid.get("table_box") or []
    if not (isinstance(actual_box, list) and len(actual_box) == 4 and isinstance(structure_box, list) and len(structure_box) == 4):
        return _blocked_structure_slot_assignment(
            strategy=strategy,
            template=template,
            blockers=["structure_alignment_box_missing"],
            facility_id=facility_id,
            skeleton_rows=skeleton_rows,
            signature_validation=signature_validation,
            grid={"structure_grid": structure_grid, "actual_grid": actual_grid_data},
        )
    try:
        page_image = _render_pdf_page_image(pdf_bytes, template)
    except HakodateOcrUnavailable:
        return _blocked_structure_slot_assignment(
            strategy=strategy,
            template=template,
            blockers=["structure_slot_pdf_render_unavailable"],
            facility_id=facility_id,
            skeleton_rows=skeleton_rows,
            signature_validation=signature_validation,
            grid={"structure_grid": structure_grid, "actual_grid": actual_grid_data},
        )
    if page_image is None:
        return _blocked_structure_slot_assignment(
            strategy=strategy,
            template=template,
            blockers=["structure_slot_pdf_page_missing"],
            facility_id=facility_id,
            skeleton_rows=skeleton_rows,
            signature_validation=signature_validation,
            grid={"structure_grid": structure_grid, "actual_grid": actual_grid_data},
        )
    actual_columns_raw = [float(value) for value in (actual_grid_data.get("column_edges") or [])]
    actual_rows_raw = [float(value) for value in (actual_grid_data.get("row_edges") or [])]
    match_tolerance = float(template.get("hakodate_grid_line_match_tolerance") or 0.04)
    match_tolerance = max(0.005, min(match_tolerance, 0.12))
    actual_columns, column_blocker = _match_detected_edges_to_expected(
        detected_edges=actual_columns_raw,
        expected_cell_count=len(structure_columns) - 1,
        table_start=float(actual_box[0]),
        table_end=float(actual_box[2]),
        tolerance=match_tolerance,
    )
    column_alignment_evidence: dict[str, Any] = {
        "method": "detected_edges_exact_or_template_matched",
        "raw_detected_edge_count": len(actual_columns_raw),
    }
    column_edges_source = "detected_template_matched"
    if column_blocker:
        actual_columns, column_alignment_evidence, column_blocker = _detect_dense_vertical_edges_from_table_image(
            page_image,
            table_box=actual_box,
            expected_edge_count=len(structure_columns),
            structure_edges=structure_columns,
            column_band=(
                float(template.get("hakodate_dense_column_band_start") or 0.0),
                float(template.get("hakodate_dense_column_band_end") or 0.9),
            ),
        )
        column_edges_source = "dense_vertical_lines"
    if column_blocker:
        actual_columns, column_alignment_evidence, column_blocker = _project_structure_edges_to_detected_table_box(
            structure_edges=structure_columns,
            structure_start=float(structure_box[0]),
            structure_end=float(structure_box[2]),
            detected_start=float(actual_box[0]),
            detected_end=float(actual_box[2]),
            expected_edge_count=len(structure_columns),
        )
        column_edges_source = "structure_table_box_projection"
    actual_rows, row_blocker = _match_detected_edges_to_expected(
        detected_edges=actual_rows_raw,
        expected_cell_count=len(structure_rows) - 1,
        table_start=float(actual_box[1]),
        table_end=float(actual_box[3]),
        tolerance=match_tolerance,
    )
    row_alignment_evidence: dict[str, Any] = {
        "method": "detected_edges_exact_or_template_matched",
        "raw_detected_edge_count": len(actual_rows_raw),
    }
    row_edges_source = "detected_template_matched"
    if row_blocker:
        actual_rows, row_alignment_evidence, row_blocker = _detect_dense_horizontal_edges_from_table_image(
            page_image,
            table_box=actual_box,
            expected_edge_count=len(structure_rows),
            row_band=(
                float(template.get("hakodate_dense_row_band_start") or 0.36),
                float(template.get("hakodate_dense_row_band_end") or 0.98),
            ),
        )
        row_edges_source = "dense_horizontal_lines"
    if row_blocker:
        actual_rows, row_alignment_evidence, row_blocker = _project_structure_edges_to_detected_table_box(
            structure_edges=structure_rows,
            structure_start=float(structure_box[1]),
            structure_end=float(structure_box[3]),
            detected_start=float(actual_box[1]),
            detected_end=float(actual_box[3]),
            expected_edge_count=len(structure_rows),
        )
        row_edges_source = "structure_table_box_projection"
    alignment_blockers = []
    if column_blocker:
        alignment_blockers.append(f"structure_slot_column_alignment_{column_blocker}")
    if row_blocker:
        alignment_blockers.append(f"structure_slot_row_alignment_{row_blocker}")
    alignment_confidence = float(actual_grid_data.get("confidence") or 0.0)
    if alignment_blockers:
        return _blocked_structure_slot_assignment(
            strategy=strategy,
            template=template,
            blockers=alignment_blockers,
            warnings=["structure_slot_assignment", "structure_slot_alignment_blocked"],
            facility_id=facility_id,
            skeleton_rows=skeleton_rows,
            signature_validation=signature_validation,
            grid={
                "structure_grid": structure_grid,
                "actual_grid": {
                    **actual_grid_data,
                    "structure_slot_column_edges_source": "unmatched",
                    "structure_slot_row_edges_source": "unmatched",
                    "structure_slot_column_alignment_evidence": column_alignment_evidence,
                    "structure_slot_row_alignment_evidence": row_alignment_evidence,
                },
            },
        )
    if column_edges_source == "structure_table_box_projection" or row_edges_source == "structure_table_box_projection":
        return _blocked_structure_slot_assignment(
            strategy=strategy,
            template=template,
            blockers=["structure_slot_alignment_unverified_by_actual_cell_grid"],
            warnings=["structure_slot_assignment", "structure_slot_alignment_blocked"],
            facility_id=facility_id,
            skeleton_rows=skeleton_rows,
            signature_validation=signature_validation,
            grid={
                "structure_grid": structure_grid,
                "actual_grid": actual_grid_data,
                "actual_structure_slot_grid": {
                    "table_box": actual_box,
                    "column_edges": actual_columns,
                    "row_edges": actual_rows,
                    "column_edges_source": column_edges_source,
                    "row_edges_source": row_edges_source,
                    "column_alignment_evidence": column_alignment_evidence,
                    "row_alignment_evidence": row_alignment_evidence,
                    "confidence": alignment_confidence,
                },
            },
        )
    column_slots = _column_slots_from_worksheet(
        worksheet,
        col_count=len(structure_columns) - 1,
        template=template,
    )
    physical_row_map = _workbook_physical_row_map(worksheet, row_count=len(structure_rows) - 1)
    merged_cells = _worksheet_merged_cell_map(worksheet)
    quantity_columns = order_form_service._worksheet_quantity_column_indexes(worksheet)  # noqa: SLF001
    structure_merge_regions = _worksheet_merge_regions_for_grid(
        worksheet,
        row_edges=structure_rows,
        column_edges=structure_columns,
        quantity_columns=quantity_columns,
    )
    actual_merge_regions = _worksheet_merge_regions_for_grid(
        worksheet,
        row_edges=actual_rows,
        column_edges=actual_columns,
        quantity_columns=quantity_columns,
    )
    actual_quantity_merge_regions = [
        region
        for region in actual_merge_regions
        if bool(region.get("is_quantity")) and int(region.get("row_span") or 0) > 1
    ]
    skeleton_by_index = {
        idx: row
        for idx, row in enumerate(skeleton_rows or [])
        if isinstance(row, dict)
    }

    assignments: list[dict[str, Any]] = []
    review: list[dict[str, Any]] = []
    rejected: list[dict[str, Any]] = []
    seen: set[tuple[str, str]] = set()
    cell_regions: list[dict[str, Any]] = []
    debug_dir = str(template.get("hakodate_debug_output_dir") or "").strip() or None
    alignment_requires_review = (
        column_edges_source == "structure_table_box_projection"
        or row_edges_source == "structure_table_box_projection"
    )
    for row_index, row_meta in physical_row_map.items():
        row_meta = physical_row_map.get(row_index)
        if not isinstance(row_meta, dict):
            continue
        data_index = max(0, int(row_index) - 2)
        skeleton_row = skeleton_by_index.get(data_index)
        if not isinstance(skeleton_row, dict):
            continue
        row_id = str(skeleton_row.get("row_id") or skeleton_row.get("id") or f"grid_row_{row_index}").strip()
        for col_index, slot in enumerate(column_slots):
            if not isinstance(slot, dict) or slot.get("role") != "quantity":
                continue
            field = str(slot.get("slot_name") or "").strip()
            if not field.startswith("qty."):
                continue
            worksheet_col = int(slot.get("worksheet_col_index") or (col_index + 1))
            worksheet_row = int(row_meta.get("worksheet_row") or 0)
            structure_cell_box, merged_cell = _merged_or_single_cell_bbox(
                row_index=row_index,
                col_index=col_index,
                worksheet_row=worksheet_row,
                worksheet_col=worksheet_col,
                row_edges=structure_rows,
                column_edges=structure_columns,
                merged_cells=merged_cells,
            )
            fax_cell_box, _actual_merged_cell = _merged_or_single_cell_bbox(
                row_index=row_index,
                col_index=col_index,
                worksheet_row=worksheet_row,
                worksheet_col=worksheet_col,
                row_edges=actual_rows,
                column_edges=actual_columns,
                merged_cells=merged_cells,
            )
            fax_cell_center = [
                (fax_cell_box[0] + fax_cell_box[2]) / 2.0,
                (fax_cell_box[1] + fax_cell_box[3]) / 2.0,
            ]
            structure_cell_center = [
                (structure_cell_box[0] + structure_cell_box[2]) / 2.0,
                (structure_cell_box[1] + structure_cell_box[3]) / 2.0,
            ]
            evidence = {
                "method": "template_structure_grid_to_detected_fax_grid_edges",
                "structure_table_box": structure_box,
                "fax_table_box": actual_box,
                "structure_column_edges_source": structure_grid.get("column_edges_source"),
                "structure_row_edges_source": structure_grid.get("row_edges_source"),
                "fax_column_edges_source": column_edges_source,
                "fax_row_edges_source": row_edges_source,
                "fax_column_alignment_evidence": column_alignment_evidence,
                "fax_row_alignment_evidence": row_alignment_evidence,
                "fax_grid_confidence": alignment_confidence,
            }
            sheet_cell = (
                f"{get_column_letter(worksheet_col)}"
                f"{worksheet_row}"
            )
            cell_id = f"r{row_index}_c{col_index}_{field}"
            cell_region = {
                "row_id": row_id,
                "grid_row_index": row_index,
                "data_row_index": row_index - 2,
                "col_index": col_index,
                "field": field,
                "field_label": str(slot.get("label") or field),
                "sheet_cell": sheet_cell,
                "date": skeleton_row.get("date") or row_meta.get("effective_date") or row_meta.get("date"),
                "daypart": skeleton_row.get("daypart") or row_meta.get("daypart"),
                "menu_name": skeleton_row.get("menu_name") or row_meta.get("menu_name"),
                "structure_cell_bbox": structure_cell_box,
                "structure_cell_center": structure_cell_center,
                "fax_cell_bbox": fax_cell_box,
                "fax_cell_center": fax_cell_center,
                "alignment_confidence": alignment_confidence,
                "alignment_evidence": evidence,
            }
            if merged_cell:
                cell_region["merged_cell"] = merged_cell
            cell_regions.append(cell_region)
            try:
                ocr = _ocr_quantity_cell(
                    page_image=page_image,
                    fax_cell_bbox=fax_cell_box,
                    template=template,
                    cell_id=cell_id,
                    debug_dir=debug_dir,
                )
            except HakodateOcrUnavailable:
                return _blocked_structure_slot_assignment(
                    strategy=strategy,
                    template=template,
                    blockers=["structure_slot_cell_ocr_unavailable"],
                    facility_id=facility_id,
                    skeleton_rows=skeleton_rows,
                    signature_validation=signature_validation,
                    grid={"structure_grid": structure_grid, "actual_grid": actual_grid_data},
                )
            base_item = {
                **cell_region,
                "value_text": str(ocr.text or "").strip(),
                "value_normalized": ocr.normalized,
                "ocr_scope": "cell_crop",
                "ocr_engine_bbox_used_for_assignment": False,
                "ocr_cell_crop_bbox_px": ocr.crop_bbox_px,
                "ocr_cell_crop_image_path": ocr.crop_image_path,
                "confidence": round(max(0.0, min(0.99, alignment_confidence)), 3),
                "structure_row": row_meta,
                "slot": slot,
            }
            if not ocr.normalized:
                rejected.append({**base_item, "decision": "rejected", "reason": "empty_cell"})
                continue
            if not _is_numeric_text(ocr.normalized):
                rejected.append({**base_item, "decision": "rejected", "reason": "non_numeric_cell_ocr"})
                continue
            key = (row_id, field)
            if alignment_requires_review:
                review.append(
                    {
                        **base_item,
                        "decision": "review",
                        "reason": "alignment_not_verified_by_actual_cell_grid",
                    }
                )
                continue
            if key in seen:
                review.append({**base_item, "decision": "review", "reason": "duplicate_quantity_cell_ocr"})
                continue
            seen.add(key)
            assignments.append({**base_item, "decision": "assigned"})
    overlay_path = _save_structure_slot_overlay(
        page_image=page_image,
        cells=cell_regions,
        actual_box=actual_box,
        structure_box=structure_box,
        actual_columns=actual_columns,
        actual_rows=actual_rows,
        worksheet=worksheet,
        merge_regions=actual_merge_regions,
        debug_dir=debug_dir,
    )
    if debug_dir:
        try:
            from pathlib import Path

            Path(debug_dir).mkdir(parents=True, exist_ok=True)
            (Path(debug_dir) / "cell_regions.json").write_text(
                json.dumps(cell_regions, ensure_ascii=False, indent=2),
                encoding="utf-8",
            )
        except Exception as exc:  # noqa: BLE001
            logger.warning("Failed to save Hakodate cell regions: {}", str(exc))
    status = "auto_assignable"
    if review:
        status = "review_required"
    if not assignments and not review:
        status = "blocked"
    return {
        "version": "1",
        "strategy": strategy,
        "status": status,
        "blockers": [] if status != "blocked" else ["no_structure_slot_quantity_assignments"],
        "warnings": [
            "structure_slot_assignment",
            *(["structure_slot_alignment_requires_review"] if alignment_requires_review else []),
        ],
        "fields": derive_row_fields_from_template(template),
        "quantity_fields": [slot.get("slot_name") for slot in column_slots if slot.get("role") == "quantity"],
        "template_signature": signature_validation,
        "header_rows": 2,
        "grid": {
            "structure_grid": structure_grid,
            "actual_grid": actual_grid_data,
            "actual_structure_slot_grid": {
                "table_box": actual_box,
                "column_edges": actual_columns,
                "row_edges": actual_rows,
                "column_edges_source": column_edges_source,
                "row_edges_source": row_edges_source,
                "column_alignment_evidence": column_alignment_evidence,
                "row_alignment_evidence": row_alignment_evidence,
                "confidence": alignment_confidence,
                "merged_quantity_cells": actual_quantity_merge_regions,
            },
            "structure_merge_regions": structure_merge_regions,
            "column_slots": column_slots,
            "physical_row_map_count": len(physical_row_map),
        },
        "artifacts": {
            "overlay_image_path": overlay_path,
            "cell_regions_path": f"{debug_dir}/cell_regions.json" if debug_dir else None,
        },
        "assignments": assignments,
        "review_candidates": review,
        "rejected_candidates": rejected,
        "metrics": {
            "ocr_cell_count": len(cell_regions),
            "ocr_numeric_candidate_count": len(assignments) + len(review),
            "assigned_count": len(assignments),
            "review_count": len(review),
            "rejected_count": len(rejected),
            "structure_grid_column_count": len(structure_columns) - 1,
            "structure_grid_row_count": len(structure_rows) - 1,
            "merged_quantity_cell_count": len(actual_quantity_merge_regions),
        },
    }


def _extract_tesseract_tokens(pdf_bytes: bytes, template: dict[str, Any]) -> list[HakodateToken]:
    _ = pdf_bytes, template
    raise HakodateOcrUnavailable("tesseract_ocr_removed")


def build_hakodate_assignment(
    *,
    tokens: list[HakodateToken | dict[str, Any]],
    grid: GridDetectionResult | dict[str, Any],
    template: dict[str, Any],
    strategy: str | None = None,
    skeleton_rows: list[dict[str, Any]] | None = None,
) -> dict[str, Any]:
    resolved_strategy = normalize_quantity_assignment_strategy(
        strategy or resolve_quantity_assignment_strategy(template)
    )
    fields = derive_row_fields_from_template(template)
    quantity_indexes = _quantity_field_indexes(fields)
    grid_data = _grid_payload(grid)
    column_edges = [float(value) for value in (grid_data.get("column_edges") or [])]
    row_edges = [float(value) for value in (grid_data.get("row_edges") or [])]
    header_rows = _header_row_count(template)
    synthetic_grid_edges = _uses_synthetic_grid_edges(grid_data)

    warnings: list[str] = []
    blockers: list[str] = []
    if resolved_strategy == "legacy":
        blockers.append("hakodate_strategy_not_enabled")
    signature_validation: dict[str, Any] | None = None
    if resolved_strategy != "legacy":
        signature_validation = validate_facility_template_signature(
            template,
            facility_id=str(template.get("facility_id") or "").strip() or None,
            skeleton_rows=skeleton_rows,
            require_signature=True,
        )
        blockers.extend(signature_validation.get("blockers") or [])
        warnings.extend(signature_validation.get("warnings") or [])
    if not fields:
        blockers.append("template_fields_missing")
    if not quantity_indexes:
        blockers.append("quantity_fields_missing")
    if len(column_edges) < 2:
        blockers.append("grid_column_edges_missing")
    if len(row_edges) < 2:
        blockers.append("grid_row_edges_missing")
    if fields and len(column_edges) >= 2 and len(column_edges) - 1 != len(fields):
        warnings.append("grid_column_count_mismatch")

    normalized_tokens: list[HakodateToken] = []
    for token in tokens or []:
        if isinstance(token, HakodateToken):
            normalized_tokens.append(token)
            continue
        if not isinstance(token, dict):
            continue
        try:
            bbox = token.get("bbox") or token.get("box") or []
            if isinstance(bbox, list) and len(bbox) == 4:
                x0, y0, x1, y1 = [float(value) for value in bbox]
                x = float(token.get("x") if token.get("x") is not None else (x0 + x1) / 2.0)
                y = float(token.get("y") if token.get("y") is not None else (y0 + y1) / 2.0)
            else:
                x = float(token.get("x"))
                y = float(token.get("y"))
                bbox = [x, y, x, y]
        except Exception:
            continue
        normalized_tokens.append(
            HakodateToken(
                text=str(token.get("text") or ""),
                x=x,
                y=y,
                bbox=[float(value) for value in bbox],
                confidence=(
                    float(token.get("confidence"))
                    if token.get("confidence") is not None
                    else None
                ),
            )
        )

    assignments: list[dict[str, Any]] = []
    rejected: list[dict[str, Any]] = []
    review: list[dict[str, Any]] = []
    if blockers:
        return {
            "version": "1",
            "strategy": resolved_strategy,
            "status": "blocked",
            "blockers": blockers,
            "warnings": warnings,
            "fields": fields,
            "quantity_fields": [fields[idx] for idx in quantity_indexes if idx < len(fields)],
            "template_signature": signature_validation,
            "grid": grid_data,
            "assignments": [],
            "review_candidates": [],
            "rejected_candidates": [],
            "metrics": {
                "ocr_numeric_candidate_count": sum(1 for token in normalized_tokens if _is_numeric_text(token.text)),
                "assigned_count": 0,
                "review_count": 0,
                "rejected_count": 0,
            },
        }

    skeleton_by_data_index = {
        idx: row
        for idx, row in enumerate(skeleton_rows or [])
        if isinstance(row, dict)
    }
    seen_assignment_keys: dict[tuple[int, str], int] = {}
    min_edge_margin = float(template.get("hakodate_min_edge_margin_ratio", 0.08) or 0.08)
    for token in normalized_tokens:
        if not _is_numeric_text(token.text):
            continue
        row_index = _find_band_index(float(token.y), row_edges)
        col_index = _find_band_index(float(token.x), column_edges)
        candidate_base = {
            "token": _token_payload(token),
            "row_index": row_index,
            "col_index": col_index,
        }
        if row_index is None or col_index is None:
            rejected.append({**candidate_base, "reason": "outside_grid"})
            continue
        if row_index < header_rows:
            rejected.append({**candidate_base, "reason": "header_row"})
            continue
        if col_index not in quantity_indexes or col_index >= len(fields):
            rejected.append({**candidate_base, "reason": "non_quantity_cell"})
            continue
        cell_box = _cell_bbox(
            row_index=row_index,
            col_index=col_index,
            row_edges=row_edges,
            column_edges=column_edges,
        )
        distance = _distance_to_center(token, cell_box)
        margin_ratio = _edge_margin_ratio(token, cell_box)
        field = fields[col_index]
        data_row_index = row_index - header_rows
        skeleton_row = skeleton_by_data_index.get(data_row_index, {})
        row_id = (
            str(skeleton_row.get("row_id") or "").strip()
            or str(skeleton_row.get("id") or "").strip()
            or f"grid_row_{data_row_index}"
        )
        item = {
            "row_id": row_id,
            "grid_row_index": row_index,
            "data_row_index": data_row_index,
            "col_index": col_index,
            "field": field,
            "field_label": _field_label(field),
            "value_text": str(token.text or "").strip(),
            "value_normalized": _normalize_digits(token.text),
            "ocr_bbox": list(token.bbox),
            "ocr_center": [float(token.x), float(token.y)],
            "cell_bbox": cell_box,
            "cell_center": [(cell_box[0] + cell_box[2]) / 2.0, (cell_box[1] + cell_box[3]) / 2.0],
            "distance_norm": round(distance, 6),
            "edge_margin_ratio": round(margin_ratio, 6),
            "confidence": round(min(0.99, max(0.0, 0.72 + margin_ratio)), 3),
            "decision": "assigned",
            "date": skeleton_row.get("date"),
            "daypart": skeleton_row.get("daypart"),
            "menu_name": skeleton_row.get("menu_name"),
        }
        if margin_ratio < min_edge_margin:
            item["decision"] = "review"
            item["reason"] = "near_cell_boundary"
            review.append(item)
            continue
        if synthetic_grid_edges:
            item["decision"] = "review"
            item["reason"] = "synthetic_grid_edges_require_review"
            review.append(item)
            continue
        assignment_key = (data_row_index, field)
        if assignment_key in seen_assignment_keys:
            first_index = seen_assignment_keys[assignment_key]
            if 0 <= first_index < len(assignments):
                first = dict(assignments[first_index])
                first["decision"] = "review"
                first["reason"] = "duplicate_quantity_candidates"
                review.append(first)
                assignments[first_index] = {**assignments[first_index], "decision": "superseded_by_review"}
            item["decision"] = "review"
            item["reason"] = "duplicate_quantity_candidates"
            review.append(item)
            continue
        seen_assignment_keys[assignment_key] = len(assignments)
        assignments.append(item)

    active_assignments = [item for item in assignments if item.get("decision") == "assigned"]
    status = "auto_assignable"
    if review:
        status = "review_required"
    if not active_assignments and not review:
        status = "blocked"
        blockers.append("no_quantity_assignments")
    return {
        "version": "1",
        "strategy": resolved_strategy,
        "status": status,
        "blockers": blockers,
        "warnings": warnings,
        "fields": fields,
        "quantity_fields": [fields[idx] for idx in quantity_indexes if idx < len(fields)],
        "template_signature": signature_validation,
        "header_rows": header_rows,
        "grid": grid_data,
        "assignments": active_assignments,
        "review_candidates": review,
        "rejected_candidates": rejected,
        "metrics": {
            "ocr_numeric_candidate_count": len(active_assignments) + len(review) + len(rejected),
            "assigned_count": len(active_assignments),
            "review_count": len(review),
            "rejected_count": len(rejected),
            "grid_column_count": max(0, len(column_edges) - 1),
            "grid_row_count": max(0, len(row_edges) - 1),
        },
    }


def build_hakodate_assignment_from_pdf(
    *,
    pdf_bytes: bytes,
    template: dict[str, Any],
    strategy: str | None = None,
    skeleton_rows: list[dict[str, Any]] | None = None,
) -> dict[str, Any]:
    resolved_strategy = normalize_quantity_assignment_strategy(
        strategy or resolve_quantity_assignment_strategy(template)
    )
    if resolved_strategy == "legacy":
        return build_hakodate_assignment(
            tokens=[],
            grid={"column_edges": [], "row_edges": [], "confidence": 0.0},
            template=template,
            strategy=resolved_strategy,
            skeleton_rows=skeleton_rows,
        )
    alignment_mode = str(template.get("hakodate_alignment_mode") or "structure_slot").strip()
    if alignment_mode != "structure_slot":
        return {
            "version": "1",
            "strategy": resolved_strategy,
            "status": "blocked",
            "blockers": ["hakodate_structure_slot_required"],
            "warnings": ["hakodate_detected_grid_token_assignment_disabled"],
            "fields": derive_row_fields_from_template(template),
            "quantity_fields": [],
            "grid": None,
            "assignments": [],
            "review_candidates": [],
            "rejected_candidates": [],
            "metrics": {
                "ocr_numeric_candidate_count": 0,
                "assigned_count": 0,
                "review_count": 0,
                "rejected_count": 0,
            },
        }
    structure_slot_result = build_structure_slot_assignment_from_pdf(
        pdf_bytes=pdf_bytes,
        template=template,
        strategy=resolved_strategy,
        skeleton_rows=skeleton_rows,
    )
    if isinstance(structure_slot_result, dict):
        return structure_slot_result
    return _blocked_structure_slot_assignment(
        strategy=resolved_strategy,
        template=template,
        blockers=["structure_slot_assignment_unavailable"],
        warnings=["structure_slot_assignment", "hakodate_detected_grid_token_assignment_disabled"],
        skeleton_rows=skeleton_rows,
    )
