import csv
import json
import math
import calendar
import re
import time
import zipfile
from copy import copy
from datetime import date as dt_date, datetime, timedelta
from functools import lru_cache
from io import BytesIO
from pathlib import Path
from typing import Dict, Any
from xml.etree import ElementTree as ET
from xml.sax.saxutils import escape as xml_escape
from uuid import uuid4

import pandas as pd
from loguru import logger
from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter

from src.db import session_scope
from src.models.output import Bag, LabelRow, DeliveryNote, ManufacturingAggregateRow
from src.models.order_output_artifact import OrderBaggingResult
from src.models.order_sheet_draft import OrderSheetDraft
from src.models.order_workflow_state import OrderWorkflowState
from src.services.order_service import get_order_by_id, get_order_menu_snapshot
from src.services import (
    config_service,
    menu_service,
    menu_rule_service,
    order_service,
    daily_output_override_service,
    order_output_artifact_service,
)
from src.services.menu_vocabulary import bucket_diet_type_for_aggregation
from src.services.storage_service import load_bytes_from_uri

OUTPUT_DIR = Path("/tmp/orders-outputs")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
_OUTPUT_LOOKUP_CACHE_SECONDS = 300

DATA_DIR = Path(__file__).resolve().parents[1] / "data"
DAILY_DELIVERY_REFERENCE_TEMPLATE = DATA_DIR / "delivery_note_templates" / "daily_delivery_note_reference.xlsx"
_XLSX_MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_XLSX_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
_XLSX_PACKAGE_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
_XLSX_CONTENT_TYPE_NS = "http://schemas.openxmlformats.org/package/2006/content-types"
ET.register_namespace("", _XLSX_MAIN_NS)
ET.register_namespace("r", _XLSX_REL_NS)

DAILY_DELIVERY_SHEET_BY_FACILITY_ID = {
    "FAC00001": "大和なでしこ",
    "FAC00002": "なごみ",
    "FAC00003": "春日苑",
    "FAC00004": "ふれあいの丘",
    "FAC00005": "池袋病院",
    "FAC00006": "アイテラス",
    "FAC00007": "百々家",
    "FAC00008": "佐古",
    "FAC00009": "そよかぜ",
    "FAC00010": "山城",
    "FAC00011": "四万十ピア",
    "FAC00012": "グランフォレスト",
    "FAC00013": "いこいの森",
    "FAC00014": "さくら",
    "FAC00015": "四万十ピア",
    "FAC00016": "いこいの森",
    "FAC636208": "長生苑",
}

DAILY_LABEL_SHEET_BY_FACILITY_ID = {
    "FAC00001": "大和なでしこ",
    "FAC00002": "なごみ",
    "FAC00003": "春日苑",
    "FAC00004": "ふれあいの丘",
    "FAC00005": "池袋病院",
    "FAC00006": "藍TERRACE",
    "FAC00007": "百々家",
    "FAC00008": "佐古グループホーム",
    "FAC00009": "そよかぜ",
    "FAC00010": "山城グループホーム",
    "FAC00012": "グランフォレスト",
    "FAC00014": "湘南さくら病院",
    "FAC00015": "四万十ピア",
    "FAC00016": "いこいの森",
    "FAC636208": "ケアホーム長生苑",
}

DAILY_LABEL_SHEET_ORDER = [
    "メニュー",
    "藍TERRACE",
    "百々家",
    "池袋病院",
    "大和なでしこ",
    "春日苑",
    "四万十ピア",
    "山城グループホーム",
    "佐古グループホーム",
    "なごみ",
    "いこいの森",
    "そよかぜ",
    "グランフォレスト",
    "ふれあいの丘",
    "湘南さくら病院",
    "ケアホーム長生苑",
]

DAILY_LABEL_COMPARABLE_DIETS_BY_FACILITY_ID = {
    "FAC00001": {"no_meat", "no_fish"},
    "FAC00002": {"no_meat", "no_fish"},
    "FAC00003": {"soft", "mixer", "no_fish"},
    "FAC00004": {"daycare", "staff", "no_meat", "no_fish", "no_fried", "forbidden_other"},
    "FAC00006": {"soft", "mixer", "no_meat", "no_fish"},
    "FAC00007": {"no_meat", "no_fish"},
    "FAC00008": {"soft", "mixer"},
    "FAC00009": {"soft", "mixer"},
    "FAC00010": {"soft", "mixer"},
    "FAC00011": {"no_meat", "no_fish"},
    "FAC00012": {"no_meat", "no_fish"},
    "FAC00013": {"diabetes", "no_fish"},
    "FAC00014": {"staff", "no_meat", "no_fish", "sesame_allergy"},
    "FAC00015": {"no_meat", "no_fish"},
    "FAC00016": {"diabetes", "no_fish"},
    "FAC636208": {"mixer", "no_meat", "no_fish"},
}

_EXPANDED_CELL_COPY_ENABLED_CACHE: dict[tuple[str, str], bool] = {}

DEFAULT_LABEL_FIELDS = [
    "呼び出し番号",
    "発行枚数",
    "賞味期限",
    "時間",
    "メニュー",
    "温・冷",
    "商品名１",
    "商品名２",
    "内容量",
    "内容詳細",
    "",
]
LEGACY_LABEL_FIELDS = {
    "facility_name",
    "expiry_date",
    "storage_mode",
    "meal_slot",
    "menu_category",
    "product_name",
    "quantity",
    "details",
    "maker_info",
    "notice",
}

_GARNISH_SPLIT_RE = re.compile(r"\s*(?:添え|添[)）:：])\s*")

_DAILY_LABEL_MENU_DEFAULTS = {
    "ごぼうと竹輪の煮物": ("朝", "副菜①", "温菜", 70, "g"),
    "竹輪の煮物": ("朝", "副菜①", "温菜", 70, "g"),
    "いんげんの味噌和え": ("朝", "副菜②", "冷菜", 40, "g"),
    "豚肉と白菜のすき煮": ("昼", "主菜", "温菜", 100, "g"),
    "さつま芋の天ぷら": ("昼", "副菜①", "温菜", 2, "個"),
    "さつまいもレモン煮": ("昼", "副菜①", "温菜", 40, "g"),
    "ﾌﾞﾛｯｺﾘｰのちりめん和え": ("昼", "副菜②", "冷菜", 40, "g"),
    "煮込みハンバーグ": ("夕", "主菜", "温菜", 100, "g"),
    "ジャーマンポテト": ("夕", "副菜①", "温菜", 40, "g"),
    "ほうれん草の和え物": ("夕", "副菜②", "冷菜", 40, "g"),
    "ソース": ("夕", "ソース", "冷菜", 5, "g"),
}

DAILY_LABEL_SHEET_MAX_ROWS = {
    "メニュー": 603,
    "藍TERRACE": 627,
    "百々家": 622,
    "池袋病院": 619,
    "大和なでしこ": 623,
    "春日苑": 620,
    "四万十ピア": 623,
    "山城グループホーム": 622,
    "佐古グループホーム": 607,
    "なごみ": 623,
    "いこいの森": 619,
    "そよかぜ": 637,
    "グランフォレスト": 622,
    "ふれあいの丘": 629,
    "湘南さくら病院": 627,
    "ケアホーム長生苑": 621,
}

DAILY_LABEL_MENU_ROWS = [
    ("朝", "副菜①", "温菜", "ごぼうと竹輪の煮物", "", 70),
    ("朝", "副菜②", "冷菜", "いんげんの味噌和え", "", 40),
    ("昼", "主菜", "温菜", "豚肉と白菜のすき煮", "", 100),
    ("昼", "副菜①", "温菜", "", "", 40),
    ("昼", "副菜①", "温菜", "さつま芋の天ぷら", "", "2個"),
    ("昼", "副菜②", "冷菜", "ﾌﾞﾛｯｺﾘｰのちりめん和え", "", 40),
    ("夕", "主菜", "温菜", "煮込みハンバーグ", "", 100),
    ("夕", "ソース", "温菜", "", "", 50),
    ("夕", "副菜①", "温菜", "ジャーマンポテト", "", 40),
    ("夕", "副菜②", "冷菜", "ほうれん草の和え物", "", 40),
]


def _ensure_date(value):
    if value is None:
        return None
    if hasattr(value, "isoformat"):
        return value
    try:
        return pd.to_datetime(value).date()
    except Exception:
        return None


def _serialize_for_json(value: Any) -> Any:
    if isinstance(value, dict):
        return {k: _serialize_for_json(v) for k, v in value.items()}
    if isinstance(value, list):
        return [_serialize_for_json(item) for item in value]
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return value


def _format_number(value: float | int | None) -> str:
    if value is None:
        return ""
    try:
        num = float(value)
    except Exception:
        return str(value)
    if num.is_integer():
        return str(int(num))
    return f"{num:.2f}".rstrip("0").rstrip(".")


def _format_jp_date(value: Any) -> str:
    if value is None:
        return ""
    if hasattr(value, "year") and hasattr(value, "month") and hasattr(value, "day"):
        return f"{value.year}年{value.month}月{value.day}日"
    try:
        parsed = pd.to_datetime(value).date()
        return f"{parsed.year}年{parsed.month}月{parsed.day}日"
    except Exception:
        return str(value)


def _coerce_to_date(value: Any) -> dt_date | None:
    if value is None:
        return None
    if isinstance(value, dt_date):
        return value
    try:
        return pd.to_datetime(value).date()
    except Exception:
        return None


def _add_months(base_date: dt_date, months: int) -> dt_date:
    if months == 0:
        return base_date
    month_index = (base_date.month - 1) + months
    target_year = base_date.year + (month_index // 12)
    target_month = (month_index % 12) + 1
    target_day = min(base_date.day, calendar.monthrange(target_year, target_month)[1])
    return dt_date(target_year, target_month, target_day)


def _resolve_label_expiry_date(
    meal_date: Any,
    label_profile: dict | None,
) -> Any:
    base_date = _coerce_to_date(meal_date)
    if base_date is None:
        return meal_date
    profile = label_profile if isinstance(label_profile, dict) else {}
    raw_months = profile.get("expiry_offset_months", 0)
    raw_days = profile.get("expiry_offset_days", 0)
    try:
        offset_months = int(raw_months or 0)
    except Exception:
        offset_months = 0
    try:
        offset_days = int(raw_days or 0)
    except Exception:
        offset_days = 0
    expiry_date = _add_months(base_date, offset_months)
    if offset_days:
        expiry_date = expiry_date + timedelta(days=offset_days)
    return expiry_date


def _normalize_temp_label(temp: str | None) -> str:
    if not temp:
        return ""
    value = str(temp)
    if "冷" in value:
        return "冷菜"
    if "温" in value:
        return "温菜"
    lowered = value.lower()
    if lowered in {"hot", "warm"}:
        return "温菜"
    if lowered in {"cold", "chilled"}:
        return "冷菜"
    return value


def _normalize_unit_type(unit_type: str | None) -> str | None:
    if not unit_type:
        return None
    raw = str(unit_type).strip()
    lowered = raw.lower()
    if "g" in lowered or "グラム" in raw:
        return "g"
    if "切" in raw or "枚" in raw or lowered in {"cut", "slice", "slices"}:
        return "切"
    if "個" in raw or lowered in {"count", "piece", "pieces"}:
        return "個"
    return raw


_AREA_TRANSLATION = str.maketrans("０１２３４５６７８９ｆＦ", "0123456789fF")


def _normalize_diet_key(value: str | None) -> str | None:
    if not value:
        return None
    raw = str(value).strip()
    if not raw:
        return None
    lowered = raw.lower()
    if ("袋" in raw or "bag" in lowered) and (
        "常食" in raw or "通常" in raw or lowered in {"regular_bag", "bag"}
    ):
        return "regular_bag"
    if "常食" in raw or "通常" in raw or lowered in {"regular", "standard"}:
        return "regular"
    if "軟菜" in raw or "やわ" in raw or "ﾔﾜ" in raw or "ヤワ" in raw or lowered in {"soft"}:
        return "soft"
    if "ミキサ" in raw or "ﾐｷｻ" in raw or lowered in {"mixer"}:
        return "mixer"
    if "通所" in raw or lowered in {"daycare"}:
        return "regular"
    if "職員" in raw or lowered in {"staff"}:
        return "regular"
    if "揚げ物禁" in raw or "揚物禁" in raw or lowered in {"no_fried", "nofried"}:
        return "no_fried"
    if (
        ("肉" in raw or "meat" in lowered)
        and ("卵" in raw or "玉子" in raw or "egg" in lowered)
        and ("魚" in raw or "鯖" in raw or "さば" in raw or "fish" in lowered)
    ) or "肉卵魚禁" in raw:
        return "forbidden_other"
    if "禁" in raw and "肉" in raw:
        return "no_meat"
    if "禁" in raw and "魚" in raw:
        return "no_fish"
    return lowered


def _normalize_delivery_diet_key(value: str | None) -> str | None:
    if not value:
        return None
    raw = str(value).strip()
    lowered = raw.lower()
    if "糖尿" in raw or lowered in {"diabetes", "diabetic"}:
        return "diabetes"
    if "通所" in raw or lowered == "daycare":
        return "daycare"
    if "職員" in raw or lowered == "staff":
        return "staff"
    if "小口" in raw:
        return "小口"
    return _normalize_diet_key(raw)


def _build_output_diet_type_map(facility_config: dict | None) -> dict[tuple[str, str], str]:
    if not isinstance(facility_config, dict):
        return {}
    override = facility_config.get("fax_template_override")
    if not isinstance(override, dict):
        return {}
    columns = override.get("columns")
    if not isinstance(columns, list):
        return {}
    result: dict[tuple[str, str], str] = {}
    for column in columns:
        if not isinstance(column, dict):
            continue
        if str(column.get("role") or "").strip() != "quantity":
            continue
        source_diet = _normalize_diet_key(column.get("diet_type"))
        source_area = str(column.get("area_id") or "X").strip().upper() or "X"
        output_diet = _normalize_diet_key(
            column.get("output_diet_type")
            or column.get("aggregation_diet_type")
            or column.get("daily_output_diet_type")
        )
        if not source_diet or not output_diet or output_diet == source_diet:
            continue
        result[(source_diet, source_area)] = output_diet
    return result


def _apply_output_diet_type_overrides(
    lines: list[dict],
    facility_config: dict | None,
) -> list[dict]:
    overrides = _build_output_diet_type_map(facility_config)
    if not overrides:
        return lines
    updated_lines: list[dict] = []
    for line in lines:
        diet_key = _normalize_diet_key(line.get("diet_type"))
        area_key = str(line.get("area_id") or "X").strip().upper() or "X"
        output_diet = overrides.get((diet_key or "", area_key)) or overrides.get((diet_key or "", "X"))
        if not output_diet:
            updated_lines.append(line)
            continue
        updated = dict(line)
        updated.setdefault("source_diet_type", line.get("diet_type"))
        updated["diet_type"] = output_diet
        updated_lines.append(updated)
    return updated_lines


def _normalize_area_id(value: str | None) -> str | None:
    if not value:
        return None
    text = str(value).strip()
    if not text:
        return None
    text = text.translate(_AREA_TRANSLATION)
    text = re.sub(r"[\\s　]+", "", text)
    text = text.replace("階", "F").replace("ｆ", "F").replace("f", "F")
    match = re.search(r"(\\d)F", text, re.IGNORECASE)
    if match:
        return f"{match.group(1)}F"
    return text.upper()


def _build_area_alias_map(facility_config: dict | None) -> dict[str, str]:
    if not facility_config:
        return {}
    areas = facility_config.get("areas") or []
    if not isinstance(areas, list):
        return {}
    aliases: dict[str, str] = {}
    for area in areas:
        if not isinstance(area, dict):
            continue
        area_id = area.get("area_id") or area.get("id") or ""
        name = area.get("name") or ""
        canonical = _normalize_area_id(area_id) or _normalize_area_id(name)
        if not canonical:
            continue
        for candidate in (area_id, name):
            norm = _normalize_area_id(candidate)
            if norm:
                aliases[norm] = canonical
    return aliases


def _resolve_area_key(value: str | None, aliases: dict[str, str]) -> str | None:
    norm = _normalize_area_id(value)
    if not norm:
        return None
    return aliases.get(norm, norm)


def _infer_delivery_column_meta(name: str | None) -> tuple[str | None, str | None]:
    if not name:
        return None, None
    raw = str(name)
    if "揚げ物" in raw or "揚物" in raw:
        diet = "no_fried"
    elif "肉禁" in raw:
        diet = "no_meat"
    elif "魚禁" in raw:
        diet = "no_fish"
    elif "その他" in raw:
        diet = "forbidden_other"
    elif "禁食" in raw:
        diet = "禁食"
    else:
        diet = _normalize_delivery_diet_key(raw)
    area = None
    match = re.search(r"(\\d)\\s*(?:f|ｆ|Ｆ|階)", raw, re.IGNORECASE)
    if match:
        area = f"{match.group(1)}F"
    elif "月" in raw:
        area = "3F"
    elif "花" in raw:
        area = "2F"
    return diet, area


def _extract_qty_and_unit(value: Any, unit_type: str | None) -> tuple[float | None, str | None]:
    if value is None:
        return None, _normalize_unit_type(unit_type)
    if isinstance(value, (int, float)):
        return float(value), _normalize_unit_type(unit_type)
    text = str(value).strip()
    if not text:
        return None, _normalize_unit_type(unit_type)
    match = re.search(r"[-+]?[0-9]*\\.?[0-9]+", text)
    qty = float(match.group()) if match else None
    inferred_unit = None
    if "g" in text or "ｇ" in text or "グラム" in text:
        inferred_unit = "g"
    elif "切" in text or "枚" in text:
        inferred_unit = "切"
    elif "個" in text:
        inferred_unit = "個"
    return qty, _normalize_unit_type(unit_type) or inferred_unit


def _format_amount(value: float | int | None, unit_type: str | None) -> str:
    if value is None:
        return ""
    suffix = _normalize_unit_type(unit_type)
    formatted = _format_number(value)
    if suffix:
        return f"{formatted}{suffix}"
    return formatted


def _format_servings(quantity: float | int | None) -> str:
    if quantity is None:
        return ""
    return f"{_format_number(quantity)}人前"


def _label_area_suffix(area_id: Any, facility_id: Any = None) -> str:
    text = str(area_id or "").strip()
    normalized = _normalize_area_id(text)
    facility = str(facility_id or "").strip()
    if facility in {"FAC00008", "FAC00009", "FAC00010"}:
        if normalized == "2F":
            return "2階"
        if normalized == "3F":
            return "3階"
    if normalized == "2F":
        return "花"
    if normalized == "3F":
        return "月"
    return ""


def _label_menu_defaults(menu_name: Any) -> tuple[str | None, str | None, str | None, float | None, str | None]:
    name = str(menu_name or "").strip()
    if name in _DAILY_LABEL_MENU_DEFAULTS:
        daypart, category, temp, qty, unit = _DAILY_LABEL_MENU_DEFAULTS[name]
        return daypart, category, temp, float(qty), unit
    return None, None, None, None, None


def _daily_label_display_diet_key(bag: dict) -> str:
    source_diet = str(bag.get("source_diet_type") or "").strip()
    raw_diet = str(bag.get("diet_type") or "").strip()
    for candidate in (source_diet, raw_diet):
        if not candidate:
            continue
        if "糖尿" in candidate or candidate.lower() in {"diabetes", "diabetic"}:
            return "diabetes"
        if "通所" in candidate or candidate.lower() == "daycare":
            return "daycare"
        if "職員" in candidate or candidate.lower() == "staff":
            return "staff"
        if "ゴマ" in candidate or "ごま" in candidate or candidate.lower() in {"sesame_allergy", "sesame"}:
            return "sesame_allergy"
        normalized = _normalize_diet_key(candidate)
        if normalized:
            return normalized
    return ""


def _daily_label_category_suffix_for_bag(bag: dict) -> str:
    diet_key = _daily_label_display_diet_key(bag)
    if not diet_key or diet_key in {"regular", "regular_bag", "standard"}:
        return ""
    facility_id = str(bag.get("facility") or "").strip()
    configured_diets = bag.get("daily_label_comparable_diet_types")
    comparable_diets = (
        {str(item).strip() for item in configured_diets if str(item).strip()}
        if isinstance(configured_diets, list)
        else DAILY_LABEL_COMPARABLE_DIETS_BY_FACILITY_ID.get(facility_id, set())
    )
    if diet_key not in comparable_diets:
        return ""
    return {
        "soft": "軟菜",
        "mixer": "ミキサー",
        "soft_mixer": "ミキサー",
        "daycare": "通所",
        "staff": "職員",
        "diabetes": "糖尿",
        "no_meat": "肉禁",
        "no_fish": "魚禁",
        "no_fried": "揚げ物禁",
        "forbidden_other": "禁",
        "sesame_allergy": "ごま禁",
    }.get(diet_key, "")


def _daily_label_comparable_diet_types_for_facility(facility_config: dict | None, facility_id: str | None) -> list[str]:
    if isinstance(facility_config, dict):
        configured = facility_config.get("daily_label_comparable_diet_types")
        if isinstance(configured, list):
            return [str(item).strip() for item in configured if str(item).strip()]
    return sorted(DAILY_LABEL_COMPARABLE_DIETS_BY_FACILITY_ID.get(str(facility_id or "").strip(), set()))


def _apply_daily_label_facility_rules_to_bags(
    bags: list[dict],
    facility_config: dict | None,
    facility_id: str | None,
) -> list[dict]:
    comparable_diets = _daily_label_comparable_diet_types_for_facility(facility_config, facility_id)
    if not comparable_diets:
        return bags
    updated_bags: list[dict] = []
    for bag in bags:
        updated = dict(bag)
        updated["daily_label_comparable_diet_types"] = comparable_diets
        updated_bags.append(updated)
    return updated_bags


def _append_daily_label_category_suffix(category: str, suffix: str) -> str:
    if not category or not suffix:
        return category
    if suffix in category:
        return category
    return f"{category}（{suffix}）"


def _normalize_daily_label_category_text(category: str) -> str:
    text = str(category or "").strip()
    if text in {"主", "主菜", "主Ａ", "主A"}:
        return "主菜"
    if text in {"副", "副①", "副1", "副菜1", "副菜①"}:
        return "副菜①"
    if text in {"副②", "副2", "副菜2", "副菜②"}:
        return "副菜②"
    return text


def _label_category_for_bag(bag: dict, product_name: str, diet_type: str) -> str:
    _daypart, default_category, _temp, _qty, _unit = _label_menu_defaults(product_name)
    category = str(bag.get("menu_category") or default_category or "").strip()
    category = _normalize_daily_label_category_text(category)
    if category == "副菜":
        category = default_category or category
    if not category:
        category = default_category or ""
    suffix = _daily_label_category_suffix_for_bag(bag)
    if category in {"添え", "付属品"}:
        parent_category = _normalize_daily_label_category_text(
            bag.get("parent_menu_category")
            or bag.get("garnish_parent_category")
            or bag.get("accessory_parent_category")
            or ""
        )
        if parent_category.startswith("主菜") or parent_category.startswith("副菜"):
            category = f"{parent_category} 添え"
    if suffix and (category.startswith("主菜") or category.startswith("副菜")):
        category = _append_daily_label_category_suffix(category, suffix)
    return category


def _daily_label_display_product(product_name: str, diet_type: str) -> str:
    if diet_type not in {"mixer", "soft", "soft_mixer"}:
        return product_name
    replacements = {
        "ごぼうと竹輪の煮物": "竹輪の煮物",
        "さつま芋の天ぷら": "さつまいもレモン煮",
    }
    return replacements.get(product_name, product_name)


def _daily_label_amount_cell(amount: float | None, unit: str | None) -> Any:
    if amount is None:
        return ""
    return _format_amount(amount, unit)


def _is_daily_label_excluded_bag(bag: dict) -> bool:
    diet_key = _daily_label_display_diet_key(bag)
    return diet_key in {"no_meat", "no_fish", "no_fried", "forbidden_other", "forbidden"}


def _resolve_label_fields(label_profile: dict) -> tuple[list[str], str]:
    fields = label_profile.get("label_fields")
    if isinstance(fields, list) and any(field in LEGACY_LABEL_FIELDS for field in fields):
        return fields, "legacy"
    return list(DEFAULT_LABEL_FIELDS), "jp"

def _safe_qty(line: dict, zero_as_empty: bool) -> float | None:
    qty = line.get("quantity_corrected")
    if qty is None:
        qty = line.get("quantity_original")
    if qty is None:
        qty = line.get("quantity")
    if qty is None:
        return None
    if zero_as_empty and qty <= 0:
        return None
    return qty


def _delivery_lines_have_structural_slots(lines: list[dict], quantity_rules: dict) -> bool:
    zero_as_empty = quantity_rules.get("zero_as_empty", True)
    seen_quantity_line = False
    for line in lines:
        if _safe_qty(line, zero_as_empty) is None:
            continue
        seen_quantity_line = True
        if not _ensure_date(line.get("date")):
            return False
        if not _normalize_delivery_daypart(line.get("daypart")):
            return False
        if not _normalize_delivery_category_label(line.get("menu_category")):
            return False
    return seen_quantity_line


def _format_menu_unit(qty: float | int | None, unit_type: str | None) -> str | None:
    if qty is None or unit_type is None:
        return None
    try:
        qty_value = float(qty)
    except Exception:
        return None
    if qty_value.is_integer():
        qty_str = str(int(qty_value))
    else:
        qty_str = str(qty_value)
    normalized = _normalize_unit_type(unit_type)
    suffix = "g" if normalized == "g" else ("個" if normalized == "個" else ("切" if normalized == "切" else normalized))
    return f"{qty_str}{suffix}"


def _build_label_details(bag: dict) -> str:
    parts: list[str] = []
    area = bag.get("area_id")
    if area:
        parts.append(str(area))
    unit_str = _format_menu_unit(bag.get("menu_qty_per_serving"), bag.get("menu_unit_type"))
    if unit_str:
        parts.append(unit_str)
    temp = bag.get("menu_temp_type")
    if temp:
        parts.append(str(temp))
    return " / ".join(parts)


def _build_menu_name_aliases(value: object) -> list[str]:
    text = str(value or "").strip().strip("　")
    if not text:
        return []
    aliases = [text]
    match = _GARNISH_SPLIT_RE.search(text)
    if match:
        base = text[:match.start()].strip().strip("　")
        if base and base not in aliases:
            aliases.append(base)
    return aliases


def _output_lookup_cache_bucket() -> int:
    return int(time.time() // _OUTPUT_LOOKUP_CACHE_SECONDS)


@lru_cache(maxsize=512)
def _cached_menu_items_for_week(
    week_value: str,
    facility_id: str,
    cache_bucket: int,
) -> tuple[tuple[tuple[str, Any], ...], ...]:
    del cache_bucket
    items = order_service._collect_menu_items_for_week(week_value, facility_id or None)  # noqa: SLF001
    return tuple(tuple(sorted(dict(item).items())) for item in items if isinstance(item, dict))


@lru_cache(maxsize=512)
def _cached_menu_entries_for_week(
    week_value: str,
    facility_id: str,
    cache_bucket: int,
) -> tuple[tuple[tuple[str, Any], ...], ...]:
    del cache_bucket
    entries = order_service._collect_menu_entries_for_week(week_value, facility_id or None)  # noqa: SLF001
    return tuple(tuple(sorted(dict(entry).items())) for entry in entries if isinstance(entry, dict))


@lru_cache(maxsize=512)
def _cached_menu_defaults(
    names: tuple[str, ...],
    facility_id: str,
    cache_bucket: int,
) -> tuple[tuple[str, tuple[tuple[str, Any], ...]], ...]:
    del cache_bucket
    defaults = menu_service.resolve_menu_defaults(list(names), facility_id or None)
    return tuple(
        (name, tuple(sorted(dict(defaults.get(name, {})).items())))
        for name in names
    )


@lru_cache(maxsize=32)
def _cached_active_menu_rules(cache_bucket: int) -> tuple[tuple[tuple[str, Any], ...], ...]:
    del cache_bucket
    rules = menu_rule_service.list_active_rules()
    return tuple(tuple(sorted(dict(rule).items())) for rule in rules if isinstance(rule, dict))


def _cached_tuple_rows_to_dicts(rows: tuple[tuple[tuple[str, Any], ...], ...]) -> list[dict]:
    return [dict(row) for row in rows]


def _collect_cached_menu_items_for_week(week_value: str | None, facility_id: str | None) -> list[dict]:
    if not week_value:
        return []
    rows = _cached_menu_items_for_week(
        str(week_value),
        str(facility_id or ""),
        _output_lookup_cache_bucket(),
    )
    return _cached_tuple_rows_to_dicts(rows)


def _collect_cached_menu_entries_for_week(week_value: str | None, facility_id: str | None) -> list[dict]:
    if not week_value:
        return []
    rows = _cached_menu_entries_for_week(
        str(week_value),
        str(facility_id or ""),
        _output_lookup_cache_bucket(),
    )
    return _cached_tuple_rows_to_dicts(rows)


def _resolve_cached_menu_defaults(names: list[str], facility_id: str | None) -> dict[str, dict]:
    unique_names = tuple(dict.fromkeys(str(name or "").strip() for name in names if str(name or "").strip()))
    if not unique_names:
        return {}
    rows = _cached_menu_defaults(
        unique_names,
        str(facility_id or ""),
        _output_lookup_cache_bucket(),
    )
    return {name: dict(payload) for name, payload in rows}


def _list_cached_active_menu_rules() -> list[dict]:
    rows = _cached_active_menu_rules(_output_lookup_cache_bucket())
    return _cached_tuple_rows_to_dicts(rows)


def _menu_item_matches_context(item: dict, line: dict) -> tuple[int, int, str] | None:
    item_daypart = _normalize_delivery_daypart(item.get("daypart"))
    line_daypart = _normalize_delivery_daypart(line.get("daypart"))
    item_category = _normalize_delivery_category_label(item.get("category"))
    line_category = _normalize_delivery_category_label(line.get("menu_category"))
    item_diet = str(item.get("diet_type") or "").strip()
    line_diet = str(line.get("diet_type") or line.get("menu_diet_type") or "").strip()
    score = 0
    specificity = 0
    if item_daypart:
        specificity += 1
        if line_daypart and item_daypart == line_daypart:
            score += 80
        elif line_daypart:
            return None
    if item_category:
        specificity += 1
        if line_category and item_category == line_category:
            score += 40
        elif line_category and _is_specific_delivery_category(item_category):
            score -= 4
    if item_diet:
        specificity += 1
        if line_diet and item_diet == line_diet:
            score += 20
        elif line_diet and item_diet not in {"regular", "常食", "普通食"}:
            return None
        elif line_diet:
            score -= 2
    return score, specificity, str(item.get("id") or "")


def _select_menu_item_for_line(line: dict, menu_items_by_key: dict[str, list[dict]]) -> dict | None:
    name_key = _normalize_menu_key(line.get("menu_name"))
    candidates = menu_items_by_key.get(name_key) or []
    if not candidates:
        return None
    ranked: list[tuple[tuple[int, int, str], dict]] = []
    for item in candidates:
        score = _menu_item_matches_context(item, line)
        if score is not None:
            ranked.append((score, item))
    if not ranked:
        return None
    ranked.sort(key=lambda pair: pair[0], reverse=True)
    return ranked[0][1]


def _apply_menu_overrides(lines: list[dict], menu_items: list[dict]) -> list[dict]:
    if not menu_items:
        return lines
    index: dict[str, list[dict]] = {}
    for item in menu_items:
        for alias in _build_menu_name_aliases(item.get("name")):
            key = _normalize_menu_key(alias)
            if key:
                index.setdefault(key, []).append(item)
    if not index:
        return lines
    enriched: list[dict] = []
    for line in lines:
        item = _select_menu_item_for_line(line, index)
        if not item:
            enriched.append(line)
            continue
        updated = dict(line)
        if item.get("daypart") and not updated.get("daypart"):
            updated["daypart"] = item.get("daypart")
        item_category = _normalize_delivery_category_label(item.get("category"))
        if item_category and (
            _is_specific_delivery_category(item_category)
            or not _is_specific_delivery_category(updated.get("menu_category"))
        ):
            updated["menu_category"] = item_category
        if item.get("unit_type"):
            updated["menu_unit_type"] = item.get("unit_type")
        if item.get("qty_per_serving") is not None:
            updated["menu_qty_per_serving"] = item.get("qty_per_serving")
            updated["_menu_qty_source_daypart"] = item.get("daypart")
            updated["_menu_qty_source_category"] = item.get("category")
            updated["_monthly_menu_item_override_applied"] = True
        if item.get("bag_max_qty") is not None:
            updated["menu_bag_max_qty"] = item.get("bag_max_qty")
        if item.get("bag_max_unit"):
            updated["menu_bag_max_unit"] = item.get("bag_max_unit")
        if item.get("temp_type"):
            updated["menu_temp_type"] = item.get("temp_type")
        if item.get("condiments") is not None:
            updated["condiments"] = item.get("condiments")
        enriched.append(updated)
    return enriched


def _apply_menu_snapshot(lines: list[dict], snapshot_items: dict) -> list[dict]:
    if not snapshot_items:
        return lines
    index: dict[str, dict] = {}
    for name, item in snapshot_items.items():
        for alias in _build_menu_name_aliases(name):
            key = _normalize_menu_key(alias)
            if key and key not in index:
                index[key] = item
    enriched: list[dict] = []
    for line in lines:
        name_key = _normalize_menu_key(line.get("menu_name"))
        item = index.get(name_key)
        if not item:
            enriched.append(line)
            continue
        updated = dict(line)
        if item.get("daypart") and not updated.get("daypart"):
            updated["daypart"] = item.get("daypart")
        item_category = _normalize_delivery_category_label(item.get("category"))
        if item_category and (
            _is_specific_delivery_category(item_category)
            or not _is_specific_delivery_category(updated.get("menu_category"))
        ):
            updated["menu_category"] = item_category
        if item.get("unit_type"):
            updated["menu_unit_type"] = item.get("unit_type")
        if item.get("qty_per_serving") is not None:
            updated["menu_qty_per_serving"] = item.get("qty_per_serving")
            updated["_menu_qty_source_daypart"] = item.get("daypart")
            updated["_menu_qty_source_category"] = item.get("category")
            updated["_monthly_menu_item_override_applied"] = True
        if item.get("bag_max_qty") is not None:
            updated["menu_bag_max_qty"] = item.get("bag_max_qty")
        if item.get("bag_max_unit"):
            updated["menu_bag_max_unit"] = item.get("bag_max_unit")
        if item.get("temp_type"):
            updated["menu_temp_type"] = item.get("temp_type")
        if item.get("condiments") is not None:
            updated["condiments"] = item.get("condiments")
        enriched.append(updated)
    return enriched


def _build_menu_entry_indexes(menu_entries: list[dict]) -> tuple[dict[tuple[str, str, str], dict], dict[tuple[str, str], list[dict]]]:
    exact: dict[tuple[str, str, str], dict] = {}
    by_date_name: dict[tuple[str, str], list[dict]] = {}
    for entry in menu_entries:
        line_date = _ensure_date(entry.get("menu_date"))
        if not line_date:
            continue
        name_key = _normalize_menu_key(entry.get("name"))
        if not name_key:
            continue
        date_key = line_date.isoformat()
        by_date_name.setdefault((date_key, name_key), []).append(entry)
        daypart_key = _normalize_delivery_daypart(entry.get("daypart"))
        if daypart_key:
            exact[(date_key, daypart_key, name_key)] = entry
    return exact, by_date_name


def _resolve_menu_entry_override(
    line: dict,
    exact_index: dict[tuple[str, str, str], dict],
    date_name_index: dict[tuple[str, str], list[dict]],
) -> dict | None:
    line_date = _ensure_date(line.get("date"))
    if not line_date:
        return None
    name_key = _normalize_menu_key(line.get("menu_name"))
    if not name_key:
        return None
    date_key = line_date.isoformat()
    daypart_key = _normalize_delivery_daypart(line.get("daypart"))
    if daypart_key:
        matched = exact_index.get((date_key, daypart_key, name_key))
        if matched:
            return matched
    candidates = date_name_index.get((date_key, name_key)) or []
    if len(candidates) == 1:
        return candidates[0]
    return None


def _apply_menu_entry_overrides(lines: list[dict], menu_entries: list[dict]) -> list[dict]:
    if not menu_entries:
        return lines
    exact_index, date_name_index = _build_menu_entry_indexes(menu_entries)
    if not exact_index and not date_name_index:
        return lines
    enriched: list[dict] = []
    for line in lines:
        entry = _resolve_menu_entry_override(line, exact_index, date_name_index)
        if not entry:
            enriched.append(line)
            continue
        updated = dict(line)
        entry_daypart = _normalize_delivery_daypart(entry.get("daypart"))
        if entry_daypart:
            updated["daypart"] = entry_daypart
        if entry.get("category"):
            updated["menu_category"] = entry.get("category")
        enriched.append(updated)
    return enriched


def _normalize_output_daypart(value: object) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    if text.startswith("朝"):
        return "朝"
    if text.startswith("昼"):
        return "昼"
    if text.startswith("夕"):
        return "夕"
    return text


def _menu_entry_category_label(entry: dict) -> str:
    normalized = _normalize_delivery_category_label(entry.get("category"))
    daypart = _normalize_output_daypart(entry.get("daypart"))
    try:
        slot_index = int(entry.get("slot_index"))
    except Exception:
        return normalized
    if daypart == "朝" and slot_index > 0:
        return {1: "副菜1", 2: "副菜2"}.get(slot_index, normalized)
    if normalized in {"副菜1", "副菜2", "主菜"}:
        return normalized
    if daypart in {"朝", "昼"}:
        return {0: "副菜1", 1: "副菜2"}.get(slot_index, normalized)
    if daypart == "夕":
        return {0: "副菜1", 1: "副菜2", 2: "主菜"}.get(slot_index, normalized)
    return normalized


def _is_specific_delivery_category(value: object) -> bool:
    return _normalize_delivery_category_label(value) in {"副菜1", "副菜2", "主菜"}


def _build_menu_entry_indexes(menu_entries: list[dict]) -> tuple[dict[tuple[str, str, str], dict], dict[tuple[str, str], dict]]:
    by_exact: dict[tuple[str, str, str], dict] = {}
    by_date_name: dict[tuple[str, str], dict | None] = {}
    for entry in menu_entries:
        menu_date = str(entry.get("menu_date") or "").strip()
        menu_name = str(entry.get("name") or "").strip()
        daypart = _normalize_output_daypart(entry.get("daypart"))
        if not menu_date or not menu_name:
            continue
        for alias in _build_menu_name_aliases(menu_name):
            alias_key = _normalize_menu_key(alias)
            if not alias_key:
                continue
            if daypart:
                by_exact[(menu_date, daypart, alias_key)] = entry
            date_name_key = (menu_date, alias_key)
            existing = by_date_name.get(date_name_key)
            if existing is None:
                by_date_name[date_name_key] = entry
            elif existing != entry:
                by_date_name[date_name_key] = {}
    resolved_by_date_name = {
        key: value
        for key, value in by_date_name.items()
        if isinstance(value, dict) and value
    }
    return by_exact, resolved_by_date_name


def _apply_menu_entry_overrides(lines: list[dict], menu_entries: list[dict]) -> list[dict]:
    if not menu_entries:
        return lines
    by_exact, by_date_name = _build_menu_entry_indexes(menu_entries)
    if not by_exact and not by_date_name:
        return lines
    enriched: list[dict] = []
    for line in lines:
        menu_name_key = _normalize_menu_key(line.get("menu_name"))
        line_date = _ensure_date(line.get("date"))
        date_key = line_date.isoformat() if line_date else ""
        daypart = _normalize_output_daypart(line.get("daypart"))
        entry = None
        if date_key and daypart and menu_name_key:
            entry = by_exact.get((date_key, daypart, menu_name_key))
        if entry is None and date_key and menu_name_key:
            entry = by_date_name.get((date_key, menu_name_key))
        if not entry:
            enriched.append(line)
            continue
        updated = dict(line)
        if entry.get("daypart") and not daypart:
            updated["daypart"] = _normalize_output_daypart(entry.get("daypart"))
        entry_category = _menu_entry_category_label(entry)
        if entry_category:
            updated["menu_category"] = entry_category
        updated["_monthly_entry_override_applied"] = True
        enriched.append(updated)
    return enriched


def _normalize_category_key(value: object) -> str:
    return str(value or "").strip()


def _clear_stale_menu_qty_from_monthly_entry(lines: list[dict]) -> list[dict]:
    enriched: list[dict] = []
    for line in lines:
        if (
            not line.get("_monthly_entry_override_applied")
            or line.get("_monthly_menu_item_override_applied")
            or line.get("menu_qty_per_serving") is None
        ):
            enriched.append(line)
            continue
        source_daypart = _normalize_output_daypart(line.get("_menu_qty_source_daypart"))
        current_daypart = _normalize_output_daypart(line.get("daypart"))
        source_category = _normalize_category_key(line.get("_menu_qty_source_category"))
        current_category = _normalize_category_key(line.get("menu_category"))
        daypart_conflict = bool(source_daypart and current_daypart and source_daypart != current_daypart)
        category_conflict = bool(source_category and current_category and source_category != current_category)
        if not daypart_conflict and not category_conflict:
            enriched.append(line)
            continue
        updated = dict(line)
        updated["menu_qty_per_serving"] = None
        enriched.append(updated)
    return enriched


def _apply_garnish_defaults(lines: list[dict]) -> list[dict]:
    enriched: list[dict] = []
    for line in lines:
        if _normalize_category_key(line.get("menu_category")) != "添え" or line.get("menu_qty_per_serving") is not None:
            enriched.append(line)
            continue
        updated = dict(line)
        updated["menu_unit_type"] = _normalize_unit_type(updated.get("menu_unit_type")) or "g"
        updated["menu_qty_per_serving"] = 30
        enriched.append(updated)
    return enriched


def _line_order_key(line: dict) -> tuple[int, str]:
    raw_index = line.get("_order_index")
    if raw_index is None:
        raw_index = line.get("source_row_index")
    try:
        order_index = int(raw_index)
    except Exception:
        order_index = 1_000_000
    return order_index, str(line.get("menu_name") or "")


def _apply_garnish_parent_categories(lines: list[dict]) -> list[dict]:
    grouped: dict[tuple[Any, str], list[dict]] = {}
    for line in lines:
        daypart = _normalize_output_daypart(line.get("daypart"))
        if not daypart:
            continue
        grouped.setdefault((line.get("date"), daypart), []).append(line)

    parent_by_id: dict[int, str] = {}
    for daypart_lines in grouped.values():
        current_parent = ""
        for line in sorted(daypart_lines, key=_line_order_key):
            category = _normalize_delivery_category_label(line.get("menu_category"))
            if category == "添え":
                if current_parent:
                    parent_by_id[id(line)] = current_parent
                continue
            normalized = _normalize_daily_label_category_text(category)
            if normalized.startswith("主菜") or normalized.startswith("副菜"):
                current_parent = normalized

    enriched: list[dict] = []
    for line in lines:
        if (
            _normalize_delivery_category_label(line.get("menu_category")) == "添え"
            and not line.get("parent_menu_category")
            and parent_by_id.get(id(line))
        ):
            updated = dict(line)
            updated["parent_menu_category"] = parent_by_id[id(line)]
            enriched.append(updated)
        else:
            enriched.append(line)
    return enriched


def _label_meal_slot_sequence(daypart: str) -> list[str]:
    if daypart == "朝":
        return ["副菜1", "副菜2"]
    if daypart in {"昼", "夕"}:
        return ["主菜", "副菜1", "副菜2"]
    return []


def _apply_label_meal_slot_categories(lines: list[dict]) -> list[dict]:
    grouped: dict[tuple[Any, str], list[dict]] = {}
    for line in lines:
        daypart = _normalize_output_daypart(line.get("daypart"))
        if not daypart:
            continue
        grouped.setdefault((line.get("date"), daypart), []).append(line)

    slot_by_id: dict[int, str] = {}
    for (_date_value, daypart), daypart_lines in grouped.items():
        slot_sequence = _label_meal_slot_sequence(daypart)
        if not slot_sequence:
            continue
        menu_slot_by_name: dict[str, str] = {}
        menu_names: list[str] = []
        for line in sorted(daypart_lines, key=_line_order_key):
            category = _normalize_delivery_category_label(line.get("menu_category"))
            if category == "添え":
                continue
            menu_name = str(line.get("menu_name") or "").strip()
            if not menu_name:
                continue
            if menu_name not in menu_slot_by_name:
                menu_names.append(menu_name)
                if len(menu_names) <= len(slot_sequence):
                    menu_slot_by_name[menu_name] = slot_sequence[len(menu_names) - 1]
                else:
                    menu_slot_by_name[menu_name] = category
            slot = menu_slot_by_name.get(menu_name)
            if slot:
                slot_by_id[id(line)] = slot

    enriched: list[dict] = []
    for line in lines:
        slot = slot_by_id.get(id(line))
        if slot:
            updated = dict(line)
            updated["menu_category"] = slot
            enriched.append(updated)
        else:
            enriched.append(line)
    return enriched


def _apply_menu_master_defaults(lines: list[dict], facility_id: str | None) -> list[dict]:
    menu_names = [str(line.get("menu_name") or "").strip() for line in lines if str(line.get("menu_name") or "").strip()]
    defaults = _resolve_cached_menu_defaults(menu_names, facility_id)
    if not defaults:
        return lines
    enriched: list[dict] = []
    for line in lines:
        payload = defaults.get(str(line.get("menu_name") or "").strip()) or {}
        if not payload:
            enriched.append(line)
            continue
        updated = dict(line)
        if not updated.get("menu_unit_type") and payload.get("unit_type"):
            updated["menu_unit_type"] = payload.get("unit_type")
        if updated.get("menu_qty_per_serving") is None and payload.get("qty_per_serving") is not None:
            updated["menu_qty_per_serving"] = payload.get("qty_per_serving")
        if not updated.get("menu_temp_type") and payload.get("temp_type"):
            updated["menu_temp_type"] = payload.get("temp_type")
        if payload.get("condiments") is not None:
            updated["condiments"] = payload.get("condiments")
        payload_category = _normalize_delivery_category_label(payload.get("category"))
        if not updated.get("_monthly_entry_override_applied") and payload_category and (
            _is_specific_delivery_category(payload_category)
            or not _is_specific_delivery_category(updated.get("menu_category"))
        ):
            updated["menu_category"] = payload_category
        enriched.append(updated)
    return enriched


def _normalize_condiments(value: object) -> list[str]:
    if not value:
        return []
    if isinstance(value, list):
        return [str(v).strip() for v in value if str(v).strip()]
    if isinstance(value, str):
        return [value.strip()] if value.strip() else []
    return []


def _apply_condiment_lines(lines: list[dict]) -> list[dict]:
    enriched: list[dict] = []
    for line in lines:
        updated = dict(line)
        condiments = _normalize_condiments(updated.get("condiments"))
        if condiments and not updated.get("note"):
            if any("ソース" in label for label in condiments):
                updated["note"] = "ソース"
        enriched.append(updated)
        if not condiments:
            continue
        for label in condiments:
            condiment_line = dict(updated)
            condiment_line["menu_name"] = label
            condiment_line["menu_category"] = "添え"
            condiment_line["parent_menu_name"] = updated.get("menu_name")
            condiment_line["parent_menu_category"] = updated.get("menu_category")
            condiment_line["condiments"] = []
            condiment_line["bag_type"] = "condiment"
            for field in (
                "menu_qty_per_serving",
                "menu_unit_type",
                "menu_temp_type",
                "actual_amount",
                "actual_unit_type",
                "menu_bag_max_qty",
                "menu_bag_max_unit",
                "_menu_qty_source_daypart",
                "_menu_qty_source_category",
                "_monthly_menu_item_override_applied",
            ):
                condiment_line.pop(field, None)
            enriched.append(condiment_line)
    return enriched


def _build_condiment_map(menu_names: list[str], facility_id: str | None) -> dict[str, list[str]]:
    if not menu_names:
        return {}
    defaults = _resolve_cached_menu_defaults(menu_names, facility_id)
    condiment_map: dict[str, list[str]] = {}
    for name in menu_names:
        payload = defaults.get(name, {})
        condiments = _normalize_condiments(payload.get("condiments"))
        if condiments:
            condiment_map[name] = condiments
    return condiment_map


def _apply_condiment_note(row: dict, condiments: list[str]) -> dict:
    if not condiments:
        return row
    row["_delivery_condiments"] = condiments
    return row


def _append_condiments_to_delivery_menu_name(menu_name: object, condiments: object) -> str:
    text = str(menu_name or "").strip()
    labels = _normalize_condiments(condiments)
    if not labels:
        return text
    suffix = " ".join(f"添）{label}" for label in labels)
    return f"{text} {suffix}".strip()


def _resolve_bag_types(facility_config: dict | None) -> list[dict]:
    bag_types = []
    if isinstance(facility_config, dict):
        bag_types = facility_config.get("bag_types") or []
    if not isinstance(bag_types, list):
        return []
    normalized: list[dict] = []
    for entry in bag_types:
        if not isinstance(entry, dict):
            continue
        bag_id = entry.get("bag_type_id")
        max_qty = entry.get("max_qty")
        unit = entry.get("unit")
        if not bag_id or max_qty is None:
            continue
        try:
            max_qty_value = float(max_qty)
        except Exception:
            continue
        normalized.append(
            {
                "bag_type_id": str(bag_id),
                "label": entry.get("label") or str(bag_id),
                "max_qty": max_qty_value,
                "unit": _normalize_unit_type(unit or "g") or "g",
            }
        )
    normalized.sort(key=lambda item: item["max_qty"])
    return normalized


def _apply_bag_size_defaults(lines: list[dict], bag_types: list[dict]) -> list[dict]:
    if not bag_types:
        return lines
    largest_by_unit: dict[str, dict] = {}
    for entry in bag_types:
        unit = _normalize_unit_type(entry.get("unit")) or "g"
        selected = largest_by_unit.get(unit)
        if not selected or float(entry.get("max_qty", 0)) > float(selected.get("max_qty", 0)):
            largest_by_unit[unit] = entry
    for line in lines:
        if line.get("bag_type") == "condiment":
            continue
        unit = _normalize_unit_type(line.get("menu_unit_type"))
        if not unit:
            continue
        if line.get("menu_bag_max_qty") is not None:
            continue
        largest = largest_by_unit.get(unit)
        if not largest:
            continue
        line["menu_bag_max_qty"] = largest["max_qty"]
        line["menu_bag_max_unit"] = largest.get("unit") or unit
    return lines


def _assign_bag_type_for_bags(bags: list[dict], bag_types: list[dict]) -> list[dict]:
    if not bag_types:
        return bags
    bag_types_by_unit: dict[str, list[dict]] = {}
    for entry in bag_types:
        unit = _normalize_unit_type(entry.get("unit")) or "g"
        bag_types_by_unit.setdefault(unit, []).append(entry)
    for entries in bag_types_by_unit.values():
        entries.sort(key=lambda item: item["max_qty"])
    for bag in bags:
        if bag.get("bag_type") == "condiment":
            continue
        unit = _normalize_unit_type(bag.get("menu_unit_type"))
        per_qty = bag.get("menu_qty_per_serving")
        servings = bag.get("quantity")
        if not unit or per_qty is None or servings is None:
            if not bag.get("bag_type"):
                bag["bag_type"] = "standard"
            continue
        candidates = bag_types_by_unit.get(unit)
        if not candidates:
            if not bag.get("bag_type"):
                bag["bag_type"] = "standard"
            continue
        try:
            total_weight = float(per_qty) * float(servings)
        except Exception:
            if not bag.get("bag_type"):
                bag["bag_type"] = "standard"
            continue
        selected = None
        for entry in candidates:
            if total_weight <= entry["max_qty"]:
                selected = entry
                break
        if not selected:
            selected = candidates[-1]
        bag["bag_type"] = selected["bag_type_id"]
    return bags


def _exception_applies(rule: dict, line: dict) -> bool:
    menu_pattern = rule.get("menu_pattern")
    if menu_pattern:
        if not _match_menu_pattern(line.get("menu_name") or "", menu_pattern, rule.get("match_type")):
            return False
    if rule.get("daypart") and rule.get("daypart") != line.get("daypart"):
        return False
    if rule.get("diet_type") and rule.get("diet_type") != line.get("diet_type"):
        return False
    if rule.get("category") and rule.get("category") != line.get("menu_category"):
        return False
    return True


def _apply_bagging_exceptions(lines: list[dict], facility_config: dict | None) -> list[dict]:
    if not isinstance(facility_config, dict):
        return lines
    exceptions = facility_config.get("bagging_exceptions") or []
    if not isinstance(exceptions, list) or not exceptions:
        return lines
    enriched = [dict(line) for line in lines]
    for rule in exceptions:
        if not isinstance(rule, dict):
            continue
        rule_type = rule.get("type") or rule.get("id")
        if rule_type == "quantity_multiplier":
            multiplier = rule.get("multiplier")
            try:
                multiplier = float(multiplier)
            except Exception:
                continue
            if multiplier <= 0:
                continue
            for line in enriched:
                if not _exception_applies(rule, line):
                    continue
                if line.get("quantity_original") is not None:
                    line["quantity_original"] = float(line["quantity_original"]) * multiplier
                if line.get("quantity_corrected") is not None:
                    line["quantity_corrected"] = float(line["quantity_corrected"]) * multiplier
                line.setdefault("change_note", "袋分け例外適用")
    return enriched


def build_order_lines_for_outputs(
    order: dict,
    *,
    include_expanded_copy: bool = True,
    allow_stale_draft_lines: bool = False,
    timings: dict[str, float] | None = None,
) -> list[dict]:
    total_started = time.perf_counter()
    facility_id = order.get("facility")
    week_value = (
        str(order.get("stored_week_value") or "").strip()
        or str(order.get("week_value") or "").strip()
        or str(order.get("persisted_week_value") or "").strip()
        or str(order.get("week") or "").strip()
        or str(order.get("week_code") or "").strip()
    )
    facility_config_started = time.perf_counter()
    facility_config = config_service.get_facility_config(facility_id) if facility_id else None
    if timings is not None:
        timings["build_order_lines_facility_config_ms"] = round(
            (time.perf_counter() - facility_config_started) * 1000,
            1,
        )
    raw_lines = order.get("lines", [])
    order_id = order.get("id")
    workflow_state = order.get("workflow_state") if isinstance(order.get("workflow_state"), dict) else {}
    workflow_warnings = {
        str(item).strip()
        for item in (workflow_state.get("warnings_json") or workflow_state.get("warnings") or [])
        if str(item).strip()
    }
    workflow_blockers = {
        str(item).strip()
        for item in (workflow_state.get("blockers_json") or workflow_state.get("blockers") or [])
        if str(item).strip()
    }
    draft_newer_than_lines = bool(
        workflow_state.get("ocr_draft_newer_than_lines")
        or workflow_state.get("draft_newer_than_lines")
        or "draft_newer_than_lines" in workflow_warnings
        or "draft_newer_than_lines" in workflow_blockers
    )
    use_existing_lines_for_stale_draft = (
        allow_stale_draft_lines
        and draft_newer_than_lines
        and isinstance(raw_lines, list)
        and bool(raw_lines)
    )
    if use_existing_lines_for_stale_draft:
        logger.warning(
            "Daily output read used existing order lines without workflow draft materialization",
            order_id=order_id,
            facility_id=facility_id,
            workflow_state=workflow_state.get("state"),
        )
    else:
        workflow_started = time.perf_counter()
        workflow_v2_lines = _workflow_v2_lines_for_outputs(order, raw_lines)
        if timings is not None:
            timings["build_order_lines_workflow_v2_ms"] = round(
                (time.perf_counter() - workflow_started) * 1000,
                1,
            )
        if workflow_v2_lines is not None:
            raw_lines = workflow_v2_lines
    week_sheet_name = order_service._week_sheet_name_from_week_value(week_value)  # noqa: SLF001
    facility_cache_key = (str(facility_id or ""), str(week_sheet_name or ""))
    expanded_copy_enabled = False
    if include_expanded_copy:
        if facility_cache_key in _EXPANDED_CELL_COPY_ENABLED_CACHE:
            expanded_copy_enabled = _EXPANDED_CELL_COPY_ENABLED_CACHE[facility_cache_key]
        else:
            expanded_copy_enabled = order_service._expanded_cell_same_daypart_copy_enabled(  # noqa: SLF001
                facility_config,
                week_sheet_name=week_sheet_name,
            )
            _EXPANDED_CELL_COPY_ENABLED_CACHE[facility_cache_key] = expanded_copy_enabled
    if expanded_copy_enabled:
        if order_id:
            expanded_started = time.perf_counter()
            materialization_candidate = order_service.build_confirm_materialization_candidate(order_id)
            if timings is not None:
                timings["build_order_lines_expanded_copy_ms"] = round(
                    (time.perf_counter() - expanded_started) * 1000,
                    1,
                )
            candidate_lines = (
                materialization_candidate.get("lines")
                if isinstance(materialization_candidate, dict)
                and not materialization_candidate.get("error")
                else None
            )
            if isinstance(candidate_lines, list) and candidate_lines:
                raw_lines = candidate_lines
    enrich_started = time.perf_counter()
    raw_lines = order_service._apply_change_override_priority_to_lines(raw_lines)  # noqa: SLF001
    raw_lines = _apply_garnish_lines(raw_lines)
    if timings is not None:
        timings["build_order_lines_pre_menu_ms"] = round(
            (time.perf_counter() - enrich_started) * 1000,
            1,
        )
    menu_items_started = time.perf_counter()
    menu_items = _collect_cached_menu_items_for_week(week_value, facility_id)
    if timings is not None:
        timings["build_order_lines_menu_items_ms"] = round(
            (time.perf_counter() - menu_items_started) * 1000,
            1,
        )
    menu_entries_started = time.perf_counter()
    menu_entries = _collect_cached_menu_entries_for_week(week_value, facility_id)
    if timings is not None:
        timings["build_order_lines_menu_entries_ms"] = round(
            (time.perf_counter() - menu_entries_started) * 1000,
            1,
        )
    snapshot_started = time.perf_counter()
    snapshot = get_order_menu_snapshot(order.get("id"))
    if timings is not None:
        timings["build_order_lines_snapshot_ms"] = round(
            (time.perf_counter() - snapshot_started) * 1000,
            1,
        )
    snapshot_items = snapshot.get("menu_items") if isinstance(snapshot, dict) else None
    if snapshot_items:
        order_lines = _apply_menu_snapshot(raw_lines, snapshot_items)
    else:
        order_lines = raw_lines
    overrides_started = time.perf_counter()
    order_lines = _apply_menu_entry_overrides(order_lines, menu_entries)
    order_lines = _apply_output_diet_type_overrides(order_lines, facility_config)
    # Current monthly/menu-master settings must win over stale confirmed snapshots.
    order_lines = _apply_menu_overrides(order_lines, menu_items)
    order_lines = _clear_stale_menu_qty_from_monthly_entry(order_lines)
    order_lines = _apply_menu_rules(order_lines, facility_id)
    order_lines = _apply_garnish_defaults(order_lines)
    order_lines = _apply_menu_master_defaults(order_lines, facility_id)
    order_lines = _apply_builtin_menu_defaults(order_lines)
    order_lines = daily_output_override_service.apply_overrides_to_lines(order_lines, facility_id)
    order_lines = _apply_bagging_exceptions(order_lines, facility_config)
    order_lines = _apply_condiment_lines(order_lines)
    order_lines = _apply_garnish_parent_categories(order_lines)
    order_lines = _apply_garnish_defaults(order_lines)
    order_lines = _apply_menu_master_defaults(order_lines, facility_id)
    order_lines = _apply_builtin_menu_defaults(order_lines)
    order_lines = _apply_label_meal_slot_categories(order_lines)
    bag_types = _resolve_bag_types(facility_config)
    order_lines = _apply_bag_size_defaults(order_lines, bag_types)
    if timings is not None:
        timings["build_order_lines_apply_overrides_ms"] = round(
            (time.perf_counter() - overrides_started) * 1000,
            1,
        )
        timings["build_order_lines_total_ms"] = round(
            (time.perf_counter() - total_started) * 1000,
            1,
        )
    return order_lines


def _workflow_v2_lines_for_outputs(order: dict, raw_lines: object) -> list[dict] | None:
    order_id = str(order.get("id") or "").strip()
    if not order_id:
        return None
    try:
        with session_scope() as session:
            workflow = session.get(OrderWorkflowState, order_id)
            if workflow is None or not workflow.draft_id:
                return None
            draft = session.get(OrderSheetDraft, workflow.draft_id)
            if draft is None or draft.order_id != order_id:
                if _workflow_v2_output_state_requires_saved_sheet(workflow.state):
                    raise ValueError("workflow-v2 saved sheet is missing")
                return None
            candidate = _workflow_v2_materialization_candidate(session, order, workflow, draft)
            if not isinstance(candidate, dict):
                if _workflow_v2_output_state_requires_saved_sheet(workflow.state):
                    raise ValueError("workflow-v2 saved sheet materialization is missing")
                return None
            error = str(candidate.get("error") or "").strip()
            if error:
                if _workflow_v2_output_state_requires_saved_sheet(workflow.state) or not raw_lines:
                    raise ValueError(f"workflow-v2 saved sheet materialization failed: {error}")
                return None
            lines = candidate.get("lines")
            if isinstance(lines, list) and lines:
                return [dict(line) for line in lines if isinstance(line, dict)]
            if _workflow_v2_output_state_requires_saved_sheet(workflow.state) or not raw_lines:
                raise ValueError("workflow-v2 saved sheet produced no output lines")
    except ValueError:
        raise
    except Exception as exc:  # noqa: BLE001
        if not raw_lines:
            raise ValueError(f"workflow-v2 output state lookup failed: {exc}") from exc
        logger.warning("Workflow-v2 output state lookup skipped", order_id=order_id, error=str(exc))
    return None


def _workflow_v2_output_state_requires_saved_sheet(state: object) -> bool:
    return str(state or "").strip() in {
        "sheet_saved",
        "bagging_ready",
        "output_review",
        "confirmed",
    }


def _workflow_v2_materialization_candidate(
    session,
    order: dict,
    workflow: OrderWorkflowState,
    draft: OrderSheetDraft,
) -> dict | None:
    meta = workflow.secondary_actions_json if isinstance(workflow.secondary_actions_json, dict) else {}
    workflow_meta = meta.get("workflow_v2") if isinstance(meta.get("workflow_v2"), dict) else {}
    workflow_meta = order_output_artifact_service.enrich_workflow_meta_with_artifacts(session, workflow_meta)
    bagging_result = workflow_meta.get("bagging_result") if isinstance(workflow_meta.get("bagging_result"), dict) else None
    if bagging_result is None:
        bagging_id = str(workflow_meta.get("bagging_result_id") or "").strip()
        if bagging_id:
            artifact = session.get(OrderBaggingResult, bagging_id)
            if artifact is not None and isinstance(artifact.payload_json, dict):
                bagging_result = artifact.payload_json
    if isinstance(bagging_result, dict):
        source_saved_sheet_id = str(bagging_result.get("source_saved_sheet_id") or "").strip()
        template_version_id = str(bagging_result.get("template_version_id") or "").strip()
        draft_template_version_id = str(draft.template_version_id or "").strip()
        if source_saved_sheet_id and source_saved_sheet_id != draft.id:
            raise ValueError("workflow-v2 bagging result source does not match saved sheet")
        if template_version_id and draft_template_version_id and template_version_id != draft_template_version_id:
            raise ValueError("workflow-v2 bagging result template does not match saved sheet")
        candidate = bagging_result.get("materialization_candidate")
        if isinstance(candidate, dict):
            return candidate
    return order_service._build_materialization_candidate_from_draft_record(  # noqa: SLF001
        str(order.get("id") or ""),
        draft_record=_workflow_v2_draft_record_for_outputs(draft),
        facility_id=order.get("facility") or order.get("facility_code"),
        existing_week_code=order.get("stored_week_value") or order.get("week_value") or order.get("week") or order.get("week_code"),
        received_at=order.get("received_at"),
    )


def _workflow_v2_draft_record_for_outputs(draft: OrderSheetDraft) -> dict:
    return {
        "id": draft.id,
        "order_id": draft.order_id,
        "base_evidence_run_id": draft.base_evidence_run_id,
        "base_template_resolution_id": draft.base_template_resolution_id,
        "base_menu_snapshot_id": draft.base_menu_snapshot_id,
        "draft_sheet_json": draft.draft_sheet_json if isinstance(draft.draft_sheet_json, dict) else {},
        "draft_state": str(draft.draft_state or "saved").strip() or "saved",
        "blockers_json": list(draft.blockers_json or []),
        "warnings_json": list(draft.warnings_json or []),
    }


def _normalize_rule_text(value: str | None) -> str:
    if not value:
        return ""
    return re.sub(r"\s+", "", str(value)).lower()


def _split_garnish_name(value: str | None) -> tuple[str | None, str | None]:
    if not value:
        return None, None
    text = str(value).strip()
    if not text:
        return None, None
    match = _GARNISH_SPLIT_RE.search(text)
    if not match:
        return text, None
    base = text[: match.start()].strip().strip("　")
    garnish = text[match.end() :].strip().strip("　")
    if not base:
        return text, None
    if not garnish:
        return base, None
    return base, garnish


def _apply_garnish_lines(lines: list[dict]) -> list[dict]:
    enriched: list[dict] = []
    for line in lines:
        name = (line.get("menu_name") or "").strip()
        base, garnish = _split_garnish_name(name)
        if not garnish:
            if base and base != name:
                updated = dict(line)
                updated["menu_name"] = base
                enriched.append(updated)
            else:
                enriched.append(line)
            continue
        updated = dict(line)
        updated["menu_name"] = base
        enriched.append(updated)
        garnish_line = dict(line)
        garnish_line["menu_name"] = garnish
        garnish_line["menu_category"] = "添え"
        garnish_line["parent_menu_name"] = base
        garnish_line["parent_menu_category"] = line.get("menu_category")
        for field in (
            "menu_qty_per_serving",
            "menu_unit_type",
            "actual_amount",
            "actual_unit_type",
            "menu_bag_max_qty",
            "menu_bag_max_unit",
            "_menu_qty_source_daypart",
            "_menu_qty_source_category",
            "_monthly_menu_item_override_applied",
            "_monthly_entry_override_applied",
        ):
            garnish_line.pop(field, None)
        enriched.append(garnish_line)
    return enriched


def _match_menu_pattern(menu_name: str, pattern: str, match_type: str | None) -> bool:
    if not pattern:
        return False
    if match_type == "regex":
        try:
            return re.search(pattern, menu_name) is not None
        except re.error:
            return False
    normalized_menu = _normalize_rule_text(menu_name)
    normalized_pattern = _normalize_rule_text(pattern)
    if match_type == "exact":
        return normalized_menu == normalized_pattern
    return normalized_pattern in normalized_menu


def _rule_applies(rule: dict, line: dict, facility_id: str | None) -> bool:
    if rule.get("rule_type") == "facility":
        if not facility_id or not rule.get("facility_id"):
            return False
        if rule.get("facility_id") != facility_id:
            return False
    if rule.get("rule_type") in {"menu", "facility"}:
        menu_name = line.get("menu_name") or ""
        if not _match_menu_pattern(
            menu_name,
            rule.get("menu_pattern") or "",
            rule.get("match_type"),
        ):
            return False
    if rule.get("daypart"):
        rule_daypart = _normalize_output_daypart(rule.get("daypart"))
        line_daypart = _normalize_output_daypart(line.get("daypart"))
        if rule_daypart != line_daypart:
            return False
    if rule.get("category") and rule.get("category") != line.get("menu_category"):
        return False
    if rule.get("diet_type") and rule.get("diet_type") != line.get("diet_type"):
        return False
    return True


def _apply_menu_rules(lines: list[dict], facility_id: str | None) -> list[dict]:
    rules = _list_cached_active_menu_rules()
    if not rules:
        return lines
    type_weight = {"global": 100, "menu": 200, "facility": 300}
    enriched: list[dict] = []
    for line in lines:
        matches = [
            rule
            for rule in rules
            if _rule_applies(rule, line, facility_id)
        ]
        if not matches:
            enriched.append(line)
            continue
        selected = max(
            matches,
            key=lambda rule: type_weight.get(rule.get("rule_type"), 0) + int(rule.get("priority") or 0),
        )
        updated = dict(line)
        if selected.get("unit_type"):
            updated["menu_unit_type"] = selected.get("unit_type")
        if selected.get("qty_per_serving") is not None:
            updated["menu_qty_per_serving"] = selected.get("qty_per_serving")
        enriched.append(updated)
    return enriched


def _apply_builtin_menu_defaults(lines: list[dict]) -> list[dict]:
    defaults = menu_rule_service.DEFAULT_GLOBAL_RULES
    if not defaults:
        return lines
    enriched: list[dict] = []
    for line in lines:
        default_daypart, default_category, _default_temp, _default_qty, _default_unit = _label_menu_defaults(
            line.get("menu_name")
        )
        updated = dict(line)
        if default_daypart and not updated.get("daypart"):
            updated["daypart"] = default_daypart
        if default_category and not updated.get("_monthly_entry_override_applied"):
            updated["menu_category"] = _normalize_delivery_category_label(default_category)
        if updated.get("menu_qty_per_serving") is not None:
            enriched.append(updated)
            continue
        line_daypart = _normalize_output_daypart(updated.get("daypart"))
        line_category = _normalize_category_key(updated.get("menu_category"))
        matched = next(
            (
                rule
                for rule in defaults
                if _normalize_output_daypart(rule.get("daypart")) == line_daypart
                and _normalize_category_key(rule.get("category")) == line_category
            ),
            None,
        )
        if not matched:
            enriched.append(updated)
            continue
        updated["menu_unit_type"] = matched.get("unit_type")
        updated["menu_qty_per_serving"] = matched.get("qty_per_serving")
        enriched.append(updated)
    return enriched


def _build_bags(order: dict, packaging_policy: dict, quantity_rules: dict) -> list[dict]:
    split_key = packaging_policy.get(
        "split_key",
        ["facility", "date", "daypart", "menu_name", "diet_type", "area_id", "bag_type"],
    )
    zero_as_empty = quantity_rules.get("zero_as_empty", True)

    grouped: dict[tuple, dict] = {}
    for line in order.get("lines", []):
        line_date = _ensure_date(line.get("date"))
        qty = _safe_qty(line, zero_as_empty)
        if qty is None:
            continue
        is_condiment = line.get("bag_type") == "condiment"
        default_bag_type = line.get("bag_type") or "standard"
        menu_category = "付属品" if is_condiment else line.get("menu_category")
        key_parts: list[Any] = []
        for part in split_key:
            if part == "facility":
                key_parts.append(order.get("facility"))
                continue
            if part == "date":
                key_parts.append(line_date)
                continue
            if part == "menu_name":
                key_parts.append("condiment" if is_condiment else line.get("menu_name"))
                continue
            if part == "bag_type":
                key_parts.append("condiment" if is_condiment else default_bag_type)
                continue
            if is_condiment and part in {"diet_type", "area_id"}:
                # 付属品は area/diet を跨いで1袋にまとめる。
                key_parts.append("__condiment__")
                continue
            key_parts.append(line.get(part))
        key = tuple(key_parts)
        if key not in grouped:
            grouped[key] = {
                "order_id": order["id"],
                "facility": order.get("facility"),
                "date": line_date,
                "daypart": line.get("daypart"),
                "menu_name": line.get("menu_name"),
                "menu_category": menu_category,
                "diet_type": None if is_condiment else line.get("diet_type"),
                "area_id": None if is_condiment else line.get("area_id"),
                "bag_type": "condiment" if is_condiment else default_bag_type,
                "parent_menu_name": line.get("parent_menu_name"),
                "parent_menu_category": line.get("parent_menu_category"),
                "menu_unit_type": line.get("menu_unit_type"),
                "menu_qty_per_serving": line.get("menu_qty_per_serving"),
                "menu_bag_max_qty": line.get("menu_bag_max_qty"),
                "menu_bag_max_unit": line.get("menu_bag_max_unit"),
                "menu_temp_type": line.get("menu_temp_type"),
                "quantity": 0.0,
                "_condiment_names": set(),
                "_source_refs": [],
            }
        if is_condiment:
            name_value = (line.get("menu_name") or "").strip()
            if name_value:
                grouped[key]["_condiment_names"].add(name_value)
            if not grouped[key].get("parent_menu_name") and line.get("parent_menu_name"):
                grouped[key]["parent_menu_name"] = line.get("parent_menu_name")
            if not grouped[key].get("parent_menu_category") and line.get("parent_menu_category"):
                grouped[key]["parent_menu_category"] = line.get("parent_menu_category")
        grouped[key]["quantity"] += float(qty)
        source_ref = {
            "order_id": order.get("id"),
            "source_row_index": line.get("source_row_index"),
            "source_col_index": line.get("source_col_index"),
            "source_field": line.get("source_field"),
            "value_source": line.get("value_source"),
            "was_user_edited": line.get("was_user_edited"),
            "ocr_confidence": line.get("ocr_confidence"),
            "ocr_confidence_tier": line.get("ocr_confidence_tier"),
            "cell_image_ref": line.get("cell_image_ref"),
            "quantity": float(qty),
        }
        grouped[key]["_source_refs"].append(source_ref)
    result = list(grouped.values())
    for bag in result:
        if bag.get("bag_type") == "condiment":
            names = sorted(bag.pop("_condiment_names", set()))
            if names:
                bag["menu_name"] = " / ".join(names)
            if not bag.get("menu_category"):
                bag["menu_category"] = "付属品"
    return result


def _max_servings_for_bag(bag: dict) -> int | None:
    bag_max = bag.get("menu_bag_max_qty")
    per_serving = bag.get("menu_qty_per_serving")
    if bag_max is None or per_serving is None:
        return None
    try:
        bag_max_value = float(bag_max)
        per_value = float(per_serving)
    except Exception:
        return None
    if bag_max_value <= 0 or per_value <= 0:
        return None
    unit = _normalize_unit_type(bag.get("menu_unit_type"))
    bag_unit = _normalize_unit_type(bag.get("menu_bag_max_unit")) or unit
    if bag_unit and unit and bag_unit != unit:
        return None
    max_servings = int(math.floor(bag_max_value / per_value))
    if max_servings <= 0:
        return None
    return max_servings


def _split_bags_by_max(bags: list[dict]) -> list[dict]:
    split: list[dict] = []
    for bag in bags:
        max_servings = _max_servings_for_bag(bag)
        if not max_servings:
            split.append(bag)
            continue
        remaining = bag.get("quantity") or 0
        try:
            remaining = float(remaining)
        except Exception:
            split.append(bag)
            continue
        if remaining <= max_servings:
            split.append(bag)
            continue
        while remaining > 0:
            chunk = max_servings if remaining > max_servings else remaining
            next_bag = dict(bag)
            next_bag["quantity"] = chunk
            split.append(next_bag)
            remaining -= chunk
    return split


def _serialize_bag_payload_rows(rows: list[dict]) -> list[dict]:
    payload = [
        {
            "date": bag.get("date").isoformat() if bag.get("date") else None,
            "daypart": bag.get("daypart"),
            "menu_name": bag.get("menu_name"),
            "menu_category": bag.get("menu_category"),
            "diet_type": bag.get("diet_type"),
            "area_id": bag.get("area_id"),
            "bag_type": bag.get("bag_type"),
            "quantity": bag.get("quantity"),
            "source_refs": bag.get("_source_refs") or bag.get("source_refs") or [],
        }
        for bag in rows
    ]
    payload.sort(
        key=lambda row: (
            row.get("date") or "",
            row.get("daypart") or "",
            row.get("menu_name") or "",
            row.get("diet_type") or "",
            row.get("area_id") or "",
            row.get("bag_type") or "",
        )
    )
    return payload


def build_bag_rows_for_outputs(
    order: dict,
    *,
    order_lines: list[dict] | None = None,
    facility_config: dict | None = None,
) -> list[dict]:
    facility_id = order.get("facility")
    resolved_facility_config = facility_config
    if not resolved_facility_config and facility_id:
        resolved_facility_config = config_service.get_facility_config(facility_id)
    if not resolved_facility_config:
        logger.warning("Facility config missing", facility_id=facility_id)
        resolved_facility_config = {}

    packaging_policy = resolved_facility_config.get("packaging_policy", {})
    quantity_rules = config_service.load_ingest_policy().get("quantity_rules", {})

    resolved_lines = order_lines if isinstance(order_lines, list) else build_order_lines_for_outputs(order)
    order_for_outputs = {**order, "lines": resolved_lines}

    bags = _split_bags_by_max(_build_bags(order_for_outputs, packaging_policy, quantity_rules))
    week_value = (
        str(order.get("stored_week_value") or "").strip()
        or str(order.get("week_value") or "").strip()
        or str(order.get("persisted_week_value") or "").strip()
        or str(order.get("week") or "").strip()
        or str(order.get("week_code") or "").strip()
    )
    menu_items = _collect_cached_menu_items_for_week(week_value, facility_id)
    bags = _apply_menu_overrides(bags, menu_items)
    bags = _clear_stale_menu_qty_from_monthly_entry(bags)
    bags = _apply_menu_master_defaults(bags, facility_id)
    bags = _apply_builtin_menu_defaults(bags)
    bags = _apply_label_meal_slot_categories(bags)
    bag_types = _resolve_bag_types(resolved_facility_config)
    bags = _assign_bag_type_for_bags(bags, bag_types)
    return _apply_daily_label_facility_rules_to_bags(bags, resolved_facility_config, facility_id)


def build_bag_payload_for_outputs(
    order: dict,
    *,
    order_lines: list[dict] | None = None,
    facility_config: dict | None = None,
) -> list[dict]:
    return _serialize_bag_payload_rows(
        build_bag_rows_for_outputs(
            order,
            order_lines=order_lines,
            facility_config=facility_config,
        )
    )


def _label_payload_legacy(bag: dict, label_profile: dict, facility_name: str | None) -> dict:
    fixed_text = label_profile.get("fixed_text", {})
    expiry_rule = label_profile.get("expiry_rule", "meal_date")
    expiry_date = bag.get("date")
    if expiry_rule == "meal_date" and expiry_date:
        expiry_value = _resolve_label_expiry_date(expiry_date, label_profile)
    else:
        expiry_value = _resolve_label_expiry_date(expiry_date, label_profile)
    menu_category = bag.get("menu_category") or bag.get("diet_type")
    return {
        "facility_name": facility_name,
        "expiry_date": expiry_value,
        "storage_mode": label_profile.get("storage_mode"),
        "meal_slot": bag.get("daypart"),
        "menu_category": menu_category,
        "product_name": bag.get("menu_name"),
        "quantity": bag.get("quantity"),
        "details": _build_label_details(bag),
        "maker_info": fixed_text.get("maker_name"),
        "notice": fixed_text.get("notice"),
    }

def _label_payload_jp(bag: dict, label_profile: dict | None = None) -> dict:
    product_name = bag.get("menu_name") or ""
    expiry_value = _resolve_label_expiry_date(bag.get("date"), label_profile)
    default_daypart, _default_category, default_temp, default_qty, default_unit = _label_menu_defaults(product_name)
    raw_per_qty = bag.get("menu_qty_per_serving")
    raw_unit = bag.get("menu_unit_type")
    per_qty, unit = _extract_qty_and_unit(
        raw_per_qty if raw_per_qty is not None else default_qty,
        raw_unit or default_unit,
    )
    servings = bag.get("quantity")
    total_qty = None
    if per_qty is not None and servings is not None:
        try:
            total_qty = float(per_qty) * float(servings)
        except Exception:
            total_qty = None
    diet_type = str(bag.get("diet_type") or "").strip()
    display_product_name = _daily_label_display_product(str(product_name), diet_type)
    menu_value = _label_category_for_bag(bag, str(product_name), diet_type)
    # 重複表示を避けるため、メニュー列は分類（主菜/副菜など）を優先して扱う。
    if menu_value and product_name and str(menu_value).strip() == str(product_name).strip():
        menu_value = ""
    total_amount = _daily_label_amount_cell(total_qty, unit)
    if not total_amount and servings not in (None, ""):
        total_amount = _format_number(servings)
    area_suffix = _label_area_suffix(bag.get("area_id"), bag.get("facility"))
    time_value = str(bag.get("daypart") or default_daypart or "").strip()
    if area_suffix:
        time_value = f"{time_value}　{area_suffix}".strip()
    source_refs = bag.get("_source_refs") or bag.get("source_refs") or []
    source_row_indexes: list[int] = []
    if isinstance(source_refs, list):
        for ref in source_refs:
            if not isinstance(ref, dict):
                continue
            try:
                source_row_indexes.append(int(ref.get("source_row_index")))
            except Exception:
                continue
    return {
        "呼び出し番号": "",
        "発行枚数": 1,
        "賞味期限": _format_jp_date(expiry_value),
        "時間": time_value,
        "メニュー": menu_value,
        "温・冷": _normalize_temp_label(bag.get("menu_temp_type") or default_temp),
        "商品名１": display_product_name,
        "商品名２": "",
        "内容量": total_amount,
        "内容詳細": _daily_label_amount_cell(per_qty, unit),
        "": _format_servings(servings),
        "_sort_date": expiry_value.isoformat() if hasattr(expiry_value, "isoformat") else str(expiry_value or ""),
        "_sort_source_row": min(source_row_indexes) if source_row_indexes else None,
        "_sort_diet": _daily_label_display_diet_key(bag),
        "_sort_area": _normalize_area_id(bag.get("area_id")) or "",
    }


def _merge_label_rows(rows: list[dict], fields: list[str]) -> list[dict]:
    if not rows:
        return []
    rows = _split_daily_label_serving_rows(rows)
    group_fields = [field for field in fields if field not in {"呼び出し番号", "発行枚数"}]
    grouped: dict[tuple, dict] = {}
    counts: dict[tuple, int] = {}
    for row in rows:
        key = tuple(row.get(field, "") for field in group_fields)
        if key not in grouped:
            grouped[key] = dict(row)
            counts[key] = 0
        counts[key] += 1
    merged = []
    for key, row in grouped.items():
        row["発行枚数"] = counts.get(key, 1)
        merged.append(row)
    daypart_order = {"朝": 0, "昼": 1, "夕": 2}
    category_order = {
        "主菜": 0,
        "主菜（軟菜）": 1,
        "主菜（ミキサー）": 2,
        "副菜①": 3,
        "副菜①（軟菜）": 4,
        "副菜①（ミキサー）": 5,
        "副菜②": 6,
        "副菜②（軟菜）": 7,
        "副菜②（ミキサー）": 8,
        "副菜": 9,
        "副菜（軟菜）": 10,
        "副菜（ミキサー）": 11,
        "ソース": 12,
    }
    diet_order = {
        "": 0,
        "regular": 0,
        "regular_bag": 0,
        "standard": 0,
        "soft": 1,
        "soft_mixer": 1,
        "mixer": 2,
    }
    area_order = {"": 0, "X": 0, "2F": 1, "3F": 2}

    def split_time_text(value: Any) -> tuple[str, str]:
        text = str(value or "").strip()
        if not text:
            return "", ""
        daypart = text[:1]
        area = text[1:].strip(" 　")
        return daypart, area

    def normalize_category_for_sort(value: Any) -> str:
        text = str(value or "").strip()
        if text in category_order:
            return text
        if text.startswith("主菜"):
            if "ミキサー" in text:
                return "主菜（ミキサー）"
            if "軟菜" in text:
                return "主菜（軟菜）"
            return "主菜"
        if text.startswith("副菜①"):
            if "ミキサー" in text:
                return "副菜①（ミキサー）"
            if "軟菜" in text:
                return "副菜①（軟菜）"
            return "副菜①"
        if text.startswith("副菜②"):
            if "ミキサー" in text:
                return "副菜②（ミキサー）"
            if "軟菜" in text:
                return "副菜②（軟菜）"
            return "副菜②"
        if text.startswith("副菜"):
            if "ミキサー" in text:
                return "副菜（ミキサー）"
            if "軟菜" in text:
                return "副菜（軟菜）"
            return "副菜"
        return text

    def variant_for_sort(value: Any, explicit_diet: Any = None) -> str:
        diet = str(explicit_diet or "").strip()
        if diet:
            return diet
        text = str(value or "").strip()
        if "ミキサー" in text:
            return "mixer"
        if "軟菜" in text:
            return "soft"
        return "regular"

    def area_for_sort(row: dict, time_text: str) -> str:
        explicit = str(row.get("_sort_area") or "").strip()
        if explicit:
            return explicit
        _daypart, area = split_time_text(time_text)
        normalized = _normalize_area_id(area)
        return normalized or area

    def sort_key(row: dict) -> tuple:
        time_text = str(row.get("時間") or "")
        daypart, area = split_time_text(time_text)
        category = normalize_category_for_sort(row.get("メニュー"))
        variant = variant_for_sort(row.get("メニュー"), row.get("_sort_diet"))
        area_key = area_for_sort(row, time_text)
        try:
            source_row_sort = int(row.get("_sort_source_row"))
        except Exception:
            source_row_sort = 999999
        try:
            serving_sort = -float(row.get("") or 0)
        except Exception:
            serving_sort = 0
        return (
            row.get("_sort_date") or row.get("賞味期限", ""),
            daypart_order.get(daypart, 99),
            source_row_sort,
            category_order.get(category, 99),
            row.get("商品名１", ""),
            diet_order.get(variant, 50),
            area_order.get(area_key, 50),
            time_text,
            serving_sort,
        )

    merged.sort(
        key=sort_key
    )
    for row in merged:
        for key in list(row.keys()):
            if str(key).startswith("_sort_"):
                row.pop(key, None)
    return merged


def _daily_label_max_servings(row: dict) -> int | None:
    product = str(row.get("商品名１") or "")
    category = str(row.get("メニュー") or "")
    if "ミキサー" in category:
        return None
    if product in {"ごぼうと竹輪の煮物", "竹輪の煮物"}:
        return 20
    if product == "豚肉と白菜のすき煮":
        return 15
    if product == "さつま芋の天ぷら":
        return 10
    if product == "煮込みハンバーグ":
        return 10
    return None


def _split_daily_label_serving_rows(rows: list[dict]) -> list[dict]:
    split: list[dict] = []
    for row in rows:
        max_servings = _daily_label_max_servings(row)
        servings = row.get("")
        try:
            servings_value = float(servings)
        except Exception:
            split.append(row)
            continue
        if not max_servings or servings_value <= max_servings:
            split.append(row)
            continue
        detail_value = row.get("内容詳細")
        try:
            per_qty = float(detail_value)
        except Exception:
            match = re.search(r"[-+]?[0-9]*\.?[0-9]+", str(detail_value or ""))
            per_qty = float(match.group(0)) if match else None
        remaining = servings_value
        while remaining > 0:
            chunk = float(max_servings) if remaining > max_servings else remaining
            next_row = dict(row)
            next_row[""] = int(chunk) if chunk.is_integer() else chunk
            if per_qty is not None:
                total = per_qty * chunk
                next_row["内容量"] = f"{_format_number(total)}個" if "個" in str(detail_value or "") else total
            split.append(next_row)
            remaining -= chunk
    return split


def _write_label_csv(path: Path, labels: list[dict], label_fields: list[str]) -> None:
    fieldnames = label_fields or (list(labels[0].keys()) if labels else [])
    with path.open("w", newline="", encoding="cp932", errors="replace") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for label in labels:
            writer.writerow({k: label.get(k, "") for k in fieldnames})

def _normalize_cell_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value)
    text = re.sub(r"[\\s　]+", "", text)
    return text


def _find_delivery_header_row(ws, columns: list[dict]) -> int | None:
    targets = []
    for col in columns:
        header = col.get("header") or col.get("name")
        if header:
            targets.append(_normalize_cell_text(header))
    best_row = None
    best_hits = 0
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        row_text = [_normalize_cell_text(cell.value) for cell in row if cell.value is not None]
        if not row_text:
            continue
        hits = sum(1 for target in targets if any(target in cell for cell in row_text))
        if hits > best_hits:
            best_hits = hits
            best_row = row[0].row
    return best_row


def _build_delivery_column_map(ws, header_row: int, columns: list[dict]) -> dict[str, int]:
    column_map: dict[str, int] = {}
    header_cells = list(ws[header_row])
    for col in columns:
        name = col.get("name")
        if not name:
            continue
        column_index = col.get("column_index")
        if isinstance(column_index, int) and column_index > 0:
            column_map[name] = column_index
            continue
        header = col.get("header") or name
        normalized = _normalize_cell_text(header)
        if not normalized:
            continue
        for cell in header_cells:
            cell_text = _normalize_cell_text(cell.value)
            if normalized and normalized in cell_text:
                column_map[name] = cell.col_idx
                break
    return column_map


def _apply_delivery_configured_headers(
    ws,
    header_row: int,
    columns: list[dict],
    column_map: dict[str, int],
    *,
    data_start_row: int | None = None,
) -> None:
    if header_row < 1:
        return
    configured_indexes = {
        int(col.get("column_index"))
        for col in columns
        if isinstance(col.get("column_index"), int) and int(col.get("column_index")) > 0
    }
    if not configured_indexes:
        return
    for col in columns:
        name = col.get("name")
        if not name:
            continue
        col_idx = column_map.get(name)
        if not col_idx:
            continue
        header = col.get("header") or name
        cell = _resolve_merged_cell(ws, header_row, col_idx)
        merged = next(
            (
                merged
                for merged in ws.merged_cells.ranges
                if merged.min_row <= header_row <= merged.max_row
                and merged.min_col <= col_idx <= merged.max_col
            ),
            None,
        )
        if merged:
            covered = [
                configured_col
                for configured_col in configured_indexes
                if merged.min_col <= configured_col <= merged.max_col
            ]
            if len(covered) > 1:
                continue
        if not isinstance(cell, MergedCell):
            cell.value = header
    start_row = data_start_row or _delivery_start_row(ws, header_row)
    intended_headers_by_col = {
        int(col.get("column_index")): _normalize_cell_text(col.get("header") or col.get("name"))
        for col in columns
        if isinstance(col.get("column_index"), int) and int(col.get("column_index")) > 0
    }
    for merged in ws.merged_cells.ranges:
        if merged.min_row >= header_row or merged.max_row >= start_row:
            continue
        covered = [
            col_idx
            for col_idx in configured_indexes
            if merged.min_col <= col_idx <= merged.max_col
        ]
        if not covered:
            continue
        cell = ws.cell(row=merged.min_row, column=merged.min_col)
        cell_text = _normalize_cell_text(cell.value)
        if not cell_text:
            continue
        intended = {intended_headers_by_col.get(col_idx, "") for col_idx in covered}
        if cell_text not in intended:
            cell.value = ""
    for row_idx in range(header_row + 1, start_row):
        for col_idx in configured_indexes:
            cell = _resolve_merged_cell(ws, row_idx, col_idx)
            if not isinstance(cell, MergedCell):
                cell.value = ""


def _delivery_start_row(ws, header_row: int) -> int:
    max_row = header_row
    for merged in ws.merged_cells.ranges:
        if merged.min_row <= header_row <= merged.max_row:
            max_row = max(max_row, merged.max_row)
    return max_row + 1


def _normalize_delivery_daypart(value: Any) -> str:
    if not value:
        return ""
    text = str(value)
    if "朝" in text:
        return "朝"
    if "昼" in text:
        return "昼"
    if "夕" in text or "夜" in text:
        return "夕"
    return text


def _normalize_ocr_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value)
    text = re.sub(r"<br\s*/?>", " ", text, flags=re.IGNORECASE)
    text = text.replace("\\(", "(").replace("\\)", ")")
    text = re.sub(r"\s+", " ", text).strip()
    return text


def _normalize_menu_key(value: Any) -> str:
    return _normalize_ocr_text(value)


def _normalize_ocr_daypart(value: Any) -> str:
    text = _normalize_ocr_text(value)
    if not text:
        return ""
    if "朝" in text or "明" in text:
        return "朝"
    if "昼" in text or "星" in text or "a" in text or "中" in text:
        return "昼"
    if "夕" in text or "タ" in text or "夜" in text:
        return "夕"
    return text


def _normalize_ocr_category(value: Any) -> str:
    text = _normalize_ocr_text(value)
    if not text:
        return ""
    if "主" in text:
        if "A" in text or "Ａ" in text:
            return "主Ａ"
        if "B" in text or "Ｂ" in text:
            return "主Ｂ"
        return "主"
    if "副" in text:
        if "2" in text or "②" in text:
            return "副②"
        return "副①"
    if text.isdigit():
        if "2" in text:
            return "副②"
        if "1" in text:
            return "副①"
    return ""


def _parse_ocr_date(value: str, year_hint: int | None, fallback: dt_date | None) -> dt_date | None:
    text = _normalize_ocr_text(value)
    if not text:
        return fallback
    match = re.search(r"(\d{4})[-/](\d{1,2})[-/](\d{1,2})", text)
    if match:
        try:
            return dt_date(int(match.group(1)), int(match.group(2)), int(match.group(3)))
        except ValueError:
            return fallback
    match = re.search(r"(\d{1,2})[月/](\d{1,2})", text)
    if match:
        month = int(match.group(1))
        day = int(match.group(2))
        year = year_hint or (fallback.year if fallback else None)
        if not year:
            return fallback
        try:
            return dt_date(year, month, day)
        except ValueError:
            return fallback
    if text.isdigit():
        try:
            serial = int(text)
        except ValueError:
            return fallback
        if serial > 10000:
            base = dt_date(1899, 12, 30)
            return base + timedelta(days=serial)
    return fallback


def _parse_ocr_quantity(value: Any) -> float | None:
    text = _normalize_ocr_text(value)
    if not text:
        return None
    match = re.search(r"[-+]?[0-9]*\.?[0-9]+", text)
    if not match:
        return None
    try:
        return float(match.group())
    except ValueError:
        return None


def _parse_ocr_table_rows(table_raw: str, year_hint: int | None) -> list[dict]:
    rows: list[dict] = []
    current_date: dt_date | None = None
    current_daypart = ""
    in_table = False
    header_seen = False
    for line in table_raw.splitlines():
        raw_line = line.lstrip()
        if not in_table and raw_line.startswith("|") and "日付" in raw_line and "献立" in raw_line:
            in_table = True
            header_seen = True
            continue
        if not in_table:
            continue
        if raw_line.startswith("|-"):
            continue
        if not raw_line.startswith("|"):
            if header_seen:
                break
            continue
        cells = [cell.strip() for cell in raw_line.split("|")[1:-1]]
        if len(cells) < 4:
            continue
        date_cell = _normalize_ocr_text(cells[0])
        if date_cell:
            parsed_date = _parse_ocr_date(date_cell, year_hint, current_date)
            if parsed_date:
                current_date = parsed_date
        daypart_cell = _normalize_ocr_text(cells[1])
        if daypart_cell:
            current_daypart = _normalize_ocr_daypart(daypart_cell)
        category_cell = _normalize_ocr_category(cells[2])
        menu_name = _normalize_ocr_text(cells[3])
        if not menu_name:
            continue
        qty_values = [None] * 6
        note_value = _normalize_ocr_text(cells[10]) if len(cells) >= 11 else ""
        if len(cells) >= 10:
            qty_cells = cells[4:10]
            if len(cells) >= 11 and not cells[4] and sum(1 for c in qty_cells if _normalize_ocr_text(c)) == 5:
                qty_cells = cells[5:10] + [None]
            qty_values = [
                _parse_ocr_quantity(qty_cells[0]),
                _parse_ocr_quantity(qty_cells[1]),
                _parse_ocr_quantity(qty_cells[2]),
                _parse_ocr_quantity(qty_cells[3]),
                _parse_ocr_quantity(qty_cells[4]),
                _parse_ocr_quantity(qty_cells[5]),
            ]
        rows.append(
            {
                "date": current_date,
                "daypart": current_daypart,
                "category": category_cell,
                "menu_name": menu_name,
                "qty_regular_2f": qty_values[0],
                "qty_regular_3f": qty_values[1],
                "qty_soft_2f": qty_values[2],
                "qty_soft_3f": qty_values[3],
                "qty_mixer_2f": qty_values[4],
                "qty_mixer_3f": qty_values[5],
                "note": note_value,
            }
        )
    return rows


def _resolve_ocr_year_hint(order: dict) -> int | None:
    year_hint = None
    for line in order.get("lines", []):
        date_val = _ensure_date(line.get("date"))
        if date_val:
            year_hint = date_val.year
            break
    if not year_hint:
        week_value = order.get("week") or order.get("week_code")
        week_text = str(week_value) if week_value is not None else ""
        match = re.search(r"(\d{4})", week_text)
        if match:
            year_hint = int(match.group(1))
    if not year_hint:
        received_at = order.get("received_at")
        if isinstance(received_at, dt_date):
            year_hint = received_at.year
        elif isinstance(received_at, str):
            match = re.search(r"(\d{4})", received_at)
            if match:
                year_hint = int(match.group(1))
    return year_hint


def _load_order_ocr_payload(order_id: str) -> dict[str, Any] | None:
    parsed, _ = order_service.get_ocr_output(order_id)
    if isinstance(parsed, dict):
        return parsed
    raw_text, _ = order_service.get_ocr_raw_text(order_id)
    if not isinstance(raw_text, str) or not raw_text.strip():
        return None
    raw_text = raw_text.strip()
    if raw_text.startswith("{"):
        try:
            parsed_raw = json.loads(raw_text)
        except Exception:  # noqa: BLE001
            parsed_raw = None
        if isinstance(parsed_raw, dict):
            return parsed_raw
    return {"table_raw": raw_text}


def _entry_quantity_key(diet: str | None, area: str | None) -> str | None:
    diet_key = _normalize_diet_key(diet)
    area_key = str(area or "").strip().lower()
    if not diet_key or not area_key:
        return None
    return f"{diet_key}_{area_key}"


def _legacy_entry_quantity_field(diet: str | None, area: str | None) -> str | None:
    diet_key = _normalize_diet_key(diet)
    area_key = str(area or "").strip().upper()
    if diet_key not in {"regular", "soft", "mixer"}:
        return None
    if area_key not in {"2F", "3F"}:
        return None
    return f"qty_{diet_key}_{area_key.lower()}"


def _sum_quantity_map(quantity_map: dict[str, Any], keys: tuple[str, ...]) -> float | None:
    total = 0.0
    matched = False
    for key in keys:
        qty = _parse_ocr_quantity(quantity_map.get(key))
        if qty is None:
            continue
        total += qty
        matched = True
    if not matched:
        return None
    return total if total > 0 else None


def _generic_quantity_total(quantity_map: dict[str, Any], diet_key: str | None) -> float | None:
    normalized_diet = _normalize_diet_key(diet_key)
    if normalized_diet == "regular":
        override_qty = (
            _parse_ocr_quantity(quantity_map.get("change_2_x"))
            if "change_2_x" in quantity_map
            else None
        )
        if override_qty is None and "change_1_x" in quantity_map:
            override_qty = _parse_ocr_quantity(quantity_map.get("change_1_x"))
        if override_qty is not None:
            return override_qty
        return _sum_quantity_map(quantity_map, ("regular_x", "regular_bag_x", "staff_x", "daycare_x"))
    if normalized_diet == "regular_bag":
        return _sum_quantity_map(quantity_map, ("regular_bag_x",))
    if normalized_diet == "no_fried":
        return _sum_quantity_map(quantity_map, ("no_fried_x",))
    if normalized_diet == "soft":
        return _sum_quantity_map(quantity_map, ("soft_x",))
    if normalized_diet == "mixer":
        return _sum_quantity_map(quantity_map, ("mixer_x",))
    if normalized_diet in {"禁食", "forbidden"}:
        return _sum_quantity_map(quantity_map, ("forbidden_x", "no_meat_x", "no_fish_x", "forbidden_other_x"))
    return None


def _resolve_payload_template(payload: dict[str, Any], template: dict[str, Any] | None) -> dict[str, Any] | None:
    resolved = template if isinstance(template, dict) else None
    if not isinstance(payload, dict):
        return resolved
    template_id = payload.get("template_id")
    if not isinstance(template_id, str):
        classification = payload.get("classification")
        if isinstance(classification, dict):
            template_id = classification.get("matched_template_id")
    if not isinstance(template_id, str):
        return resolved
    template_id = template_id.strip()
    if not template_id:
        return resolved
    if isinstance(resolved, dict) and resolved.get("template_id") == template_id:
        return resolved
    registry = config_service.load_fax_template_registry()
    matched = registry.get(template_id)
    if isinstance(matched, dict) and matched:
        return matched
    return resolved


def _extract_ocr_entries_from_structured_payload(
    payload: dict[str, Any],
    template: dict[str, Any] | None,
    year_hint: int | None,
) -> list[dict[str, Any]]:
    if not isinstance(payload, dict):
        return []
    template = _resolve_payload_template(payload, template)
    if not isinstance(template, dict):
        return []
    rows = order_service._extract_sheet_rows_from_payload(payload, template)
    if not rows:
        return []
    cell_issues = order_service._extract_payload_cell_issues(payload, template)
    issues_by_row: dict[int, list[dict[str, Any]]] = {}
    for issue in cell_issues:
        if not isinstance(issue, dict):
            continue
        row_index = issue.get("source_row_index")
        if not isinstance(row_index, int) or row_index < 0:
            continue
        issues_by_row.setdefault(row_index, []).append(dict(issue))
    fields, field_index = order_service._build_sheet_fields_and_indexes(template)
    if not fields:
        return []
    date_idx = field_index.get("date_mmdd")
    if date_idx is None:
        date_idx = field_index.get("date")
    daypart_idx = field_index.get("daypart")
    menu_idx = field_index.get("menu")
    if menu_idx is None:
        menu_idx = field_index.get("menu_name")
    note_idx = field_index.get("remarks")
    if note_idx is None:
        note_idx = field_index.get("note")

    quantity_fields: list[tuple[int, str, str]] = []
    for idx, field in enumerate(fields):
        diet, area = order_service._quantity_meta_from_field(field)
        if not diet or not area:
            continue
        quantity_fields.append((idx, diet, area))

    current_date: dt_date | None = None
    current_daypart = ""
    entries: list[dict[str, Any]] = []
    for idx, row in enumerate(rows):
        if not isinstance(row, list):
            continue
        if date_idx is not None and date_idx < len(row):
            parsed_date = _parse_ocr_date(str(row[date_idx] or ""), year_hint, current_date)
            if parsed_date:
                current_date = parsed_date
        if daypart_idx is not None and daypart_idx < len(row):
            daypart_value = _normalize_ocr_daypart(row[daypart_idx])
            if daypart_value:
                current_daypart = daypart_value
        menu_name = ""
        if menu_idx is not None and menu_idx < len(row):
            menu_name = _normalize_ocr_text(row[menu_idx])
        if not menu_name:
            continue
        note = ""
        if note_idx is not None and note_idx < len(row):
            note = _normalize_ocr_text(row[note_idx])
        quantity_map: dict[str, float] = {}
        payload_entry: dict[str, Any] = {
            "date": current_date,
            "daypart": current_daypart,
            "category": "",
            "menu_name": menu_name,
            "note": note,
            "index": idx,
            "quantity_map": quantity_map,
            "source": "structured_rows",
        }
        if idx in issues_by_row:
            payload_entry["ocr_issues"] = issues_by_row[idx]
            payload_entry["needs_review"] = True
        for col_idx, diet, area in quantity_fields:
            if col_idx >= len(row):
                continue
            qty = _parse_ocr_quantity(row[col_idx])
            if qty is None:
                continue
            key = _entry_quantity_key(diet, area)
            if key:
                quantity_map[key] = qty
            legacy_field = _legacy_entry_quantity_field(diet, area)
            if legacy_field:
                payload_entry[legacy_field] = qty
        entries.append(payload_entry)
    return entries


def _lookup_ocr_entry_quantity(entry: dict[str, Any], diet_key: str | None, area_key: str | None) -> float | None:
    if not isinstance(entry, dict):
        return None
    quantity_map = entry.get("quantity_map")
    normalized_diet = _normalize_diet_key(diet_key)
    normalized_area = str(area_key or "").strip().upper()
    if isinstance(quantity_map, dict):
        key = _entry_quantity_key(normalized_diet, area_key)
        if key and key in quantity_map:
            exact_qty = _parse_ocr_quantity(quantity_map.get(key))
            if normalized_area not in {"", "X"} or normalized_diet not in {"regular", "禁食", "forbidden"}:
                return exact_qty
        if normalized_area in {"", "X"}:
            generic_total = _generic_quantity_total(quantity_map, normalized_diet)
            if generic_total is not None:
                return generic_total
    if normalized_diet == "regular":
        if normalized_area == "2F":
            return _parse_ocr_quantity(entry.get("qty_regular_2f"))
        if normalized_area == "3F":
            return _parse_ocr_quantity(entry.get("qty_regular_3f"))
        qty_2f = _parse_ocr_quantity(entry.get("qty_regular_2f")) or 0.0
        qty_3f = _parse_ocr_quantity(entry.get("qty_regular_3f")) or 0.0
        total = qty_2f + qty_3f
        return total if total > 0 else None
    if normalized_diet == "soft":
        if normalized_area == "2F":
            return _parse_ocr_quantity(entry.get("qty_soft_2f"))
        if normalized_area == "3F":
            return _parse_ocr_quantity(entry.get("qty_soft_3f"))
        qty_2f = _parse_ocr_quantity(entry.get("qty_soft_2f")) or 0.0
        qty_3f = _parse_ocr_quantity(entry.get("qty_soft_3f")) or 0.0
        total = qty_2f + qty_3f
        return total if total > 0 else None
    if normalized_diet == "mixer":
        if normalized_area == "2F":
            return _parse_ocr_quantity(entry.get("qty_mixer_2f"))
        if normalized_area == "3F":
            return _parse_ocr_quantity(entry.get("qty_mixer_3f"))
        qty_2f = _parse_ocr_quantity(entry.get("qty_mixer_2f")) or 0.0
        qty_3f = _parse_ocr_quantity(entry.get("qty_mixer_3f")) or 0.0
        total = qty_2f + qty_3f
        return total if total > 0 else None
    return None


def _build_ocr_menu_meta(order: dict, facility_config: dict | None = None) -> dict[str, object]:
    order_id = order.get("id")
    if not order_id:
        return {}
    parsed = _load_order_ocr_payload(str(order_id))
    if not isinstance(parsed, dict):
        return {}
    year_hint = _resolve_ocr_year_hint(order)
    template = facility_config.get("fax_template") if isinstance(facility_config, dict) else None
    entries = _extract_ocr_entries_from_structured_payload(parsed, template, year_hint)
    if not entries:
        table_raw = parsed.get("table_raw")
        if not isinstance(table_raw, str) or not table_raw.strip():
            return {}
        entries = _parse_ocr_table_rows(table_raw, year_hint)
    meta: dict[tuple[dt_date, str], dict] = {}
    normalized_entries: list[dict] = []
    for idx, entry in enumerate(entries):
        date_val = entry.get("date")
        menu_name = _normalize_menu_key(entry.get("menu_name"))
        if not date_val or not menu_name:
            continue
        key = (date_val, menu_name)
        payload = {
            "date": date_val,
            "menu_name": entry.get("menu_name"),
            "daypart": entry.get("daypart"),
            "category": entry.get("category"),
            "index": idx,
            "qty_regular_2f": entry.get("qty_regular_2f"),
            "qty_regular_3f": entry.get("qty_regular_3f"),
            "qty_soft_2f": entry.get("qty_soft_2f"),
            "qty_soft_3f": entry.get("qty_soft_3f"),
            "qty_mixer_2f": entry.get("qty_mixer_2f"),
            "qty_mixer_3f": entry.get("qty_mixer_3f"),
            "quantity_map": dict(entry.get("quantity_map") or {}),
            "note": entry.get("note"),
            "source": entry.get("source") or "table_raw",
            "ocr_issues": list(entry.get("ocr_issues") or []),
            "needs_review": bool(entry.get("needs_review")),
        }
        normalized_entries.append(payload)
        if key in meta:
            continue
        meta[key] = payload
    return {
        "by_menu": meta,
        "entries": normalized_entries,
        "issues": [
            issue
            for entry in normalized_entries
            for issue in (entry.get("ocr_issues") or [])
            if isinstance(issue, dict)
        ],
        "review_required_count": sum(1 for entry in normalized_entries if entry.get("needs_review")),
    }


def _resolve_delivery_cell(row: dict, column: dict) -> Any:
    source = column.get("source") or column.get("name")
    value = None
    if source == "quantity":
        value = row.get(column.get("name", ""))
    else:
        value = row.get(source)
        if value is None and column.get("name"):
            value = row.get(column["name"])
    if source == "daypart":
        return _normalize_delivery_daypart(value)
    return value


def _delivery_daypart_sort_value(value: Any) -> int:
    return {"朝": 0, "昼": 1, "夕": 2}.get(_normalize_delivery_daypart(value), 99)


def _normalize_delivery_category_label(value: Any) -> str:
    text = str(value or "").strip()
    compact = text.replace(" ", "").replace("　", "")
    if compact in {"副1", "副①", "副菜1", "副菜①", "副菜一"}:
        return "副菜1"
    if compact in {"副2", "副②", "副菜2", "副菜②", "副菜二"}:
        return "副菜2"
    if compact in {"主", "主菜", "主菜1", "主菜①"}:
        return "主菜"
    return text


def _delivery_category_sort_value(daypart: Any, category: Any) -> int:
    label = _normalize_delivery_category_label(category)
    normalized_daypart = _normalize_delivery_daypart(daypart)
    if label == "主菜":
        return 0 if normalized_daypart in {"昼", "夕"} else 2
    if label == "副菜1":
        return 1 if normalized_daypart in {"昼", "夕"} else 0
    if label == "副菜2":
        return 2 if normalized_daypart in {"昼", "夕"} else 1
    return 90


def _delivery_category_from_menu_default(daypart: Any, menu_name: Any, current_category: Any) -> str:
    normalized = _normalize_delivery_category_label(current_category)
    if normalized in {"副菜1", "副菜2", "添え"}:
        return normalized
    default_daypart, default_category, _temp, _qty, _unit = _label_menu_defaults(menu_name)
    if default_category and _normalize_delivery_daypart(default_daypart) == _normalize_delivery_daypart(daypart):
        default_label = _normalize_delivery_category_label(default_category)
        if default_label in {"副菜1", "副菜2", "主菜"}:
            return default_label
    return normalized


def _delivery_display_category_label(daypart: Any, category: Any) -> str:
    label = _normalize_delivery_category_label(category)
    if label == "副菜1":
        return "副①"
    if label == "副菜2":
        return "副②"
    if label == "主菜":
        return "主Ａ" if _normalize_delivery_daypart(daypart) == "昼" else "主"
    return label


def _delivery_reference_slot_sequence(daypart: Any) -> list[str]:
    normalized_daypart = _normalize_delivery_daypart(daypart)
    if normalized_daypart == "朝":
        return ["副菜1", "副菜2"]
    if normalized_daypart == "昼":
        return ["主菜", "副菜1", "副菜2"]
    if normalized_daypart == "夕":
        return ["主菜", "副菜1", "副菜2"]
    return []


def _apply_delivery_reference_slot_categories(rows: list[dict]) -> None:
    grouped: dict[tuple[Any, str], list[dict]] = {}
    for row in rows:
        if _normalize_delivery_category_label(row.get("menu_category")) == "添え":
            continue
        daypart = _normalize_delivery_daypart(row.get("daypart"))
        if not daypart:
            continue
        grouped.setdefault((row.get("date"), daypart), []).append(row)

    for (_date_value, daypart), daypart_rows in grouped.items():
        slot_sequence = _delivery_reference_slot_sequence(daypart)
        if len(daypart_rows) != len(slot_sequence):
            continue
        daypart_rows.sort(
            key=lambda row: (
                row.get("_order_index") if row.get("_order_index") is not None else 1_000_000,
                row.get("menu_name") or "",
            )
        )
        for row, category in zip(daypart_rows, slot_sequence, strict=True):
            row["menu_category"] = category


def _format_delivery_preview_value(value: Any) -> Any:
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def _is_blank_cell_value(value: Any) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def _format_reference_quantity_value(value: Any, original_value: Any) -> Any:
    if _is_blank_cell_value(value):
        if original_value == 0:
            return 0
        return value
    if not isinstance(value, (int, float)):
        return value
    if float(value) == 0 and _is_blank_cell_value(original_value):
        return None
    if isinstance(original_value, str) and re.search(r"\d", original_value):
        return re.sub(r"[-+]?[0-9]*\.?[0-9]+", _format_number(value), original_value, count=1)
    return value


def _copy_cell_style(source, target) -> None:
    target.font = copy(source.font)
    target.border = copy(source.border)
    target.fill = copy(source.fill)
    target.number_format = copy(source.number_format)
    target.protection = copy(source.protection)
    target.alignment = copy(source.alignment)


def _resolve_merged_cell(ws, row: int, column: int):
    cell = ws.cell(row=row, column=column)
    if not isinstance(cell, MergedCell):
        return cell
    for merged in ws.merged_cells.ranges:
        if merged.min_row <= row <= merged.max_row and merged.min_col <= column <= merged.max_col:
            return ws.cell(row=merged.min_row, column=merged.min_col)
    return cell


def _apply_delivery_facility_name(ws, facility_name: str | None) -> None:
    if not facility_name:
        return
    target_cell = None
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            if cell.value is None:
                continue
            text = _normalize_cell_text(cell.value)
            if "施設名" in text:
                target_cell = _resolve_merged_cell(ws, cell.row + 1, cell.col_idx)
                break
        if target_cell:
            break
    if target_cell:
        target_cell.value = facility_name


def _resolve_delivery_column_index(columns: list[dict], column_map: dict[str, int], source: str) -> int | None:
    for col in columns:
        if col.get("source") != source:
            continue
        name = col.get("name")
        if name and name in column_map:
            return column_map[name]
    return None


def _resolve_delivery_menu_column(columns: list[dict], column_map: dict[str, int]) -> int | None:
    for col in columns:
        name = str(col.get("name") or "")
        if col.get("source") == "menu_name" or "献立" in name:
            col_idx = column_map.get(col.get("name"))
            if col_idx:
                return col_idx
    return None


def _normalize_slot_label(value: Any) -> str:
    text = _normalize_cell_text(value)
    if not text:
        return ""
    if "主" in text:
        if "A" in text or "Ａ" in text:
            return "主Ａ"
        if "B" in text or "Ｂ" in text:
            return "主Ｂ"
        return "主"
    if "副" in text:
        if "2" in text or "②" in text:
            return "副②"
        return "副①"
    return text


def _find_delivery_slot_rows(ws, menu_col_idx: int) -> list[int]:
    slots: list[int] = []
    for row in range(1, ws.max_row + 1):
        cell = ws.cell(row=row, column=menu_col_idx)
        text = _normalize_cell_text(cell.value)
        if not text:
            continue
        if "副" in text or "主" in text:
            slots.append(row)
    return slots


def _find_daypart_rows(ws, daypart_col_idx: int) -> dict[str, int]:
    rows: dict[str, int] = {}
    for row in range(1, ws.max_row + 1):
        text = _normalize_cell_text(ws.cell(row=row, column=daypart_col_idx).value)
        if not text:
            continue
        normalized = _normalize_delivery_daypart(text)
        if normalized and normalized not in rows:
            rows[normalized] = row
    return rows


def _format_delivery_weekday(value: dt_date | None) -> str:
    if not value:
        return ""
    weekdays = ["月", "火", "水", "木", "金", "土", "日"]
    return f"({weekdays[value.weekday()]})"


def _build_delivery_slot_map(
    ws,
    slot_rows: list[int],
    menu_col_idx: int,
    daypart_col_idx: int | None,
) -> dict[tuple[str, str], int]:
    slot_map: dict[tuple[str, str], int] = {}
    current_daypart = ""
    for row in slot_rows:
        if daypart_col_idx:
            daypart_text = _normalize_cell_text(ws.cell(row=row, column=daypart_col_idx).value)
            if daypart_text:
                current_daypart = _normalize_delivery_daypart(daypart_text)
        slot_label = _normalize_slot_label(ws.cell(row=row, column=menu_col_idx).value)
        if current_daypart and slot_label:
            slot_map[(current_daypart, slot_label)] = row
    return slot_map


def _build_delivery_slot_label_map(ws, slot_rows: list[int], menu_col_idx: int) -> dict[str, list[int]]:
    label_map: dict[str, list[int]] = {}
    for row in slot_rows:
        slot_label = _normalize_slot_label(ws.cell(row=row, column=menu_col_idx).value)
        if not slot_label:
            continue
        label_map.setdefault(slot_label, []).append(row)
    return label_map


def _build_delivery_slot_label_map_by_daypart(
    ws,
    slot_rows: list[int],
    menu_col_idx: int,
    daypart_col_idx: int | None,
) -> dict[str, dict[str, list[int]]]:
    label_map: dict[str, dict[str, list[int]]] = {}
    current_daypart = ""
    for row in slot_rows:
        if daypart_col_idx:
            daypart_text = _normalize_cell_text(ws.cell(row=row, column=daypart_col_idx).value)
            if daypart_text:
                current_daypart = _normalize_delivery_daypart(daypart_text)
        slot_label = _normalize_slot_label(ws.cell(row=row, column=menu_col_idx).value)
        if current_daypart and slot_label:
            label_map.setdefault(current_daypart, {}).setdefault(slot_label, []).append(row)
    return label_map


def _build_delivery_slot_menu_map_by_daypart(
    ws,
    slot_rows: list[int],
    menu_name_col_idx: int,
    daypart_col_idx: int | None,
) -> dict[str, dict[str, list[int]]]:
    menu_map: dict[str, dict[str, list[int]]] = {}
    current_daypart = ""
    for row_idx in slot_rows:
        if daypart_col_idx:
            daypart_text = _normalize_cell_text(ws.cell(row=row_idx, column=daypart_col_idx).value)
            if daypart_text:
                current_daypart = _normalize_delivery_daypart(daypart_text)
        menu_key = _normalize_menu_key(ws.cell(row=row_idx, column=menu_name_col_idx).value)
        if current_daypart and menu_key:
            menu_map.setdefault(current_daypart, {}).setdefault(menu_key, []).append(row_idx)
    return menu_map


def _write_delivery_slot_row(
    ws,
    row_idx: int,
    row: dict,
    columns: list[dict],
    column_map: dict[str, int],
    include_menu_name: bool,
) -> None:
    for col in columns:
        name = col.get("name")
        if not name:
            continue
        col_idx = column_map.get(name)
        if not col_idx:
            continue
        source = col.get("source")
        if source in {"menu_name", "menu_display"}:
            if source == "menu_name" and not include_menu_name:
                continue
            value = row.get("menu_display") if source == "menu_display" else row.get("menu_name")
        elif source == "quantity":
            value = row.get(name)
        elif source == "note":
            value = row.get("note")
        else:
            continue
        cell = _resolve_merged_cell(ws, row_idx, col_idx)
        if isinstance(cell, MergedCell):
            continue
        cell.value = "" if value is None else value


def _clear_delivery_slot_row(
    ws,
    row_idx: int,
    columns: list[dict],
    column_map: dict[str, int],
    include_menu_name: bool,
) -> None:
    for col in columns:
        name = col.get("name")
        if not name:
            continue
        source = col.get("source")
        if source not in {"menu_name", "menu_display", "quantity", "note"}:
            continue
        if source == "menu_name" and not include_menu_name:
            continue
        col_idx = column_map.get(name)
        if not col_idx:
            continue
        cell = _resolve_merged_cell(ws, row_idx, col_idx)
        if isinstance(cell, MergedCell):
            continue
        cell.value = ""


def _assign_delivery_rows_to_slots(
    rows_for_date: list[dict],
    slot_rows: list[int],
    slot_map: dict[tuple[str, str], int],
    slot_label_map: dict[str, list[int]],
    slot_label_map_by_daypart: dict[str, dict[str, list[int]]],
    slot_rows_by_daypart: dict[str, list[int]],
    slot_menu_map_by_daypart: dict[str, dict[str, list[int]]] | None = None,
) -> dict[int, dict]:
    assignments: dict[int, dict] = {}
    used_rows: set[int] = set()
    pending: list[dict] = []
    rows_for_date.sort(
        key=lambda row: (
            row.get("_order_index") if row.get("_order_index") is not None else 1_000_000,
            row.get("menu_name") or "",
        )
    )
    for row in rows_for_date:
        daypart = _normalize_delivery_daypart(row.get("daypart"))
        menu_key = _normalize_menu_key(row.get("menu_name"))
        target_row = None
        if daypart and menu_key and slot_menu_map_by_daypart:
            for candidate in slot_menu_map_by_daypart.get(daypart, {}).get(menu_key, []):
                if candidate not in used_rows:
                    target_row = candidate
                    break
        slot_label = _normalize_slot_label(row.get("menu_category"))
        if not target_row:
            target_row = slot_map.get((daypart, slot_label)) if daypart and slot_label else None
        if not target_row and slot_label:
            candidates = slot_label_map_by_daypart.get(daypart, {}).get(slot_label, []) if daypart else []
            if not candidates and not daypart:
                candidates = slot_label_map.get(slot_label, [])
            for candidate in candidates:
                if candidate not in used_rows:
                    target_row = candidate
                    break
        if not target_row and daypart:
            for candidate in slot_rows_by_daypart.get(daypart, []):
                if candidate not in used_rows:
                    target_row = candidate
                    break
        if target_row and target_row not in used_rows:
            assignments[target_row] = row
            used_rows.add(target_row)
        else:
            pending.append(row)
    if slot_rows_by_daypart:
        for row in pending:
            daypart = _normalize_delivery_daypart(row.get("daypart"))
            candidates = slot_rows if not daypart else slot_rows_by_daypart.get(daypart, [])
            for candidate in candidates:
                if candidate not in used_rows:
                    assignments[candidate] = row
                    used_rows.add(candidate)
                    break
    else:
        pending_iter = iter(pending)
        for slot_row in slot_rows:
            if slot_row in assignments:
                continue
            row = next(pending_iter, None)
            if row:
                assignments[slot_row] = row
                used_rows.add(slot_row)
    return assignments


def _write_delivery_note(
    path: Path,
    rows: list[dict],
    columns: list[dict],
    template_uri: str | None,
    include_menu_name: bool,
    sheet_name: str | None = None,
    facility_name: str | None = None,
) -> None:
    if not template_uri:
        export_columns = [
            (
                str(col.get("name") or "").strip(),
                str(col.get("header") or col.get("name") or "").strip(),
            )
            for col in columns
            if isinstance(col, dict) and str(col.get("name") or "").strip()
        ]
        export_headers = [header or name for name, header in export_columns]
        if not rows:
            df = pd.DataFrame(columns=export_headers)
        else:
            df = pd.DataFrame(
                [
                    {header or name: row.get(name) for name, header in export_columns}
                    for row in rows
                ],
                columns=export_headers,
            )
        df.to_excel(path, index=False)
        return

    try:
        template_bytes = load_bytes_from_uri(template_uri)
    except Exception as exc:  # noqa: BLE001
        logger.warning("Delivery template load failed", template_uri=template_uri, error=str(exc))
        if not rows:
            df = pd.DataFrame(columns=[col["name"] for col in columns])
        else:
            df = pd.DataFrame(rows)
        df.to_excel(path, index=False)
        return
    workbook = load_workbook(BytesIO(template_bytes))
    if sheet_name and sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
    else:
        ws = workbook.active
    _apply_delivery_facility_name(ws, facility_name)

    header_row = _find_delivery_header_row(ws, columns)
    if not header_row:
        header_row = 1
    column_map = _build_delivery_column_map(ws, header_row, columns)
    menu_col_idx = _resolve_delivery_menu_column(columns, column_map)
    slot_rows = _find_delivery_slot_rows(ws, menu_col_idx) if menu_col_idx else []
    if slot_rows:
        # Slot templates keep the editable column labels immediately above the
        # first menu row; fixed title/company rows can otherwise win detection.
        header_row = max(1, min(slot_rows) - 1)
    _apply_delivery_configured_headers(
        ws,
        header_row,
        columns,
        column_map,
        data_start_row=min(slot_rows) if slot_rows else None,
    )
    if slot_rows:
        original_sheet_count = len(workbook.worksheets)
        for name in list(workbook.sheetnames):
            if workbook[name] is not ws:
                workbook.remove(workbook[name])
        rows_by_date: dict[dt_date | None, list[dict]] = {}
        for row in rows:
            rows_by_date.setdefault(row.get("date"), []).append(row)
        dates = [date_val for date_val in rows_by_date.keys() if date_val]
        dates.sort()
        daypart_col_idx = _resolve_delivery_column_index(columns, column_map, "daypart")
        date_col_idx = _resolve_delivery_column_index(columns, column_map, "date")
        slot_map = (
            _build_delivery_slot_map(ws, slot_rows, menu_col_idx, daypart_col_idx)
            if daypart_col_idx
            else {}
        )
        slot_label_map = _build_delivery_slot_label_map(ws, slot_rows, menu_col_idx)
        slot_label_map_by_daypart = (
            _build_delivery_slot_label_map_by_daypart(ws, slot_rows, menu_col_idx, daypart_col_idx)
            if daypart_col_idx
            else {}
        )
        slot_rows_by_daypart: dict[str, list[int]] = {}
        if daypart_col_idx:
            current_daypart = ""
            for row in slot_rows:
                daypart_text = _normalize_cell_text(ws.cell(row=row, column=daypart_col_idx).value)
                if daypart_text:
                    current_daypart = _normalize_delivery_daypart(daypart_text)
                if current_daypart:
                    slot_rows_by_daypart.setdefault(current_daypart, []).append(row)
        if not dates:
            if len(workbook.worksheets) == original_sheet_count:
                _patch_template_package_with_workbook_values(
                    workbook,
                    template_bytes=template_bytes,
                    output_path=path,
                )
            else:
                workbook.save(path)
            return
        for idx, date_val in enumerate(dates):
            target_ws = ws if idx == 0 else workbook.copy_worksheet(ws)
            if idx == 0:
                target_ws.title = date_val.isoformat()
            else:
                title = date_val.isoformat()
                if title in workbook.sheetnames:
                    title = f"{title}-{idx + 1}"
                target_ws.title = title[:31]
            _apply_delivery_facility_name(target_ws, facility_name)
            if date_col_idx:
                if daypart_col_idx:
                    daypart_rows = _find_daypart_rows(target_ws, daypart_col_idx)
                    morning_row = daypart_rows.get("朝") or slot_rows[0]
                    evening_row = daypart_rows.get("夕")
                else:
                    morning_row = slot_rows[0]
                    evening_row = None
                date_cell = _resolve_merged_cell(target_ws, morning_row, date_col_idx)
                if not isinstance(date_cell, MergedCell):
                    date_cell.value = date_val
                if evening_row:
                    weekday_cell = _resolve_merged_cell(target_ws, evening_row, date_col_idx)
                    if not isinstance(weekday_cell, MergedCell):
                        weekday_cell.value = _format_delivery_weekday(date_val)
            assignments = _assign_delivery_rows_to_slots(
                rows_by_date.get(date_val, []),
                slot_rows,
                slot_map,
                slot_label_map,
                slot_label_map_by_daypart,
                slot_rows_by_daypart,
            )
            for slot_row in slot_rows:
                row_payload = assignments.get(slot_row)
                if row_payload:
                    _write_delivery_slot_row(
                        target_ws,
                        slot_row,
                        row_payload,
                        columns,
                        column_map,
                        include_menu_name,
                    )
                else:
                    _clear_delivery_slot_row(target_ws, slot_row, columns, column_map, include_menu_name)
        if len(workbook.worksheets) == original_sheet_count:
            _patch_template_package_with_workbook_values(
                workbook,
                template_bytes=template_bytes,
                output_path=path,
            )
        else:
            workbook.save(path)
        return

    start_row = _delivery_start_row(ws, header_row)

    for idx, row in enumerate(rows):
        target_row = start_row + idx
        for col in columns:
            name = col.get("name")
            if not name:
                continue
            if not include_menu_name and col.get("source") == "menu_name":
                continue
            col_idx = column_map.get(name)
            if not col_idx:
                continue
            cell = _resolve_merged_cell(ws, target_row, col_idx)
            template_cell = _resolve_merged_cell(ws, start_row, col_idx)
            _copy_cell_style(template_cell, cell)
            if isinstance(cell, MergedCell):
                continue
            cell.value = _resolve_delivery_cell(row, col)

    _patch_template_package_with_workbook_values(
        workbook,
        template_bytes=template_bytes,
        output_path=path,
    )

def _delivery_menu_entry_date(entry: dict) -> dt_date | None:
    return _ensure_date(entry.get("date") or entry.get("menu_date"))


def _delivery_menu_entry_name(entry: dict) -> object:
    return entry.get("menu_name") or entry.get("name")


def _delivery_hidden_condiments_for_menu(menu_name: object) -> list[str]:
    text = str(menu_name or "")
    if any(token in text for token in ("カツ", "ｶﾂ", "フライ", "ﾌﾗｲ")):
        return ["ソース"]
    return []


def _delivery_split_menu_name(menu_name: object) -> tuple[str, list[str]]:
    base, garnish = _split_garnish_name(str(menu_name or "").strip())
    condiments: list[str] = []
    if garnish:
        condiments.append(garnish)
    for condiment in _delivery_hidden_condiments_for_menu(base or menu_name):
        if condiment not in condiments:
            condiments.append(condiment)
    return base or str(menu_name or "").strip(), condiments


def _build_delivery_rows(
    order: dict,
    template: dict,
    quantity_rules: dict,
    facility_config: dict | None = None,
    menu_meta: dict[str, object] | None = None,
    allow_ocr_menu_meta: bool = True,
    timings: dict[str, float] | None = None,
) -> list[dict]:
    setup_started = time.perf_counter()
    columns = template.get("columns", [])
    zero_as_empty = quantity_rules.get("zero_as_empty", True)
    facility_id = order.get("facility")
    area_aliases = _build_area_alias_map(facility_config)
    quantity_columns: list[dict[str, str | None]] = []
    for col in columns:
        if col.get("source") != "quantity":
            continue
        name = col.get("name")
        if not name:
            continue
        diet_type = col.get("diet_type")
        area_id = col.get("area_id")
        if not diet_type or not area_id:
            inferred_diet, inferred_area = _infer_delivery_column_meta(name)
            diet_type = diet_type or inferred_diet
            area_id = area_id or inferred_area
        quantity_columns.append(
            {
                "name": name,
                "diet_key": _normalize_delivery_diet_key(diet_type),
                "area_key": _resolve_area_key(area_id, area_aliases),
            }
        )
    if timings is not None:
        timings["build_rows_setup_ms"] = round((time.perf_counter() - setup_started) * 1000, 1)
    prefer_ocr_rows = bool(template.get("prefer_ocr_raw_rows", False))
    if allow_ocr_menu_meta and not (isinstance(menu_meta, dict) and menu_meta.get("entries")):
        ocr_started = time.perf_counter()
        menu_meta = _build_ocr_menu_meta(order, facility_config)
        if timings is not None:
            timings["build_rows_ocr_menu_meta_ms"] = round((time.perf_counter() - ocr_started) * 1000, 1)
    entries = menu_meta.get("entries") if isinstance(menu_meta, dict) else None
    entry_defaults = _resolve_cached_menu_defaults(
        [
            str(_delivery_menu_entry_name(entry) or "").strip()
            for entry in entries
            if isinstance(entry, dict) and str(_delivery_menu_entry_name(entry) or "").strip()
        ],
        facility_id,
    ) if isinstance(entries, list) else {}
    rows: dict[tuple, dict] = {}
    menu_names = []
    aggregate_started = time.perf_counter()
    for line_index, line in enumerate(order.get("lines", [])):
        line_date = _ensure_date(line.get("date"))
        qty = _safe_qty(line, zero_as_empty)
        if qty is None:
            continue
        raw_menu_name = line.get("menu_name")
        menu_name, split_condiments = _delivery_split_menu_name(raw_menu_name)
        if menu_name:
            menu_names.append(menu_name)
        menu_key = _normalize_menu_key(menu_name)
        meta_map = menu_meta.get("by_menu") if isinstance(menu_meta, dict) else None
        meta = meta_map.get((line_date, menu_key)) if meta_map else None
        daypart_value = line.get("daypart") or line.get("menu_category")
        daypart_key = _normalize_delivery_daypart(daypart_value) or daypart_value
        if meta and not daypart_value:
            daypart_value = meta.get("daypart") or daypart_value
        menu_category = line.get("menu_category") or (meta.get("category") if meta else None)
        order_index = meta.get("index") if meta else line.get("_order_index", line.get("order_index", line_index))
        key = (line_date, daypart_key, menu_name)
        row = rows.setdefault(
            key,
            {
                "date": line_date,
                "daypart": daypart_key,
                "menu_name": menu_name,
                "menu_category": menu_category,
                "menu_display": "",
                "_order_index": order_index,
                "_delivery_condiments": [],
            },
        )
        if menu_category and not row.get("menu_category"):
            row["menu_category"] = menu_category
        if order_index is not None and row.get("_order_index") is None:
            row["_order_index"] = order_index
        line_condiments = _normalize_condiments(line.get("condiments"))
        for condiment in split_condiments:
            if condiment not in line_condiments:
                line_condiments.append(condiment)
        for condiment in line_condiments:
            if condiment not in row["_delivery_condiments"]:
                row["_delivery_condiments"].append(condiment)
        condiment_rows: list[dict] = []
        for condiment in line_condiments:
            condiment_key = (line_date, daypart_key, condiment)
            condiment_rows.append(
                rows.setdefault(
                    condiment_key,
                    {
                        "date": line_date,
                        "daypart": daypart_key,
                        "menu_name": condiment,
                        "menu_category": "添え",
                        "menu_display": "",
                        "_order_index": order_index,
                        "_delivery_condiments": [],
                    },
                )
            )
        line_diet_key = _normalize_delivery_diet_key(line.get("diet_type"))
        line_area_key = _resolve_area_key(line.get("area_id"), area_aliases)
        for col in quantity_columns:
            col_diet_key = col.get("diet_key")
            if col_diet_key and col_diet_key != line_diet_key:
                if not (line_diet_key == "no_fried" and col_diet_key == "forbidden_other"):
                    continue
            col_area_key = col.get("area_key")
            if col_area_key and col_area_key != line_area_key:
                continue
            name = col.get("name")
            if not name:
                continue
            row[name] = (row.get(name) or 0) + float(qty)
            for condiment_row in condiment_rows:
                condiment_row[name] = (condiment_row.get(name) or 0) + float(qty)
    if isinstance(entries, list) and entries:
        allowed_dates = {
            _ensure_date(line.get("date"))
            for line in order.get("lines", [])
            if _ensure_date(line.get("date")) is not None
        }
        for entry in entries:
            if not isinstance(entry, dict):
                continue
            entry_date = _delivery_menu_entry_date(entry)
            if allowed_dates and entry_date not in allowed_dates:
                continue
            raw_menu_name = _delivery_menu_entry_name(entry)
            menu_name, split_condiments = _delivery_split_menu_name(raw_menu_name)
            menu_key = _normalize_menu_key(menu_name)
            if not entry_date or not menu_key:
                continue
            daypart_value = entry.get("daypart")
            daypart_key = _normalize_delivery_daypart(daypart_value) or daypart_value
            menu_category = entry.get("category")
            key = (entry_date, daypart_key, menu_name)
            row = rows.setdefault(
                key,
                {
                    "date": entry_date,
                    "daypart": daypart_key,
                    "menu_name": menu_name,
                    "menu_category": menu_category,
                    "menu_display": "",
                    "_order_index": entry.get("index"),
                    "_delivery_condiments": [],
                },
            )
            entry_condiments = _normalize_condiments(entry.get("condiments"))
            if not entry_condiments:
                entry_condiments = _normalize_condiments(entry_defaults.get(str(menu_name or "").strip(), {}).get("condiments"))
            for condiment in split_condiments:
                if condiment not in entry_condiments:
                    entry_condiments.append(condiment)
            for condiment in entry_condiments:
                if condiment not in row["_delivery_condiments"]:
                    row["_delivery_condiments"].append(condiment)
                condiment_key = (entry_date, daypart_key, condiment)
                rows.setdefault(
                    condiment_key,
                    {
                        "date": entry_date,
                        "daypart": daypart_key,
                        "menu_name": condiment,
                        "menu_category": "添え",
                        "menu_display": "",
                        "_order_index": entry.get("index"),
                        "_delivery_condiments": [],
                    },
                )
    if timings is not None:
        timings["build_rows_aggregate_ms"] = round((time.perf_counter() - aggregate_started) * 1000, 1)
    finalize_started = time.perf_counter()
    default_payloads = _resolve_cached_menu_defaults(menu_names, facility_id)
    for row in rows.values():
        if row.get("_delivery_condiments"):
            continue
        payload = default_payloads.get(str(row.get("menu_name") or "").strip()) or {}
        for condiment in _normalize_condiments(payload.get("condiments")):
            if condiment not in row["_delivery_condiments"]:
                row["_delivery_condiments"].append(condiment)
    result = list(rows.values())
    for row in result:
        row["menu_category"] = _delivery_category_from_menu_default(
            row.get("daypart"),
            row.get("menu_name"),
            row.get("menu_category"),
        )
    _apply_delivery_reference_slot_categories(result)
    result.sort(
        key=lambda row: (
            row.get("date") or "",
            _delivery_daypart_sort_value(row.get("daypart")),
            _delivery_category_sort_value(row.get("daypart"), row.get("menu_category")),
            row.get("_order_index") if row.get("_order_index") is not None else 1_000_000,
            row.get("menu_name") or "",
        )
    )
    for row in result:
        row["menu_category"] = _delivery_display_category_label(row.get("daypart"), row.get("menu_category"))
        condiments = _normalize_condiments(row.get("_delivery_condiments"))
        _apply_condiment_note(row, condiments)
        delivery_menu_name = _append_condiments_to_delivery_menu_name(
            row.get("menu_name"),
            row.get("_delivery_condiments"),
        )
        if row.get("menu_category"):
            row["menu_display"] = f"{row.get('menu_category')} {delivery_menu_name}".strip()
        else:
            row["menu_display"] = delivery_menu_name
    if timings is not None:
        timings["build_rows_finalize_ms"] = round((time.perf_counter() - finalize_started) * 1000, 1)
    return result


def _build_delivery_rows_for_bundle(
    order: dict,
    template: dict,
    quantity_rules: dict,
    facility_config: dict | None,
    menu_meta: dict[str, object] | None,
    *,
    allow_ocr_menu_meta: bool,
) -> list[dict]:
    try:
        return _build_delivery_rows(
            order,
            template,
            quantity_rules,
            facility_config,
            menu_meta,
            allow_ocr_menu_meta=allow_ocr_menu_meta,
        )
    except TypeError as exc:
        if "allow_ocr_menu_meta" not in str(exc):
            raise
        return _build_delivery_rows(order, template, quantity_rules, facility_config, menu_meta)


def _build_label_rows(
    bags: list[dict],
    label_profile: dict,
    facility_name: str | None,
) -> tuple[list[dict], list[str], str]:
    label_fields, label_format = _resolve_label_fields(label_profile)
    if label_format == "legacy":
        labels = [
            _label_payload_legacy(bag, label_profile, facility_name)
            for bag in bags
            if not _is_daily_label_excluded_bag(bag)
        ]
        return labels, label_fields, label_format
    labels = [_label_payload_jp(bag, label_profile) for bag in bags if not _is_daily_label_excluded_bag(bag)]
    merged = _merge_label_rows(labels, label_fields)
    return merged, label_fields, label_format


def _build_total_rows(
    order_lines: list[dict],
    label_profile: dict,
    facility_name: str | None,
    quantity_rules: dict,
) -> tuple[list[dict], list[str], str]:
    label_fields, label_format = _resolve_label_fields(label_profile)
    zero_as_empty = quantity_rules.get("zero_as_empty", True)
    grouped: dict[tuple, dict] = {}
    for line in order_lines:
        line_date = _ensure_date(line.get("date"))
        qty = _safe_qty(line, zero_as_empty)
        if qty is None:
            continue
        key = (
            line_date,
            line.get("daypart"),
            line.get("menu_category"),
            line.get("menu_name"),
            line.get("menu_temp_type"),
            line.get("menu_qty_per_serving"),
            line.get("menu_unit_type"),
        )
        row = grouped.setdefault(
            key,
            {
                "date": line_date,
                "daypart": line.get("daypart"),
                "menu_category": line.get("menu_category"),
                "menu_name": line.get("menu_name"),
                "menu_temp_type": line.get("menu_temp_type"),
                "menu_qty_per_serving": line.get("menu_qty_per_serving"),
                "menu_unit_type": line.get("menu_unit_type"),
                "quantity": 0.0,
            },
        )
        row["quantity"] += float(qty)
    if label_format == "legacy":
        labels = [
            _label_payload_legacy(row, label_profile, facility_name) for row in grouped.values()
        ]
        return labels, label_fields, label_format
    labels = [_label_payload_jp(row, label_profile) for row in grouped.values()]
    merged = _merge_label_rows(labels, label_fields)
    for row in merged:
        row["発行枚数"] = ""
    return merged, label_fields, label_format


def _write_aggregate_csv(path: Path, rows: list[dict], label_fields: list[str]) -> None:
    fieldnames = label_fields or (list(rows[0].keys()) if rows else [])
    with path.open("w", newline="", encoding="cp932", errors="replace") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow({k: row.get(k, "") for k in fieldnames})


def _safe_filename_segment(value: object, fallback: str) -> str:
    text = str(value or "").strip()
    if not text:
        return fallback
    text = re.sub(r"[^\w.-]+", "_", text)
    text = text.strip("._")
    return text or fallback


def _safe_sheet_title(value: object, fallback: str, used_titles: set[str]) -> str:
    text = str(value or "").strip() or fallback
    text = re.sub(r'[\\/*?:\[\]]+', "_", text)
    text = re.sub(r"\s+", " ", text).strip()
    if not text:
        text = fallback
    base = text[:31] or fallback
    title = base
    suffix = 2
    while title in used_titles:
        tail = f"-{suffix}"
        title = f"{base[: max(31 - len(tail), 1)]}{tail}"
        suffix += 1
    used_titles.add(title)
    return title


def _copy_sheet_contents(source_ws, target_ws) -> None:
    for row in source_ws.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            target = target_ws.cell(row=cell.row, column=cell.column)
            target.value = cell.value
            _copy_cell_style(cell, target)
    for merged_range in source_ws.merged_cells.ranges:
        target_ws.merge_cells(str(merged_range))
    for col_key, dim in source_ws.column_dimensions.items():
        target_dim = target_ws.column_dimensions[col_key]
        target_dim.width = dim.width
        target_dim.hidden = dim.hidden
        target_dim.bestFit = dim.bestFit
    for row_key, dim in source_ws.row_dimensions.items():
        target_dim = target_ws.row_dimensions[row_key]
        target_dim.height = dim.height
        target_dim.hidden = dim.hidden
    target_ws.sheet_view.showGridLines = source_ws.sheet_view.showGridLines
    target_ws.freeze_panes = source_ws.freeze_panes
    target_ws.sheet_format.defaultRowHeight = source_ws.sheet_format.defaultRowHeight
    target_ws.sheet_format.defaultColWidth = source_ws.sheet_format.defaultColWidth
    target_ws.page_margins = copy(source_ws.page_margins)
    target_ws.print_options = copy(source_ws.print_options)
    target_ws.page_setup = copy(source_ws.page_setup)


def _cell_text_from_merged(ws, row: int, column: int) -> str:
    cell = _resolve_merged_cell(ws, row, column)
    return _normalize_cell_text(cell.value)


def _daily_delivery_header_text(ws, column: int) -> str:
    parts: list[str] = []
    seen: set[str] = set()
    for row in range(8, min(ws.max_row, 11) + 1):
        text = _cell_text_from_merged(ws, row, column)
        if text and text not in seen:
            parts.append(text)
            seen.add(text)
    return "".join(parts)


def _daily_delivery_column_meta(ws) -> list[dict]:
    columns: list[dict] = [
        {"name": "日付", "source": "date", "column_index": 1},
        {"name": "区分", "source": "daypart", "column_index": 2},
        {"name": "献立区分", "source": "menu_category", "column_index": 3},
        {"name": "メニュー名", "source": "menu_display", "column_index": 4},
    ]
    used_names = {str(col["name"]) for col in columns}
    for col_idx in range(5, ws.max_column + 1):
        header = _daily_delivery_header_text(ws, col_idx)
        if not header:
            continue
        if "保管温度" in header or "賞味期限" in header:
            continue
        if "備考" in header:
            name = "備考"
            suffix = 2
            while name in used_names:
                name = f"備考{suffix}"
                suffix += 1
            used_names.add(name)
            columns.append({"name": name, "source": "note", "column_index": col_idx})
            continue
        diet_type = None
        area_id = None
        if "2F" in header:
            area_id = "2F"
        elif "3F" in header:
            area_id = "3F"
        elif "月" in header:
            area_id = "3F"
        elif "花" in header:
            area_id = "2F"
        if "肉禁" in header:
            diet_type = "no_meat"
        elif "魚禁" in header:
            diet_type = "no_fish"
        elif "揚げ物" in header:
            diet_type = "no_fried"
        elif "その他" in header:
            diet_type = "forbidden_other"
        elif "禁食" in header:
            diet_type = "禁食"
        elif "常食" in header:
            diet_type = "regular"
        elif "小口" in header:
            diet_type = "小口"
        elif "通所" in header:
            diet_type = "daycare"
        elif "職員" in header:
            diet_type = "staff"
        elif "糖尿" in header:
            diet_type = "diabetes"
        elif "軟菜" in header:
            diet_type = "soft"
        elif "ミキサ" in header or "ﾐｷｻ" in header:
            diet_type = "mixer"
        elif header == "袋分け":
            diet_type = "regular_bag"
        if not diet_type:
            continue
        name = header
        suffix = 2
        while name in used_names:
            name = f"{header}_{suffix}"
            suffix += 1
        used_names.add(name)
        columns.append(
            {
                "name": name,
                "source": "quantity",
                "diet_type": diet_type,
                "area_id": area_id,
                "column_index": col_idx,
            }
        )
    return columns


def _daily_delivery_table_max_column(ws) -> int:
    columns = _daily_delivery_column_meta(ws)
    if not columns:
        return 4
    return max(int(col.get("column_index") or 0) for col in columns)


def _clear_daily_delivery_sheet_data(ws) -> None:
    note_columns = {
        int(col.get("column_index"))
        for col in _daily_delivery_column_meta(ws)
        if col.get("source") == "note" and isinstance(col.get("column_index"), int)
    }
    for row_idx in range(12, min(ws.max_row, 19) + 1):
        for col_idx in range(5, ws.max_column + 1):
            if col_idx in note_columns:
                continue
            cell = _resolve_merged_cell(ws, row_idx, col_idx)
            if isinstance(cell, MergedCell):
                continue
            if isinstance(cell.value, str) and cell.value.startswith("="):
                continue
            cell.value = None


def _materialize_daily_delivery_static_cells(ws, display_ws) -> None:
    if display_ws is None:
        return
    # openpyxl drops cached formula values on save. The reference daily delivery
    # workbook uses formulas for static date/menu labels, so materialize those
    # display values before writing quantities.
    for row_idx in range(12, min(ws.max_row, 19) + 1):
        for col_idx in range(1, 5):
            display_value = display_ws.cell(row=row_idx, column=col_idx).value
            if display_value is None:
                continue
            cell = _resolve_merged_cell(ws, row_idx, col_idx)
            if isinstance(cell, MergedCell):
                continue
            cell.value = display_value


def _restore_daily_delivery_table_borders(ws, template_ws) -> None:
    if template_ws is None:
        return
    max_col = max(ws.max_column, template_ws.max_column)
    for row_idx in range(12, min(max(ws.max_row, template_ws.max_row), 19) + 1):
        for col_idx in range(1, max_col + 1):
            target = ws.cell(row=row_idx, column=col_idx)
            if isinstance(target, MergedCell):
                continue
            source = template_ws.cell(row=row_idx, column=col_idx)
            target.border = copy(source.border)


def _xlsx_tag(name: str) -> str:
    return f"{{{_XLSX_MAIN_NS}}}{name}"


def _xlsx_package_rel_tag(name: str) -> str:
    return f"{{{_XLSX_PACKAGE_REL_NS}}}{name}"


def _xlsx_content_type_tag(name: str) -> str:
    return f"{{{_XLSX_CONTENT_TYPE_NS}}}{name}"


def _excel_serial_date(value: dt_date | datetime) -> float:
    date_value = value.date() if isinstance(value, datetime) else value
    return float((date_value - dt_date(1899, 12, 30)).days)


def _worksheet_paths_by_name(xlsx_parts: dict[str, bytes]) -> dict[str, str]:
    workbook_root = ET.fromstring(xlsx_parts["xl/workbook.xml"])
    rels_root = ET.fromstring(xlsx_parts["xl/_rels/workbook.xml.rels"])
    rel_targets = {rel.attrib["Id"]: rel.attrib["Target"] for rel in rels_root}
    paths: dict[str, str] = {}
    sheets = workbook_root.find(_xlsx_tag("sheets"))
    if sheets is None:
        return paths
    for sheet in sheets:
        sheet_name = str(sheet.attrib.get("name") or "")
        rel_id = sheet.attrib.get(f"{{{_XLSX_REL_NS}}}id")
        target = rel_targets.get(str(rel_id or ""))
        if not sheet_name or not target:
            continue
        paths[sheet_name] = target.lstrip("/") if target.startswith("/") else f"xl/{target}"
    return paths


def _shared_strings(xlsx_parts: dict[str, bytes]) -> list[str]:
    data = xlsx_parts.get("xl/sharedStrings.xml")
    if not data:
        return []
    root = ET.fromstring(data)
    values: list[str] = []
    for item in root.findall(_xlsx_tag("si")):
        texts = [node.text or "" for node in item.findall(f".//{_xlsx_tag('t')}")]
        values.append("".join(texts))
    return values


def _style_ids_with_bottom_border(xlsx_parts: dict[str, bytes]) -> set[int]:
    data = xlsx_parts.get("xl/styles.xml")
    if not data:
        return set()
    root = ET.fromstring(data)
    borders = list(root.find(_xlsx_tag("borders")) or [])
    style_ids: set[int] = set()
    for style_idx, xf in enumerate(list(root.find(_xlsx_tag("cellXfs")) or [])):
        border_id_text = xf.attrib.get("borderId")
        if border_id_text is None or not border_id_text.isdigit():
            continue
        border_id = int(border_id_text)
        if border_id >= len(borders):
            continue
        bottom = borders[border_id].find(_xlsx_tag("bottom"))
        if bottom is not None and bottom.attrib.get("style"):
            style_ids.add(style_idx)
    return style_ids


def _cell_column_index(coordinate: str) -> int | None:
    match = re.match(r"([A-Z]+)", str(coordinate or ""))
    if not match:
        return None
    return column_index_from_string(match.group(1))


def _set_xml_attr_text(attrs: str, name: str, value: str) -> str:
    if re.search(rf'\s{name}="[^"]*"', attrs):
        return re.sub(rf'\s{name}="[^"]*"', f' {name}="{value}"', attrs, count=1)
    return f'{attrs} {name}="{value}"'


def _remove_xml_attr_text(attrs: str, name: str) -> str:
    return re.sub(rf'\s{name}="[^"]*"', "", attrs)


def _patch_sheet_column_width_xml_text(sheet_xml: str, column_index: int, width: float | None) -> str:
    if not width:
        return sheet_xml
    pattern = re.compile(r'(<col\b[^>]*\bmin="' + str(column_index) + r'"[^>]*\bmax="' + str(column_index) + r'"[^>]*)/?>')

    def repl(match: re.Match[str]) -> str:
        attrs = match.group(1).rstrip("/")
        attrs = _set_xml_attr_text(attrs, "width", _format_number(width))
        attrs = _set_xml_attr_text(attrs, "customWidth", "1")
        return f"{attrs}/>"

    return pattern.sub(repl, sheet_xml, count=1)


def _xml_cell_inner_value(value: Any) -> tuple[str | None, str]:
    if _is_blank_cell_value(value):
        return None, ""
    if isinstance(value, (datetime, dt_date)):
        return "n", f"<v>{_format_number(_excel_serial_date(value))}</v>"
    if isinstance(value, bool):
        return "b", f"<v>{'1' if value else '0'}</v>"
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return "n", f"<v>{_format_number(value)}</v>"
    return "inlineStr", f"<is><t>{xml_escape(str(value))}</t></is>"


def _xml_cell_with_value(attrs: str, value: Any) -> str:
    attrs = attrs.rstrip("/")
    attrs = _remove_xml_attr_text(attrs, "t")
    cell_type, inner = _xml_cell_inner_value(value)
    if cell_type:
        attrs = _set_xml_attr_text(attrs, "t", cell_type)
    if not inner:
        return f"<c{attrs}/>"
    return f"<c{attrs}>{inner}</c>"


def _replace_xml_cell_value(sheet_xml: str, coordinate: str, value: Any) -> str:
    escaped_coordinate = re.escape(coordinate)
    pattern = re.compile(
        rf'<c(?P<attrs>[^>]*\br="{escaped_coordinate}"[^>]*)(?:/>|>.*?</c>)',
        re.DOTALL,
    )

    def replace(match: re.Match[str]) -> str:
        return _xml_cell_with_value(match.group("attrs"), value)

    next_xml, count = pattern.subn(replace, sheet_xml, count=1)
    if count:
        return next_xml

    row_match = re.search(
        rf'(<row(?P<attrs>[^>]*\br="{re.escape(str(re.sub(r"^[A-Z]+", "", coordinate)))}"[^>]*)>)(?P<body>.*?)(</row>)',
        sheet_xml,
        re.DOTALL,
    )
    if row_match is None:
        return sheet_xml
    new_cell = _xml_cell_with_value(f' r="{coordinate}"', value)
    return (
        sheet_xml[: row_match.end("body")]
        + new_cell
        + sheet_xml[row_match.end("body") :]
    )


def _patch_daily_delivery_evening_bottom_border_xml_text(sheet_xml: str, bottom_style_ids: set[int]) -> str:
    if not bottom_style_ids:
        return sheet_xml
    row_match = re.search(r'(<row[^>]*\br="19"[^>]*>)(?P<body>.*?)(</row>)', sheet_xml, re.DOTALL)
    if row_match is None:
        return sheet_xml
    body = row_match.group("body")
    cell_pattern = re.compile(r'<c(?P<attrs>[^>]*\br="[A-Z]+19"[^>]*)(?:/>|>.*?</c>)', re.DOTALL)
    cells: list[tuple[int, str, str, int | None]] = []
    source_styles: list[tuple[int, int]] = []
    for match in cell_pattern.finditer(body):
        attrs = match.group("attrs")
        coord_match = re.search(r'\br="([^"]+)"', attrs)
        style_match = re.search(r'\bs="([^"]+)"', attrs)
        col_idx = _cell_column_index(coord_match.group(1) if coord_match else "")
        style_id = int(style_match.group(1)) if style_match and style_match.group(1).isdigit() else None
        if col_idx is None:
            continue
        cells.append((match.start(), match.end(), attrs, col_idx, style_id))  # type: ignore[arg-type]
        if style_id in bottom_style_ids:
            source_styles.append((col_idx, int(style_id)))
    if not source_styles:
        return sheet_xml
    next_body = body
    for start, end, attrs, col_idx, style_id in reversed(cells):  # type: ignore[misc]
        if style_id in bottom_style_ids:
            continue
        nearest_style = min(source_styles, key=lambda item: abs(item[0] - col_idx))[1]
        original_cell = next_body[start:end]
        patched_attrs = _set_xml_attr_text(attrs, "s", str(nearest_style))
        patched_cell = original_cell.replace(f"<c{attrs}", f"<c{patched_attrs}", 1)
        next_body = next_body[:start] + patched_cell + next_body[end:]
    return sheet_xml[: row_match.start("body")] + next_body + sheet_xml[row_match.end("body") :]


def _xlsx_parts_from_bytes(template_bytes: bytes) -> dict[str, bytes]:
    with zipfile.ZipFile(BytesIO(template_bytes), "r") as source:
        return {name: source.read(name) for name in source.namelist()}


def _write_xlsx_parts(output_path: Path, parts: dict[str, bytes]) -> None:
    with zipfile.ZipFile(output_path, "w", zipfile.ZIP_DEFLATED) as target:
        for name, content in parts.items():
            target.writestr(name, content)


def _drop_calc_chain(parts: dict[str, bytes]) -> None:
    if "xl/calcChain.xml" not in parts:
        return
    parts.pop("xl/calcChain.xml", None)
    rels_path = "xl/_rels/workbook.xml.rels"
    if rels_path in parts:
        rels_root = ET.fromstring(parts[rels_path])
        for rel in list(rels_root.findall(_xlsx_package_rel_tag("Relationship"))):
            if str(rel.attrib.get("Target") or "").endswith("calcChain.xml"):
                rels_root.remove(rel)
        parts[rels_path] = ET.tostring(rels_root, encoding="utf-8", xml_declaration=True)
    content_types_path = "[Content_Types].xml"
    if content_types_path in parts:
        content_root = ET.fromstring(parts[content_types_path])
        for override in list(content_root.findall(_xlsx_content_type_tag("Override"))):
            if override.attrib.get("PartName") == "/xl/calcChain.xml":
                content_root.remove(override)
        parts[content_types_path] = ET.tostring(content_root, encoding="utf-8", xml_declaration=True)


def _ensure_xml_row(sheet_data, row_idx: int):
    rows = {int(row.attrib["r"]): row for row in sheet_data.findall(_xlsx_tag("row")) if row.attrib.get("r")}
    row = rows.get(row_idx)
    if row is not None:
        return row
    row = ET.Element(_xlsx_tag("row"), {"r": str(row_idx)})
    inserted = False
    for index, existing in enumerate(list(sheet_data)):
        existing_r = int(existing.attrib.get("r", "0") or 0)
        if existing_r > row_idx:
            sheet_data.insert(index, row)
            inserted = True
            break
    if not inserted:
        sheet_data.append(row)
    return row


def _ensure_xml_cell(row, coordinate: str):
    cells = {cell.attrib.get("r"): cell for cell in row.findall(_xlsx_tag("c"))}
    cell = cells.get(coordinate)
    if cell is not None:
        return cell
    cell = ET.Element(_xlsx_tag("c"), {"r": coordinate})
    inserted = False
    for index, existing in enumerate(list(row)):
        existing_ref = str(existing.attrib.get("r") or "")
        if existing_ref and existing_ref > coordinate:
            row.insert(index, cell)
            inserted = True
            break
    if not inserted:
        row.append(cell)
    return cell


def _set_xml_cell_value(cell, value: Any) -> None:
    for child in list(cell):
        if child.tag in {_xlsx_tag("f"), _xlsx_tag("v"), _xlsx_tag("is")}:
            cell.remove(child)
    cell.attrib.pop("t", None)
    if _is_blank_cell_value(value):
        return
    if isinstance(value, (datetime, dt_date)):
        cell.attrib["t"] = "n"
        node = ET.SubElement(cell, _xlsx_tag("v"))
        node.text = _format_number(_excel_serial_date(value))
        return
    if isinstance(value, bool):
        cell.attrib["t"] = "b"
        node = ET.SubElement(cell, _xlsx_tag("v"))
        node.text = "1" if value else "0"
        return
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        cell.attrib["t"] = "n"
        node = ET.SubElement(cell, _xlsx_tag("v"))
        node.text = _format_number(value)
        return
    cell.attrib["t"] = "inlineStr"
    inline = ET.SubElement(cell, _xlsx_tag("is"))
    text = ET.SubElement(inline, _xlsx_tag("t"))
    text.text = str(value)


def _is_delivery_static_artifact_value(value: Any) -> bool:
    if value in {"v", "V", 0, "0"}:
        return True
    return False


def _remove_delivery_static_artifacts(workbook: Workbook) -> None:
    for ws in workbook.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue
                if _is_delivery_static_artifact_value(cell.value):
                    cell.value = None


def _patch_template_package_with_workbook_values(
    workbook: Workbook,
    *,
    template_bytes: bytes,
    output_path: Path,
    min_row: int = 1,
    max_row: int | None = None,
) -> None:
    parts = _xlsx_parts_from_bytes(template_bytes)
    _drop_calc_chain(parts)
    shared_strings = _shared_strings(parts)
    bottom_style_ids = _style_ids_with_bottom_border(parts)
    workbook_root = ET.fromstring(parts["xl/workbook.xml"])
    sheet_paths = _worksheet_paths_by_name(parts)
    sheet_nodes = list(workbook_root.find(_xlsx_tag("sheets")) or [])
    if len(workbook.worksheets) != len(sheet_nodes):
        raise ValueError("template-preserving xlsx output cannot add or remove sheets")
    ordered_paths: list[str] = []
    workbook_xml_changed = False
    for sheet_node in sheet_nodes:
        rel_id = sheet_node.attrib.get(f"{{{_XLSX_REL_NS}}}id")
        rels_root = ET.fromstring(parts["xl/_rels/workbook.xml.rels"])
        rel_targets = {rel.attrib["Id"]: rel.attrib["Target"] for rel in rels_root}
        target = rel_targets.get(str(rel_id or ""))
        ordered_paths.append(target.lstrip("/") if target.startswith("/") else f"xl/{target}")
    for index, ws in enumerate(workbook.worksheets):
        if sheet_nodes[index].attrib.get("name") != ws.title:
            sheet_nodes[index].attrib["name"] = ws.title
            workbook_xml_changed = True
        sheet_path = sheet_paths.get(ws.title) or ordered_paths[index]
        if not sheet_path or sheet_path not in parts:
            raise ValueError(f"template-preserving xlsx output cannot add or rename sheet: {ws.title}")
        row_stop = max_row if max_row is not None else ws.max_row
        row_indices = set(range(min_row, min(ws.max_row, row_stop) + 1))
        original_root = ET.fromstring(parts[sheet_path])
        original_sheet_data = original_root.find(_xlsx_tag("sheetData"))
        if original_sheet_data is not None:
            for original_row in original_sheet_data.findall(_xlsx_tag("row")):
                original_row_idx = int(original_row.attrib.get("r", "0") or 0)
                for original_cell in original_row.findall(_xlsx_tag("c")):
                    value_node = original_cell.find(_xlsx_tag("v"))
                    inline_node = original_cell.find(_xlsx_tag("is"))
                    original_value = value_node.text if value_node is not None else None
                    if (
                        original_cell.attrib.get("t") == "s"
                        and original_value is not None
                        and str(original_value).isdigit()
                    ):
                        index = int(original_value)
                        if 0 <= index < len(shared_strings):
                            original_value = shared_strings[index]
                    if inline_node is not None:
                        text_node = inline_node.find(_xlsx_tag("t"))
                        original_value = text_node.text if text_node is not None else original_value
                    if _is_delivery_static_artifact_value(original_value):
                        row_indices.add(original_row_idx)
        sheet_xml = parts[sheet_path].decode("utf-8")
        for row_idx in sorted(row_indices):
            for col_idx in range(1, ws.max_column + 1):
                coordinate = f"{get_column_letter(col_idx)}{row_idx}"
                sheet_xml = _replace_xml_cell_value(sheet_xml, coordinate, ws.cell(row=row_idx, column=col_idx).value)
        for col_key, dim in ws.column_dimensions.items():
            col_idx = column_index_from_string(col_key)
            sheet_xml = _patch_sheet_column_width_xml_text(sheet_xml, col_idx, dim.width)
        sheet_xml = _patch_daily_delivery_evening_bottom_border_xml_text(sheet_xml, bottom_style_ids)
        parts[sheet_path] = sheet_xml.encode("utf-8")
    if workbook_xml_changed:
        parts["xl/workbook.xml"] = ET.tostring(workbook_root, encoding="utf-8", xml_declaration=True)
    _write_xlsx_parts(output_path, parts)


def _save_reference_daily_delivery_workbook_preserving_template_package(
    workbook: Workbook,
    output_path: Path,
) -> None:
    workbook.save(output_path)


def _reference_delivery_sheet_name(facility_code: str | None, facility_name: str | None) -> str | None:
    code = str(facility_code or "").strip()
    if code in DAILY_DELIVERY_SHEET_BY_FACILITY_ID:
        return DAILY_DELIVERY_SHEET_BY_FACILITY_ID[code]
    name = str(facility_name or "").strip()
    if name:
        title = re.sub(r'[\\/*?:\[\]]+', "_", name)
        return (re.sub(r"\s+", " ", title).strip() or name)[:31]
    return code or None


def _write_reference_daily_delivery_sheet(
    ws,
    *,
    rows: list[dict],
    target_date: dt_date,
    display_ws=None,
) -> None:
    _materialize_daily_delivery_static_cells(ws, display_ws)
    columns = _daily_delivery_column_meta(ws)
    column_map = {
        str(col.get("name")): int(col.get("column_index"))
        for col in columns
        if col.get("name") and isinstance(col.get("column_index"), int)
    }
    slot_rows = list(range(12, min(ws.max_row, 19) + 1))
    menu_col_idx = 3
    daypart_col_idx = 2
    slot_source_ws = display_ws or ws
    slot_map = _build_delivery_slot_map(slot_source_ws, slot_rows, menu_col_idx, daypart_col_idx)
    slot_label_map = _build_delivery_slot_label_map(slot_source_ws, slot_rows, menu_col_idx)
    slot_label_map_by_daypart = _build_delivery_slot_label_map_by_daypart(
        slot_source_ws, slot_rows, menu_col_idx, daypart_col_idx
    )
    slot_menu_map_by_daypart = _build_delivery_slot_menu_map_by_daypart(slot_source_ws, slot_rows, 4, daypart_col_idx)
    slot_rows_by_daypart: dict[str, list[int]] = {"朝": [], "昼": [], "夕": []}
    current_daypart = ""
    for row_idx in slot_rows:
        daypart_text = _normalize_cell_text(ws.cell(row=row_idx, column=daypart_col_idx).value)
        if daypart_text:
            current_daypart = _normalize_delivery_daypart(daypart_text)
        if current_daypart:
            slot_rows_by_daypart.setdefault(current_daypart, []).append(row_idx)
    assignments = _assign_delivery_rows_to_slots(
        [dict(row) for row in rows if _ensure_date(row.get("date")) == target_date],
        slot_rows,
        slot_map,
        slot_label_map,
        slot_label_map_by_daypart,
        slot_rows_by_daypart,
        slot_menu_map_by_daypart,
    )
    for row_idx in slot_rows:
        row_payload = assignments.get(row_idx)
        for col in columns:
            source = col.get("source")
            if source not in {"date", "daypart", "menu_category", "menu_display"}:
                continue
            col_idx = int(col.get("column_index") or 0)
            if not col_idx:
                continue
            cell = ws.cell(row=row_idx, column=col_idx)
            if isinstance(cell, MergedCell):
                continue
            display_value = display_ws.cell(row=row_idx, column=col_idx).value if display_ws is not None else cell.value
            if source == "date":
                if row_idx == slot_rows[0]:
                    cell.value = target_date
                elif isinstance(display_value, str) and re.fullmatch(r"\([月火水木金土日]\)", display_value.strip()):
                    cell.value = _format_delivery_weekday(target_date)
                continue
            if not row_payload:
                continue
            if source == "daypart":
                if _is_blank_cell_value(display_value):
                    continue
                cell.value = _normalize_delivery_daypart(row_payload.get("daypart"))
            elif source == "menu_category":
                cell.value = row_payload.get("menu_category") or ""
            elif source == "menu_display":
                cell.value = row_payload.get("menu_display") or row_payload.get("menu_name") or ""
        for col in columns:
            source = col.get("source")
            if source not in {"quantity", "note"}:
                continue
            name = str(col.get("name") or "")
            col_idx = column_map.get(name)
            if not col_idx:
                continue
            cell = _resolve_merged_cell(ws, row_idx, col_idx)
            if isinstance(cell, MergedCell):
                continue
            original_value = display_ws.cell(row=row_idx, column=col_idx).value if display_ws is not None else None
            if not row_payload:
                cell.value = 0 if original_value == 0 else None
                continue
            value = row_payload.get(name) if source == "quantity" else row_payload.get("note")
            if source == "quantity":
                value = _format_reference_quantity_value(value, original_value)
            elif _is_blank_cell_value(value):
                continue
            cell.value = "" if value is None else value
    _restore_daily_delivery_table_borders(ws, display_ws)


def _create_reference_daily_delivery_workbook(
    *,
    target_date: dt_date,
    grouped_outputs: dict[str, dict[str, Any]],
) -> Workbook:
    if not DAILY_DELIVERY_REFERENCE_TEMPLATE.exists():
        raise ValueError(f"daily delivery reference template not found: {DAILY_DELIVERY_REFERENCE_TEMPLATE}")
    workbook = load_workbook(DAILY_DELIVERY_REFERENCE_TEMPLATE)
    if not workbook.worksheets:
        raise ValueError(f"daily delivery reference template has no sheets: {DAILY_DELIVERY_REFERENCE_TEMPLATE}")
    template_ws = workbook.worksheets[0]
    template_ws.title = "テンプレート"
    _remove_delivery_static_artifacts(workbook)
    template_ws.column_dimensions["D"].width = max(template_ws.column_dimensions["D"].width or 0, 36)
    _clear_daily_delivery_sheet_data(template_ws)
    template_columns = _daily_delivery_column_meta(template_ws)
    used_titles: set[str] = set()
    for group in grouped_outputs.values():
        sheet_name = _reference_delivery_sheet_name(group.get("facility_code"), group.get("facility_name"))
        if not sheet_name:
            continue
        ws = workbook.copy_worksheet(template_ws)
        ws.title = _safe_sheet_title(sheet_name, "納品書", used_titles)
        facility_name = group.get("facility_name") or (group.get("facility_config") or {}).get("facility_name")
        _apply_delivery_facility_name(ws, facility_name)
        sheet_template = {
            "columns": template_columns,
            "prefer_ocr_raw_rows": bool((group.get("invoice_template") or {}).get("prefer_ocr_raw_rows", False)),
        }
        rows: list[dict] = []
        for ctx in group.get("contexts", []):
            rows.extend(
                _build_delivery_rows_for_bundle(
                    ctx["order_for_outputs"],
                    sheet_template,
                    {**ctx["quantity_rules"], "zero_as_empty": False},
                    ctx["facility_config"],
                    ctx.get("ocr_menu_meta"),
                    allow_ocr_menu_meta=False,
                )
            )
        merged_rows = _merge_delivery_bundle_rows(rows, sheet_template)
        _write_reference_daily_delivery_sheet(
            ws,
            rows=merged_rows,
            target_date=target_date,
            display_ws=None,
        )
    workbook.remove(template_ws)
    if not workbook.worksheets:
        ws = workbook.create_sheet("納品書")
        _write_reference_daily_delivery_sheet(ws, rows=[], target_date=target_date, display_ws=None)
    return workbook


def _apply_daily_label_sheet_shape(ws, max_rows: int | None = None) -> None:
    thin = Side(style="thin", color="000000")
    hair = Side(style="hair", color="000000")
    header_fill = PatternFill("solid", fgColor="FFFF99")
    cold_fill = PatternFill("solid", fgColor="CCFFFF")
    warm_fill = PatternFill("solid", fgColor="FFCCFF")
    header_font = Font(name="ＭＳ Ｐゴシック", size=11, bold=False)
    body_font = Font(name="ＭＳ Ｐゴシック", size=11)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    for row_idx in range(1, (max_rows or ws.max_row) + 1):
        ws.row_dimensions[row_idx].height = 25.5
    if max_rows and ws.max_row < max_rows:
        for row_idx in range(ws.max_row + 1, max_rows + 1):
            ws.cell(row=row_idx, column=11, value="")
    ws.row_dimensions[1].height = 39
    for col_idx, width in enumerate([8.125, 6.5, 15.25, 6.375, 14.625, 10, 22.5, 13.75, 9, 12, 10], start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 80), min_col=1, max_col=11):
        for cell in row:
            cell.font = body_font
            cell.alignment = left if cell.column in {7, 8} else center
            if cell.row == 1:
                cell.font = header_font
                cell.fill = header_fill
            elif cell.column == 6 and str(cell.value or "") == "冷菜":
                cell.fill = cold_fill
            elif cell.column == 6 and str(cell.value or "") == "温菜":
                cell.fill = warm_fill
            cell.border = Border(
                left=thin if cell.column == 1 else hair,
                right=thin if cell.column == 11 else hair,
                top=thin if cell.row == 1 else hair,
                bottom=thin if cell.row in {1, ws.max_row} else hair,
            )


def _populate_daily_label_menu_sheet(ws, target_date: dt_date) -> None:
    ws.append(["製造日", "賞味期限", "時間", "メニュー", "温・冷", "商品名１", "商品名２", "内容詳細", "赤字は触らない", "", ""])
    manufacture_date = target_date - timedelta(days=4)
    for row_idx, (daypart, menu_category, temp, product_name, product_name2, detail) in enumerate(
        DAILY_LABEL_MENU_ROWS,
        start=2,
    ):
        ws.cell(row=row_idx, column=1, value=manufacture_date if row_idx == 2 else "")
        ws.cell(row=row_idx, column=2, value=target_date if row_idx == 2 else "")
        ws.cell(row=row_idx, column=3, value=daypart)
        ws.cell(row=row_idx, column=4, value=menu_category)
        ws.cell(row=row_idx, column=5, value=temp)
        ws.cell(row=row_idx, column=6, value=product_name)
        ws.cell(row=row_idx, column=7, value=product_name2)
        ws.cell(row=row_idx, column=8, value=detail)
    _apply_daily_label_sheet_shape(ws, DAILY_LABEL_SHEET_MAX_ROWS.get(ws.title))


def _populate_label_sheet(ws, fieldnames: list[str], rows: list[dict]) -> None:
    ws.append(fieldnames)
    for row in rows:
        ws.append([row.get(field, "") for field in fieldnames])
    _apply_daily_label_sheet_shape(ws, DAILY_LABEL_SHEET_MAX_ROWS.get(ws.title))


def _daily_label_sheet_name(facility_code: str | None, facility_name: str | None) -> str:
    code = str(facility_code or "").strip()
    if code in DAILY_LABEL_SHEET_BY_FACILITY_ID:
        return DAILY_LABEL_SHEET_BY_FACILITY_ID[code]
    name = str(facility_name or code or "").strip()
    return name or "ラベル"


def _daily_label_group_sort_key(group: dict) -> tuple[int, str]:
    sheet_name = _daily_label_sheet_name(group.get("facility_code"), group.get("facility_name"))
    try:
        return DAILY_LABEL_SHEET_ORDER.index(sheet_name), sheet_name
    except ValueError:
        return len(DAILY_LABEL_SHEET_ORDER), sheet_name


def _line_to_label_bag(order_id: str, facility_code: str | None, line: dict) -> dict:
    qty = line.get("quantity_corrected")
    if qty is None:
        qty = line.get("quantity_original")
    return {
        "order_id": order_id,
        "facility": facility_code,
        "date": _ensure_date(line.get("date")),
        "daypart": line.get("daypart"),
        "menu_name": line.get("menu_name"),
        "menu_category": line.get("menu_category"),
        "diet_type": line.get("diet_type"),
        "area_id": line.get("area_id"),
        "bag_type": line.get("bag_type"),
        "menu_unit_type": line.get("menu_unit_type"),
        "menu_qty_per_serving": line.get("menu_qty_per_serving"),
        "menu_temp_type": line.get("menu_temp_type"),
        "quantity": qty if qty is not None else 0,
    }


def _append_zero_quantity_label_bags(group: dict, ctx: dict, target_date: dt_date) -> None:
    existing_keys = {
        (
            _ensure_date(bag.get("date")),
            bag.get("daypart"),
            bag.get("menu_name"),
            bag.get("diet_type"),
            bag.get("area_id"),
        )
        for bag in group.get("bags", [])
    }
    order_id = str(ctx.get("order_for_outputs", {}).get("id") or ctx.get("order", {}).get("id") or "")
    facility_code = str(ctx.get("order_for_outputs", {}).get("facility") or "").strip()
    for line in ctx.get("order_lines", []):
        line_date = _ensure_date(line.get("date"))
        if line_date != target_date:
            continue
        qty = line.get("quantity_corrected")
        if qty is None:
            qty = line.get("quantity_original")
        try:
            qty_value = float(qty or 0)
        except Exception:
            qty_value = 0.0
        if qty_value != 0:
            continue
        key = (
            line_date,
            line.get("daypart"),
            line.get("menu_name"),
            line.get("diet_type"),
            line.get("area_id"),
        )
        if key in existing_keys:
            continue
        group.setdefault("bags", []).append(_line_to_label_bag(order_id, facility_code, line))
        existing_keys.add(key)


def _create_daily_labels_sheet(
    workbook,
    used_titles: set[str],
    *,
    title_seed: str,
    bags: list[dict],
    label_profile: dict,
    facility_name: str | None,
) -> str:
    labels, label_fields, _ = _build_label_rows(bags, label_profile, facility_name)
    if not labels:
        raise ValueError("label rows not found for target date")
    ws = workbook.create_sheet(title=_safe_sheet_title(title_seed, "ラベル", used_titles))
    _populate_label_sheet(ws, label_fields, labels)
    return ws.title


def _safe_output_filename(value: str, fallback: str) -> str:
    text = str(value or "").strip() or fallback
    text = re.sub(r'[\\/:*?"<>|]+', "_", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text or fallback


def _create_daily_label_workbook_bytes(
    *,
    target_date: dt_date,
    group: dict,
) -> tuple[bytes, str]:
    workbook = Workbook()
    workbook.remove(workbook.active)
    used_titles: set[str] = set()
    sheet_name = _create_daily_labels_sheet(
        workbook,
        used_titles,
        title_seed=_daily_label_sheet_name(group.get("facility_code"), group.get("facility_name")),
        bags=group["bags"],
        label_profile=group["label_profile"],
        facility_name=group["facility_config"].get("facility_name"),
    )
    buffer = BytesIO()
    workbook.save(buffer)
    facility_name = str(group.get("facility_name") or group.get("facility_code") or sheet_name or "facility").strip()
    filename = _safe_output_filename(f"{target_date.isoformat()}_{facility_name}_labels.xlsx", "daily_labels.xlsx")
    return buffer.getvalue(), filename


def _create_daily_delivery_sheet(
    workbook,
    used_titles: set[str],
    *,
    title_seed: str,
    rows: list[dict],
    invoice_template: dict,
    facility_name: str | None,
    order_id: str,
) -> str:
    if not rows:
        raise ValueError("delivery rows not found for target date")
    if invoice_template.get("template_uri"):
        raise ValueError("templated delivery notes cannot be embedded into a rebuilt daily bundle workbook")
    temp_path = OUTPUT_DIR / f"{order_id}_daily_bundle_{uuid4().hex}.xlsx"
    try:
        _write_delivery_note(
            temp_path,
            rows,
            invoice_template.get("columns", []),
            invoice_template.get("template_uri"),
            bool(invoice_template.get("include_menu_name", False)),
            invoice_template.get("sheet_name"),
            facility_name,
        )
        rendered = load_workbook(temp_path)
        source_ws = rendered.active
        ws = workbook.create_sheet(title=_safe_sheet_title(title_seed, "納品書", used_titles))
        _copy_sheet_contents(source_ws, ws)
        return ws.title
    finally:
        try:
            temp_path.unlink(missing_ok=True)
        except Exception:
            pass


_WEEKLY_WEIGHT_DAYPART_SLOTS = [
    ("朝", ["副①", "副②"]),
    ("昼", ["主Ａ", "副①", "副②"]),
    ("夕", ["主", "副①", "副②"]),
]
_WEEKLY_WEIGHT_REGULAR_DIETS = {"regular", "regular_bag", "staff", "daycare", "1600kcal"}
_WEEKLY_WEIGHT_SOFT_MIXER_DIETS = {"soft", "mixer", "soft_mixer"}


def _weekly_weight_start(target_date: dt_date) -> dt_date:
    return target_date - timedelta(days=(target_date.weekday() + 1) % 7)


def _weekly_weight_sheet_title(week_start: dt_date) -> str:
    week_end = week_start + timedelta(days=6)
    return f"{week_start.month}月{week_start.day}日～{week_end.month}月{week_end.day}日"


def _weekly_weight_reference_layout_path() -> Path | None:
    relative_path = Path("input_example") / "2026.0512" / "May 10-16 2026 Weight.xlsx"
    for parent in Path(__file__).resolve().parents:
        candidate = parent / relative_path
        if candidate.exists():
            return candidate
    return None


def _normalize_weekly_weight_slot(value: Any) -> str:
    text = str(value or "").strip()
    compact = re.sub(r"\s+", "", text)
    if not compact:
        return ""
    if compact in {"副1", "副①", "副一"} or "副①" in compact or "副1" in compact:
        return "副①"
    if compact in {"副2", "副②", "副二"} or "副②" in compact or "副2" in compact:
        return "副②"
    if "主" in compact and ("A" in compact.upper() or "Ａ" in compact):
        return "主Ａ"
    if "主" in compact:
        return "主"
    return text


def _weekly_weight_slot_for_line(line: dict) -> str:
    for key in ("menu_category", "slot_label", "category", "menu_slot", "_menu_slot_label"):
        slot = _normalize_weekly_weight_slot(line.get(key))
        if slot:
            return slot
    return ""


def _weekly_weight_format_amount(amounts: dict[str, float]) -> Any:
    literal = amounts.get("__literal__") if isinstance(amounts, dict) else None
    if literal not in (None, ""):
        return literal
    if not amounts:
        return None
    main_unit = amounts.get("__main_unit__")
    main_count = amounts.get("__main_count__")
    garnish_amount = amounts.get("__garnish_amount__")
    if main_unit not in (None, "") and main_count not in (None, ""):
        main = f"{_format_number(float(main_count))}{main_unit}"
        try:
            garnish_value = float(garnish_amount)
        except (TypeError, ValueError):
            garnish_value = 0.0
        if math.isfinite(garnish_value) and garnish_value > 0:
            garnish_unit = str(amounts.get("__garnish_unit__") or "g")
            display_value = round(garnish_value / 1000, 1) if garnish_unit == "g" else garnish_value
            garnish_label = str(amounts.get("__garnish_label__") or "")
            separator = str(amounts.get("__garnish_separator__") or "、")
            return f"{main}{separator}{garnish_label}{_format_number(display_value)}"
        return main
    parts: list[str] = []
    for unit, raw_value in sorted(amounts.items(), key=lambda item: item[0]):
        if str(unit).startswith("__"):
            continue
        try:
            value = float(raw_value)
        except (TypeError, ValueError):
            continue
        if not math.isfinite(value):
            continue
        if unit == "g":
            parts.append(round(value / 1000, 1))
        else:
            parts.append(f"{_format_number(value)}{unit}")
    if not parts:
        return None
    if len(parts) == 1:
        return parts[0]
    return "、".join(str(part) for part in parts)


def _weekly_weight_rule_garnish_label(menu_name: str, diet: str) -> tuple[str, str]:
    separator = "、" if diet in {"soft", "mixer", "soft_mixer"} else "＋"
    if "煮込みハンバーグ" in menu_name:
        return ("", "、") if diet in {"soft", "mixer", "soft_mixer"} else ("ソース", "＋")
    if "アジのちゃんちゃん焼き" in menu_name:
        return "野菜", separator
    if any(name in menu_name for name in ("鶏唐揚げ", "サバの塩焼き", "チキンカツ")):
        return "添", "、"
    return "", separator


def _weekly_weight_add_amount(target: dict[str, float], line: dict, quantity: float) -> None:
    literal = line.get("actual_amount_label") or line.get("weekly_weight_amount_label")
    if literal not in (None, ""):
        target["__literal__"] = str(literal)
        return
    rule_amounts = order_service._daily_bag_rule_amounts(line, quantity)  # noqa: SLF001
    if rule_amounts:
        main_items = [(unit, value) for unit, value in rule_amounts.items() if unit != "g"]
        if main_items:
            main_unit, main_count = main_items[0]
            target["__main_unit__"] = str(main_unit)
            target["__main_count__"] = round(float(target.get("__main_count__", 0.0)) + float(main_count), 4)
            if "g" in rule_amounts:
                target["__garnish_unit__"] = "g"
                target["__garnish_amount__"] = round(float(target.get("__garnish_amount__", 0.0)) + float(rule_amounts["g"]), 4)
                garnish_label, separator = _weekly_weight_rule_garnish_label(
                    str(line.get("menu_name") or ""),
                    bucket_diet_type_for_aggregation(line.get("diet_type")) or "",
                )
                target["__garnish_label__"] = garnish_label
                target["__garnish_separator__"] = separator
            return
        for unit, amount in rule_amounts.items():
            target[unit] = round(target.get(unit, 0.0) + float(amount), 4)
        return
    unit = _normalize_unit_type(line.get("menu_unit_type") or line.get("actual_unit_type"))
    if not unit:
        return
    amount = None
    try:
        per_serving = line.get("menu_qty_per_serving")
        if per_serving is not None:
            per_value = float(per_serving)
            if math.isfinite(per_value) and per_value >= 0:
                amount = per_value * quantity
    except (TypeError, ValueError):
        amount = None
    if amount is None:
        try:
            actual_amount = float(line.get("actual_amount"))
            if math.isfinite(actual_amount) and actual_amount >= 0:
                amount = actual_amount
        except (TypeError, ValueError):
            amount = None
    if amount is not None:
        target[unit] = round(target.get(unit, 0.0) + float(amount), 4)


_WEEKLY_WEIGHT_MERGED_RANGES = [
    "A2:D2",
    "A3:E3",
    "A4:E5",
    "A7:A10",
    "B7:C10",
    "D7:D10",
    "E7:E10",
    "F7:F10",
    "G7:G10",
    "H7:H10",
    "I7:I10",
    "A11:A15",
    "B11:B12",
    "B13:B15",
    "A16:A18",
    "B16:B18",
    "A19:A23",
    "B19:B20",
    "B21:B23",
    "A24:A26",
    "B24:B26",
    "A27:A31",
    "B27:B28",
    "B29:B31",
    "A32:A34",
    "B32:B34",
    "A35:A39",
    "B35:B36",
    "B37:B39",
    "A40:A42",
    "B40:B42",
    "A43:A47",
    "B43:B44",
    "B45:B47",
    "A48:A50",
    "B48:B50",
    "A51:A55",
    "B51:B52",
    "B53:B55",
    "A56:A58",
    "B56:B58",
    "A59:A63",
    "B59:B60",
    "B61:B63",
    "A64:A66",
    "B64:B66",
    "A67:C67",
    "E67:H67",
]


def _build_weekly_weight_workbook_shell(week_start: dt_date) -> Workbook:
    from openpyxl.cell.cell import MergedCell
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    reference_layout = _weekly_weight_reference_layout_path()
    if reference_layout is not None:
        workbook = load_workbook(reference_layout)
        ws = workbook.worksheets[0]
        ws.title = _weekly_weight_sheet_title(week_start)
        setattr(workbook, "_weekly_weight_reference_layout", True)
        for row_idx in range(11, 67):
            for col_idx in range(3, 10):
                cell = ws.cell(row=row_idx, column=col_idx)
                if not isinstance(cell, MergedCell):
                    cell.value = None
        return workbook

    workbook = Workbook()
    ws = workbook.active
    ws.title = _weekly_weight_sheet_title(week_start)
    for merged_range in _WEEKLY_WEIGHT_MERGED_RANGES:
        ws.merge_cells(merged_range)

    widths = {"A": 8.5, "B": 3.625, "C": 7.125, "D": 54.5, "E": 15.625, "F": 29.0, "G": 15.625, "H": 25.25, "I": 27.25, "J": 9.0}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    row_heights = {1: 20.65, 2: 35.85, 3: 35.85, 4: 35.85, 5: 35.85, 6: 35.85, 7: 18.75, 8: 15.6, 9: 18.75, 10: 8.65, 67: 34.5, 86: 18.0}
    for row_idx in range(11, 67):
        row_heights[row_idx] = 29.1
    for row_idx, height in row_heights.items():
        ws.row_dimensions[row_idx].height = height
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = 9
    ws.page_setup.scale = 40
    ws.page_margins.left = 0.2362204724409449
    ws.page_margins.right = 0.2362204724409449
    ws.page_margins.top = 0.35433070866141736
    ws.page_margins.bottom = 0.1968503937007874
    ws.page_margins.header = 0.31496062992125984
    ws.page_margins.footer = 0.31496062992125984
    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = True
    ws.print_area = "A1:I66"

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="D9EAD3")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    for row_idx in range(7, 68):
        for col_idx in range(1, 10):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = border
            cell.alignment = left if col_idx in {4, 9} else center
    ws["A4"] = "各メニューの重量"
    ws["A4"].font = Font(size=14, bold=True)
    ws["A4"].alignment = center
    for col_idx, value in ((1, "日　付"), (2, "区　分"), (4, "献立"), (5, "常食"), (6, "重量(㎏）"), (7, "軟菜＋ミキサー"), (8, "重量(㎏）"), (9, "軟菜別メニュー")):
        cell = ws.cell(row=7, column=col_idx)
        cell.value = value
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = center
    ws.cell(row=86, column=9).value = None
    ws.cell(row=86, column=9).border = Border()
    return workbook


def _xml_escape(value: str) -> str:
    return (
        str(value)
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
        .replace("'", "&apos;")
    )


def _patch_weekly_weight_package(path: Path, *, sheet_title: str) -> None:
    content_types = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types"><Default Extension="bin" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.printerSettings"/><Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/><Default Extension="xml" ContentType="application/xml"/><Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/><Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/><Override PartName="/xl/theme/theme1.xml" ContentType="application/vnd.openxmlformats-officedocument.theme+xml"/><Override PartName="/xl/styles.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.styles+xml"/><Override PartName="/xl/sharedStrings.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sharedStrings+xml"/><Override PartName="/xl/calcChain.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.calcChain+xml"/><Override PartName="/docProps/core.xml" ContentType="application/vnd.openxmlformats-package.core-properties+xml"/><Override PartName="/docProps/app.xml" ContentType="application/vnd.openxmlformats-officedocument.extended-properties+xml"/><Override PartName="/xl/persons/person.xml" ContentType="application/vnd.ms-excel.person+xml"/></Types>"""
    workbook_xml = f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"><fileVersion appName="xl" lastEdited="5" lowestEdited="7" rupBuild="9303"/><workbookPr/><bookViews><workbookView xWindow="2025" yWindow="-150" windowWidth="9630" windowHeight="11925" tabRatio="702"/></bookViews><sheets><sheet name="{_xml_escape(sheet_title)}" sheetId="8" r:id="rId1"/></sheets><definedNames><definedName name="aaa">#REF!</definedName><definedName name="ColumnTitle1">#REF!</definedName><definedName name="LastDay_Week">MAX(#REF!)</definedName><definedName name="LastDayOfMonth_Week">DAY(EOMONTH(DATE(#REF!,WkMonthNum,1),0))</definedName><definedName name="MoMonth">#REF!</definedName><definedName name="MoMonthNum">MONTH(DATEVALUE(MoMonth&amp;&quot;/1&quot;))</definedName><definedName name="MoWeek2">#REF!</definedName><definedName name="MoWeek3">#REF!</definedName><definedName name="MoWeek4">#REF!</definedName><definedName name="MoWeek5">#REF!</definedName><definedName name="MoYear">#REF!</definedName><definedName name="_xlnm.Print_Area" localSheetId="0">'{_xml_escape(sheet_title)}'!$A$1:$I$66</definedName><definedName name="WkMonth">#REF!</definedName><definedName name="WkMonthNum">MONTH(DATEVALUE(#REF!&amp;&quot;/1&quot;))</definedName><definedName name="WkMonthView">#REF!</definedName><definedName name="WkWeek">LEFT(RIGHT(#REF!,3),1)</definedName></definedNames><calcPr calcId="145621"/></workbook>"""
    workbook_rels = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships"><Relationship Id="rId3" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles" Target="styles.xml"/><Relationship Id="rId2" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/theme" Target="theme/theme1.xml"/><Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/><Relationship Id="rId11" Type="http://schemas.microsoft.com/office/2017/10/relationships/person" Target="persons/person.xml"/><Relationship Id="rId5" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/calcChain" Target="calcChain.xml"/><Relationship Id="rId4" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/sharedStrings" Target="sharedStrings.xml"/></Relationships>"""
    sheet_rels = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships"><Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/printerSettings" Target="../printerSettings/printerSettings1.bin"/></Relationships>"""
    app_xml = f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Properties xmlns="http://schemas.openxmlformats.org/officeDocument/2006/extended-properties" xmlns:vt="http://schemas.openxmlformats.org/officeDocument/2006/docPropsVTypes"><Application>Microsoft Excel</Application><DocSecurity>0</DocSecurity><ScaleCrop>false</ScaleCrop><HeadingPairs><vt:vector size="4" baseType="variant"><vt:variant><vt:lpstr>ワークシート</vt:lpstr></vt:variant><vt:variant><vt:i4>1</vt:i4></vt:variant><vt:variant><vt:lpstr>名前付き一覧</vt:lpstr></vt:variant><vt:variant><vt:i4>1</vt:i4></vt:variant></vt:vector></HeadingPairs><TitlesOfParts><vt:vector size="2" baseType="lpstr"><vt:lpstr>{_xml_escape(sheet_title)}</vt:lpstr><vt:lpstr>'{_xml_escape(sheet_title)}'!Print_Area</vt:lpstr></vt:vector></TitlesOfParts><LinksUpToDate>false</LinksUpToDate><SharedDoc>false</SharedDoc><HyperlinksChanged>false</HyperlinksChanged><AppVersion>14.0300</AppVersion></Properties>"""
    calc_chain = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<calcChain xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"><c r="A19" i="8"/><c r="A27" i="8"/><c r="A35" i="8"/><c r="A43" i="8"/><c r="A51" i="8"/><c r="A59" i="8"/></calcChain>"""
    shared_strings = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" count="0" uniqueCount="0"/>"""
    person = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<personList xmlns="http://schemas.microsoft.com/office/spreadsheetml/2017/richdata"><person displayName="PCUSER" id="{00000000-0000-0000-0000-000000000000}" userId="PCUSER"/></personList>"""
    printer_settings = bytes(4664)

    replacements = {
        "[Content_Types].xml": content_types.encode("utf-8"),
        "xl/workbook.xml": workbook_xml.encode("utf-8"),
        "xl/_rels/workbook.xml.rels": workbook_rels.encode("utf-8"),
        "xl/worksheets/_rels/sheet1.xml.rels": sheet_rels.encode("utf-8"),
        "docProps/app.xml": app_xml.encode("utf-8"),
        "xl/calcChain.xml": calc_chain.encode("utf-8"),
        "xl/sharedStrings.xml": shared_strings.encode("utf-8"),
        "xl/persons/person.xml": person.encode("utf-8"),
        "xl/printerSettings/printerSettings1.bin": printer_settings,
    }
    temp_path = path.with_suffix(".patched.xlsx")
    with zipfile.ZipFile(path, "r") as source, zipfile.ZipFile(temp_path, "w", compression=zipfile.ZIP_DEFLATED) as target:
        written: set[str] = set()
        ordered_names = [
            "[Content_Types].xml",
            "_rels/.rels",
            "docProps/app.xml",
            "docProps/core.xml",
            "xl/_rels/workbook.xml.rels",
            "xl/calcChain.xml",
            "xl/persons/person.xml",
            "xl/printerSettings/printerSettings1.bin",
            "xl/sharedStrings.xml",
            "xl/styles.xml",
            "xl/theme/theme1.xml",
            "xl/workbook.xml",
            "xl/worksheets/_rels/sheet1.xml.rels",
            "xl/worksheets/sheet1.xml",
        ]
        for name in ordered_names:
            data = replacements.get(name)
            if data is None:
                data = source.read(name)
            target.writestr(name, data)
            written.add(name)
        for name in source.namelist():
            if name not in written:
                target.writestr(name, source.read(name))
    temp_path.replace(path)


def _weekly_weight_collect_rows(target_date: dt_date, *, status: str | None = None) -> dict[tuple[dt_date, str, str], dict]:
    week_start = _weekly_weight_start(target_date)
    rows: dict[tuple[dt_date, str, str], dict] = {}
    for offset in range(7):
        current_date = week_start + timedelta(days=offset)
        for order_summary in order_service.list_orders_by_line_date(current_date, status=status):
            order_id = str(order_summary.get("id") or "").strip()
            if not order_id:
                continue
            ctx = _prepare_output_context_for_bundle(
                order_id,
                include_bags=False,
                include_ocr_menu_meta=False,
                include_expanded_copy=True,
            )
            for line in ctx.get("order_lines") or []:
                if _ensure_date(line.get("date")) != current_date:
                    continue
                quantity = _safe_qty(line, True)
                try:
                    quantity_value = float(quantity)
                except (TypeError, ValueError):
                    continue
                if not math.isfinite(quantity_value) or quantity_value <= 0:
                    continue
                daypart = _normalize_output_daypart(line.get("daypart"))
                slot = _weekly_weight_slot_for_line(line)
                if not daypart or not slot:
                    continue
                row = rows.setdefault(
                    (current_date, daypart, slot),
                    {
                        "regular_menu": "",
                        "regular_quantity": 0.0,
                        "regular_amounts": {},
                        "soft_mixer_menu": "",
                        "soft_mixer_quantity": 0.0,
                        "soft_mixer_amounts": {},
                    },
                )
                diet = _normalize_diet_key(line.get("diet_type")) or ""
                menu_name = str(line.get("menu_name") or "").strip()
                if diet in _WEEKLY_WEIGHT_SOFT_MIXER_DIETS:
                    row["soft_mixer_menu"] = row["soft_mixer_menu"] or menu_name
                    row["soft_mixer_quantity"] = round(row["soft_mixer_quantity"] + quantity_value, 4)
                    _weekly_weight_add_amount(row["soft_mixer_amounts"], line, quantity_value)
                elif diet in _WEEKLY_WEIGHT_REGULAR_DIETS or not diet:
                    row["regular_menu"] = row["regular_menu"] or menu_name
                    row["regular_quantity"] = round(row["regular_quantity"] + quantity_value, 4)
                    _weekly_weight_add_amount(row["regular_amounts"], line, quantity_value)
    return rows


def build_weekly_weight_summary_workbook(target_date: dt_date, *, status: str | None = None) -> Path:
    rows_by_key = _weekly_weight_collect_rows(target_date, status=status)
    week_start = _weekly_weight_start(target_date)
    workbook = _build_weekly_weight_workbook_shell(week_start)
    ws = workbook.worksheets[0]
    uses_reference_layout = bool(getattr(workbook, "_weekly_weight_reference_layout", False))
    row_idx = 11
    weekdays = ["(月)", "(火)", "（水）", "(木)", "(金)", "(土)", "（日）"]
    for offset in range(7):
        current_date = week_start + timedelta(days=offset)
        day_start = row_idx
        for daypart, slots in _WEEKLY_WEIGHT_DAYPART_SLOTS:
            part_start = row_idx
            for slot in slots:
                payload = rows_by_key.get((current_date, daypart, slot), {})
                ws.cell(row=row_idx, column=3).value = slot
                ws.cell(row=row_idx, column=4).value = payload.get("regular_menu") or payload.get("soft_mixer_menu") or None
                ws.cell(row=row_idx, column=5).value = payload.get("regular_quantity") if payload else None
                ws.cell(row=row_idx, column=6).value = _weekly_weight_format_amount(payload.get("regular_amounts") or {})
                ws.cell(row=row_idx, column=7).value = payload.get("soft_mixer_quantity") if payload else None
                ws.cell(row=row_idx, column=8).value = _weekly_weight_format_amount(payload.get("soft_mixer_amounts") or {})
                soft_menu = payload.get("soft_mixer_menu") or ""
                regular_menu = payload.get("regular_menu") or ""
                ws.cell(row=row_idx, column=9).value = soft_menu if soft_menu and soft_menu != regular_menu else None
                row_idx += 1
            if not uses_reference_layout:
                ws.cell(row=part_start, column=2).value = daypart
        if offset == 0 or not uses_reference_layout:
            ws.cell(row=day_start, column=1).value = datetime(current_date.year, current_date.month, current_date.day)
        if not uses_reference_layout:
            ws.cell(row=day_start + 5, column=1).value = weekdays[current_date.weekday()]
    output_path = OUTPUT_DIR / f"{_weekly_weight_sheet_title(week_start)} Weight.xlsx"
    workbook.save(output_path)
    try:
        _patch_weekly_weight_package(output_path, sheet_title=_weekly_weight_sheet_title(week_start))
    except Exception:
        if rows_by_key:
            raise
        logger.warning("weekly weight package patch failed for empty workbook; returning saved workbook")
    return output_path


def _merge_delivery_bundle_rows(rows: list[dict], invoice_template: dict | None) -> list[dict]:
    columns = invoice_template.get("columns", []) if isinstance(invoice_template, dict) else []
    quantity_names = [
        str(col.get("name") or "").strip()
        for col in columns
        if col.get("source") == "quantity" and str(col.get("name") or "").strip()
    ]
    if not rows or not quantity_names:
        return rows

    merged: dict[tuple[object, object, object, object], dict] = {}
    for row in rows:
        key = (
            _ensure_date(row.get("date")),
            row.get("daypart"),
            row.get("menu_category"),
            row.get("menu_name"),
        )
        existing = merged.get(key)
        if existing is None:
            merged[key] = dict(row)
            continue
        target = existing
        if target.get("_order_index") is None and row.get("_order_index") is not None:
            target["_order_index"] = row.get("_order_index")
        if not target.get("menu_display") and row.get("menu_display"):
            target["menu_display"] = row.get("menu_display")
        if not target.get("note") and row.get("note"):
            target["note"] = row.get("note")
        for name in quantity_names:
            value = row.get(name)
            if value in (None, ""):
                continue
            target[name] = (target.get(name) or 0) + float(value)

    result = list(merged.values())
    result.sort(
        key=lambda row: (
            row.get("date") or "",
            row.get("_order_index") if row.get("_order_index") is not None else 1_000_000,
            row.get("menu_name") or "",
        )
    )
    return result


def _invoice_quantity_signature(column: dict, area_aliases: dict[str, str]) -> tuple[str | None, str | None] | None:
    if not isinstance(column, dict) or column.get("source") != "quantity":
        return None
    diet_type = column.get("diet_type")
    area_id = column.get("area_id")
    if not diet_type or not area_id:
        inferred_diet, inferred_area = _infer_delivery_column_meta(column.get("name") or column.get("header"))
        diet_type = diet_type or inferred_diet
        area_id = area_id or inferred_area
    diet_key = _normalize_delivery_diet_key(diet_type)
    area_key = _resolve_area_key(area_id, area_aliases) if area_id else None
    if not diet_key:
        return None
    return diet_key, area_key


def _fax_quantity_signature(column: dict, area_aliases: dict[str, str]) -> tuple[str | None, str | None] | None:
    if not isinstance(column, dict) or str(column.get("role") or "").strip().lower() != "quantity":
        return None
    if column.get("delivery_enabled") is False:
        return None
    diet_key = _normalize_delivery_diet_key(column.get("diet_type"))
    if diet_key == "placeholder" and str(column.get("header") or "").strip() in {"-", "－"}:
        diet_key = "regular_bag"
    area_key = _resolve_area_key(column.get("area_id"), area_aliases) if column.get("area_id") else None
    if not diet_key:
        return None
    return diet_key, area_key


def _delivery_invoice_header(column: dict, signature: tuple[str | None, str | None]) -> str:
    delivery_header = str(column.get("delivery_header") or "").strip()
    if delivery_header:
        return delivery_header
    raw_header = str(column.get("header") or "").strip()
    diet_key, area_key = signature
    diet_labels = {
        "regular": "常食",
        "regular_bag": "小口",
        "soft": "軟菜",
        "mixer": "ミキサー",
        "daycare": "通所",
        "staff": "職員",
        "diabetes": "糖尿",
        "no_meat": "肉禁",
        "no_fish": "魚禁",
        "no_fried": "揚げ物禁",
        "forbidden_other": "その他禁食",
        "sesame_allergy": "ゴマアレルギー",
    }
    label = diet_labels.get(str(diet_key or ""), str(column.get("header") or column.get("name") or diet_key or "").strip())
    header_group = str(column.get("header_group") or "").strip()
    if header_group and header_group != label:
        label = f"{header_group} {label}"
    if area_key and area_key != "X":
        return f"{label}{area_key}"
    if raw_header and raw_header not in {"-", "－"} and not header_group:
        return raw_header
    return label


def _build_invoice_template_from_fax_columns(facility_config: dict | None) -> dict:
    if not isinstance(facility_config, dict):
        return {}
    fax_template = facility_config.get("fax_template") or {}
    fax_columns = (fax_template.get("columns") or [])
    if not isinstance(fax_columns, list):
        return {}
    extra_columns = fax_template.get("delivery_extra_columns") or []
    if not isinstance(extra_columns, list):
        extra_columns = []
    area_aliases = _build_area_alias_map(facility_config)
    fax_quantity_columns = [
        column
        for column in list(fax_columns) + list(extra_columns)
        if _fax_quantity_signature(column, area_aliases)
    ]
    fax_signatures = {
        signature
        for column in fax_quantity_columns
        if (signature := _fax_quantity_signature(column, area_aliases))
    }
    if not fax_signatures:
        return {}

    derived_columns: list[dict] = [
        {"name": "日付", "header": "日付", "source": "date"},
        {"name": "区分", "header": "区分", "source": "daypart"},
        {"name": "献立区分", "header": "献立区分", "source": "menu_category"},
        {"name": "メニュー名", "header": "メニュー名", "source": "menu_name"},
    ]
    seen: set[tuple[str | None, str | None]] = set()
    for column in fax_quantity_columns:
        signature = _fax_quantity_signature(column, area_aliases)
        if not signature or signature in seen:
            continue
        seen.add(signature)
        diet_key, area_key = signature
        if diet_key in {"change_1", "change_2", "unknown"}:
            continue
        name = f"qty.{diet_key}_{str(area_key or 'x').lower()}"
        derived_columns.append(
            {
                "name": name,
                "header": _delivery_invoice_header(column, signature),
                "source": "quantity",
                "diet_type": diet_key,
                "area_id": area_key or "X",
            }
        )
    for column in list(fax_columns) + list(extra_columns):
        if not isinstance(column, dict) or column.get("delivery_enabled") is False:
            continue
        delivery_source = str(column.get("delivery_source") or "").strip()
        delivery_header = str(column.get("delivery_header") or "").strip()
        if delivery_source not in {"note", "static"}:
            continue
        header = delivery_header or str(column.get("header") or column.get("name") or "").strip()
        if not header:
            continue
        derived_columns.append(
            {
                "name": str(column.get("delivery_name") or column.get("name") or header).strip(),
                "header": header,
                "source": delivery_source,
            }
        )
    derived_columns.append({"name": "備考欄", "header": "備考欄", "source": "note"})

    return {
        "columns": derived_columns,
        "include_menu_name": True,
    }


def build_daily_output_bundle(
    target_date: dt_date,
    *,
    bundle_type: str = "both",
    status: str | None = None,
) -> tuple[Path, dict]:
    normalized_type = str(bundle_type or "").strip().lower()
    if normalized_type not in {"labels", "delivery", "both"}:
        raise ValueError("bundle_type must be labels, delivery, or both")

    orders = order_service.list_orders_by_line_date(target_date, status=status)
    stamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    bundle_name = f"daily_outputs_{target_date.isoformat()}_{normalized_type}_{stamp}.xlsx"
    bundle_path = OUTPUT_DIR / bundle_name

    manifest_items: list[dict] = []
    grouped_outputs: dict[str, dict[str, Any]] = {}
    workbook = Workbook()
    workbook.remove(workbook.active)
    used_titles: set[str] = set()

    for order_summary in orders:
        order_id = str(order_summary.get("id") or "").strip()
        if not order_id:
            continue
        facility_code = str(order_summary.get("facility") or "").strip()
        facility_name = ""
        if facility_code:
            try:
                facility_config = config_service.get_facility_config(facility_code) or {}
                facility_name = str(facility_config.get("facility_name") or "").strip()
            except Exception:
                facility_name = ""
        try:
            ctx = _prepare_output_context_for_bundle(
                order_id,
                include_bags=normalized_type in {"labels", "both"},
                include_ocr_menu_meta=normalized_type != "delivery",
                include_expanded_copy=normalized_type != "delivery",
            )
            facility_config = ctx.get("facility_config") or {}
            facility_name = str(facility_config.get("facility_name") or facility_name or "").strip()
            facility_code = str(ctx.get("order_for_outputs", {}).get("facility") or facility_code or "").strip()
            group_key = facility_code or facility_name or order_id
            sheet_seed = facility_name or facility_code or order_id
            group = grouped_outputs.get(group_key)
            if not group:
                group = {
                    "facility_code": facility_code or None,
                    "facility_name": facility_name or None,
                    "sheet_seed": sheet_seed,
                    "order_ids": [],
                    "label_profile": ctx["label_profile"],
                    "invoice_template": ctx["invoice_template"],
                    "quantity_rules": ctx["quantity_rules"],
                    "facility_config": facility_config,
                    "ocr_menu_meta": ctx.get("ocr_menu_meta"),
                    "contexts": [],
                    "bags": [],
                    "delivery_rows": [],
                }
                grouped_outputs[group_key] = group

            group["order_ids"].append(order_id)
            group["contexts"].append(ctx)
            if normalized_type in {"labels", "both"}:
                filtered_bags = [
                    bag for bag in ctx["bags"] if _ensure_date(bag.get("date")) == target_date
                ]
                group["bags"].extend(filtered_bags)
                if normalized_type == "labels":
                    _append_zero_quantity_label_bags(group, ctx, target_date)
            if normalized_type in {"delivery", "both"}:
                delivery_rows = _build_delivery_rows_for_bundle(
                    ctx["order_for_outputs"],
                    ctx["invoice_template"],
                    ctx["quantity_rules"],
                    facility_config,
                    ctx.get("ocr_menu_meta"),
                    allow_ocr_menu_meta=normalized_type != "delivery",
                )
                filtered_delivery_rows = [
                    row for row in delivery_rows if _ensure_date(row.get("date")) == target_date
                ]
                group["delivery_rows"].extend(filtered_delivery_rows)
        except Exception as exc:  # noqa: BLE001
            manifest_items.append(
                {
                    "order_id": order_id,
                    "facility_code": facility_code or None,
                    "facility_name": facility_name or None,
                    "status": "error",
                    "error": str(exc),
                    "files": [],
                }
            )

    use_reference_daily_delivery = False
    if use_reference_daily_delivery:
        workbook = _create_reference_daily_delivery_workbook(
            target_date=target_date,
            grouped_outputs=grouped_outputs,
        )
        if normalized_type == "both":
            used_titles = set(workbook.sheetnames)
            if "メニュー" not in workbook.sheetnames:
                menu_ws = workbook.create_sheet(title="メニュー", index=0)
                _populate_daily_label_menu_sheet(menu_ws, target_date)
                used_titles.add("メニュー")
            for group in sorted(grouped_outputs.values(), key=_daily_label_group_sort_key):
                filtered_bags = group.get("bags") or []
                if not filtered_bags:
                    continue
                sheet_name = _create_daily_labels_sheet(
                    workbook,
                    used_titles,
                    title_seed=f"ラベル_{_daily_label_sheet_name(group.get('facility_code'), group.get('facility_name'))}",
                    bags=filtered_bags,
                    label_profile=group["label_profile"],
                    facility_name=group["facility_config"].get("facility_name"),
                )
                group.setdefault("_reference_bundle_label_sheets", []).append(sheet_name)
        workbook.save(bundle_path)
        manifest_items.extend(
            {
                "order_ids": list(group["order_ids"]),
                "facility_code": group["facility_code"],
                "facility_name": group["facility_name"],
                "status": "ok",
                "files": [
                    item
                    for item in [
                        _reference_delivery_sheet_name(group.get("facility_code"), group.get("facility_name")),
                        *list(group.get("_reference_bundle_label_sheets") or []),
                    ]
                    if item
                ],
            }
            for group in grouped_outputs.values()
        )
        summary = {
            "target_date": target_date.isoformat(),
            "bundle_type": normalized_type,
            "total_orders": len(orders),
            "success_orders": len(grouped_outputs),
            "error_orders": sum(1 for item in manifest_items if item.get("status") == "error"),
            "items": manifest_items,
        }
        return bundle_path, summary

    if normalized_type == "labels" and "メニュー" not in workbook.sheetnames:
        menu_ws = workbook.create_sheet(title="メニュー")
        _populate_daily_label_menu_sheet(menu_ws, target_date)

    if normalized_type == "labels":
        zip_path = OUTPUT_DIR / f"daily_outputs_{target_date.isoformat()}_labels_{stamp}.zip"
        success_count = 0
        with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
            for group in sorted(grouped_outputs.values(), key=_daily_label_group_sort_key):
                item_payload: dict[str, object] = {
                    "order_ids": list(group["order_ids"]),
                    "facility_code": group["facility_code"],
                    "facility_name": group["facility_name"],
                    "status": "ok",
                    "files": [],
                }
                try:
                    if not group.get("bags"):
                        item_payload["status"] = "empty"
                        item_payload["error"] = "label rows not found for target date"
                        manifest_items.append(item_payload)
                        continue
                    workbook_bytes, filename = _create_daily_label_workbook_bytes(
                        target_date=target_date,
                        group=group,
                    )
                    archive.writestr(filename, workbook_bytes)
                    item_payload["files"].append(filename)
                    success_count += 1
                except Exception as exc:  # noqa: BLE001
                    item_payload["status"] = "error"
                    item_payload["error"] = str(exc)
                manifest_items.append(item_payload)
        if success_count == 0:
            raise ValueError("対象日の出力対象がありません")
        manifest = {
            "date": target_date.isoformat(),
            "bundle_type": normalized_type,
            "status_filter": status,
            "created_at": datetime.utcnow().isoformat(),
            "total_orders": len(manifest_items),
            "success_orders": success_count,
            "error_orders": max(len(manifest_items) - success_count, 0),
            "items": manifest_items,
            "file_format": "zip",
        }
        return zip_path, manifest

    success_count = 0
    output_groups = (
        sorted(grouped_outputs.values(), key=_daily_label_group_sort_key)
        if normalized_type == "labels"
        else list(grouped_outputs.values())
    )
    for group in output_groups:
        item_payload: dict[str, object] = {
            "order_ids": list(group["order_ids"]),
            "facility_code": group["facility_code"],
            "facility_name": group["facility_name"],
            "status": "ok",
            "files": [],
        }
        try:
            has_label_rows = bool(group["bags"])
            merged_delivery_rows = _merge_delivery_bundle_rows(
                group["delivery_rows"],
                group["invoice_template"],
            ) if normalized_type in {"delivery", "both"} else []
            has_delivery_rows = bool(merged_delivery_rows)

            should_create_label_sheet = normalized_type in {"labels", "both"} and has_label_rows
            should_create_delivery_sheet = normalized_type in {"delivery", "both"} and has_delivery_rows

            if normalized_type == "labels" and not should_create_label_sheet:
                item_payload["status"] = "empty"
                item_payload["error"] = "label rows not found for target date"
                manifest_items.append(item_payload)
                continue
            if normalized_type == "delivery" and not should_create_delivery_sheet:
                item_payload["status"] = "empty"
                item_payload["error"] = "delivery rows not found for target date"
                manifest_items.append(item_payload)
                continue
            if normalized_type == "both" and not should_create_label_sheet and not should_create_delivery_sheet:
                item_payload["status"] = "empty"
                item_payload["error"] = "no bundle rows found for target date"
                manifest_items.append(item_payload)
                continue

            if should_create_label_sheet:
                label_title_seed = (
                    _daily_label_sheet_name(group.get("facility_code"), group.get("facility_name"))
                    if normalized_type == "labels"
                    else f"ラベル_{group['sheet_seed']}"
                )
                sheet_name = _create_daily_labels_sheet(
                    workbook,
                    used_titles,
                    title_seed=label_title_seed,
                    bags=group["bags"],
                    label_profile=group["label_profile"],
                    facility_name=group["facility_config"].get("facility_name"),
                )
                item_payload["files"].append(sheet_name)
            if should_create_delivery_sheet:
                delivery_title_seed = (
                    group["sheet_seed"]
                    if normalized_type == "delivery"
                    else f"納品書_{group['sheet_seed']}"
                )
                sheet_name = _create_daily_delivery_sheet(
                    workbook,
                    used_titles,
                    title_seed=delivery_title_seed,
                    rows=merged_delivery_rows,
                    invoice_template=group["invoice_template"],
                    facility_name=group["facility_config"].get("facility_name"),
                    order_id=str(group["order_ids"][0]),
                )
                item_payload["files"].append(sheet_name)
            success_count += 1
        except Exception as exc:  # noqa: BLE001
            item_payload["status"] = "error"
            item_payload["error"] = str(exc)
        manifest_items.append(item_payload)

    if normalized_type == "labels" and success_count == 0:
        raise ValueError("対象日の出力対象がありません")
    if not workbook.sheetnames:
        raise ValueError("対象日の出力対象がありません")

    workbook.save(bundle_path)
    manifest = {
        "date": target_date.isoformat(),
        "bundle_type": normalized_type,
        "status_filter": status,
        "created_at": datetime.utcnow().isoformat(),
        "total_orders": len(manifest_items),
        "success_orders": success_count,
        "error_orders": max(len(manifest_items) - success_count, 0),
        "items": manifest_items,
        "file_format": "xlsx",
    }

    return bundle_path, manifest


def _prepare_output_context(
    order_id: str,
    *,
    include_bags: bool = True,
    include_ocr_menu_meta: bool = True,
    include_expanded_copy: bool = True,
    timings: dict[str, float] | None = None,
) -> dict:
    order_started = time.perf_counter()
    order = get_order_by_id(order_id)
    if timings is not None:
        timings["prepare_get_order_ms"] = round((time.perf_counter() - order_started) * 1000, 1)
    if not order:
        raise ValueError("order not found")

    facility_id = order.get("facility")
    facility_started = time.perf_counter()
    facility_config = config_service.get_facility_config(facility_id) if facility_id else None
    if timings is not None:
        timings["prepare_facility_config_ms"] = round((time.perf_counter() - facility_started) * 1000, 1)
    if not facility_config:
        logger.warning("Facility config missing", facility_id=facility_id)
        facility_config = {}

    packaging_policy = facility_config.get("packaging_policy", {})
    label_profile = facility_config.get("label_profile", {})
    invoice_template = _build_invoice_template_from_fax_columns(facility_config)
    policy_started = time.perf_counter()
    quantity_rules = config_service.load_ingest_policy().get("quantity_rules", {})
    if timings is not None:
        timings["prepare_ingest_policy_ms"] = round((time.perf_counter() - policy_started) * 1000, 1)

    order_lines_started = time.perf_counter()
    order_lines = build_order_lines_for_outputs(
        order,
        include_expanded_copy=include_expanded_copy,
        timings=timings,
    )
    if timings is not None:
        timings["prepare_order_lines_ms"] = round((time.perf_counter() - order_lines_started) * 1000, 1)
    order_for_outputs = {**order, "lines": order_lines}
    week_value = (
        str(order.get("stored_week_value") or "").strip()
        or str(order.get("week_value") or "").strip()
        or str(order.get("persisted_week_value") or "").strip()
        or str(order.get("week") or "").strip()
        or str(order.get("week_code") or "").strip()
    )
    ocr_started = time.perf_counter()
    has_structural_slots = _delivery_lines_have_structural_slots(order_lines, quantity_rules)
    if include_ocr_menu_meta and not has_structural_slots:
        ocr_menu_meta = _build_ocr_menu_meta(order, facility_config)
    else:
        ocr_menu_meta = {}
    if timings is not None:
        timings["prepare_ocr_menu_meta_ms"] = round((time.perf_counter() - ocr_started) * 1000, 1)
        timings["prepare_ocr_menu_meta_skipped"] = 1.0 if has_structural_slots else 0.0

    bags = []
    if include_bags:
        bags_started = time.perf_counter()
        bags = _split_bags_by_max(_build_bags(order_for_outputs, packaging_policy, quantity_rules))
        menu_items = _collect_cached_menu_items_for_week(week_value, facility_id)
        bags = _apply_menu_overrides(bags, menu_items)
        bags = _clear_stale_menu_qty_from_monthly_entry(bags)
        bags = _apply_menu_master_defaults(bags, facility_id)
        bags = _apply_builtin_menu_defaults(bags)
        bags = _apply_label_meal_slot_categories(bags)
        bag_types = _resolve_bag_types(facility_config)
        bags = _assign_bag_type_for_bags(bags, bag_types)
        bags = _apply_daily_label_facility_rules_to_bags(bags, facility_config, facility_id)
        if timings is not None:
            timings["prepare_bags_ms"] = round((time.perf_counter() - bags_started) * 1000, 1)
    return {
        "order": order,
        "facility_config": facility_config,
        "label_profile": label_profile,
        "invoice_template": invoice_template,
        "quantity_rules": quantity_rules,
        "order_lines": order_lines,
        "order_for_outputs": order_for_outputs,
        "ocr_menu_meta": ocr_menu_meta,
        "bags": bags,
    }


def _prepare_output_context_for_bundle(
    order_id: str,
    *,
    include_bags: bool,
    include_ocr_menu_meta: bool,
    include_expanded_copy: bool,
) -> dict:
    try:
        return _prepare_output_context(
            order_id,
            include_bags=include_bags,
            include_ocr_menu_meta=include_ocr_menu_meta,
            include_expanded_copy=include_expanded_copy,
            timings=None,
        )
    except TypeError as exc:
        if (
            "include_bags" not in str(exc)
            and "include_ocr_menu_meta" not in str(exc)
            and "include_expanded_copy" not in str(exc)
        ):
            raise
        return _prepare_output_context(order_id)


def build_output_preview(order_id: str, output_type: str) -> Dict[str, Any]:
    ctx = _prepare_output_context(order_id, include_bags=output_type in {"labels", "aggregate"})
    label_profile = ctx["label_profile"]
    invoice_template = ctx["invoice_template"]
    quantity_rules = ctx["quantity_rules"]
    facility_name = ctx["facility_config"].get("facility_name")
    bags = ctx["bags"]

    label_path = OUTPUT_DIR / f"{order_id}_labels.csv"
    delivery_path = OUTPUT_DIR / f"{order_id}_delivery.xlsx"
    agg_path = OUTPUT_DIR / f"{order_id}_aggregate.csv"

    if output_type == "labels":
        labels, label_fields, _ = _build_label_rows(bags, label_profile, facility_name)
        _write_label_csv(label_path, labels, label_fields)
        return {"labels": str(label_path)}
    if output_type == "delivery":
        delivery_rows = _build_delivery_rows(
            ctx["order_for_outputs"],
            invoice_template,
            quantity_rules,
            ctx["facility_config"],
            ctx.get("ocr_menu_meta"),
        )
        include_menu_name = bool(invoice_template.get("include_menu_name", False))
        _write_delivery_note(
            delivery_path,
            delivery_rows,
            invoice_template.get("columns", []),
            invoice_template.get("template_uri"),
            include_menu_name,
            invoice_template.get("sheet_name"),
            facility_name,
        )
        return {"delivery_note": str(delivery_path)}
    if output_type == "aggregate":
        total_rows, total_fields, _ = _build_total_rows(
            ctx["order_lines"], label_profile, facility_name, quantity_rules
        )
        _write_aggregate_csv(agg_path, total_rows, total_fields)
        return {"aggregate": str(agg_path)}
    raise ValueError(f"invalid output type: {output_type}")


def build_delivery_preview(
    order_id: str,
    *,
    include_diagnostics: bool = True,
    target_date: dt_date | None = None,
) -> dict:
    total_started = time.perf_counter()
    timings: dict[str, float] = {}
    context_started = time.perf_counter()
    ctx = _prepare_output_context(
        order_id,
        include_bags=False,
        include_expanded_copy=False,
        timings=timings,
    )
    timings["prepare_context_ms"] = round((time.perf_counter() - context_started) * 1000, 1)
    if target_date is not None:
        filtered_lines = [
            line
            for line in ctx["order_for_outputs"].get("lines", [])
            if _ensure_date(line.get("date")) == target_date
        ]
        ctx["order_for_outputs"] = {**ctx["order_for_outputs"], "lines": filtered_lines}
        timings["target_date_filtered_lines"] = float(len(filtered_lines))
        order = ctx["order"]
        week_value = (
            str(order.get("stored_week_value") or "").strip()
            or str(order.get("week_value") or "").strip()
            or str(order.get("persisted_week_value") or "").strip()
            or str(order.get("week") or "").strip()
            or str(order.get("week_code") or "").strip()
        )
        menu_entries = [
            entry
            for entry in _collect_cached_menu_entries_for_week(week_value, order.get("facility"))
            if _delivery_menu_entry_date(entry) == target_date
        ]
        if menu_entries:
            menu_meta = dict(ctx.get("ocr_menu_meta") or {})
            existing_entries = menu_meta.get("entries") if isinstance(menu_meta.get("entries"), list) else []
            menu_meta["entries"] = [*existing_entries, *menu_entries]
            ctx["ocr_menu_meta"] = menu_meta
        timings["target_date_menu_entries"] = float(len(menu_entries))
    invoice_template = ctx["invoice_template"]
    quantity_rules = ctx["quantity_rules"]
    include_menu_name = bool(invoice_template.get("include_menu_name", False))
    columns = invoice_template.get("columns", [])
    display_headers = []
    for col in columns:
        name = col.get("name")
        if not name:
            continue
        display_headers.append(col.get("header") or name)
    rows_started = time.perf_counter()
    rows = _build_delivery_rows(
        ctx["order_for_outputs"],
        invoice_template,
        quantity_rules,
        ctx["facility_config"],
        ctx.get("ocr_menu_meta"),
        allow_ocr_menu_meta=False,
        timings=timings,
    )
    timings["build_rows_ms"] = round((time.perf_counter() - rows_started) * 1000, 1)
    render_started = time.perf_counter()
    raw_rows = [dict(row) for row in rows]
    preview_rows = []
    for row in rows:
        rendered = []
        for col in columns:
            if not col.get("name"):
                continue
            if not include_menu_name and col.get("source") == "menu_name":
                rendered.append("")
                continue
            rendered.append(_format_delivery_preview_value(_resolve_delivery_cell(row, col)))
        preview_rows.append(rendered)
    timings["render_rows_ms"] = round((time.perf_counter() - render_started) * 1000, 1)
    ocr_entries = ctx.get("ocr_menu_meta", {}).get("entries", []) if isinstance(ctx.get("ocr_menu_meta"), dict) else []
    table_raw_len = None
    if include_diagnostics:
        diagnostics_started = time.perf_counter()
        parsed, _ = order_service.get_ocr_output(order_id)
        table_raw = parsed.get("table_raw") if isinstance(parsed, dict) else None
        table_raw_len = len(table_raw) if isinstance(table_raw, str) else None
        timings["diagnostics_ms"] = round((time.perf_counter() - diagnostics_started) * 1000, 1)
    timings["total_ms"] = round((time.perf_counter() - total_started) * 1000, 1)
    return {
        "headers": display_headers,
        "rows": preview_rows,
        "columns": columns,
        "raw_rows": raw_rows,
        "timings": timings,
        "ocr_entry_count": len(ocr_entries),
        "ocr_table_raw_len": table_raw_len,
    }


def build_outputs(order_id: str) -> Dict[str, Any]:
    ctx = _prepare_output_context(order_id)
    order = ctx["order"]
    label_profile = ctx["label_profile"]
    invoice_template = ctx["invoice_template"]
    quantity_rules = ctx["quantity_rules"]
    order_lines = ctx["order_lines"]
    order_for_outputs = ctx["order_for_outputs"]
    bags = ctx["bags"]
    labels, label_fields, _ = _build_label_rows(
        bags, label_profile, ctx["facility_config"].get("facility_name")
    )

    label_path = OUTPUT_DIR / f"{order_id}_labels.csv"
    delivery_path = OUTPUT_DIR / f"{order_id}_delivery.xlsx"
    agg_path = OUTPUT_DIR / f"{order_id}_aggregate.csv"

    _write_label_csv(label_path, labels, label_fields)

    delivery_rows = _build_delivery_rows(
        order_for_outputs,
        invoice_template,
        quantity_rules,
        ctx["facility_config"],
        ctx.get("ocr_menu_meta"),
    )
    include_menu_name = bool(invoice_template.get("include_menu_name", False))
    _write_delivery_note(
        delivery_path,
        delivery_rows,
        invoice_template.get("columns", []),
        invoice_template.get("template_uri"),
        include_menu_name,
        invoice_template.get("sheet_name"),
        ctx["facility_config"].get("facility_name"),
    )

    total_rows, total_fields, _ = _build_total_rows(
        order_lines, label_profile, ctx["facility_config"].get("facility_name"), quantity_rules
    )
    _write_aggregate_csv(agg_path, total_rows, total_fields)

    with session_scope() as session:
        session.query(Bag).filter(Bag.order_id == order_id).delete()
        session.query(LabelRow).filter(LabelRow.order_id == order_id).delete()
        session.query(DeliveryNote).filter(DeliveryNote.order_id == order_id).delete()

        for bag in bags:
            bag_id = f"BAG{uuid4().hex[:8]}"
            session.add(
                Bag(
                    id=bag_id,
                    order_id=order_id,
                    date=bag.get("date"),
                    daypart=bag.get("daypart"),
                    menu_name=bag.get("menu_name"),
                    diet_type=bag.get("diet_type"),
                    area_id=bag.get("area_id"),
                    bag_type=bag.get("bag_type"),
                    quantity=bag.get("quantity"),
                )
            )
        labels_payload = _serialize_for_json(labels)
        delivery_rows_payload = _serialize_for_json(delivery_rows)
        for label in labels_payload:
            session.add(
                LabelRow(
                    id=f"LAB{uuid4().hex[:8]}",
                    order_id=order_id,
                    bag_id=None,
                    payload_json=label,
                )
            )
        session.add(
            DeliveryNote(
                id=f"INV{uuid4().hex[:8]}",
                order_id=order_id,
                facility_code=order.get("facility") or "",
                date=None,
                file_uri=str(delivery_path),
                payload_json={"rows": delivery_rows_payload},
            )
        )
        for row in bags:
            session.add(
                ManufacturingAggregateRow(
                    id=f"MAG{uuid4().hex[:8]}",
                    week_code=order.get("week") or "",
                    facility_code=order.get("facility") or "",
                    menu_name=row.get("menu_name"),
                    diet_type=row.get("diet_type"),
                    area_id=row.get("area_id"),
                    bag_type=row.get("bag_type"),
                    quantity=row.get("quantity") or 0,
                )
            )

    return {
        "order_id": order_id,
        "labels": str(label_path),
        "delivery_note": str(delivery_path),
        "aggregate": str(agg_path),
    }


def rebuild_bags(order_id: str) -> Dict[str, Any]:
    order = get_order_by_id(order_id)
    if not order:
        raise ValueError("order not found")

    facility_id = order.get("facility")
    facility_config = config_service.get_facility_config(facility_id) if facility_id else None
    if not facility_config:
        logger.warning("Facility config missing", facility_id=facility_id)
        facility_config = {}
    order_lines = build_order_lines_for_outputs(order)
    bags = build_bag_rows_for_outputs(order, order_lines=order_lines, facility_config=facility_config)
    payload = []

    with session_scope() as session:
        session.query(Bag).filter(Bag.order_id == order_id).delete()
        for bag in bags:
            bag_id = f"BAG{uuid4().hex[:8]}"
            session.add(
                Bag(
                    id=bag_id,
                    order_id=order_id,
                    date=bag.get("date"),
                    daypart=bag.get("daypart"),
                    menu_name=bag.get("menu_name"),
                    diet_type=bag.get("diet_type"),
                    area_id=bag.get("area_id"),
                    bag_type=bag.get("bag_type"),
                    quantity=bag.get("quantity"),
                )
            )
            payload.append({"id": bag_id, **_serialize_bag_payload_rows([bag])[0]})
    return {"order_id": order_id, "generated": bool(payload), "bags": payload}
