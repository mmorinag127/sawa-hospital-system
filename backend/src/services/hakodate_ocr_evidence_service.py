from __future__ import annotations

from dataclasses import asdict, dataclass, field
import hashlib
import re
from typing import Any

from openpyxl.utils import get_column_letter


_FULLWIDTH_DIGIT_TRANS = str.maketrans("０１２３４５６７８９，．", "0123456789,.")


@dataclass(frozen=True)
class OcrEvidenceRecord:
    evidence_id: str
    run_id: str
    engine: str
    source_scope: str
    raw_text: str
    normalized_value: str
    confidence: float | None = None
    source_bbox: list[float] | None = None
    center: list[float] | None = None
    candidate_type: str = "unknown"
    raw_payload_ref: str | None = None
    engine_metadata: dict[str, Any] = field(default_factory=dict)


@dataclass(frozen=True)
class TargetCell:
    target_cell_id: str
    sheet_cell: str
    worksheet_row: int
    worksheet_col: int
    semantic_field: str
    bbox: list[float]
    center: list[float]
    metadata: dict[str, Any] = field(default_factory=dict)


@dataclass(frozen=True)
class AssignedOcrResult:
    target_cell_id: str
    sheet_cell: str
    worksheet_row: int
    worksheet_col: int
    semantic_field: str
    evidence_ids: list[str]
    assigned_value: str
    assignment_confidence: float | None
    assignment_state: str
    raw_texts: list[str]


def normalize_ocr_value(value: object) -> str:
    text = str(value or "").strip().translate(_FULLWIDTH_DIGIT_TRANS)
    text = re.sub(r"[^0-9,.]", "", text)
    return text.strip(".,")


def _stable_id(parts: list[object]) -> str:
    payload = "\u001f".join(str(part) for part in parts)
    digest = hashlib.sha1(payload.encode("utf-8")).hexdigest()
    return digest[:16]


def _as_float_box(value: object) -> list[float] | None:
    if not isinstance(value, (list, tuple)) or len(value) != 4:
        return None
    try:
        x0, y0, x1, y1 = [float(item) for item in value]
    except Exception:
        return None
    if x1 < x0 or y1 < y0:
        return None
    return [x0, y0, x1, y1]


def _center_from_box(box: list[float] | None) -> list[float] | None:
    if not box:
        return None
    return [(float(box[0]) + float(box[2])) / 2.0, (float(box[1]) + float(box[3])) / 2.0]


def _as_center(value: object, *, fallback_box: list[float] | None = None) -> list[float] | None:
    if isinstance(value, (list, tuple)) and len(value) == 2:
        try:
            return [float(value[0]), float(value[1])]
        except Exception:
            return None
    return _center_from_box(fallback_box)


def evidence_from_records(
    records: list[dict[str, Any]],
    *,
    run_id: str,
    engine: str,
    source_scope: str,
    candidate_type: str = "numeric",
    raw_payload_ref: str | None = None,
) -> list[dict[str, Any]]:
    evidence: list[dict[str, Any]] = []
    for index, record in enumerate(records):
        raw_text = str(record.get("raw_text") or record.get("text") or "").strip()
        if not raw_text:
            continue
        source_bbox = _as_float_box(record.get("source_bbox") or record.get("bbox"))
        center = _as_center(record.get("center"), fallback_box=source_bbox)
        evidence_id = str(
            record.get("evidence_id")
            or _stable_id([run_id, engine, source_scope, index, raw_text, source_bbox, center])
        )
        confidence = record.get("confidence")
        try:
            confidence_value = float(confidence) if confidence is not None else None
        except Exception:
            confidence_value = None
        evidence.append(
            asdict(
                OcrEvidenceRecord(
                    evidence_id=evidence_id,
                    run_id=run_id,
                    engine=engine,
                    source_scope=source_scope,
                    raw_text=raw_text,
                    normalized_value=str(record.get("normalized_value") or normalize_ocr_value(raw_text)),
                    confidence=confidence_value,
                    source_bbox=source_bbox,
                    center=center,
                    candidate_type=str(record.get("candidate_type") or candidate_type),
                    raw_payload_ref=str(record.get("raw_payload_ref") or raw_payload_ref or ""),
                    engine_metadata=dict(record.get("engine_metadata") or {}),
                )
            )
        )
    return evidence


def target_cells_from_regions(regions: list[dict[str, Any]]) -> list[dict[str, Any]]:
    cells: list[dict[str, Any]] = []
    for index, region in enumerate(regions):
        box = _as_float_box(region.get("bbox"))
        if not box:
            continue
        center = _as_center(region.get("center"), fallback_box=box)
        if not center:
            continue
        target_id = str(region.get("target_cell_id") or region.get("sheet_cell") or region.get("region_id") or "")
        if not target_id:
            target_id = f"target_{index}"
        cells.append(
            asdict(
                TargetCell(
                    target_cell_id=target_id,
                    sheet_cell=str(region.get("sheet_cell") or target_id),
                    worksheet_row=int(region.get("worksheet_row") or 0),
                    worksheet_col=int(region.get("worksheet_col") or 0),
                    semantic_field=str(region.get("semantic_field") or region.get("field") or ""),
                    bbox=box,
                    center=center,
                    metadata={
                        key: value
                        for key, value in region.items()
                        if key
                        not in {
                            "target_cell_id",
                            "sheet_cell",
                            "worksheet_row",
                            "worksheet_col",
                            "semantic_field",
                            "field",
                            "bbox",
                            "center",
                        }
                    },
                )
            )
        )
    return cells


def _contains_point(box: list[float], center: list[float]) -> bool:
    return float(box[0]) <= float(center[0]) <= float(box[2]) and float(box[1]) <= float(center[1]) <= float(box[3])


def _area(box: list[float]) -> float:
    return max(0.0, float(box[2]) - float(box[0])) * max(0.0, float(box[3]) - float(box[1]))


def _mean_confidence(records: list[dict[str, Any]]) -> float | None:
    values = [float(record["confidence"]) for record in records if record.get("confidence") is not None]
    if not values:
        return None
    return round(sum(values) / len(values), 4)


def assign_evidence_to_target_cells(
    *,
    evidence_records: list[dict[str, Any]],
    target_cells: list[dict[str, Any]],
) -> dict[str, Any]:
    targets = target_cells_from_regions(target_cells)
    by_target: dict[str, list[dict[str, Any]]] = {str(target["target_cell_id"]): [] for target in targets}
    unassigned: list[dict[str, Any]] = []
    blockers: list[str] = []

    for evidence in evidence_records:
        center = _as_center(evidence.get("center"), fallback_box=_as_float_box(evidence.get("source_bbox")))
        if not center:
            unassigned.append({**evidence, "unassigned_reason": "missing_center"})
            continue
        matches = [target for target in targets if _contains_point(target["bbox"], center)]
        if not matches:
            unassigned.append({**evidence, "unassigned_reason": "outside_target_cells"})
            continue
        matches.sort(key=lambda target: (_area(target["bbox"]), str(target["target_cell_id"])))
        best = matches[0]
        if len(matches) > 1 and _area(matches[0]["bbox"]) == _area(matches[1]["bbox"]):
            unassigned.append(
                {
                    **evidence,
                    "unassigned_reason": "ambiguous_target_cell",
                    "candidate_target_cell_ids": [str(target["target_cell_id"]) for target in matches],
                }
            )
            blockers.append(f"ambiguous evidence target: {evidence.get('evidence_id')}")
            continue
        by_target[str(best["target_cell_id"])].append(evidence)

    assignments: list[dict[str, Any]] = []
    for target in targets:
        target_id = str(target["target_cell_id"])
        records = by_target.get(target_id) or []
        values = [str(record.get("normalized_value") or "") for record in records if str(record.get("normalized_value") or "")]
        unique_values = sorted(set(values))
        if not records:
            state = "blank"
            assigned_value = ""
        elif len(unique_values) <= 1:
            state = "assigned" if unique_values else "blank_evidence"
            assigned_value = unique_values[0] if unique_values else ""
        else:
            state = "conflict"
            assigned_value = ""
            blockers.append(f"conflicting evidence for target: {target_id}")
        assignments.append(
            asdict(
                AssignedOcrResult(
                    target_cell_id=target_id,
                    sheet_cell=str(target["sheet_cell"]),
                    worksheet_row=int(target["worksheet_row"]),
                    worksheet_col=int(target["worksheet_col"]),
                    semantic_field=str(target["semantic_field"]),
                    evidence_ids=[str(record.get("evidence_id")) for record in records],
                    assigned_value=assigned_value,
                    assignment_confidence=_mean_confidence(records),
                    assignment_state=state,
                    raw_texts=[str(record.get("raw_text") or "") for record in records],
                )
            )
        )

    return {
        "assignments": assignments,
        "unassigned_evidence": unassigned,
        "blockers": sorted(set(blockers)),
        "summary": {
            "target_cell_count": len(targets),
            "evidence_count": len(evidence_records),
            "assigned_target_count": sum(1 for item in assignments if item["assignment_state"] == "assigned"),
            "conflict_target_count": sum(1 for item in assignments if item["assignment_state"] == "conflict"),
            "unassigned_evidence_count": len(unassigned),
            "blocker_count": len(set(blockers)),
        },
    }


def sheet_output_from_assigned_results(
    *,
    assignments: list[dict[str, Any]],
    blockers: list[str] | None = None,
    unassigned_evidence: list[dict[str, Any]] | None = None,
) -> dict[str, Any]:
    cells: dict[str, dict[str, Any]] = {}
    row_indexes: set[int] = set()
    col_indexes: set[int] = set()
    output_blockers = list(blockers or [])
    for assignment in assignments:
        sheet_cell = str(assignment.get("sheet_cell") or "").strip()
        if not sheet_cell:
            output_blockers.append(f"assignment missing sheet_cell: {assignment.get('target_cell_id')}")
            continue
        try:
            worksheet_row = int(assignment.get("worksheet_row") or 0)
            worksheet_col = int(assignment.get("worksheet_col") or 0)
        except Exception:
            output_blockers.append(f"assignment has invalid worksheet position: {sheet_cell}")
            continue
        if worksheet_row <= 0 or worksheet_col <= 0:
            output_blockers.append(f"assignment has non-positive worksheet position: {sheet_cell}")
            continue
        if sheet_cell in cells:
            output_blockers.append(f"duplicate sheet cell assignment: {sheet_cell}")
            continue
        row_indexes.add(worksheet_row)
        col_indexes.add(worksheet_col)
        state = str(assignment.get("assignment_state") or "")
        if state == "conflict":
            output_blockers.append(f"conflicting assignment cannot enter sheet: {sheet_cell}")
        cells[sheet_cell] = {
            "sheet_cell": sheet_cell,
            "worksheet_row": worksheet_row,
            "worksheet_col": worksheet_col,
            "column_letter": get_column_letter(worksheet_col),
            "semantic_field": assignment.get("semantic_field"),
            "value_text": assignment.get("assigned_value") or "",
            "value_normalized": assignment.get("assigned_value") or "",
            "assignment_state": state,
            "assignment_confidence": assignment.get("assignment_confidence"),
            "target_cell_id": assignment.get("target_cell_id"),
            "evidence_ids": list(assignment.get("evidence_ids") or []),
            "raw_texts": list(assignment.get("raw_texts") or []),
        }

    sorted_cols = sorted(col_indexes)
    rows: list[dict[str, Any]] = []
    for row_index in sorted(row_indexes):
        values_by_column: dict[str, str] = {}
        cells_by_column: dict[str, dict[str, Any]] = {}
        states_by_column: dict[str, str] = {}
        for col_index in sorted_cols:
            col_letter = get_column_letter(col_index)
            sheet_cell = f"{col_letter}{row_index}"
            cell = cells.get(sheet_cell)
            values_by_column[col_letter] = str((cell or {}).get("value_text") or "")
            states_by_column[col_letter] = str((cell or {}).get("assignment_state") or "missing")
            if cell is not None:
                cells_by_column[col_letter] = cell
        rows.append(
            {
                "worksheet_row": row_index,
                "values_by_column": values_by_column,
                "states_by_column": states_by_column,
                "cells_by_column": cells_by_column,
            }
        )

    unassigned = list(unassigned_evidence or [])
    if unassigned:
        output_blockers.append(f"unassigned OCR evidence exists: {len(unassigned)}")
    return {
        "cells": cells,
        "rows": rows,
        "columns": [get_column_letter(index) for index in sorted_cols],
        "unassigned_evidence": unassigned,
        "blockers": sorted(set(output_blockers)),
        "summary": {
            "cell_count": len(cells),
            "row_count": len(rows),
            "column_count": len(sorted_cols),
            "blocker_count": len(set(output_blockers)),
            "unassigned_evidence_count": len(unassigned),
            "conflict_cell_count": sum(1 for cell in cells.values() if cell["assignment_state"] == "conflict"),
        },
    }
