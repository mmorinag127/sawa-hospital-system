from __future__ import annotations

from copy import copy, deepcopy
from datetime import UTC, date, datetime
import hashlib
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.units import pixels_to_EMU
from openpyxl.worksheet.worksheet import Worksheet

from src.services import config_service, menu_service, order_form_service, sheet_week_service


DATA_DIR = Path(__file__).resolve().parents[1] / "data"
MASTER_TEMPLATE_PATH = DATA_DIR / "order_form_master_templates" / "master_layout_template.xlsx"
MASTER_SHEET_NAME = "master_layout"
FACILITY_TEMPLATE_SHEET_NAME = "facility_template"
GENERATED_START_COL = 5  # E
MASTER_GENERATED_END_COL = 12  # L in the canonical common source workbook
HEADER_TOP_ROW = 7
HEADER_GROUP_BOTTOM_ROW = 8
HEADER_LEAF_TOP_ROW = 9
HEADER_BOTTOM_ROW = 10
BODY_START_ROW = 11
BODY_END_ROW = 66
PRINT_END_ROW = 69
FACILITY_NAME_RANGE = "A4:E5"
SOURCE_GENERATED_PIXEL_PROFILE = (172, 144, 224, 144, 172, 144, 141, 144)
SOURCE_GENERATED_PIXEL_WIDTH = sum(SOURCE_GENERATED_PIXEL_PROFILE)
DEFAULT_DPI = 144
EMU_PER_PIXEL = 9525
DAYPART_ORDER = {"朝": 0, "昼": 1, "夕": 2}


class FacilityTemplateBuildError(ValueError):
    pass


def _style_side(style: str = "thin") -> Side:
    return Side(style=style, color="000000")


def _set_border(cell, *, left=None, right=None, top=None, bottom=None) -> None:
    current = cell.border
    cell.border = Border(
        left=left or current.left,
        right=right or current.right,
        top=top or current.top,
        bottom=bottom or current.bottom,
    )


def _copy_side(side) -> Side | None:
    if side is None or not getattr(side, "style", None):
        return None
    return copy(side)


def _column_label(column: dict[str, Any]) -> str:
    for key in ("header", "display_name", "label", "name"):
        value = str(column.get(key) or "").strip()
        if value:
            return value
    role = str(column.get("role") or "").strip()
    return role or "列"


def _column_name(column: dict[str, Any], fallback_index: int) -> str:
    value = str(column.get("name") or "").strip()
    if value:
        return value
    role = str(column.get("role") or "").strip() or "column"
    return f"{role}_{fallback_index}"


def _normalise_columns(columns: Any) -> list[dict[str, Any]]:
    if not isinstance(columns, list):
        return []
    normalised: list[dict[str, Any]] = []
    for fallback_index, raw in enumerate(columns):
        if not isinstance(raw, dict):
            continue
        column = deepcopy(raw)
        column["index"] = int(column.get("index") if column.get("index") is not None else fallback_index)
        column["role"] = str(column.get("role") or "").strip()
        column["header"] = _column_label(column)
        column["name"] = _column_name(column, fallback_index)
        normalised.append(column)
    return sorted(normalised, key=lambda item: int(item.get("index") or 0))


def _columns_from_facility_config(facility_config: dict[str, Any]) -> list[dict[str, Any]]:
    fax_template = facility_config.get("fax_template") if isinstance(facility_config, dict) else None
    if isinstance(fax_template, dict) and isinstance(fax_template.get("columns"), list):
        return _normalise_columns(fax_template.get("columns"))
    override = facility_config.get("fax_template_override") if isinstance(facility_config, dict) else None
    if isinstance(override, dict) and isinstance(override.get("columns"), list):
        return _normalise_columns(override.get("columns"))
    return []


def _generated_columns(columns: list[dict[str, Any]]) -> list[dict[str, Any]]:
    menu_index = next(
        (
            int(column.get("index") or 0)
            for column in columns
            if str(column.get("role") or "").strip() in {"menu_name", "menu"}
        ),
        2,
    )
    result: list[dict[str, Any]] = []
    for column in columns:
        role = str(column.get("role") or "").strip()
        try:
            column_index = int(column.get("index") or 0)
        except Exception:
            column_index = 0
        if role in {"date", "daypart", "menu_name", "menu"}:
            continue
        if role == "aux" and column_index <= menu_index:
            continue
        if role == "aux" and column_index > menu_index:
            result.append(column)
            continue
        if role in {"quantity", "note", "remarks"} or str(column.get("name") or "").strip() == "remarks":
            result.append(column)
    return result


def _unmerge_generated_area(ws: Worksheet, end_col: int) -> None:
    for merged_range in list(ws.merged_cells.ranges):
        min_col, min_row, max_col, max_row = merged_range.bounds
        if max_row < HEADER_TOP_ROW or min_row > BODY_END_ROW:
            continue
        if max_col < GENERATED_START_COL or min_col > end_col:
            continue
        ws.unmerge_cells(str(merged_range))


def _ensure_generated_capacity(ws: Worksheet, required_count: int) -> int:
    if required_count < 1:
        raise FacilityTemplateBuildError("facility_template_generated_columns_missing")
    capacity = MASTER_GENERATED_END_COL - GENERATED_START_COL + 1
    if required_count > capacity:
        ws.insert_cols(MASTER_GENERATED_END_COL + 1, required_count - capacity)
    return GENERATED_START_COL + required_count - 1


def _trim_unused_generated_columns(ws: Worksheet, end_col: int) -> None:
    if end_col >= MASTER_GENERATED_END_COL:
        return
    ws.delete_cols(end_col + 1, MASTER_GENERATED_END_COL - end_col)


def _normalize_static_right_merges(ws: Worksheet, end_col: int) -> None:
    for row in range(2, 7):
        for merged_range in list(ws.merged_cells.ranges):
            min_col, min_row, max_col, max_row = merged_range.bounds
            if min_row == row and max_row == row and min_col == 7 and max_col == MASTER_GENERATED_END_COL:
                value = ws.cell(row, min_col).value
                alignment = copy(ws.cell(row, min_col).alignment)
                font = copy(ws.cell(row, min_col).font)
                ws.unmerge_cells(str(merged_range))
                ws.merge_cells(start_row=row, start_column=min_col, end_row=row, end_column=end_col)
                cell = ws.cell(row, min_col)
                cell.value = value
                cell.alignment = alignment
                cell.font = font
                break


def _clear_generated_area(ws: Worksheet, end_col: int) -> None:
    for row in range(HEADER_TOP_ROW, BODY_END_ROW + 1):
        for col in range(GENERATED_START_COL, end_col + 1):
            cell = ws.cell(row, col)
            if isinstance(cell, MergedCell):
                continue
            if HEADER_TOP_ROW <= row <= HEADER_BOTTOM_ROW:
                cell.value = None
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _write_header(ws: Worksheet, generated_columns: list[dict[str, Any]], end_col: int) -> None:
    header_fill = PatternFill(fill_type="solid", start_color="F2F2F2", end_color="F2F2F2")
    thick = _style_side("thick")
    thin = _style_side("thin")

    col_by_offset = {
        offset: GENERATED_START_COL + offset
        for offset, _column in enumerate(generated_columns)
    }
    last_logical_col = GENERATED_START_COL + len(generated_columns) - 1
    offset = 0
    while offset < len(generated_columns):
        column = generated_columns[offset]
        group = str(column.get("header_group") or "").strip()
        group_end = offset
        if group:
            while (
                group_end + 1 < len(generated_columns)
                and str(generated_columns[group_end + 1].get("header_group") or "").strip() == group
            ):
                group_end += 1

        if group and group_end > offset:
            start_col = col_by_offset[offset]
            stop_col = col_by_offset[group_end]
            ws.merge_cells(
                start_row=HEADER_TOP_ROW,
                start_column=start_col,
                end_row=HEADER_GROUP_BOTTOM_ROW,
                end_column=stop_col,
            )
            ws.cell(HEADER_TOP_ROW, start_col).value = group
            for inner_offset in range(offset, group_end + 1):
                leaf_col = col_by_offset[inner_offset]
                ws.merge_cells(
                    start_row=HEADER_LEAF_TOP_ROW,
                    start_column=leaf_col,
                    end_row=HEADER_BOTTOM_ROW,
                    end_column=leaf_col,
                )
                ws.cell(HEADER_LEAF_TOP_ROW, leaf_col).value = _column_label(
                    generated_columns[inner_offset]
                )
        else:
            col = col_by_offset[offset]
            role = str(column.get("role") or "").strip()
            name = str(column.get("name") or "").strip()
            merge_end_col = end_col if col == last_logical_col and (role in {"note", "remarks"} or name == "remarks") else col
            ws.merge_cells(
                start_row=HEADER_TOP_ROW,
                start_column=col,
                end_row=HEADER_BOTTOM_ROW,
                end_column=merge_end_col,
            )
            ws.cell(HEADER_TOP_ROW, col).value = _column_label(column)
        offset = group_end + 1

    for row in range(HEADER_TOP_ROW, HEADER_BOTTOM_ROW + 1):
        for col in range(GENERATED_START_COL, end_col + 1):
            cell = ws.cell(row, col)
            cell.fill = header_fill
            cell.font = Font(name="Yu Gothic", bold=True, size=10)
            _set_border(cell, left=thin, right=thin, top=thin, bottom=thin)

    for col in range(GENERATED_START_COL, end_col + 1):
        _set_border(ws.cell(HEADER_TOP_ROW, col), top=thick)
        _set_border(ws.cell(HEADER_BOTTOM_ROW, col), bottom=thick)
    for row in range(HEADER_TOP_ROW, HEADER_BOTTOM_ROW + 1):
        _set_border(ws.cell(row, GENERATED_START_COL), left=thick)
        _set_border(ws.cell(row, end_col), right=thick)


def _write_body_grid(ws: Worksheet, end_col: int) -> None:
    thick = _style_side("thick")
    thin = _style_side("thin")
    for row in range(BODY_START_ROW, BODY_END_ROW + 1):
        row_top = _copy_side(ws.cell(row, 4).border.top)
        row_bottom = _copy_side(ws.cell(row, 4).border.bottom)
        for col in range(GENERATED_START_COL, end_col + 1):
            cell = ws.cell(row, col)
            _set_border(cell, left=thin, right=thin, top=row_top, bottom=row_bottom)
        _set_border(ws.cell(row, GENERATED_START_COL), left=thick)
        _set_border(ws.cell(row, end_col), right=thick)
    for col in range(GENERATED_START_COL, end_col + 1):
        _set_border(ws.cell(BODY_END_ROW, col), bottom=thick)


def _week_sheet_name_from_week_value(week_value: object | None) -> str | None:
    _month_id, start_date, end_date = sheet_week_service.parse_sheet_week_value(week_value)
    if not isinstance(start_date, date) or not isinstance(end_date, date):
        return None
    return f"{start_date.month}月{start_date.day}日～{end_date.month}月{end_date.day}日"


def _source_index_for_generated_column(column: dict[str, Any], *, generated_offset: int) -> int:
    try:
        return int(column.get("source_index"))
    except Exception:
        return GENERATED_START_COL - 1 + int(generated_offset)


def _apply_source_body_merges(
    ws: Worksheet,
    *,
    facility_config: dict[str, Any],
    generated_columns: list[dict[str, Any]],
    week_value: object | None,
    end_col: int,
) -> int:
    week_sheet_name = _week_sheet_name_from_week_value(week_value)
    if not week_sheet_name:
        return 0
    try:
        source_workbook_name = order_form_service.resolve_facility_source_workbook_name_for_week_sheet(
            facility_config,
            week_sheet_name,
        )
        source_path = order_form_service._resolve_source_workbook_path(source_workbook_name)  # noqa: SLF001
        source_workbook = load_workbook(source_path, data_only=True)
    except Exception:
        return 0
    try:
        if week_sheet_name not in source_workbook.sheetnames:
            return 0
        source_ws = source_workbook[week_sheet_name]
        source_to_generated_col = {
            _source_index_for_generated_column(column, generated_offset=offset) + 1: GENERATED_START_COL + offset
            for offset, column in enumerate(generated_columns)
        }
        applied = 0
        for merged_range in list(source_ws.merged_cells.ranges):
            min_col, min_row, max_col, max_row = merged_range.bounds
            if min_col != max_col:
                continue
            if max_row <= min_row:
                continue
            if min_row < BODY_START_ROW or max_row > BODY_END_ROW:
                continue
            target_col = source_to_generated_col.get(int(min_col))
            if target_col is None or target_col < GENERATED_START_COL or target_col > end_col:
                continue
            ws.merge_cells(
                start_row=int(min_row),
                start_column=int(target_col),
                end_row=int(max_row),
                end_column=int(target_col),
            )
            applied += 1
        return applied
    finally:
        source_workbook.close()


def _generated_widths(generated_columns: list[dict[str, Any]]) -> list[float]:
    count = len(generated_columns)
    if count <= 0:
        return []
    if count == len(SOURCE_GENERATED_PIXEL_PROFILE):
        desired_pixels = list(SOURCE_GENERATED_PIXEL_PROFILE)
    elif count >= 3:
        fixed = list(SOURCE_GENERATED_PIXEL_PROFILE[:2])
        distributed = _distribute_pixels(SOURCE_GENERATED_PIXEL_WIDTH - sum(fixed), count - len(fixed))
        desired_pixels = fixed + distributed
    else:
        desired_pixels = _distribute_pixels(SOURCE_GENERATED_PIXEL_WIDTH, count)
    return [_column_width_from_render_pixels(px) for px in desired_pixels]


def _distribute_pixels(total: int, count: int) -> list[int]:
    if count <= 0:
        return []
    base = total // count
    remainder = total - base * count
    return [base + (1 if index >= count - remainder else 0) for index in range(count)]


def _column_width_to_render_pixels(width: float, *, dpi: int = 144) -> int:
    return max(8, int(round((float(width) * 7.0 + 5.0) * float(dpi) / 96.0)))


def _column_width_from_render_pixels(pixels: int, *, dpi: int = 144) -> float:
    return max(1.0, ((float(pixels) * 96.0 / float(dpi)) - 5.0) / 7.0)


def _row_height_to_render_pixels(height_points: float | None, *, dpi: int = DEFAULT_DPI) -> int:
    points = float(height_points if height_points is not None else 15.0)
    return max(12, int(round(points * float(dpi) / 72.0)))


def _worksheet_positions(ws: Worksheet, *, max_col: int, max_row: int, dpi: int = DEFAULT_DPI) -> tuple[dict[int, int], dict[int, int]]:
    x_positions = {1: 0}
    current_x = 0
    for col in range(1, max_col + 1):
        dim = ws.column_dimensions[get_column_letter(col)]
        current_x += _column_width_to_render_pixels(float(dim.width or 8.43), dpi=dpi)
        x_positions[col + 1] = current_x

    y_positions = {1: 0}
    current_y = 0
    default_row_height = ws.sheet_format.defaultRowHeight
    for row in range(1, max_row + 1):
        dim = ws.row_dimensions[row]
        current_y += _row_height_to_render_pixels(dim.height or default_row_height, dpi=dpi)
        y_positions[row + 1] = current_y
    return x_positions, y_positions


def _anchor_offset_to_render_pixels(value: int | float | None, *, dpi: int = DEFAULT_DPI) -> int:
    return int(round(float(value or 0) / float(EMU_PER_PIXEL) * float(dpi) / 96.0))


def _freeze_images_to_master_render_size(ws: Worksheet) -> None:
    """Make top-area images independent from generated quantity-column widths."""
    images = list(getattr(ws, "_images", []))
    if not images:
        return
    x_positions, y_positions = _worksheet_positions(
        ws,
        max_col=MASTER_GENERATED_END_COL,
        max_row=BODY_END_ROW,
    )
    for image in images:
        anchor = getattr(image, "anchor", None)
        start = getattr(anchor, "_from", None)
        end = getattr(anchor, "to", None)
        if start is None or end is None:
            continue
        start_col = int(getattr(start, "col", 0)) + 1
        start_row = int(getattr(start, "row", 0)) + 1
        end_col = int(getattr(end, "col", 0)) + 1
        end_row = int(getattr(end, "row", 0)) + 1
        if start_col not in x_positions or end_col not in x_positions or start_row not in y_positions or end_row not in y_positions:
            continue
        start_x = x_positions[start_col] + _anchor_offset_to_render_pixels(getattr(start, "colOff", 0))
        start_y = y_positions[start_row] + _anchor_offset_to_render_pixels(getattr(start, "rowOff", 0))
        end_x = x_positions[end_col] + _anchor_offset_to_render_pixels(getattr(end, "colOff", 0))
        end_y = y_positions[end_row] + _anchor_offset_to_render_pixels(getattr(end, "rowOff", 0))
        target_width = max(1, int(round((end_x - start_x) * 96.0 / float(DEFAULT_DPI))))
        target_height = max(1, int(round((end_y - start_y) * 96.0 / float(DEFAULT_DPI))))
        image.width = target_width
        image.height = target_height
        image.anchor = OneCellAnchor(
            _from=deepcopy(start),
            ext=XDRPositiveSize2D(
                cx=pixels_to_EMU(target_width),
                cy=pixels_to_EMU(target_height),
            ),
        )


def _set_generated_widths(ws: Worksheet, generated_columns: list[dict[str, Any]], end_col: int) -> None:
    widths = _generated_widths(generated_columns)
    expected_count = end_col - GENERATED_START_COL + 1
    if len(widths) != expected_count:
        raise FacilityTemplateBuildError(
            f"generated_width_count_mismatch:{len(widths)}!={expected_count}"
        )
    for offset, _width in enumerate(widths):
        col = GENERATED_START_COL + offset
        ws.column_dimensions[get_column_letter(col)].width = widths[offset]


def _write_facility_name(ws: Worksheet, facility_name: str) -> None:
    text = str(facility_name or "").strip()
    if not text:
        return
    for merged_range in list(ws.merged_cells.ranges):
        if str(merged_range) == FACILITY_NAME_RANGE:
            break
    else:
        ws.merge_cells(FACILITY_NAME_RANGE)
    cell = ws["A4"]
    cell.value = text
    cell.font = Font(name="Yu Gothic", bold=True, size=18)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False, shrink_to_fit=True)


def _cell_value_for_digest(value: object) -> object:
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return value


def _side_for_digest(side: Any) -> dict[str, Any]:
    return {
        "style": getattr(side, "style", None),
        "color": getattr(getattr(side, "color", None), "rgb", None),
    }


def _cell_style_for_digest(cell: Any) -> dict[str, Any]:
    alignment = cell.alignment
    font = cell.font
    fill = cell.fill
    border = cell.border
    return {
        "alignment": {
            "horizontal": alignment.horizontal,
            "vertical": alignment.vertical,
            "wrap_text": alignment.wrap_text,
            "shrink_to_fit": alignment.shrink_to_fit,
            "text_rotation": alignment.textRotation,
        },
        "font": {
            "name": font.name,
            "size": font.sz,
            "bold": font.bold,
            "italic": font.italic,
            "color": getattr(getattr(font, "color", None), "rgb", None),
        },
        "fill": {
            "fill_type": fill.fill_type,
            "fgColor": getattr(fill.fgColor, "rgb", None),
            "bgColor": getattr(fill.bgColor, "rgb", None),
        },
        "border": {
            "left": _side_for_digest(border.left),
            "right": _side_for_digest(border.right),
            "top": _side_for_digest(border.top),
            "bottom": _side_for_digest(border.bottom),
        },
        "number_format": cell.number_format,
    }


def _worksheet_digest_payload(ws: Worksheet, *, max_row: int, max_col: int) -> dict[str, Any]:
    cells: list[dict[str, Any]] = []
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            if isinstance(cell, MergedCell):
                continue
            cells.append(
                {
                    "row": row,
                    "col": col,
                    "value": _cell_value_for_digest(cell.value),
                    "style": _cell_style_for_digest(cell),
                }
            )
    images: list[dict[str, Any]] = []
    for image in getattr(ws, "_images", []) or []:
        anchor = getattr(image, "anchor", None)
        marker = getattr(anchor, "_from", None)
        images.append(
            {
                "row": getattr(marker, "row", None),
                "col": getattr(marker, "col", None),
                "rowOff": getattr(marker, "rowOff", None),
                "colOff": getattr(marker, "colOff", None),
                "width": getattr(image, "width", None),
                "height": getattr(image, "height", None),
            }
        )
    return {
        "title": ws.title,
        "print_area": str(ws.print_area or ""),
        "max_row": max_row,
        "max_col": max_col,
        "merged_ranges": sorted(str(item) for item in ws.merged_cells.ranges),
        "rows": [
            {
                "row": row,
                "height": ws.row_dimensions[row].height,
                "hidden": ws.row_dimensions[row].hidden,
            }
            for row in range(1, max_row + 1)
        ],
        "columns": [
            {
                "col": col,
                "letter": get_column_letter(col),
                "width": ws.column_dimensions[get_column_letter(col)].width,
                "hidden": ws.column_dimensions[get_column_letter(col)].hidden,
            }
            for col in range(1, max_col + 1)
        ],
        "images": sorted(images, key=lambda item: (item.get("row") or 0, item.get("col") or 0)),
        "cells": cells,
    }


def _stable_digest(value: object) -> str:
    payload = json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"), default=str)
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def _file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def build_facility_template_diagnostics(
    *,
    facility_config: dict[str, Any],
    week_value: object | None = None,
    week_menu_entries: Any | None = None,
    master_template_path: Path | str = MASTER_TEMPLATE_PATH,
) -> dict[str, Any]:
    master_path = Path(master_template_path)
    workbook = build_facility_template_workbook(
        facility_config=facility_config,
        week_value=week_value,
        week_menu_entries=week_menu_entries,
        master_template_path=master_path,
    )
    worksheet = workbook[FACILITY_TEMPLATE_SHEET_NAME]
    schema = workbook["generated_template_schema"]
    end_col = worksheet.max_column
    schema_rows = [
        [_cell_value_for_digest(value) for value in row]
        for row in schema.iter_rows(values_only=True)
    ]
    schema_rows_for_digest = [
        ([row[0], "<generated_at_utc>", *row[2:]] if row and row[0] == "generated_at_utc" else row)
        for row in schema_rows
    ]
    facility_payload = _worksheet_digest_payload(
        worksheet,
        max_row=PRINT_END_ROW,
        max_col=end_col,
    )
    canonical_payload = {
        "facility_template": facility_payload,
        "generated_template_schema": schema_rows_for_digest,
    }
    return {
        "master_template_sha256": _file_sha256(master_path),
        "master_template_name": master_path.name,
        "facility_id": str(facility_config.get("facility_id") or facility_config.get("id") or "").strip(),
        "week_value": str(week_value or ""),
        "facility_template_canonical_digest": _stable_digest(canonical_payload),
        "facility_template_payload_digest": _stable_digest(facility_payload),
        "schema_digest": _stable_digest(schema_rows_for_digest),
        "generated_end_col": end_col,
        "generated_end_letter": get_column_letter(end_col),
        "schema": schema_rows,
    }


def _parse_menu_date(value: object) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value or "").strip()
    if not text:
        return None
    try:
        return date.fromisoformat(text[:10])
    except Exception:
        return None


def _normalize_week_menu_entries(*, entries: Any, week_value: object) -> list[dict[str, Any]]:
    month_id, start_date, end_date = sheet_week_service.parse_sheet_week_value(week_value)
    if not month_id or not isinstance(start_date, date) or not isinstance(end_date, date):
        return []
    if not isinstance(entries, list):
        return []
    normalized: list[dict[str, Any]] = []
    for index, entry in enumerate(entries):
        if not isinstance(entry, dict):
            continue
        menu_date = _parse_menu_date(entry.get("menu_date"))
        daypart = str(entry.get("daypart") or "").strip()
        menu_name = str(entry.get("name") or "").strip()
        if not isinstance(menu_date, date) or not daypart or not menu_name:
            continue
        if not (start_date <= menu_date <= end_date):
            continue
        normalized.append(
            {
                **entry,
                "_menu_date_obj": menu_date,
                "_source_index": index,
            }
        )
    return sorted(
        normalized,
        key=lambda item: (
            item["_menu_date_obj"],
            DAYPART_ORDER.get(str(item.get("daypart") or "").strip(), 99),
            int(item.get("sort_order") or item.get("display_order") or item.get("order") or item["_source_index"]),
            int(item["_source_index"]),
        ),
    )


def _collect_week_menu_entries(*, facility_id: str, week_value: object) -> list[dict[str, Any]]:
    month_id, start_date, end_date = sheet_week_service.parse_sheet_week_value(week_value)
    if not month_id or not isinstance(start_date, date) or not isinstance(end_date, date):
        return []
    payload = menu_service.get_menu_for_facility(month_id, facility_id)
    entries = payload.get("entries") if isinstance(payload, dict) else None
    return _normalize_week_menu_entries(entries=entries, week_value=week_value)


def _weekday_label(menu_date: date) -> str:
    labels = ["月", "火", "水", "木", "金", "土", "日"]
    return labels[menu_date.weekday()]


def _clear_body_identity_values(ws: Worksheet) -> None:
    for row in range(BODY_START_ROW, BODY_END_ROW + 1):
        for col in range(1, 5):
            cell = ws.cell(row=row, column=col)
            if isinstance(cell, MergedCell):
                continue
            cell.value = None


def _write_cell_if_writable(ws: Worksheet, *, row: int, column: int, value: object) -> None:
    cell = ws.cell(row=row, column=column)
    if isinstance(cell, MergedCell):
        return
    cell.value = value


def _write_week_menu_identity(ws: Worksheet, entries: list[dict[str, Any]]) -> int:
    _clear_body_identity_values(ws)
    row_idx = BODY_START_ROW
    current_date: date | None = None
    date_start_row = BODY_START_ROW
    current_daypart = ""
    written = 0
    for entry in entries:
        menu_date = entry.get("_menu_date_obj")
        if not isinstance(menu_date, date):
            continue
        if row_idx > BODY_END_ROW:
            raise FacilityTemplateBuildError("facility_template_week_menu_exceeds_supported_rows")
        daypart = str(entry.get("daypart") or "").strip()
        category = str(entry.get("category") or "").strip()
        name = str(entry.get("name") or "").strip()
        if current_date != menu_date:
            if current_date is not None and row_idx - 1 > date_start_row:
                _write_cell_if_writable(ws, row=row_idx - 1, column=1, value=_weekday_label(current_date))
            current_date = menu_date
            date_start_row = row_idx
            current_daypart = ""
            _write_cell_if_writable(ws, row=row_idx, column=1, value=menu_date)
        _write_cell_if_writable(ws, row=row_idx, column=2, value=daypart if current_daypart != daypart else None)
        _write_cell_if_writable(ws, row=row_idx, column=3, value=category)
        _write_cell_if_writable(ws, row=row_idx, column=4, value=name)
        current_daypart = daypart
        row_idx += 1
        written += 1
    if current_date is not None and row_idx - 1 > date_start_row:
        _write_cell_if_writable(ws, row=row_idx - 1, column=1, value=_weekday_label(current_date))
    return written


def _apply_week_menu_identity(
    ws: Worksheet,
    *,
    facility_config: dict[str, Any],
    week_value: object | None,
    week_menu_entries: Any | None = None,
) -> int:
    if not week_value:
        return 0
    facility_id = str(facility_config.get("facility_id") or facility_config.get("id") or "").strip()
    if not facility_id:
        return 0
    entries = (
        _normalize_week_menu_entries(entries=week_menu_entries, week_value=week_value)
        if week_menu_entries is not None
        else _collect_week_menu_entries(facility_id=facility_id, week_value=week_value)
    )
    if not entries:
        return 0
    return _write_week_menu_identity(ws, entries)


def _append_schema_sheet(wb, *, facility_config: dict[str, Any], generated_columns: list[dict[str, Any]], end_col: int) -> None:
    if "generated_template_schema" in wb.sheetnames:
        del wb["generated_template_schema"]
    schema = wb.create_sheet("generated_template_schema")
    schema.sheet_state = "hidden"
    rows = [
        ["key", "value"],
        ["generated_at_utc", datetime.now(UTC).isoformat()],
        ["source", "master_layout_template"],
        ["facility_id", str(facility_config.get("facility_id") or facility_config.get("id") or "")],
        ["facility_name", str(facility_config.get("facility_name") or facility_config.get("name") or "")],
        ["generated_start_col", GENERATED_START_COL],
        ["generated_end_col", end_col],
        ["generated_column_count", len(generated_columns)],
        ["source_generated_pixel_width", SOURCE_GENERATED_PIXEL_WIDTH],
    ]
    for row in rows:
        schema.append(row)
    schema.append([])
    schema.append(["index", "role", "header", "header_group", "name", "diet_type", "area_id", "source_index"])
    for offset, column in enumerate(generated_columns):
        schema.append(
            [
                offset,
                column.get("role"),
                column.get("header"),
                column.get("header_group"),
                column.get("name"),
                column.get("diet_type"),
                column.get("area_id"),
                column.get("source_index"),
            ]
        )


def build_facility_template_workbook(
    *,
    facility_config: dict[str, Any],
    master_template_path: Path | str = MASTER_TEMPLATE_PATH,
    week_value: object | None = None,
    week_menu_entries: Any | None = None,
):
    master_path = Path(master_template_path)
    if not master_path.exists():
        raise FacilityTemplateBuildError(f"master_template_not_found:{master_path}")
    columns = _columns_from_facility_config(facility_config)
    generated_columns = _generated_columns(columns)
    if not generated_columns:
        raise FacilityTemplateBuildError("facility_template_generated_columns_missing")

    wb = load_workbook(master_path)
    if MASTER_SHEET_NAME not in wb.sheetnames:
        raise FacilityTemplateBuildError(f"master_sheet_not_found:{MASTER_SHEET_NAME}")
    ws = wb[MASTER_SHEET_NAME]
    ws.title = FACILITY_TEMPLATE_SHEET_NAME
    _freeze_images_to_master_render_size(ws)

    facility_name = str(facility_config.get("facility_name") or facility_config.get("name") or "").strip()
    _write_facility_name(ws, facility_name)
    written_menu_rows = _apply_week_menu_identity(
        ws,
        facility_config=facility_config,
        week_value=week_value,
        week_menu_entries=week_menu_entries,
    )

    end_col = _ensure_generated_capacity(ws, len(generated_columns))
    _normalize_static_right_merges(ws, end_col)
    _unmerge_generated_area(ws, max(end_col, MASTER_GENERATED_END_COL))
    _trim_unused_generated_columns(ws, end_col)
    _clear_generated_area(ws, end_col)
    _write_header(ws, generated_columns, end_col)
    _write_body_grid(ws, end_col)
    applied_body_merges = _apply_source_body_merges(
        ws,
        facility_config=facility_config,
        generated_columns=generated_columns,
        week_value=week_value,
        end_col=end_col,
    )
    _set_generated_widths(ws, generated_columns, end_col)
    ws.print_area = f"A1:{get_column_letter(end_col)}{PRINT_END_ROW}"
    _append_schema_sheet(wb, facility_config=facility_config, generated_columns=generated_columns, end_col=end_col)
    wb["generated_template_schema"].append(["week_value", str(week_value or "")])
    wb["generated_template_schema"].append(["week_menu_rows", written_menu_rows])
    wb["generated_template_schema"].append(["source_body_merged_ranges", applied_body_merges])
    return wb


def build_facility_template_xlsx(
    *,
    facility_config: dict[str, Any],
    output_path: Path | str,
    master_template_path: Path | str = MASTER_TEMPLATE_PATH,
    week_value: object | None = None,
    week_menu_entries: Any | None = None,
) -> Path:
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    wb = build_facility_template_workbook(
        facility_config=facility_config,
        master_template_path=master_template_path,
        week_value=week_value,
        week_menu_entries=week_menu_entries,
    )
    wb.save(output)
    return output


def build_facility_template_xlsx_for_facility(
    *,
    facility_id: str,
    output_path: Path | str,
    master_template_path: Path | str = MASTER_TEMPLATE_PATH,
    week_value: object | None = None,
    week_menu_entries: Any | None = None,
) -> Path:
    facility_config = config_service.get_facility_config(facility_id)
    if not facility_config:
        raise FacilityTemplateBuildError(f"facility_not_found:{facility_id}")
    return build_facility_template_xlsx(
        facility_config=facility_config,
        output_path=output_path,
        master_template_path=master_template_path,
        week_value=week_value,
        week_menu_entries=week_menu_entries,
    )
