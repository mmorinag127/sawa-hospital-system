from __future__ import annotations

from copy import deepcopy
from datetime import date, datetime, timedelta
import hashlib
import json
from pathlib import Path
import os
import re
from typing import Any
from urllib.parse import urlparse

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter, range_boundaries
from sqlalchemy import select

from src.db import session_scope
from src.models.order import Order
from src.models.order_version import OrderVersion
from src.models.order_sheet_draft import OrderSheetDraft
from src.models.order_workflow_state import OrderWorkflowState
from src.services import config_service, draft_sheet_service, menu_service, sheet_week_service
from src.services.ingest_policy import parse_date_string
from src.services.storage_service import (
    get_default_output_bucket,
    load_bytes_from_uri,
    save_artifact_bytes_to_gcs,
)


_MONTH_ID_RE = re.compile(r"^\d{4}-\d{2}$")
_OUTPUT_DIR = Path(os.getenv("ORDER_FORM_OUTPUT_DIR", "/tmp/order-form-outputs"))
_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
_FAX_OUTPUT_DIR = Path(os.getenv("FAX_ORDER_FORM_OUTPUT_DIR", "/tmp/fax-order-form-prototypes"))
_FAX_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
_DEFAULT_WEEK_SHEET = "3月22日～3月28日"
_ORDER_FORM_BODY_START_ROW = 11
_ORDER_FORM_BODY_END_ROW = 67
_ORDER_FORM_DEADLINE_SEARCH_MAX_ROW = 6
_ORDER_FORM_DEADLINE_SEARCH_MAX_COL = 16
_BOTTOM_MARKER_ROW = 69
_WEEKDAY_LABELS = ["（月）", "（火）", "（水）", "（木）", "（金）", "（土）", "（日）"]
_MARKER_FILL = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
_META_FILL = PatternFill(start_color="E9EEF5", end_color="E9EEF5", fill_type="solid")
_THIN_BORDER = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000"),
)
_ORDER_FORM_SOURCE_TEMPLATE_ASSET_DIR = Path(__file__).resolve().parents[1] / "data" / "order_form_source_workbooks"
_ORDER_FORM_SOURCE_TEMPLATE_MANIFEST = _ORDER_FORM_SOURCE_TEMPLATE_ASSET_DIR / "manifest.json"
_SOURCE_WORKBOOK_MATERIALIZED_DIR = Path(
    os.getenv("FAX_SOURCE_TEMPLATE_MATERIALIZED_DIR", "/tmp/fax-source-template-workbooks")
)
_SOURCE_WORKBOOK_UPLOAD_DIR = Path(
    os.getenv("FAX_SOURCE_TEMPLATE_UPLOAD_DIR", "/tmp/fax-source-template-uploads")
)


class WeekEntriesWriteResult(int):
    def __new__(
        cls,
        written_rows: int,
        *,
        source_rows: int,
        overflow_entries: list[dict[str, str]],
    ):
        obj = int.__new__(cls, written_rows)
        obj.written_rows = int(written_rows)
        obj.source_rows = int(source_rows)
        obj.overflow_entries = overflow_entries
        return obj

    @property
    def overflow_rows(self) -> int:
        return len(self.overflow_entries)


def _resolve_fax_source_template_dir() -> Path:
    configured = os.getenv("FAX_SOURCE_TEMPLATE_DIR", "").strip()
    if configured:
        return Path(configured)
    return _ORDER_FORM_SOURCE_TEMPLATE_ASSET_DIR


def _format_generated_at(value: datetime | None = None) -> str:
    return (value or datetime.utcnow()).strftime("%Y-%m-%d %H:%M")


def _current_order_version_metadata(session, order_id: str) -> dict[str, Any]:
    if not hasattr(session, "execute"):
        return {}
    order = session.get(Order, order_id) if hasattr(session, "get") else None
    try:
        rows = (
            session.execute(
                select(OrderVersion)
                .where(OrderVersion.order_id == order_id)
                .order_by(OrderVersion.version_no.desc(), OrderVersion.received_at.desc(), OrderVersion.id.desc())
            )
            .scalars()
            .all()
        )
    except Exception:
        return {}
    if not rows:
        if order is None or not str(getattr(order, "current_document_id", "") or "").strip():
            return {}
        received_at = getattr(order, "received_at", None)
        return {
            "fax_version_no": 1,
            "fax_version_count": 1,
            "fax_version_document_id": getattr(order, "current_document_id", "") or "",
            "fax_version_message_id": getattr(order, "message_id", "") or "",
            "fax_version_received_at": received_at.isoformat() if hasattr(received_at, "isoformat") else "",
        }
    current = next((row for row in rows if bool(row.is_current)), rows[0])
    return {
        "fax_version_no": int(current.version_no or 0),
        "fax_version_count": len(rows),
        "fax_version_document_id": current.document_id,
        "fax_version_message_id": current.message_id,
        "fax_version_received_at": current.received_at.isoformat() if current.received_at else "",
    }


def _load_fax_source_manifest() -> dict[str, dict]:
    if not _ORDER_FORM_SOURCE_TEMPLATE_MANIFEST.exists():
        raise RuntimeError(
            f"order-form source manifest not found: {_ORDER_FORM_SOURCE_TEMPLATE_MANIFEST}"
        )
    payload = json.loads(_ORDER_FORM_SOURCE_TEMPLATE_MANIFEST.read_text(encoding="utf-8"))
    families = payload.get("families")
    if not isinstance(families, dict) or not families:
        raise RuntimeError(
            f"order-form source manifest has no families: {_ORDER_FORM_SOURCE_TEMPLATE_MANIFEST}"
        )
    normalized: dict[str, dict] = {}
    for template_id, raw_spec in families.items():
        if not isinstance(raw_spec, dict):
            raise RuntimeError(f"invalid order-form source spec: {template_id}")
        family_label = str(raw_spec.get("family_label") or "").strip()
        if not family_label:
            raise RuntimeError(f"missing family_label for order-form source spec: {template_id}")
        month_sources_raw = raw_spec.get("month_sources") or {}
        month_sources: dict[str, str] = {}
        if isinstance(month_sources_raw, dict):
            for month_id, filename in month_sources_raw.items():
                normalized_month = _normalize_month_id(month_id)
                filename_text = str(filename or "").strip()
                if not filename_text:
                    raise RuntimeError(
                        f"missing filename for order-form source spec: {template_id} month={month_id}"
                    )
                month_sources[normalized_month] = filename_text
        source_workbook = str(raw_spec.get("source_workbook") or "").strip()
        if not month_sources and not source_workbook:
            raise RuntimeError(
                f"order-form source spec must define month_sources or source_workbook: {template_id}"
            )
        normalized[str(template_id).strip()] = {
            "family_label": family_label,
            "month_sources": month_sources,
            "source_workbook": source_workbook,
        }
    return normalized


def list_order_form_patterns() -> list[dict]:
    return config_service.get_order_form_patterns()


def _normalize_month_id(month_id: str) -> str:
    value = str(month_id or "").strip()
    if not _MONTH_ID_RE.match(value):
        raise ValueError("month_id must be YYYY-MM")
    return value


_FAX_SOURCE_TEMPLATE_DIR = _resolve_fax_source_template_dir()
_FAX_FAMILY_SOURCE_MAP = _load_fax_source_manifest()
_SOURCE_WORKBOOK_SHEET_CACHE: dict[str, tuple[str, ...]] = {}
_SOURCE_WORKBOOK_MERGED_QUANTITY_CACHE: dict[tuple[str, str], bool] = {}
_ORDER_FORM_HEADER_ROWS = (7, 8, 9)
_ORDER_FORM_QUANTITY_HEADER_KEYWORDS = (
    "常食",
    "糖尿",
    "軟菜",
    "ミキサ",
    "ﾐｷｻ",
    "禁食",
    "肉禁",
    "魚禁",
    "職員",
    "通所",
    "変更",
    "袋分",
    "合計",
    "その他",
    "2F",
    "3F",
    "２F",
    "３F",
    "２Ｆ",
    "３Ｆ",
    "1回",
    "2回",
    "3回",
    "１回",
    "２回",
    "３回",
)
_ORDER_FORM_NON_QUANTITY_HEADER_KEYWORDS = (
    "日付",
    "区分",
    "献立",
    "メニュー",
    "備考",
)


def _resolve_pattern(facility: dict, pattern_id: str | None) -> dict:
    if pattern_id:
        pattern = config_service.get_order_form_pattern(pattern_id)
        if pattern:
            return pattern
    facility_pattern = facility.get("order_form_pattern_id")
    if isinstance(facility_pattern, str) and facility_pattern.strip():
        pattern = config_service.get_order_form_pattern(facility_pattern.strip())
        if pattern:
            return pattern
    patterns = config_service.get_order_form_patterns()
    if patterns:
        return dict(patterns[0])
    return {"pattern_id": "PATTERN_A", "label": "標準A", "marker_cells": []}


def _safe_source_workbook_suffix(value: str) -> str:
    suffix = Path(str(value or "")).suffix.lower()
    return suffix if suffix in {".xlsx", ".xlsm"} else ".xlsx"


def _materialize_source_workbook_uri(uri: str) -> Path:
    data = load_bytes_from_uri(uri)
    digest = hashlib.sha256(data).hexdigest()
    target = _SOURCE_WORKBOOK_MATERIALIZED_DIR / f"{digest}{_safe_source_workbook_suffix(uri)}"
    target.parent.mkdir(parents=True, exist_ok=True)
    if not target.exists():
        target.write_bytes(data)
    return target


def _resolve_source_workbook_path(source_workbook_name: str) -> Path:
    source_ref = str(source_workbook_name or "").strip()
    parsed = urlparse(source_ref)
    if parsed.scheme in {"gs", "file"}:
        source_path = _materialize_source_workbook_uri(source_ref)
    else:
        direct_path = Path(source_ref)
        source_path = direct_path if direct_path.is_absolute() else _FAX_SOURCE_TEMPLATE_DIR / source_ref
    if not source_path.exists():
        raise ValueError(
            f"source workbook not found: {source_workbook_name} (dir={_FAX_SOURCE_TEMPLATE_DIR})"
        )
    return source_path


def _resolve_source_workbook_name_for_month(fax_template_id: str, month_id: str) -> str:
    spec = _resolve_fax_family_spec(fax_template_id)
    month_sources = spec.get("month_sources") or {}
    normalized_month = _normalize_month_id(month_id)
    if normalized_month in month_sources:
        return str(month_sources[normalized_month])
    source_workbook = str(spec.get("source_workbook") or "").strip()
    if source_workbook:
        return source_workbook
    raise ValueError(
        f"source workbook not configured for fax_template_id={fax_template_id} month_id={normalized_month}"
    )


def _resolve_source_workbook_name_for_template_clone(fax_template_id: str, month_id: str) -> str:
    spec = _resolve_fax_family_spec(fax_template_id)
    month_sources = spec.get("month_sources") or {}
    normalized_month = _normalize_month_id(month_id)
    if normalized_month in month_sources:
        return str(month_sources[normalized_month])
    earlier_months = [source_month for source_month in month_sources if source_month <= normalized_month]
    if earlier_months:
        return str(month_sources[sorted(earlier_months)[-1]])
    if month_sources:
        return str(month_sources[sorted(month_sources)[-1]])
    source_workbook = str(spec.get("source_workbook") or "").strip()
    if source_workbook:
        return source_workbook
    raise ValueError(
        f"source workbook not configured for fax_template_id={fax_template_id} month_id={normalized_month}"
    )


def _facility_source_workbook_names(facility: dict | None) -> list[str]:
    if not isinstance(facility, dict):
        return []
    source_names: list[str] = []
    raw_month_sources = facility.get("order_form_month_sources")
    if isinstance(raw_month_sources, dict):
        for month_id in sorted(raw_month_sources):
            source_workbook = str(raw_month_sources.get(month_id) or "").strip()
            if source_workbook:
                source_names.append(source_workbook)
    raw_month_source_uris = facility.get("order_form_month_source_uris")
    if isinstance(raw_month_source_uris, dict):
        for month_id in sorted(raw_month_source_uris):
            source_workbook = str(raw_month_source_uris.get(month_id) or "").strip()
            if source_workbook:
                source_names.append(source_workbook)
    source_workbook = str(facility.get("order_form_source_workbook") or "").strip()
    if source_workbook:
        source_names.append(source_workbook)
    source_workbook_uri = str(facility.get("order_form_source_workbook_uri") or "").strip()
    if source_workbook_uri:
        source_names.append(source_workbook_uri)
    result: list[str] = []
    seen: set[str] = set()
    for source_workbook in source_names:
        if source_workbook in seen:
            continue
        seen.add(source_workbook)
        result.append(source_workbook)
    return result


def _resolve_facility_source_workbook_name_for_month(
    facility: dict,
    *,
    fax_template_id: str,
    month_id: str,
) -> str:
    normalized_month = _normalize_month_id(month_id)
    raw_month_sources = facility.get("order_form_month_sources")
    if isinstance(raw_month_sources, dict):
        source_workbook = str(raw_month_sources.get(normalized_month) or "").strip()
        if source_workbook:
            return source_workbook
    raw_month_source_uris = facility.get("order_form_month_source_uris")
    if isinstance(raw_month_source_uris, dict):
        source_workbook = str(raw_month_source_uris.get(normalized_month) or "").strip()
        if source_workbook:
            return source_workbook
    source_workbook = str(facility.get("order_form_source_workbook") or "").strip()
    if source_workbook:
        return source_workbook
    source_workbook_uri = str(facility.get("order_form_source_workbook_uri") or "").strip()
    if source_workbook_uri:
        return source_workbook_uri
    return _resolve_source_workbook_name_for_month(fax_template_id, normalized_month)


def _resolve_facility_source_workbook_name_for_template_clone(
    facility: dict,
    *,
    fax_template_id: str,
    month_id: str,
) -> str:
    normalized_month = _normalize_month_id(month_id)
    raw_month_sources = facility.get("order_form_month_sources")
    if isinstance(raw_month_sources, dict):
        source_workbook = str(raw_month_sources.get(normalized_month) or "").strip()
        if source_workbook:
            return source_workbook
        earlier_months = [
            str(source_month)
            for source_month, source_name in raw_month_sources.items()
            if str(source_month) <= normalized_month and str(source_name or "").strip()
        ]
        if earlier_months:
            return str(raw_month_sources[sorted(earlier_months)[-1]])
    raw_month_source_uris = facility.get("order_form_month_source_uris")
    if isinstance(raw_month_source_uris, dict):
        source_workbook = str(raw_month_source_uris.get(normalized_month) or "").strip()
        if source_workbook:
            return source_workbook
        earlier_months = [
            str(source_month)
            for source_month, source_name in raw_month_source_uris.items()
            if str(source_month) <= normalized_month and str(source_name or "").strip()
        ]
        if earlier_months:
            return str(raw_month_source_uris[sorted(earlier_months)[-1]])
    source_workbook = str(facility.get("order_form_source_workbook") or "").strip()
    if source_workbook:
        return source_workbook
    source_workbook_uri = str(facility.get("order_form_source_workbook_uri") or "").strip()
    if source_workbook_uri:
        return source_workbook_uri
    return _resolve_source_workbook_name_for_template_clone(fax_template_id, normalized_month)


def _source_workbook_sheetnames(source_workbook_name: str) -> tuple[str, ...]:
    source_path = _resolve_source_workbook_path(source_workbook_name)
    cache_key = str(source_path)
    cached = _SOURCE_WORKBOOK_SHEET_CACHE.get(cache_key)
    if cached is not None:
        return cached
    workbook = load_workbook(source_path, read_only=True)
    try:
        sheetnames = tuple(workbook.sheetnames)
    finally:
        workbook.close()
    _SOURCE_WORKBOOK_SHEET_CACHE[cache_key] = sheetnames
    return sheetnames


def _normalize_order_form_header_text(value: object) -> str:
    return re.sub(r"[\s　]+", "", str(value or "")).strip()


def _worksheet_effective_merged_values(worksheet) -> dict[tuple[int, int], object]:
    values: dict[tuple[int, int], object] = {}
    for merged_range in worksheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
        anchor_value = worksheet.cell(row=min_row, column=min_col).value
        for row_idx in range(min_row, max_row + 1):
            for col_idx in range(min_col, max_col + 1):
                values[(row_idx, col_idx)] = anchor_value
    return values


def _worksheet_quantity_column_indexes(worksheet) -> set[int]:
    merged_values = _worksheet_effective_merged_values(worksheet)
    quantity_columns: set[int] = set()
    for col_idx in range(1, worksheet.max_column + 1):
        header_blob = "".join(
            _normalize_order_form_header_text(
                worksheet.cell(row=row_idx, column=col_idx).value
                or merged_values.get((row_idx, col_idx))
            )
            for row_idx in _ORDER_FORM_HEADER_ROWS
        )
        if not header_blob:
            continue
        if any(keyword in header_blob for keyword in _ORDER_FORM_NON_QUANTITY_HEADER_KEYWORDS):
            continue
        if any(keyword in header_blob for keyword in _ORDER_FORM_QUANTITY_HEADER_KEYWORDS):
            quantity_columns.add(col_idx)
    return quantity_columns


def _worksheet_has_vertical_merged_quantity_cells(worksheet) -> bool:
    quantity_columns = _worksheet_quantity_column_indexes(worksheet)
    if not quantity_columns:
        return False
    for merged_range in worksheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
        if max_row <= min_row:
            continue
        if max_row < _ORDER_FORM_BODY_START_ROW or min_row > _ORDER_FORM_BODY_END_ROW:
            continue
        if any(col_idx in quantity_columns for col_idx in range(min_col, max_col + 1)):
            return True
    return False


def source_workbook_has_vertical_merged_quantity_cells(
    source_workbook_name: str,
    *,
    week_sheet_name: str | None = None,
) -> bool:
    source_path = _resolve_source_workbook_path(source_workbook_name)
    cache_key = (str(source_path), str(week_sheet_name or ""))
    cached = _SOURCE_WORKBOOK_MERGED_QUANTITY_CACHE.get(cache_key)
    if cached is not None:
        return cached
    workbook = load_workbook(source_path, data_only=True)
    try:
        if week_sheet_name:
            if week_sheet_name not in workbook.sheetnames:
                raise ValueError(f"week sheet not found in source workbook: {week_sheet_name}")
            worksheets = [workbook[week_sheet_name]]
        else:
            worksheets = list(workbook.worksheets)
        detected = any(_worksheet_has_vertical_merged_quantity_cells(worksheet) for worksheet in worksheets)
    finally:
        workbook.close()
    _SOURCE_WORKBOOK_MERGED_QUANTITY_CACHE[cache_key] = detected
    return detected


def facility_template_has_vertical_merged_quantity_cells(
    facility: dict[str, Any] | None,
    *,
    week_sheet_name: str | None = None,
) -> bool:
    _ = week_sheet_name
    if not isinstance(facility, dict):
        return False
    if isinstance(facility.get("expanded_cell_same_daypart_copy_enabled"), bool):
        return bool(facility.get("expanded_cell_same_daypart_copy_enabled"))
    template_candidates = [
        facility.get("fax_template_override"),
        facility.get("fax_template"),
    ]
    for template in template_candidates:
        if not isinstance(template, dict):
            continue
        policy = template.get("body_merge_policy")
        if isinstance(policy, dict):
            mode = str(policy.get("mode") or "").strip().lower()
            columns = [str(item or "").strip() for item in (policy.get("columns") or []) if str(item or "").strip()]
            if mode == "daypart" and columns:
                return True
        for column in template.get("columns") or []:
            if not isinstance(column, dict):
                continue
            body_merge = str(column.get("body_merge") or column.get("body_merge_mode") or "").strip().lower()
            if body_merge == "daypart":
                return True
    return False


def _resolve_source_workbook_name_for_week_sheet(fax_template_id: str, week_sheet_name: str) -> str:
    spec = _resolve_fax_family_spec(fax_template_id)
    month_sources = spec.get("month_sources") or {}
    for month_id in sorted(month_sources):
        source_workbook_name = str(month_sources[month_id])
        if week_sheet_name in _source_workbook_sheetnames(source_workbook_name):
            return source_workbook_name
    source_workbook = str(spec.get("source_workbook") or "").strip()
    if source_workbook:
        return source_workbook
    raise ValueError(
        f"week sheet not configured for fax_template_id={fax_template_id}: {week_sheet_name}"
    )


def resolve_facility_source_workbook_name_for_week_sheet(facility: dict, week_sheet_name: str) -> str:
    fax_template_id = str(_infer_fax_template_id_from_facility(facility) or "").strip()
    if not fax_template_id:
        raise ValueError("facility fax_template_id not found")
    for source_workbook_name in _facility_source_workbook_names(facility):
        if week_sheet_name in _source_workbook_sheetnames(source_workbook_name):
            return source_workbook_name
    return _resolve_source_workbook_name_for_week_sheet(fax_template_id, week_sheet_name)


def save_facility_source_workbook_upload(
    *,
    facility_id: str,
    filename: str,
    data: bytes,
) -> dict[str, Any]:
    if not data:
        raise ValueError("source workbook is empty")
    suffix = _safe_source_workbook_suffix(filename)
    digest = hashlib.sha256(data).hexdigest()
    safe_name = _sanitize_filename_fragment(Path(filename or "source_workbook").stem)
    object_name = f"{safe_name}_{digest[:12]}{suffix}"
    bucket = get_default_output_bucket()
    if bucket:
        uri = save_artifact_bytes_to_gcs(
            bucket,
            f"facility-template-{facility_id}",
            object_name,
            data,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        storage = "gcs"
    else:
        _SOURCE_WORKBOOK_UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
        target = _SOURCE_WORKBOOK_UPLOAD_DIR / object_name
        target.write_bytes(data)
        uri = str(target)
        storage = "local"
    # Validate the uploaded workbook before it is referenced by facility config.
    workbook_path = _materialize_source_workbook_uri(f"file://{uri}") if uri.startswith("/") else _resolve_source_workbook_path(uri)
    workbook = load_workbook(workbook_path, read_only=True)
    try:
        sheetnames = list(workbook.sheetnames)
    finally:
        workbook.close()
    if not sheetnames:
        raise ValueError("source workbook has no sheets")
    return {
        "uri": uri,
        "storage": storage,
        "filename": filename,
        "sha256": digest,
        "sheetnames": sheetnames,
    }


def _resolve_facility(facility_id: str) -> dict:
    facility = config_service.get_facility_config(facility_id)
    if not facility:
        facility = config_service.get_facility_by_id(facility_id)
    if not facility:
        raise ValueError("facility not found")
    return facility


def _infer_fax_template_id_from_facility(facility: dict) -> str | None:
    explicit = str(facility.get("fax_template_id") or "").strip()
    if explicit:
        return explicit

    facility_text = " ".join(
        [
            str(facility.get("facility_name") or ""),
            *[str(item or "") for item in (facility.get("aliases") or [])],
        ]
    )
    if "ふれあい" in facility_text:
        return "fax_layout_regular_staff_daycare_other_forbidden_v1"
    if "池袋病院" in facility_text:
        return "fax_layout_soft_packaging_forbidden_v1"

    columns = (
        ((facility.get("fax_template_override") or {}).get("columns") or [])
        or ((facility.get("invoice_template") or {}).get("columns") or [])
    )
    headers: list[str] = []
    diet_types: set[str] = set()
    area_ids: set[str] = set()
    for column in columns:
        if not isinstance(column, dict):
            continue
        header = str(column.get("header") or column.get("name") or "").strip()
        if header:
            headers.append(header)
        diet_type = str(column.get("diet_type") or "").strip()
        if diet_type:
            diet_types.add(diet_type)
        area_id = str(column.get("area_id") or "").strip()
        if area_id:
            area_ids.add(area_id.lower())

    header_blob = " ".join(headers)
    if "糖尿" in header_blob or "diabetes" in diet_types or "糖尿" in diet_types:
        return "fax_layout_regular_diabetes_v1"
    if "池袋病院" in facility_text or (
        "軟菜" in header_blob
        and ("袋分け" in header_blob or "袋分" in header_blob)
        and "常食" not in header_blob
    ):
        return "fax_layout_soft_packaging_forbidden_v1"
    if "ふれあい" in facility_text or "通所" in header_blob or "その他" in header_blob:
        return "fax_layout_regular_staff_daycare_other_forbidden_v1"
    if {"2f", "3f"} & area_ids or "2F" in header_blob or "3F" in header_blob:
        return "fax_layout_floor_2f3f_v1"
    if (
        ("regular_bag" in diet_types or "袋分け" in header_blob or "袋分" in header_blob)
        and {"soft", "mixer"} & diet_types
    ):
        return "fax_layout_regular_soft_mixer_forbidden_v1"
    if {"staff", "daycare"} & diet_types or "職員" in header_blob or "通所" in header_blob:
        return "fax_layout_regular_staff_daycare_v1"
    if {"no_meat", "no_fish", "change_1", "change_2"} & diet_types or "禁食" in header_blob:
        return "fax_layout_regular_forbidden_v1"
    return "fax_layout_regular_forbidden_v1"


def _parse_menu_date(value: object) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value or "").strip()
    if not text:
        return None
    try:
        return date.fromisoformat(text[:10])
    except Exception:
        return None


def _collect_menu_entries(month_id: str, facility_id: str) -> list[dict]:
    payload = menu_service.get_menu_for_facility(month_id, facility_id)
    if not payload:
        raise ValueError("monthly menu not found")
    entries = payload.get("entries")
    if not isinstance(entries, list) or not entries:
        raise ValueError("monthly menu entries not found")
    normalized: list[dict] = []
    for entry in entries:
        if not isinstance(entry, dict):
            continue
        menu_date = _parse_menu_date(entry.get("menu_date"))
        daypart = entry.get("daypart")
        name = entry.get("name")
        if not menu_date or not daypart or not name:
            continue
        normalized.append(
            {
                **entry,
                "_menu_date_obj": menu_date,
            }
        )
    if not normalized:
        raise ValueError("no usable menu entries")
    return normalized


def _month_end(month_start: date) -> date:
    if month_start.month == 12:
        next_month = date(month_start.year + 1, 1, 1)
    else:
        next_month = date(month_start.year, month_start.month + 1, 1)
    return next_month - timedelta(days=1)


def _build_week_ranges_for_month(month_id: str) -> list[tuple[date, date]]:
    month_start = date.fromisoformat(f"{month_id}-01")
    month_end = _month_end(month_start)
    ranges: list[tuple[date, date]] = []
    current = month_start
    while current <= month_end:
        days_until_saturday = (5 - current.weekday()) % 7
        week_end = min(current + timedelta(days=days_until_saturday), month_end)
        ranges.append((current, week_end))
        current = week_end + timedelta(days=1)
    return ranges


def _format_week_sheet_name(start_date: date, end_date: date) -> str:
    return f"{start_date.month}月{start_date.day}日～{end_date.month}月{end_date.day}日"


def _select_entries_for_range(entries: list[dict], start_date: date, end_date: date) -> list[dict]:
    return [
        entry
        for entry in entries
        if isinstance(entry.get("_menu_date_obj"), date) and start_date <= entry["_menu_date_obj"] <= end_date
    ]


def _clone_sheet_images(source_worksheet, target_worksheet) -> None:
    for image in getattr(source_worksheet, "_images", []):
        target_worksheet.add_image(deepcopy(image))


def _ensure_workbook_sheet_count(workbook: Workbook, template_sheet_name: str, target_count: int) -> None:
    if target_count < 1:
        raise ValueError("target_count must be >= 1")
    template_sheet = workbook[template_sheet_name]
    while len(workbook.sheetnames) < target_count:
        copied = workbook.copy_worksheet(template_sheet)
        _clone_sheet_images(template_sheet, copied)
    while len(workbook.sheetnames) > target_count:
        del workbook[workbook.sheetnames[-1]]


def _clear_week_sheet_body(worksheet) -> None:
    for row in worksheet.iter_rows(
        min_row=_ORDER_FORM_BODY_START_ROW,
        max_row=_ORDER_FORM_BODY_END_ROW,
        min_col=1,
        max_col=worksheet.max_column,
    ):
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            cell.value = None


def _daypart_key(value: object) -> str:
    return str(value or "").strip()


def _week_menu_daypart_key(value: object) -> str:
    text = str(value or "").strip()
    if text.startswith("朝"):
        return "朝"
    if text.startswith("昼"):
        return "昼"
    if text.startswith("夕"):
        return "夕"
    return text


def _week_menu_daypart_capacity(value: object) -> int | None:
    key = _week_menu_daypart_key(value)
    if key == "朝":
        return 2
    if key in {"昼", "夕"}:
        return 3
    return None


def _week_entry_overflow_payload(entry: dict) -> dict[str, str]:
    menu_date = entry.get("_menu_date_obj")
    return {
        "date": menu_date.isoformat() if isinstance(menu_date, date) else str(entry.get("menu_date") or ""),
        "daypart": str(entry.get("daypart") or ""),
        "category": str(entry.get("category") or ""),
        "name": str(entry.get("name") or ""),
    }


def _split_week_entries_for_order_form(week_entries: list[dict]) -> tuple[list[dict], list[dict[str, str]], int]:
    valid_entries = [entry for entry in week_entries if isinstance(entry.get("_menu_date_obj"), date)]
    body_capacity = _ORDER_FORM_BODY_END_ROW - _ORDER_FORM_BODY_START_ROW + 1
    if len(valid_entries) <= body_capacity:
        return valid_entries, [], len(valid_entries)

    physical_entries: list[dict] = []
    overflow_entries: list[dict[str, str]] = []
    counts_by_daypart: dict[tuple[date, str], int] = {}
    for entry in valid_entries:
        menu_date = entry["_menu_date_obj"]
        daypart_key = _week_menu_daypart_key(entry.get("daypart"))
        capacity = _week_menu_daypart_capacity(entry.get("daypart"))
        if capacity is None:
            physical_entries.append(entry)
            continue
        counter_key = (menu_date, daypart_key)
        counts_by_daypart[counter_key] = counts_by_daypart.get(counter_key, 0) + 1
        if counts_by_daypart[counter_key] > capacity:
            overflow_entries.append(_week_entry_overflow_payload(entry))
            continue
        physical_entries.append(entry)

    if len(physical_entries) > body_capacity:
        overflow_entries.extend(_week_entry_overflow_payload(entry) for entry in physical_entries[body_capacity:])
        physical_entries = physical_entries[:body_capacity]

    return physical_entries, overflow_entries, len(valid_entries)


def _write_facility_name_in_box(worksheet, facility_name: str) -> None:
    text = str(facility_name or "").strip()
    name_length = len(text)
    if name_length <= 10:
        font_size = 18
    elif name_length <= 16:
        font_size = 16
    elif name_length <= 24:
        font_size = 14
    elif name_length <= 32:
        font_size = 12
    else:
        font_size = 10
    worksheet["A4"] = text
    worksheet["A4"].font = Font(name="Meiryo", size=font_size, bold=True)
    worksheet["A4"].alignment = Alignment(horizontal="left", vertical="center", shrink_to_fit=True, indent=1)


def _weekday_label(menu_date: date) -> str:
    return _WEEKDAY_LABELS[menu_date.weekday()]


def _column_merged_ranges(worksheet, col_idx: int) -> list[tuple[int, int, int, int]]:
    ranges: list[tuple[int, int, int, int]] = []
    for merged_range in worksheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
        if min_col <= col_idx <= max_col:
            ranges.append((min_col, min_row, max_col, max_row))
    return sorted(ranges, key=lambda item: (item[1], item[0], item[3], item[2]))


def _write_weekday_label_for_date_block(
    worksheet,
    *,
    date_start_row: int,
    date_end_row: int,
    menu_date: date,
) -> None:
    date_range: tuple[int, int, int, int] | None = None
    col_ranges = _column_merged_ranges(worksheet, 1)
    for merged in col_ranges:
        _min_col, min_row, _max_col, max_row = merged
        if min_row <= date_start_row <= max_row:
            date_range = merged
            break
    target_row = date_end_row
    if date_range is not None:
        _min_col, _min_row, _max_col, date_range_end = date_range
        for merged in col_ranges:
            _min_col, min_row, _max_col, max_row = merged
            if min_row > date_range_end and min_row <= date_end_row <= max_row:
                target_row = min_row
                break
    _write_to_cell_or_merged_anchor(
        worksheet,
        row_idx=target_row,
        col_idx=1,
        value=_weekday_label(menu_date),
    )


def _write_week_entries(worksheet, week_entries: list[dict]) -> WeekEntriesWriteResult:
    row_idx = _ORDER_FORM_BODY_START_ROW
    current_date: date | None = None
    date_start_row = _ORDER_FORM_BODY_START_ROW
    current_daypart = ""
    written_rows = 0
    physical_entries, overflow_entries, source_rows = _split_week_entries_for_order_form(week_entries)

    for entry in physical_entries:
        menu_date = entry.get("_menu_date_obj")
        if not isinstance(menu_date, date):
            continue
        daypart = str(entry.get("daypart") or "").strip()
        category = str(entry.get("category") or "")
        menu_name = str(entry.get("name") or "")

        if current_date != menu_date:
            if current_date is not None and row_idx - 1 > date_start_row:
                _write_weekday_label_for_date_block(
                    worksheet,
                    date_start_row=date_start_row,
                    date_end_row=row_idx - 1,
                    menu_date=current_date,
                )
            current_date = menu_date
            date_start_row = row_idx
            current_daypart = ""
            _write_to_cell_or_merged_anchor(worksheet, row_idx=row_idx, col_idx=1, value=menu_date)

        _write_to_cell_or_merged_anchor(
            worksheet,
            row_idx=row_idx,
            col_idx=2,
            value=daypart if current_daypart != daypart else None,
        )
        _write_to_cell_or_merged_anchor(worksheet, row_idx=row_idx, col_idx=3, value=category)
        _write_to_cell_or_merged_anchor(worksheet, row_idx=row_idx, col_idx=4, value=menu_name)

        current_daypart = daypart
        row_idx += 1
        written_rows += 1

    if current_date is not None and row_idx - 1 > date_start_row:
        _write_weekday_label_for_date_block(
            worksheet,
            date_start_row=date_start_row,
            date_end_row=row_idx - 1,
            menu_date=current_date,
        )
    return WeekEntriesWriteResult(
        written_rows,
        source_rows=source_rows,
        overflow_entries=overflow_entries,
    )


def _sheet_field_indexes(fields: list[Any]) -> dict[str, int]:
    result: dict[str, int] = {}
    for idx, field in enumerate(fields):
        normalized = str(field or "").strip()
        if normalized and normalized not in result:
            result[normalized] = idx
    return result


def _normalize_sheet_label(value: object) -> str:
    return re.sub(r"[\s　・]+", "", str(value or "")).strip().lower()


def _sheet_field_label(fields: list[Any], header: list[Any], idx: int) -> str:
    header_text = str(header[idx] or "").strip() if idx < len(header) else ""
    return header_text or str(fields[idx] or "").strip()


def _first_field_index(field_indexes: dict[str, int], candidates: tuple[str, ...]) -> int | None:
    for candidate in candidates:
        if candidate in field_indexes:
            return field_indexes[candidate]
    return None


def _sheet_cell(row: list[Any], idx: int | None) -> object:
    if idx is None or idx < 0 or idx >= len(row):
        return None
    return row[idx]


def _coerce_quantity_cell(value: object) -> object:
    text = str(value or "").strip()
    if not text:
        return None
    normalized = text.replace(",", "")
    try:
        number = float(normalized)
    except ValueError:
        return text
    if number.is_integer():
        return int(number)
    return number


def _field_role(field: object, label: object = "") -> str | None:
    field_text = str(field or "").strip().lower()
    label_text = _normalize_sheet_label(label)
    token = f"{field_text} {label_text}"
    if field_text in {"date", "menu_date", "date_mmdd"} or label_text in {"日付", "日にち"}:
        return "date"
    if field_text in {"daypart", "meal"} or label_text in {"区分", "食区分"}:
        return "daypart"
    if field_text in {"category", "menu_category"} or label_text in {"献立区分", "副区分", "区分2"}:
        return "category"
    if field_text in {"menu_name", "name", "menu"} or label_text in {"献立", "メニュー"}:
        return "menu"
    if field_text in {"remarks", "remark", "note", "notes"} or "備考" in label_text:
        return "remarks"
    if "合計" in label_text or "total" in token:
        return "total"
    if "常食" in label_text or "regular" in token:
        return "regular"
    if "通所" in label_text or "daycare" in token:
        return "daycare"
    if "職員" in label_text or "staff" in token:
        return "staff"
    if "肉禁" in label_text or "nomeat" in token or "no_meat" in token:
        return "no_meat"
    if "魚禁" in label_text or "nofish" in token or "no_fish" in token:
        return "no_fish"
    if "その他" in label_text or "forbiddenother" in token or "forbidden_other" in token:
        return "forbidden_other"
    if "揚禁" in label_text or "nofried" in token or "no_fried" in token:
        return "no_fried"
    if "変更①" in label_text or "変更1" in label_text or "change_1" in token:
        return "change_1"
    if "変更②" in label_text or "変更2" in label_text or "change_2" in token:
        return "change_2"
    if "placeholder" in token or "unknown" in token or "不明" in label_text:
        return None
    return None


def _worksheet_quantity_column_role_map(worksheet) -> dict[str, int]:
    merged_values = _worksheet_effective_merged_values(worksheet)
    result: dict[str, int] = {}
    for col_idx in sorted(_worksheet_quantity_column_indexes(worksheet)):
        header_parts = [
            worksheet.cell(row=row_idx, column=col_idx).value or merged_values.get((row_idx, col_idx))
            for row_idx in _ORDER_FORM_HEADER_ROWS
        ]
        role = _field_role("", "".join(str(part or "") for part in header_parts))
        if role and role not in result:
            result[role] = col_idx
    if "forbidden_other" in result and "no_fried" not in result:
        result["no_fried"] = result["forbidden_other"]
    return result


def _write_to_cell_or_merged_anchor(worksheet, *, row_idx: int, col_idx: int, value: object) -> None:
    cell = worksheet.cell(row=row_idx, column=col_idx)
    if not isinstance(cell, MergedCell):
        cell.value = value
        return
    for merged_range in worksheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
        if min_row <= row_idx <= max_row and min_col <= col_idx <= max_col:
            anchor = worksheet.cell(row=min_row, column=min_col)
            if anchor.value in (None, "") or anchor.value == value:
                anchor.value = value
            return


def _write_saved_sheet_rows_to_order_form(
    worksheet,
    *,
    fields: list[Any],
    header: list[Any] | None = None,
    rows: list[Any],
    received_at: datetime,
) -> int:
    header_values = list(header or [])
    role_indexes: dict[str, int] = {}
    quantity_field_roles: list[tuple[int, str]] = []
    structural_roles = {"date", "daypart", "category", "menu", "remarks"}
    for idx, field in enumerate(fields):
        label = _sheet_field_label(fields, header_values, idx)
        role = _field_role(field, label)
        if role in structural_roles and role not in role_indexes:
            role_indexes[role] = idx
        elif role and role not in structural_roles:
            quantity_field_roles.append((idx, role))

    quantity_columns_by_role = _worksheet_quantity_column_role_map(worksheet)
    if not quantity_columns_by_role:
        raise ValueError("order form quantity columns not found")

    row_idx = _ORDER_FORM_BODY_START_ROW
    current_date: date | None = None
    date_start_row = _ORDER_FORM_BODY_START_ROW
    current_daypart = ""
    written_rows = 0
    for raw_row in rows:
        if not isinstance(raw_row, list):
            continue
        if row_idx > _ORDER_FORM_BODY_END_ROW:
            raise ValueError("saved sheet exceeds supported template rows")
        date_idx = role_indexes.get("date")
        daypart_idx = role_indexes.get("daypart")
        category_idx = role_indexes.get("category")
        menu_idx = role_indexes.get("menu")
        date_value = _sheet_cell(raw_row, date_idx)
        parsed_date = parse_date_string(str(date_value or ""), received_at) if date_value not in (None, "") else None
        daypart = str(_sheet_cell(raw_row, daypart_idx) or "").strip()
        category = str(_sheet_cell(raw_row, category_idx) or "").strip()
        menu_name = str(_sheet_cell(raw_row, menu_idx) or "").strip()
        if not any([parsed_date, daypart, category, menu_name]) and not any(
            str(_sheet_cell(raw_row, idx) or "").strip()
            for idx, _role in quantity_field_roles
        ):
            continue
        if parsed_date and current_date != parsed_date:
            if current_date is not None and row_idx - 1 > date_start_row:
                _write_weekday_label_for_date_block(
                    worksheet,
                    date_start_row=date_start_row,
                    date_end_row=row_idx - 1,
                    menu_date=current_date,
                )
            current_date = parsed_date
            date_start_row = row_idx
            current_daypart = ""
            _write_to_cell_or_merged_anchor(worksheet, row_idx=row_idx, col_idx=1, value=parsed_date)
        _write_to_cell_or_merged_anchor(
            worksheet,
            row_idx=row_idx,
            col_idx=2,
            value=daypart if daypart and current_daypart != daypart else None,
        )
        _write_to_cell_or_merged_anchor(worksheet, row_idx=row_idx, col_idx=3, value=category or None)
        _write_to_cell_or_merged_anchor(worksheet, row_idx=row_idx, col_idx=4, value=menu_name or None)
        for qty_idx, role in quantity_field_roles:
            col_idx = quantity_columns_by_role.get(role)
            if col_idx is None:
                continue
            value = _coerce_quantity_cell(_sheet_cell(raw_row, qty_idx))
            if value is not None:
                _write_to_cell_or_merged_anchor(worksheet, row_idx=row_idx, col_idx=col_idx, value=value)
        if daypart:
            current_daypart = daypart
        row_idx += 1
        written_rows += 1
    if current_date is not None and row_idx - 1 > date_start_row:
        _write_weekday_label_for_date_block(
            worksheet,
            date_start_row=date_start_row,
            date_end_row=row_idx - 1,
            menu_date=current_date,
        )
    return written_rows


def _set_deadline_text_for_week(worksheet, start_date: date) -> None:
    deadline = start_date - timedelta(days=16)
    deadline_text = f"締切日{deadline.month}月{deadline.day}日まで"
    for row_idx in range(1, _ORDER_FORM_DEADLINE_SEARCH_MAX_ROW + 1):
        for col_idx in range(1, _ORDER_FORM_DEADLINE_SEARCH_MAX_COL + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            value = str(cell.value or "")
            if value.startswith("締切日"):
                cell.value = deadline_text
                return


def _append_monthly_metadata_sheet(
    workbook: Workbook,
    *,
    source_workbook_name: str,
    facility_id: str,
    facility_name: str,
    month_id: str,
    pattern: dict,
    fax_template_id: str,
    family_label: str,
    week_ranges: list[tuple[date, date]],
    entry_count: int,
    week_write_results: list[WeekEntriesWriteResult] | None = None,
) -> None:
    if "設定" in workbook.sheetnames:
        del workbook["設定"]
    meta = workbook.create_sheet("設定")
    meta.sheet_state = "hidden"
    meta.append(["key", "value"])
    meta.append(["generated_at_utc", datetime.utcnow().isoformat()])
    meta.append(["source_workbook", source_workbook_name])
    meta.append(["facility_id", facility_id])
    meta.append(["facility_name", facility_name])
    meta.append(["month_id", month_id])
    meta.append(["pattern_id", str(pattern.get("pattern_id") or "")])
    meta.append(["pattern_label", str(pattern.get("label") or "")])
    meta.append(["fax_template_id", fax_template_id])
    meta.append(["family_label", family_label])
    meta.append(["sheet_count", len(week_ranges)])
    meta.append(["entry_count", entry_count])
    total_overflow_rows = sum(int(result.overflow_rows) for result in (week_write_results or []))
    meta.append(["week_menu_overflow_rows", total_overflow_rows])
    for index, (start_date, end_date) in enumerate(week_ranges, start=1):
        meta.append([f"week_{index}_sheet_name", _format_week_sheet_name(start_date, end_date)])
        meta.append([f"week_{index}_start", start_date.isoformat()])
        meta.append([f"week_{index}_end", end_date.isoformat()])
        if week_write_results and index <= len(week_write_results):
            result = week_write_results[index - 1]
            meta.append([f"week_{index}_written_rows", int(result.written_rows)])
            meta.append([f"week_{index}_source_rows", int(result.source_rows)])
            meta.append([f"week_{index}_overflow_rows", int(result.overflow_rows)])
            if result.overflow_rows:
                meta.append(
                    [
                        f"week_{index}_overflow_entries",
                        json.dumps(result.overflow_entries, ensure_ascii=False, sort_keys=True),
                    ]
                )


def _build_monthly_fax_order_form_workbook(
    *,
    facility: dict,
    month_id: str,
    entries: list[dict],
    pattern: dict,
) -> Workbook:
    fax_template_id = str(_infer_fax_template_id_from_facility(facility) or "").strip()
    if not fax_template_id:
        raise ValueError("facility fax_template_id not found")
    spec = _resolve_fax_family_spec(fax_template_id)
    source_workbook_name = _resolve_facility_source_workbook_name_for_month(
        facility,
        fax_template_id=fax_template_id,
        month_id=month_id,
    )
    source_path = _resolve_source_workbook_path(source_workbook_name)

    workbook = load_workbook(source_path)
    template_sheet_name = _DEFAULT_WEEK_SHEET if _DEFAULT_WEEK_SHEET in workbook.sheetnames else workbook.sheetnames[0]
    week_ranges = _build_week_ranges_for_month(month_id)
    _ensure_workbook_sheet_count(workbook, template_sheet_name, len(week_ranges))

    facility_id = str(facility.get("facility_id") or facility.get("id") or "")
    facility_name = str(facility.get("facility_name") or facility.get("name") or facility_id)
    family_label = str(spec["family_label"])
    week_write_results: list[WeekEntriesWriteResult] = []

    for index, (start_date, end_date) in enumerate(week_ranges):
        worksheet = workbook.worksheets[index]
        sheet_name = _format_week_sheet_name(start_date, end_date)
        worksheet.title = sheet_name
        _clear_week_sheet_body(worksheet)
        _write_facility_name_in_box(worksheet, facility_name)
        _set_deadline_text_for_week(worksheet, start_date)
        week_write_results.append(
            _write_week_entries(worksheet, _select_entries_for_range(entries, start_date, end_date))
        )
        _apply_fax_metadata_header(
            worksheet,
            fax_template_id=fax_template_id,
            facility_id=facility_id,
            facility_name=facility_name,
            week_sheet_name=sheet_name,
            family_label=family_label,
        )
        _apply_fax_markers(worksheet)
        _apply_bottom_instruction_strip(worksheet, fax_template_id=fax_template_id, base_label="monthly")
        _extend_print_area(worksheet, bottom_row=_BOTTOM_MARKER_ROW)

    _append_monthly_metadata_sheet(
        workbook,
        source_workbook_name=source_workbook_name,
        facility_id=facility_id,
        facility_name=facility_name,
        month_id=month_id,
        pattern=pattern,
        fax_template_id=fax_template_id,
        family_label=family_label,
        week_ranges=week_ranges,
        entry_count=len(entries),
        week_write_results=week_write_results,
    )
    workbook.active = 0
    return workbook


def build_order_form_excel(
    *,
    facility_id: str,
    month_id: str,
    pattern_id: str | None = None,
) -> Path:
    normalized_month = _normalize_month_id(month_id)
    facility = _resolve_facility(facility_id)
    entries = _collect_menu_entries(normalized_month, facility_id)
    pattern = _resolve_pattern(facility, pattern_id)
    resolved_fax_template_id = _infer_fax_template_id_from_facility(facility)
    wb = _build_monthly_fax_order_form_workbook(
        facility=facility,
        month_id=normalized_month,
        entries=entries,
        pattern=pattern,
    )

    file_pattern = str(resolved_fax_template_id or pattern.get("pattern_id") or "PATTERN_A")
    stamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    output = _OUTPUT_DIR / f"order_form_{facility_id}_{normalized_month}_{file_pattern}_{stamp}.xlsx"
    wb.save(output)
    return output


def _dates_from_saved_sheet(fields: list[Any], rows: list[Any], received_at: datetime) -> list[date]:
    field_indexes = _sheet_field_indexes(fields)
    date_idx = _first_field_index(field_indexes, ("date", "menu_date", "日付"))
    if date_idx is None:
        return []
    dates: list[date] = []
    for row in rows:
        if not isinstance(row, list):
            continue
        parsed = parse_date_string(str(_sheet_cell(row, date_idx) or ""), received_at)
        if parsed:
            dates.append(parsed)
    return dates


def _resolve_week_range_for_saved_sheet(
    *,
    week_code: object,
    fields: list[Any],
    rows: list[Any],
    received_at: datetime,
) -> tuple[str, date, date]:
    month_id, start_date, end_date = sheet_week_service.parse_sheet_week_value(week_code)
    if month_id and isinstance(start_date, date) and isinstance(end_date, date):
        return month_id, start_date, end_date
    saved_dates = _dates_from_saved_sheet(fields, rows, received_at)
    if saved_dates:
        first_date = min(saved_dates)
        for candidate_start, candidate_end in _build_week_ranges_for_month(first_date.strftime("%Y-%m")):
            if candidate_start <= first_date <= candidate_end:
                return candidate_start.strftime("%Y-%m"), candidate_start, candidate_end
    if month_id:
        ranges = _build_week_ranges_for_month(month_id)
        if ranges:
            start, end = ranges[0]
            return month_id, start, end
    raise ValueError("order week is unresolved")


def build_saved_sheet_order_form_excel(*, order_id: str) -> Path:
    normalized_order_id = str(order_id or "").strip()
    if not normalized_order_id:
        raise ValueError("order_id is required")
    generated_at = datetime.utcnow()
    fax_version: dict[str, Any] = {}
    with session_scope() as session:
        order = session.get(Order, normalized_order_id)
        if not order:
            raise ValueError("order not found")
        fax_version = _current_order_version_metadata(session, normalized_order_id)
        facility_id = str(order.facility_code or "").strip()
        week_code = order.week_code
        received_at = order.received_at or datetime.utcnow()
        workflow = session.get(OrderWorkflowState, normalized_order_id)
        workflow_draft_id = str(getattr(workflow, "draft_id", None) or "").strip() or None
        draft_row = session.get(OrderSheetDraft, workflow_draft_id) if workflow_draft_id else None
        if draft_row is not None and draft_row.order_id == normalized_order_id:
            draft = {
                "id": draft_row.id,
                "draft_sheet_json": draft_row.draft_sheet_json if isinstance(draft_row.draft_sheet_json, dict) else {},
            }
        else:
            draft = None
    if not facility_id:
        raise ValueError("facility missing")

    if not isinstance(draft, dict):
        draft = draft_sheet_service.get_latest_sheet_draft(normalized_order_id)
    draft_payload = draft.get("draft_sheet_json") if isinstance(draft, dict) else None
    if not isinstance(draft_payload, dict):
        raise ValueError("saved sheet not found")
    fields = draft_payload.get("fields")
    header = draft_payload.get("header")
    rows = draft_payload.get("rows")
    if not isinstance(fields, list) or not isinstance(rows, list) or not rows:
        raise ValueError("saved sheet rows not found")

    facility = _resolve_facility(facility_id)
    fax_template_id = str(_infer_fax_template_id_from_facility(facility) or "").strip()
    if not fax_template_id:
        raise ValueError("facility fax_template_id not found")
    spec = _resolve_fax_family_spec(fax_template_id)
    month_id, week_start, week_end = _resolve_week_range_for_saved_sheet(
        week_code=week_code,
        fields=fields,
        rows=rows,
        received_at=received_at,
    )
    week_sheet_name = _format_week_sheet_name(week_start, week_end)
    try:
        source_workbook_name = resolve_facility_source_workbook_name_for_week_sheet(facility, week_sheet_name)
    except ValueError as exc:
        if "week sheet not configured" not in str(exc):
            raise
        source_workbook_name = _resolve_facility_source_workbook_name_for_template_clone(
            facility,
            fax_template_id=fax_template_id,
            month_id=month_id,
        )
    source_path = _resolve_source_workbook_path(source_workbook_name)
    workbook = load_workbook(source_path)
    _ensure_week_sheet_from_template(workbook, week_sheet_name)
    _keep_only_target_sheet(workbook, week_sheet_name)
    worksheet = workbook[week_sheet_name]

    facility_name = str(facility.get("facility_name") or facility.get("name") or facility_id)
    _clear_week_sheet_body(worksheet)
    _write_facility_name(worksheet, facility_name)
    _write_facility_name_in_box(worksheet, facility_name)
    _set_deadline_text_for_week(worksheet, week_start)
    written_rows = _write_saved_sheet_rows_to_order_form(
        worksheet,
        fields=fields,
        header=header if isinstance(header, list) else None,
        rows=rows,
        received_at=received_at,
    )
    if written_rows <= 0:
        raise ValueError("saved sheet has no writable rows")
    _apply_fax_metadata_header(
        worksheet,
        fax_template_id=fax_template_id,
        facility_id=facility_id,
        facility_name=facility_name,
        week_sheet_name=week_sheet_name,
        family_label=str(spec["family_label"]),
        generated_at=generated_at,
        label="FAX読取シートExcel作成",
        fax_version=fax_version,
    )
    _apply_fax_markers(worksheet)
    _apply_bottom_instruction_strip(worksheet, fax_template_id=fax_template_id, base_label="saved_sheet")
    _extend_print_area(worksheet, bottom_row=_BOTTOM_MARKER_ROW)
    _append_hidden_metadata_sheet(
        workbook,
        source_workbook_name=source_workbook_name,
        facility_id=facility_id,
        facility_name=facility_name,
        fax_template_id=fax_template_id,
        family_label=str(spec["family_label"]),
        week_sheet_name=week_sheet_name,
        base_label="saved_sheet",
        generated_at=generated_at,
        fax_version=fax_version,
    )
    metadata = workbook["設定"]
    metadata.append(["order_id", normalized_order_id])
    metadata.append(["month_id", month_id])
    metadata.append(["saved_sheet_draft_id", str(draft.get("id") or "") if isinstance(draft, dict) else ""])
    metadata.append(["saved_sheet_row_count", written_rows])

    stamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    safe_facility_id = _sanitize_filename_fragment(facility_id)
    safe_week = _sanitize_filename_fragment(week_sheet_name)
    output = _OUTPUT_DIR / f"fax_read_sheet_excel_{normalized_order_id}_{safe_facility_id}_{safe_week}_{stamp}.xlsx"
    workbook.save(output)
    return output


def list_fax_order_form_template_specs() -> list[dict]:
    specs: list[dict] = []
    for template_id, payload in _FAX_FAMILY_SOURCE_MAP.items():
        month_sources = dict(payload.get("month_sources") or {})
        specs.append(
            {
                "fax_template_id": template_id,
                "source_workbook": payload.get("source_workbook") or next(iter(month_sources.values()), ""),
                "month_sources": month_sources,
                "family_label": payload["family_label"],
            }
        )
    return specs


def build_fax_base_template_excel(
    *,
    fax_template_id: str,
    week_sheet_name: str = _DEFAULT_WEEK_SHEET,
    output_dir: Path | str | None = None,
) -> Path:
    return _render_fax_order_form_workbook(
        source_workbook_name=_resolve_source_workbook_name_for_week_sheet(fax_template_id, week_sheet_name),
        week_sheet_name=week_sheet_name,
        facility_name="施設名記入欄",
        facility_id="BASE",
        fax_template_id=fax_template_id,
        family_label=str(_resolve_fax_family_spec(fax_template_id)["family_label"]),
        base_label="base",
        output_dir=output_dir,
    )


def build_fax_order_form_excel(
    *,
    facility_id: str,
    week_sheet_name: str = _DEFAULT_WEEK_SHEET,
    output_dir: Path | str | None = None,
) -> Path:
    facility = config_service.get_facility_config(facility_id)
    if not facility:
        raise ValueError("facility not found")
    fax_template_id = str(_infer_fax_template_id_from_facility(facility) or "").strip()
    if not fax_template_id:
        raise ValueError("facility fax_template_id not found")
    spec = _resolve_fax_family_spec(fax_template_id)
    facility_name = str(facility.get("facility_name") or facility.get("name") or facility_id)
    return _render_fax_order_form_workbook(
        source_workbook_name=resolve_facility_source_workbook_name_for_week_sheet(facility, week_sheet_name),
        week_sheet_name=week_sheet_name,
        facility_name=facility_name,
        facility_id=facility_id,
        fax_template_id=fax_template_id,
        family_label=str(spec["family_label"]),
        base_label="facility",
        output_dir=output_dir,
    )


def _clear_order_form_body_values(
    worksheet,
    *,
    start_row: int = _ORDER_FORM_BODY_START_ROW,
    end_row: int = _ORDER_FORM_BODY_END_ROW,
    min_col: int = 1,
    max_col: int | None = None,
) -> None:
    target_max_col = max_col if max_col is not None else worksheet.max_column
    for row in worksheet.iter_rows(
        min_row=start_row,
        max_row=end_row,
        min_col=min_col,
        max_col=target_max_col,
    ):
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            cell.value = None


def build_fax_structure_only_excel(
    *,
    facility_id: str,
    week_sheet_name: str = _DEFAULT_WEEK_SHEET,
    output_dir: Path | str | None = None,
) -> Path:
    from src.services import master_order_form_template_service  # noqa: PLC0415

    _ = week_sheet_name
    facility = config_service.get_facility_config(facility_id)
    if not facility:
        raise ValueError("facility not found")
    stamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    safe_facility_id = _sanitize_filename_fragment(facility_id)
    output_path = _resolve_fax_output_dir(output_dir) / f"fax_order_form_structure_only_{safe_facility_id}_{stamp}.xlsx"
    return master_order_form_template_service.build_facility_template_xlsx(
        facility_config=facility,
        output_path=output_path,
    )


def _resolve_fax_family_spec(fax_template_id: str) -> dict:
    template_key = str(fax_template_id or "").strip()
    spec = _FAX_FAMILY_SOURCE_MAP.get(template_key)
    if spec:
        return dict(spec)
    raise ValueError(f"unsupported fax_template_id for prototype generation: {template_key}")


def _render_fax_order_form_workbook(
    *,
    source_workbook_name: str,
    week_sheet_name: str,
    facility_name: str,
    facility_id: str,
    fax_template_id: str,
    family_label: str,
    base_label: str,
    output_dir: Path | str | None,
) -> Path:
    source_path = _resolve_source_workbook_path(source_workbook_name)
    workbook = load_workbook(source_path)
    if week_sheet_name not in workbook.sheetnames:
        raise ValueError(f"week sheet not found in source workbook: {week_sheet_name}")
    _keep_only_target_sheet(workbook, week_sheet_name)
    worksheet = workbook[week_sheet_name]

    _write_facility_name(worksheet, facility_name)
    _apply_fax_metadata_header(
        worksheet,
        fax_template_id=fax_template_id,
        facility_id=facility_id,
        facility_name=facility_name,
        week_sheet_name=week_sheet_name,
        family_label=family_label,
    )
    _apply_fax_markers(worksheet)
    _apply_bottom_instruction_strip(worksheet, fax_template_id=fax_template_id, base_label=base_label)
    _extend_print_area(worksheet, bottom_row=_BOTTOM_MARKER_ROW)
    _append_hidden_metadata_sheet(
        workbook,
        source_workbook_name=source_workbook_name,
        facility_id=facility_id,
        facility_name=facility_name,
        fax_template_id=fax_template_id,
        family_label=family_label,
        week_sheet_name=week_sheet_name,
        base_label=base_label,
    )

    stamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    safe_template_id = _sanitize_filename_fragment(fax_template_id)
    safe_facility_id = _sanitize_filename_fragment(facility_id)
    safe_week = _sanitize_filename_fragment(week_sheet_name)
    output_name = f"fax_order_form_{base_label}_{safe_facility_id}_{safe_week}_{safe_template_id}_{stamp}.xlsx"
    output_path = _resolve_fax_output_dir(output_dir) / output_name
    workbook.save(output_path)
    return output_path


def _render_fax_structure_only_workbook(
    *,
    source_workbook_name: str,
    week_sheet_name: str,
    facility_name: str,
    facility_id: str,
    fax_template_id: str,
    family_label: str,
    output_dir: Path | str | None,
) -> Path:
    source_path = _resolve_source_workbook_path(source_workbook_name)
    workbook = load_workbook(source_path)
    if week_sheet_name not in workbook.sheetnames:
        raise ValueError(f"week sheet not found in source workbook: {week_sheet_name}")
    _keep_only_target_sheet(workbook, week_sheet_name)
    worksheet = workbook[week_sheet_name]

    _clear_order_form_body_values(worksheet)
    _write_facility_name(worksheet, facility_name)
    _extend_print_area(worksheet, bottom_row=_BOTTOM_MARKER_ROW)
    _append_hidden_metadata_sheet(
        workbook,
        source_workbook_name=source_workbook_name,
        facility_id=facility_id,
        facility_name=facility_name,
        fax_template_id=fax_template_id,
        family_label=family_label,
        week_sheet_name=week_sheet_name,
        base_label="structure_only",
    )

    stamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    safe_template_id = _sanitize_filename_fragment(fax_template_id)
    safe_facility_id = _sanitize_filename_fragment(facility_id)
    safe_week = _sanitize_filename_fragment(week_sheet_name)
    output_name = (
        f"fax_order_form_structure_only_{safe_facility_id}_{safe_week}_{safe_template_id}_{stamp}.xlsx"
    )
    output_path = _resolve_fax_output_dir(output_dir) / output_name
    workbook.save(output_path)
    return output_path


def _resolve_fax_output_dir(output_dir: Path | str | None) -> Path:
    if output_dir is None:
        path = _FAX_OUTPUT_DIR
    else:
        path = Path(output_dir)
    path.mkdir(parents=True, exist_ok=True)
    return path


def _keep_only_target_sheet(workbook: Workbook, week_sheet_name: str) -> None:
    for sheet_name in list(workbook.sheetnames):
        if sheet_name == week_sheet_name:
            continue
        del workbook[sheet_name]
    workbook.active = 0


def _ensure_week_sheet_from_template(workbook: Workbook, week_sheet_name: str) -> None:
    if week_sheet_name in workbook.sheetnames:
        return
    template_sheet_name = _DEFAULT_WEEK_SHEET if _DEFAULT_WEEK_SHEET in workbook.sheetnames else workbook.sheetnames[0]
    template_sheet = workbook[template_sheet_name]
    copied = workbook.copy_worksheet(template_sheet)
    _clone_sheet_images(template_sheet, copied)
    copied.title = week_sheet_name


def _write_facility_name(worksheet, facility_name: str) -> None:
    worksheet["A3"] = facility_name
    worksheet["A3"].font = Font(name="Meiryo", size=11, bold=True)
    worksheet["A3"].alignment = Alignment(horizontal="left", vertical="center")


def _apply_fax_metadata_header(
    worksheet,
    *,
    fax_template_id: str,
    facility_id: str,
    facility_name: str,
    week_sheet_name: str,
    family_label: str,
    generated_at: datetime | None = None,
    label: str = "発注書作成",
    fax_version: dict[str, Any] | None = None,
) -> None:
    _ = fax_template_id
    _ = facility_id
    _ = facility_name
    _ = week_sheet_name
    _ = family_label
    version_text = ""
    if isinstance(fax_version, dict) and fax_version.get("fax_version_no"):
        version_text = f" / FAX v{fax_version.get('fax_version_no')} of {fax_version.get('fax_version_count') or 1}"
    worksheet.row_dimensions[1].height = 18
    _safe_merge(worksheet, "B1:K1")
    header_cell = worksheet["B1"]
    header_cell.value = f"{label}: {_format_generated_at(generated_at)}{version_text}"
    header_cell.font = Font(name="Meiryo", size=8, bold=True)
    header_cell.alignment = Alignment(horizontal="center", vertical="center")
    header_cell.fill = PatternFill(fill_type=None)
    header_cell.border = _THIN_BORDER


def _apply_fax_markers(worksheet) -> None:
    worksheet.row_dimensions[_BOTTOM_MARKER_ROW].height = 18
    for cell_ref in ("A1", "L1", f"A{_BOTTOM_MARKER_ROW}", f"L{_BOTTOM_MARKER_ROW}"):
        cell = worksheet[cell_ref]
        cell.value = None
        cell.fill = _MARKER_FILL
        cell.border = _THIN_BORDER


def _apply_bottom_instruction_strip(worksheet, *, fax_template_id: str, base_label: str) -> None:
    _ = fax_template_id
    _ = base_label
    _safe_merge(worksheet, f"B{_BOTTOM_MARKER_ROW}:K{_BOTTOM_MARKER_ROW}")
    info_cell = worksheet[f"B{_BOTTOM_MARKER_ROW}"]
    info_cell.value = None
    info_cell.font = Font(name="Meiryo", size=8, bold=True)
    info_cell.alignment = Alignment(horizontal="center", vertical="center")
    info_cell.fill = PatternFill(fill_type=None)
    info_cell.border = _THIN_BORDER


def _extend_print_area(worksheet, *, bottom_row: int) -> None:
    print_area = worksheet.print_area
    if not print_area:
        worksheet.print_area = f"A1:L{bottom_row}"
        return
    area_ref = print_area.split("!", 1)[-1].replace("$", "")
    min_col, min_row, max_col, max_row = range_boundaries(area_ref)
    max_row = max(max_row, bottom_row)
    worksheet.print_area = (
        f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
    )


def _append_hidden_metadata_sheet(
    workbook: Workbook,
    *,
    source_workbook_name: str,
    facility_id: str,
    facility_name: str,
    fax_template_id: str,
    family_label: str,
    week_sheet_name: str,
    base_label: str,
    generated_at: datetime | None = None,
    fax_version: dict[str, Any] | None = None,
) -> None:
    if "設定" in workbook.sheetnames:
        del workbook["設定"]
    meta = workbook.create_sheet("設定")
    meta.sheet_state = "hidden"
    meta.append(["key", "value"])
    generated_at_value = generated_at or datetime.utcnow()
    meta.append(["generated_at_utc", generated_at_value.isoformat()])
    meta.append(["source_workbook", source_workbook_name])
    meta.append(["facility_id", facility_id])
    meta.append(["facility_name", facility_name])
    meta.append(["fax_template_id", fax_template_id])
    meta.append(["family_label", family_label])
    meta.append(["week_sheet_name", week_sheet_name])
    meta.append(["mode", base_label])
    if isinstance(fax_version, dict):
        for key in (
            "fax_version_no",
            "fax_version_count",
            "fax_version_document_id",
            "fax_version_message_id",
            "fax_version_received_at",
        ):
            if fax_version.get(key) not in (None, ""):
                meta.append([key, fax_version.get(key)])


def _safe_merge(worksheet, cell_range: str) -> None:
    if cell_range in {str(item) for item in worksheet.merged_cells.ranges}:
        return
    worksheet.merge_cells(cell_range)


def _sanitize_filename_fragment(value: str) -> str:
    text = str(value or "").strip()
    safe = re.sub(r"[^0-9A-Za-z_-]+", "_", text)
    return safe.strip("_") or "value"
