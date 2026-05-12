from __future__ import annotations

from collections import Counter, defaultdict
from dataclasses import asdict, dataclass
import json
import os
import re
import sys
import time
from pathlib import Path
from typing import Any
from urllib.parse import urlparse

import cv2
import numpy as np
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont

from src.services import config_service, hakodate_assignment_service
from src.services.storage_service import load_bytes_from_uri
from src.services.hakodate_fixed_quad_registration_service import (
    build_fixed_quad_template_registration,
    rectify_fax_to_template_grid,
    render_pdf_page_to_bgr,
    render_template_pdf_to_canvas,
    resolve_fixed_quad_px_for_manifest_item,
    resolve_template_axes_from_manifest_or_image,
)
from src.services.hakodate_step_review_pipeline_service import (
    TARGET_RULE,
    WEEK_SHEET_NAME,
    _align_axes,
    _bbox_quad_points,
    _bgr_to_rgb_image,
    _draw_merge_aware_grid,
    _draw_quad_points,
    _draw_target_regions,
    _make_review_canvas,
    _post_menu_target_regions,
    _source_template_name,
    _split_line_masks,
    _write_pdf_from_pages,
)


DEFAULT_CELL_OCR_ENGINE = "yomitoku_contact_sheet_batch"
_FULLWIDTH_DIGIT_TRANS = str.maketrans("０１２３４５６７８９，．", "0123456789,.")
_LOCAL_YOMITOKU_ANALYZER = None
_LOCAL_YOMITOKU_ANALYZER_KEY: tuple[str, bool] | None = None


@dataclass(frozen=True)
class HakodateCellOcrCaseResult:
    page: int
    facility_code: str
    order_id: str
    fax_pdf: str
    template_pdf: str
    source_template: str
    ocr_engine: str
    target_rule: str
    physical_region_count: int
    logical_assignment_count: int
    recognized_region_count: int
    recognized_assignment_count: int
    metrics: dict[str, Any]
    outputs: dict[str, str]


def _normalize_digits(value: object) -> str:
    text = str(value or "").strip().translate(_FULLWIDTH_DIGIT_TRANS)
    text = re.sub(r"[^0-9,.]", "", text)
    return text.strip(".,")


def _safe_int_box(
    box: list[float],
    *,
    width: int,
    height: int,
    margin_ratio: float,
) -> tuple[int, int, int, int] | None:
    if not isinstance(box, list) or len(box) != 4:
        return None
    try:
        x0, y0, x1, y1 = [float(value) for value in box]
    except Exception:
        return None
    if x1 <= x0 or y1 <= y0:
        return None
    margin_x = (x1 - x0) * max(0.0, min(margin_ratio, 0.35))
    margin_y = (y1 - y0) * max(0.0, min(margin_ratio, 0.35))
    ix0 = max(0, min(width, int(round(x0 + margin_x))))
    iy0 = max(0, min(height, int(round(y0 + margin_y))))
    ix1 = max(0, min(width, int(round(x1 - margin_x))))
    iy1 = max(0, min(height, int(round(y1 - margin_y))))
    if ix1 <= ix0 or iy1 <= iy0:
        return None
    return ix0, iy0, ix1, iy1


def _expanded_cell_box(
    box: list[float],
    *,
    width: int,
    height: int,
    pad_x_px: int,
    pad_y_px: int,
) -> tuple[int, int, int, int] | None:
    if not isinstance(box, list) or len(box) != 4:
        return None
    try:
        x0, y0, x1, y1 = [float(value) for value in box]
    except Exception:
        return None
    if x1 <= x0 or y1 <= y0:
        return None
    pad_x = max(0, min(int(pad_x_px), 24))
    pad_y = max(0, min(int(pad_y_px), 24))
    ix0 = max(0, min(width, int(np.floor(x0 - pad_x))))
    iy0 = max(0, min(height, int(np.floor(y0 - pad_y))))
    ix1 = max(0, min(width, int(np.ceil(x1 + pad_x))))
    iy1 = max(0, min(height, int(np.ceil(y1 + pad_y))))
    if ix1 <= ix0 or iy1 <= iy0:
        return None
    return ix0, iy0, ix1, iy1


def _erase_known_cell_frame(
    gray: np.ndarray,
    *,
    cell_box: list[float],
    crop_box: tuple[int, int, int, int],
) -> np.ndarray:
    crop_x0, crop_y0, _crop_x1, _crop_y1 = crop_box
    x0, y0, x1, y1 = [float(value) for value in cell_box]
    rel_x0 = int(round(x0 - crop_x0))
    rel_y0 = int(round(y0 - crop_y0))
    rel_x1 = int(round(x1 - crop_x0))
    rel_y1 = int(round(y1 - crop_y0))
    height, width = gray.shape[:2]
    thickness = max(4, int(round(min(max(1, x1 - x0), max(1, y1 - y0)) * 0.075)))
    erased = gray.copy()

    def fill_rect(rx0: int, ry0: int, rx1: int, ry1: int) -> None:
        ax0 = max(0, min(width, rx0))
        ay0 = max(0, min(height, ry0))
        ax1 = max(0, min(width, rx1))
        ay1 = max(0, min(height, ry1))
        if ax1 > ax0 and ay1 > ay0:
            erased[ay0:ay1, ax0:ax1] = 255

    fill_rect(rel_x0 - thickness, rel_y0 - thickness, rel_x1 + thickness, rel_y0 + thickness)
    fill_rect(rel_x0 - thickness, rel_y1 - thickness, rel_x1 + thickness, rel_y1 + thickness)
    fill_rect(rel_x0 - thickness, rel_y0 - thickness, rel_x0 + thickness, rel_y1 + thickness)
    fill_rect(rel_x1 - thickness, rel_y0 - thickness, rel_x1 + thickness, rel_y1 + thickness)
    return erased


def _remove_small_noise_only(gray: np.ndarray) -> np.ndarray:
    _threshold, binary = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    ink = 255 - binary
    num_labels, labels, stats, _centroids = cv2.connectedComponentsWithStats(ink, connectivity=8)
    cleaned = gray.copy()
    for label in range(1, num_labels):
        area = int(stats[label, cv2.CC_STAT_AREA])
        if area < 5:
            cleaned[labels == label] = 255
    return cleaned


def _preprocess_cell_crop(
    crop_bgr: np.ndarray,
    *,
    cell_box: list[float],
    crop_box: tuple[int, int, int, int],
) -> Image.Image:
    gray = cv2.cvtColor(crop_bgr, cv2.COLOR_BGR2GRAY) if crop_bgr.ndim == 3 else crop_bgr
    frame_removed = _erase_known_cell_frame(gray, cell_box=cell_box, crop_box=crop_box)
    cleaned = _remove_small_noise_only(frame_removed)
    return Image.fromarray(cleaned).convert("RGB")


def build_cell_contact_sheet(
    *,
    rectified_fax_bgr: np.ndarray,
    regions: list[dict[str, Any]],
    slot_width: int = 220,
    slot_height: int = 130,
    columns: int = 12,
    crop_pad_x_px: int = 1,
    crop_pad_y_px: int = 8,
) -> tuple[Image.Image, list[dict[str, Any]]]:
    height, width = rectified_fax_bgr.shape[:2]
    usable_regions: list[dict[str, Any]] = []
    row_count = max(1, (len(regions) + columns - 1) // columns)
    sheet = Image.new("RGB", (columns * slot_width, row_count * slot_height), "white")
    for slot_index, region in enumerate(regions):
        box = region.get("bbox")
        if not isinstance(box, list):
            continue
        px_box = _expanded_cell_box(
            box,
            width=width,
            height=height,
            pad_x_px=crop_pad_x_px,
            pad_y_px=crop_pad_y_px,
        )
        if not px_box:
            continue
        x0, y0, x1, y1 = px_box
        crop = rectified_fax_bgr[y0:y1, x0:x1]
        if crop.size == 0:
            continue
        crop_image = _preprocess_cell_crop(crop, cell_box=box, crop_box=px_box)
        max_w = slot_width - 12
        max_h = slot_height - 12
        scale = min(max_w / max(1, crop_image.width), max_h / max(1, crop_image.height))
        scale = max(0.2, min(scale, 5.0))
        crop_image = crop_image.resize(
            (
                max(1, int(round(crop_image.width * scale))),
                max(1, int(round(crop_image.height * scale))),
            ),
            Image.Resampling.BICUBIC,
        )
        slot_col = slot_index % columns
        slot_row = slot_index // columns
        slot_x = slot_col * slot_width
        slot_y = slot_row * slot_height
        paste_x = slot_x + (slot_width - crop_image.width) // 2
        paste_y = slot_y + (slot_height - crop_image.height) // 2
        sheet.paste(crop_image, (paste_x, paste_y))
        usable_regions.append(
            {
                **region,
                "ocr_contact_slot_index": slot_index,
                "ocr_contact_slot": [slot_x, slot_y, slot_x + slot_width, slot_y + slot_height],
                "ocr_contact_crop_box": [paste_x, paste_y, paste_x + crop_image.width, paste_y + crop_image.height],
                "ocr_cell_crop_bbox_px": [x0, y0, x1, y1],
            }
        )
    return sheet, usable_regions


def _workspace_root() -> Path:
    path = Path(__file__).resolve()
    for candidate in (path.parents[3], path.parents[2]):
        if (candidate / "tmp" / "hakodate_text_recognizer_trial_20260428").exists():
            return candidate
    for candidate in (path.parents[3], path.parents[2]):
        if (candidate / "tmp").exists() or (candidate / "backend").exists():
            return candidate
    return path.parents[2]


def _load_yomitoku_ocr_words():
    ocr_pipeline_root = _workspace_root() / "ocr_pipeline"
    if str(ocr_pipeline_root) not in sys.path:
        sys.path.insert(0, str(ocr_pipeline_root))
    try:
        from app.yomitoku_runner import ocr_image_words  # noqa: PLC0415

        return ocr_image_words
    except ModuleNotFoundError as exc:
        if exc.name not in {"app", "app.yomitoku_runner"}:
            raise
    return _local_yomitoku_ocr_words


def _coerce_jsonable(value: object) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, dict):
        return {str(key): _coerce_jsonable(item) for key, item in value.items()}
    if isinstance(value, (list, tuple)):
        return [_coerce_jsonable(item) for item in value]
    if hasattr(value, "tolist"):
        try:
            return _coerce_jsonable(value.tolist())
        except Exception:
            pass
    if hasattr(value, "item"):
        try:
            return value.item()
        except Exception:
            pass
    return str(value)


def _serialize_yomitoku_results(results: object) -> dict[str, Any] | None:
    if isinstance(results, dict):
        return _coerce_jsonable(results)
    for method_name, kwargs in (
        ("model_dump", {"mode": "python"}),
        ("model_dump", {}),
        ("to_dict", {}),
        ("dict", {}),
        ("to_json", {}),
        ("json", {}),
    ):
        method = getattr(results, method_name, None)
        if not callable(method):
            continue
        try:
            serialized = method(**kwargs)
        except TypeError:
            try:
                serialized = method()
            except Exception:
                continue
        except Exception:
            continue
        if isinstance(serialized, str):
            try:
                serialized = json.loads(serialized)
            except Exception:
                continue
        if isinstance(serialized, dict):
            return _coerce_jsonable(serialized)
    return None


def _normalize_yomitoku_box(box: object, *, width: int, height: int) -> list[float] | None:
    if not isinstance(box, (list, tuple)) or len(box) != 4:
        return None
    try:
        x0, y0, x1, y1 = [float(item) for item in box]
    except Exception:
        return None
    if max(abs(x0), abs(y0), abs(x1), abs(y1)) > 1.0:
        if width <= 0 or height <= 0:
            return None
        x0 /= float(width)
        x1 /= float(width)
        y0 /= float(height)
        y1 /= float(height)
    x0 = max(0.0, min(1.0, x0))
    x1 = max(0.0, min(1.0, x1))
    y0 = max(0.0, min(1.0, y0))
    y1 = max(0.0, min(1.0, y1))
    return [x0, y0, x1, y1]


def _center_from_yomitoku_points(points: object) -> tuple[float, float] | None:
    if not isinstance(points, list) or not points:
        return None
    xs: list[float] = []
    ys: list[float] = []
    for point in points:
        if not isinstance(point, (list, tuple)) or len(point) < 2:
            continue
        try:
            xs.append(float(point[0]))
            ys.append(float(point[1]))
        except Exception:
            continue
    if not xs or not ys:
        return None
    return sum(xs) / len(xs), sum(ys) / len(ys)


def _analysis_to_yomitoku_words(
    analysis: dict[str, Any] | None,
    *,
    width: int,
    height: int,
) -> list[dict[str, Any]]:
    if not isinstance(analysis, dict):
        return []
    raw_words = analysis.get("words")
    if not isinstance(raw_words, list):
        return []
    words: list[dict[str, Any]] = []
    for raw_word in raw_words:
        if not isinstance(raw_word, dict):
            continue
        text = str(raw_word.get("content") or raw_word.get("contents") or "").strip()
        if not text:
            continue
        box = _normalize_yomitoku_box(raw_word.get("box"), width=width, height=height)
        if box is not None:
            x_center = (box[0] + box[2]) / 2.0
            y_center = (box[1] + box[3]) / 2.0
        else:
            center = _center_from_yomitoku_points(raw_word.get("points"))
            if center is None or width <= 0 or height <= 0:
                continue
            x_center = center[0] / float(width)
            y_center = center[1] / float(height)
        words.append(
            {
                "text": text,
                "x": max(0.0, min(1.0, float(x_center))),
                "y": max(0.0, min(1.0, float(y_center))),
                "box": box,
            }
        )
    return words


def _get_local_yomitoku_analyzer(device: str, visualize: bool):
    global _LOCAL_YOMITOKU_ANALYZER  # noqa: PLW0603
    global _LOCAL_YOMITOKU_ANALYZER_KEY  # noqa: PLW0603
    key = (device, visualize)
    if _LOCAL_YOMITOKU_ANALYZER is not None and _LOCAL_YOMITOKU_ANALYZER_KEY == key:
        return _LOCAL_YOMITOKU_ANALYZER
    if not os.getenv("HF_HOME"):
        os.environ["HF_HOME"] = "/tmp/hf"
    from yomitoku import DocumentAnalyzer  # noqa: PLC0415

    configs = {
        "ocr": {
            "text_detector": {},
            "text_recognizer": {},
        },
        "layout_analyzer": {
            "layout_parser": {},
            "table_structure_recognizer": {},
        },
    }
    _LOCAL_YOMITOKU_ANALYZER = DocumentAnalyzer(configs=configs, device=device, visualize=visualize)
    _LOCAL_YOMITOKU_ANALYZER_KEY = key
    return _LOCAL_YOMITOKU_ANALYZER


def _local_yomitoku_ocr_words(image_bgr: np.ndarray, *, device: str) -> list[dict[str, Any]]:
    if image_bgr.ndim == 2:
        image_bgr = cv2.cvtColor(image_bgr, cv2.COLOR_GRAY2BGR)
    analyzer = _get_local_yomitoku_analyzer(device, False)
    results, _ocr_vis, _layout_vis = analyzer(image_bgr)
    analysis = _serialize_yomitoku_results(results)
    height, width = image_bgr.shape[:2]
    return _analysis_to_yomitoku_words(analysis, width=width, height=height)


def _words_to_region_text(words: list[dict[str, Any]]) -> str:
    if not words:
        return ""
    ordered = sorted(
        words,
        key=lambda item: (
            round(float(item.get("y") or 0.0), 4),
            round(float(item.get("x") or 0.0), 4),
        ),
    )
    lines: list[list[dict[str, Any]]] = []
    current: list[dict[str, Any]] = []
    current_y: float | None = None
    for word in ordered:
        y = float(word.get("y") or 0.0)
        if current_y is None or abs(y - current_y) <= 0.025:
            current.append(word)
            current_y = y if current_y is None else (current_y + y) / 2.0
            continue
        lines.append(current)
        current = [word]
        current_y = y
    if current:
        lines.append(current)
    rendered: list[str] = []
    for line in lines:
        line.sort(key=lambda item: float(item.get("x") or 0.0))
        text = " ".join(str(item.get("text") or "").strip() for item in line if str(item.get("text") or "").strip())
        if text:
            rendered.append(text)
    return "\n".join(rendered).strip()


def assign_yomitoku_words_to_contact_regions(
    *,
    words: list[dict[str, Any]],
    regions: list[dict[str, Any]],
    sheet_size: tuple[int, int],
) -> list[dict[str, Any]]:
    width, height = sheet_size
    by_region_id: dict[str, list[dict[str, Any]]] = {str(region.get("region_id")): [] for region in regions}
    for raw_word in words:
        text = str(raw_word.get("text") or "").strip()
        if not text:
            continue
        try:
            cx = float(raw_word.get("x") or 0.0) * width
            cy = float(raw_word.get("y") or 0.0) * height
        except Exception:
            continue
        for region in regions:
            slot = region.get("ocr_contact_slot")
            if not isinstance(slot, list) or len(slot) != 4:
                continue
            if float(slot[0]) <= cx <= float(slot[2]) and float(slot[1]) <= cy <= float(slot[3]):
                by_region_id.setdefault(str(region.get("region_id")), []).append(
                    {
                        "text": text,
                        "x": cx,
                        "y": cy,
                        "normalized_digits": _normalize_digits(text),
                        "source_word": raw_word,
                    }
                )
                break
    assigned: list[dict[str, Any]] = []
    for region in regions:
        region_words = by_region_id.get(str(region.get("region_id"))) or []
        text = _words_to_region_text(region_words)
        assigned.append(
            {
                **region,
                "ocr_text": text,
                "ocr_normalized": _normalize_digits(text),
                "ocr_words": region_words,
                "ocr_word_count": len(region_words),
            }
        )
    return assigned


def sheet_assignments_from_ocr_regions(regions: list[dict[str, Any]]) -> list[dict[str, Any]]:
    assignments: list[dict[str, Any]] = []
    for region in regions:
        logical_targets = region.get("logical_targets") or []
        if not isinstance(logical_targets, list):
            logical_targets = []
        for target in logical_targets:
            if not isinstance(target, dict):
                continue
            assignments.append(
                {
                    "sheet_cell": target.get("sheet_cell"),
                    "worksheet_row": target.get("worksheet_row"),
                    "worksheet_col": target.get("worksheet_col"),
                    "grid_row_index": target.get("grid_row_index"),
                    "grid_col_index": target.get("grid_col_index"),
                    "field": target.get("field"),
                    "field_label": target.get("field_label"),
                    "date": target.get("date"),
                    "daypart": target.get("daypart"),
                    "menu_name": target.get("menu_name"),
                    "value_text": region.get("ocr_text") or "",
                    "value_normalized": region.get("ocr_normalized") or "",
                    "source_region_id": region.get("region_id"),
                    "source_region_bbox": region.get("bbox"),
                    "source_region_crop_bbox_px": region.get("ocr_cell_crop_bbox_px"),
                    "merged_cell": region.get("merged_cell"),
                }
            )
    return assignments


def sheet_value_grid_from_assignments(assignments: list[dict[str, Any]]) -> dict[str, Any]:
    cells: dict[str, dict[str, Any]] = {}
    row_indexes: set[int] = set()
    col_indexes: set[int] = set()
    for assignment in assignments:
        sheet_cell = str(assignment.get("sheet_cell") or "").strip()
        if not sheet_cell:
            continue
        try:
            worksheet_row = int(assignment.get("worksheet_row") or 0)
            worksheet_col = int(assignment.get("worksheet_col") or 0)
        except Exception:
            continue
        if worksheet_row <= 0 or worksheet_col <= 0:
            continue
        row_indexes.add(worksheet_row)
        col_indexes.add(worksheet_col)
        cells[sheet_cell] = {
            "sheet_cell": sheet_cell,
            "worksheet_row": worksheet_row,
            "worksheet_col": worksheet_col,
            "column_letter": get_column_letter(worksheet_col),
            "value_text": assignment.get("value_text") or "",
            "value_normalized": assignment.get("value_normalized") or "",
            "field": assignment.get("field"),
            "field_label": assignment.get("field_label"),
            "date": assignment.get("date"),
            "daypart": assignment.get("daypart"),
            "menu_name": assignment.get("menu_name"),
            "source_region_id": assignment.get("source_region_id"),
            "merged_cell": assignment.get("merged_cell"),
        }
    sorted_cols = sorted(col_indexes)
    rows: list[dict[str, Any]] = []
    for row_index in sorted(row_indexes):
        row_values: dict[str, str] = {}
        row_cells: dict[str, dict[str, Any]] = {}
        for col_index in sorted_cols:
            col_letter = get_column_letter(col_index)
            sheet_cell = f"{col_letter}{row_index}"
            cell = cells.get(sheet_cell)
            row_values[col_letter] = str((cell or {}).get("value_text") or "")
            if cell is not None:
                row_cells[col_letter] = cell
        rows.append(
            {
                "worksheet_row": row_index,
                "values_by_column": row_values,
                "cells_by_column": row_cells,
            }
        )
    return {
        "cells": cells,
        "rows": rows,
        "columns": [get_column_letter(index) for index in sorted_cols],
    }


def validate_cell_ocr_mapping(
    *,
    ocr_regions: list[dict[str, Any]],
    sheet_assignments: list[dict[str, Any]],
    sheet_values: dict[str, Any],
) -> dict[str, Any]:
    errors: list[str] = []
    region_by_id = {str(region.get("region_id")): region for region in ocr_regions}
    assignments_by_region: dict[str, list[dict[str, Any]]] = defaultdict(list)
    sheet_cell_counts: Counter[str] = Counter()
    for assignment in sheet_assignments:
        region_id = str(assignment.get("source_region_id"))
        assignments_by_region[region_id].append(assignment)
        sheet_cell = str(assignment.get("sheet_cell") or "").strip()
        if sheet_cell:
            sheet_cell_counts[sheet_cell] += 1

    cells = sheet_values.get("cells") if isinstance(sheet_values, dict) else {}
    if not isinstance(cells, dict):
        errors.append("sheet_values.cells is missing")
        cells = {}

    duplicate_cells = sorted(cell for cell, count in sheet_cell_counts.items() if count != 1)
    if duplicate_cells:
        errors.append(f"duplicate or missing sheet assignment uniqueness: {duplicate_cells[:12]}")

    if len(cells) != len(sheet_assignments):
        errors.append(f"sheet cell count mismatch: cells={len(cells)} assignments={len(sheet_assignments)}")

    unknown_region_ids = sorted(
        region_id for region_id in assignments_by_region.keys() if region_id not in region_by_id
    )
    if unknown_region_ids:
        errors.append(f"assignments reference unknown regions: {unknown_region_ids[:12]}")

    missing_region_ids = sorted(
        region_id for region_id in region_by_id.keys() if region_id not in assignments_by_region
    )
    if missing_region_ids:
        errors.append(f"regions without assignments: {missing_region_ids[:12]}")

    for region_id, region in region_by_id.items():
        expected_targets = region.get("logical_targets") or []
        if not isinstance(expected_targets, list):
            expected_targets = []
        actual_assignments = assignments_by_region.get(region_id) or []
        if len(actual_assignments) != len(expected_targets):
            errors.append(
                f"{region_id}: logical target count mismatch "
                f"expected={len(expected_targets)} actual={len(actual_assignments)}"
            )

        region_text = str(region.get("ocr_text") or "")
        region_normalized = str(region.get("ocr_normalized") or "")
        region_words = region.get("ocr_words") or []
        if not isinstance(region_words, list):
            errors.append(f"{region_id}: ocr_words is not a list")
            region_words = []
        reparsed_text = _words_to_region_text(region_words)
        reparsed_normalized = _normalize_digits(reparsed_text)
        if reparsed_text != region_text:
            errors.append(f"{region_id}: raw OCR words do not reparse to ocr_text")
        if reparsed_normalized != region_normalized:
            errors.append(f"{region_id}: raw OCR words do not reparse to ocr_normalized")
        if int(region.get("ocr_word_count") or 0) != len(region_words):
            errors.append(f"{region_id}: ocr_word_count does not match ocr_words length")
        for assignment in actual_assignments:
            sheet_cell = str(assignment.get("sheet_cell") or "").strip()
            if not sheet_cell:
                errors.append(f"{region_id}: assignment missing sheet_cell")
                continue
            if str(assignment.get("value_text") or "") != region_text:
                errors.append(f"{region_id}/{sheet_cell}: value_text does not match region OCR text")
            if str(assignment.get("value_normalized") or "") != region_normalized:
                errors.append(f"{region_id}/{sheet_cell}: value_normalized does not match region OCR normalized")
            cell = cells.get(sheet_cell)
            if cell is None:
                errors.append(f"{region_id}/{sheet_cell}: missing from sheet_values.cells")
                continue
            if str(cell.get("value_text") or "") != region_text:
                errors.append(f"{region_id}/{sheet_cell}: sheet_values value_text does not match assignment")
            if str(cell.get("value_normalized") or "") != region_normalized:
                errors.append(f"{region_id}/{sheet_cell}: sheet_values value_normalized does not match assignment")

        slot = region.get("ocr_contact_slot")
        if isinstance(slot, list) and len(slot) == 4:
            sx0, sy0, sx1, sy1 = [float(value) for value in slot]
            for word in region_words:
                try:
                    wx = float(word.get("x") or 0.0)
                    wy = float(word.get("y") or 0.0)
                except Exception:
                    errors.append(f"{region_id}: OCR word has invalid coordinates")
                    continue
                if not (sx0 <= wx <= sx1 and sy0 <= wy <= sy1):
                    errors.append(f"{region_id}: OCR word lies outside assigned contact slot")

    return {
        "ok": not errors,
        "error_count": len(errors),
        "errors": errors[:50],
        "region_count": len(ocr_regions),
        "assignment_count": len(sheet_assignments),
        "sheet_cell_count": len(cells),
        "assigned_word_count": sum(len(region.get("ocr_words") or []) for region in ocr_regions),
        "recognized_region_count": sum(1 for region in ocr_regions if str(region.get("ocr_text") or "").strip()),
        "recognized_assignment_count": sum(
            1 for assignment in sheet_assignments if str(assignment.get("value_text") or "").strip()
        ),
    }


def _load_overlay_font(size: int = 18) -> ImageFont.FreeTypeFont | ImageFont.ImageFont:
    candidates = [
        "/System/Library/Fonts/ヒラギノ角ゴシック W3.ttc",
        "/System/Library/Fonts/ヒラギノ角ゴシック W4.ttc",
        "/Library/Fonts/Arial Unicode.ttf",
        "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
    ]
    for path in candidates:
        try:
            if Path(path).exists():
                return ImageFont.truetype(path, size=size)
        except Exception:
            continue
    return ImageFont.load_default()


def _draw_text_safe(
    draw: ImageDraw.ImageDraw,
    xy: tuple[int, int],
    text: str,
    *,
    fill: tuple[int, int, int, int],
    font: ImageFont.FreeTypeFont | ImageFont.ImageFont,
) -> None:
    try:
        draw.text(xy, text, fill=fill, font=font)
    except Exception:
        fallback = text.encode("ascii", "ignore").decode("ascii") or "?"
        draw.text(xy, fallback, fill=fill, font=font)


def draw_ocr_results_overlay(
    *,
    target_overlay: Image.Image,
    ocr_regions: list[dict[str, Any]],
) -> Image.Image:
    image = target_overlay.convert("RGBA")
    layer = Image.new("RGBA", image.size, (0, 0, 0, 0))
    draw = ImageDraw.Draw(layer)
    font = _load_overlay_font(20)
    for region in ocr_regions:
        value = str(region.get("ocr_normalized") or region.get("ocr_text") or "").strip()
        if not value:
            continue
        box = region.get("bbox")
        if not isinstance(box, list) or len(box) != 4:
            continue
        x0, y0, x1, y1 = [int(round(float(value))) for value in box]
        cx = int(round((x0 + x1) / 2.0))
        cy = int(round((y0 + y1) / 2.0))
        label = value[:12]
        try:
            text_box = draw.textbbox((0, 0), label, font=font)
            text_w = max(1, text_box[2] - text_box[0])
            text_h = max(1, text_box[3] - text_box[1])
        except Exception:
            text_w = 28
            text_h = 20
        tx0 = min(max(cx + 12, 0), max(0, image.width - text_w - 1))
        ty0 = min(max(cy - text_h - 8, 0), max(0, image.height - text_h - 1))
        _draw_text_safe(draw, (tx0, ty0), label, fill=(220, 0, 0, 255), font=font)
    return Image.alpha_composite(image, layer).convert("RGB")


def _accepted_header_intersection_points(axis_evidence: dict[str, Any] | None) -> list[dict[str, Any]]:
    if not isinstance(axis_evidence, dict):
        return []
    match = axis_evidence.get("header_intersection_x_match")
    if not isinstance(match, dict):
        return []
    points = match.get("header_intersection_points")
    if not isinstance(points, list):
        return []
    if not match.get("used"):
        return [point for point in points if isinstance(point, dict)]
    x_clusters = match.get("fax_x_clusters")
    y_clusters = match.get("fax_y_clusters")
    if not isinstance(x_clusters, list) or not isinstance(y_clusters, list):
        return [point for point in points if isinstance(point, dict)]

    accepted_x_point_indexes: set[int] = set()
    for cluster in x_clusters:
        if not isinstance(cluster, dict):
            continue
        for point_index in cluster.get("point_indexes") or []:
            try:
                accepted_x_point_indexes.add(int(point_index))
            except Exception:
                continue

    accepted_y_point_indexes: set[int] = set()
    for cluster in y_clusters:
        if not isinstance(cluster, dict):
            continue
        for point_index in cluster.get("point_indexes") or []:
            try:
                accepted_y_point_indexes.add(int(point_index))
            except Exception:
                continue

    accepted: list[dict[str, Any]] = []
    for index in sorted(accepted_x_point_indexes & accepted_y_point_indexes):
        if 0 <= index < len(points) and isinstance(points[index], dict):
            accepted.append(points[index])
    return accepted


def _draw_header_intersections_overlay(
    *,
    image: Image.Image,
    axis_evidence: dict[str, Any] | None,
) -> Image.Image:
    points = _accepted_header_intersection_points(axis_evidence)
    if not points:
        return image
    base = image.convert("RGBA")
    layer = Image.new("RGBA", base.size, (0, 0, 0, 0))
    draw = ImageDraw.Draw(layer)
    for point in points:
        try:
            x = int(round(float(point.get("x"))))
            y = int(round(float(point.get("y"))))
        except Exception:
            continue
        if not (0 <= x < base.width and 0 <= y < base.height):
            continue
        draw.ellipse((x - 7, y - 7, x + 7, y + 7), outline=(255, 145, 0, 255), width=3)
        draw.ellipse((x - 2, y - 2, x + 2, y + 2), fill=(255, 145, 0, 255))
    return Image.alpha_composite(base, layer).convert("RGB")


def _bgr_from_pil(image: Image.Image) -> np.ndarray:
    rgb = np.array(image.convert("RGB"))
    return cv2.cvtColor(rgb, cv2.COLOR_RGB2BGR)


def _run_yomitoku_contact_sheet(contact_sheet: Image.Image, *, device: str) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    ocr_image_words = _load_yomitoku_ocr_words()
    t0 = time.perf_counter()
    words = ocr_image_words(_bgr_from_pil(contact_sheet), device=device)
    elapsed = time.perf_counter() - t0
    return words, {
        "engine": DEFAULT_CELL_OCR_ENGINE,
        "device": device,
        "ocr_seconds": round(float(elapsed), 3),
        "raw_word_count": len(words),
    }


def build_hakodate_best_method_for_manifest_item(
    *,
    item: dict[str, Any],
    page: int,
    draft_sheet: dict[str, Any],
    output_dir: Path,
    render_width: int = 1864,
) -> tuple[HakodateCellOcrCaseResult, Image.Image]:
    """Production-compatible entry point for the accepted Hakodate cell OCR pipeline.

    Keep the live entry point on the same accepted runtime that produced the
    locally reviewed best_method overlay/records artifacts. Path materialization
    is handled before dispatch; the accepted runtime logic itself is unchanged.
    """
    runtime_item = _materialize_best_method_item_paths(item=item, output_dir=output_dir)
    from src.hakodate_best_method_runtime.render_best_method_overlay_all_facilities import (
        build_best_method_for_manifest_item,
    )

    summary, review_page = build_best_method_for_manifest_item(
        item=runtime_item,
        page_index=page,
        draft_sheet=draft_sheet,
        output_dir=output_dir,
        render_width=render_width,
    )
    outputs = {
        key: str(value)
        for key, value in (summary.get("outputs") or {}).items()
        if value is not None
    }
    metrics = dict(summary.get("metrics") or {})
    result = HakodateCellOcrCaseResult(
        page=page,
        facility_code=str(runtime_item.get("facility_code") or ""),
        order_id=str(runtime_item.get("order_id") or ""),
        fax_pdf=str(runtime_item.get("fax_pdf") or ""),
        template_pdf=str(runtime_item.get("template_pdf") or ""),
        source_template=str(summary.get("source_template") or ""),
        ocr_engine=str(summary.get("engine") or "opencv_knn_leave_one_out_k5"),
        target_rule=TARGET_RULE,
        physical_region_count=int(metrics.get("numeric_eval_cell_count") or 0),
        logical_assignment_count=int(metrics.get("numeric_eval_cell_count") or 0),
        recognized_region_count=int(metrics.get("pred_nonempty_count") or 0),
        recognized_assignment_count=int(metrics.get("pred_nonempty_count") or 0),
        metrics=metrics,
        outputs=outputs,
    )
    return result, review_page


def _materialize_best_method_item_paths(*, item: dict[str, Any], output_dir: Path) -> dict[str, Any]:
    """Resolve deployment/storage paths only; keep the accepted OCR runtime logic untouched."""
    resolved = dict(item)
    materialized_dir = output_dir / "_inputs"
    for key, suffix in (("fax_pdf", ".pdf"), ("template_pdf", ".pdf"), ("step2_png", ".png")):
        raw = str(resolved.get(key) or "").strip()
        if not raw or Path(raw).exists():
            continue
        parsed = urlparse(raw)
        if parsed.scheme not in {"gs", "file"}:
            continue
        materialized_dir.mkdir(parents=True, exist_ok=True)
        target = materialized_dir / f"{key}{suffix}"
        target.write_bytes(load_bytes_from_uri(raw))
        resolved[key] = str(target)
    return resolved


def _fax_template_for_manifest_item(item: dict[str, Any], facility_code: str) -> dict[str, Any] | None:
    embedded = item.get("fax_template")
    if isinstance(embedded, dict):
        return embedded
    template_id = (
        str(
            item.get("fax_template_id")
            or item.get("template_id")
            or item.get("resolved_template_id")
            or ""
        ).strip()
        or None
    )
    try:
        facility_config = config_service.get_facility_config_for_template(facility_code, template_id)
    except Exception:
        facility_config = None
    fax_template = facility_config.get("fax_template") if isinstance(facility_config, dict) else None
    return fax_template if isinstance(fax_template, dict) else None


def _build_preprocess_for_ocr(
    *,
    item: dict[str, Any],
    page: int,
    render_width: int,
) -> dict[str, Any]:
    facility_code = str(item["facility_code"])
    order_id = str(item["order_id"])
    existing_step2 = cv2.imread(item["step2_png"])
    if existing_step2 is None:
        raise ValueError(f"step2 canvas not found: {item['step2_png']}")
    canvas_height, canvas_width = existing_step2.shape[:2]
    template = render_template_pdf_to_canvas(item["template_pdf"], width=canvas_width, height=canvas_height)
    template_xs, template_ys, _all_xs, _all_ys = resolve_template_axes_from_manifest_or_image(
        item=item,
        template_image=template,
        manifest_template_bbox=item["template_bbox"],
    )
    week_sheet_name = str(item.get("week_sheet_name") or WEEK_SHEET_NAME).strip() or WEEK_SHEET_NAME
    worksheet = hakodate_assignment_service._worksheet_for_manifest_structure_template(  # noqa: SLF001
        item=item,
        facility_id=facility_code,
        week_sheet_name=week_sheet_name,
    )
    fax_template = _fax_template_for_manifest_item(item, facility_code)
    quad_px, quad_source, quad_estimate = resolve_fixed_quad_px_for_manifest_item(item)
    registration, _step_images_np = build_fixed_quad_template_registration(
        facility_code=facility_code,
        order_id=order_id,
        fax_pdf=item["fax_pdf"],
        template_pdf=item["template_pdf"],
        quad_px=quad_px,
        manifest_template_bbox=item["template_bbox"],
        canvas_width=canvas_width,
        canvas_height=canvas_height,
        render_width=render_width,
        quad_source=quad_source,
        output_dir=None,
        template_axes_x=template_xs,
        template_axes_y=template_ys,
    )
    original = render_pdf_page_to_bgr(item["fax_pdf"], width=render_width)
    table_bbox = registration.template_outer_grid_bbox_used
    raw_rectified = rectify_fax_to_template_grid(
        original,
        quad_px=quad_px,
        table_bbox=table_bbox,
        canvas_width=canvas_width,
        canvas_height=canvas_height,
    )
    horizontal_line_mask, _vertical_line_mask = _split_line_masks(raw_rectified)
    aligned_xs, aligned_ys, axis_evidence, _axis_match_image = _align_axes(
        rectified_fax=raw_rectified,
        template_xs=template_xs,
        template_ys=template_ys,
        worksheet=worksheet,
        fax_template=fax_template,
    )
    grid_overlay, merge_evidence = _draw_merge_aware_grid(
        worksheet=worksheet,
        rectified_fax=raw_rectified,
        xs=aligned_xs,
        ys=aligned_ys,
        horizontal_line_mask=horizontal_line_mask,
    )
    target_regions, target_evidence = _post_menu_target_regions(
        worksheet=worksheet,
        column_edges=[float(value) for value in aligned_xs],
        row_edges=[float(value) for value in aligned_ys],
        fax_template=fax_template,
        horizontal_line_mask=horizontal_line_mask,
    )
    target_overlay = _draw_target_regions(grid_overlay=grid_overlay, regions=target_regions)
    target_overlay = _draw_header_intersections_overlay(image=target_overlay, axis_evidence=axis_evidence)
    rectified_quad_points = _bbox_quad_points(table_bbox)
    return {
        "page": page,
        "facility_code": facility_code,
        "order_id": order_id,
        "worksheet": worksheet,
        "raw_rectified": raw_rectified,
        "target_regions": target_regions,
        "rectified_quad_points": rectified_quad_points,
        "template_outer_grid_bbox_used": table_bbox,
        "target_overlay": _draw_quad_points(target_overlay, rectified_quad_points, prefix="Q"),
        "source_template": _source_template_name(facility_code),
        "axis_evidence": {
            **axis_evidence,
            "merge": merge_evidence,
            "target": target_evidence,
            "quad_estimate": quad_estimate,
        },
    }


def build_hakodate_cell_ocr_for_manifest_item(
    *,
    item: dict[str, Any],
    page: int,
    output_dir: Path,
    render_width: int = 1864,
    device: str = "cpu",
) -> tuple[HakodateCellOcrCaseResult, Image.Image]:
    t0 = time.perf_counter()
    pre = _build_preprocess_for_ocr(item=item, page=page, render_width=render_width)
    facility_code = str(pre["facility_code"])
    order_id = str(pre["order_id"])
    case_dir = output_dir / f"{page:02d}_{facility_code}_{order_id}"
    case_dir.mkdir(parents=True, exist_ok=True)
    contact_sheet, usable_regions = build_cell_contact_sheet(
        rectified_fax_bgr=pre["raw_rectified"],
        regions=pre["target_regions"],
    )
    words, ocr_metrics = _run_yomitoku_contact_sheet(contact_sheet, device=device)
    ocr_regions = assign_yomitoku_words_to_contact_regions(
        words=words,
        regions=usable_regions,
        sheet_size=contact_sheet.size,
    )
    sheet_assignments = sheet_assignments_from_ocr_regions(ocr_regions)
    sheet_values = sheet_value_grid_from_assignments(sheet_assignments)
    parse_validation = validate_cell_ocr_mapping(
        ocr_regions=ocr_regions,
        sheet_assignments=sheet_assignments,
        sheet_values=sheet_values,
    )
    recognized_regions = [region for region in ocr_regions if str(region.get("ocr_text") or "").strip()]
    recognized_assignments = [
        assignment for assignment in sheet_assignments if str(assignment.get("value_text") or "").strip()
    ]
    ocr_overlay = draw_ocr_results_overlay(
        target_overlay=pre["target_overlay"],
        ocr_regions=ocr_regions,
    )
    total_seconds = time.perf_counter() - t0
    metrics = {
        **ocr_metrics,
        "total_seconds": round(float(total_seconds), 3),
        "physical_region_count": len(ocr_regions),
        "logical_assignment_count": len(sheet_assignments),
        "recognized_region_count": len(recognized_regions),
        "recognized_assignment_count": len(recognized_assignments),
        "cells_per_ocr_second": round(len(ocr_regions) / max(float(ocr_metrics["ocr_seconds"]), 1e-6), 3),
        "parse_validation": parse_validation,
    }
    if not parse_validation["ok"]:
        raise ValueError(f"OCR parse validation failed for {facility_code} {order_id}: {parse_validation['errors']}")
    contact_sheet_path = case_dir / "cell_contact_sheet.png"
    overlay_path = case_dir / "cell_ocr_overlay.png"
    regions_path = case_dir / "cell_ocr_regions.json"
    assignments_path = case_dir / "cell_ocr_sheet_assignments.json"
    sheet_values_path = case_dir / "cell_ocr_sheet_values.json"
    contact_sheet.save(contact_sheet_path)
    ocr_overlay.save(overlay_path)
    regions_path.write_text(json.dumps(ocr_regions, ensure_ascii=False, indent=2), encoding="utf-8")
    assignments_path.write_text(json.dumps(sheet_assignments, ensure_ascii=False, indent=2), encoding="utf-8")
    sheet_values_path.write_text(json.dumps(sheet_values, ensure_ascii=False, indent=2), encoding="utf-8")
    details = [
        f"source_template={pre['source_template']}",
        f"target_rule={TARGET_RULE}",
        (
            f"engine={DEFAULT_CELL_OCR_ENGINE} device={device} "
            f"regions={len(ocr_regions)} logical={len(sheet_assignments)} "
            f"recognized={len(recognized_regions)} logical_recognized={len(recognized_assignments)} "
            f"ocr_sec={ocr_metrics['ocr_seconds']}"
        ),
        "red centers: OCR target cells / red digit labels: OCR result assigned by contact-sheet slot",
    ]
    review_page = _make_review_canvas(
        title="Hakodate cell OCR result on target red points",
        facility_code=facility_code,
        order_id=order_id,
        image=ocr_overlay,
        details=details,
    )
    review_page_path = case_dir / "cell_ocr_review_page.png"
    review_page.save(review_page_path)
    result = HakodateCellOcrCaseResult(
        page=page,
        facility_code=facility_code,
        order_id=order_id,
        fax_pdf=str(item["fax_pdf"]),
        template_pdf=str(item["template_pdf"]),
        source_template=str(pre["source_template"]),
        ocr_engine=DEFAULT_CELL_OCR_ENGINE,
        target_rule=TARGET_RULE,
        physical_region_count=len(ocr_regions),
        logical_assignment_count=len(sheet_assignments),
        recognized_region_count=len(recognized_regions),
        recognized_assignment_count=len(recognized_assignments),
        metrics=metrics,
        outputs={
            "contact_sheet": str(contact_sheet_path),
            "overlay": str(overlay_path),
            "review_page": str(review_page_path),
            "ocr_regions": str(regions_path),
            "sheet_assignments": str(assignments_path),
            "sheet_values": str(sheet_values_path),
        },
    )
    return result, review_page


def build_all_facility_hakodate_cell_ocr_pdf(
    *,
    manifest_path: Path,
    output_dir: Path,
    render_width: int = 1864,
    device: str = "cpu",
) -> dict[str, Any]:
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    items = manifest.get("results") if isinstance(manifest, dict) else manifest
    if not isinstance(items, list):
        raise ValueError("manifest results are missing")
    output_dir.mkdir(parents=True, exist_ok=True)
    pages: list[Image.Image] = []
    results: list[dict[str, Any]] = []
    t0 = time.perf_counter()
    for page, item in enumerate(items, start=1):
        result, review_page = build_hakodate_cell_ocr_for_manifest_item(
            item=item,
            page=page,
            output_dir=output_dir,
            render_width=render_width,
            device=device,
        )
        pages.append(review_page)
        results.append(asdict(result))
    pdf_path = output_dir / "cell_ocr_results_on_red_points_all14.pdf"
    _write_pdf_from_pages(pages, pdf_path)
    summary = {
        "count": len(results),
        "ocr_engine": DEFAULT_CELL_OCR_ENGINE,
        "target_rule": TARGET_RULE,
        "total_elapsed_seconds": round(float(time.perf_counter() - t0), 3),
        "total_physical_region_count": sum(int(item["physical_region_count"]) for item in results),
        "total_logical_assignment_count": sum(int(item["logical_assignment_count"]) for item in results),
        "total_recognized_region_count": sum(int(item["recognized_region_count"]) for item in results),
        "total_recognized_assignment_count": sum(int(item["recognized_assignment_count"]) for item in results),
        "total_parse_validation_error_count": sum(
            int(item["metrics"]["parse_validation"]["error_count"]) for item in results
        ),
        "pdf": str(pdf_path),
        "results": results,
    }
    summary_path = output_dir / "cell_ocr_summary.json"
    summary_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    return summary
