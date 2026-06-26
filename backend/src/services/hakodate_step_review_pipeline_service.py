from __future__ import annotations

from dataclasses import asdict, dataclass
import itertools
import json
from pathlib import Path
from typing import Any

import cv2
import numpy as np
from openpyxl.utils import get_column_letter
from openpyxl.utils import range_boundaries
from PIL import Image, ImageDraw, ImageFont

from src.services import config_service, hakodate_assignment_service
from src.services.hakodate_fixed_quad_registration_service import (
    build_fixed_quad_template_registration,
    render_pdf_page_to_bgr,
    render_template_pdf_to_canvas,
    rectify_fax_to_template_grid,
    resolve_fixed_quad_px_for_manifest_item,
    resolve_template_axes_from_manifest_or_image,
)


WEEK_SHEET_NAME = "4月26日～4月30日"
TARGET_RULE = "all columns strictly right of menu(献立)"
STEP_REVIEW_BODY_START_ROW = 11
STEP_REVIEW_BODY_END_ROW = 67
STEP_REVIEW_HEADER_BANDS = 2


@dataclass(frozen=True)
class HakodateStepReviewResult:
    page: int
    facility_code: str
    order_id: str
    fax_pdf: str
    template_pdf: str
    source_template: str
    target_rule: str
    target_worksheet_cols: list[int]
    region_count: int
    logical_target_count: int
    label_counts: dict[str, int]
    axis_evidence: dict[str, Any]
    outputs: dict[str, str]


def _source_template_name(facility_id: str) -> str:
    facility = config_service.get_facility_config(facility_id)
    if not facility:
        return ""
    template = facility.get("fax_template") if isinstance(facility.get("fax_template"), dict) else {}
    return str(facility.get("fax_template_id") or template.get("template_id") or "").strip()


def _bgr_to_rgb_image(image: np.ndarray) -> Image.Image:
    return Image.fromarray(cv2.cvtColor(image, cv2.COLOR_BGR2RGB)).convert("RGB")


def _write_pdf_from_pages(pages: list[Image.Image], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if pages:
        pages[0].save(path, save_all=True, append_images=pages[1:])


def _make_review_canvas(
    *,
    title: str,
    facility_code: str,
    order_id: str,
    image: Image.Image,
    details: list[str] | None = None,
    target_width: int = 1600,
) -> Image.Image:
    scale = min(1.0, float(target_width) / max(1, image.width))
    if scale < 1.0:
        image = image.resize(
            (int(round(image.width * scale)), int(round(image.height * scale))),
            Image.Resampling.BICUBIC,
        )
    header_h = 110 + 18 * len(details or [])
    canvas = Image.new("RGB", (image.width + 40, image.height + header_h), "white")
    draw = ImageDraw.Draw(canvas)
    font = ImageFont.load_default()
    draw.text((20, 14), f"{title}  {facility_code} {order_id}", fill=(0, 0, 0), font=font)
    y = 42
    for detail in details or []:
        draw.text((20, y), detail, fill=(0, 0, 0), font=font)
        y += 18
    canvas.paste(image, (20, header_h - 20))
    return canvas


def _bbox_quad_points(bbox: list[float] | list[int]) -> list[tuple[float, float]]:
    x0, y0, x1, y1 = [float(value) for value in bbox]
    return [(x0, y0), (x1, y0), (x1, y1), (x0, y1)]


def _draw_quad_points(
    image: Image.Image,
    points: list[tuple[float, float]] | list[list[float]],
    *,
    prefix: str,
) -> Image.Image:
    out = image.convert("RGBA")
    layer = Image.new("RGBA", out.size, (0, 0, 0, 0))
    draw = ImageDraw.Draw(layer)
    colors = [(255, 0, 255, 230), (0, 80, 255, 230), (0, 170, 0, 230), (255, 140, 0, 230)]
    labels = ["TL", "TR", "BR", "BL"]
    pts = [(float(point[0]), float(point[1])) for point in points]
    if len(pts) == 4:
        draw.line([*pts, pts[0]], fill=(255, 0, 255, 180), width=4)
    font = ImageFont.load_default()
    for point, label, color in zip(pts, labels, colors):
        x, y = point
        draw.ellipse((x - 12, y - 12, x + 12, y + 12), outline=color, width=4)
        draw.ellipse((x - 4, y - 4, x + 4, y + 4), fill=color)
        draw.text((x + 14, y + 12), f"{prefix}{label}", fill=color, font=font)
    return Image.alpha_composite(out, layer).convert("RGB")


def _split_line_masks(image_bgr: np.ndarray) -> tuple[np.ndarray, np.ndarray]:
    gray = cv2.cvtColor(image_bgr, cv2.COLOR_BGR2GRAY)
    gray = cv2.fastNlMeansDenoising(gray, None, h=7, templateWindowSize=7, searchWindowSize=21)
    _thr, inv = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)
    height, width = inv.shape[:2]
    h_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (max(18, width // 55), 1))
    v_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1, max(18, height // 75)))
    h_lines = cv2.morphologyEx(inv, cv2.MORPH_OPEN, h_kernel, iterations=1)
    v_lines = cv2.morphologyEx(inv, cv2.MORPH_OPEN, v_kernel, iterations=1)
    return h_lines, v_lines


def dewarp_rectified_y_to_template_rows(
    rectified: np.ndarray,
    *,
    source_ys: list[float],
    template_ys: list[int | float],
) -> tuple[np.ndarray, dict[str, Any]]:
    if len(source_ys) != len(template_ys) or len(source_ys) < 3:
        return rectified, {
            "applied": False,
            "reason": "row_dewarp_axis_count_mismatch",
            "source_count": len(source_ys),
            "template_count": len(template_ys),
        }
    source = np.array([float(value) for value in source_ys], dtype=np.float32)
    target = np.array([float(value) for value in template_ys], dtype=np.float32)
    if not np.all(np.diff(source) > 0) or not np.all(np.diff(target) > 0):
        return rectified, {"applied": False, "reason": "row_dewarp_axes_not_monotonic"}
    source_heights = np.diff(source)
    target_heights = np.diff(target)
    if float(np.min(source_heights)) <= 4.0 or float(np.min(target_heights)) <= 4.0:
        return rectified, {"applied": False, "reason": "row_dewarp_axis_gap_too_small"}
    offsets = source - target
    max_abs_offset = float(np.max(np.abs(offsets)))
    if max_abs_offset <= 2.0:
        return rectified, {
            "applied": False,
            "reason": "row_dewarp_offsets_within_tolerance",
            "max_abs_offset": round(max_abs_offset, 3),
        }

    height, width = rectified.shape[:2]
    output_y = np.arange(height, dtype=np.float32)
    source_y = np.interp(output_y, target, source).astype(np.float32)
    source_y[: int(max(0, round(float(target[0]))))] = output_y[: int(max(0, round(float(target[0]))))]
    tail_start = int(min(height, max(0, round(float(target[-1])))))
    if tail_start < height:
        source_y[tail_start:] = output_y[tail_start:]
    map_y = np.repeat(source_y[:, None], width, axis=1).astype(np.float32)
    map_x = np.repeat(np.arange(width, dtype=np.float32)[None, :], height, axis=0)
    dewarped = cv2.remap(
        rectified,
        map_x,
        map_y,
        interpolation=cv2.INTER_CUBIC,
        borderMode=cv2.BORDER_CONSTANT,
        borderValue=(255, 255, 255),
    )
    return dewarped, {
        "applied": True,
        "method": "vertical_row_axis_dewarp_after_quad_rectification",
        "source_count": len(source_ys),
        "template_count": len(template_ys),
        "max_abs_offset": round(max_abs_offset, 3),
        "mean_abs_offset": round(float(np.mean(np.abs(offsets))), 3),
        "source_height_quality": _row_height_outlier_evidence([float(value) for value in source.tolist()]),
    }


def snap_regions_x_to_local_fax_rulings(
    rectified: np.ndarray,
    regions: list[dict[str, Any]],
    *,
    snap_y: bool = True,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    target_regions = [
        region
        for region in regions
        if isinstance(region.get("bbox"), list) and len(region["bbox"]) == 4
    ]
    if not target_regions:
        return regions, {"applied": False, "reason": "no_target_regions"}
    original_boundaries = sorted(
        {int(round(float(region["bbox"][0]))) for region in target_regions}
        | {int(round(float(region["bbox"][2]))) for region in target_regions}
    )
    if len(original_boundaries) < 2:
        return regions, {"applied": False, "reason": "insufficient_boundaries"}
    row_boundaries = sorted(
        {int(round(float(region["bbox"][1]))) for region in target_regions}
        | {int(round(float(region["bbox"][3]))) for region in target_regions}
    )
    if len(row_boundaries) < 2:
        return regions, {"applied": False, "reason": "insufficient_row_boundaries"}

    gray = cv2.cvtColor(rectified, cv2.COLOR_BGR2GRAY) if rectified.ndim == 3 else rectified
    if rectified.ndim == 3:
        blue, green, red = cv2.split(rectified)
        dark_mask = ((gray < 95) & (green < 130) & (red < 130) & (blue < 130)).astype(np.uint8)
    else:
        dark_mask = (gray < 95).astype(np.uint8)
    height, width = dark_mask.shape[:2]
    line_source = rectified if rectified.ndim == 3 else cv2.cvtColor(rectified, cv2.COLOR_GRAY2BGR)
    horizontal_mask, _vertical_mask = _split_line_masks(line_source)

    def projection_centers(projection: np.ndarray, *, threshold: float, min_len: int) -> list[float]:
        if projection.size == 0:
            return []
        centers: list[float] = []
        in_segment = False
        start = 0
        for index, value in enumerate(projection):
            if value >= threshold and not in_segment:
                start = index
                in_segment = True
            elif (value < threshold or index == len(projection) - 1) and in_segment:
                end = index if value < threshold else index + 1
                if end - start >= min_len:
                    values = projection[start:end]
                    if values.size and float(values.sum()) > 0.0:
                        positions = np.arange(start, end, dtype=np.float32)
                        centers.append(float(np.average(positions, weights=values)))
                in_segment = False
        return centers

    def nearest(value: float, candidates: list[float], *, max_distance: float) -> float | None:
        if not candidates:
            return None
        best = min(candidates, key=lambda candidate: abs(float(candidate) - value))
        return float(best) if abs(float(best) - value) <= max_distance else None

    def detected_y_edges_for_x_span(x0: float, x1: float) -> list[float]:
        span_width = max(1.0, float(x1) - float(x0))
        inset = min(6.0, span_width * 0.22)
        span_left = max(0, int(round(float(x0) + inset)))
        span_right = min(width, int(round(float(x1) - inset)))
        if span_right <= span_left:
            return []
        detected: list[float] = []
        line_projection = horizontal_mask[:, span_left:span_right].sum(axis=1).astype(np.float32) / 255.0
        if float(line_projection.max(initial=0.0)) > 0.0:
            line_smooth = np.convolve(line_projection, np.ones(5, dtype=np.float32) / 5.0, mode="same")
            line_threshold = max(3.0, float(line_smooth.max(initial=0.0)) * 0.34)
            detected.extend(projection_centers(line_smooth, threshold=line_threshold, min_len=2))
        dark_projection = dark_mask[:, span_left:span_right].sum(axis=1).astype(np.float32)
        if float(dark_projection.max(initial=0.0)) > 0.0:
            dark_smooth = np.convolve(dark_projection, np.ones(5, dtype=np.float32) / 5.0, mode="same")
            dark_threshold = max(4.0, float(dark_smooth.max(initial=0.0)) * 0.44)
            detected.extend(projection_centers(dark_smooth, threshold=dark_threshold, min_len=2))
        if not detected:
            return []
        merged: list[float] = []
        for value in sorted(float(item) for item in detected):
            if merged and abs(float(value) - float(merged[-1])) <= 2.0:
                merged[-1] = (float(merged[-1]) + float(value)) / 2.0
            else:
                merged.append(float(value))
        return merged

    row_edge_snaps: dict[int, list[float | None]] = {}
    row_debug: list[dict[str, Any]] = []
    def snapped_x_edges_for_y(y_value: float) -> list[float | None] | None:
        band_top = max(0, int(round(float(y_value) - 24.0)))
        band_bottom = min(height, int(round(float(y_value) + 24.0)))
        if band_bottom <= band_top:
            return None
        projection = dark_mask[band_top:band_bottom, :].sum(axis=0).astype(np.float32)
        if float(projection.max(initial=0.0)) <= 0.0:
            return None
        smooth = np.convolve(projection, np.ones(5, dtype=np.float32) / 5.0, mode="same")
        threshold = max(7.0, float(smooth.max(initial=0.0)) * 0.36)
        centers = projection_centers(smooth, threshold=threshold, min_len=2)
        return [
            nearest(float(boundary), centers, max_distance=18.0)
            for boundary in original_boundaries
        ]

    for row_boundary in row_boundaries:
        snapped = snapped_x_edges_for_y(float(row_boundary))
        if snapped is None:
            continue
        deltas = [
            None if snapped_x is None else float(snapped_x) - float(boundary)
            for boundary, snapped_x in zip(original_boundaries, snapped)
        ]
        matched = [delta for delta in deltas if delta is not None]
        row_edge_snaps[int(row_boundary)] = snapped
        row_debug.append(
            {
                "row_y": int(row_boundary),
                "matched_count": len(matched),
                "mean_delta": None if not matched else round(float(np.mean(matched)), 3),
                "max_abs_delta": None if not matched else round(float(max(abs(delta) for delta in matched)), 3),
            }
        )

    y_span_left = max(0, min(original_boundaries))
    y_span_right = min(width, max(original_boundaries))
    y_projection = dark_mask[:, y_span_left:y_span_right].sum(axis=1).astype(np.float32)
    y_smooth = np.convolve(y_projection, np.ones(5, dtype=np.float32) / 5.0, mode="same")
    y_threshold = max(15.0, float(y_smooth.max(initial=0.0)) * 0.35)
    detected_y_edges = projection_centers(y_smooth, threshold=y_threshold, min_len=2)
    y_edge_snaps = {
        int(boundary): nearest(float(boundary), detected_y_edges, max_distance=18.0)
        for boundary in row_boundaries
    }
    top_boundary = int(min(row_boundaries))
    bottom_boundary = int(max(row_boundaries))
    if y_edge_snaps.get(top_boundary) is None:
        top_candidates = [
            float(value)
            for value in detected_y_edges
            if float(top_boundary) - 80.0 <= float(value) <= float(top_boundary) + 80.0
        ]
        if top_candidates:
            y_edge_snaps[top_boundary] = min(
                top_candidates,
                key=lambda value: abs(float(value) - float(top_boundary)),
            )
    bottom_candidates = [
        float(value)
        for value in detected_y_edges
        if float(bottom_boundary) - 160.0 <= float(value) <= float(bottom_boundary) + 20.0
    ]
    if bottom_candidates:
        # The bottom table ruling can be far above the template edge after
        # local warp correction.  Prefer the lowest detected long ruling in
        # the lower band; near-template dark peaks can be edge noise rather
        # than the actual bottom horizontal ruling.
        y_edge_snaps[bottom_boundary] = max(bottom_candidates)
    for boundary in (top_boundary, bottom_boundary):
        snapped_y = y_edge_snaps.get(boundary)
        if snapped_y is None or abs(float(snapped_y) - float(boundary)) <= 1.0:
            continue
        moved_snaps = snapped_x_edges_for_y(float(snapped_y))
        if moved_snaps is not None:
            row_edge_snaps[boundary] = moved_snaps

    boundary_index = {boundary: index for index, boundary in enumerate(original_boundaries)}
    snapped_regions: list[dict[str, Any]] = []
    snapped_count = 0
    fallback_count = 0
    for region in regions:
        box = region.get("bbox")
        if not isinstance(box, list) or len(box) != 4:
            snapped_regions.append(region)
            continue
        x0, y0, x1, y1 = [float(value) for value in box]
        left_key = int(round(x0))
        right_key = int(round(x1))
        top_key = int(round(y0))
        bottom_key = int(round(y1))
        left_index = boundary_index.get(left_key)
        right_index = boundary_index.get(right_key)
        top_snaps = row_edge_snaps.get(top_key)
        bottom_snaps = row_edge_snaps.get(bottom_key)
        snapped_left_values: list[float] = []
        snapped_right_values: list[float] = []
        if top_snaps and left_index is not None and left_index < len(top_snaps) and top_snaps[left_index] is not None:
            snapped_left_values.append(float(top_snaps[left_index]))
        if bottom_snaps and left_index is not None and left_index < len(bottom_snaps) and bottom_snaps[left_index] is not None:
            snapped_left_values.append(float(bottom_snaps[left_index]))
        if top_snaps and right_index is not None and right_index < len(top_snaps) and top_snaps[right_index] is not None:
            snapped_right_values.append(float(top_snaps[right_index]))
        if bottom_snaps and right_index is not None and right_index < len(bottom_snaps) and bottom_snaps[right_index] is not None:
            snapped_right_values.append(float(bottom_snaps[right_index]))
        if snap_y:
            snapped_top = y_edge_snaps.get(top_key)
            snapped_bottom = y_edge_snaps.get(bottom_key)
        else:
            snapped_top = float(y0)
            snapped_bottom = float(y1)
        local_y_edges = detected_y_edges_for_x_span(x0, x1) if snap_y else []
        local_top = nearest(float(top_key), local_y_edges, max_distance=24.0) if snap_y else None
        local_bottom = nearest(float(bottom_key), local_y_edges, max_distance=24.0) if snap_y else None
        expected_height = max(1.0, float(y1) - float(y0))
        local_height_ok = (
            local_top is not None
            and local_bottom is not None
            and expected_height * 0.72 <= float(local_bottom) - float(local_top) <= expected_height * 1.34
        )
        if (
            local_top is not None
            and local_bottom is not None
            and float(local_bottom) > float(local_top) + 8.0
            and local_height_ok
        ):
            snapped_top = local_top
            snapped_bottom = local_bottom
        if (
            len(snapped_left_values) < 2
            or len(snapped_right_values) < 2
            or snapped_top is None
            or snapped_bottom is None
        ):
            fallback_count += 1
            snapped_regions.append(region)
            continue
        snapped_x0 = float(np.mean(snapped_left_values))
        snapped_x1 = float(np.mean(snapped_right_values))
        snapped_y0 = float(snapped_top)
        snapped_y1 = float(snapped_bottom)
        if snapped_x1 <= snapped_x0 + 8.0 or snapped_y1 <= snapped_y0 + 8.0:
            fallback_count += 1
            snapped_regions.append(region)
            continue
        snapped_count += 1
        snapped_regions.append(
            {
                **region,
                "bbox": [snapped_x0, snapped_y0, snapped_x1, snapped_y1],
                "local_grid_snap": {
                    "source_bbox": [x0, y0, x1, y1],
                    "snapped_bbox": [snapped_x0, snapped_y0, snapped_x1, snapped_y1],
                    "x_delta": [snapped_x0 - x0, snapped_x1 - x1],
                    "y_delta": [snapped_y0 - y0, snapped_y1 - y1],
                    "method": "row_edge_local_fax_ruling_snap_v2",
                    "y_snap_enabled": bool(snap_y),
                    "local_y_snap_applied": bool(snap_y and local_height_ok),
                },
            }
        )
    required_min = max(1, int(round(float(len(target_regions)) * 0.55)))
    if snapped_count < required_min:
        return regions, {
            "applied": False,
            "reason": "local_grid_snap_insufficient_matches",
            "original_boundaries": original_boundaries,
            "row_boundary_count": len(row_boundaries),
            "snapped_region_count": snapped_count,
            "fallback_region_count": fallback_count,
            "required_min_snapped_region_count": required_min,
            "row_debug": row_debug,
        }
    return snapped_regions, {
        "applied": True,
        "method": "row_edge_local_fax_ruling_snap_v1",
        "y_snap_enabled": bool(snap_y),
        "original_boundaries": original_boundaries,
        "row_boundary_count": len(row_boundaries),
        "outer_y_snap": {
            "top_template_y": top_boundary,
            "top_snapped_y": y_edge_snaps.get(top_boundary),
            "bottom_template_y": bottom_boundary,
            "bottom_snapped_y": y_edge_snaps.get(bottom_boundary),
        },
        "snapped_region_count": snapped_count,
        "fallback_region_count": fallback_count,
        "row_debug": row_debug,
    }


def _clustered_projection_positions(
    projection: np.ndarray,
    *,
    threshold_ratio: float = 0.18,
    min_value: float = 18.0,
    max_gap: int = 3,
) -> list[int]:
    if projection.size == 0 or float(projection.max()) <= 0:
        return []
    threshold = max(min_value, float(projection.max()) * threshold_ratio)
    indices = np.where(projection >= threshold)[0].tolist()
    if not indices:
        return []
    groups: list[list[int]] = [[indices[0]]]
    for value in indices[1:]:
        if value - groups[-1][-1] > max_gap:
            groups.append([value])
        else:
            groups[-1].append(value)
    return [int(round(sum(group) / len(group))) for group in groups]


def _header_bounds_for_axis_alignment(table_bbox: list[int], template_ys: list[int]) -> tuple[int, int, int, int]:
    x0, _table_y0, x1, _table_y1 = [int(round(float(value))) for value in table_bbox]
    sorted_ys = sorted(int(round(float(value))) for value in template_ys)
    if len(sorted_ys) >= 3:
        y0 = sorted_ys[0] - 18
        y1 = sorted_ys[2] + 12
    else:
        y0 = table_bbox[1] - 18
        y1 = table_bbox[1] + 240
    return x0, max(0, int(y0)), x1, max(0, int(y1))


def _cluster_header_intersection_points(mask: np.ndarray) -> list[dict[str, Any]]:
    count, _labels, stats, centroids = cv2.connectedComponentsWithStats(mask, connectivity=8)
    points: list[dict[str, Any]] = []
    for label in range(1, count):
        x, y, w, h, area = [int(value) for value in stats[label]]
        if area < 4:
            continue
        cx, cy = centroids[label]
        points.append(
            {
                "x": round(float(cx), 2),
                "y": round(float(cy), 2),
                "bbox": [x, y, w, h],
                "area": area,
            }
        )
    return sorted(points, key=lambda item: (float(item["y"]), float(item["x"])))


def _detect_header_intersections_for_axis_alignment(
    rectified_bgr: np.ndarray,
    *,
    table_bbox: list[int],
    template_ys: list[int],
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    h_mask, v_mask = _split_line_masks(rectified_bgr)
    x0, y0, x1, y1 = _header_bounds_for_axis_alignment(table_bbox, template_ys)
    h = h_mask[y0:y1, x0:x1]
    v = v_mask[y0:y1, x0:x1]
    if h.size == 0 or v.size == 0:
        return [], {"reason": "empty_header_roi", "header_roi": [x0, y0, x1, y1]}
    h_dilated = cv2.dilate(h, cv2.getStructuringElement(cv2.MORPH_RECT, (7, 5)), iterations=1)
    v_dilated = cv2.dilate(v, cv2.getStructuringElement(cv2.MORPH_RECT, (5, 7)), iterations=1)
    intersections = cv2.bitwise_and(h_dilated, v_dilated)
    points = _cluster_header_intersection_points(intersections)
    for point in points:
        point["x"] = round(float(point["x"]) + x0, 2)
        point["y"] = round(float(point["y"]) + y0, 2)
        bx, by, bw, bh = point["bbox"]
        point["bbox"] = [int(bx + x0), int(by + y0), int(bw), int(bh)]
    return points, {
        "header_roi": [x0, y0, x1, y1],
        "horizontal_pixels": int(np.count_nonzero(h)),
        "vertical_pixels": int(np.count_nonzero(v)),
        "intersection_count": len(points),
    }


def _cluster_header_axis_values(points: list[dict[str, Any]], axis: str, tolerance_px: float = 18.0) -> list[dict[str, Any]]:
    indexed_values = sorted(
        [(index, float(point[axis])) for index, point in enumerate(points)],
        key=lambda item: item[1],
    )
    clusters: list[list[tuple[int, float]]] = []
    for point_index, value in indexed_values:
        if not clusters:
            clusters.append([(point_index, value)])
            continue
        center = float(np.median([item[1] for item in clusters[-1]]))
        if abs(value - center) <= tolerance_px:
            clusters[-1].append((point_index, value))
        else:
            clusters.append([(point_index, value)])
    result: list[dict[str, Any]] = []
    for cluster_index, cluster in enumerate(clusters):
        values = [item[1] for item in cluster]
        result.append(
            {
                "cluster_index": cluster_index,
                "value": round(float(np.median(values)), 2),
                "count": len(cluster),
                "point_indexes": [item[0] for item in cluster],
            }
        )
    return result


def _filter_header_x_clusters_by_y_coverage(
    *,
    x_clusters: list[dict[str, Any]],
    y_clusters: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    point_to_y: dict[int, int] = {}
    for cluster in y_clusters:
        for point_index in cluster["point_indexes"]:
            point_to_y[int(point_index)] = int(cluster["cluster_index"])

    accepted: list[dict[str, Any]] = []
    rejected: list[dict[str, Any]] = []
    for cluster in x_clusters:
        y_cluster_indexes = sorted(
            {point_to_y[int(point_index)] for point_index in cluster["point_indexes"] if int(point_index) in point_to_y}
        )
        enriched = dict(cluster)
        enriched["raw_cluster_index"] = int(cluster["cluster_index"])
        enriched["covered_y_cluster_indexes"] = y_cluster_indexes
        if len(y_cluster_indexes) >= 2:
            enriched["cluster_index"] = len(accepted)
            accepted.append(enriched)
        else:
            enriched["reject_reason"] = "x_cluster_touches_less_than_two_header_y_levels"
            rejected.append(enriched)
    return accepted, rejected


def _filter_header_x_clusters_by_table_span(
    *,
    x_clusters: list[dict[str, Any]],
    vertical_mask: np.ndarray,
    table_y0: int,
    table_y1: int,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    accepted: list[dict[str, Any]] = []
    rejected: list[dict[str, Any]] = []
    for cluster in x_clusters:
        evidence = _vertical_line_span_evidence(
            vertical_mask,
            x=int(round(float(cluster["value"]))),
            y0=table_y0,
            y1=table_y1,
        )
        enriched = dict(cluster)
        enriched["table_span_evidence"] = evidence
        if evidence.get("valid"):
            enriched["cluster_index"] = len(accepted)
            accepted.append(enriched)
        else:
            enriched["reject_reason"] = "x_cluster_not_full_table_vertical_line"
            rejected.append(enriched)
    return accepted, rejected


def _structural_match_header_x_clusters(
    *,
    template_xs: list[int],
    x_clusters: list[dict[str, Any]],
) -> tuple[list[float] | None, dict[str, Any]]:
    """Align header-derived X clusters to template axes without requiring equal counts."""
    n = len(template_xs)
    m = len(x_clusters)
    if n < 2 or m < 2:
        return None, {
            "used": False,
            "reason": "insufficient_header_x_clusters_for_structural_match",
            "template_count": n,
            "cluster_count": m,
        }

    template = [float(value) for value in template_xs]
    clusters = [float(cluster["value"]) for cluster in x_clusters]

    expected_match_count = min(n, m)
    if n >= m:
        cluster_index_sets = [tuple(range(m))]
        template_index_sets = itertools.combinations(range(n), expected_match_count)
    else:
        cluster_index_sets = itertools.combinations(range(m), expected_match_count)
        template_index_sets = [tuple(range(n))]

    best_score = float("inf")
    best_cluster_indexes: tuple[int, ...] | None = None
    best_template_indexes: tuple[int, ...] | None = None
    checked = 0
    max_combinations = 250_000
    for cluster_indexes in cluster_index_sets:
        cluster_values = np.array([clusters[index] for index in cluster_indexes], dtype=np.float32)
        cluster_span = max(1.0, float(cluster_values[-1] - cluster_values[0]))
        cluster_gaps = np.diff(cluster_values) / cluster_span
        for template_indexes in template_index_sets:
            checked += 1
            if checked > max_combinations:
                break
            template_values = np.array([template[index] for index in template_indexes], dtype=np.float32)
            template_span = max(1.0, float(template_values[-1] - template_values[0]))
            template_gaps = np.diff(template_values) / template_span
            gap_score = float(np.mean(np.abs(cluster_gaps - template_gaps))) if len(cluster_gaps) else 0.0
            scale = cluster_span / template_span
            projected = cluster_values[0] + (template_values - template_values[0]) * scale
            affine_residual = float(np.mean(np.abs(projected - cluster_values))) / cluster_span
            score = gap_score + 0.35 * affine_residual
            if score < best_score:
                best_score = score
                best_cluster_indexes = tuple(int(index) for index in cluster_indexes)
                best_template_indexes = tuple(int(index) for index in template_indexes)
        if checked > max_combinations:
            break

    if best_cluster_indexes is None or best_template_indexes is None:
        return None, {
            "used": False,
            "reason": "no_structural_header_gap_match",
            "template_count": n,
            "cluster_count": m,
            "checked": checked,
        }

    matched_pairs = list(zip(best_cluster_indexes, best_template_indexes, strict=False))
    skipped_template_indexes = [index for index in range(n) if index not in set(best_template_indexes)]
    skipped_cluster_indexes = [index for index in range(m) if index not in set(best_cluster_indexes)]

    min_matches = max(4, min(n, m) - 1)
    if len(matched_pairs) < min_matches:
        return None, {
            "used": False,
            "reason": "too_few_structural_header_matches",
            "match_count": len(matched_pairs),
            "required_match_count": min_matches,
            "skipped_template_indexes": skipped_template_indexes,
            "skipped_cluster_indexes": skipped_cluster_indexes,
            "score": round(float(best_score), 6),
        }

    corrected: list[float | None] = [None for _ in range(n)]
    for cluster_index, template_index in matched_pairs:
        corrected[template_index] = clusters[cluster_index]
    corrected[0] = float(template_xs[0])
    corrected[-1] = float(template_xs[-1])

    known_indexes = [index for index, value in enumerate(corrected) if value is not None]
    if len(known_indexes) < 2:
        return None, {
            "used": False,
            "reason": "insufficient_known_axes_after_structural_match",
            "match_count": len(matched_pairs),
        }
    known_template = np.array([template[index] for index in known_indexes], dtype=np.float32)
    known_corrected = np.array([float(corrected[index]) for index in known_indexes], dtype=np.float32)
    all_template = np.array(template, dtype=np.float32)
    interpolated = np.interp(all_template, known_template, known_corrected).astype(np.float32).tolist()
    interpolated[0] = float(template_xs[0])
    interpolated[-1] = float(template_xs[-1])
    if not np.all(np.diff(np.array(interpolated, dtype=np.float32)) > 0):
        return None, {
            "used": False,
            "reason": "structural_header_match_not_monotonic",
            "matched_pairs": matched_pairs,
            "skipped_template_indexes": skipped_template_indexes,
            "skipped_cluster_indexes": skipped_cluster_indexes,
        }
    return [float(value) for value in interpolated], {
        "used": True,
        "reason": "applied_structural_header_x_match",
        "match_count": len(matched_pairs),
        "score": round(float(best_score), 6),
        "checked": checked,
        "matched_pairs": [
            {
                "cluster_index": int(cluster_index),
                "template_index": int(template_index),
                "cluster_x": round(float(clusters[cluster_index]), 3),
                "template_x": round(float(template[template_index]), 3),
            }
            for cluster_index, template_index in matched_pairs
        ],
        "skipped_template_indexes": skipped_template_indexes,
        "skipped_cluster_indexes": skipped_cluster_indexes,
    }


def _header_intersection_correct_xs(
    *,
    rectified_fax: np.ndarray,
    template_xs: list[int],
    template_ys: list[int],
) -> tuple[list[float] | None, dict[str, Any]]:
    if len(template_xs) < 5 or len(template_ys) < 3:
        return None, {"used": False, "reason": "insufficient_template_axes"}
    table_bbox = [template_xs[0], template_ys[0], template_xs[-1], template_ys[-1]]
    points, detection = _detect_header_intersections_for_axis_alignment(
        rectified_fax,
        table_bbox=table_bbox,
        template_ys=template_ys,
    )
    _h_mask, v_mask = _split_line_masks(rectified_fax)
    v_mask = cv2.dilate(v_mask, cv2.getStructuringElement(cv2.MORPH_RECT, (2, 5)), iterations=1)
    raw_x_clusters = _cluster_header_axis_values(points, "x")
    y_clusters = _cluster_header_axis_values(points, "y")
    span_x_clusters, span_rejected_x_clusters = _filter_header_x_clusters_by_table_span(
        x_clusters=raw_x_clusters,
        vertical_mask=v_mask,
        table_y0=max(0, int(round(float(template_ys[0])))),
        table_y1=min(v_mask.shape[0], int(round(float(template_ys[-1])))),
    )
    filtered_x_clusters, rejected_x_clusters = _filter_header_x_clusters_by_y_coverage(
        x_clusters=raw_x_clusters,
        y_clusters=y_clusters,
    )
    span_and_y_x_clusters, span_and_y_rejected = _filter_header_x_clusters_by_y_coverage(
        x_clusters=span_x_clusters,
        y_clusters=y_clusters,
    )
    x_clusters = raw_x_clusters
    x_cluster_source = "raw_x_clusters"
    if len(span_and_y_x_clusters) == len(template_xs):
        x_clusters = span_and_y_x_clusters
        rejected_x_clusters = span_rejected_x_clusters + span_and_y_rejected
        x_cluster_source = "filtered_by_table_span_and_header_y_coverage"
    elif len(span_x_clusters) == len(template_xs):
        x_clusters = span_x_clusters
        rejected_x_clusters = span_rejected_x_clusters
        x_cluster_source = "filtered_by_table_span"
    elif len(raw_x_clusters) > len(template_xs) and len(filtered_x_clusters) == len(template_xs):
        x_clusters = filtered_x_clusters
        x_cluster_source = "filtered_extra_x_clusters_by_header_y_coverage"
    elif len(raw_x_clusters) > len(template_xs):
        fixed_left_count = 4
        fixed_right_count = 1
        span_source = span_x_clusters if len(span_x_clusters) >= fixed_left_count + fixed_right_count else raw_x_clusters
        left_clusters = [dict(cluster) for cluster in span_source[:fixed_left_count]]
        right_clusters = [dict(cluster) for cluster in span_source[-fixed_right_count:]]
        rest_filtered, rest_rejected = _filter_header_x_clusters_by_y_coverage(
            x_clusters=span_source[fixed_left_count:-fixed_right_count],
            y_clusters=y_clusters,
        )
        left_and_filtered = left_clusters + rest_filtered + right_clusters
        if len(left_and_filtered) == len(template_xs):
            x_clusters = []
            for cluster in left_and_filtered:
                enriched = dict(cluster)
                enriched["cluster_index"] = len(x_clusters)
                enriched.setdefault("raw_cluster_index", int(cluster["cluster_index"]))
                x_clusters.append(enriched)
            rejected_x_clusters = span_rejected_x_clusters + rest_rejected
            x_cluster_source = "fixed_outer_then_filtered_extra_x_clusters_by_header_y_coverage"

    evidence = {
        "used": False,
        "method": "header_intersection_structural_x_axis_correction",
        **detection,
        "raw_fax_x_cluster_count": len(raw_x_clusters),
        "span_filtered_fax_x_cluster_count": len(span_x_clusters),
        "fax_x_cluster_count": len(x_clusters),
        "fax_y_cluster_count": len(y_clusters),
        "template_x_count": len(template_xs),
        "template_header_y_count": min(3, len(template_ys)),
        "fax_x_cluster_source": x_cluster_source,
        "header_intersection_points": points,
        "raw_fax_x_clusters": raw_x_clusters,
        "span_filtered_fax_x_clusters": span_x_clusters,
        "fax_x_clusters": x_clusters,
        "rejected_fax_x_clusters": rejected_x_clusters,
    }
    if len(y_clusters) != min(3, len(template_ys)):
        evidence["reason"] = "header_y_cluster_count_mismatch"
        return None, evidence
    structural_corrected, structural_evidence = _structural_match_header_x_clusters(
        template_xs=template_xs,
        x_clusters=x_clusters,
    )
    evidence["structural_match"] = structural_evidence
    if structural_corrected is not None:
        corrected = structural_corrected
    else:
        if len(x_clusters) != len(template_xs):
            evidence["reason"] = "header_x_cluster_count_mismatch"
            return None, evidence
        corrected = [float(cluster["value"]) for cluster in x_clusters]
        corrected[0] = float(template_xs[0])
        corrected[-1] = float(template_xs[-1])
    if not np.all(np.diff(np.array(corrected, dtype=np.float32)) > 0):
        evidence["reason"] = "corrected_x_not_monotonic"
        return None, evidence
    evidence["used"] = True
    evidence["reason"] = "applied"
    evidence["corrected_xs"] = [round(float(value), 3) for value in corrected]
    evidence["x_offsets_px"] = [round(float(value) - float(template_xs[index]), 3) for index, value in enumerate(corrected)]
    return corrected, evidence


def normalize_header_axis_override(
    value: object,
    *,
    expected_count: int,
    canvas_width: int,
) -> list[float] | None:
    if not isinstance(value, dict):
        return None
    coordinate_space = value.get("coordinate_space")
    if not isinstance(coordinate_space, dict):
        return None
    if str(coordinate_space.get("mode") or "").strip() != "template_canvas":
        return None
    try:
        width_value = int(coordinate_space.get("width"))
    except (TypeError, ValueError):
        return None
    if width_value != int(canvas_width):
        return None
    raw_xs = value.get("corrected_xs")
    if not isinstance(raw_xs, list) or len(raw_xs) != int(expected_count):
        return None
    xs: list[float] = []
    for raw in raw_xs:
        try:
            x = float(raw)
        except (TypeError, ValueError):
            return None
        if x < 0 or x > float(canvas_width):
            return None
        xs.append(round(x, 3))
    if not np.all(np.diff(np.array(xs, dtype=np.float32)) > 0):
        return None
    return xs


def normalize_row_axis_override(
    value: object,
    *,
    expected_count: int,
    canvas_height: int,
) -> list[float] | None:
    if not isinstance(value, dict):
        return None
    coordinate_space = value.get("coordinate_space")
    if not isinstance(coordinate_space, dict):
        return None
    if str(coordinate_space.get("mode") or "").strip() != "template_canvas":
        return None
    try:
        height_value = int(coordinate_space.get("height"))
    except (TypeError, ValueError):
        return None
    if height_value != int(canvas_height):
        return None
    raw_ys = value.get("corrected_ys")
    if not isinstance(raw_ys, list) or len(raw_ys) < int(expected_count):
        return None
    ys: list[float] = []
    for raw in raw_ys:
        try:
            y = float(raw)
        except (TypeError, ValueError):
            return None
        if y < 0 or y > float(canvas_height):
            return None
        ys.append(round(y, 3))
    if not np.all(np.diff(np.array(ys, dtype=np.float32)) > 0):
        return None
    if len(ys) > int(expected_count):
        fitted, _fit_evidence = _fit_extra_y_clusters_to_template_count(
            template_ys=[float(value) for value in range(int(expected_count))],
            clusters=[float(value) for value in ys],
        )
        if fitted is None or len(fitted) != int(expected_count):
            return None
        return [round(float(value), 3) for value in fitted]
    return ys


def _row_height_quality_score(row_edges: list[float]) -> tuple[int, float, float]:
    heights = np.diff(np.array([float(value) for value in row_edges], dtype=np.float64))
    if heights.size == 0:
        return (9999, 9999.0, 9999.0)
    median = float(np.median(heights))
    if median <= 0:
        return (9999, 9999.0, 9999.0)
    ratios = np.array([float(height) / median for height in heights], dtype=np.float64)
    outlier_count = int(np.count_nonzero((ratios < 0.55) | (ratios > 1.75)))
    max_deviation = float(np.max(np.abs(ratios - 1.0)))
    spread = float(np.max(heights) - np.min(heights))
    return (outlier_count, max_deviation, spread)


def _fit_extra_y_clusters_to_template_count(
    *,
    template_ys: list[float],
    clusters: list[float],
) -> tuple[list[float] | None, dict[str, Any]]:
    n = len(template_ys)
    m = len(clusters)
    if n < 2 or m < n:
        return None, {"used": False, "reason": "insufficient_clusters_for_extra_row_fit", "template_count": n, "cluster_count": m}
    if m == n:
        return [float(value) for value in clusters], {
            "used": False,
            "reason": "cluster_count_matches_template",
            "template_count": n,
            "cluster_count": m,
            "skipped_cluster_indexes": [],
        }
    if m - n > 3:
        return None, {
            "used": False,
            "reason": "too_many_extra_row_clusters_for_safe_fit",
            "template_count": n,
            "cluster_count": m,
            "extra_count": m - n,
        }

    template = np.array([float(value) for value in template_ys], dtype=np.float64)
    cluster_values = np.array([float(value) for value in clusters], dtype=np.float64)
    template_span = max(1.0, float(template[-1] - template[0]))
    cluster_span = max(1.0, float(cluster_values[-1] - cluster_values[0]))
    projected = cluster_values[0] + ((template - template[0]) / template_span) * cluster_span
    skip_count = m - n
    best: tuple[tuple[int, float, float, float], list[int], list[float]] | None = None
    for skipped_tuple in itertools.combinations(range(m), skip_count):
        skipped = set(skipped_tuple)
        selected_indexes = [index for index in range(m) if index not in skipped]
        selected = [float(cluster_values[index]) for index in selected_indexes]
        if len(selected) != n or not np.all(np.diff(np.array(selected, dtype=np.float64)) > 0):
            continue
        height_score = _row_height_quality_score(selected)
        alignment_cost = 0.0
        for template_index, cluster_index in enumerate(selected_indexes):
            alignment_cost += abs(float(projected[template_index]) - float(cluster_values[cluster_index])) / cluster_span
            if template_index > 0:
                template_gap = (template[template_index] - template[template_index - 1]) / template_span
                cluster_gap = (cluster_values[cluster_index] - cluster_values[selected_indexes[template_index - 1]]) / cluster_span
                alignment_cost += 0.45 * abs(float(template_gap) - float(cluster_gap))
        score = (*height_score, float(alignment_cost))
        if best is None or score < best[0]:
            best = (score, selected_indexes, selected)
    if best is None:
        return None, {"used": False, "reason": "no_safe_extra_row_cluster_fit", "template_count": n, "cluster_count": m}
    _score, selected_indexes, selected = best
    skipped_indexes = [index for index in range(m) if index not in set(selected_indexes)]
    return selected, {
        "used": True,
        "reason": "extra_row_clusters_fitted_by_row_height_quality",
        "method": "exhaustive_extra_y_cluster_skip_selection",
        "template_count": n,
        "cluster_count": m,
        "extra_count": skip_count,
        "selected_cluster_indexes": [int(index) for index in selected_indexes],
        "skipped_cluster_indexes": [int(index) for index in skipped_indexes],
        "row_height_score": {
            "outlier_count": int(_score[0]),
            "max_ratio_deviation": round(float(_score[1]), 6),
            "height_spread": round(float(_score[2]), 3),
            "alignment_cost": round(float(_score[3]), 6),
        },
    }


def _enrich_y_clusters_with_column_votes(
    *,
    points: list[dict[str, Any]],
    y_clusters: list[dict[str, Any]],
    corrected_xs: list[float],
) -> list[dict[str, Any]]:
    enriched: list[dict[str, Any]] = []
    for cluster in y_clusters:
        column_votes: list[int] = []
        for point_index in cluster.get("point_indexes") or []:
            try:
                point = points[int(point_index)]
                x = float(point["x"])
            except (TypeError, ValueError, IndexError, KeyError):
                continue
            nearest_index = min(range(len(corrected_xs)), key=lambda index: abs(float(corrected_xs[index]) - x))
            if abs(float(corrected_xs[nearest_index]) - x) <= 30.0:
                column_votes.append(nearest_index)
        unique_columns = sorted(set(column_votes))
        item = dict(cluster)
        item["column_votes"] = unique_columns
        item["column_vote_count"] = len(unique_columns)
        item["menu_support"] = any(index in unique_columns for index in (3, 4, 5))
        item["quantity_support_count"] = len([index for index in unique_columns if index >= 5])
        enriched.append(item)
    return enriched


def _filter_structural_y_clusters(y_clusters: list[dict[str, Any]]) -> list[dict[str, Any]]:
    accepted: list[dict[str, Any]] = []
    for cluster in y_clusters:
        vote_count = int(cluster.get("column_vote_count") or 0)
        quantity_count = int(cluster.get("quantity_support_count") or 0)
        menu_support = bool(cluster.get("menu_support"))
        if vote_count >= 4 or quantity_count >= 3 or (menu_support and vote_count >= 2):
            item = dict(cluster)
            item["cluster_index"] = len(accepted)
            accepted.append(item)
    return accepted


def _detect_table_intersections_for_row_axis(
    rectified_bgr: np.ndarray,
    *,
    corrected_xs: list[float],
    template_ys: list[int],
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    h_mask, v_mask = _split_line_masks(rectified_bgr)
    x0 = max(0, int(round(float(corrected_xs[0]))) - 14)
    x1 = min(rectified_bgr.shape[1], int(round(float(corrected_xs[-1]))) + 14)
    y0 = max(0, int(round(float(template_ys[0]))) - 18)
    y1 = min(rectified_bgr.shape[0], int(round(float(template_ys[-1]))) + 18)
    h = h_mask[y0:y1, x0:x1]
    v = v_mask[y0:y1, x0:x1]
    if h.size == 0 or v.size == 0:
        return [], {"reason": "empty_table_roi", "table_roi": [x0, y0, x1, y1]}
    h_dilated = cv2.dilate(h, cv2.getStructuringElement(cv2.MORPH_RECT, (9, 5)), iterations=1)
    v_dilated = cv2.dilate(v, cv2.getStructuringElement(cv2.MORPH_RECT, (5, 9)), iterations=1)
    intersections = cv2.bitwise_and(h_dilated, v_dilated)
    points = _cluster_header_intersection_points(intersections)
    for point in points:
        point["x"] = round(float(point["x"]) + x0, 2)
        point["y"] = round(float(point["y"]) + y0, 2)
        bx, by, bw, bh = point["bbox"]
        point["bbox"] = [int(bx + x0), int(by + y0), int(bw), int(bh)]
    return points, {
        "method": "full_table_intersection_structural_y_axis_correction",
        "table_roi": [x0, y0, x1, y1],
        "menu_roi": [
            max(0, int(round(float(corrected_xs[3]))) - 14) if len(corrected_xs) > 5 else x0,
            y0,
            min(rectified_bgr.shape[1], int(round(float(corrected_xs[5]))) + 14) if len(corrected_xs) > 5 else x1,
            y1,
        ],
        "horizontal_pixels": int(np.count_nonzero(h)),
        "vertical_pixels": int(np.count_nonzero(v)),
        "intersection_count": len(points),
    }


def _ordered_match_y_clusters_to_template(
    *,
    template_ys: list[int],
    y_clusters: list[dict[str, Any]],
) -> tuple[list[float] | None, dict[str, Any]]:
    n = len(template_ys)
    m = len(y_clusters)
    if n < 2 or m < 2:
        return None, {"used": False, "reason": "insufficient_row_y_clusters", "template_count": n, "cluster_count": m}
    template = np.array([float(value) for value in template_ys], dtype=np.float64)
    clusters = np.array([float(cluster["value"]) for cluster in y_clusters], dtype=np.float64)
    template_span = max(1.0, float(template[-1] - template[0]))
    cluster_span = max(1.0, float(clusters[-1] - clusters[0]))
    projected = clusters[0] + ((template - template[0]) / template_span) * cluster_span
    precomputed_corrected: list[float] | None = None
    if n >= m:
        dp = np.full((n + 1, m + 1), np.inf, dtype=np.float64)
        prev: list[list[tuple[int, int, str] | None]] = [[None for _ in range(m + 1)] for _ in range(n + 1)]
        dp[0, 0] = 0.0
        skip_penalty = 0.001
        for i in range(n):
            for j in range(m + 1):
                base = dp[i, j]
                if not np.isfinite(base):
                    continue
                if j < m:
                    cost = abs(projected[i] - clusters[j]) / cluster_span
                    if j > 0 and i > 0:
                        template_gap = (template[i] - template[i - 1]) / template_span
                        cluster_gap = (clusters[j] - clusters[j - 1]) / cluster_span
                        cost += 0.45 * abs(template_gap - cluster_gap)
                    if base + cost < dp[i + 1, j + 1]:
                        dp[i + 1, j + 1] = base + cost
                        prev[i + 1][j + 1] = (i, j, "match")
                remaining_template = n - (i + 1)
                remaining_cluster = m - j
                if remaining_template >= remaining_cluster and base + skip_penalty < dp[i + 1, j]:
                    dp[i + 1, j] = base + skip_penalty
                    prev[i + 1][j] = (i, j, "skip_template")
        if not np.isfinite(dp[n, m]):
            return None, {"used": False, "reason": "no_ordered_row_y_match", "template_count": n, "cluster_count": m}
        i, j = n, m
        matched_pairs: list[tuple[int, int]] = []
        while i > 0 or j > 0:
            step = prev[i][j]
            if step is None:
                break
            pi, pj, action = step
            if action == "match":
                matched_pairs.append((j - 1, i - 1))
            i, j = pi, pj
        matched_pairs.reverse()
        score = round(float(dp[n, m] / max(1, len(matched_pairs))), 6)
    else:
        fitted_clusters, fit_evidence = _fit_extra_y_clusters_to_template_count(
            template_ys=[float(value) for value in template_ys],
            clusters=[float(value) for value in clusters.tolist()],
        )
        if fitted_clusters is not None:
            precomputed_corrected = [float(value) for value in fitted_clusters]
            selected_cluster_indexes = [
                int(index)
                for index in (fit_evidence.get("selected_cluster_indexes") or [])
                if isinstance(index, int) or str(index).isdigit()
            ]
            if len(selected_cluster_indexes) != n:
                selected_cluster_indexes = []
                last = -1
                for fitted in fitted_clusters:
                    for index in range(last + 1, m):
                        if abs(float(clusters[index]) - float(fitted)) < 0.001:
                            selected_cluster_indexes.append(index)
                            last = index
                            break
            matched_pairs = [(cluster_index, template_index) for template_index, cluster_index in enumerate(selected_cluster_indexes)]
            score = float((fit_evidence.get("row_height_score") or {}).get("alignment_cost") or 0.0)
        else:
            selected_cluster_indexes = np.linspace(0, m - 1, n).round().astype(int).tolist()
            matched_pairs = [(cluster_index, template_index) for template_index, cluster_index in enumerate(selected_cluster_indexes)]
            score = None
    min_matches = min(min(n, m), max(10, min(n, m) - 1))
    if len(matched_pairs) < min_matches:
        return None, {
            "used": False,
            "reason": "too_few_row_y_structural_matches",
            "template_count": n,
            "cluster_count": m,
            "match_count": len(matched_pairs),
            "required_match_count": min_matches,
        }
    corrected: list[float | None] = [None for _ in range(n)]
    for cluster_index, template_index in matched_pairs:
        corrected[template_index] = float(clusters[cluster_index])
    if precomputed_corrected is not None:
        interpolated = [float(value) for value in precomputed_corrected]
    else:
        corrected[0] = float(template[0])
        corrected[-1] = float(template[-1])
        known_indexes = [index for index, value in enumerate(corrected) if value is not None]
        interpolated = np.interp(
            template,
            np.array([template[index] for index in known_indexes], dtype=np.float64),
            np.array([float(corrected[index]) for index in known_indexes], dtype=np.float64),
        ).tolist()
        interpolated[0] = float(template[0])
        interpolated[-1] = float(template[-1])
    if not np.all(np.diff(np.array(interpolated, dtype=np.float64)) > 0):
        return None, {"used": False, "reason": "row_y_structural_match_not_monotonic"}
    matched_template_indexes = {template_index for _cluster_index, template_index in matched_pairs}
    matched_cluster_indexes = {cluster_index for cluster_index, _template_index in matched_pairs}
    evidence = {
        "used": True,
        "reason": "applied_full_table_intersection_structural_y_match",
        "method": "ordered_dp_full_table_y_intersection_match",
        "score": score,
        "match_count": len(matched_pairs),
        "template_count": n,
        "cluster_count": m,
        "matched_pairs": [
            {
                "cluster_index": int(cluster_index),
                "template_index": int(template_index),
                "cluster_y": round(float(clusters[cluster_index]), 3),
                "template_y": round(float(template[template_index]), 3),
            }
            for cluster_index, template_index in matched_pairs
        ],
        "skipped_template_indexes": [index for index in range(n) if index not in matched_template_indexes],
        "skipped_cluster_indexes": [index for index in range(m) if index not in matched_cluster_indexes],
    }
    if m > n:
        evidence["method"] = "ordered_full_table_y_intersection_match_with_extra_cluster_fit"
        if "fit_evidence" in locals() and isinstance(fit_evidence, dict):
            evidence["extra_cluster_fit"] = fit_evidence
    return [float(value) for value in interpolated], evidence


def _row_height_outlier_evidence(row_edges: list[float]) -> dict[str, Any]:
    heights = np.diff(np.array([float(value) for value in row_edges], dtype=np.float64))
    if heights.size == 0:
        return {"manual_review_required": True, "reason": "row_heights_empty"}
    median = float(np.median(heights))
    if median <= 0:
        return {"manual_review_required": True, "reason": "row_height_median_invalid"}
    outliers: list[dict[str, Any]] = []
    for index, height in enumerate(heights.tolist()):
        ratio = float(height) / median
        if ratio < 0.55 or ratio > 1.75:
            outliers.append({"row_band_index": index, "height": round(float(height), 3), "median_ratio": round(ratio, 3)})
    return {
        "manual_review_required": bool(outliers),
        "reason": "row_height_outlier_detected" if outliers else "row_heights_within_tolerance",
        "median_height": round(median, 3),
        "min_height": round(float(np.min(heights)), 3),
        "max_height": round(float(np.max(heights)), 3),
        "outlier_count": len(outliers),
        "outliers": outliers,
    }


def _row_intersection_correct_ys(
    *,
    rectified_fax: np.ndarray,
    corrected_xs: list[float],
    template_ys: list[int],
) -> tuple[list[float] | None, dict[str, Any]]:
    if len(corrected_xs) < 2 or len(template_ys) < 2:
        return None, {"used": False, "reason": "insufficient_axes_for_row_intersection_match"}
    points, detection = _detect_table_intersections_for_row_axis(
        rectified_fax,
        corrected_xs=[float(value) for value in corrected_xs],
        template_ys=template_ys,
    )
    raw_y_clusters = _cluster_header_axis_values(points, "y", tolerance_px=18.0)
    voted_y_clusters = _enrich_y_clusters_with_column_votes(
        points=points,
        y_clusters=raw_y_clusters,
        corrected_xs=[float(value) for value in corrected_xs],
    )
    y_clusters = _filter_structural_y_clusters(voted_y_clusters)
    corrected, structural_evidence = _ordered_match_y_clusters_to_template(template_ys=template_ys, y_clusters=y_clusters)
    evidence = {
        "used": False,
        "method": "full_table_intersection_structural_y_axis_correction",
        **detection,
        "raw_fax_y_cluster_count": len(raw_y_clusters),
        "fax_y_cluster_count": len(y_clusters),
        "template_y_count": len(template_ys),
        "intersection_points": points,
        "raw_fax_y_clusters": voted_y_clusters,
        "fax_y_clusters": y_clusters,
        "structural_match": structural_evidence,
    }
    if corrected is None:
        evidence["reason"] = structural_evidence.get("reason") if isinstance(structural_evidence, dict) else "row_y_match_failed"
        return None, evidence
    height_evidence = _row_height_outlier_evidence(corrected)
    evidence["row_height_quality"] = height_evidence
    evidence["corrected_ys"] = [round(float(value), 3) for value in corrected]
    evidence["y_offsets_px"] = [round(float(value) - float(template_ys[index]), 3) for index, value in enumerate(corrected)]
    evidence["used"] = True
    if height_evidence.get("manual_review_required"):
        evidence["reason"] = "applied_with_row_height_manual_review_required"
        return corrected, evidence
    evidence["reason"] = "applied"
    return corrected, evidence


def _draw_line_extraction(rectified: np.ndarray) -> tuple[Image.Image, dict[str, Any]]:
    h_mask, v_mask = _split_line_masks(rectified)
    image = _bgr_to_rgb_image(rectified).convert("RGBA")
    layer = Image.new("RGBA", image.size, (0, 0, 0, 0))
    overlay = np.array(layer)
    overlay[h_mask > 0] = (255, 0, 0, 170)
    overlay[v_mask > 0] = (0, 80, 255, 170)
    composed = Image.alpha_composite(image, Image.fromarray(overlay, "RGBA")).convert("RGB")
    return composed, {
        "method": "rectified_fax_line_extraction",
        "horizontal_pixels": int(np.count_nonzero(h_mask)),
        "vertical_pixels": int(np.count_nonzero(v_mask)),
    }


def _single_cell_bbox(
    *,
    row_index: int,
    col_index: int,
    row_edges: list[float],
    column_edges: list[float],
) -> list[float]:
    return [
        float(column_edges[col_index]),
        float(row_edges[row_index]),
        float(column_edges[col_index + 1]),
        float(row_edges[row_index + 1]),
    ]


def _step_review_worksheet_row_to_grid_index(worksheet_row: int) -> int | None:
    if 7 <= worksheet_row <= 8:
        return 0
    if 9 <= worksheet_row <= 10:
        return 1
    if STEP_REVIEW_BODY_START_ROW <= worksheet_row <= STEP_REVIEW_BODY_END_ROW:
        return STEP_REVIEW_HEADER_BANDS + (worksheet_row - STEP_REVIEW_BODY_START_ROW)
    return None


def _step_review_physical_row_map(worksheet: Any, *, row_count: int) -> dict[int, dict[str, Any]]:
    rows: dict[int, dict[str, Any]] = {}
    last_date = ""
    for row_index in range(STEP_REVIEW_HEADER_BANDS, row_count):
        worksheet_row = STEP_REVIEW_BODY_START_ROW + (row_index - STEP_REVIEW_HEADER_BANDS)
        if worksheet_row > STEP_REVIEW_BODY_END_ROW:
            break
        date_text = hakodate_assignment_service._normalize_slot_text(worksheet.cell(row=worksheet_row, column=1).value)  # noqa: SLF001
        if date_text and not hakodate_assignment_service._is_weekday_only(date_text):  # noqa: SLF001
            last_date = date_text
        rows[row_index] = {
            "worksheet_row": worksheet_row,
            "row_index": row_index,
            "date": date_text,
            "effective_date": last_date,
            "daypart": hakodate_assignment_service._normalize_slot_text(worksheet.cell(row=worksheet_row, column=2).value),  # noqa: SLF001
            "aux": hakodate_assignment_service._normalize_slot_text(worksheet.cell(row=worksheet_row, column=3).value),  # noqa: SLF001
            "menu_name": hakodate_assignment_service._normalize_slot_text(worksheet.cell(row=worksheet_row, column=4).value),  # noqa: SLF001
            "menu_key": hakodate_assignment_service._normalize_menu_key(worksheet.cell(row=worksheet_row, column=4).value),  # noqa: SLF001
        }
    return rows


def _step_review_merged_or_single_cell_bbox(
    *,
    row_index: int,
    col_index: int,
    worksheet_row: int,
    worksheet_col: int,
    row_edges: list[float],
    column_edges: list[float],
    merged_cells: dict[tuple[int, int], dict[str, Any]],
) -> tuple[list[float], dict[str, Any] | None]:
    merged = merged_cells.get((worksheet_row, worksheet_col))
    if not merged:
        return _single_cell_bbox(
            row_index=row_index,
            col_index=col_index,
            row_edges=row_edges,
            column_edges=column_edges,
        ), None
    start_row_index = _step_review_worksheet_row_to_grid_index(int(merged["min_row"]))
    end_row_index = _step_review_worksheet_row_to_grid_index(int(merged["max_row"]))
    if start_row_index is None or end_row_index is None:
        return _single_cell_bbox(
            row_index=row_index,
            col_index=col_index,
            row_edges=row_edges,
            column_edges=column_edges,
        ), None
    start_col_index = max(0, int(merged["min_col"]) - 1)
    end_col_index = max(0, int(merged["max_col"]) - 1)
    if (
        start_col_index >= len(column_edges) - 1
        or end_col_index >= len(column_edges) - 1
        or start_row_index >= len(row_edges) - 1
        or end_row_index >= len(row_edges) - 1
        or end_col_index < start_col_index
        or end_row_index < start_row_index
    ):
        return _single_cell_bbox(
            row_index=row_index,
            col_index=col_index,
            row_edges=row_edges,
            column_edges=column_edges,
        ), None
    return [
        float(column_edges[start_col_index]),
        float(row_edges[start_row_index]),
        float(column_edges[end_col_index + 1]),
        float(row_edges[end_row_index + 1]),
    ], dict(merged)


def _step_review_merge_regions_for_grid(
    worksheet: Any,
    *,
    row_edges: list[float],
    column_edges: list[float],
    quantity_columns: set[int] | None = None,
) -> list[dict[str, Any]]:
    quantity_columns = set(quantity_columns or set())
    regions: list[dict[str, Any]] = []
    for merged_range in worksheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
        start_row_index = _step_review_worksheet_row_to_grid_index(int(min_row))
        end_row_index = _step_review_worksheet_row_to_grid_index(int(max_row))
        if start_row_index is None or end_row_index is None:
            continue
        start_col_index = max(0, int(min_col) - 1)
        end_col_index = max(0, int(max_col) - 1)
        if (
            start_col_index >= len(column_edges) - 1
            or end_col_index >= len(column_edges) - 1
            or start_row_index >= len(row_edges) - 1
            or end_row_index >= len(row_edges) - 1
            or end_col_index < start_col_index
            or end_row_index < start_row_index
        ):
            continue
        regions.append(
            {
                "range": str(merged_range),
                "min_col": int(min_col),
                "min_row": int(min_row),
                "max_col": int(max_col),
                "max_row": int(max_row),
                "row_span": int(end_row_index - start_row_index + 1),
                "worksheet_row_span": int(max_row - min_row + 1),
                "col_span": int(max_col - min_col + 1),
                "start_col_index": start_col_index,
                "end_col_index": end_col_index,
                "start_row_index": start_row_index,
                "end_row_index": end_row_index,
                "bbox": [
                    float(column_edges[start_col_index]),
                    float(row_edges[start_row_index]),
                    float(column_edges[end_col_index + 1]),
                    float(row_edges[end_row_index + 1]),
                ],
                "is_quantity": any(
                    col_idx in quantity_columns
                    for col_idx in range(int(min_col), int(max_col) + 1)
                ),
            }
        )
    return regions


def _physical_internal_horizontal_lines(
    *,
    horizontal_line_mask: np.ndarray | None,
    row_edges: list[float],
    bbox: list[float],
    start_row_index: int,
    row_span: int,
    min_line_ratio: float = 0.65,
    search_px: int = 4,
) -> list[dict[str, Any]]:
    if horizontal_line_mask is None or row_span <= 1 or len(bbox) != 4:
        return []
    x0, _y0, x1, _y1 = [float(value) for value in bbox]
    width = max(1.0, float(x1 - x0))
    x_start = max(0, int(round(x0)))
    x_end = min(horizontal_line_mask.shape[1], int(round(x1)))
    if x_end <= x_start:
        return []
    hits: list[dict[str, Any]] = []
    for boundary_idx in range(start_row_index + 1, start_row_index + row_span):
        if boundary_idx >= len(row_edges):
            continue
        expected_y = int(round(float(row_edges[boundary_idx])))
        best_y = expected_y
        best_ratio = 0.0
        for y in range(max(0, expected_y - search_px), min(horizontal_line_mask.shape[0], expected_y + search_px + 1)):
            row = horizontal_line_mask[y : y + 1, x_start:x_end]
            if row.size == 0:
                continue
            ratio = float(row.sum() / 255.0) / width
            if ratio > best_ratio:
                best_ratio = ratio
                best_y = y
        if best_ratio >= min_line_ratio:
            hits.append(
                {
                    "boundary_row_index": int(boundary_idx),
                    "expected_y": round(float(row_edges[boundary_idx]), 3),
                    "detected_y": int(best_y),
                    "line_ratio": round(float(best_ratio), 4),
                }
            )
    return hits


def _match_axis_by_order(
    predicted: np.ndarray,
    detected: np.ndarray,
    *,
    max_distance: float,
) -> tuple[np.ndarray, dict[str, Any]]:
    if len(predicted) == 0 or len(detected) == 0:
        return np.zeros_like(predicted), {"matched": 0, "detected": int(len(detected))}
    pred_order = np.argsort(predicted)
    det_order = np.argsort(detected)
    predicted_sorted = predicted[pred_order]
    detected_sorted = detected[det_order]
    pairs: list[tuple[int, int, float]] = []
    for pi, p in enumerate(predicted_sorted):
        distances = np.abs(detected_sorted - p)
        di = int(np.argmin(distances))
        distance = float(distances[di])
        if distance <= max_distance:
            pairs.append((pi, di, distance))
    pairs.sort(key=lambda item: item[2])
    used_p: set[int] = set()
    used_d: set[int] = set()
    matched: list[tuple[float, float]] = []
    for pi, di, _distance in pairs:
        if pi in used_p or di in used_d:
            continue
        used_p.add(pi)
        used_d.add(di)
        matched.append((float(predicted_sorted[pi]), float(detected_sorted[di])))
    matched.sort(key=lambda item: item[0])
    if len(matched) < 2:
        return np.zeros_like(predicted), {"matched": len(matched), "detected": int(len(detected))}
    xp = np.array([p for p, _d in matched], dtype=np.float32)
    fp = np.array([d - p for p, d in matched], dtype=np.float32)
    offsets_sorted = np.interp(predicted_sorted, xp, fp)
    offsets_sorted = np.clip(offsets_sorted, -max_distance, max_distance)
    offsets = np.zeros_like(predicted)
    offsets[pred_order] = offsets_sorted.astype(np.float32)
    return offsets, {
        "matched": int(len(matched)),
        "detected": int(len(detected)),
        "mean_abs_offset": round(float(np.mean(np.abs(offsets))), 3),
        "max_abs_offset": round(float(np.max(np.abs(offsets))), 3),
    }


def _axis_identity_bounds(values: np.ndarray, idx: int) -> tuple[float, float]:
    if len(values) == 1:
        return float(values[idx] - 80.0), float(values[idx] + 80.0)
    if idx == 0:
        gap = float(values[1] - values[0])
        return float(values[0] - 0.55 * gap), float((values[0] + values[1]) / 2.0)
    if idx == len(values) - 1:
        gap = float(values[-1] - values[-2])
        return float((values[-2] + values[-1]) / 2.0), float(values[-1] + 0.55 * gap)
    return float((values[idx - 1] + values[idx]) / 2.0), float((values[idx] + values[idx + 1]) / 2.0)


def _nearest_line_positions_in_rectified(
    rectified_bgr: np.ndarray,
    xs: list[int],
    ys: list[int],
) -> tuple[np.ndarray, np.ndarray, dict[str, Any]]:
    h_mask, v_mask = _split_line_masks(rectified_bgr)
    h_mask = cv2.dilate(h_mask, cv2.getStructuringElement(cv2.MORPH_RECT, (9, 3)), iterations=1)
    v_mask = cv2.dilate(v_mask, cv2.getStructuringElement(cv2.MORPH_RECT, (3, 9)), iterations=1)
    xs_arr = np.array(xs, dtype=np.float32)
    ys_arr = np.array(ys, dtype=np.float32)
    source_x = np.tile(xs_arr.reshape(1, -1), (len(ys), 1)).astype(np.float32)
    source_y = np.tile(ys_arr.reshape(-1, 1), (1, len(xs))).astype(np.float32)
    h_hits = 0
    v_hits = 0
    height, width = h_mask.shape[:2]

    def axis_band(values: list[int], idx: int, *, floor: int, cap: int, scale: float) -> int:
        if len(values) < 2:
            return floor
        gaps: list[float] = []
        if idx > 0:
            gaps.append(float(values[idx] - values[idx - 1]))
        if idx < len(values) - 1:
            gaps.append(float(values[idx + 1] - values[idx]))
        return int(max(floor, min(cap, min(gaps) * scale)))

    def fill_invalid_bands(values: np.ndarray, valid: np.ndarray) -> np.ndarray:
        out = values.copy()
        valid_indices = np.where(valid)[0]
        if len(valid_indices) == 0:
            return out
        all_indices = np.arange(values.shape[0], dtype=np.float32)
        for col in range(values.shape[1]):
            out[:, col] = np.interp(
                all_indices,
                valid_indices.astype(np.float32),
                values[valid_indices, col].astype(np.float32),
            )
        return out

    horizontal_valid = np.zeros(len(xs), dtype=bool)
    for col_idx, x in enumerate(xs):
        half_w = axis_band(xs, col_idx, floor=8, cap=42, scale=0.42)
        x0 = max(0, int(round(x)) - half_w)
        x1 = min(width, int(round(x)) + half_w + 1)
        roi = h_mask[:, x0:x1]
        if roi.size == 0:
            continue
        projection = roi.sum(axis=1) / 255.0
        detected = np.array(
            _clustered_projection_positions(
                projection,
                threshold_ratio=0.12,
                min_value=max(5.0, (x1 - x0) * 0.16),
            ),
            dtype=np.float32,
        )
        offsets, match_ev = _match_axis_by_order(ys_arr, detected, max_distance=34.0)
        if int(match_ev["matched"]) < max(3, len(ys) // 6):
            continue
        source_y[:, col_idx] = ys_arr + offsets
        horizontal_valid[col_idx] = True
        h_hits += int(match_ev["matched"])

    if horizontal_valid.any():
        source_y = fill_invalid_bands(source_y.T, horizontal_valid).T
        source_y = cv2.GaussianBlur(source_y.astype(np.float32), (3, 1), 0)

    vertical_valid = np.zeros(len(ys), dtype=bool)
    for row_idx, y in enumerate(ys):
        half_h = axis_band(ys, row_idx, floor=7, cap=28, scale=0.45)
        y0 = max(0, int(round(y)) - half_h)
        y1 = min(height, int(round(y)) + half_h + 1)
        roi = v_mask[y0:y1, :]
        if roi.size == 0:
            continue
        projection = roi.sum(axis=0) / 255.0
        detected = np.array(
            _clustered_projection_positions(
                projection,
                threshold_ratio=0.12,
                min_value=max(5.0, (y1 - y0) * 0.16),
            ),
            dtype=np.float32,
        )
        offsets, match_ev = _match_axis_by_order(xs_arr, detected, max_distance=48.0)
        if int(match_ev["matched"]) < max(3, len(xs) // 4):
            continue
        source_x[row_idx, :] = xs_arr + offsets
        vertical_valid[row_idx] = True
        v_hits += int(match_ev["matched"])

    if vertical_valid.any():
        source_x = fill_invalid_bands(source_x, vertical_valid)
        source_x = cv2.GaussianBlur(source_x.astype(np.float32), (1, 5), 0)

    for row_idx in range(source_x.shape[0]):
        for col_idx in range(source_x.shape[1]):
            lower, upper = _axis_identity_bounds(xs_arr, col_idx)
            source_x[row_idx, col_idx] = min(max(source_x[row_idx, col_idx], lower), upper)
        source_x[row_idx] = np.maximum.accumulate(source_x[row_idx])
    for col_idx in range(source_y.shape[1]):
        for row_idx in range(source_y.shape[0]):
            lower, upper = _axis_identity_bounds(ys_arr, row_idx)
            source_y[row_idx, col_idx] = min(max(source_y[row_idx, col_idx], lower), upper)
        source_y[:, col_idx] = np.maximum.accumulate(source_y[:, col_idx])
    return source_x, source_y, {
        "method": "fixed_template_grid_ordered_band_registration",
        "horizontal_hits": int(h_hits),
        "vertical_hits": int(v_hits),
        "horizontal_valid_bands": int(horizontal_valid.sum()),
        "vertical_valid_bands": int(vertical_valid.sum()),
        "grid_points": int(len(xs) * len(ys)),
        "template_horizontal_lines": int(len(ys)),
        "template_vertical_lines": int(len(xs)),
    }


def _vertical_line_span_evidence(
    vertical_mask: np.ndarray,
    *,
    x: int,
    y0: int,
    y1: int,
    band_count: int = 14,
) -> dict[str, Any]:
    height = max(1, int(y1 - y0))
    x0 = max(0, int(x) - 2)
    x1 = min(vertical_mask.shape[1], int(x) + 3)
    strip = vertical_mask[max(0, y0) : min(vertical_mask.shape[0], y1), x0:x1]
    if strip.size == 0:
        return {
            "coverage_ratio": 0.0,
            "hit_band_ratio": 0.0,
            "top_band_hit": False,
            "bottom_band_hit": False,
            "valid": False,
        }
    row_hits = (strip.sum(axis=1) / 255.0) >= 1.0
    coverage_ratio = float(row_hits.sum()) / float(height)
    bands = np.array_split(row_hits, max(1, int(band_count)))
    band_hits = [bool(band.size and band.any()) for band in bands]
    hit_band_ratio = float(sum(1 for hit in band_hits if hit)) / float(len(band_hits))
    top_band_hit = any(band_hits[:2])
    bottom_band_hit = any(band_hits[-2:])
    return {
        "coverage_ratio": round(coverage_ratio, 4),
        "hit_band_ratio": round(hit_band_ratio, 4),
        "top_band_hit": top_band_hit,
        "bottom_band_hit": bottom_band_hit,
        "valid": bool(hit_band_ratio >= 0.72 and top_band_hit and bottom_band_hit),
    }


def _detect_vertical_candidates(rectified_bgr: np.ndarray, table_bbox: list[int]) -> list[float]:
    _h_mask, v_mask = _split_line_masks(rectified_bgr)
    v_mask = cv2.dilate(v_mask, cv2.getStructuringElement(cv2.MORPH_RECT, (2, 5)), iterations=1)
    x0, y0, x1, y1 = [int(round(v)) for v in table_bbox]
    roi = v_mask[max(0, y0) : min(v_mask.shape[0], y1), max(0, x0) : min(v_mask.shape[1], x1)]
    projection = roi.sum(axis=0) / 255.0
    xs = _clustered_projection_positions(
        projection,
        threshold_ratio=0.08,
        min_value=max(6.0, (y1 - y0) * 0.030),
    )
    candidates = [float(x0 + x) for x in xs]
    return [
        candidate
        for candidate in candidates
        if _vertical_line_span_evidence(
            v_mask,
            x=int(round(candidate)),
            y0=max(0, y0),
            y1=min(v_mask.shape[0], y1),
        )["valid"]
    ]


def _gap_order_match(template_xs: list[int], candidates: list[float]) -> tuple[list[float], dict[str, Any]]:
    expected_count = len(template_xs)
    if expected_count < 2:
        return [float(value) for value in template_xs], {"used": False, "reason": "not_enough_template_axes"}
    sorted_candidates = sorted(float(value) for value in candidates)
    if len(sorted_candidates) < expected_count:
        return [float(value) for value in template_xs], {
            "used": False,
            "reason": "candidate_count_less_than_template_count",
            "candidate_count": len(sorted_candidates),
            "template_count": expected_count,
        }
    template = np.array(template_xs, dtype=np.float32)
    template_gaps = np.diff(template)
    template_norm = template_gaps / max(1.0, float(template[-1] - template[0]))
    best_score = float("inf")
    best_indexes: tuple[int, ...] | None = None
    max_combinations = 250_000
    checked = 0
    if len(sorted_candidates) == expected_count:
        combinations_iter = [tuple(range(expected_count))]
    else:
        # The accepted pipeline keeps the detected outer frame fixed. Extra
        # vertical candidates are only allowed to be dropped inside the frame.
        combinations_iter = (
            (0, *middle, len(sorted_candidates) - 1)
            for middle in itertools.combinations(range(1, len(sorted_candidates) - 1), expected_count - 2)
        )
    for indexes in combinations_iter:
        checked += 1
        if checked > max_combinations:
            break
        selected = np.array([sorted_candidates[index] for index in indexes], dtype=np.float32)
        if not np.all(np.diff(selected) > 0):
            continue
        gaps = np.diff(selected)
        span = max(1.0, float(selected[-1] - selected[0]))
        norm = gaps / span
        score = float(np.mean(np.abs(norm - template_norm)))
        score += 0.0001 * abs(span - float(template[-1] - template[0])) / max(1.0, float(template[-1] - template[0]))
        if score < best_score:
            best_score = score
            best_indexes = indexes
    if best_indexes is None:
        return [float(value) for value in template_xs], {"used": False, "reason": "no_valid_gap_order_match"}
    selected_values = [sorted_candidates[index] for index in best_indexes]
    skipped_indexes = [index for index in range(len(sorted_candidates)) if index not in best_indexes]
    return selected_values, {
        "used": True,
        "selected_indexes": list(best_indexes),
        "skipped_indexes": skipped_indexes,
        "selected_values": [round(float(value), 3) for value in selected_values],
        "skipped_values": [round(float(sorted_candidates[index]), 3) for index in skipped_indexes],
        "score": round(best_score, 6),
    }


def _dedupe_close_line_candidates(candidates: list[float], *, min_gap_px: float = 58.0) -> list[float]:
    sorted_candidates = sorted(float(value) for value in candidates)
    if not sorted_candidates:
        return []
    groups: list[list[float]] = [[sorted_candidates[0]]]
    for value in sorted_candidates[1:]:
        if value - groups[-1][-1] < min_gap_px:
            groups[-1].append(value)
        else:
            groups.append([value])
    return [float(group[0]) for group in groups]


def _post_menu_boundary_preserving_xs(
    *,
    worksheet: Any | None,
    matched_xs: list[float],
    fax_x_candidates: list[float],
    fax_template: dict[str, Any] | None = None,
) -> tuple[list[float], dict[str, Any]]:
    if worksheet is None or len(matched_xs) < 3:
        return matched_xs, {"used": False, "reason": "worksheet_unavailable"}
    col_count = min(int(getattr(worksheet, "max_column", 0) or 0), len(matched_xs) - 1)
    if col_count < 2:
        return matched_xs, {"used": False, "reason": "not_enough_columns"}
    slots = hakodate_assignment_service._column_slots_from_worksheet(  # noqa: SLF001
        worksheet,
        col_count=col_count,
        template=fax_template,
    )
    menu_col = _menu_column_from_slots(slots, worksheet)
    post_start_edge_index = int(menu_col)
    if post_start_edge_index <= 0 or post_start_edge_index >= len(matched_xs) - 1:
        return matched_xs, {
            "used": False,
            "reason": "post_menu_edge_out_of_range",
            "menu_worksheet_col": menu_col,
            "edge_count": len(matched_xs),
        }

    previous_edge = float(matched_xs[post_start_edge_index - 1])
    right_edge = float(matched_xs[-1])
    post_candidates = [
        float(value)
        for value in fax_x_candidates
        if previous_edge + 40.0 <= float(value) <= right_edge + 45.0
    ]
    deduped = _dedupe_close_line_candidates(post_candidates)
    required_count = len(matched_xs) - post_start_edge_index
    if len(deduped) < required_count:
        return matched_xs, {
            "used": False,
            "reason": "post_menu_candidate_count_shortfall",
            "menu_worksheet_col": menu_col,
            "required_count": required_count,
            "candidate_count": len(deduped),
            "candidates": [round(float(value), 3) for value in deduped],
        }

    if len(deduped) == required_count:
        selected_post_edges = deduped
    else:
        # Preserve the first physical quantity boundary after the menu.  If the
        # detected FAX has more post-menu lines than the canonical target grid,
        # the surplus belongs to a trailing absorbed column such as remarks,
        # not to the first quantity column.
        selected_post_edges = [*deduped[: required_count - 1], deduped[-1]]
    adjusted = [float(value) for value in matched_xs]
    adjusted[post_start_edge_index:] = [float(value) for value in selected_post_edges]
    if len(adjusted) != len(matched_xs) or not np.all(np.diff(np.array(adjusted, dtype=np.float32)) > 0):
        return matched_xs, {
            "used": False,
            "reason": "post_menu_adjustment_not_monotonic",
            "menu_worksheet_col": menu_col,
            "selected_post_edges": [round(float(value), 3) for value in selected_post_edges],
        }
    return adjusted, {
        "used": True,
        "menu_worksheet_col": menu_col,
        "post_start_edge_index": post_start_edge_index,
        "required_count": required_count,
        "candidate_count": len(deduped),
        "raw_candidates": [round(float(value), 3) for value in post_candidates],
        "deduped_candidates": [round(float(value), 3) for value in deduped],
        "selected_post_edges": [round(float(value), 3) for value in selected_post_edges],
        "previous_first_post_edge": round(float(matched_xs[post_start_edge_index]), 3),
        "adjusted_first_post_edge": round(float(adjusted[post_start_edge_index]), 3),
    }


def _align_axes(
    *,
    rectified_fax: np.ndarray,
    template_xs: list[int],
    template_ys: list[int],
    worksheet: Any | None = None,
    fax_template: dict[str, Any] | None = None,
    header_axis_override: dict[str, Any] | None = None,
    row_axis_override: dict[str, Any] | None = None,
) -> tuple[list[float], list[float], dict[str, Any], Image.Image]:
    table_bbox = [template_xs[0], template_ys[0], template_xs[-1], template_ys[-1]]
    source_x, source_y, grid_evidence = _nearest_line_positions_in_rectified(rectified_fax, template_xs, template_ys)
    fax_x_candidates = _detect_vertical_candidates(rectified_fax, table_bbox)
    matched_xs, x_match = _gap_order_match(template_xs, fax_x_candidates)
    if not x_match.get("used"):
        matched_xs = np.maximum.accumulate(np.median(source_x, axis=0)).astype(np.float32).tolist()
    header_corrected_xs, header_x_match = _header_intersection_correct_xs(
        rectified_fax=rectified_fax,
        template_xs=template_xs,
        template_ys=template_ys,
    )
    if header_corrected_xs is not None:
        matched_xs = [float(value) for value in header_corrected_xs]
    manual_header_xs = normalize_header_axis_override(
        header_axis_override,
        expected_count=len(template_xs),
        canvas_width=rectified_fax.shape[1],
    )
    if manual_header_xs is not None:
        matched_xs = [float(value) for value in manual_header_xs]
        header_x_match = {
            **(header_x_match if isinstance(header_x_match, dict) else {}),
            "used": True,
            "method": "operator_header_axis_override",
            "reason": "applied_operator_header_axis_override",
            "corrected_xs": [round(float(value), 3) for value in matched_xs],
            "template_x_count": len(template_xs),
            "coordinate_space": {
                "mode": "template_canvas",
                "width": int(rectified_fax.shape[1]),
                "height": int(rectified_fax.shape[0]),
            },
        }
    matched_xs, post_menu_x_match = _post_menu_boundary_preserving_xs(
        worksheet=worksheet,
        matched_xs=[float(value) for value in matched_xs],
        fax_x_candidates=fax_x_candidates,
        fax_template=fax_template,
    )
    adjusted_ys = np.maximum.accumulate(np.median(source_y, axis=1)).astype(np.float32).tolist()
    row_corrected_ys, row_y_match = _row_intersection_correct_ys(
        rectified_fax=rectified_fax,
        corrected_xs=[float(value) for value in matched_xs],
        template_ys=template_ys,
    )
    if row_corrected_ys is not None:
        adjusted_ys = [float(value) for value in row_corrected_ys]
    manual_row_ys = normalize_row_axis_override(
        row_axis_override,
        expected_count=len(template_ys),
        canvas_height=rectified_fax.shape[0],
    )
    if manual_row_ys is not None:
        adjusted_ys = [float(value) for value in manual_row_ys]
        row_y_match = {
            **(row_y_match if isinstance(row_y_match, dict) else {}),
            "used": True,
            "method": "operator_row_axis_override",
            "reason": "applied_operator_row_axis_override",
            "corrected_ys": [round(float(value), 3) for value in adjusted_ys],
            "template_y_count": len(template_ys),
            "row_height_quality": _row_height_outlier_evidence(adjusted_ys),
            "coordinate_space": {
                "mode": "template_canvas",
                "width": int(rectified_fax.shape[1]),
                "height": int(rectified_fax.shape[0]),
            },
        }
    axis_image = _draw_axis_match(
        rectified_fax,
        template_xs=template_xs,
        template_ys=template_ys,
        fax_x_candidates=fax_x_candidates,
        matched_xs=matched_xs,
        adjusted_ys=adjusted_ys,
        header_x_match=header_x_match,
    )
    evidence = {
        "x_match": x_match,
        "header_intersection_x_match": header_x_match,
        "row_intersection_y_match": row_y_match,
        "post_menu_x_match": post_menu_x_match,
        "grid": grid_evidence,
        "x_gaps": [round(float(b - a), 3) for a, b in zip(matched_xs, matched_xs[1:])],
        "y_gaps": [round(float(b - a), 3) for a, b in zip(adjusted_ys, adjusted_ys[1:])],
        "counts": {
            "template_x": len(template_xs),
            "aligned_x": len(matched_xs),
            "template_y": len(template_ys),
            "aligned_y": len(adjusted_ys),
        },
    }
    return matched_xs, adjusted_ys, evidence, axis_image


def _draw_axis_match(
    rectified_fax: np.ndarray,
    *,
    template_xs: list[int],
    template_ys: list[int],
    fax_x_candidates: list[float],
    matched_xs: list[float],
    adjusted_ys: list[float],
    header_x_match: dict[str, Any] | None = None,
) -> Image.Image:
    image = _bgr_to_rgb_image(rectified_fax).convert("RGBA")
    layer = Image.new("RGBA", image.size, (0, 0, 0, 0))
    draw = ImageDraw.Draw(layer)
    y0, y1 = int(round(template_ys[0])), int(round(template_ys[-1]))
    x0, x1 = int(round(template_xs[0])), int(round(template_xs[-1]))
    for x in matched_xs:
        draw.line((int(round(x)), y0, int(round(x)), y1), fill=(0, 190, 0, 230), width=3)
    for y in adjusted_ys:
        draw.line((x0, int(round(y)), x1, int(round(y))), fill=(0, 190, 0, 210), width=2)
    return Image.alpha_composite(image, layer).convert("RGB")


def _draw_row_intersections_overlay(
    *,
    image: Image.Image,
    axis_evidence: dict[str, Any],
) -> Image.Image:
    match = axis_evidence.get("row_intersection_y_match") if isinstance(axis_evidence, dict) else {}
    if not isinstance(match, dict):
        return image
    points = match.get("intersection_points")
    if not isinstance(points, list):
        points = []
    clusters = match.get("fax_y_clusters")
    if not isinstance(clusters, list):
        clusters = []
    corrected_ys = match.get("corrected_ys")
    if not isinstance(corrected_ys, list):
        corrected_ys = []
    table_roi = match.get("table_roi")
    menu_roi = match.get("menu_roi")
    overlay = image.convert("RGBA")
    layer = Image.new("RGBA", overlay.size, (0, 0, 0, 0))
    draw = ImageDraw.Draw(layer)
    if isinstance(table_roi, list) and len(table_roi) == 4:
        x0, y0, x1, y1 = [int(round(float(value))) for value in table_roi]
        draw.rectangle([x0, y0, x1, y1], outline=(255, 140, 0, 150), width=4)
    if isinstance(menu_roi, list) and len(menu_roi) == 4:
        x0, y0, x1, y1 = [int(round(float(value))) for value in menu_roi]
        draw.rectangle([x0, y0, x1, y1], outline=(255, 140, 0, 240), width=6)
    for point in points:
        if not isinstance(point, dict):
            continue
        try:
            x = float(point["x"])
            y = float(point["y"])
        except (TypeError, ValueError, KeyError):
            continue
        draw.ellipse([x - 4, y - 4, x + 4, y + 4], outline=(255, 0, 0, 210), width=2)
    x_start = 0
    x_end = overlay.size[0]
    if isinstance(table_roi, list) and len(table_roi) == 4:
        x_start = int(round(float(table_roi[0])))
        x_end = int(round(float(table_roi[2])))
    for cluster in clusters:
        if not isinstance(cluster, dict):
            continue
        try:
            y = float(cluster.get("value"))
        except (TypeError, ValueError):
            continue
        vote_count = int(cluster.get("column_vote_count") or 0)
        width = 1 if vote_count < 4 else 3
        color = (255, 0, 0, 110) if vote_count < 4 else (255, 0, 0, 190)
        draw.line([x_start, y, x_end, y], fill=color, width=width)
    for raw_y in corrected_ys:
        try:
            y = float(raw_y)
        except (TypeError, ValueError):
            continue
        draw.line([0, y, overlay.size[0], y], fill=(0, 190, 70, 210), width=2)
    return Image.alpha_composite(overlay, layer).convert("RGB")


def _draw_merge_aware_grid(
    *,
    worksheet: Any,
    rectified_fax: np.ndarray,
    xs: list[float],
    ys: list[float],
    horizontal_line_mask: np.ndarray | None = None,
) -> tuple[Image.Image, dict[str, Any]]:
    image = _bgr_to_rgb_image(rectified_fax).convert("RGBA")
    layer = Image.new("RGBA", image.size, (0, 0, 0, 0))
    draw = ImageDraw.Draw(layer)
    quantity_columns = hakodate_assignment_service._worksheet_quantity_column_indexes(worksheet)  # noqa: SLF001
    merge_regions = _step_review_merge_regions_for_grid(
        worksheet,
        row_edges=[float(value) for value in ys],
        column_edges=[float(value) for value in xs],
        quantity_columns=quantity_columns,
    )
    segments = hakodate_assignment_service._merge_aware_grid_line_segments(  # noqa: SLF001
        row_edges=[float(value) for value in ys],
        column_edges=[float(value) for value in xs],
        merge_regions=merge_regions,
    )
    for segment in segments:
        start = segment.get("start") or []
        end = segment.get("end") or []
        if len(start) != 2 or len(end) != 2:
            continue
        draw.line(
            (int(round(float(start[0]))), int(round(float(start[1]))), int(round(float(end[0]))), int(round(float(end[1])))),
            fill=(0, 180, 0, 230),
            width=2,
        )
    return Image.alpha_composite(image, layer).convert("RGB"), {
        "merge_region_count": len(merge_regions),
        "retained_merge_region_count": len(merge_regions),
        "physically_split_merge_region_count": 0,
        "physically_split_merge_regions": [],
        "merge_aware_segment_count": len(segments),
        "merged_ranges": [str(region.get("range") or "") for region in merge_regions],
    }


def _menu_column_from_slots(slots: list[dict[str, Any]], worksheet: Any) -> int:
    for slot in slots:
        if slot.get("role") == "menu_name":
            return int(slot.get("worksheet_col_index") or 4)
    for row in (7, 8, 9):
        for col in range(1, int(worksheet.max_column or 0) + 1):
            if "献立" in str(worksheet.cell(row=row, column=col).value or ""):
                return col
    return 4


def _post_menu_target_regions(
    *,
    worksheet: Any,
    column_edges: list[float],
    row_edges: list[float],
    fax_template: dict[str, Any] | None = None,
    horizontal_line_mask: np.ndarray | None = None,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    col_count = min(int(worksheet.max_column or 0), len(column_edges) - 1)
    slots = hakodate_assignment_service._column_slots_from_worksheet(  # noqa: SLF001
        worksheet,
        col_count=col_count,
        template=fax_template,
    )
    slot_by_col = {int(slot.get("worksheet_col_index") or 0): slot for slot in slots if isinstance(slot, dict)}
    menu_col = _menu_column_from_slots(slots, worksheet)
    canonical_target_cols = [
        col
        for col in range(menu_col + 1, col_count + 1)
        if str((slot_by_col.get(col) or {}).get("canonical_source") or "").strip() == "facility_fax_template"
        and (
            str((slot_by_col.get(col) or {}).get("role") or "").strip() in {"quantity", "note"}
            or str((slot_by_col.get(col) or {}).get("slot_name") or "").strip().startswith("qty.")
            or str((slot_by_col.get(col) or {}).get("slot_name") or "").strip() in {"note", "remarks"}
        )
    ]
    # The OCR target map is a geometric map: all physical cells to the right of
    # the menu column must stay visible to OCR.  Downstream sheet/materialization
    # logic decides how totals, notes, and helper columns aggregate.
    target_cols = list(range(menu_col + 1, col_count + 1))
    physical_row_map = _step_review_physical_row_map(worksheet, row_count=len(row_edges) - 1)
    merged_cells = hakodate_assignment_service._worksheet_merged_cell_map(worksheet)  # noqa: SLF001
    by_region_id: dict[str, dict[str, Any]] = {}
    blank_menu_row_count = 0
    for row_index, row_meta in sorted(physical_row_map.items()):
        worksheet_row = int(row_meta.get("worksheet_row") or 0)
        if worksheet_row <= 0 or int(row_index) >= len(row_edges) - 1:
            continue
        if not str(row_meta.get("menu_key") or "").strip():
            blank_menu_row_count += 1
            continue
        for worksheet_col in target_cols:
            col_index = worksheet_col - 1
            if col_index >= len(column_edges) - 1:
                continue
            slot = slot_by_col.get(worksheet_col) or {}
            label = str(slot.get("label") or "").strip()
            role = str(slot.get("role") or "post_menu").strip()
            field = str(slot.get("slot_name") or "").strip()
            if not label or label == "spacer":
                label = f"空白列({get_column_letter(worksheet_col)})"
            if not field or field == "spacer":
                field = f"post_menu.{get_column_letter(worksheet_col)}"
            bbox, merged = _step_review_merged_or_single_cell_bbox(
                row_index=int(row_index),
                col_index=col_index,
                worksheet_row=worksheet_row,
                worksheet_col=worksheet_col,
                row_edges=row_edges,
                column_edges=column_edges,
                merged_cells=merged_cells,
            )
            sheet_cell = f"{get_column_letter(worksheet_col)}{worksheet_row}"
            physical_split_from_excel_merge: dict[str, Any] | None = None
            region_id = str(merged.get("range") if merged else sheet_cell)
            logical_target = {
                "sheet_cell": sheet_cell,
                "worksheet_row": worksheet_row,
                "worksheet_col": worksheet_col,
                "grid_row_index": int(row_index),
                "grid_col_index": col_index,
                "role": role,
                "field": field,
                "field_label": label,
                "date": row_meta.get("effective_date") or row_meta.get("date"),
                "daypart": row_meta.get("daypart"),
                "menu_name": row_meta.get("menu_name"),
            }
            if region_id not in by_region_id:
                by_region_id[region_id] = {
                    "region_id": region_id,
                    "sheet_cell": sheet_cell,
                    "worksheet_row": worksheet_row,
                    "worksheet_col": worksheet_col,
                    "grid_row_index": int(row_index),
                    "grid_col_index": col_index,
                    "role": role,
                    "field": field,
                    "field_label": label,
                    "bbox": bbox,
                    "merged_cell": merged,
                    "physical_split_from_excel_merge": physical_split_from_excel_merge,
                    "logical_targets": [],
                    "covered_sheet_cells": [],
                }
            region = by_region_id[region_id]
            region["logical_targets"].append(logical_target)
            if sheet_cell not in region["covered_sheet_cells"]:
                region["covered_sheet_cells"].append(sheet_cell)
    regions = list(by_region_id.values())
    label_counts: dict[str, int] = {}
    for region in regions:
        label = str(region.get("field_label") or "")
        label_counts[label] = label_counts.get(label, 0) + max(1, len(region.get("logical_targets") or []))
    return regions, {
        "target_rule": TARGET_RULE,
        "menu_worksheet_col": menu_col,
        "target_worksheet_cols": target_cols,
        "canonical_target_worksheet_cols": canonical_target_cols,
        "template_column_restricted": False,
        "target_selection_mode": "all_physical_columns_right_of_menu",
        "blank_menu_row_count": blank_menu_row_count,
        "region_count": len(regions),
        "logical_target_count": sum(len(region.get("logical_targets") or []) for region in regions),
        "label_counts": label_counts,
        "physical_split_excel_merge_count": 0,
        "physical_split_excel_merge_ranges": [],
    }


def _draw_target_regions(
    *,
    grid_overlay: Image.Image,
    regions: list[dict[str, Any]],
) -> Image.Image:
    image = grid_overlay.convert("RGBA")
    layer = Image.new("RGBA", image.size, (0, 0, 0, 0))
    draw = ImageDraw.Draw(layer)
    font = ImageFont.load_default()
    for region in regions:
        box = region.get("bbox")
        if not isinstance(box, list) or len(box) != 4:
            continue
        x0, y0, x1, y1 = [int(round(float(value))) for value in box]
        cx = int(round((x0 + x1) / 2.0))
        cy = int(round((y0 + y1) / 2.0))
        draw.rectangle((x0, y0, x1, y1), outline=(0, 120, 255, 150), width=1)
        draw.ellipse((cx - 8, cy - 8, cx + 8, cy + 8), fill=(255, 255, 255, 230))
        draw.ellipse((cx - 6, cy - 6, cx + 6, cy + 6), fill=(255, 0, 0, 255))
        label = str(region.get("field_label") or "")
        if label:
            draw.text((cx + 5, cy - 5), label[:8], fill=(220, 0, 0, 210), font=font)
    return Image.alpha_composite(image, layer).convert("RGB")


def build_hakodate_step_review_for_manifest_item(
    *,
    item: dict[str, Any],
    page: int,
    output_dir: Path,
    render_width: int = 1864,
) -> tuple[HakodateStepReviewResult, dict[str, Image.Image]]:
    facility_code = str(item["facility_code"])
    order_id = str(item["order_id"])
    existing_step2 = cv2.imread(item["step2_png"])
    if existing_step2 is None:
        raise ValueError(f"step2 canvas not found: {item['step2_png']}")
    canvas_height, canvas_width = existing_step2.shape[:2]
    template = render_template_pdf_to_canvas(item["template_pdf"], width=canvas_width, height=canvas_height)
    template_xs, template_ys, _all_xs, _all_ys = resolve_template_axes_from_manifest_or_image(
        item=item,
        template_image=template,
        manifest_template_bbox=item["template_bbox"],
    )
    week_sheet_name = str(item.get("week_sheet_name") or WEEK_SHEET_NAME).strip() or WEEK_SHEET_NAME
    selected_template_id = str(
        item.get("template_id")
        or item.get("fax_template_id")
        or item.get("resolved_template_id")
        or ""
    ).strip() or None
    facility_config = config_service.get_facility_config_for_template(
        facility_code,
        selected_template_id,
    )
    fax_template = (
        facility_config.get("fax_template")
        if isinstance(facility_config, dict) and isinstance(facility_config.get("fax_template"), dict)
        else None
    )
    worksheet = hakodate_assignment_service._worksheet_for_manifest_structure_template(  # noqa: SLF001
        item=item,
        facility_id=facility_code,
        week_sheet_name=week_sheet_name,
    )
    quad_px, quad_source, quad_estimate = resolve_fixed_quad_px_for_manifest_item(item, render_width=render_width)
    registration, step_images_np = build_fixed_quad_template_registration(
        facility_code=facility_code,
        order_id=order_id,
        fax_pdf=item["fax_pdf"],
        template_pdf=item["template_pdf"],
        quad_px=quad_px,
        manifest_template_bbox=item["template_bbox"],
        canvas_width=canvas_width,
        canvas_height=canvas_height,
        render_width=render_width,
        quad_source=quad_source,
        output_dir=None,
        template_axes_x=template_xs,
        template_axes_y=template_ys,
    )
    original = render_pdf_page_to_bgr(item["fax_pdf"], width=render_width)
    # step2 image contains the red bbox annotation. Rebuild the raw rectified FAX
    # by using the accepted registration service's output dimensions and quad.
    table_bbox = registration.template_outer_grid_bbox_used
    rectified_quad_points = _bbox_quad_points(table_bbox)

    raw_rectified = rectify_fax_to_template_grid(
        original,
        quad_px=quad_px,
        table_bbox=table_bbox,
        canvas_width=canvas_width,
        canvas_height=canvas_height,
    )
    horizontal_line_mask, _vertical_line_mask = _split_line_masks(raw_rectified)
    line_extraction, line_evidence = _draw_line_extraction(raw_rectified)
    aligned_xs, aligned_ys, axis_evidence, axis_match_image = _align_axes(
        rectified_fax=raw_rectified,
        template_xs=template_xs,
        template_ys=template_ys,
        worksheet=worksheet,
        fax_template=fax_template,
        header_axis_override=item.get("header_axis_override") if isinstance(item.get("header_axis_override"), dict) else None,
        row_axis_override=item.get("row_axis_override") if isinstance(item.get("row_axis_override"), dict) else None,
    )
    grid_overlay, merge_evidence = _draw_merge_aware_grid(
        worksheet=worksheet,
        rectified_fax=raw_rectified,
        xs=aligned_xs,
        ys=aligned_ys,
        horizontal_line_mask=horizontal_line_mask,
    )
    target_regions, target_evidence = _post_menu_target_regions(
        worksheet=worksheet,
        column_edges=[float(value) for value in aligned_xs],
        row_edges=[float(value) for value in aligned_ys],
        fax_template=fax_template,
        horizontal_line_mask=horizontal_line_mask,
    )
    target_overlay = _draw_target_regions(
        grid_overlay=grid_overlay,
        regions=target_regions,
    )
    target_overlay = _draw_row_intersections_overlay(image=target_overlay, axis_evidence=axis_evidence)
    source_template = _source_template_name(facility_code)
    axis_evidence = {
        **axis_evidence,
        "line_extraction": line_evidence,
        "merge": merge_evidence,
        "target": target_evidence,
        "quad_estimate": quad_estimate,
    }
    details = [
        f"source_template={source_template}",
        f"target_rule={TARGET_RULE}",
        f"targets={target_evidence['region_count']} logical={target_evidence['logical_target_count']} cols={target_evidence['target_worksheet_cols']}",
        "quad markers: source on STEP1, mapped target corners on STEP2-STEP6",
    ]
    images = {
        "step1": _make_review_canvas(
            title="STEP1 original FAX + accepted 4 points",
            facility_code=facility_code,
            order_id=order_id,
            image=_bgr_to_rgb_image(step_images_np["step1"]),
            details=details,
        ),
        "step2": _make_review_canvas(
            title="STEP2 FAX rectified by accepted 4 points",
            facility_code=facility_code,
            order_id=order_id,
            image=_draw_quad_points(_bgr_to_rgb_image(step_images_np["step2"]), rectified_quad_points, prefix="Q"),
            details=details,
        ),
        "step3": _make_review_canvas(
            title="STEP3 extracted FAX lines on rectified FAX",
            facility_code=facility_code,
            order_id=order_id,
            image=_draw_quad_points(line_extraction, rectified_quad_points, prefix="Q"),
            details=details,
        ),
        "step4": _make_review_canvas(
            title="STEP4 final adopted axes only (green)",
            facility_code=facility_code,
            order_id=order_id,
            image=_draw_quad_points(axis_match_image, rectified_quad_points, prefix="Q"),
            details=details,
        ),
        "step5": _make_review_canvas(
            title="STEP5 rectified FAX + merge-aware green grid",
            facility_code=facility_code,
            order_id=order_id,
            image=_draw_quad_points(grid_overlay, rectified_quad_points, prefix="Q"),
            details=details,
        ),
        "step6": _make_review_canvas(
            title="STEP6 post-menu target cells: red centers, blue boxes",
            facility_code=facility_code,
            order_id=order_id,
            image=_draw_quad_points(target_overlay, rectified_quad_points, prefix="Q"),
            details=details,
        ),
    }
    case_dir = output_dir / f"{page:02d}_{facility_code}_{order_id}"
    case_dir.mkdir(parents=True, exist_ok=True)
    outputs: dict[str, str] = {}
    for key, image in images.items():
        path = case_dir / f"{key}.png"
        image.save(path)
        outputs[key] = str(path)
    regions_path = case_dir / "target_regions.json"
    regions_path.write_text(json.dumps(target_regions, ensure_ascii=False, indent=2), encoding="utf-8")
    outputs["target_regions"] = str(regions_path)
    return (
        HakodateStepReviewResult(
            page=page,
            facility_code=facility_code,
            order_id=order_id,
            fax_pdf=str(item["fax_pdf"]),
            template_pdf=str(item["template_pdf"]),
            source_template=source_template,
            target_rule=TARGET_RULE,
            target_worksheet_cols=list(target_evidence["target_worksheet_cols"]),
            region_count=int(target_evidence["region_count"]),
            logical_target_count=int(target_evidence["logical_target_count"]),
            label_counts=dict(target_evidence["label_counts"]),
            axis_evidence=axis_evidence,
            outputs=outputs,
        ),
        images,
    )


def build_all_facility_hakodate_step_review_pdfs(
    *,
    manifest_path: Path,
    output_dir: Path,
    render_width: int = 1864,
) -> dict[str, Any]:
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    items = manifest.get("results") if isinstance(manifest, dict) else manifest
    if not isinstance(items, list):
        raise ValueError("manifest results are missing")
    output_dir.mkdir(parents=True, exist_ok=True)
    pages_by_step: dict[str, list[Image.Image]] = {f"step{idx}": [] for idx in range(1, 7)}
    results: list[dict[str, Any]] = []
    for page, item in enumerate(items, start=1):
        result, images = build_hakodate_step_review_for_manifest_item(
            item=item,
            page=page,
            output_dir=output_dir,
            render_width=render_width,
        )
        for key, image in images.items():
            pages_by_step[key].append(image)
        results.append(asdict(result))
    pdfs: dict[str, str] = {}
    names = {
        "step1": "step1_original_fax_accepted_quad_all14.pdf",
        "step2": "step2_rectified_fax_by_accepted_quad_all14.pdf",
        "step3": "step3_extracted_fax_lines_all14.pdf",
        "step4": "step4_final_adopted_axes_only_all14.pdf",
        "step5": "step5_rectified_merge_aware_grid_all14.pdf",
        "step6": "step6_post_menu_target_red_points_all14.pdf",
    }
    for key, pages in pages_by_step.items():
        pdf_path = output_dir / names[key]
        _write_pdf_from_pages(pages, pdf_path)
        pdfs[key] = str(pdf_path)
    summary = {
        "count": len(results),
        "target_rule": TARGET_RULE,
        "pdfs": pdfs,
        "results": results,
    }
    summary_path = output_dir / "summary.json"
    summary_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    return summary
