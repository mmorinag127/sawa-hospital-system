import pathlib
import sys
from datetime import datetime
from types import SimpleNamespace

import pytest
from openpyxl import Workbook, load_workbook

ROOT = pathlib.Path(__file__).resolve().parents[2]
sys.path.append(str(ROOT))

from src.services import order_form_service  # noqa: E402


def _metadata_rows(workbook):
    return {key: value for key, value in workbook["設定"].iter_rows(min_row=2, values_only=True)}


def _build_source_workbook(path: pathlib.Path, sheet_name: str) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    worksheet["A2"] = "【発　注　連　絡　表】"
    worksheet["G3"] = "締切日1月1日まで"
    worksheet["A4"] = ""
    worksheet["D7"] = "献立"
    workbook.save(path)


class _FakeSessionScope:
    def __init__(self, order):
        self.order = order

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc, tb):
        return False

    def get(self, _model, order_id):
        if self.order and self.order.id == order_id:
            return self.order
        return None


def test_build_week_ranges_for_month_matches_real_sample_boundaries() -> None:
    march = [
        order_form_service._format_week_sheet_name(start_date, end_date)
        for start_date, end_date in order_form_service._build_week_ranges_for_month("2026-03")
    ]
    april = [
        order_form_service._format_week_sheet_name(start_date, end_date)
        for start_date, end_date in order_form_service._build_week_ranges_for_month("2026-04")
    ]
    may = [
        order_form_service._format_week_sheet_name(start_date, end_date)
        for start_date, end_date in order_form_service._build_week_ranges_for_month("2026-05")
    ]

    assert march == [
        "3月1日～3月7日",
        "3月8日～3月14日",
        "3月15日～3月21日",
        "3月22日～3月28日",
        "3月29日～3月31日",
    ]
    assert april == [
        "4月1日～4月4日",
        "4月5日～4月11日",
        "4月12日～4月18日",
        "4月19日～4月25日",
        "4月26日～4月30日",
    ]
    assert may == [
        "5月1日～5月2日",
        "5月3日～5月9日",
        "5月10日～5月16日",
        "5月17日～5月23日",
        "5月24日～5月30日",
        "5月31日～5月31日",
    ]


def test_build_fax_base_template_excel_adds_guides_and_keeps_logo(tmp_path):
    output = order_form_service.build_fax_base_template_excel(
        fax_template_id="fax_layout_regular_forbidden_v1",
        week_sheet_name="3月22日～3月28日",
        output_dir=tmp_path,
    )

    workbook = load_workbook(output)
    worksheet = workbook["3月22日～3月28日"]

    assert "A1" in {cell.coordinate for row in worksheet["A1:L1"] for cell in row if cell.fill.fill_type}
    assert worksheet["A3"].value == "施設名記入欄"
    assert "TEMPLATE=fax_layout_regular_forbidden_v1" in worksheet["B1"].value
    assert "$L$69" in str(worksheet.print_area)
    assert len(getattr(worksheet, "_images", [])) == 1
    assert workbook["設定"].sheet_state == "hidden"


def test_build_fax_order_form_excel_uses_facility_config_and_hidden_metadata(tmp_path, monkeypatch):
    monkeypatch.setattr(
        order_form_service.config_service,
        "get_facility_config",
        lambda facility_id: {
            "facility_id": facility_id,
            "facility_name": "テスト施設",
            "fax_template_id": "fax_layout_regular_diabetes_v1",
        },
    )

    output = order_form_service.build_fax_order_form_excel(
        facility_id="FACTEST01",
        week_sheet_name="3月22日～3月28日",
        output_dir=tmp_path,
    )

    workbook = load_workbook(output)
    worksheet = workbook["3月22日～3月28日"]
    metadata = workbook["設定"]
    rows = {key: value for key, value in metadata.iter_rows(min_row=2, values_only=True)}

    assert worksheet["A3"].value == "テスト施設"
    assert "FACILITY=FACTEST01:テスト施設" in worksheet["B1"].value
    assert rows["facility_id"] == "FACTEST01"
    assert rows["facility_name"] == "テスト施設"
    assert rows["fax_template_id"] == "fax_layout_regular_diabetes_v1"


def test_build_order_form_excel_creates_weekly_sheets_and_metadata(tmp_path, monkeypatch):
    source_name = "source.xlsx"
    _build_source_workbook(tmp_path / source_name, order_form_service._DEFAULT_WEEK_SHEET)

    monkeypatch.setattr(order_form_service, "_FAX_SOURCE_TEMPLATE_DIR", tmp_path)
    monkeypatch.setattr(order_form_service, "_OUTPUT_DIR", tmp_path)
    monkeypatch.setitem(
        order_form_service._FAX_FAMILY_SOURCE_MAP,
        "fax_layout_regular_forbidden_v1",
        {
            "source_workbook": source_name,
            "family_label": "共通・禁食2種",
        },
    )
    monkeypatch.setattr(
        order_form_service.config_service,
        "get_facility_config",
        lambda facility_id: {
            "facility_id": facility_id,
            "facility_name": "週次テスト施設",
            "fax_template_id": "fax_layout_regular_forbidden_v1",
        },
    )
    monkeypatch.setattr(
        order_form_service.menu_service,
        "get_menu_for_facility",
        lambda month_id, facility_id: {
            "entries": [
                {"menu_date": "2026-03-01", "daypart": "朝", "category": "副①", "name": "A"},
                {"menu_date": "2026-03-01", "daypart": "昼", "category": "主", "name": "B"},
                {"menu_date": "2026-03-08", "daypart": "朝", "category": "副①", "name": "C"},
                {"menu_date": "2026-03-29", "daypart": "夕", "category": "主", "name": "D"},
                {"menu_date": "2026-03-31", "daypart": "朝", "category": "副①", "name": "E"},
            ]
        },
    )

    output = order_form_service.build_order_form_excel(facility_id="FACTEST01", month_id="2026-03")

    workbook = load_workbook(output)
    assert workbook.sheetnames[:5] == [
        "3月1日～3月7日",
        "3月8日～3月14日",
        "3月15日～3月21日",
        "3月22日～3月28日",
        "3月29日～3月31日",
    ]

    first_sheet = workbook["3月1日～3月7日"]
    last_sheet = workbook["3月29日～3月31日"]
    rows = _metadata_rows(workbook)

    assert first_sheet["A4"].value == "週次テスト施設"
    assert "WEEK=3月1日～3月7日" in first_sheet["B1"].value
    assert first_sheet["A11"].value.date() == datetime(2026, 3, 1).date()
    assert first_sheet["B11"].value == "朝"
    assert first_sheet["C12"].value == "主"
    assert first_sheet["D12"].value == "B"
    assert first_sheet["G3"].value == "締切日2月13日まで"
    assert last_sheet["A11"].value.date() == datetime(2026, 3, 29).date()
    assert last_sheet["A12"].value.date() == datetime(2026, 3, 31).date()
    assert rows["month_id"] == "2026-03"
    assert rows["sheet_count"] == 5
    assert rows["entry_count"] == 5
    assert rows["week_1_sheet_name"] == "3月1日～3月7日"
    assert workbook["設定"].sheet_state == "hidden"


def test_build_saved_sheet_order_form_excel_writes_saved_quantities_to_week_form(tmp_path, monkeypatch):
    source_name = "saved-source.xlsx"
    week_sheet = "3月1日～3月7日"
    _build_source_workbook(tmp_path / source_name, week_sheet)
    source_workbook = load_workbook(tmp_path / source_name)
    source_ws = source_workbook[week_sheet]
    source_ws["E7"] = "常食"
    source_ws["F7"] = "肉禁"
    source_workbook.save(tmp_path / source_name)

    order = SimpleNamespace(
        id="ORD-SAVED-SHEET-001",
        facility_code="FACTEST01",
        week_code="2026-03@2026-03-01~2026-03-07",
        received_at=datetime(2026, 3, 1, 9, 0, 0),
    )
    monkeypatch.setattr(order_form_service, "_FAX_SOURCE_TEMPLATE_DIR", tmp_path)
    monkeypatch.setattr(order_form_service, "_OUTPUT_DIR", tmp_path)
    monkeypatch.setattr(order_form_service, "session_scope", lambda: _FakeSessionScope(order))
    monkeypatch.setitem(
        order_form_service._FAX_FAMILY_SOURCE_MAP,
        "fax_layout_regular_forbidden_v1",
        {
            "source_workbook": source_name,
            "family_label": "共通・禁食2種",
        },
    )
    monkeypatch.setattr(
        order_form_service.config_service,
        "get_facility_config",
        lambda facility_id: {
            "facility_id": facility_id,
            "facility_name": "保存済み施設",
            "fax_template_id": "fax_layout_regular_forbidden_v1",
            "order_form_source_workbook": source_name,
        },
    )
    monkeypatch.setattr(
        order_form_service.draft_sheet_service,
        "get_latest_sheet_draft",
        lambda order_id: {
            "id": "ODR-DRAFT-001",
            "draft_sheet_json": {
                "fields": ["date", "daypart", "category", "menu_name", "qty.regular_x", "qty.no_meat_x"],
                "rows": [
                    ["3/1", "朝", "主", "テスト献立A", "12", "3"],
                    ["3/1", "昼", "副", "テスト献立B", "4", ""],
                ],
            },
        },
    )

    output = order_form_service.build_saved_sheet_order_form_excel(order_id=order.id)

    workbook = load_workbook(output)
    worksheet = workbook[week_sheet]
    metadata = _metadata_rows(workbook)
    assert worksheet["A4"].value == "保存済み施設"
    assert worksheet["A11"].value.date() == datetime(2026, 3, 1).date()
    assert worksheet["B11"].value == "朝"
    assert worksheet["C11"].value == "主"
    assert worksheet["D11"].value == "テスト献立A"
    assert worksheet["E11"].value == 12
    assert worksheet["F11"].value == 3
    assert worksheet["B12"].value == "昼"
    assert worksheet["D12"].value == "テスト献立B"
    assert worksheet["E12"].value == 4
    assert metadata["mode"] == "saved_sheet"
    assert metadata["order_id"] == order.id
    assert metadata["saved_sheet_draft_id"] == "ODR-DRAFT-001"


def test_clear_week_sheet_body_preserves_quantity_body_merges_for_diabetes_template() -> None:
    source_path = order_form_service._resolve_source_workbook_path("いこいの森プラス　2604.xlsx")
    workbook = load_workbook(source_path)
    worksheet = workbook["4月26日～4月30日"]

    assert "E11:E12" in {str(item) for item in worksheet.merged_cells.ranges}
    assert "F11:F12" in {str(item) for item in worksheet.merged_cells.ranges}

    order_form_service._clear_week_sheet_body(worksheet)
    merged_ranges = {str(item) for item in worksheet.merged_cells.ranges}

    assert "E11:E12" in merged_ranges
    assert "F11:F12" in merged_ranges
    assert worksheet["E11"].value is None
    assert worksheet["F11"].value is None


def test_facility_template_merge_detection_uses_facility_settings_only() -> None:
    name_only_facility = {
        "facility_id": "FAC00007",
        "facility_name": "ゆうゆう（株）百々家",
        "fax_template_id": "fax_layout_regular_forbidden_v1",
    }

    assert not order_form_service.facility_template_has_vertical_merged_quantity_cells(
        name_only_facility,
        week_sheet_name="4月26日～4月30日",
    )


def test_facility_template_merge_detection_uses_body_merge_policy() -> None:
    facility = {
        "facility_id": "FAC00007",
        "facility_name": "ゆうゆう（株）百々家",
        "fax_template_id": "fax_layout_regular_forbidden_v1",
        "fax_template_override": {
            "body_merge_policy": {
                "mode": "daypart",
                "columns": ["qty.regular_x"],
            },
        },
    }

    assert order_form_service.facility_template_has_vertical_merged_quantity_cells(
        facility,
        week_sheet_name="4月26日～4月30日",
    )


def test_facility_template_scan_uses_registered_source_workbook_uri(tmp_path) -> None:
    source_path = tmp_path / "registered.xlsx"
    _build_source_workbook(source_path, "4月26日～4月30日")
    facility = {
        "facility_id": "FACURI",
        "facility_name": "登録テンプレート施設",
        "fax_template_id": "fax_layout_regular_forbidden_v1",
        "order_form_source_workbook_uri": str(source_path),
    }

    assert order_form_service.resolve_facility_source_workbook_name_for_week_sheet(
        facility,
        "4月26日～4月30日",
    ) == str(source_path)


def test_facility_template_scan_month_uri_replaces_default_source(tmp_path) -> None:
    default_path = tmp_path / "default.xlsx"
    month_path = tmp_path / "month.xlsx"
    _build_source_workbook(default_path, "3月22日～3月28日")
    _build_source_workbook(month_path, "4月26日～4月30日")
    facility = {
        "facility_id": "FACURI",
        "facility_name": "登録テンプレート施設",
        "fax_template_id": "fax_layout_regular_forbidden_v1",
        "order_form_source_workbook_uri": str(default_path),
        "order_form_month_source_uris": {
            "2026-04": str(month_path),
        },
    }

    assert order_form_service.resolve_facility_source_workbook_name_for_week_sheet(
        facility,
        "4月26日～4月30日",
    ) == str(month_path)


def test_build_fax_order_form_excel_uses_explicit_facility_source_workbook(tmp_path, monkeypatch):
    monkeypatch.setattr(
        order_form_service.config_service,
        "get_facility_config",
        lambda facility_id: {
            "facility_id": facility_id,
            "facility_name": "ゆうゆう（株）百々家",
            "fax_template_id": "fax_layout_regular_forbidden_v1",
            "order_form_month_sources": {
                "2026-03": "百々家 2603.xlsx",
                "2026-04": "百々家 2604.xlsx",
            },
        },
    )

    output = order_form_service.build_fax_order_form_excel(
        facility_id="FACMOMO",
        week_sheet_name="4月26日～4月30日",
        output_dir=tmp_path,
    )

    workbook = load_workbook(output)
    worksheet = workbook["4月26日～4月30日"]
    metadata = _metadata_rows(workbook)

    assert metadata["source_workbook"] == "百々家 2604.xlsx"
    assert "E11:E12" in {str(item) for item in worksheet.merged_cells.ranges}


def test_clear_week_sheet_body_uses_header_detected_quantity_columns_for_sibling_template() -> None:
    source_path = order_form_service._resolve_source_workbook_path("百々家 2604.xlsx")
    workbook = load_workbook(source_path)
    worksheet = workbook["4月26日～4月30日"]

    assert "E11:E12" in {str(item) for item in worksheet.merged_cells.ranges}

    order_form_service._clear_week_sheet_body(worksheet)
    merged_ranges = {str(item) for item in worksheet.merged_cells.ranges}

    assert "E11:E12" in merged_ranges
    assert worksheet["E11"].value is None


def test_build_order_form_excel_supports_six_week_months(tmp_path, monkeypatch):
    source_name = "source.xlsx"
    _build_source_workbook(tmp_path / source_name, order_form_service._DEFAULT_WEEK_SHEET)

    monkeypatch.setattr(order_form_service, "_FAX_SOURCE_TEMPLATE_DIR", tmp_path)
    monkeypatch.setattr(order_form_service, "_OUTPUT_DIR", tmp_path)
    monkeypatch.setitem(
        order_form_service._FAX_FAMILY_SOURCE_MAP,
        "fax_layout_regular_forbidden_v1",
        {
            "source_workbook": source_name,
            "family_label": "共通・禁食2種",
        },
    )
    monkeypatch.setattr(
        order_form_service.config_service,
        "get_facility_config",
        lambda facility_id: {
            "facility_id": facility_id,
            "facility_name": "6週施設",
            "fax_template_id": "fax_layout_regular_forbidden_v1",
        },
    )
    monkeypatch.setattr(
        order_form_service.menu_service,
        "get_menu_for_facility",
        lambda month_id, facility_id: {
            "entries": [
                {"menu_date": "2026-05-01", "daypart": "朝", "category": "副①", "name": "A"},
                {"menu_date": "2026-05-31", "daypart": "夕", "category": "主", "name": "B"},
            ]
        },
    )

    output = order_form_service.build_order_form_excel(facility_id="FAC6WEEK", month_id="2026-05")

    workbook = load_workbook(output)
    assert workbook.sheetnames[:6] == [
        "5月1日～5月2日",
        "5月3日～5月9日",
        "5月10日～5月16日",
        "5月17日～5月23日",
        "5月24日～5月30日",
        "5月31日～5月31日",
    ]
    assert _metadata_rows(workbook)["sheet_count"] == 6


@pytest.mark.parametrize(
    ("fax_template_id", "facility_name"),
    [
        ("fax_layout_regular_forbidden_v1", "共通確認"),
        ("fax_layout_floor_2f3f_v1", "GH確認"),
        ("fax_layout_regular_soft_mixer_forbidden_v1", "藍確認"),
        ("fax_layout_regular_staff_daycare_v1", "湘南確認"),
        ("fax_layout_regular_diabetes_v1", "糖尿確認"),
        ("fax_layout_regular_staff_daycare_other_forbidden_v1", "ふれあい確認"),
        ("fax_layout_soft_packaging_forbidden_v1", "池袋確認"),
    ],
)
def test_build_order_form_excel_supports_all_fax_families(tmp_path, monkeypatch, fax_template_id, facility_name):
    monkeypatch.setattr(order_form_service, "_OUTPUT_DIR", tmp_path)
    monkeypatch.setattr(
        order_form_service.config_service,
        "get_facility_config",
        lambda facility_id: {
            "facility_id": facility_id,
            "facility_name": facility_name,
            "fax_template_id": fax_template_id,
        },
    )
    monkeypatch.setattr(
        order_form_service.menu_service,
        "get_menu_for_facility",
        lambda month_id, facility_id: {
            "entries": [
                {"menu_date": "2026-03-01", "daypart": "朝", "category": "副①", "name": "A"},
                {"menu_date": "2026-03-02", "daypart": "昼", "category": "主", "name": "B"},
                {"menu_date": "2026-03-29", "daypart": "夕", "category": "副②", "name": "C"},
            ]
        },
    )

    output = order_form_service.build_order_form_excel(facility_id="FACTESTALL", month_id="2026-03")
    workbook = load_workbook(output)

    assert workbook.sheetnames[:5] == [
        "3月1日～3月7日",
        "3月8日～3月14日",
        "3月15日～3月21日",
        "3月22日～3月28日",
        "3月29日～3月31日",
    ]
    assert workbook[workbook.sheetnames[0]]["A4"].value == facility_name
    assert fax_template_id in workbook[workbook.sheetnames[0]]["B1"].value


def test_infer_fax_template_id_from_facility_falls_back_from_invoice_columns() -> None:
    assert (
        order_form_service._infer_fax_template_id_from_facility(
            {
                "facility_name": "大和なでしこ",
                "invoice_template": {
                    "columns": [
                        {"header": "常食", "diet_type": "regular"},
                        {"header": "禁食", "diet_type": "禁食"},
                    ]
                },
            }
        )
        == "fax_layout_regular_forbidden_v1"
    )
    assert (
        order_form_service._infer_fax_template_id_from_facility(
            {
                "facility_name": "佐古",
                "invoice_template": {
                    "columns": [
                        {"header": "常食2F", "diet_type": "regular", "area_id": "2F"},
                        {"header": "常食3F", "diet_type": "regular", "area_id": "3F"},
                        {"header": "軟菜", "diet_type": "soft"},
                    ]
                },
            }
        )
        == "fax_layout_floor_2f3f_v1"
    )


def test_build_order_form_excel_uses_inferred_template_id_in_output_name(tmp_path, monkeypatch):
    source_name = "source.xlsx"
    _build_source_workbook(tmp_path / source_name, order_form_service._DEFAULT_WEEK_SHEET)

    monkeypatch.setattr(order_form_service, "_FAX_SOURCE_TEMPLATE_DIR", tmp_path)
    monkeypatch.setattr(order_form_service, "_OUTPUT_DIR", tmp_path)
    monkeypatch.setitem(
        order_form_service._FAX_FAMILY_SOURCE_MAP,
        "fax_layout_floor_2f3f_v1",
        {
            "source_workbook": source_name,
            "family_label": "2F/3F分割",
        },
    )
    monkeypatch.setattr(
        order_form_service.config_service,
        "get_facility_config",
        lambda facility_id: {
            "facility_id": facility_id,
            "facility_name": "推定施設",
            "invoice_template": {
                "columns": [
                    {"header": "常食2F", "diet_type": "regular", "area_id": "2F"},
                    {"header": "常食3F", "diet_type": "regular", "area_id": "3F"},
                    {"header": "軟菜", "diet_type": "soft"},
                ]
            },
        },
    )
    monkeypatch.setattr(
        order_form_service.menu_service,
        "get_menu_for_facility",
        lambda month_id, facility_id: {
            "entries": [
                {"menu_date": "2026-03-01", "daypart": "朝", "category": "副①", "name": "A"},
            ]
        },
    )

    output = order_form_service.build_order_form_excel(facility_id="FACINF01", month_id="2026-03")

    assert "fax_layout_floor_2f3f_v1" in output.name


def test_packaged_order_form_sources_exist_for_manifest() -> None:
    specs = order_form_service.list_fax_order_form_template_specs()
    assert specs
    for spec in specs:
        month_sources = spec.get("month_sources") or {}
        assert month_sources
        for filename in month_sources.values():
            assert order_form_service._resolve_source_workbook_path(filename).exists()


def test_build_order_form_excel_uses_month_specific_packaged_source_workbook(tmp_path, monkeypatch):
    monkeypatch.setattr(order_form_service, "_OUTPUT_DIR", tmp_path)
    monkeypatch.setattr(
        order_form_service.config_service,
        "get_facility_config",
        lambda facility_id: {
            "facility_id": facility_id,
            "facility_name": "4月施設",
            "fax_template_id": "fax_layout_regular_forbidden_v1",
        },
    )
    monkeypatch.setattr(
        order_form_service.menu_service,
        "get_menu_for_facility",
        lambda month_id, facility_id: {
            "entries": [
                {"menu_date": "2026-04-01", "daypart": "朝", "category": "副①", "name": "A"},
                {"menu_date": "2026-04-26", "daypart": "昼", "category": "主", "name": "B"},
            ]
        },
    )

    output = order_form_service.build_order_form_excel(facility_id="FACAPR01", month_id="2026-04")

    workbook = load_workbook(output)
    rows = _metadata_rows(workbook)
    assert rows["source_workbook"] == "共通　2604.xlsx"
    assert workbook.sheetnames[:5] == [
        "4月1日～4月4日",
        "4月5日～4月11日",
        "4月12日～4月18日",
        "4月19日～4月25日",
        "4月26日～4月30日",
    ]


def test_build_fax_order_form_excel_uses_month_specific_source_for_week_sheet(tmp_path, monkeypatch):
    monkeypatch.setattr(
        order_form_service.config_service,
        "get_facility_config",
        lambda facility_id: {
            "facility_id": facility_id,
            "facility_name": "4月FAX施設",
            "fax_template_id": "fax_layout_regular_forbidden_v1",
        },
    )

    output = order_form_service.build_fax_order_form_excel(
        facility_id="FACAPR02",
        week_sheet_name="4月1日～4月4日",
        output_dir=tmp_path,
    )

    workbook = load_workbook(output)
    metadata = _metadata_rows(workbook)
    assert metadata["source_workbook"] == "共通　2604.xlsx"
    assert workbook.sheetnames == ["4月1日～4月4日", "設定"]


def test_build_fax_structure_only_excel_clears_body_values(tmp_path, monkeypatch):
    monkeypatch.setattr(
        order_form_service.config_service,
        "get_facility_config",
        lambda facility_id: {
            "facility_id": facility_id,
            "facility_name": "構造比較施設",
            "fax_template_id": "fax_layout_regular_forbidden_v1",
            "fax_template_override": {
                "columns": [
                    {"role": "date", "header": "日付", "name": "date_mmdd", "source_index": 0},
                    {"role": "daypart", "header": "区分", "name": "daypart", "source_index": 1},
                    {"role": "menu", "header": "献立", "name": "menu", "source_index": 2},
                    {"role": "quantity", "header": "常食", "name": "qty.regular_x", "source_index": 3},
                    {"role": "quantity", "header": "肉禁", "name": "qty.no_meat_x", "source_index": 4},
                ],
            },
        },
    )

    output = order_form_service.build_fax_structure_only_excel(
        facility_id="FACSTRUCT01",
        week_sheet_name="4月5日～4月11日",
        output_dir=tmp_path,
    )

    generated = load_workbook(output)
    generated_sheet = generated["facility_template"]
    schema = generated["generated_template_schema"]
    schema_rows = {str(row[0].value): row[1].value for row in schema.iter_rows(min_row=1, max_col=2) if row[0].value}

    assert generated_sheet["A4"].value == "構造比較施設"
    assert generated_sheet["B11"].value == "朝"
    assert generated_sheet["E11"].value is None
    assert schema_rows["source"] == "master_layout_template"
    assert schema_rows["facility_id"] == "FACSTRUCT01"


def test_build_order_form_excel_blocks_when_month_specific_source_missing(tmp_path, monkeypatch):
    monkeypatch.setattr(order_form_service, "_OUTPUT_DIR", tmp_path)
    monkeypatch.setitem(
        order_form_service._FAX_FAMILY_SOURCE_MAP,
        "fax_layout_missing_month_v1",
        {
            "family_label": "欠損確認",
            "month_sources": {"2026-03": "source.xlsx"},
            "source_workbook": "",
        },
    )
    monkeypatch.setattr(
        order_form_service.config_service,
        "get_facility_config",
        lambda facility_id: {
            "facility_id": facility_id,
            "facility_name": "欠損施設",
            "fax_template_id": "fax_layout_missing_month_v1",
        },
    )
    monkeypatch.setattr(
        order_form_service.menu_service,
        "get_menu_for_facility",
        lambda month_id, facility_id: {
            "entries": [
                {"menu_date": "2026-05-01", "daypart": "朝", "category": "副①", "name": "A"},
            ]
        },
    )

    with pytest.raises(
        ValueError,
        match="source workbook not configured for fax_template_id=fax_layout_missing_month_v1 month_id=2026-05",
    ):
        order_form_service.build_order_form_excel(facility_id="FACMISS01", month_id="2026-05")
