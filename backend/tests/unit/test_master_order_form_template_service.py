from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor
from openpyxl.utils import get_column_letter, range_boundaries

from src.services import config_service
from src.services import master_order_form_template_service as service


def _facility_config(columns: list[dict]) -> dict:
    return {
        "facility_id": "FACTEST",
        "facility_name": "テスト施設",
        "fax_template": {
            "columns": columns,
        },
    }


def _generated_render_width(worksheet, count: int) -> int:
    return sum(
        service._column_width_to_render_pixels(
            worksheet.column_dimensions[get_column_letter(col)].width
        )
        for col in range(service.GENERATED_START_COL, service.GENERATED_START_COL + count)
    )


def _assert_no_top_static_merge_extends_past_generated_end(worksheet, end_col: int) -> None:
    for merged_range in worksheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
        if min_col == 7 and 2 <= min_row <= 6 and max_row <= 6:
            assert max_col <= end_col


def test_build_facility_template_from_master_writes_quantity_columns(tmp_path: Path) -> None:
    output = service.build_facility_template_xlsx(
        facility_config=_facility_config(
            [
                {"index": 0, "role": "date", "header": "日付", "name": "date_mmdd"},
                {"index": 1, "role": "daypart", "header": "区分", "name": "daypart"},
                {"index": 2, "role": "menu_name", "header": "献立", "name": "menu"},
                {"index": 3, "role": "quantity", "header": "常食", "name": "qty.regular_x"},
                {"index": 4, "role": "quantity", "header": "肉禁", "name": "qty.no_meat_x"},
                {"index": 5, "role": "quantity", "header": "魚禁", "name": "qty.no_fish_x"},
                {"index": 6, "role": "quantity", "header": "変更1", "name": "qty.change_1_x"},
                {"index": 7, "role": "quantity", "header": "変更2", "name": "qty.change_2_x"},
                {"index": 8, "role": "note", "header": "備考欄", "name": "remarks"},
            ]
        ),
        output_path=tmp_path / "facility_template.xlsx",
    )

    workbook = load_workbook(output)
    worksheet = workbook["facility_template"]
    schema = workbook["generated_template_schema"]

    assert worksheet["A4"].value == "テスト施設"
    assert "A4:E5" in {str(item) for item in worksheet.merged_cells.ranges}
    assert worksheet["A4"].alignment.horizontal == "center"
    assert worksheet["A4"].alignment.shrink_to_fit is True
    assert worksheet["A4"].font.bold is True
    assert worksheet["E7"].value == "常食"
    assert worksheet["F7"].value == "肉禁"
    assert worksheet["G7"].value == "魚禁"
    assert worksheet["H7"].value == "変更1"
    assert worksheet["I7"].value == "変更2"
    assert worksheet["J7"].value == "備考欄"
    assert _generated_render_width(worksheet, 6) == service.SOURCE_GENERATED_PIXEL_WIDTH
    assert str(worksheet.print_area) == "'facility_template'!$A$1:$J$69"
    assert schema.sheet_state == "hidden"
    assert schema["B8"].value == 6
    assert schema["B9"].value == service.SOURCE_GENERATED_PIXEL_WIDTH


def test_build_facility_template_from_master_merges_consecutive_header_groups(tmp_path: Path) -> None:
    output = service.build_facility_template_xlsx(
        facility_config=_facility_config(
            [
                {"index": 0, "role": "date", "header": "日付", "name": "date_mmdd"},
                {"index": 1, "role": "daypart", "header": "区分", "name": "daypart"},
                {"index": 2, "role": "menu_name", "header": "献立", "name": "menu"},
                {"index": 3, "role": "quantity", "header": "常食", "name": "qty.regular_x"},
                {
                    "index": 4,
                    "role": "quantity",
                    "header": "肉禁",
                    "header_group": "禁食",
                    "name": "qty.no_meat_x",
                },
                {
                    "index": 5,
                    "role": "quantity",
                    "header": "魚禁",
                    "header_group": "禁食",
                    "name": "qty.no_fish_x",
                },
                {"index": 6, "role": "note", "header": "備考欄", "name": "remarks"},
            ]
        ),
        output_path=tmp_path / "facility_template_grouped.xlsx",
    )

    workbook = load_workbook(output)
    worksheet = workbook["facility_template"]
    merged_ranges = {str(item) for item in worksheet.merged_cells.ranges}

    assert "F7:G8" in merged_ranges
    assert "F9:F10" in merged_ranges
    assert "G9:G10" in merged_ranges
    assert worksheet["F7"].value == "禁食"
    assert worksheet["F9"].value == "肉禁"
    assert worksheet["G9"].value == "魚禁"
    assert worksheet["E7"].value == "常食"
    assert "H7:H10" in merged_ranges
    assert worksheet["H7"].value == "備考欄"
    assert _generated_render_width(worksheet, 4) == service.SOURCE_GENERATED_PIXEL_WIDTH


def test_build_facility_template_from_master_keeps_post_menu_aux_columns(tmp_path: Path) -> None:
    output = service.build_facility_template_xlsx(
        facility_config=_facility_config(
            [
                {"index": 0, "role": "date", "header": "日付", "name": "date_mmdd"},
                {"index": 1, "role": "daypart", "header": "区分", "name": "daypart"},
                {"index": 2, "role": "aux", "header": "副区分", "name": "aux.col_2"},
                {"index": 3, "role": "menu_name", "header": "献立", "name": "menu"},
                {"index": 4, "role": "aux", "header": "合計", "name": "aux.col_4"},
                {"index": 5, "role": "quantity", "header": "常食", "name": "qty.regular_x"},
                {"index": 6, "role": "note", "header": "備考欄", "name": "remarks"},
            ]
        ),
        output_path=tmp_path / "facility_template_with_total.xlsx",
    )

    workbook = load_workbook(output)
    worksheet = workbook["facility_template"]

    assert worksheet["E7"].value == "合計"
    assert worksheet["F7"].value == "常食"
    assert worksheet["G7"].value == "備考欄"
    assert str(worksheet.print_area) == "'facility_template'!$A$1:$G$69"


def test_build_facility_template_from_master_supports_all_master_facilities(tmp_path: Path) -> None:
    master = config_service.load_facility_master()
    facility_ids = [facility["facility_id"] for facility in master.get("facilities", []) if facility.get("facility_id")]

    assert facility_ids
    for facility in master.get("facilities", []):
        facility_id = facility.get("facility_id")
        if not facility_id:
            continue
        resolved = config_service._build_facility_config(facility_id=facility_id, facility=facility)  # noqa: SLF001
        output = service.build_facility_template_xlsx(
            facility_config=resolved,
            output_path=tmp_path / f"{facility_id}_facility_template.xlsx",
        )
        workbook = load_workbook(output)
        worksheet = workbook["facility_template"]
        schema = workbook["generated_template_schema"]

        assert worksheet["E7"].value
        assert schema["B8"].value >= 1
        assert _generated_render_width(worksheet, schema["B8"].value) == service.SOURCE_GENERATED_PIXEL_WIDTH
        end_col = service.GENERATED_START_COL + int(schema["B8"].value) - 1
        _assert_no_top_static_merge_extends_past_generated_end(worksheet, end_col)


def test_build_facility_template_from_master_uses_explicit_header_groups_only(tmp_path: Path) -> None:
    output = service.build_facility_template_xlsx(
        facility_config=_facility_config(
            [
                {"index": 0, "role": "date", "header": "日付", "name": "date_mmdd"},
                {"index": 1, "role": "daypart", "header": "区分", "name": "daypart"},
                {"index": 2, "role": "menu_name", "header": "献立", "name": "menu"},
                {"index": 3, "role": "quantity", "header": "常食2F", "name": "qty.regular_2f"},
                {"index": 4, "role": "quantity", "header": "常食3F", "name": "qty.regular_3f"},
                {"index": 5, "role": "note", "header": "備考", "name": "remarks"},
            ]
        ),
        output_path=tmp_path / "facility_template_ungrouped.xlsx",
    )

    workbook = load_workbook(output)
    worksheet = workbook["facility_template"]
    merged_ranges = {str(item) for item in worksheet.merged_cells.ranges}

    assert "E7:F7" not in merged_ranges
    assert worksheet["E7"].value == "常食2F"
    assert worksheet["F7"].value == "常食3F"
    assert _generated_render_width(worksheet, 3) == service.SOURCE_GENERATED_PIXEL_WIDTH


def test_build_facility_template_from_master_applies_configured_two_level_headers(tmp_path: Path) -> None:
    master = config_service.load_facility_master()
    facility = next(item for item in master.get("facilities", []) if item.get("facility_id") == "FAC00003")
    resolved = config_service._build_facility_config(facility_id="FAC00003", facility=facility)  # noqa: SLF001

    output = service.build_facility_template_xlsx(
        facility_config=resolved,
        output_path=tmp_path / "FAC00003_facility_template.xlsx",
    )

    workbook = load_workbook(output)
    worksheet = workbook["facility_template"]
    merged_ranges = {str(item) for item in worksheet.merged_cells.ranges}

    assert "E7:F8" in merged_ranges
    assert "G7:H8" in merged_ranges
    assert "I7:J8" in merged_ranges
    assert "K7:P7" in merged_ranges
    assert "K8:L8" in merged_ranges
    assert "M8:N8" in merged_ranges
    assert "O8:P8" in merged_ranges
    assert worksheet["E7"].value == "常食"
    assert worksheet["E9"].value == "花"
    assert worksheet["F9"].value == "月"
    assert worksheet["G7"].value == "軟菜"
    assert worksheet["G9"].value == "花"
    assert worksheet["H9"].value == "月"
    assert worksheet["I7"].value == "ミキサー"
    assert worksheet["I9"].value == "花"
    assert worksheet["J9"].value == "月"
    assert worksheet["K7"].value == "禁食、魚"
    assert worksheet["K8"].value == "常食"
    assert worksheet["K9"].value == "花"
    assert worksheet["L9"].value == "月"
    assert worksheet["M8"].value == "軟菜"
    assert worksheet["O8"].value == "ミキサー"
    assert _generated_render_width(worksheet, workbook["generated_template_schema"]["B8"].value) == service.SOURCE_GENERATED_PIXEL_WIDTH


def test_build_facility_template_from_master_freezes_logo_size_across_column_counts(tmp_path: Path) -> None:
    outputs = []
    for facility_id in ("FAC00001", "FAC00011"):
        outputs.append(
            service.build_facility_template_xlsx_for_facility(
                facility_id=facility_id,
                output_path=tmp_path / f"{facility_id}_facility_template.xlsx",
            )
        )

    image_payloads = []
    for output in outputs:
        workbook = load_workbook(output)
        worksheet = workbook["facility_template"]
        images = list(getattr(worksheet, "_images", []))
        assert len(images) == 1
        image = images[0]
        assert isinstance(image.anchor, OneCellAnchor)
        image_payloads.append(
            (
                image.anchor.ext.cx,
                image.anchor.ext.cy,
                image.anchor._from.col,
                image.anchor._from.row,
                image.anchor._from.colOff,
                image.anchor._from.rowOff,
            )
        )

    assert image_payloads[0] == image_payloads[1]


def test_build_facility_template_from_master_can_embed_week_menu_identity(tmp_path: Path, monkeypatch) -> None:
    def fake_get_menu_for_facility(month_id: str, facility_id: str) -> dict:
        assert month_id == "2026-04"
        assert facility_id == "FACTEST"
        return {
            "entries": [
                {"menu_date": "2026-04-26", "daypart": "朝", "category": "朝①", "name": "大豆のトマト煮"},
                {"menu_date": "2026-04-26", "daypart": "朝", "category": "朝②", "name": "胡瓜のサラダ"},
                {"menu_date": "2026-04-26", "daypart": "昼", "category": "主A", "name": "サワラの揚げ浸し"},
                {"menu_date": "2026-04-27", "daypart": "朝", "category": "朝①", "name": "じゃが芋の煮物"},
                {"menu_date": "2026-05-01", "daypart": "朝", "category": "朝①", "name": "範囲外"},
            ]
        }

    monkeypatch.setattr(service.menu_service, "get_menu_for_facility", fake_get_menu_for_facility)

    output = service.build_facility_template_xlsx(
        facility_config=_facility_config(
            [
                {"index": 0, "role": "date", "header": "日付", "name": "date_mmdd"},
                {"index": 1, "role": "daypart", "header": "区分", "name": "daypart"},
                {"index": 2, "role": "menu_name", "header": "献立", "name": "menu"},
                {"index": 3, "role": "quantity", "header": "常食", "name": "qty.regular_x"},
                {"index": 4, "role": "quantity", "header": "肉禁", "name": "qty.no_meat_x"},
                {"index": 5, "role": "note", "header": "備考欄", "name": "remarks"},
            ]
        ),
        output_path=tmp_path / "facility_template_with_week_menu.xlsx",
        week_value="2026-04@2026-04-26~2026-04-30",
    )

    workbook = load_workbook(output)
    worksheet = workbook["facility_template"]
    schema = workbook["generated_template_schema"]

    assert worksheet["D11"].value == "大豆のトマト煮"
    assert worksheet["D12"].value == "胡瓜のサラダ"
    assert worksheet["D13"].value == "サワラの揚げ浸し"
    assert worksheet["D14"].value == "じゃが芋の煮物"
    assert worksheet["D15"].value is None
    schema_values = {row[0]: row[1] for row in schema.iter_rows(values_only=True) if row and row[0]}
    assert schema_values["week_value"] == "2026-04@2026-04-26~2026-04-30"
    assert schema_values["week_menu_rows"] == 4


def test_build_facility_template_from_master_keeps_week_menu_overflow_as_review_metadata(
    tmp_path: Path,
    monkeypatch,
) -> None:
    dayparts = [
        ("朝", ["朝①", "朝②"]),
        ("昼", ["主菜", "副菜", "副菜"]),
        ("夕", ["主菜", "副菜", "副菜"]),
    ]
    entries: list[dict[str, str]] = []
    for day in range(5, 12):
        menu_date = f"2026-07-{day:02d}"
        for daypart, categories in dayparts:
            for index, category in enumerate(categories, start=1):
                entries.append(
                    {
                        "menu_date": menu_date,
                        "daypart": daypart,
                        "category": category,
                        "name": f"{menu_date}-{daypart}-{index}",
                    }
                )
        if day == 7:
            entries.insert(
                len(entries) - 3,
                {
                    "menu_date": "2026-07-07",
                    "daypart": "昼",
                    "category": "",
                    "name": "小松菜のおかか和え",
                },
            )

    def fake_get_menu_for_facility(month_id: str, facility_id: str) -> dict:
        assert month_id == "2026-07"
        assert facility_id == "FACTEST"
        return {"entries": entries}

    monkeypatch.setattr(service.menu_service, "get_menu_for_facility", fake_get_menu_for_facility)

    output = service.build_facility_template_xlsx(
        facility_config=_facility_config(
            [
                {"index": 0, "role": "date", "header": "日付", "name": "date_mmdd"},
                {"index": 1, "role": "daypart", "header": "区分", "name": "daypart"},
                {"index": 2, "role": "menu_name", "header": "献立", "name": "menu"},
                {"index": 3, "role": "quantity", "header": "常食", "name": "qty.regular_x"},
                {"index": 4, "role": "quantity", "header": "肉禁", "name": "qty.no_meat_x"},
                {"index": 5, "role": "note", "header": "備考欄", "name": "remarks"},
            ]
        ),
        output_path=tmp_path / "facility_template_with_week_menu_overflow.xlsx",
        week_value="2026-07@2026-07-05~2026-07-11",
    )

    workbook = load_workbook(output)
    worksheet = workbook["facility_template"]
    schema_values = {
        row[0]: row[1]
        for row in workbook["generated_template_schema"].iter_rows(values_only=True)
        if row and row[0]
    }

    assert worksheet.cell(row=service.BODY_END_ROW, column=4).value == "2026-07-11-夕-3"
    assert schema_values["week_menu_rows"] == 57
    assert schema_values["week_menu_source_rows"] == 57
    assert schema_values["week_menu_overflow_rows"] == 0
    overflow_entries = json.loads(schema_values["week_menu_overflow_entries"])
    assert overflow_entries == []


def test_build_facility_template_from_master_applies_configured_body_merges(tmp_path: Path) -> None:
    master = config_service.load_facility_master()
    facility = next(item for item in master.get("facilities", []) if item.get("facility_id") == "FAC00007")
    resolved = config_service._build_facility_config(facility_id="FAC00007", facility=facility)  # noqa: SLF001

    output = service.build_facility_template_xlsx(
        facility_config=resolved,
        output_path=tmp_path / "FAC00007_facility_template.xlsx",
        week_value="2026-04@2026-04-26~2026-04-30",
    )

    workbook = load_workbook(output)
    worksheet = workbook["facility_template"]
    merged_ranges = {str(item) for item in worksheet.merged_cells.ranges}
    schema_values = {row[0]: row[1] for row in workbook["generated_template_schema"].iter_rows(values_only=True) if row and row[0]}

    assert "E43:E44" in merged_ranges
    assert "E45:E47" in merged_ranges
    assert schema_values["configured_body_merged_ranges"] == len(
        [
            item
            for item in merged_ranges
            if range_boundaries(item)[0] >= service.GENERATED_START_COL
            and range_boundaries(item)[1] >= service.BODY_START_ROW
            and range_boundaries(item)[3] <= service.BODY_END_ROW
        ]
    )
    assert schema_values["configured_body_merged_range_details"]


def test_build_facility_template_from_master_applies_body_merges_without_week_value(tmp_path: Path) -> None:
    master = config_service.load_facility_master()
    facility = next(item for item in master.get("facilities", []) if item.get("facility_id") == "FAC00016")
    resolved = config_service._build_facility_config(facility_id="FAC00016", facility=facility)  # noqa: SLF001

    output = service.build_facility_template_xlsx(
        facility_config=resolved,
        output_path=tmp_path / "FAC00016_facility_template.xlsx",
    )

    workbook = load_workbook(output)
    worksheet = workbook["facility_template"]
    merged_ranges = {str(item) for item in worksheet.merged_cells.ranges}
    schema_values = {row[0]: row[1] for row in workbook["generated_template_schema"].iter_rows(values_only=True) if row and row[0]}

    assert "E11:E12" in merged_ranges
    assert "F11:F12" in merged_ranges
    assert "E13:E15" in merged_ranges
    assert "F13:F15" in merged_ranges
    assert "G11:G12" not in merged_ranges
    assert schema_values["configured_body_merged_ranges"] > 0


def test_build_facility_template_from_master_rejects_large_cell_without_targets(tmp_path: Path) -> None:
    facility_config = _facility_config(
        [
            {"index": 0, "role": "date", "header": "日付", "name": "date_mmdd"},
            {"index": 1, "role": "daypart", "header": "区分", "name": "daypart"},
            {"index": 2, "role": "menu_name", "header": "献立", "name": "menu"},
            {"index": 3, "role": "quantity", "header": "常食", "name": "qty.regular_x"},
            {"index": 4, "role": "quantity", "header": "肉禁", "name": "qty.no_meat_x"},
            {"index": 5, "role": "note", "header": "備考欄", "name": "remarks"},
        ]
    )
    facility_config["fax_template"]["body_merge_policy"] = {"mode": "daypart", "required": True}

    try:
        service.build_facility_template_xlsx(
            facility_config=facility_config,
            output_path=tmp_path / "invalid_body_merge.xlsx",
        )
    except service.FacilityTemplateBuildError as exc:
        assert str(exc) == "facility_template_body_merge_columns_missing"
    else:
        raise AssertionError("expected body merge target validation failure")


def test_build_facility_template_diagnostics_is_stable_for_same_inputs(monkeypatch) -> None:
    def fake_get_menu_for_facility(month_id: str, facility_id: str) -> dict:
        assert month_id == "2026-04"
        assert facility_id == "FACTEST"
        return {
            "entries": [
                {"menu_date": "2026-04-26", "daypart": "朝", "category": "朝①", "name": "大豆のトマト煮"},
                {"menu_date": "2026-04-26", "daypart": "昼", "category": "主A", "name": "サワラの揚げ浸し"},
            ]
        }

    monkeypatch.setattr(service.menu_service, "get_menu_for_facility", fake_get_menu_for_facility)
    facility_config = _facility_config(
        [
            {"index": 0, "role": "date", "header": "日付", "name": "date_mmdd"},
            {"index": 1, "role": "daypart", "header": "区分", "name": "daypart"},
            {"index": 2, "role": "menu_name", "header": "献立", "name": "menu"},
            {"index": 3, "role": "quantity", "header": "常食", "name": "qty.regular_x"},
            {"index": 4, "role": "quantity", "header": "肉禁", "name": "qty.no_meat_x"},
            {"index": 5, "role": "note", "header": "備考欄", "name": "remarks"},
        ]
    )

    first = service.build_facility_template_diagnostics(
        facility_config=facility_config,
        week_value="2026-04@2026-04-26~2026-04-30",
    )
    second = service.build_facility_template_diagnostics(
        facility_config=facility_config,
        week_value="2026-04@2026-04-26~2026-04-30",
    )

    assert first["master_template_sha256"] == second["master_template_sha256"]
    assert first["facility_template_canonical_digest"] == second["facility_template_canonical_digest"]
    assert first["schema_digest"] == second["schema_digest"]
    assert first["generated_end_letter"] == "G"
