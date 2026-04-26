from src.services import hakodate_assignment_service
from src.services.hakodate_assignment_service import build_hakodate_assignment
from openpyxl import Workbook
from pathlib import Path
from PIL import Image
import pytest


_REAL_PDF_FIELDS = [
    "date",
    "daypart",
    "aux",
    "menu_name",
    "qty.regular_x",
    "qty.staff_x",
    "qty.no_meat_x",
    "qty.no_fish_x",
    "qty.change_1_x",
    "qty.change_2_x",
    "note",
    "spacer",
]
_REAL_PDF_WEEK = "4月26日～4月30日"


def _template() -> dict:
    template = {
        "quantity_assignment_strategy": "hakodate",
        "hakodate_header_rows": 1,
        "hakodate_data_row_count": 2,
        "main_ocr_row_fields": [
            "date",
            "daypart",
            "menu_name",
            "qty.regular_x",
            "qty.no_meat_x",
        ],
    }
    signature = hakodate_assignment_service.build_facility_template_signature(template)
    template["hakodate_template_signature"] = signature["hash"]
    template["hakodate_template_signature_components"] = signature["components"]
    return template


def _grid() -> dict:
    return {
        "table_box": [0.0, 0.0, 1.0, 1.0],
        "column_edges": [0.0, 0.2, 0.4, 0.6, 0.8, 1.0],
        "row_edges": [0.0, 0.2, 0.4, 0.6],
        "confidence": 0.9,
    }


def _legacy_grid_pdf_template() -> dict:
    template = _template()
    template["hakodate_alignment_mode"] = "detected_grid"
    return template


def _structure_slot_template(skeleton_rows: list[dict] | None = None) -> dict:
    template = {
        "facility_id": "FAC_TEST",
        "quantity_assignment_strategy": "hakodate",
        "hakodate_alignment_mode": "structure_slot",
        "hakodate_header_rows": 2,
        "main_ocr_row_fields": [
            "date",
            "daypart",
            "aux",
            "menu_name",
            "qty.regular_x",
        ],
    }
    signature = hakodate_assignment_service.build_facility_template_signature(
        template,
        facility_id="FAC_TEST",
        skeleton_rows=skeleton_rows or [{"row_id": "row-a", "menu_name": "献立A"}],
    )
    template["hakodate_template_signature"] = signature["hash"]
    template["hakodate_template_signature_components"] = signature["components"]
    return template


def _structure_slot_worksheet():
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.cell(row=7, column=1, value="日付")
    worksheet.cell(row=7, column=2, value="区分")
    worksheet.cell(row=7, column=3, value="補助")
    worksheet.cell(row=7, column=4, value="献立")
    worksheet.cell(row=7, column=5, value="常食")
    worksheet.cell(row=11, column=1, value="4/26")
    worksheet.cell(row=11, column=2, value="朝")
    worksheet.cell(row=11, column=4, value="献立A")
    return worksheet


def _merged_structure_slot_worksheet():
    worksheet = _structure_slot_worksheet()
    worksheet.cell(row=12, column=1, value="4/26")
    worksheet.cell(row=12, column=2, value="朝")
    worksheet.cell(row=12, column=4, value="献立B")
    worksheet.merge_cells("E11:E12")
    return worksheet


def _real_pdf_skeleton_rows(facility_id: str) -> list[dict]:
    worksheet = hakodate_assignment_service._source_worksheet_for_structure_template(  # noqa: SLF001
        facility_id=facility_id,
        week_sheet_name=_REAL_PDF_WEEK,
    )
    rows = hakodate_assignment_service._workbook_physical_row_map(  # noqa: SLF001
        worksheet,
        row_count=59,
    )
    return [
        {
            "row_id": f"{facility_id}-r{row_index}",
            "date": row.get("effective_date") or row.get("date") or "",
            "daypart": row.get("daypart") or "",
            "menu_name": row.get("menu_name") or "",
        }
        for row_index, row in sorted(rows.items())
    ]


def _real_pdf_template(facility_id: str, debug_dir: Path) -> tuple[dict, list[dict]]:
    skeleton_rows = _real_pdf_skeleton_rows(facility_id)
    template = {
        "facility_id": facility_id,
        "quantity_assignment_strategy": "hakodate",
        "hakodate_alignment_mode": "structure_slot",
        "hakodate_week_sheet_name": _REAL_PDF_WEEK,
        "hakodate_header_rows": 2,
        "hakodate_data_row_count": len(skeleton_rows),
        "hakodate_structure_dpi": 200,
        "hakodate_debug_output_dir": str(debug_dir),
        "main_ocr_row_fields": list(_REAL_PDF_FIELDS),
    }
    signature = hakodate_assignment_service.build_facility_template_signature(
        template,
        facility_id=facility_id,
        skeleton_rows=skeleton_rows,
    )
    template["hakodate_template_signature"] = signature["hash"]
    template["hakodate_template_signature_components"] = signature["components"]
    return template, skeleton_rows


def test_hakodate_assignment_places_numeric_token_in_quantity_cell() -> None:
    result = build_hakodate_assignment(
        tokens=[
            {"text": "12", "x": 0.7, "y": 0.3, "bbox": [0.68, 0.28, 0.72, 0.32]},
            {"text": "99", "x": 0.5, "y": 0.3, "bbox": [0.48, 0.28, 0.52, 0.32]},
        ],
        grid=_grid(),
        template=_template(),
        skeleton_rows=[
            {"row_id": "row-a", "date": "2026-04-26", "daypart": "朝", "menu_name": "献立A"},
        ],
    )

    assert result["status"] == "auto_assignable"
    assert result["metrics"]["assigned_count"] == 1
    assert result["assignments"][0]["row_id"] == "row-a"
    assert result["assignments"][0]["field"] == "qty.regular_x"
    assert result["assignments"][0]["value_normalized"] == "12"
    assert result["rejected_candidates"][0]["reason"] == "non_quantity_cell"


def test_hakodate_assignment_reviews_boundary_and_duplicate_candidates() -> None:
    result = build_hakodate_assignment(
        tokens=[
            {"text": "10", "x": 0.605, "y": 0.3, "bbox": [0.60, 0.28, 0.61, 0.32]},
            {"text": "11", "x": 0.7, "y": 0.5, "bbox": [0.68, 0.48, 0.72, 0.52]},
            {"text": "12", "x": 0.72, "y": 0.5, "bbox": [0.70, 0.48, 0.74, 0.52]},
        ],
        grid=_grid(),
        template=_template(),
        skeleton_rows=[{"row_id": "row-a"}, {"row_id": "row-b"}],
    )

    reasons = {item.get("reason") for item in result["review_candidates"]}
    assert result["status"] == "review_required"
    assert "near_cell_boundary" in reasons
    assert "duplicate_quantity_candidates" in reasons
    assert result["metrics"]["review_count"] == 3


def test_hakodate_assignment_blocks_when_strategy_is_legacy() -> None:
    template = _template()
    template["quantity_assignment_strategy"] = "legacy"

    result = build_hakodate_assignment(
        tokens=[{"text": "12", "x": 0.7, "y": 0.3, "bbox": [0.68, 0.28, 0.72, 0.32]}],
        grid=_grid(),
        template=template,
    )

    assert result["status"] == "blocked"
    assert "hakodate_strategy_not_enabled" in result["blockers"]


def test_hakodate_assignment_blocks_when_template_signature_is_missing() -> None:
    template = _template()
    template.pop("hakodate_template_signature", None)
    template.pop("hakodate_template_signature_components", None)

    result = build_hakodate_assignment(
        tokens=[{"text": "12", "x": 0.7, "y": 0.3, "bbox": [0.68, 0.28, 0.72, 0.32]}],
        grid=_grid(),
        template=template,
    )

    assert result["status"] == "blocked"
    assert "hakodate_template_signature_missing" in result["blockers"]


def test_hakodate_assignment_blocks_when_template_signature_is_stale() -> None:
    template = _template()
    template["main_ocr_row_fields"].append("qty.mixer_x")

    result = build_hakodate_assignment(
        tokens=[{"text": "12", "x": 0.7, "y": 0.3, "bbox": [0.68, 0.28, 0.72, 0.32]}],
        grid=_grid(),
        template=template,
    )

    assert result["status"] == "blocked"
    assert "template_stale_due_to_facility_category_change" in result["blockers"]


def test_hakodate_template_regeneration_candidate_summarizes_quantity_diff() -> None:
    template = _template()
    template["main_ocr_row_fields"].append("qty.mixer_x")

    candidate = hakodate_assignment_service.build_facility_template_regeneration_candidate(template)

    assert candidate["status"] == "candidate_ready"
    assert candidate["reason"] == "template_stale_due_to_facility_category_change"
    assert candidate["diff"]["quantity_fields_added"] == ["qty.mixer_x"]
    assert candidate["candidate_template"]["hakodate_template_signature"] == candidate["candidate_signature"]


def test_hakodate_assignment_from_pdf_blocks_detected_grid_mode_before_token_ocr(monkeypatch) -> None:
    monkeypatch.setattr(
        hakodate_assignment_service,
        "detect_table_grid",
        lambda _pdf_bytes, _template: (_ for _ in ()).throw(AssertionError("detected grid mode must not run")),
    )

    def raise_unavailable(_pdf_bytes, _template):
        raise AssertionError("detected grid mode must not run token OCR")

    monkeypatch.setattr(
        hakodate_assignment_service,
        "_extract_tesseract_tokens",
        raise_unavailable,
    )

    result = hakodate_assignment_service.build_hakodate_assignment_from_pdf(
        pdf_bytes=b"%PDF-1.4",
        template=_legacy_grid_pdf_template(),
        strategy="hakodate",
    )

    assert result["status"] == "blocked"
    assert "hakodate_structure_slot_required" in result["blockers"]
    assert "hakodate_detected_grid_token_assignment_disabled" in result["warnings"]


def test_hakodate_assignment_from_pdf_blocks_detected_grid_template_fallback_path(monkeypatch) -> None:
    monkeypatch.setattr(
        hakodate_assignment_service,
        "detect_table_grid",
        lambda _pdf_bytes, _template: None,
    )
    monkeypatch.setattr(
        hakodate_assignment_service,
        "_extract_tesseract_tokens",
        lambda _pdf_bytes, _template: [
            {"text": "12", "x": 0.7, "y": 0.5, "bbox": [0.68, 0.48, 0.72, 0.52]},
        ],
    )
    template = _legacy_grid_pdf_template()
    template["table_box"] = [0.0, 0.0, 1.0, 1.0]

    result = hakodate_assignment_service.build_hakodate_assignment_from_pdf(
        pdf_bytes=b"%PDF-1.4",
        template=template,
        strategy="hakodate",
        skeleton_rows=[{"row_id": "row-a"}, {"row_id": "row-b"}],
    )

    assert result["status"] == "blocked"
    assert result["assignments"] == []
    assert result["review_candidates"] == []
    assert "hakodate_structure_slot_required" in result["blockers"]
    assert "hakodate_detected_grid_token_assignment_disabled" in result["warnings"]


def test_hakodate_assignment_from_pdf_blocks_detected_grid_unmatched_edges_path(monkeypatch) -> None:
    monkeypatch.setattr(
        hakodate_assignment_service,
        "detect_table_grid",
        lambda _pdf_bytes, _template: {
            "table_box": [0.0, 0.0, 1.0, 1.0],
            "column_edges": [0.0, 0.5, 1.0],
            "row_edges": [0.0, 0.5, 1.0],
            "confidence": 0.8,
        },
    )
    monkeypatch.setattr(
        hakodate_assignment_service,
        "_extract_tesseract_tokens",
        lambda _pdf_bytes, _template: [
            {"text": "12", "x": 0.7, "y": 0.5, "bbox": [0.68, 0.48, 0.72, 0.52]},
        ],
    )

    result = hakodate_assignment_service.build_hakodate_assignment_from_pdf(
        pdf_bytes=b"%PDF-1.4",
        template=_legacy_grid_pdf_template(),
        strategy="hakodate",
        skeleton_rows=[{"row_id": "row-a"}, {"row_id": "row-b"}],
    )

    assert result["status"] == "blocked"
    assert result["assignments"] == []
    assert result["review_candidates"] == []
    assert "hakodate_structure_slot_required" in result["blockers"]


def test_hakodate_assignment_from_pdf_blocks_detected_grid_template_matched_edges_path(monkeypatch) -> None:
    monkeypatch.setattr(
        hakodate_assignment_service,
        "detect_table_grid",
        lambda _pdf_bytes, _template: {
            "table_box": [0.0, 0.0, 1.0, 1.0],
            "column_edges": [0.0, 0.2, 0.4, 0.6, 0.61, 0.8, 1.0],
            "row_edges": [0.0, 0.33, 0.66, 0.67, 1.0],
            "confidence": 0.8,
        },
    )
    monkeypatch.setattr(
        hakodate_assignment_service,
        "_extract_tesseract_tokens",
        lambda _pdf_bytes, _template: [
            {"text": "12", "x": 0.7, "y": 0.5, "bbox": [0.68, 0.48, 0.72, 0.52]},
        ],
    )

    result = hakodate_assignment_service.build_hakodate_assignment_from_pdf(
        pdf_bytes=b"%PDF-1.4",
        template=_legacy_grid_pdf_template(),
        strategy="hakodate",
        skeleton_rows=[{"row_id": "row-a"}, {"row_id": "row-b"}],
    )

    assert result["status"] == "blocked"
    assert result["assignments"] == []
    assert "hakodate_structure_slot_required" in result["blockers"]


def test_structure_slot_assignment_ocr_each_quantity_cell_without_token_ocr(monkeypatch) -> None:
    skeleton_rows = [{"row_id": "row-a", "date": "2026-04-26", "daypart": "朝", "menu_name": "献立A"}]
    structure_grid = {
        "table_box": [0.0, 0.0, 1.0, 1.0],
        "column_edges": [0.0, 0.2, 0.4, 0.6, 0.8, 1.0],
        "row_edges": [0.0, 0.2, 0.4, 0.6],
        "confidence": 1.0,
        "column_edges_source": "structure_workbook_geometry",
        "row_edges_source": "structure_workbook_geometry",
    }
    monkeypatch.setattr(
        hakodate_assignment_service,
        "_structure_grid_for_facility_template",
        lambda **_kwargs: (structure_grid, _structure_slot_worksheet()),
    )
    monkeypatch.setattr(
        hakodate_assignment_service,
        "detect_table_grid",
        lambda _pdf_bytes, _template: {
            "table_box": [0.1, 0.1, 0.9, 0.7],
            "column_edges": [0.1, 0.23, 0.38, 0.56, 0.73, 0.9],
            "row_edges": [0.1, 0.24, 0.45, 0.7],
            "confidence": 0.86,
        },
    )

    def fail_token_ocr(_pdf_bytes, _template):
        raise AssertionError("structure_slot must not use whole-table token OCR")

    monkeypatch.setattr(hakodate_assignment_service, "_extract_tesseract_tokens", fail_token_ocr)
    monkeypatch.setattr(
        hakodate_assignment_service,
        "_render_pdf_page_image",
        lambda _pdf_bytes, _template: Image.new("RGB", (1000, 1000), "white"),
    )

    seen_cell_boxes = []

    def fake_cell_ocr(**kwargs):
        seen_cell_boxes.append(kwargs["fax_cell_bbox"])
        return hakodate_assignment_service.HakodateCellOcrResult(text="12", normalized="12")

    monkeypatch.setattr(hakodate_assignment_service, "_ocr_quantity_cell", fake_cell_ocr)

    result = hakodate_assignment_service.build_structure_slot_assignment_from_pdf(
        pdf_bytes=b"%PDF-1.4",
        template=_structure_slot_template(skeleton_rows),
        strategy="hakodate",
        skeleton_rows=skeleton_rows,
    )

    assert result["status"] == "auto_assignable"
    assert len(seen_cell_boxes) == 1
    assert result["assignments"][0]["row_id"] == "row-a"
    assert result["assignments"][0]["field"] == "qty.regular_x"
    assert result["assignments"][0]["sheet_cell"] == "E11"
    assert result["assignments"][0]["ocr_scope"] == "cell_crop"
    assert result["assignments"][0]["ocr_engine_bbox_used_for_assignment"] is False
    assert result["assignments"][0]["fax_cell_bbox"] == seen_cell_boxes[0]
    assert result["metrics"]["ocr_cell_count"] == 1


def test_structure_slot_assignment_uses_merged_quantity_bbox_for_each_spanned_row(monkeypatch) -> None:
    skeleton_rows = [
        {"row_id": "row-a", "date": "2026-04-26", "daypart": "朝", "menu_name": "献立A"},
        {"row_id": "row-b", "date": "2026-04-26", "daypart": "朝", "menu_name": "献立B"},
    ]
    structure_grid = {
        "table_box": [0.0, 0.0, 1.0, 1.0],
        "column_edges": [0.0, 0.2, 0.4, 0.6, 0.8, 1.0],
        "row_edges": [0.0, 0.2, 0.4, 0.6, 0.8],
        "confidence": 1.0,
        "column_edges_source": "structure_workbook_geometry",
        "row_edges_source": "structure_workbook_geometry",
    }
    monkeypatch.setattr(
        hakodate_assignment_service,
        "_structure_grid_for_facility_template",
        lambda **_kwargs: (structure_grid, _merged_structure_slot_worksheet()),
    )
    monkeypatch.setattr(
        hakodate_assignment_service,
        "detect_table_grid",
        lambda _pdf_bytes, _template: {
            "table_box": [0.1, 0.1, 0.9, 0.7],
            "column_edges": [0.1, 0.23, 0.38, 0.56, 0.73, 0.9],
            "row_edges": [0.1, 0.24, 0.45, 0.58, 0.7],
            "confidence": 0.86,
        },
    )
    monkeypatch.setattr(
        hakodate_assignment_service,
        "_render_pdf_page_image",
        lambda _pdf_bytes, _template: Image.new("RGB", (1000, 1000), "white"),
    )

    seen_cell_boxes = []

    def fake_cell_ocr(**kwargs):
        seen_cell_boxes.append(kwargs["fax_cell_bbox"])
        return hakodate_assignment_service.HakodateCellOcrResult(text="20", normalized="20")

    monkeypatch.setattr(hakodate_assignment_service, "_ocr_quantity_cell", fake_cell_ocr)

    result = hakodate_assignment_service.build_structure_slot_assignment_from_pdf(
        pdf_bytes=b"%PDF-1.4",
        template=_structure_slot_template(skeleton_rows),
        strategy="hakodate",
        skeleton_rows=skeleton_rows,
    )

    assert result["status"] == "auto_assignable"
    assert [item["row_id"] for item in result["assignments"]] == ["row-a", "row-b"]
    assert [item["sheet_cell"] for item in result["assignments"]] == ["E11", "E12"]
    assert seen_cell_boxes == [[0.73, 0.45, 0.9, 0.7], [0.73, 0.45, 0.9, 0.7]]
    assert {item["value_normalized"] for item in result["assignments"]} == {"20"}
    assert result["assignments"][0]["merged_cell"]["range"] == "E11:E12"
    assert result["assignments"][1]["merged_cell"]["row_span"] == 2


def test_column_slots_detect_diabetes_quantity_from_template_header() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.cell(row=7, column=1, value="日付")
    worksheet.cell(row=7, column=2, value="区分")
    worksheet.cell(row=7, column=4, value="献立")
    worksheet.cell(row=7, column=5, value="常食")
    worksheet.cell(row=7, column=6, value="糖尿")

    slots = hakodate_assignment_service._column_slots_from_worksheet(  # noqa: SLF001
        worksheet,
        col_count=6,
    )

    assert slots[4]["role"] == "quantity"
    assert slots[4]["slot_name"] == "qty.regular_x"
    assert slots[5]["role"] == "quantity"
    assert slots[5]["slot_name"] == "qty.diabetes_x"


def test_structure_slot_assignment_blocks_when_detected_table_box_is_invalid(monkeypatch) -> None:
    skeleton_rows = [{"row_id": "row-a", "date": "2026-04-26", "daypart": "朝", "menu_name": "献立A"}]
    structure_grid = {
        "table_box": [0.0, 0.0, 1.0, 1.0],
        "column_edges": [0.0, 0.2, 0.4, 0.6, 0.8, 1.0],
        "row_edges": [0.0, 0.2, 0.4, 0.6],
        "confidence": 1.0,
        "column_edges_source": "structure_workbook_geometry",
        "row_edges_source": "structure_workbook_geometry",
    }
    monkeypatch.setattr(
        hakodate_assignment_service,
        "_structure_grid_for_facility_template",
        lambda **_kwargs: (structure_grid, _structure_slot_worksheet()),
    )
    monkeypatch.setattr(
        hakodate_assignment_service,
        "detect_table_grid",
        lambda _pdf_bytes, _template: {
            "table_box": [],
            "column_edges": [0.1, 0.5, 0.9],
            "row_edges": [0.1, 0.7],
            "confidence": 0.86,
        },
    )
    monkeypatch.setattr(
        hakodate_assignment_service,
        "_extract_tesseract_tokens",
        lambda *_args, **_kwargs: (_ for _ in ()).throw(AssertionError("no token fallback")),
    )
    monkeypatch.setattr(
        hakodate_assignment_service,
        "_ocr_quantity_cell",
        lambda **_kwargs: (_ for _ in ()).throw(AssertionError("invalid table box must block before OCR")),
    )
    monkeypatch.setattr(
        hakodate_assignment_service,
        "_render_pdf_page_image",
        lambda _pdf_bytes, _template: Image.new("RGB", (1000, 1000), "white"),
    )

    result = hakodate_assignment_service.build_structure_slot_assignment_from_pdf(
        pdf_bytes=b"%PDF-1.4",
        template=_structure_slot_template(skeleton_rows),
        strategy="hakodate",
        skeleton_rows=skeleton_rows,
    )

    assert result["status"] == "blocked"
    assert result["assignments"] == []
    assert "structure_alignment_box_missing" in result["blockers"]


@pytest.mark.parametrize(
    ("case_name", "facility_id", "pdf_path", "expected_col_source", "expected_status"),
    [
        (
            "fac14_stg_corrected",
            "FAC00014",
            Path("/Users/mmorinag/Sawa/2025.12/workspace/tmp/stg_fac14_compare/stg_corrected.pdf"),
            "detected_template_matched",
            "auto_assignable",
        ),
        (
            "fac14_fax000364233",
            "FAC00014",
            Path("/Users/mmorinag/Sawa/2025.12/tmp/reupload_legacy_split/19.fax000364233_0426_0501_.pdf"),
            "detected_template_matched",
            "auto_assignable",
        ),
        (
            "fac01_degraded_document2",
            "FAC00001",
            Path("/Users/mmorinag/Downloads/Document (2).pdf"),
            "structure_table_box_projection",
            "blocked",
        ),
        (
            "fac02_sibling_stg_corrected",
            "FAC00002",
            Path("/Users/mmorinag/Sawa/2025.12/workspace/tmp/stg_fac14_compare/stg_corrected.pdf"),
            "detected_template_matched",
            "auto_assignable",
        ),
        (
            "fac07_sibling_stg_corrected",
            "FAC00007",
            Path("/Users/mmorinag/Sawa/2025.12/workspace/tmp/stg_fac14_compare/stg_corrected.pdf"),
            "detected_template_matched",
            "auto_assignable",
        ),
    ],
)
def test_structure_slot_real_pdfs_cover_representative_degraded_and_sibling_cases(
    monkeypatch,
    tmp_path,
    case_name: str,
    facility_id: str,
    pdf_path: Path,
    expected_col_source: str,
    expected_status: str,
) -> None:
    if not pdf_path.exists():
        pytest.skip(f"local verification PDF not found: {pdf_path}")

    def fake_cell_ocr(**_kwargs):
        return hakodate_assignment_service.HakodateCellOcrResult(text="1", normalized="1")

    monkeypatch.setattr(hakodate_assignment_service, "_ocr_quantity_cell", fake_cell_ocr)
    template, skeleton_rows = _real_pdf_template(facility_id, tmp_path / case_name)

    result = hakodate_assignment_service.build_structure_slot_assignment_from_pdf(
        pdf_bytes=pdf_path.read_bytes(),
        template=template,
        strategy="hakodate",
        skeleton_rows=skeleton_rows,
    )

    slot_grid = result["grid"]["actual_structure_slot_grid"]
    assert result["status"] == expected_status
    if expected_status == "blocked":
        assert result["blockers"] == ["structure_slot_alignment_unverified_by_actual_cell_grid"]
    else:
        assert result["blockers"] == []
        assert result["metrics"]["structure_grid_row_count"] == 59
        assert result["metrics"]["structure_grid_column_count"] == 12
    assert len(slot_grid["row_edges"]) == 60
    assert len(slot_grid["column_edges"]) == 13
    assert slot_grid["row_edges_source"] == "dense_horizontal_lines"
    assert slot_grid["column_edges_source"] == expected_col_source
    if expected_status == "auto_assignable":
        assert result["metrics"]["assigned_count"] > 0
        assert result["metrics"]["review_count"] == 0
    else:
        assert result["metrics"]["assigned_count"] == 0
        assert result["metrics"]["review_count"] == 0
        assert "structure_slot_alignment_blocked" in result["warnings"]
        assert "artifacts" not in result
        return
    assert Path(result["artifacts"]["overlay_image_path"]).exists()
