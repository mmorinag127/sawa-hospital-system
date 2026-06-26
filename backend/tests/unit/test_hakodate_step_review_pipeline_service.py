import numpy as np
from openpyxl import Workbook

from src.services.hakodate_step_review_pipeline_service import (
    _detect_vertical_candidates,
    _draw_merge_aware_grid,
    _fit_extra_y_clusters_to_template_count,
    _ordered_match_y_clusters_to_template,
    _physical_internal_horizontal_lines,
    _post_menu_boundary_preserving_xs,
    _post_menu_target_regions,
    _row_height_outlier_evidence,
    _step_review_merge_regions_for_grid,
    _step_review_physical_row_map,
    _step_review_worksheet_row_to_grid_index,
    dewarp_rectified_rows_by_bounded_slant,
    dewarp_rectified_y_to_template_rows,
    snap_regions_x_to_local_fax_rulings,
)


def test_physical_internal_horizontal_lines_detects_full_width_rule() -> None:
    mask = np.zeros((80, 80), dtype=np.uint8)
    mask[30, 10:50] = 255

    hits = _physical_internal_horizontal_lines(
        horizontal_line_mask=mask,
        row_edges=[10.0, 30.0, 50.0],
        bbox=[10.0, 10.0, 50.0, 50.0],
        start_row_index=0,
        row_span=2,
    )

    assert hits == [
        {
            "boundary_row_index": 1,
            "expected_y": 30.0,
            "detected_y": 30,
            "line_ratio": 1.0,
        }
    ]


def test_physical_internal_horizontal_lines_ignores_short_marks() -> None:
    mask = np.zeros((80, 80), dtype=np.uint8)
    mask[30, 10:20] = 255

    hits = _physical_internal_horizontal_lines(
        horizontal_line_mask=mask,
        row_edges=[10.0, 30.0, 50.0],
        bbox=[10.0, 10.0, 50.0, 50.0],
        start_row_index=0,
        row_span=2,
    )

    assert hits == []


def test_physical_internal_horizontal_lines_ignores_unmerged_rows() -> None:
    mask = np.zeros((80, 80), dtype=np.uint8)
    mask[30, 10:50] = 255

    hits = _physical_internal_horizontal_lines(
        horizontal_line_mask=mask,
        row_edges=[10.0, 30.0, 50.0],
        bbox=[10.0, 10.0, 50.0, 50.0],
        start_row_index=0,
        row_span=1,
    )

    assert hits == []


def test_snap_regions_uses_cell_local_y_rulings_for_slanted_rows() -> None:
    rectified = np.full((80, 120, 3), 255, dtype=np.uint8)
    for x in (40, 60, 80):
        rectified[8:55, x - 1 : x + 2] = 0
    rectified[22:24, 40:60] = 0
    rectified[42:44, 40:60] = 0
    rectified[15:17, 60:80] = 0
    rectified[35:37, 60:80] = 0
    regions = [
        {"region_id": "E11", "sheet_cell": "E11", "bbox": [40.0, 20.0, 60.0, 40.0]},
        {"region_id": "F11", "sheet_cell": "F11", "bbox": [60.0, 20.0, 80.0, 40.0]},
    ]

    snapped, evidence = snap_regions_x_to_local_fax_rulings(rectified, regions)

    by_cell = {str(region["sheet_cell"]): region for region in snapped}
    assert evidence["applied"] is True
    assert abs(float(by_cell["E11"]["bbox"][1]) - 22.0) < 1.0
    assert abs(float(by_cell["E11"]["bbox"][3]) - 42.0) < 1.0
    assert abs(float(by_cell["F11"]["bbox"][1]) - 15.0) < 1.0
    assert abs(float(by_cell["F11"]["bbox"][3]) - 35.0) < 1.0
    assert by_cell["F11"]["local_grid_snap"]["local_y_snap_applied"] is True


def test_snap_regions_preserves_slanted_cell_rulings_as_polygon() -> None:
    rectified = np.full((100, 150, 3), 255, dtype=np.uint8)
    for x in (40, 90):
        rectified[15:85, x - 1 : x + 2] = 0
    for x in range(40, 91):
        y_top = int(round(30.0 + 8.0 * ((x - 40) / 50.0)))
        y_bottom = int(round(58.0 + 8.0 * ((x - 40) / 50.0)))
        rectified[y_top - 1 : y_top + 2, x] = 0
        rectified[y_bottom - 1 : y_bottom + 2, x] = 0
    regions = [
        {"region_id": "E11", "sheet_cell": "E11", "bbox": [40.0, 30.0, 90.0, 58.0]},
    ]

    snapped, evidence = snap_regions_x_to_local_fax_rulings(rectified, regions)

    region = snapped[0]
    polygon = region["polygon"]
    assert evidence["applied"] is True
    assert region["local_grid_snap"]["local_polygon_snap_applied"] is True
    assert len(polygon) == 4
    assert abs(float(polygon[0][1]) - 30.0) < 2.5
    assert abs(float(polygon[1][1]) - 38.0) < 2.5
    assert abs(float(polygon[2][1]) - 66.0) < 2.5
    assert abs(float(polygon[3][1]) - 58.0) < 2.5
    assert float(region["bbox"][1]) <= min(float(point[1]) for point in polygon)
    assert float(region["bbox"][3]) >= max(float(point[1]) for point in polygon)


def test_snap_regions_uses_shared_row_boundary_curves_for_distorted_table() -> None:
    rectified = np.full((140, 180, 3), 255, dtype=np.uint8)
    x_edges = [30, 60, 90, 120, 150]
    for x in x_edges:
        rectified[18:112, x - 1 : x + 2] = 0
    for base_y in (30.0, 58.0, 86.0):
        for x in range(x_edges[0], x_edges[-1] + 1):
            y = int(round(base_y + 10.0 * ((x - x_edges[0]) / (x_edges[-1] - x_edges[0]))))
            rectified[y - 1 : y + 2, x] = 0
    regions = [
        {"region_id": "E11", "sheet_cell": "E11", "bbox": [30.0, 30.0, 60.0, 58.0]},
        {"region_id": "F11", "sheet_cell": "F11", "bbox": [60.0, 30.0, 90.0, 58.0]},
        {"region_id": "G11", "sheet_cell": "G11", "bbox": [90.0, 30.0, 120.0, 58.0]},
        {"region_id": "H11", "sheet_cell": "H11", "bbox": [120.0, 30.0, 150.0, 58.0]},
        {"region_id": "E12", "sheet_cell": "E12", "bbox": [30.0, 58.0, 60.0, 86.0]},
        {"region_id": "F12", "sheet_cell": "F12", "bbox": [60.0, 58.0, 90.0, 86.0]},
        {"region_id": "G12", "sheet_cell": "G12", "bbox": [90.0, 58.0, 120.0, 86.0]},
        {"region_id": "H12", "sheet_cell": "H12", "bbox": [120.0, 58.0, 150.0, 86.0]},
    ]

    snapped, evidence = snap_regions_x_to_local_fax_rulings(rectified, regions)

    assert evidence["applied"] is True
    assert evidence["row_boundary_curve_count"] == 3
    by_cell = {str(region["sheet_cell"]): region for region in snapped}
    left_polygon = by_cell["E11"]["polygon"]
    right_polygon = by_cell["H11"]["polygon"]
    assert by_cell["E11"]["local_grid_snap"]["polygon_method"] == "shared_row_boundary_curve"
    assert by_cell["H11"]["local_grid_snap"]["polygon_method"] == "shared_row_boundary_curve"
    assert abs(float(left_polygon[0][1]) - 30.0) < 3.0
    assert abs(float(left_polygon[1][1]) - 32.5) < 3.0
    assert abs(float(right_polygon[0][1]) - 37.5) < 3.0
    assert abs(float(right_polygon[1][1]) - 40.0) < 3.0


def test_snap_regions_extrapolates_row_curve_when_right_ruling_is_missing() -> None:
    rectified = np.full((140, 180, 3), 255, dtype=np.uint8)
    x_edges = [30, 60, 90, 120, 150]
    for x in x_edges:
        rectified[18:112, x - 1 : x + 2] = 0
    for base_y in (30.0, 58.0, 86.0):
        for x in range(x_edges[0], 121):
            y = int(round(base_y + 12.0 * ((x - x_edges[0]) / 90.0)))
            rectified[y - 1 : y + 2, x] = 0
    regions = [
        {"region_id": "E11", "sheet_cell": "E11", "bbox": [30.0, 30.0, 60.0, 58.0]},
        {"region_id": "F11", "sheet_cell": "F11", "bbox": [60.0, 30.0, 90.0, 58.0]},
        {"region_id": "G11", "sheet_cell": "G11", "bbox": [90.0, 30.0, 120.0, 58.0]},
        {"region_id": "H11", "sheet_cell": "H11", "bbox": [120.0, 30.0, 150.0, 58.0]},
        {"region_id": "E12", "sheet_cell": "E12", "bbox": [30.0, 58.0, 60.0, 86.0]},
        {"region_id": "F12", "sheet_cell": "F12", "bbox": [60.0, 58.0, 90.0, 86.0]},
        {"region_id": "G12", "sheet_cell": "G12", "bbox": [90.0, 58.0, 120.0, 86.0]},
        {"region_id": "H12", "sheet_cell": "H12", "bbox": [120.0, 58.0, 150.0, 86.0]},
    ]

    snapped, evidence = snap_regions_x_to_local_fax_rulings(rectified, regions)

    assert evidence["applied"] is True
    by_cell = {str(region["sheet_cell"]): region for region in snapped}
    right_polygon = by_cell["H11"]["polygon"]
    assert by_cell["H11"]["local_grid_snap"]["polygon_method"] == "shared_row_boundary_curve"
    assert float(right_polygon[1][1]) > float(right_polygon[0][1]) + 2.0
    assert abs(float(right_polygon[1][1]) - 46.0) < 4.0


def test_snap_regions_allows_shared_row_curve_on_wide_cells_without_noisy_fallback() -> None:
    rectified = np.full((140, 240, 3), 255, dtype=np.uint8)
    x_edges = [30, 70, 110, 210]
    for x in x_edges:
        rectified[18:112, x - 1 : x + 2] = 0
    for base_y in (30.0, 58.0, 86.0):
        for x in range(x_edges[0], x_edges[-1] + 1):
            y = int(round(base_y + 18.0 * ((x - x_edges[0]) / (x_edges[-1] - x_edges[0]))))
            if x <= 160:
                rectified[y - 1 : y + 2, x] = 0
    for y in range(70, 78):
        rectified[y, 150:210] = 0
    regions = [
        {"region_id": "E11", "sheet_cell": "E11", "bbox": [30.0, 30.0, 70.0, 58.0]},
        {"region_id": "F11", "sheet_cell": "F11", "bbox": [70.0, 30.0, 110.0, 58.0]},
        {"region_id": "G11", "sheet_cell": "G11", "bbox": [110.0, 30.0, 210.0, 58.0]},
        {"region_id": "E12", "sheet_cell": "E12", "bbox": [30.0, 58.0, 70.0, 86.0]},
        {"region_id": "F12", "sheet_cell": "F12", "bbox": [70.0, 58.0, 110.0, 86.0]},
        {"region_id": "G12", "sheet_cell": "G12", "bbox": [110.0, 58.0, 210.0, 86.0]},
    ]

    snapped, evidence = snap_regions_x_to_local_fax_rulings(rectified, regions)

    assert evidence["applied"] is True
    by_cell = {str(region["sheet_cell"]): region for region in snapped}
    assert by_cell["G11"]["local_grid_snap"]["polygon_method"] == "shared_row_boundary_curve"
    assert by_cell["G11"]["bbox"] != regions[2]["bbox"]
    assert abs(float(by_cell["G11"]["polygon"][1][1]) - 48.0) < 5.0


def test_snap_regions_keeps_strong_wide_cell_curve_instead_of_template_fallback() -> None:
    rectified = np.full((190, 260, 3), 255, dtype=np.uint8)
    x_edges = [30, 70, 110, 160, 210]
    for x in x_edges:
        rectified[28:170, x - 1 : x + 2] = 0
    for base_y in (40.0, 68.0, 96.0):
        for x in range(x_edges[0], x_edges[-1] + 1):
            y = int(round(base_y + 36.0 * ((x - x_edges[0]) / (x_edges[-1] - x_edges[0]))))
            rectified[y - 1 : y + 2, x] = 0
    regions = [
        {"region_id": "E11", "sheet_cell": "E11", "bbox": [30.0, 40.0, 70.0, 68.0]},
        {"region_id": "F11", "sheet_cell": "F11", "bbox": [70.0, 40.0, 110.0, 68.0]},
        {"region_id": "G11", "sheet_cell": "G11", "bbox": [110.0, 40.0, 210.0, 68.0]},
        {"region_id": "E12", "sheet_cell": "E12", "bbox": [30.0, 68.0, 70.0, 96.0]},
        {"region_id": "F12", "sheet_cell": "F12", "bbox": [70.0, 68.0, 110.0, 96.0]},
        {"region_id": "G12", "sheet_cell": "G12", "bbox": [110.0, 68.0, 210.0, 96.0]},
    ]

    snapped, evidence = snap_regions_x_to_local_fax_rulings(rectified, regions)

    assert evidence["applied"] is True
    by_cell = {str(region["sheet_cell"]): region for region in snapped}
    wide_cell = by_cell["G12"]
    assert wide_cell["local_grid_snap"]["polygon_method"] == "shared_row_boundary_curve"
    assert wide_cell["bbox"] != regions[5]["bbox"]
    assert float(wide_cell["polygon"][3][1]) > float(wide_cell["polygon"][2][1]) + 18.0
    assert abs(float(wide_cell["polygon"][2][1]) - 84.0) < 6.0


def test_snap_regions_uses_column_boundary_curves_for_vertical_distortion() -> None:
    rectified = np.full((150, 190, 3), 255, dtype=np.uint8)
    y_edges = [30, 58, 86, 114]
    x_bases = [40, 80, 120, 160]
    for y in y_edges:
        rectified[y - 1 : y + 2, 35:166] = 0
    for base_x in x_bases:
        for y in range(y_edges[0], y_edges[-1] + 1):
            x = int(round(base_x + 8.0 * ((y - y_edges[0]) / (y_edges[-1] - y_edges[0]))))
            rectified[y, x - 1 : x + 2] = 0
    regions = [
        {"region_id": "E11", "sheet_cell": "E11", "bbox": [40.0, 30.0, 80.0, 58.0]},
        {"region_id": "F11", "sheet_cell": "F11", "bbox": [80.0, 30.0, 120.0, 58.0]},
        {"region_id": "G11", "sheet_cell": "G11", "bbox": [120.0, 30.0, 160.0, 58.0]},
        {"region_id": "E12", "sheet_cell": "E12", "bbox": [40.0, 58.0, 80.0, 86.0]},
        {"region_id": "F12", "sheet_cell": "F12", "bbox": [80.0, 58.0, 120.0, 86.0]},
        {"region_id": "G12", "sheet_cell": "G12", "bbox": [120.0, 58.0, 160.0, 86.0]},
        {"region_id": "E13", "sheet_cell": "E13", "bbox": [40.0, 86.0, 80.0, 114.0]},
        {"region_id": "F13", "sheet_cell": "F13", "bbox": [80.0, 86.0, 120.0, 114.0]},
        {"region_id": "G13", "sheet_cell": "G13", "bbox": [120.0, 86.0, 160.0, 114.0]},
    ]

    snapped, evidence = snap_regions_x_to_local_fax_rulings(rectified, regions)

    assert evidence["applied"] is True
    assert evidence["column_boundary_curve_count"] >= 4
    by_cell = {str(region["sheet_cell"]): region for region in snapped}
    polygon = by_cell["E13"]["polygon"]
    assert by_cell["E13"]["local_grid_snap"]["polygon_method"] == "shared_row_and_column_boundary_curve"
    assert by_cell["E13"]["local_grid_snap"]["column_curve_applied"] is True
    assert float(polygon[3][0]) - float(polygon[0][0]) > 1.0
    assert abs(float(polygon[0][0]) - 45.0) < 3.0
    assert abs(float(polygon[3][0]) - 48.0) < 3.0


def test_snap_regions_rejects_unsafe_column_curve_kink() -> None:
    rectified = np.full((150, 190, 3), 255, dtype=np.uint8)
    y_edges = [30, 58, 86, 114]
    x_bases = [40, 80, 120, 160]
    for y in y_edges:
        rectified[y - 1 : y + 2, 35:166] = 0
    for base_x in x_bases:
        for y in range(y_edges[0], y_edges[-1] + 1):
            x = int(round(base_x + 30.0 * ((y - y_edges[0]) / (y_edges[-1] - y_edges[0]))))
            rectified[y, x - 1 : x + 2] = 0
    regions = [
        {"region_id": "E11", "sheet_cell": "E11", "bbox": [40.0, 30.0, 80.0, 58.0]},
        {"region_id": "F11", "sheet_cell": "F11", "bbox": [80.0, 30.0, 120.0, 58.0]},
        {"region_id": "G11", "sheet_cell": "G11", "bbox": [120.0, 30.0, 160.0, 58.0]},
        {"region_id": "E12", "sheet_cell": "E12", "bbox": [40.0, 58.0, 80.0, 86.0]},
        {"region_id": "F12", "sheet_cell": "F12", "bbox": [80.0, 58.0, 120.0, 86.0]},
        {"region_id": "G12", "sheet_cell": "G12", "bbox": [120.0, 58.0, 160.0, 86.0]},
        {"region_id": "E13", "sheet_cell": "E13", "bbox": [40.0, 86.0, 80.0, 114.0]},
        {"region_id": "F13", "sheet_cell": "F13", "bbox": [80.0, 86.0, 120.0, 114.0]},
        {"region_id": "G13", "sheet_cell": "G13", "bbox": [120.0, 86.0, 160.0, 114.0]},
    ]

    snapped, evidence = snap_regions_x_to_local_fax_rulings(rectified, regions)

    assert evidence["applied"] is False
    assert evidence["reason"] == "local_grid_snap_insufficient_matches"
    assert evidence["column_boundary_curve_count"] >= 4
    assert evidence["fallback_region_count"] >= 1
    assert snapped == regions


def test_snap_regions_can_preserve_template_y_after_row_dewarp() -> None:
    rectified = np.full((80, 120, 3), 255, dtype=np.uint8)
    for x in (39, 61, 81):
        rectified[8:55, x - 1 : x + 2] = 0
    rectified[22:24, 40:60] = 0
    rectified[42:44, 40:60] = 0
    rectified[15:17, 60:80] = 0
    rectified[35:37, 60:80] = 0
    regions = [
        {"region_id": "E11", "sheet_cell": "E11", "bbox": [40.0, 20.0, 60.0, 40.0]},
        {"region_id": "F11", "sheet_cell": "F11", "bbox": [60.0, 20.0, 80.0, 40.0]},
    ]

    snapped, evidence = snap_regions_x_to_local_fax_rulings(rectified, regions, snap_y=False)

    by_cell = {str(region["sheet_cell"]): region for region in snapped}
    assert evidence["applied"] is True
    assert evidence["y_snap_enabled"] is False
    assert by_cell["E11"]["bbox"][1] == 20.0
    assert by_cell["E11"]["bbox"][3] == 40.0
    assert by_cell["F11"]["bbox"][1] == 20.0
    assert by_cell["F11"]["bbox"][3] == 40.0
    assert abs(float(by_cell["E11"]["bbox"][0]) - 39.0) < 1.0
    assert abs(float(by_cell["F11"]["bbox"][2]) - 81.0) < 1.0
    assert by_cell["F11"]["local_grid_snap"]["y_snap_enabled"] is False
    assert by_cell["F11"]["local_grid_snap"]["local_y_snap_applied"] is False


def test_dewarp_rectified_y_to_template_rows_moves_source_rows_to_template_positions() -> None:
    rectified = np.full((80, 100, 3), 255, dtype=np.uint8)
    source_ys = [10.0, 24.0, 43.0, 68.0]
    template_ys = [10.0, 30.0, 50.0, 70.0]
    for y in source_ys:
        yy = int(round(y))
        rectified[yy - 1 : yy + 2, 20:80] = 0

    dewarped, evidence = dewarp_rectified_y_to_template_rows(
        rectified,
        source_ys=source_ys,
        template_ys=template_ys,
    )

    gray = dewarped[:, 20:80, 0]
    projection = (gray < 80).sum(axis=1)
    detected = [index for index, value in enumerate(projection.tolist()) if value >= 40]
    assert evidence["applied"] is True
    for expected_y in template_ys:
        assert min(abs(index - int(round(expected_y))) for index in detected) <= 1


def test_dewarp_rectified_rows_by_bounded_slant_flattens_moderate_row_tilt() -> None:
    rectified = np.full((120, 180, 3), 255, dtype=np.uint8)
    x0, x1 = 20, 150
    for x in (20, 80, 150):
        rectified[10:100, x - 1 : x + 2] = 0
    for base_y in (20, 50, 80):
        for x in range(x0, x1 + 1):
            y = int(round(float(base_y) + 5.0 * (float(x - x0) / float(x1 - x0))))
            rectified[y - 1 : y + 2, x] = 0

    dewarped, evidence = dewarp_rectified_rows_by_bounded_slant(
        rectified,
        corrected_xs=[20.0, 80.0, 150.0],
        template_ys=[20.0, 50.0, 80.0],
    )

    assert evidence["applied"] is True
    for x_start, x_end in ((22, 28), (142, 148)):
        gray = dewarped[:, x_start:x_end, 0]
        projection = (gray < 80).sum(axis=1)
        detected = [index for index, value in enumerate(projection.tolist()) if value >= 2]
        for expected_y in (20, 50, 80):
            assert min(abs(index - expected_y) for index in detected) <= 2


def test_step_review_body_rows_start_after_preserved_two_stage_header_boundary() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.cell(row=11, column=4, value="先頭メニュー")
    worksheet.cell(row=66, column=4, value="56行目メニュー")
    worksheet.cell(row=67, column=4, value="57行目メニュー")
    worksheet.cell(row=68, column=4, value="")

    rows = _step_review_physical_row_map(worksheet, row_count=59)

    assert rows[2]["worksheet_row"] == 11
    assert rows[2]["menu_name"] == "先頭メニュー"
    assert rows[57]["worksheet_row"] == 66
    assert rows[57]["menu_name"] == "56行目メニュー"
    assert rows[58]["worksheet_row"] == 67
    assert rows[58]["menu_name"] == "57行目メニュー"
    assert 59 not in rows


def test_step_review_worksheet_row_to_grid_index_preserves_two_stage_header_boundary() -> None:
    assert _step_review_worksheet_row_to_grid_index(7) == 0
    assert _step_review_worksheet_row_to_grid_index(8) == 0
    assert _step_review_worksheet_row_to_grid_index(9) == 1
    assert _step_review_worksheet_row_to_grid_index(10) == 1
    assert _step_review_worksheet_row_to_grid_index(11) == 2
    assert _step_review_worksheet_row_to_grid_index(12) == 3
    assert _step_review_worksheet_row_to_grid_index(66) == 57
    assert _step_review_worksheet_row_to_grid_index(67) == 58
    assert _step_review_worksheet_row_to_grid_index(68) is None


def test_full_table_row_axis_match_interpolates_missing_template_rows() -> None:
    template_ys = [0, 10, 20, 30, 40, 50]
    y_clusters = [
        {"cluster_index": 0, "value": 0.0},
        {"cluster_index": 1, "value": 10.0},
        {"cluster_index": 2, "value": 20.0},
        {"cluster_index": 3, "value": 40.0},
        {"cluster_index": 4, "value": 50.0},
    ]

    corrected, evidence = _ordered_match_y_clusters_to_template(
        template_ys=template_ys,
        y_clusters=y_clusters,
    )

    assert evidence["used"] is True
    assert evidence["skipped_template_indexes"] == [3]
    assert evidence["skipped_cluster_indexes"] == []
    assert corrected == [0.0, 10.0, 20.0, 30.0, 40.0, 50.0]


def test_full_table_row_axis_match_keeps_detected_boundary_when_extra_cluster_exists() -> None:
    template_ys = [0, 10, 20, 30, 40, 50]
    y_clusters = [
        {"cluster_index": 0, "value": 0.0},
        {"cluster_index": 1, "value": 10.0},
        {"cluster_index": 2, "value": 20.0},
        {"cluster_index": 3, "value": 30.0},
        {"cluster_index": 4, "value": 40.0},
        {"cluster_index": 5, "value": 50.0},
        {"cluster_index": 6, "value": 110.0},
    ]

    corrected, evidence = _ordered_match_y_clusters_to_template(
        template_ys=template_ys,
        y_clusters=y_clusters,
    )

    assert evidence["used"] is True
    assert evidence["method"] == "ordered_full_table_y_intersection_match_with_extra_cluster_fit"
    assert evidence["skipped_cluster_indexes"] == [6]
    assert corrected == [0.0, 10.0, 20.0, 30.0, 40.0, 50.0]


def test_extra_y_cluster_fit_prefers_balanced_row_heights_over_middle_skip() -> None:
    fitted, evidence = _fit_extra_y_clusters_to_template_count(
        template_ys=[0, 10, 20, 30, 40, 50],
        clusters=[0, 10, 20, 30, 40, 50, 110],
    )

    assert evidence["used"] is True
    assert evidence["skipped_cluster_indexes"] == [6]
    assert fitted == [0.0, 10.0, 20.0, 30.0, 40.0, 50.0]


def test_row_height_outlier_evidence_requires_manual_review() -> None:
    evidence = _row_height_outlier_evidence([0, 10, 20, 30, 55, 65])

    assert evidence["manual_review_required"] is True
    assert evidence["reason"] == "row_height_outlier_detected"
    assert evidence["outliers"] == [
        {"row_band_index": 3, "height": 25.0, "median_ratio": 2.5}
    ]


def test_step_review_header_merge_regions_use_grid_band_span() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.merge_cells("G7:H8")
    worksheet["G7"] = "禁食"
    worksheet.merge_cells("G9:G10")
    worksheet["G9"] = "肉禁"
    worksheet.merge_cells("H9:H10")
    worksheet["H9"] = "魚禁"
    worksheet.merge_cells("E7:E10")
    worksheet["E7"] = "常食"

    regions = {
        str(region["range"]): region
        for region in _step_review_merge_regions_for_grid(
            worksheet,
            row_edges=[float(i) for i in range(59)],
            column_edges=[float(i) for i in range(13)],
        )
    }

    assert regions["G7:H8"]["start_row_index"] == 0
    assert regions["G7:H8"]["end_row_index"] == 0
    assert regions["G7:H8"]["row_span"] == 1
    assert regions["G9:G10"]["start_row_index"] == 1
    assert regions["G9:G10"]["end_row_index"] == 1
    assert regions["G9:G10"]["row_span"] == 1
    assert regions["H9:H10"]["start_row_index"] == 1
    assert regions["E7:E10"]["start_row_index"] == 0
    assert regions["E7:E10"]["end_row_index"] == 1
    assert regions["E7:E10"]["row_span"] == 2


def test_step_review_grid_keeps_template_merge_when_fax_has_internal_line() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.merge_cells("E11:E12")
    row_edges = [float(index * 10) for index in range(59)]
    column_edges = [float(index * 10) for index in range(6)]
    horizontal_mask = np.zeros((600, 80), dtype=np.uint8)
    horizontal_mask[30, 40:50] = 255
    rectified_fax = np.full((600, 80, 3), 255, dtype=np.uint8)

    _image, evidence = _draw_merge_aware_grid(
        worksheet=worksheet,
        rectified_fax=rectified_fax,
        xs=column_edges,
        ys=row_edges,
        horizontal_line_mask=horizontal_mask,
    )

    assert evidence["merge_region_count"] == 1
    assert evidence["retained_merge_region_count"] == 1
    assert evidence["physically_split_merge_region_count"] == 0
    assert evidence["physically_split_merge_regions"] == []
    assert evidence["merged_ranges"] == ["E11:E12"]


def test_step_review_target_regions_keep_template_merge_center() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet["D7"] = "献立"
    worksheet["D11"] = "先頭メニュー"
    worksheet["D12"] = "結合内メニュー"
    worksheet.merge_cells("E11:E12")
    row_edges = [float(index * 10) for index in range(59)]
    column_edges = [float(index * 10) for index in range(6)]
    horizontal_mask = np.zeros((600, 80), dtype=np.uint8)
    horizontal_mask[30, 40:50] = 255

    regions, evidence = _post_menu_target_regions(
        worksheet=worksheet,
        column_edges=column_edges,
        row_edges=row_edges,
        horizontal_line_mask=horizontal_mask,
    )

    by_id = {str(region["region_id"]): region for region in regions}
    assert "E11:E12" in by_id
    assert "E11" not in by_id
    assert by_id["E11:E12"]["bbox"] == [40.0, 20.0, 50.0, 40.0]
    assert by_id["E11:E12"]["merged_cell"]["range"] == "E11:E12"
    assert by_id["E11:E12"]["covered_sheet_cells"] == ["E11", "E12"]
    assert evidence["physical_split_excel_merge_count"] == 0
    assert evidence["physical_split_excel_merge_ranges"] == []


def test_step_review_target_regions_include_all_post_menu_columns() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet["D7"] = "献立"
    worksheet["E7"] = "常食"
    worksheet["F7"] = "糖尿"
    worksheet["G7"] = "備考欄"
    worksheet["H7"] = "肉禁"
    worksheet["I7"] = "魚禁"
    worksheet["D11"] = "先頭メニュー"
    row_edges = [float(index * 10) for index in range(59)]
    column_edges = [float(index * 10) for index in range(10)]
    fax_template = {
        "columns": [
            {"index": 0, "role": "date", "field": "date_mmdd", "header": "日付"},
            {"index": 1, "role": "daypart", "field": "daypart", "header": "区分"},
            {"index": 3, "role": "menu_name", "field": "menu", "header": "献立"},
            {
                "index": 4,
                "role": "quantity",
                "name": "qty.regular_x",
                "diet_type": "regular",
                "area_id": "X",
                "header": "常食",
            },
            {
                "index": 5,
                "role": "quantity",
                "name": "qty.diabetes_x",
                "diet_type": "diabetes",
                "area_id": "X",
                "header": "糖尿",
            },
            {"index": 6, "role": "note", "field": "remarks", "header": "備考欄"},
        ],
    }

    regions, evidence = _post_menu_target_regions(
        worksheet=worksheet,
        column_edges=column_edges,
        row_edges=row_edges,
        fax_template=fax_template,
    )

    first_row_regions = [region for region in regions if int(region["worksheet_row"]) == 11]
    assert [region["sheet_cell"] for region in first_row_regions] == ["E11", "F11", "G11", "H11", "I11"]
    assert [region["field"] for region in first_row_regions] == [
        "qty.regular_x",
        "qty.diabetes_x",
        "note",
        "qty.no_meat_x",
        "qty.no_fish_x",
    ]
    assert evidence["template_column_restricted"] is False
    assert evidence["target_selection_mode"] == "all_physical_columns_right_of_menu"
    assert evidence["target_worksheet_cols"] == [5, 6, 7, 8, 9]
    assert evidence["canonical_target_worksheet_cols"] == [5, 6, 7]


def test_step_review_target_regions_include_57th_body_row() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet["D7"] = "献立"
    for row in range(11, 68):
        worksheet.cell(row=row, column=4, value=f"メニュー{row}")
    for col, header in zip(("E", "F", "G", "H", "I", "J", "K"), ("常食", "糖尿", "備考欄", "肉禁", "魚禁", "軟菜", "袋分け")):
        worksheet[f"{col}7"] = header
    row_edges = [float(index * 10) for index in range(60)]
    column_edges = [float(index * 10) for index in range(12)]

    regions, evidence = _post_menu_target_regions(
        worksheet=worksheet,
        column_edges=column_edges,
        row_edges=row_edges,
    )

    target_cells = {str(region["sheet_cell"]) for region in regions}
    assert evidence["logical_target_count"] == 57 * 7
    assert all(f"{col}67" in target_cells for col in ("E", "F", "G", "H", "I", "J", "K"))


def test_step_review_target_regions_skip_blank_menu_rows() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet["D7"] = "献立"
    worksheet["E7"] = "常食"
    worksheet["D11"] = "先頭メニュー"
    worksheet["D12"] = ""
    worksheet["D13"] = "次メニュー"
    row_edges = [float(index * 10) for index in range(59)]
    column_edges = [float(index * 10) for index in range(6)]

    regions, evidence = _post_menu_target_regions(
        worksheet=worksheet,
        column_edges=column_edges,
        row_edges=row_edges,
    )

    worksheet_rows = {int(region["worksheet_row"]) for region in regions}
    assert 11 in worksheet_rows
    assert 12 not in worksheet_rows
    assert 13 in worksheet_rows
    assert evidence["blank_menu_row_count"] > 0


def test_detect_vertical_candidates_rejects_partial_height_correction_line() -> None:
    rectified = np.full((260, 360, 3), 255, dtype=np.uint8)
    for x in [40, 140, 240, 340]:
        rectified[30:230, x - 1 : x + 2] = 0
    rectified[95:165, 190 - 4 : 190 + 5] = 0

    candidates = _detect_vertical_candidates(rectified, [30, 30, 350, 230])

    rounded = [int(round(value)) for value in candidates]
    assert 190 not in rounded
    assert {40, 140, 240, 340}.issubset(set(rounded))


def test_post_menu_boundary_preserving_xs_keeps_first_quantity_line() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet["D7"] = "献立"
    worksheet["E7"] = "常食"
    worksheet["F7"] = "月"
    worksheet["G7"] = "軟菜"
    worksheet["H7"] = "月"
    worksheet["I7"] = "ミキサー"
    worksheet["J7"] = "月"
    worksheet["K7"] = "備考"
    matched_xs = [15.0, 102.0, 139.0, 212.0, 1071.0, 1251.0, 1435.0, 1617.0, 1800.0, 1982.0, 2167.0, 2347.0]
    fax_candidates = [15.0, 102.0, 139.0, 212.0, 887.0, 1071.0, 1087.0, 1105.0, 1251.0, 1435.0, 1617.0, 1800.0, 1982.0, 2167.0, 2347.0]

    adjusted, evidence = _post_menu_boundary_preserving_xs(
        worksheet=worksheet,
        matched_xs=matched_xs,
        fax_x_candidates=fax_candidates,
    )

    assert evidence["used"] is True
    assert adjusted[4] == 887.0
    assert adjusted[-1] == 2347.0
    assert 1087.0 not in adjusted
    assert 1105.0 not in adjusted
    assert 2167.0 not in adjusted
