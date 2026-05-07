import pathlib
import sys

from openpyxl import Workbook
from openpyxl.drawing.image import Image as WorkbookImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.styles import Border, Side
from openpyxl.utils.units import pixels_to_EMU
from PIL import Image

ROOT = pathlib.Path(__file__).resolve().parents[2]
sys.path.append(str(ROOT))

from src.services.workbook_pdf_renderer import (  # noqa: E402
    render_workbook_to_pdf_bytes,
    render_worksheet_to_image,
    worksheet_render_geometry,
)


def test_render_workbook_to_pdf_bytes_returns_pdf_payload() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"
    worksheet.print_area = "A1:D6"
    worksheet["A1"] = "施設名"
    worksheet["A2"] = "日付"
    worksheet["B2"] = "区分"
    worksheet["C2"] = "献立"
    worksheet["D2"] = "常食"
    worksheet["A3"] = "4/5"
    worksheet["B3"] = "朝"
    worksheet["C3"] = "サンプル"
    worksheet["D3"] = "12"

    pdf_bytes = render_workbook_to_pdf_bytes(workbook, dpi=144)

    assert pdf_bytes.startswith(b"%PDF")
    assert len(pdf_bytes) > 1000


def test_render_workbook_to_pdf_bytes_clips_merged_range_that_extends_past_print_area() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"
    worksheet.print_area = "A1:M5"
    worksheet.merge_cells("M1:P1")
    worksheet["M1"] = "右にはみ出す結合セル"
    worksheet["A2"] = "ok"

    pdf_bytes = render_workbook_to_pdf_bytes(workbook, dpi=144)

    assert pdf_bytes.startswith(b"%PDF")
    assert len(pdf_bytes) > 1000


def test_render_workbook_to_pdf_bytes_renders_merged_range_with_anchor_outside_print_area() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"
    worksheet.print_area = "B1:D4"
    worksheet.merge_cells("A1:C1")
    worksheet["A1"] = "左にはみ出す結合セル"
    worksheet["B2"] = "ok"

    pdf_bytes = render_workbook_to_pdf_bytes(workbook, dpi=144)

    assert pdf_bytes.startswith(b"%PDF")
    assert len(pdf_bytes) > 1000


def test_render_worksheet_to_image_draws_shared_border_once() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.print_area = "A1:B1"
    side = Side(style="medium", color="000000")
    worksheet["A1"].border = Border(right=side)
    worksheet["B1"].border = Border(left=side)

    image = render_worksheet_to_image(worksheet, dpi=96)
    y = image.height // 2
    black_columns = [x for x in range(image.width) if image.getpixel((x, y)) == (0, 0, 0)]

    assert len(black_columns) == 2


def test_render_worksheet_to_image_does_not_draw_internal_merged_borders() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.print_area = "A1:A2"
    side = Side(style="thin", color="000000")
    worksheet["A1"].border = Border(left=side, right=side, top=side, bottom=side)
    worksheet["A2"].border = Border(left=side, right=side, top=side, bottom=side)
    worksheet.merge_cells("A1:A2")

    geometry = worksheet_render_geometry(worksheet, dpi=96)
    image = render_worksheet_to_image(worksheet, dpi=96)
    y_internal = geometry["y_positions"][2]
    x0 = geometry["x_positions"][1] + 3
    x1 = geometry["x_positions"][2] - 3
    internal_black_pixels = [
        x for x in range(x0, x1) if image.getpixel((x, y_internal)) == (0, 0, 0)
    ]

    assert internal_black_pixels == []


def test_render_worksheet_to_image_uses_one_cell_anchor_extent(tmp_path) -> None:
    source = tmp_path / "red.png"
    Image.new("RGBA", (100, 50), (255, 0, 0, 255)).save(source)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.print_area = "A1:D4"
    workbook_image = WorkbookImage(str(source))
    workbook_image.anchor = OneCellAnchor(
        _from=AnchorMarker(col=1, row=1),
        ext=XDRPositiveSize2D(cx=pixels_to_EMU(20), cy=pixels_to_EMU(10)),
    )
    worksheet.add_image(workbook_image)

    image = render_worksheet_to_image(worksheet, dpi=96)

    red_pixels = [
        (x, y)
        for y in range(image.height)
        for x in range(image.width)
        if image.getpixel((x, y)) == (255, 0, 0)
    ]
    xs = [x for x, _y in red_pixels]
    ys = [y for _x, y in red_pixels]
    assert max(xs) - min(xs) + 1 == 20
    assert max(ys) - min(ys) + 1 == 10
