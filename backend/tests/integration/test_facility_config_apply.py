import sys
import pathlib
import re
from datetime import datetime

from openpyxl import Workbook
from sqlalchemy import delete

ROOT = pathlib.Path(__file__).resolve().parents[2]
sys.path.append(str(ROOT))

from src.services import facility_service, config_service, order_service, order_workflow_v2_service  # noqa: E402
from src.services import hakodate_assignment_service  # noqa: E402
from src.services import facility_template_version_service  # noqa: E402
from src.services import template_field_schema_service  # noqa: E402
from src.db import session_scope  # noqa: E402
from src.models.facility import Facility, FacilityArea, FacilityConfig  # noqa: E402
from src.models.facility_template_version import FacilityTemplateVersion  # noqa: E402
from src.workers.ingest_mail_adapter import IngestEmailPayload  # noqa: E402


def _clear_facilities():
    with session_scope() as session:
        session.execute(delete(FacilityConfig))
        session.execute(delete(FacilityArea))
        session.execute(delete(Facility))


def test_facility_config_applies_to_resolved_config():
    _clear_facilities()
    fac = facility_service.create_facility(
        "Beta Facility",
        [{"id": "ARE100", "name": "Unit X"}],
    )
    config = {
        "packaging_policy_override": {"split_key": ["facility", "date"]},
        "label_profile_override": {"storage_mode": "frozen"},
        "invoice_template": {
            "template_uri": "gs://example/invoice.xlsx",
            "columns": [{"name": "date", "source": "date"}],
        },
    }
    assert facility_service.update_config(fac["id"], config)
    resolved = config_service.get_facility_config(fac["id"])
    assert resolved is not None
    assert resolved["packaging_policy"]["split_key"] == ["facility", "date"]
    assert resolved["label_profile"]["storage_mode"] == "frozen"
    assert resolved["invoice_template"]["columns"][0]["name"] == "date"
    assert config_service.resolve_facility_id("Beta Facility") == fac["id"]


def test_fac00002_template_columns_preserve_area_schema_without_duplicates():
    config_service.reload_configs()
    resolved = config_service.get_facility_config("FAC00002")
    assert resolved is not None
    assert resolved.get("fax_template_id") == "シルバーホームなごみ"
    template = resolved.get("fax_template") or {}
    fields = template.get("main_ocr_row_fields") or []
    assert isinstance(fields, list)
    assert len(fields) == len(set(fields))
    assert "qty.regular_x" in fields
    assert "qty.no_meat_x" in fields
    assert "qty.no_fish_x" in fields
    assert "qty.soft_x" not in fields
    assert "qty.mixer_x" not in fields
    assert "qty.change_1_x" in fields
    assert "qty.change_2_x" in fields


def test_fac00010_uses_floor_2f3f_fax_template():
    _clear_facilities()
    facility_service.list_facilities()
    config_service.reload_configs()
    resolved = config_service.get_facility_config("FAC00010")
    assert resolved is not None
    assert resolved.get("fax_template_id") == "山城"
    fields = (resolved.get("fax_template") or {}).get("main_ocr_row_fields") or []
    assert fields == [
        "date_mmdd",
        "daypart",
        "menu",
        "qty.regular_2f",
        "qty.regular_3f",
        "qty.soft_2f",
        "qty.soft_3f",
        "qty.mixer_2f",
        "qty.mixer_3f",
        "remarks",
    ]
    columns = (resolved.get("fax_template") or {}).get("columns") or []
    assert [column.get("source_index") for column in columns] == [0, 1, 3, 4, 5, 6, 7, 8, 9, 10]


def test_fac00009_uses_floor_2f3f_fax_template_from_master():
    _clear_facilities()
    facility_service.list_facilities()
    config_service.reload_configs()
    resolved = config_service.get_facility_config("FAC00009")
    assert resolved is not None
    assert resolved.get("fax_template_id") == "グループホームそよかぜ"
    assert resolved.get("fax_template_ids") == ["グループホームそよかぜ"]
    assert (resolved.get("fax_template") or {}).get("main_ocr_row_fields") == [
        "date_mmdd",
        "daypart",
        "menu",
        "qty.regular_2f",
        "qty.regular_3f",
        "qty.soft_2f",
        "qty.soft_3f",
        "qty.mixer_2f",
        "qty.mixer_3f",
        "remarks",
    ]


def test_master_facility_template_replaces_stale_template_candidates_without_operator_source():
    _clear_facilities()
    facility_service.list_facilities()
    previous_config = facility_service.get_facility_config("FAC00009") or {}
    stale_config = {
        "fax_template_id": "fax_layout_regular_soft_mixer_forbidden_v1",
        "fax_template_ids": ["fax_layout_regular_soft_mixer_forbidden_v1"],
    }

    try:
        assert facility_service.update_config("FAC00009", stale_config)
        resolved = config_service.get_facility_config("FAC00009")
        assert resolved is not None
        assert resolved.get("fax_template_id") == "グループホームそよかぜ"
        assert resolved.get("fax_template_ids") == ["グループホームそよかぜ"]
    finally:
        assert facility_service.update_config("FAC00009", previous_config)


def test_master_facility_template_wins_over_stale_db_config_without_operator_source():
    _clear_facilities()
    facility_service.list_facilities()
    previous_config = facility_service.get_facility_config("FAC00016") or {}
    stale_config = {
        "fax_template_id": "fax_layout_regular_diabetes_v1",
        "fax_template_ids": ["fax_layout_regular_diabetes_v1"],
        "fax_template_override": {
            "columns_authoritative": True,
            "columns": [
                {"index": 0, "role": "date", "header": "日付", "format": "MM/DD"},
                {"index": 1, "role": "daypart", "header": "区分"},
                {"index": 2, "role": "menu_name", "header": "メニュー"},
                {"index": 3, "role": "quantity", "header": "常食", "diet_type": "regular", "area_id": "X"},
                {"index": 4, "role": "quantity", "header": "糖尿", "diet_type": "diabetes", "area_id": "X"},
                {"index": 5, "role": "note", "header": "備考欄"},
            ],
        },
    }

    try:
        assert facility_service.update_config("FAC00016", stale_config)
        resolved = config_service.get_facility_config("FAC00016")
        assert resolved is not None
        assert resolved.get("fax_template_id") == "いこいの森プラス"
        assert resolved.get("fax_template_ids") == ["いこいの森プラス"]
        assert (resolved.get("fax_template") or {}).get("main_ocr_row_fields") == [
            "date_mmdd",
            "daypart",
            "menu",
            "qty.regular_x",
            "qty.diabetes_x",
            "qty.no_meat_x",
            "qty.no_fish_x",
            "qty.change_1_x",
            "qty.change_2_x",
            "remarks",
        ]
    finally:
        assert facility_service.update_config("FAC00016", previous_config)


def test_operator_facility_template_source_is_ignored_in_favor_of_master():
    _clear_facilities()
    facility_service.list_facilities()
    previous_config = facility_service.get_facility_config("FAC00016") or {}
    operator_config = {
        "facility_template_source": "operator_override",
        "fax_template_id": "fax_layout_regular_diabetes_v1",
        "fax_template_override": {
            "columns_authoritative": True,
            "columns": [
                {"index": 0, "role": "date", "header": "日付", "format": "MM/DD"},
                {"index": 1, "role": "daypart", "header": "区分"},
                {"index": 2, "role": "menu_name", "header": "メニュー"},
                {"index": 3, "role": "quantity", "header": "常食", "diet_type": "regular", "area_id": "X"},
                {"index": 4, "role": "quantity", "header": "糖尿", "diet_type": "diabetes", "area_id": "X"},
                {"index": 5, "role": "note", "header": "備考欄"},
            ],
        },
    }

    try:
        assert facility_service.update_config(
            "FAC00016",
            operator_config,
            allow_authoritative_column_changes=True,
        )
        resolved = config_service.get_facility_config("FAC00016")
        assert resolved is not None
        assert (resolved.get("fax_template") or {}).get("main_ocr_row_fields") == [
            "date_mmdd",
            "daypart",
            "menu",
            "qty.regular_x",
            "qty.diabetes_x",
            "qty.no_meat_x",
            "qty.no_fish_x",
            "qty.change_1_x",
            "qty.change_2_x",
            "remarks",
        ]
    finally:
        assert facility_service.update_config("FAC00016", previous_config)


def test_hakodate_slots_use_template_fields_over_header_guessing():
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.cell(row=7, column=1, value="日付")
    worksheet.cell(row=7, column=2, value="区分")
    worksheet.cell(row=7, column=4, value="献立")
    worksheet.cell(row=7, column=5, value="常食")
    worksheet.cell(row=8, column=5, value="2F")
    worksheet.cell(row=7, column=6, value="常食")
    worksheet.cell(row=8, column=6, value="3F")
    worksheet.cell(row=7, column=7, value="軟菜")
    worksheet.cell(row=8, column=7, value="2F")
    worksheet.cell(row=7, column=8, value="軟菜")
    worksheet.cell(row=8, column=8, value="3F")
    worksheet.cell(row=7, column=9, value="ミキサー")
    worksheet.cell(row=8, column=9, value="2F")
    worksheet.cell(row=7, column=10, value="ミキサー")
    worksheet.cell(row=8, column=10, value="3F")
    worksheet.cell(row=7, column=11, value="備考")
    template = {
        "columns": [
            {"index": 0, "source_index": 0, "role": "date", "header": "日付"},
            {"index": 1, "source_index": 1, "role": "daypart", "header": "区分"},
            {"index": 2, "source_index": 3, "role": "menu_name", "header": "メニュー"},
            {"index": 3, "source_index": 4, "role": "quantity", "header": "常食2F", "diet_type": "regular", "area_id": "2F"},
            {"index": 4, "source_index": 5, "role": "quantity", "header": "常食3F", "diet_type": "regular", "area_id": "3F"},
            {"index": 5, "source_index": 6, "role": "quantity", "header": "軟菜2F", "diet_type": "soft", "area_id": "2F"},
            {"index": 6, "source_index": 7, "role": "quantity", "header": "軟菜3F", "diet_type": "soft", "area_id": "3F"},
            {"index": 7, "source_index": 8, "role": "quantity", "header": "ミキサー2F", "diet_type": "mixer", "area_id": "2F"},
            {"index": 8, "source_index": 9, "role": "quantity", "header": "ミキサー3F", "diet_type": "mixer", "area_id": "3F"},
            {"index": 9, "source_index": 10, "role": "note", "header": "備考"},
        ],
    }

    slots = hakodate_assignment_service._column_slots_from_worksheet(  # noqa: SLF001
        worksheet,
        col_count=11,
        template=template,
    )
    by_col = {slot["worksheet_col_index"]: slot for slot in slots}

    assert by_col[5]["slot_name"] == "qty.regular_2f"
    assert by_col[6]["slot_name"] == "qty.regular_3f"
    assert by_col[7]["slot_name"] == "qty.soft_2f"
    assert by_col[8]["slot_name"] == "qty.soft_3f"
    assert by_col[9]["slot_name"] == "qty.mixer_2f"
    assert by_col[10]["slot_name"] == "qty.mixer_3f"
    assert by_col[11]["slot_name"] == "note"


def test_hakodate_assignment_blocks_target_fields_not_in_facility_template():
    _clear_facilities()
    facility_service.list_facilities()
    config_service.reload_configs()
    payload = {
        "hakodate_preprocessing": {
            "target_cell_map": [
                {
                    "target_cell_id": "E11",
                    "sheet_cell": "E11",
                    "worksheet_row": 11,
                    "worksheet_col": 5,
                    "semantic_field": "qty.regular_x",
                    "bbox": [10, 10, 30, 30],
                    "center": [20, 20],
                    "metadata": {
                        "truth": {
                            "row_index": 0,
                            "field": "qty.regular_x",
                        }
                    },
                }
            ]
        },
        "hakodate_ocr_evidence_records": [
            {
                "evidence_id": "ev1",
                "raw_text": "1",
                "normalized_value": "1",
                "source_bbox": [12, 12, 18, 18],
                "center": [15, 15],
                "confidence": 0.99,
            }
        ],
    }

    assignment = order_service._build_hakodate_evidence_assignment_from_payload(  # noqa: SLF001
        order_id="ORD-test-field-unmapped",
        facility_id="FAC00010",
        template_id="fax_layout_floor_2f3f_v1",
        payload=payload,
    )

    assert "hakodate_target_field_unmapped" in assignment["blockers"]


def test_hakodate_assignment_accepts_legacy_unknown_spacer_target_field():
    _clear_facilities()
    facility_service.list_facilities()
    config_service.reload_configs()
    payload = {
        "hakodate_preprocessing": {
            "target_cell_map": [
                {
                    "target_cell_id": "E11",
                    "sheet_cell": "E11",
                    "worksheet_row": 11,
                    "worksheet_col": 5,
                    "semantic_field": "qty.unknown_x",
                    "bbox": [10, 10, 30, 30],
                    "center": [20, 20],
                    "metadata": {
                        "truth": {
                            "row_index": 0,
                            "field": "qty.unknown_x",
                        }
                    },
                }
            ]
        },
        "hakodate_ocr_evidence_records": [
            {
                "evidence_id": "ev1",
                "raw_text": "1",
                "normalized_value": "1",
                "source_bbox": [12, 12, 18, 18],
                "center": [15, 15],
                "confidence": 0.99,
            }
        ],
    }

    assignment = order_service._build_hakodate_evidence_assignment_from_payload(  # noqa: SLF001
        order_id="ORD-test-legacy-spacer",
        facility_id="FAC00001",
        template_id="template-FAC00001",
        payload=payload,
    )

    assert "hakodate_target_field_unmapped" not in assignment["blockers"]


def test_explicit_quantity_diet_type_db_override_is_not_used_as_facility_template():
    _clear_facilities()
    fac = facility_service.create_facility("Diet Override Facility", [])
    config = {
        "fax_template_override": {
            "columns": [
                {"index": 0, "role": "date", "header": "日付"},
                {"index": 1, "role": "daypart", "header": "区分"},
                {"index": 2, "role": "menu_name", "header": "メニュー"},
                {
                    "index": 3,
                    "role": "quantity",
                    "header": "糖尿",
                    "diet_type": "diabetes",
                    "area_id": "X",
                },
                {
                    "index": 4,
                    "role": "quantity",
                    "header": "ゴマアレルギー",
                    "diet_type": "sesame_allergy",
                    "area_id": "X",
                },
                {"index": 5, "role": "note", "header": "備考欄"},
            ]
        }
    }
    assert facility_service.update_config(fac["id"], config)

    resolved = config_service.get_facility_config(fac["id"])
    assert resolved is not None
    template = resolved.get("fax_template") or {}
    columns = template.get("columns") or []

    assert not any(column.get("header") == "糖尿" for column in columns)
    assert not any(column.get("diet_type") == "sesame_allergy" for column in columns)
    assert "fax_template_override" not in facility_service.get_facility_config(fac["id"])


def test_placeholder_and_custom_quantity_tokens_db_override_is_stripped():
    _clear_facilities()
    fac = facility_service.create_facility("Custom Quantity Facility", [])
    config = {
        "fax_template_override": {
            "columns": [
                {"index": 0, "role": "date", "header": "日付", "format": "MM/DD"},
                {"index": 1, "role": "daypart", "header": "区分"},
                {"index": 2, "role": "menu_name", "header": "メニュー"},
                {"index": 3, "role": "quantity", "header": "-", "diet_type": "placeholder", "area_id": "X"},
                {"index": 4, "role": "quantity", "header": "お茶", "diet_type": "tea", "area_id": "X"},
                {"index": 5, "role": "quantity", "header": "事業", "diet_type": "business", "area_id": "X"},
                {"index": 6, "role": "quantity", "header": "妊娠", "diet_type": "pregnancy", "area_id": "X"},
                {"index": 7, "role": "note", "header": "備考欄"},
            ]
        }
    }
    assert facility_service.update_config(fac["id"], config)

    resolved = config_service.get_facility_config(fac["id"])
    assert resolved is not None
    template = resolved.get("fax_template") or {}
    columns = template.get("columns") or []

    assert not any(column.get("diet_type") in {"placeholder", "tea", "business", "pregnancy"} for column in columns)
    assert "fax_template_override" not in facility_service.get_facility_config(fac["id"])


def test_fac00006_uses_repeated_regular_round_columns_from_source_master():
    config_service.reload_configs()
    resolved = config_service.get_facility_config("FAC00006")
    assert resolved is not None
    assert resolved.get("fax_template_id") == "藍TERRACE"
    assert resolved.get("fax_template_ids") == ["藍TERRACE"]
    template = resolved.get("fax_template") or {}
    assert template.get("main_ocr_row_fields") == [
        "date_mmdd",
        "daypart",
        "menu",
        "qty.regular_x",
        "qty.change_1_x",
        "qty.change_2_x",
        "qty.regular_bag_x",
        "qty.soft_x",
        "qty.mixer_x",
        "qty.no_meat_x",
        "qty.no_fish_x",
        "remarks",
    ]
    override = resolved.get("fax_template_override") or {}
    assert override.get("columns_authoritative") is True
    columns = template.get("columns") or []
    assert [column.get("header") for column in columns] == [
        "日付",
        "区分",
        "メニュー",
        "1回目",
        "2回目",
        "3回目",
        "袋分け",
        "軟菜",
        "ミキサー",
        "肉禁",
        "魚禁",
        "備考欄",
    ]
    assert [column.get("source_index") for column in columns] == [0, 1, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12]
    assert [column.get("name") for column in columns[3:11]] == [
        "qty.regular_x",
        "qty.change_1_x",
        "qty.change_2_x",
        "qty.regular_bag_x",
        "qty.soft_x",
        "qty.mixer_x",
        "qty.no_meat_x",
        "qty.no_fish_x",
    ]
def test_fac00012_preserves_placeholder_spacer_and_source_indexes_from_master():
    config_service.reload_configs()
    resolved = config_service.get_facility_config("FAC00012")
    assert resolved is not None
    assert resolved.get("fax_template_id") == "グランフォレスト方上"
    override = resolved.get("fax_template_override") or {}
    assert override.get("columns_authoritative") is True
    template = resolved.get("fax_template") or {}
    assert template.get("main_ocr_row_fields") == [
        "date_mmdd",
        "daypart",
        "menu",
        "qty.regular_x",
        "qty.placeholder_x",
        "qty.no_meat_x",
        "qty.no_fish_x",
        "qty.change_1_x",
        "qty.change_2_x",
        "remarks",
    ]
    columns = template.get("columns") or []
    assert [column.get("header") for column in columns] == [
        "日付",
        "区分",
        "メニュー",
        "常食",
        "-",
        "肉禁",
        "魚禁",
        "変更1",
        "変更2",
        "備考欄",
    ]
    assert [column.get("name") for column in columns[3:9]] == [
        "qty.regular_x",
        "qty.placeholder_x",
        "qty.no_meat_x",
        "qty.no_fish_x",
        "qty.change_1_x",
        "qty.change_2_x",
    ]
    assert [column.get("source_index") for column in columns] == [
        0,
        1,
        3,
        4,
        5,
        6,
        7,
        8,
        9,
        10,
    ]


def test_fac00005_exposes_authoritative_soft_bag_forbidden_change_columns():
    config_service.reload_configs()
    resolved = config_service.get_facility_config("FAC00005")
    assert resolved is not None
    override = resolved.get("fax_template_override") or {}
    assert override.get("columns_authoritative") is True
    template = resolved.get("fax_template") or {}
    assert template.get("main_ocr_row_fields") == [
        "date_mmdd",
        "daypart",
        "menu",
        "qty.soft_x",
        "qty.regular_bag_x",
        "qty.no_meat_x",
        "qty.no_fish_x",
        "qty.change_1_x",
        "qty.change_2_x",
        "remarks",
    ]
    columns = template.get("columns") or []
    assert [column.get("role") for column in columns] == [
        "date",
        "daypart",
        "menu_name",
        "quantity",
        "quantity",
        "quantity",
        "quantity",
        "quantity",
        "quantity",
        "note",
    ]
    assert [column.get("header") for column in columns[3:9]] == [
        "軟菜",
        "袋分け",
        "肉禁",
        "魚禁",
        "変更1",
        "変更2",
    ]


def test_fac00004_template_schema_contract_uses_current_columns():
    config_service.reload_configs()
    resolved = config_service.get_facility_config("FAC00004")
    assert resolved is not None
    template = resolved.get("fax_template") or {}

    contract = template_field_schema_service.build_template_field_schema_contract(template)

    assert contract["field_count"] == 11
    assert contract["fields"] == [
        "date_mmdd",
        "daypart",
        "menu",
        "qty.regular_x",
        "qty.daycare_x",
        "qty.staff_x",
        "qty.no_meat_x",
        "qty.no_fish_x",
        "qty.no_fried_x",
        "qty.change_1_x",
        "remarks",
    ]
    assert contract["aux_fields"] == []


def test_template_schema_uses_locked_quantity_name_as_canonical_field():
    columns = [
        {"index": 0, "role": "date"},
        {"index": 1, "role": "daypart"},
        {"index": 2, "role": "menu_name"},
        {
            "index": 3,
            "role": "quantity",
            "name": "qty.no_fish_soft_2f",
            "name_locked": True,
            "diet_type": "no_fish",
            "area_id": "2F",
        },
        {
            "index": 4,
            "role": "quantity",
            "name": "qty.no_fish_mixer_2f",
            "name_locked": True,
            "diet_type": "no_fish",
            "area_id": "2F",
        },
    ]

    assert template_field_schema_service.derive_row_fields_from_columns(columns) == [
        "date_mmdd",
        "daypart",
        "menu",
        "qty.no_fish_soft_2f",
        "qty.no_fish_mixer_2f",
    ]


def test_fac00005_update_config_preserves_authoritative_master_schema():
    _clear_facilities()
    facility_service.list_facilities()

    assert facility_service.update_config("FAC00005", {"menu_override_tags": ["larger", "keep"]})

    resolved = config_service.get_facility_config("FAC00005")
    assert resolved is not None
    override = resolved.get("fax_template_override") or {}
    assert override.get("columns_authoritative") is True
    template = resolved.get("fax_template") or {}
    assert template.get("main_ocr_row_fields") == [
        "date_mmdd",
        "daypart",
        "menu",
        "qty.soft_x",
        "qty.regular_bag_x",
        "qty.no_meat_x",
        "qty.no_fish_x",
        "qty.change_1_x",
        "qty.change_2_x",
        "remarks",
    ]
    columns = template.get("columns") or []
    assert [column.get("header") for column in columns[3:9]] == [
        "軟菜",
        "袋分け",
        "肉禁",
        "魚禁",
        "変更1",
        "変更2",
    ]


def test_fac00007_uses_regular_forbidden_plus_change_columns():
    config_service.reload_configs()
    resolved = config_service.get_facility_config("FAC00007")
    assert resolved is not None
    assert resolved.get("fax_template_id") == "ゆうゆう（株）百々家"
    template = resolved.get("fax_template") or {}
    assert template.get("main_ocr_row_fields") == [
        "date_mmdd",
        "daypart",
        "menu",
        "qty.regular_x",
        "qty.placeholder_x",
        "qty.no_meat_x",
        "qty.no_fish_x",
        "qty.change_1_x",
        "qty.change_2_x",
        "remarks",
    ]


def test_fac00004_exposes_daycare_staff_and_no_fried_columns():
    config_service.reload_configs()
    resolved = config_service.get_facility_config("FAC00004")
    assert resolved is not None
    assert resolved.get("fax_template_id") == "介護老人保健施設ふれあいの丘"
    template = resolved.get("fax_template") or {}
    assert template.get("main_ocr_row_fields") == [
        "date_mmdd",
        "daypart",
        "menu",
        "qty.regular_x",
        "qty.daycare_x",
        "qty.staff_x",
        "qty.no_meat_x",
        "qty.no_fish_x",
        "qty.no_fried_x",
        "qty.change_1_x",
        "remarks",
    ]
    columns = template.get("columns") or []
    assert [columns[idx]["role"] for idx in range(5)] == ["date", "daypart", "menu_name", "quantity", "quantity"]
    assert columns[3]["diet_type"] == "regular"


def test_columns_authoritative_db_override_does_not_replace_master_fields():
    _clear_facilities()
    fac = facility_service.create_facility("Columns Authoritative Aux Facility", [])
    config = {
        "fax_template_override": {
            "columns_authoritative": True,
            "columns": [
                {"index": 0, "role": "date", "header": "日付"},
                {"index": 1, "role": "daypart", "header": "区分"},
                {"index": 2, "role": "aux", "header": "副区分"},
                {"index": 3, "role": "menu_name", "header": "メニュー"},
                {"index": 4, "role": "aux", "header": "合計"},
                {"index": 5, "role": "quantity", "header": "常食", "diet_type": "regular", "area_id": "X"},
                {"index": 6, "role": "note", "header": "備考欄"},
            ],
            "main_ocr_row_fields": [
                "date_mmdd",
                "daypart",
                "menu",
                "qty.regular_x",
                "remarks",
            ],
        }
    }
    assert facility_service.update_config(fac["id"], config)

    resolved = config_service.get_facility_config(fac["id"])
    assert resolved is not None
    template = resolved.get("fax_template") or {}
    assert template.get("main_ocr_row_fields") == [
        "date_mmdd",
        "daypart",
        "menu",
        "qty.regular_2f",
        "qty.regular_3f",
        "qty.soft_2f",
        "qty.soft_3f",
        "qty.mixer_2f",
        "qty.mixer_3f",
        "remarks",
    ]
    assert "fax_template_override" not in facility_service.get_facility_config(fac["id"])


def test_fac00014_exposes_staff_sesame_and_change_columns():
    _clear_facilities()
    facility_service.list_facilities()
    config_service.reload_configs()
    resolved = config_service.get_facility_config("FAC00014")
    assert resolved is not None
    template = resolved.get("fax_template") or {}
    assert template.get("main_ocr_row_fields") == [
        "date_mmdd",
        "daypart",
        "menu",
        "qty.regular_x",
        "qty.staff_x",
        "qty.no_meat_x",
        "qty.no_fish_x",
        "qty.sesame_allergy_x",
        "qty.change_1_x",
        "remarks",
    ]
    columns = template.get("columns") or []
    assert [columns[idx]["role"] for idx in range(3)] == ["date", "daypart", "menu_name"]
    assert columns[3]["diet_type"] == "regular"
    assert columns[4]["diet_type"] == "staff"
    assert columns[5]["diet_type"] == "no_meat"
    assert columns[6]["diet_type"] == "no_fish"
    assert columns[7]["diet_type"] == "sesame_allergy"
    assert columns[8]["diet_type"] == "change_1"
    assert columns[9]["role"] == "note"


def test_fac00014_prefers_master_override_over_stale_persisted_columns():
    _clear_facilities()
    facility_service.list_facilities()
    stale_config = {
        "fax_template_override": {
            "columns": [
                {"index": 0, "role": "date", "header": "日付", "format": "MM/DD"},
                {"index": 1, "role": "daypart", "header": "区分"},
                {"index": 2, "role": "menu_name", "header": "メニュー"},
                {"index": 3, "role": "quantity", "header": "常食", "diet_type": "regular", "area_id": "X"},
                {"index": 4, "role": "quantity", "header": "職員", "diet_type": "staff", "area_id": "X"},
                {"index": 5, "role": "quantity", "header": "肉禁", "diet_type": "no_meat", "area_id": "X"},
                {"index": 6, "role": "quantity", "header": "魚禁", "diet_type": "no_fish", "area_id": "X"},
                {"index": 7, "role": "quantity", "header": "ゴマアレルギー", "diet_type": "sesame_allergy", "area_id": "X"},
                {"index": 8, "role": "quantity", "header": "変更1", "diet_type": "change_1", "area_id": "X"},
                {"index": 9, "role": "note", "header": "備考欄"},
            ]
        }
    }
    assert facility_service.update_config("FAC00014", stale_config)

    resolved = config_service.get_facility_config("FAC00014")
    assert resolved is not None
    columns = (resolved.get("fax_template") or {}).get("columns") or []
    assert [columns[idx]["role"] for idx in range(3)] == ["date", "daypart", "menu_name"]
    assert columns[3]["diet_type"] == "regular"
    assert columns[4]["diet_type"] == "staff"
    assert columns[5]["diet_type"] == "no_meat"
    assert columns[6]["diet_type"] == "no_fish"
    assert columns[7]["diet_type"] == "sesame_allergy"
    assert columns[8]["diet_type"] == "change_1"
    assert columns[9]["role"] == "note"


def test_repo_master_overrides_stale_db_override_authoritative_columns():
    _clear_facilities()
    facility_service.list_facilities()
    stale_fac00004_config = {
        "facility_template_source": "db_override",
        "fax_template_id": "春日苑系",
        "fax_template_ids": ["春日苑系"],
        "fax_template_override": {
            "columns_authoritative": True,
            "columns": [
                {"index": 0, "role": "date", "header": "日付", "format": "MM/DD"},
                {"index": 1, "role": "daypart", "header": "区分"},
                {"index": 2, "role": "aux", "header": "副区分", "name": "sub_daypart"},
                {"index": 3, "role": "menu_name", "header": "メニュー"},
                {"index": 4, "role": "aux", "header": "合計", "name": "raw_total"},
                {"index": 5, "role": "quantity", "header": "常食", "diet_type": "regular", "area_id": "X"},
                {"index": 6, "role": "quantity", "header": "通所", "diet_type": "commuter", "area_id": "X"},
                {"index": 7, "role": "quantity", "header": "職員", "diet_type": "staff", "area_id": "X"},
                {"index": 8, "role": "quantity", "header": "肉禁", "diet_type": "no_meat", "area_id": "X"},
                {"index": 9, "role": "quantity", "header": "魚禁", "diet_type": "no_fish", "area_id": "X"},
                {"index": 10, "role": "aux", "header": "その他", "name": "raw_other"},
                {"index": 11, "role": "quantity", "header": "変更1", "diet_type": "change_1", "area_id": "X"},
                {"index": 12, "role": "note", "header": "備考欄"},
            ],
        },
    }
    stale_fac00013_config = {
        "facility_template_source": "db_override",
        "fax_template_id": "いこいの森",
        "fax_template_ids": ["いこいの森"],
        "fax_template_override": {
            "columns_authoritative": True,
            "columns": [
                {"index": 0, "role": "date", "header": "日付", "format": "MM/DD"},
                {"index": 1, "role": "daypart", "header": "区分"},
                {"index": 2, "role": "menu_name", "header": "メニュー"},
                {"index": 3, "role": "quantity", "header": "常食", "diet_type": "regular", "area_id": "X"},
                {"index": 4, "role": "quantity", "header": "肉禁", "diet_type": "no_meat", "area_id": "X"},
                {"index": 5, "role": "quantity", "header": "魚禁", "diet_type": "no_fish", "area_id": "X"},
                {"index": 6, "role": "quantity", "header": "変更1", "diet_type": "change_1", "area_id": "X"},
                {"index": 7, "role": "quantity", "header": "変更2", "diet_type": "change_2", "area_id": "X"},
                {"index": 8, "role": "note", "header": "備考欄"},
            ],
        },
    }
    assert facility_service.update_config("FAC00004", stale_fac00004_config)
    assert facility_service.update_config("FAC00013", stale_fac00013_config)

    fac00004 = config_service.get_facility_config("FAC00004")
    fac00013 = config_service.get_facility_config("FAC00013")

    fac00004_headers = [
        str(column.get("header") or "")
        for column in ((fac00004 or {}).get("fax_template") or {}).get("columns") or []
    ]
    fac00013_columns = ((fac00013 or {}).get("fax_template") or {}).get("columns") or []
    fac00013_headers = [str(column.get("header") or "") for column in fac00013_columns]
    assert "副区分" not in fac00004_headers
    assert "合計" not in fac00004_headers
    assert fac00004_headers == [
        "日付",
        "区分",
        "メニュー",
        "常食",
        "通所",
        "職員",
        "肉禁",
        "魚禁",
        "揚げ物禁",
        "変更1",
        "備考欄",
    ]
    assert fac00013_headers[3:5] == ["常食", "糖尿"]
    assert fac00013_columns[4]["diet_type"] == "diabetes"


def test_confirm_context_refreshes_stale_active_template_version_from_repo_master():
    _clear_facilities()
    facility_service.list_facilities()
    order_service.clear_all()
    stale_columns = [
        {"index": 0, "role": "date", "header": "日付", "format": "MM/DD"},
        {"index": 1, "role": "daypart", "header": "区分"},
        {"index": 2, "role": "menu_name", "header": "メニュー"},
        {"index": 3, "role": "quantity", "header": "常食", "diet_type": "regular", "area_id": "X"},
        {"index": 4, "role": "quantity", "header": "肉禁", "diet_type": "no_meat", "area_id": "X"},
        {"index": 5, "role": "quantity", "header": "魚禁", "diet_type": "no_fish", "area_id": "X"},
        {"index": 6, "role": "note", "header": "備考"},
    ]
    stale_digest = facility_template_version_service.template_digest(
        template_id="fax_layout_regular_forbidden_v1",
        columns=facility_template_version_service.normalize_template_columns(stale_columns),
    )
    with session_scope() as session:
        if session.get(Facility, "FAC00013") is None:
            session.add(Facility(id="FAC00013", name="いこいの森"))
        session.add(
            FacilityTemplateVersion(
                id="FTV_STALE_FAC00013",
                facility_id="FAC00013",
                version="1",
                status="active",
                template_id="fax_layout_regular_forbidden_v1",
                source="test-stale-active-version",
                columns_json=stale_columns,
                cells_json=[],
                template_digest=stale_digest,
                validation_json={"errors": [], "warnings": []},
                created_by="test",
                created_at=datetime(2026, 5, 5, 0, 0, 0),
                activated_at=datetime(2026, 5, 5, 0, 0, 0),
            )
        )
    order = order_service.create_order_from_ingest(
        IngestEmailPayload(
            message_id="msg-fac00013-stale-active-template",
            pdf_uri="file://fac00013-stale-active-template.pdf",
            received_at=datetime(2026, 7, 5, 12, 0, 0),
            facility_hint="FAC00013",
            week_hint="2026-07",
        ),
        lines=[],
    )

    workflow, error = order_workflow_v2_service.confirm_context(
        order_id=order["id"],
        facility_id="FAC00013",
        week_start="2026-07-05",
        week_end="2026-07-11",
        template_id="いこいの森",
    )

    assert error is None
    assert workflow is not None
    assert workflow["template_id"] == "いこいの森"
    assert workflow["template_version_id"] != "FTV_STALE_FAC00013"
    with session_scope() as session:
        active = facility_template_version_service.get_active_template_version(session, "FAC00013")
        assert active is not None
        assert active.id == workflow["template_version_id"]
        assert active.template_id == "いこいの森"
        headers = [column.get("header") for column in active.columns_json or []]
        assert headers[3:5] == ["常食", "糖尿"]


def test_fac00014_update_config_strips_stale_override_before_storage():
    _clear_facilities()
    facility_service.list_facilities()
    stale_config = {
        "fax_template_override": {
            "columns": [
                {"index": 0, "role": "date", "header": "日付", "format": "MM/DD"},
                {"index": 1, "role": "daypart", "header": "区分"},
                {"index": 2, "role": "menu_name", "header": "メニュー"},
                {"index": 3, "role": "quantity", "header": "常食", "diet_type": "regular", "area_id": "X"},
                {"index": 4, "role": "quantity", "header": "職員", "diet_type": "staff", "area_id": "X"},
                {"index": 5, "role": "quantity", "header": "肉禁", "diet_type": "no_meat", "area_id": "X"},
                {"index": 6, "role": "quantity", "header": "魚禁", "diet_type": "no_fish", "area_id": "X"},
                {"index": 7, "role": "quantity", "header": "ゴマアレルギー", "diet_type": "sesame_allergy", "area_id": "X"},
                {"index": 8, "role": "quantity", "header": "変更1", "diet_type": "change_1", "area_id": "X"},
                {"index": 9, "role": "note", "header": "備考欄"},
            ]
        }
    }
    assert facility_service.update_config("FAC00014", stale_config)

    stored = facility_service.get_facility_config("FAC00014")
    assert stored is not None
    assert "fax_template_override" not in stored
    resolved = config_service.get_facility_config("FAC00014")
    assert resolved is not None
    template = resolved.get("fax_template") or {}
    columns = template.get("columns") or []
    assert [columns[idx]["role"] for idx in range(3)] == ["date", "daypart", "menu_name"]
    assert columns[3]["diet_type"] == "regular"
    assert columns[4]["diet_type"] == "staff"
    assert columns[5]["diet_type"] == "no_meat"
    assert columns[6]["diet_type"] == "no_fish"
    assert columns[7]["diet_type"] == "sesame_allergy"
    assert columns[8]["diet_type"] == "change_1"
    assert template.get("main_ocr_row_fields") == [
        "date_mmdd",
        "daypart",
        "menu",
        "qty.regular_x",
        "qty.staff_x",
        "qty.no_meat_x",
        "qty.no_fish_x",
        "qty.sesame_allergy_x",
        "qty.change_1_x",
        "remarks",
    ]


def test_save_order_facility_template_columns_rejects_legacy_column_removal():
    _clear_facilities()
    facility_service.list_facilities()
    order_service.clear_all()

    order = order_service.create_order_from_ingest(
        IngestEmailPayload(
            message_id="msg-fac00014-delete-column",
            pdf_uri="file://fac00014-delete-column.pdf",
            received_at=datetime(2026, 4, 12, 12, 0, 0),
            facility_hint="FAC00014",
            week_hint="2026-04",
        ),
        lines=[],
    )
    assert order is not None

    resolved = config_service.get_facility_config("FAC00014")
    assert resolved is not None
    original_columns = ((resolved.get("fax_template") or {}).get("columns") or [])
    columns_without_change_1 = [
        dict(column)
        for column in original_columns
        if str(column.get("diet_type") or "").strip() != "change_1"
    ]

    result, error = order_service.save_order_facility_template_columns(
        order["id"],
        columns_without_change_1,
    )

    assert result is None
    assert error == "legacy_facility_template_columns_removed"


def test_save_order_facility_template_columns_derives_mapping_from_visible_labels_only():
    _clear_facilities()
    facility_service.list_facilities()
    order_service.clear_all()
    previous_config = facility_service.get_facility_config("FAC00010") or {}

    order = order_service.create_order_from_ingest(
        IngestEmailPayload(
            message_id="msg-facility-template-visible-labels-only",
            pdf_uri="file://facility-template-visible-labels-only.pdf",
            received_at=datetime(2026, 4, 12, 12, 0, 0),
            facility_hint="FAC00010",
            week_hint="2026-04",
        ),
        lines=[],
    )
    assert order is not None

    try:
        result, error = order_service.save_order_facility_template_columns(
            order["id"],
            [
                {"index": 0, "role": "date", "header": "日付", "name": "date"},
                {"index": 1, "role": "daypart", "header": "区分", "name": "daypart"},
                {"index": 2, "role": "menu_name", "header": "メニュー", "name": "menu"},
                {
                    "index": 3,
                    "role": "quantity",
                    "header": "常食2F",
                    "name": "qty.no_fish_3f",
                    "diet_type": "no_fish",
                    "area_id": "3F",
                    "diet_type_locked": True,
                    "area_id_locked": True,
                    "name_locked": True,
                },
                {
                    "index": 4,
                    "role": "note",
                    "header": "不明",
                    "name": "qty.unknown_x",
                    "name_locked": True,
                },
                {"index": 5, "role": "note", "header": "備考"},
            ],
        )

        assert result is None
        assert error == "legacy_facility_template_columns_removed"
    finally:
        assert facility_service.update_config("FAC00010", previous_config)


def test_generic_update_config_preserves_master_authoritative_columns_for_fac00014():
    _clear_facilities()
    facility_service.list_facilities()

    attempted_override = {
        "fax_template_override": {
            "grid_line_scale_horizontal": 18,
            "columns": [
                {"index": 0, "role": "date", "header": "日付", "format": "MM/DD"},
                {"index": 1, "role": "daypart", "header": "区分"},
                {"index": 2, "role": "menu_name", "header": "メニュー"},
                {"index": 3, "role": "quantity", "header": "常食", "diet_type": "regular", "area_id": "X"},
                {"index": 4, "role": "quantity", "header": "肉卵魚禁", "diet_type": "forbidden_other", "area_id": "X"},
                {"index": 5, "role": "note", "header": "備考欄"},
            ],
        }
    }

    assert facility_service.update_config("FAC00014", attempted_override)

    stored = facility_service.get_facility_config("FAC00014")
    assert stored is not None
    assert "fax_template_override" not in stored
    resolved = config_service.get_facility_config("FAC00014")
    assert resolved is not None
    resolved_columns = ((resolved.get("fax_template") or {}).get("columns")) or []
    assert [column.get("diet_type") for column in resolved_columns if column.get("role") == "quantity"] == [
        "regular",
        "staff",
        "no_meat",
        "no_fish",
        "sesame_allergy",
        "change_1",
    ]
    assert not any(str(column.get("diet_type") or "").strip() == "forbidden_other" for column in resolved_columns)


def test_generic_update_config_keeps_order_authored_authoritative_columns():
    _clear_facilities()
    facility_service.list_facilities()
    order_service.clear_all()

    order = order_service.create_order_from_ingest(
        IngestEmailPayload(
            message_id="msg-fac00014-preserve-authored-columns",
            pdf_uri="file://fac00014-preserve-authored-columns.pdf",
            received_at=datetime(2026, 4, 12, 12, 0, 0),
            facility_hint="FAC00014",
            week_hint="2026-04",
        ),
        lines=[],
    )
    assert order is not None

    resolved = config_service.get_facility_config("FAC00014")
    assert resolved is not None
    original_columns = ((resolved.get("fax_template") or {}).get("columns") or [])
    columns_without_change_1 = [
        dict(column)
        for column in original_columns
        if str(column.get("diet_type") or "").strip() != "change_1"
    ]

    result, error = order_service.save_order_facility_template_columns(
        order["id"],
        columns_without_change_1,
    )

    assert result is None
    assert error == "legacy_facility_template_columns_removed"
    assert facility_service.update_config("FAC00014", {"menu_override_tags": ["keep-authored-columns"]})

    stored = facility_service.get_facility_config("FAC00014")
    assert stored is not None
    assert stored.get("menu_override_tags") == ["keep-authored-columns"]
    refreshed = config_service.get_facility_config("FAC00014")
    assert refreshed is not None
    refreshed_columns = ((refreshed.get("fax_template") or {}).get("columns") or [])
    assert any(str(column.get("diet_type") or "").strip() == "change_1" for column in refreshed_columns)


def test_reconcile_fax_override_keeps_current_when_quantity_families_differ():
    current_override = {
        "columns": [
            {"index": 0, "role": "date", "header": "日付", "format": "MM/DD"},
            {"index": 1, "role": "daypart", "header": "区分"},
            {"index": 2, "role": "menu_name", "header": "メニュー"},
            {"index": 3, "role": "quantity", "header": "常食", "diet_type": "regular", "area_id": "X"},
            {"index": 4, "role": "quantity", "header": "職員", "diet_type": "staff", "area_id": "X"},
            {"index": 5, "role": "quantity", "header": "肉禁", "diet_type": "no_meat", "area_id": "X"},
            {"index": 6, "role": "quantity", "header": "魚禁", "diet_type": "no_fish", "area_id": "X"},
            {"index": 7, "role": "quantity", "header": "ゴマアレルギー", "diet_type": "sesame_allergy", "area_id": "X"},
            {"index": 8, "role": "quantity", "header": "変更1", "diet_type": "change_1", "area_id": "X"},
            {"index": 9, "role": "note", "header": "備考欄"},
        ]
    }
    master_override = {
        "columns": [
            {"index": 0, "role": "date", "header": "日付", "format": "MM/DD"},
            {"index": 1, "role": "daypart", "header": "区分"},
            {"index": 2, "role": "aux", "header": "副区分"},
            {"index": 3, "role": "menu_name", "header": "メニュー"},
            {"index": 4, "role": "quantity", "header": "常食", "diet_type": "regular", "area_id": "X"},
            {"index": 5, "role": "quantity", "header": "職員", "diet_type": "staff", "area_id": "X"},
            {"index": 6, "role": "quantity", "header": "肉禁", "diet_type": "no_meat", "area_id": "X"},
            {"index": 7, "role": "quantity", "header": "魚禁", "diet_type": "no_fish", "area_id": "X"},
            {"index": 8, "role": "quantity", "header": "ゴマアレルギー", "diet_type": "sesame_allergy", "area_id": "X"},
            {"index": 9, "role": "note", "header": "備考欄"},
        ]
    }

    reconciled = config_service._reconcile_fax_template_override(  # type: ignore[attr-defined]
        current_override,
        master_override,
        drop_redundant=False,
    )

    assert reconciled is not None
    columns = reconciled.get("columns") or []
    assert [column.get("role") for column in columns[:3]] == ["date", "daypart", "menu_name"]
    assert [column.get("diet_type") for column in columns if column.get("role") == "quantity"] == [
        "regular",
        "staff",
        "no_meat",
        "no_fish",
        "sesame_allergy",
        "change_1",
    ]


def test_fac00016_non_authoritative_stale_subset_reconciles_to_master_columns():
    config_service.reload_configs()
    previous_config = facility_service.get_facility_config("FAC00016") or {}
    stale_override = {
        "columns": [
            {"index": 0, "role": "date", "header": "日付", "format": "MM/DD"},
            {"index": 1, "role": "daypart", "header": "区分"},
            {"index": 2, "role": "menu_name", "header": "メニュー"},
            {"index": 3, "role": "quantity", "header": "常食", "diet_type": "regular", "area_id": "X"},
            {"index": 4, "role": "quantity", "header": "糖尿", "diet_type": "diabetes", "area_id": "X"},
            {"index": 5, "role": "note", "header": "備考欄"},
        ]
    }

    try:
        assert facility_service.update_config("FAC00016", {"fax_template_override": stale_override})

        resolved = config_service.get_facility_config("FAC00016")
        assert resolved is not None
        template = resolved.get("fax_template") or {}
        columns = template.get("columns") or []
        quantity_diets = [column.get("diet_type") for column in columns if column.get("role") == "quantity"]

        assert quantity_diets == ["regular", "diabetes", "no_meat", "no_fish", "change_1", "change_2"]
        assert template.get("main_ocr_row_fields") == [
            "date_mmdd",
            "daypart",
            "menu",
            "qty.regular_x",
            "qty.diabetes_x",
            "qty.no_meat_x",
            "qty.no_fish_x",
            "qty.change_1_x",
            "qty.change_2_x",
            "remarks",
        ]
    finally:
        assert facility_service.update_config("FAC00016", previous_config)


def test_authoritative_facility_override_is_stripped_from_generic_facility_config():
    _clear_facilities()
    fac = facility_service.create_facility("Authoritative Override Facility", [])
    authored_config = {
        "fax_template_override": {
            "columns_authoritative": True,
            "columns": [
                {"index": 0, "role": "date", "header": "日付", "format": "MM/DD"},
                {"index": 1, "role": "daypart", "header": "区分"},
                {"index": 2, "role": "menu_name", "header": "メニュー"},
                {"index": 3, "role": "quantity", "header": "常食特別", "diet_type": "regular", "area_id": "X"},
                {"index": 4, "role": "quantity", "header": "職員特別", "diet_type": "staff", "area_id": "X"},
                {"index": 5, "role": "note", "header": "備考欄"},
            ],
        }
    }

    assert facility_service.update_config(fac["id"], authored_config)

    stored = facility_service.get_facility_config(fac["id"])
    assert stored is not None
    assert "fax_template_override" not in stored

    resolved = config_service.get_facility_config(fac["id"])
    assert resolved is not None
    template = resolved.get("fax_template") or {}
    assert "qty.staff_x" not in (template.get("main_ocr_row_fields") or [])


def test_reconcile_fax_template_override_prefers_authoritative_master_placeholder_columns():
    current_override = {
        "columns": [
            {"index": 0, "role": "date", "header": "日付", "format": "MM/DD"},
            {"index": 1, "role": "daypart", "header": "区分"},
            {"index": 2, "role": "menu_name", "header": "メニュー"},
            {"index": 3, "role": "quantity", "header": "常食", "diet_type": "regular", "area_id": "X"},
            {"index": 4, "role": "quantity", "header": "肉禁", "diet_type": "no_meat", "area_id": "X"},
            {"index": 5, "role": "quantity", "header": "魚禁", "diet_type": "no_fish", "area_id": "X"},
            {"index": 6, "role": "quantity", "header": "変更1", "diet_type": "change_1", "area_id": "X"},
            {"index": 7, "role": "quantity", "header": "変更2", "diet_type": "change_2", "area_id": "X"},
            {"index": 8, "role": "note", "header": "備考欄"},
        ]
    }
    master_override = {
        "columns_authoritative": True,
        "columns": [
            {"index": 0, "role": "date", "header": "日付", "format": "MM/DD"},
            {"index": 1, "role": "daypart", "header": "区分"},
            {"index": 2, "role": "menu_name", "header": "メニュー", "source_index": 3},
            {"index": 3, "role": "quantity", "header": "常食", "diet_type": "regular", "area_id": "X", "source_index": 4},
            {"index": 4, "role": "quantity", "header": "-", "diet_type": "placeholder", "area_id": "X", "source_index": 5},
            {"index": 5, "role": "quantity", "header": "肉禁", "diet_type": "no_meat", "area_id": "X", "source_index": 6},
            {"index": 6, "role": "quantity", "header": "魚禁", "diet_type": "no_fish", "area_id": "X", "source_index": 7},
            {"index": 7, "role": "quantity", "header": "変更1", "diet_type": "change_1", "area_id": "X", "source_index": 8},
            {"index": 8, "role": "quantity", "header": "変更2", "diet_type": "change_2", "area_id": "X", "source_index": 9},
            {"index": 9, "role": "note", "header": "備考欄", "source_index": 10},
        ],
    }

    reconciled = config_service._reconcile_fax_template_override(  # type: ignore[attr-defined]
        current_override,
        master_override,
        drop_redundant=False,
    )

    assert reconciled is not None
    assert reconciled.get("columns_authoritative") is True
    assert reconciled.get("main_ocr_row_fields") == [
        "date_mmdd",
        "daypart",
        "menu",
        "qty.regular_x",
        "qty.placeholder_x",
        "qty.no_meat_x",
        "qty.no_fish_x",
        "qty.change_1_x",
        "qty.change_2_x",
        "remarks",
    ]
    columns = reconciled.get("columns") or []
    assert [column.get("header") for column in columns[3:9]] == [
        "常食",
        "-",
        "肉禁",
        "魚禁",
        "変更1",
        "変更2",
    ]
    assert [column.get("source_index") for column in columns] == [
        None,
        None,
        3,
        4,
        5,
        6,
        7,
        8,
        9,
        10,
    ]


def test_fac00004_update_config_strips_template_override_and_uses_master_columns():
    _clear_facilities()
    facility_service.list_facilities()
    stale_config = {
        "fax_template_override": {
            "columns": [
                {"index": 0, "role": "date", "header": "日付", "format": "MM/DD"},
                {"index": 1, "role": "daypart", "header": "区分"},
                {"index": 2, "role": "menu_name", "header": "メニュー"},
                {"index": 3, "role": "quantity", "header": "常食", "diet_type": "regular", "area_id": "X"},
                {"index": 4, "role": "quantity", "header": "通所", "diet_type": "daycare", "area_id": "X"},
                {"index": 5, "role": "quantity", "header": "職員", "diet_type": "staff", "area_id": "X"},
                {"index": 6, "role": "quantity", "header": "肉禁", "diet_type": "no_meat", "area_id": "X"},
                {"index": 7, "role": "quantity", "header": "魚禁", "diet_type": "no_fish", "area_id": "X"},
                {"index": 8, "role": "quantity", "header": "揚げ物禁", "diet_type": "no_fried", "area_id": "X"},
                {"index": 9, "role": "quantity", "header": "変更1", "diet_type": "change_1", "area_id": "X"},
                {"index": 10, "role": "note", "header": "備考欄"},
            ]
        }
    }
    assert facility_service.update_config("FAC00004", stale_config)

    stored = facility_service.get_facility_config("FAC00004")
    assert stored is not None
    assert "fax_template_override" not in stored
    resolved = config_service.get_facility_config("FAC00004")
    assert resolved is not None
    columns = ((resolved.get("fax_template") or {}).get("columns")) or []
    assert [columns[idx]["role"] for idx in range(5)] == ["date", "daypart", "menu_name", "quantity", "quantity"]
    assert columns[3]["diet_type"] == "regular"
    assert columns[4]["diet_type"] == "daycare"
    assert columns[5]["diet_type"] == "staff"
    assert columns[6]["diet_type"] == "no_meat"
    assert columns[7]["diet_type"] == "no_fish"
    assert columns[8]["diet_type"] == "no_fried"
    assert columns[9]["diet_type"] == "change_1"
    assert (resolved.get("fax_template") or {}).get("main_ocr_row_fields") == [
        "date_mmdd",
        "daypart",
        "menu",
        "qty.regular_x",
        "qty.daycare_x",
        "qty.staff_x",
        "qty.no_meat_x",
        "qty.no_fish_x",
        "qty.no_fried_x",
        "qty.change_1_x",
        "remarks",
    ]


def test_facilities_with_legacy_base_columns_have_explicit_layout_templates():
    config_service.reload_configs()
    fac1 = config_service.get_facility_config("FAC00001")
    assert fac1 is not None
    assert fac1.get("fax_template_id") == "大和なでしこ"
    assert (fac1.get("fax_template") or {}).get("main_ocr_row_fields") == [
        "date_mmdd",
        "daypart",
        "menu",
        "qty.regular_x",
        "qty.placeholder_x",
        "qty.no_meat_x",
        "qty.no_fish_x",
        "qty.change_1_x",
        "qty.change_2_x",
        "remarks",
    ]

    fac8 = config_service.get_facility_config("FAC00008")
    assert fac8 is not None
    assert fac8.get("fax_template_id") == "佐古"
    assert (fac8.get("fax_template") or {}).get("main_ocr_row_fields") == [
        "date_mmdd",
        "daypart",
        "menu",
        "qty.regular_2f",
        "qty.regular_3f",
        "qty.soft_2f",
        "qty.soft_3f",
        "qty.mixer_2f",
        "qty.mixer_3f",
        "remarks",
    ]

    fac9 = config_service.get_facility_config("FAC00009")
    assert fac9 is not None
    assert fac9.get("fax_template_id") == "グループホームそよかぜ"

    resolved = config_service.get_facility_config("FAC00010")
    assert resolved is not None
    assert resolved.get("fax_template_id") == "山城"
    template = resolved.get("fax_template") or {}
    assert (resolved.get("fax_template_override") or {}).get("grid_line_scale_horizontal") == 20
    assert (resolved.get("fax_template_override") or {}).get("columns_authoritative") is True
    assert template.get("main_ocr_row_fields") == [
        "date_mmdd",
        "daypart",
        "menu",
        "qty.regular_2f",
        "qty.regular_3f",
        "qty.soft_2f",
        "qty.soft_3f",
        "qty.mixer_2f",
        "qty.mixer_3f",
        "remarks",
    ]

    fac11 = config_service.get_facility_config("FAC00011")
    assert fac11 is not None
    assert fac11.get("fax_template_id") == "ケアハウス四万十"


def test_all_master_facilities_have_explicit_layout_template_ids():
    config_service.reload_configs()
    missing = []
    for facility in config_service.load_facility_master().get("facilities", []):
        facility_id = facility.get("facility_id")
        resolved = config_service.get_facility_config(facility_id)
        if not resolved or not resolved.get("fax_template_id"):
            missing.append(facility_id)
    assert missing == []


def test_fac00014_15_16_expose_custom_quantity_columns():
    _clear_facilities()
    facility_service.list_facilities()
    config_service.reload_configs()

    fac14 = config_service.get_facility_config("FAC00014")
    assert fac14 is not None
    assert fac14.get("fax_template_id") == "湘南さくら病院"
    assert (fac14.get("fax_template") or {}).get("main_ocr_row_fields") == [
        "date_mmdd",
        "daypart",
        "menu",
        "qty.regular_x",
        "qty.staff_x",
        "qty.no_meat_x",
        "qty.no_fish_x",
        "qty.sesame_allergy_x",
        "qty.change_1_x",
        "remarks",
    ]
    fac14_columns = (fac14.get("fax_template") or {}).get("columns") or []
    assert [column.get("role") for column in fac14_columns] == [
        "date",
        "daypart",
        "menu_name",
        "quantity",
        "quantity",
        "quantity",
        "quantity",
        "quantity",
        "quantity",
        "note",
    ]
    assert fac14_columns[2]["header"] == "メニュー"
    assert fac14_columns[3]["diet_type"] == "regular"
    assert fac14_columns[4]["diet_type"] == "staff"
    assert fac14_columns[5]["diet_type"] == "no_meat"
    assert fac14_columns[6]["diet_type"] == "no_fish"
    assert fac14_columns[7]["diet_type"] == "sesame_allergy"
    assert fac14_columns[8]["diet_type"] == "change_1"

    fac15 = config_service.get_facility_config("FAC00015")
    assert fac15 is not None
    assert fac15.get("fax_template_id") == "ケアハウス四万十ピア"
    assert (fac15.get("fax_template") or {}).get("main_ocr_row_fields") == [
        "date_mmdd",
        "daypart",
        "menu",
        "qty.regular_x",
        "qty.placeholder_x",
        "qty.no_meat_x",
        "qty.no_fish_x",
        "qty.change_1_x",
        "qty.change_2_x",
        "remarks",
    ]

    fac16 = config_service.get_facility_config("FAC00016")
    assert fac16 is not None
    assert fac16.get("fax_template_id") == "いこいの森プラス"
    assert (fac16.get("fax_template") or {}).get("main_ocr_row_fields") == [
        "date_mmdd",
        "daypart",
        "menu",
        "qty.regular_x",
        "qty.diabetes_x",
        "qty.no_meat_x",
        "qty.no_fish_x",
        "qty.change_1_x",
        "qty.change_2_x",
        "remarks",
    ]


def test_fac00003_and_fac00013_use_explicit_layout_templates():
    config_service.reload_configs()

    fac00003 = config_service.get_facility_config("FAC00003")
    assert fac00003 is not None
    assert fac00003.get("fax_template_id") == "春日苑 松茂"
    assert fac00003.get("fax_template_ids") == ["春日苑 松茂"]

    fac00013 = config_service.get_facility_config("FAC00013")
    assert fac00013 is not None
    assert fac00013.get("fax_template_id") == "いこいの森"
    assert fac00013.get("fax_template_ids") == ["いこいの森"]
    assert (fac00013.get("fax_template") or {}).get("main_ocr_row_fields") == [
        "date_mmdd",
        "daypart",
        "menu",
        "qty.regular_x",
        "qty.diabetes_x",
        "qty.no_meat_x",
        "qty.no_fish_x",
        "remarks",
    ]

    fac00014 = config_service.get_facility_config("FAC00014")
    assert fac00014 is not None
    assert fac00014.get("fax_template_id") == "湘南さくら病院"
    assert fac00014.get("fax_template_ids") == ["湘南さくら病院"]

    fac00016 = config_service.get_facility_config("FAC00016")
    assert fac00016 is not None
    assert fac00016.get("fax_template_id") == "いこいの森プラス"
    assert fac00016.get("fax_template_ids") == ["いこいの森プラス"]
    assert (fac00016.get("fax_template") or {}).get("main_ocr_row_fields") == [
        "date_mmdd",
        "daypart",
        "menu",
        "qty.regular_x",
        "qty.diabetes_x",
        "qty.no_meat_x",
        "qty.no_fish_x",
        "qty.change_1_x",
        "qty.change_2_x",
        "remarks",
    ]


def test_fac00003_fish_forbidden_columns_keep_meal_shape_fields():
    config_service.reload_configs()

    fac00003 = config_service.get_facility_config("FAC00003")
    assert fac00003 is not None
    template = fac00003.get("fax_template") or {}
    fields = template.get("main_ocr_row_fields") or []

    assert fields == [
        "date_mmdd",
        "daypart",
        "menu",
        "qty.regular_2f",
        "qty.regular_3f",
        "qty.soft_2f",
        "qty.soft_3f",
        "qty.mixer_2f",
        "qty.mixer_3f",
        "qty.no_fish_regular_2f",
        "qty.no_fish_regular_3f",
        "qty.no_fish_soft_2f",
        "qty.no_fish_soft_3f",
        "qty.no_fish_mixer_2f",
        "qty.no_fish_mixer_3f",
        "remarks",
    ]

    columns = template.get("columns") or []
    by_field = {col.get("name"): col for col in columns if isinstance(col, dict)}
    assert by_field["qty.no_fish_regular_2f"]["source_index"] == 10
    assert by_field["qty.no_fish_regular_3f"]["source_index"] == 11
    assert by_field["qty.no_fish_soft_2f"]["source_index"] == 12
    assert by_field["qty.no_fish_soft_3f"]["source_index"] == 13
    assert by_field["qty.no_fish_mixer_2f"]["source_index"] == 14
    assert by_field["qty.no_fish_mixer_3f"]["source_index"] == 15
    assert by_field["qty.no_fish_regular_2f"]["header_super_group"] == "禁食、魚"
    assert by_field["qty.no_fish_regular_2f"]["header_group"] == "常食"
    assert by_field["qty.no_fish_soft_2f"]["header_super_group"] == "禁食、魚"
    assert by_field["qty.no_fish_soft_2f"]["header_group"] == "軟菜"
    assert by_field["qty.no_fish_mixer_2f"]["header_super_group"] == "禁食、魚"
    assert by_field["qty.no_fish_mixer_2f"]["header_group"] == "ミキサー"
    assert [col for col in columns if col.get("role") == "note"][0]["source_index"] == 16


def test_facility_config_normalizes_hana_tsuki_columns_to_floor_fields():
    original = facility_service.get_facility_config("FAC00003") or {}
    next_config = dict(original)
    override = dict(next_config.get("fax_template_override") or {})
    override["columns"] = [
        {"index": 0, "role": "date", "header": "日付"},
        {"index": 1, "role": "daypart", "header": "区分"},
        {"index": 2, "role": "menu_name", "header": "メニュー"},
        {"index": 3, "role": "quantity", "header": "常食花", "diet_type": "regular", "area_id": "花"},
        {"index": 4, "role": "quantity", "header": "常食月", "diet_type": "regular", "area_id": "月"},
        {"index": 5, "role": "quantity", "header": "軟菜花", "diet_type": "soft", "area_id": "花"},
        {"index": 6, "role": "quantity", "header": "軟菜月", "diet_type": "soft", "area_id": "月"},
        {"index": 7, "role": "quantity", "header": "ミキサー花", "diet_type": "mixer", "area_id": "花"},
        {"index": 8, "role": "quantity", "header": "ミキサー月", "diet_type": "mixer", "area_id": "月"},
        {"index": 9, "role": "note", "header": "備考"},
    ]
    override.pop("main_ocr_row_fields", None)
    next_config["fax_template_override"] = override

    try:
        assert facility_service.update_config("FAC00003", next_config)
        resolved = config_service.get_facility_config("FAC00003")
        assert resolved is not None
        columns = (resolved.get("fax_template") or {}).get("columns") or []
        qty_columns = [col for col in columns if isinstance(col, dict) and str(col.get("role") or "") == "quantity"]
        assert [col.get("area_id") for col in qty_columns[:6]] == ["2F", "3F", "2F", "3F", "2F", "3F"]
        assert [col.get("name") for col in qty_columns[:6]] == [
            "qty.regular_2f",
            "qty.regular_3f",
            "qty.soft_2f",
            "qty.soft_3f",
            "qty.mixer_2f",
            "qty.mixer_3f",
        ]
        assert (resolved.get("fax_template") or {}).get("main_ocr_row_fields") == [
            "date_mmdd",
            "daypart",
            "menu",
            "qty.regular_2f",
            "qty.regular_3f",
            "qty.soft_2f",
            "qty.soft_3f",
            "qty.mixer_2f",
            "qty.mixer_3f",
            "qty.no_fish_regular_2f",
            "qty.no_fish_regular_3f",
            "qty.no_fish_soft_2f",
            "qty.no_fish_soft_3f",
            "qty.no_fish_mixer_2f",
            "qty.no_fish_mixer_3f",
            "remarks",
        ]
    finally:
        facility_service.update_config("FAC00003", original)
