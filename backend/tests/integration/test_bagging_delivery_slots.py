import pathlib
import sys
import csv
from datetime import date, datetime
from uuid import uuid4

from openpyxl import Workbook, load_workbook
from sqlalchemy import delete

ROOT = pathlib.Path(__file__).resolve().parents[2]
sys.path.append(str(ROOT))

from src.db import session_scope
from src.models.menu import MonthlyMenu, MonthlyMenuItem
from src.models.output import Bag
from src.services import config_service, facility_service, menu_service, order_service, output_builder
from src.services.output_builder import build_outputs, rebuild_bags
from src.workers.ingest_mail_adapter import IngestEmailPayload


def _reset_orders_and_menus() -> None:
    config_service.reload_configs()
    order_service.clear_all()
    with session_scope() as session:
        session.execute(delete(MonthlyMenuItem))
        session.execute(delete(MonthlyMenu))


def test_condiment_bags_are_merged_across_diet_and_area():
    _reset_orders_and_menus()

    payload = IngestEmailPayload(
        message_id="bag-condiment-001",
        pdf_uri="file:///tmp/bag-condiment.pdf",
        received_at=datetime(2025, 1, 5, 12, 0, 0),
        facility_hint="FAC00001",
        week_hint="2025-01",
    )
    lines = [
        {
            "line_id": "1",
            "date": "2025-01-06",
            "daypart": "朝",
            "menu_name": "ソース",
            "diet_type": "regular",
            "area_id": "2F",
            "bag_type": "condiment",
            "quantity_original": 3,
        },
        {
            "line_id": "2",
            "date": "2025-01-06",
            "daypart": "朝",
            "menu_name": "ソース",
            "diet_type": "soft",
            "area_id": "3F",
            "bag_type": "condiment",
            "quantity_original": 4,
        },
    ]
    order = order_service.create_order_from_ingest(payload, lines=lines)

    rebuilt = rebuild_bags(order["id"])
    condiments = [bag for bag in rebuilt["bags"] if bag.get("bag_type") == "condiment"]

    assert len(condiments) == 1
    assert condiments[0]["menu_name"] == "ソース"
    assert condiments[0]["quantity"] == 7.0
    assert condiments[0]["diet_type"] is None
    assert condiments[0]["area_id"] is None


def test_get_bag_summary_rebuilds_stale_materialized_bags():
    _reset_orders_and_menus()

    payload = IngestEmailPayload(
        message_id="bag-stale-001",
        pdf_uri="file:///tmp/bag-stale.pdf",
        received_at=datetime(2025, 2, 18, 12, 0, 0),
        facility_hint="FAC00003",
        week_hint="2025-02",
    )
    lines = [
        {
            "line_id": "1",
            "date": "2025-02-18",
            "daypart": "昼",
            "menu_name": "筑前煮",
            "diet_type": "regular",
            "area_id": "2F",
            "bag_type": "standard",
            "quantity_original": 9,
        },
        {
            "line_id": "2",
            "date": "2025-02-18",
            "daypart": "昼",
            "menu_name": "筑前煮",
            "diet_type": "regular",
            "area_id": "3F",
            "bag_type": "standard",
            "quantity_original": 8,
        },
        {
            "line_id": "3",
            "date": "2025-02-18",
            "daypart": "昼",
            "menu_name": "筑前煮",
            "diet_type": "mixer",
            "area_id": "2F",
            "bag_type": "standard",
            "quantity_original": 2,
        },
        {
            "line_id": "4",
            "date": "2025-02-18",
            "daypart": "昼",
            "menu_name": "筑前煮",
            "diet_type": "mixer",
            "area_id": "3F",
            "bag_type": "standard",
            "quantity_original": 3,
        },
    ]
    order = order_service.create_order_from_ingest(payload, lines=lines)

    with session_scope() as session:
        session.add(
            Bag(
                id=f"BAG{uuid4().hex[:8]}",
                order_id=order["id"],
                date=datetime(2025, 2, 18).date(),
                daypart="昼",
                menu_name="筑前煮",
                diet_type="regular",
                area_id=None,
                bag_type="large",
                quantity=15,
            )
        )
        session.add(
            Bag(
                id=f"BAG{uuid4().hex[:8]}",
                order_id=order["id"],
                date=datetime(2025, 2, 18).date(),
                daypart="昼",
                menu_name="筑前煮",
                diet_type="regular",
                area_id=None,
                bag_type="large",
                quantity=15,
            )
        )
        session.add(
            Bag(
                id=f"BAG{uuid4().hex[:8]}",
                order_id=order["id"],
                date=datetime(2025, 2, 18).date(),
                daypart="昼",
                menu_name="筑前煮",
                diet_type="regular",
                area_id=None,
                bag_type="large",
                quantity=9,
            )
        )

    summary, error = order_service.get_bag_summary(order["id"])

    assert error is None
    assert summary is not None
    chikuzen = [row for row in summary["bags"] if row.get("menu_name") == "筑前煮"]
    assert {(row.get("diet_type"), row.get("area_id"), row.get("quantity")) for row in chikuzen} == {
        ("mixer", "2F", 2.0),
        ("mixer", "3F", 3.0),
        ("regular", "2F", 9.0),
        ("regular", "3F", 8.0),
    }


def test_delivery_slot_template_writes_menu_display_and_clears_unassigned(tmp_path):
    _reset_orders_and_menus()

    week_id = "2025-02"
    menu_am = menu_service.create_item_stub(week_id, "メニュー朝")
    menu_service.update_item(week_id, menu_am["id"], {"category": "主A"})
    menu_pm = menu_service.create_item_stub(week_id, "メニュー夕")
    menu_service.update_item(week_id, menu_pm["id"], {"category": "主A"})

    template_path = tmp_path / "delivery_slot_template.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Template"
    ws.cell(row=1, column=1, value="日付")
    ws.cell(row=1, column=2, value="区分")
    ws.cell(row=1, column=3, value="献立")
    ws.cell(row=1, column=4, value="常食2F")

    ws.cell(row=2, column=2, value="朝")
    ws.cell(row=2, column=3, value="主A")
    ws.cell(row=2, column=4, value=999)

    ws.cell(row=3, column=3, value="副①")
    ws.cell(row=3, column=4, value=999)

    ws.cell(row=4, column=2, value="夕")
    ws.cell(row=4, column=3, value="主A")
    ws.cell(row=4, column=4, value=999)

    wb.save(template_path)

    slot_area_id = f"ARE_SLOT_{uuid4().hex[:8]}"
    facility = facility_service.create_facility(
        "Slot Writer Facility",
        [{"id": slot_area_id, "name": "2F"}],
    )
    updated = facility_service.update_config(
        facility["id"],
        {
            "invoice_template": {
                "template_uri": template_path.as_uri(),
                "sheet_name": "Template",
                "include_menu_name": False,
                "columns": [
                    {"name": "日付", "source": "date", "column_index": 1},
                    {"name": "区分", "source": "daypart", "column_index": 2},
                    {"name": "献立", "source": "menu_display", "column_index": 3},
                    {
                        "name": "常食2F",
                        "source": "quantity",
                        "diet_type": "regular",
                        "area_id": "2F",
                        "column_index": 4,
                    },
                ],
            }
        },
    )
    assert updated

    payload = IngestEmailPayload(
        message_id="delivery-slot-001",
        pdf_uri="file:///tmp/delivery-slot.pdf",
        received_at=datetime(2025, 2, 20, 9, 0, 0),
        facility_hint=facility["id"],
        week_hint=week_id,
    )
    lines = [
        {
            "line_id": "1",
            "date": "2025-02-20",
            "daypart": "朝",
            "menu_name": "メニュー朝",
            "diet_type": "regular",
            "area_id": "2F",
            "bag_type": "standard",
            "quantity_original": 5,
        },
        {
            "line_id": "2",
            "date": "2025-02-20",
            "daypart": "夕",
            "menu_name": "メニュー夕",
            "diet_type": "regular",
            "area_id": "2F",
            "bag_type": "standard",
            "quantity_original": 7,
        },
    ]
    order = order_service.create_order_from_ingest(payload, lines=lines)

    outputs = build_outputs(order["id"])
    delivery_wb = load_workbook(outputs["delivery_note"])
    ws_out = delivery_wb["2025-02-20"]

    assert ws_out.cell(row=2, column=3).value == "主A メニュー朝"
    assert ws_out.cell(row=2, column=4).value == 5
    assert ws_out.cell(row=4, column=3).value == "主A メニュー夕"
    assert ws_out.cell(row=4, column=4).value == 7
    assert ws_out.cell(row=3, column=3).value in ("", None)
    assert ws_out.cell(row=3, column=4).value in ("", None)


def test_delivery_slot_template_applies_configured_headers(tmp_path):
    template_path = tmp_path / "delivery_slot_headers.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Template"
    ws.cell(row=1, column=1, value="日付")
    ws.cell(row=1, column=2, value="区分")
    ws.cell(row=1, column=3, value="献立")
    ws.cell(row=1, column=4, value="常食2F")
    ws.cell(row=1, column=5, value="常食3F")
    ws.cell(row=1, column=6, value="軟菜")
    ws.cell(row=2, column=5, value="2F")
    ws.cell(row=2, column=6, value="3F")
    ws.cell(row=3, column=2, value="朝")
    ws.cell(row=3, column=3, value="主菜")
    wb.save(template_path)

    output_path = tmp_path / "delivery_out.xlsx"
    output_builder._write_delivery_note(
        output_path,
        [
            {
                "date": date(2026, 5, 24),
                "daypart": "朝",
                "menu_display": "主菜 Menu A",
                "常食": 48,
            }
        ],
        [
            {"name": "日付", "source": "date", "header": "日付", "column_index": 1},
            {"name": "区分", "source": "daypart", "header": "区分", "column_index": 2},
            {"name": "献立", "source": "menu_display", "header": "献立", "column_index": 3},
            {"name": "常食", "source": "quantity", "header": "常食", "column_index": 4},
            {"name": "肉禁", "source": "quantity", "header": "肉禁", "column_index": 5},
            {"name": "魚禁", "source": "quantity", "header": "魚禁", "column_index": 6},
        ],
        template_path.as_uri(),
        False,
        "Template",
        "Facility",
    )

    actual = load_workbook(output_path)
    ws_out = actual["2026-05-24"]
    assert ws_out.cell(row=2, column=4).value == "常食"
    assert ws_out.cell(row=2, column=5).value == "肉禁"
    assert ws_out.cell(row=2, column=6).value == "魚禁"


def test_ikebukuro_label_expiry_is_plus_three_months_and_content_amount_fallbacks():
    _reset_orders_and_menus()

    payload = IngestEmailPayload(
        message_id="label-ikebukuro-001",
        pdf_uri="file:///tmp/label-ikebukuro.pdf",
        received_at=datetime(2025, 1, 5, 12, 0, 0),
        facility_hint="FAC00005",
        week_hint="2025-01",
    )
    lines = [
        {
            "line_id": "1",
            "date": "2025-01-06",
            "daypart": "朝",
            "menu_name": "テストメニュー",
            "diet_type": "regular",
            "area_id": "2F",
            "bag_type": "standard",
            "quantity_original": 2,
        }
    ]
    order = order_service.create_order_from_ingest(payload, lines=lines)
    outputs = build_outputs(order["id"])

    with pathlib.Path(outputs["labels"]).open("r", newline="", encoding="cp932") as handle:
        rows = list(csv.DictReader(handle))

    assert len(rows) == 1
    assert rows[0]["賞味期限"] == "2025年4月6日"
    assert rows[0]["内容量"] == "2人前"
    assert rows[0]["実量"] in ("", None)
