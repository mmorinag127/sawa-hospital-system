import pathlib
import sys
from datetime import date as dt_date

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import MergedCell

ROOT = pathlib.Path(__file__).resolve().parents[2]
sys.path.append(str(ROOT))

from src.services import order_service, output_builder  # noqa: E402


TARGET_DATE = dt_date(2026, 3, 22)


def _sawa_root() -> pathlib.Path:
    for parent in pathlib.Path(__file__).resolve().parents:
        candidate = parent / "input_example"
        if candidate.exists():
            return parent
    return ROOT.parent.parent


def _border_signature(border) -> tuple:
    return tuple(
        (
            side.style,
            side.color.type if side.color else None,
            side.color.rgb if side.color and side.color.type == "rgb" else None,
        )
        for side in (border.left, border.right, border.top, border.bottom)
    )


def _cell_excel_signature(cell) -> tuple:
    fill = cell.fill
    font = cell.font
    alignment = cell.alignment
    return (
        cell.value,
        cell.data_type,
        cell.number_format,
        _border_signature(cell.border),
        fill.fill_type,
        fill.fgColor.type,
        fill.fgColor.rgb if fill.fgColor.type == "rgb" else fill.fgColor.indexed,
        font.name,
        font.sz,
        font.bold,
        font.italic,
        font.color.type if font.color else None,
        font.color.rgb if font.color and font.color.type == "rgb" else None,
        alignment.horizontal,
        alignment.vertical,
        alignment.wrap_text,
        alignment.shrink_to_fit,
    )


def _dotted_table_border_cells(ws) -> list[str]:
    dotted_cells = []
    max_col = output_builder._daily_delivery_table_max_column(ws)  # noqa: SLF001
    for row_idx in range(12, min(ws.max_row, 19) + 1):
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if isinstance(cell, MergedCell):
                continue
            border = cell.border
            if any(side.style == "dotted" for side in (border.left, border.right, border.top, border.bottom)):
                dotted_cells.append(cell.coordinate)
    return dotted_cells


def _dimension_signature(ws) -> tuple:
    return (
        tuple((key, item.width, item.hidden) for key, item in sorted(ws.column_dimensions.items())),
        tuple((key, item.height, item.hidden) for key, item in sorted(ws.row_dimensions.items())),
    )


def _page_signature(ws) -> tuple:
    def _rounded(value):
        return round(value, 10) if isinstance(value, float) else value

    return (
        ws.sheet_view.showGridLines,
        ws.freeze_panes,
        ws.page_setup.orientation,
        ws.page_setup.paperSize,
        ws.page_setup.scale,
        ws.page_setup.fitToWidth,
        ws.page_setup.fitToHeight,
        _rounded(ws.page_margins.left),
        _rounded(ws.page_margins.right),
        _rounded(ws.page_margins.top),
        _rounded(ws.page_margins.bottom),
        _rounded(ws.page_margins.header),
        _rounded(ws.page_margins.footer),
        ws.print_options.horizontalCentered,
        ws.print_options.verticalCentered,
        tuple(ws.print_area),
    )


def _make_context(order_id: str, facility_code: str, facility_name: str, menu_name: str) -> dict:
    return {
        "bags": [{"date": TARGET_DATE, "menu_name": menu_name}],
        "label_profile": {},
        "facility_config": {"facility_name": facility_name},
        "order_for_outputs": {"id": order_id, "facility": facility_code, "lines": []},
        "delivery_source_for_outputs": {"id": order_id, "facility": facility_code, "lines": []},
        "invoice_template": {
            "columns": [
                {"name": "日付", "source": "date"},
                {"name": "区分", "source": "daypart"},
                {"name": "メニュー", "source": "menu_name"},
                {"name": "常食", "source": "quantity"},
            ],
            "template_uri": None,
            "include_menu_name": True,
            "sheet_name": None,
        },
        "quantity_rules": {"zero_as_empty": True},
    }


def test_write_delivery_note_blocks_when_template_uri_missing(tmp_path):
    with pytest.raises(ValueError, match="delivery_template_uri_required"):
        output_builder._write_delivery_note(
            tmp_path / "delivery.xlsx",
            [{"date": TARGET_DATE, "menu_name": "献立A"}],
            [{"name": "メニュー", "source": "menu_name"}],
            None,
            True,
        )


def test_build_outputs_download_path_does_not_write_canonical_rows(tmp_path, monkeypatch):
    monkeypatch.setattr(output_builder, "OUTPUT_DIR", tmp_path)
    context = _make_context("ORD-1", "FAC001", "そよかぜ", "献立A")
    context["order"] = {"id": "ORD-1", "facility": "FAC001", "week": "2026-03"}
    context["order_lines"] = [
        {
            "date": TARGET_DATE,
            "daypart": "朝",
            "menu_name": "献立A",
            "diet_type": "regular",
            "area_id": "common",
            "quantity_original": 2,
        }
    ]
    context["order_for_outputs"] = {**context["order"], "lines": context["order_lines"]}
    context["bags"] = [
        {
            "date": TARGET_DATE,
            "daypart": "朝",
            "menu_name": "献立A",
            "diet_type": "regular",
            "area_id": "common",
            "bag_type": "standard",
            "quantity": 2,
        }
    ]
    context["delivery_source_for_outputs"] = {**context["order"], "lines": context["bags"]}
    monkeypatch.setattr(output_builder, "_prepare_output_context", lambda order_id: context)

    def _blocked_session_scope():
        raise AssertionError("download output path must not mutate canonical output rows")

    def _fake_write_delivery_note(path, rows, columns, template_uri, include_menu_name, sheet_name=None, facility_name=None):
        workbook = Workbook()
        ws = workbook.active
        ws["A1"] = "delivery"
        workbook.save(path)

    monkeypatch.setattr(output_builder, "session_scope", _blocked_session_scope)
    monkeypatch.setattr(output_builder.menu_service, "resolve_menu_defaults", lambda *_args, **_kwargs: {})
    monkeypatch.setattr(output_builder, "_write_delivery_note", _fake_write_delivery_note)

    outputs = output_builder.build_outputs("ORD-1")

    assert pathlib.Path(outputs["labels"]).exists()
    assert pathlib.Path(outputs["delivery_note"]).exists()
    assert pathlib.Path(outputs["aggregate"]).exists()


def test_prepare_output_context_derives_invoice_columns_when_template_has_no_quantity_overlap(monkeypatch):
    monkeypatch.setattr(
        output_builder,
        "get_order_by_id",
        lambda _order_id: {
            "id": "ORD-FUREAI",
            "facility": "FAC00004",
            "week": "2026-05",
            "lines": [
                {
                    "date": TARGET_DATE,
                    "daypart": "昼",
                    "menu_name": "献立A",
                    "diet_type": "regular",
                    "area_id": "X",
                    "quantity_original": 5,
                }
            ],
        },
    )
    monkeypatch.setattr(
        output_builder.config_service,
        "get_facility_config",
        lambda _facility_id: {
            "facility_name": "介護老人保健施設ふれあいの丘",
            "fax_template": {
                "columns": [
                    {"role": "quantity", "header": "常食", "diet_type": "regular", "area_id": "X"},
                    {"role": "quantity", "header": "通所", "diet_type": "daycare", "area_id": "X"},
                    {"role": "quantity", "header": "職員", "diet_type": "staff", "area_id": "X"},
                    {"role": "quantity", "header": "魚禁", "diet_type": "no_fish", "area_id": "X"},
                ]
            },
            "invoice_template": {
                "columns": [
                    {"name": "日付", "source": "date"},
                    {"name": "常食2F", "source": "quantity", "diet_type": "regular", "area_id": "2F"},
                    {"name": "ミキサー2F", "source": "quantity", "diet_type": "mixer", "area_id": "2F"},
                ],
                "include_menu_name": False,
            },
        },
    )
    monkeypatch.setattr(output_builder.config_service, "load_ingest_policy", lambda: {"quantity_rules": {"zero_as_empty": True}})
    monkeypatch.setattr(output_builder, "_build_ocr_menu_meta", lambda *_args, **_kwargs: {})

    ctx = output_builder._prepare_output_context("ORD-FUREAI", include_bags=False, include_ocr_menu_meta=False)
    columns = ctx["invoice_template"]["columns"]

    assert [column.get("header") for column in columns] == [
        "日付",
        "区分",
        "献立区分",
        "メニュー名",
        "常食",
        "通所",
        "職員",
        "魚禁",
        "備考欄",
    ]
    rows = output_builder._build_delivery_rows(
        ctx["order_for_outputs"],
        ctx["invoice_template"],
        ctx["quantity_rules"],
        ctx["facility_config"],
        {},
        allow_ocr_menu_meta=False,
    )
    assert rows[0]["qty.regular_x"] == 5


def test_build_daily_output_bundle_labels_groups_orders_per_facility(tmp_path, monkeypatch):
    monkeypatch.setattr(output_builder, "OUTPUT_DIR", tmp_path)
    monkeypatch.setattr(
        output_builder.order_service,
        "list_orders_by_line_date",
        lambda target_date, status=None: [
            {"id": "ORD-1", "facility": "FAC001"},
            {"id": "ORD-2", "facility": "FAC001"},
            {"id": "ORD-3", "facility": "FAC002"},
        ],
    )
    monkeypatch.setattr(
        output_builder.config_service,
        "get_facility_config",
        lambda facility_code: {
            "FAC001": {"facility_name": "そよかぜ"},
            "FAC002": {"facility_name": "大和なでしこ"},
        }.get(facility_code, {}),
    )
    monkeypatch.setattr(
        output_builder,
        "_prepare_output_context",
        lambda order_id: {
            "ORD-1": _make_context("ORD-1", "FAC001", "そよかぜ", "献立A"),
            "ORD-2": _make_context("ORD-2", "FAC001", "そよかぜ", "献立B"),
            "ORD-3": _make_context("ORD-3", "FAC002", "大和なでしこ", "献立C"),
        }[order_id],
    )
    monkeypatch.setattr(
        output_builder,
        "_build_label_rows",
        lambda bags, label_profile, facility_name: (
            [{"メニュー": bag.get("menu_name") or ""} for bag in bags],
            ["メニュー"],
            "jp",
        ),
    )

    bundle_path, summary = output_builder.build_daily_output_bundle(
        TARGET_DATE,
        bundle_type="labels",
    )

    assert bundle_path.suffix == ".xlsx"
    assert summary["file_format"] == "xlsx"
    assert summary["success_orders"] == 2
    workbook = load_workbook(bundle_path)
    assert workbook.sheetnames == ["メニュー", "大和なでしこ", "そよかぜ"]
    assert workbook["そよかぜ"]["A2"].value == "献立A"
    assert workbook["そよかぜ"]["A3"].value == "献立B"
    assert workbook["大和なでしこ"]["A2"].value == "献立C"


def test_build_order_lines_for_outputs_uses_newer_draft_materialization(monkeypatch):
    order = {
        "id": "ORD-DRAFT",
        "facility": "FAC001",
        "lines": [
            {
                "date": TARGET_DATE,
                "daypart": "朝",
                "menu_name": "stale",
                "diet_type": "regular",
                "area_id": "X",
                "quantity_original": 1,
            }
        ],
        "workflow_state": {"state": "apply_ready", "warnings": ["draft_newer_than_lines"]},
    }
    monkeypatch.setattr(output_builder.config_service, "get_facility_config", lambda facility_code: {})
    monkeypatch.setattr(output_builder.order_service, "_apply_change_override_priority_to_lines", lambda lines: lines)
    monkeypatch.setattr(output_builder.order_service, "_collect_menu_entries_for_week", lambda *args, **kwargs: [])
    monkeypatch.setattr(output_builder.order_service, "_collect_menu_items_for_week", lambda *args, **kwargs: [])
    monkeypatch.setattr(output_builder, "get_order_menu_snapshot", lambda order_id: None)
    monkeypatch.setattr(output_builder.daily_output_override_service, "apply_overrides_to_lines", lambda lines, facility_id: lines)
    monkeypatch.setattr(
        output_builder,
        "_build_nonwriting_draft_materialization_candidate",
        lambda order_id, **kwargs: {
            "error": None,
            "lines": [
                {
                    "date": TARGET_DATE,
                    "daypart": "朝",
                    "menu_name": "draft",
                    "diet_type": "regular",
                    "area_id": "X",
                    "quantity_original": 7,
                }
            ],
        },
    )

    lines = output_builder.build_order_lines_for_outputs(order, include_expanded_copy=False)

    assert [(line["menu_name"], line["quantity_original"]) for line in lines] == [("draft", 7)]


def test_build_order_lines_for_outputs_does_not_force_materialization_for_apply_ready_only(monkeypatch):
    order = {
        "id": "ORD-APPLY-READY",
        "facility": "FAC001",
        "lines": [
            {
                "date": TARGET_DATE,
                "daypart": "朝",
                "menu_name": "current",
                "diet_type": "regular",
                "area_id": "X",
                "quantity_original": 3,
            }
        ],
        "workflow_state": {"state": "apply_ready", "warnings": []},
    }
    monkeypatch.setattr(output_builder.config_service, "get_facility_config", lambda facility_code: {})
    monkeypatch.setattr(output_builder.order_service, "_apply_change_override_priority_to_lines", lambda lines: lines)
    monkeypatch.setattr(output_builder.order_service, "_collect_menu_entries_for_week", lambda *args, **kwargs: [])
    monkeypatch.setattr(output_builder.order_service, "_collect_menu_items_for_week", lambda *args, **kwargs: [])
    monkeypatch.setattr(output_builder, "get_order_menu_snapshot", lambda order_id: None)
    monkeypatch.setattr(output_builder.daily_output_override_service, "apply_overrides_to_lines", lambda lines, facility_id: lines)
    monkeypatch.setattr(
        output_builder,
        "_build_nonwriting_draft_materialization_candidate",
        lambda order_id, **kwargs: (_ for _ in ()).throw(AssertionError("materialization should not run")),
    )

    lines = output_builder.build_order_lines_for_outputs(order, include_expanded_copy=False)

    assert [(line["menu_name"], line["quantity_original"]) for line in lines] == [("current", 3)]


def test_build_order_lines_for_outputs_blocks_when_newer_draft_cannot_materialize(monkeypatch):
    order = {
        "id": "ORD-DRAFT",
        "facility": "FAC001",
        "lines": [],
        "workflow_state": {"state": "apply_ready", "warnings": ["draft_newer_than_lines"]},
    }
    monkeypatch.setattr(output_builder.config_service, "get_facility_config", lambda facility_code: {})
    monkeypatch.setattr(
        output_builder,
        "_build_nonwriting_draft_materialization_candidate",
        lambda order_id, **kwargs: {"error": "draft_materialization_mismatch", "lines": []},
    )

    try:
        output_builder.build_order_lines_for_outputs(order, include_expanded_copy=False)
    except ValueError as exc:
        assert "draft_newer_than_lines requires materialized draft lines" in str(exc)
    else:
        raise AssertionError("expected draft materialization blocker")


def test_nonwriting_materialization_rebuilds_blank_weekly_menu_from_canonical_bootstrap(monkeypatch):
    blank_draft = {
        "id": "DRF-BLANK",
        "base_evidence_run_id": "OEV-CURRENT",
        "draft_sheet_json": {
            "source": "weekly_menu",
            "fields": ["date_mmdd", "daypart", "menu", "qty.regular_x"],
            "rows": [["05/10", "昼", "献立A", ""]],
        },
    }
    bootstrap_sheet = {
        "source": "weekly_menu+identity",
        "fields": ["date_mmdd", "daypart", "menu", "qty.regular_x"],
        "rows": [["05/10", "昼", "献立A", "12"]],
    }
    calls = []

    monkeypatch.setattr(output_builder.order_service, "get_current_sheet_context", lambda *args, **kwargs: None)
    monkeypatch.setattr(output_builder.order_service, "get_latest_sheet_draft", lambda *args, **kwargs: blank_draft)
    monkeypatch.setattr(output_builder.order_service, "_source_uses_weekly_menu_shell", lambda source: str(source).startswith("weekly_menu"))
    monkeypatch.setattr(output_builder.order_service, "get_ocr_evidence_run", lambda evidence_run_id: {"id": evidence_run_id})
    monkeypatch.setattr(
        output_builder.order_service,
        "_build_canonical_bootstrap_sheet",
        lambda order_id, **kwargs: (bootstrap_sheet, None),
    )
    monkeypatch.setattr(
        output_builder.order_service,
        "_build_transient_draft_record",
        lambda order_id, sheet: {"id": None, "draft_sheet_json": sheet},
    )

    def fake_materialize(order_id, *, draft_record, facility_id, existing_week_code, received_at):
        calls.append(draft_record)
        if draft_record is blank_draft:
            return {"error": "draft_lines_empty", "lines": []}
        return {
            "error": None,
            "lines": [
                {
                    "date": TARGET_DATE,
                    "daypart": "昼",
                    "menu_name": "献立A",
                    "diet_type": "regular",
                    "quantity_original": 12,
                }
            ],
        }

    monkeypatch.setattr(output_builder.order_service, "_build_materialization_candidate_from_draft_record", fake_materialize)

    candidate = output_builder._build_nonwriting_draft_materialization_candidate(
        "ORD-DRAFT",
        facility_id="FAC001",
        week_value="2026-05@2026-05-10~2026-05-16",
        received_at=None,
    )

    assert len(calls) == 1
    assert calls[0]["draft_sheet_json"] is bootstrap_sheet
    assert candidate["error"] is None
    assert candidate["lines"][0]["quantity_original"] == 12


def test_build_order_lines_for_outputs_can_allow_stale_lines_for_audit(monkeypatch):
    order = {
        "id": "ORD-DRAFT",
        "facility": "FAC001",
        "lines": [
            {
                "date": TARGET_DATE,
                "daypart": "朝",
                "menu_name": "current",
                "diet_type": "regular",
                "area_id": "X",
                "quantity_original": 4,
            }
        ],
        "workflow_state": {"warnings": ["draft_newer_than_lines"]},
    }
    monkeypatch.setattr(output_builder.config_service, "get_facility_config", lambda facility_code: {})
    monkeypatch.setattr(output_builder.order_service, "_apply_change_override_priority_to_lines", lambda lines: lines)
    monkeypatch.setattr(output_builder.order_service, "_collect_menu_entries_for_week", lambda *args, **kwargs: [])
    monkeypatch.setattr(output_builder.order_service, "_collect_menu_items_for_week", lambda *args, **kwargs: [])
    monkeypatch.setattr(output_builder, "get_order_menu_snapshot", lambda order_id: None)
    monkeypatch.setattr(output_builder.daily_output_override_service, "apply_overrides_to_lines", lambda lines, facility_id: lines)
    monkeypatch.setattr(
        output_builder,
        "_build_nonwriting_draft_materialization_candidate",
        lambda order_id, **kwargs: {"error": "draft_lines_empty", "lines": []},
    )

    lines = output_builder.build_order_lines_for_outputs(
        order,
        include_expanded_copy=False,
        allow_stale_draft_lines=True,
    )

    assert [(line["menu_name"], line["quantity_original"]) for line in lines] == [("current", 4)]


def test_build_order_lines_for_outputs_accepts_stale_lines_keyword_without_materializing(monkeypatch):
    order = {
        "id": "ORD-DAILY-STABLE",
        "facility": "FAC001",
        "lines": [
            {
                "date": TARGET_DATE,
                "daypart": "朝",
                "menu_name": "current",
                "diet_type": "regular",
                "area_id": "X",
                "quantity_original": 4,
            }
        ],
        "workflow_state": {"warnings": ["draft_newer_than_lines"]},
    }
    monkeypatch.setattr(output_builder.config_service, "get_facility_config", lambda facility_code: {})
    monkeypatch.setattr(output_builder.order_service, "_apply_change_override_priority_to_lines", lambda lines: lines)
    monkeypatch.setattr(output_builder.order_service, "_collect_menu_entries_for_week", lambda *args, **kwargs: [])
    monkeypatch.setattr(output_builder.order_service, "_collect_menu_items_for_week", lambda *args, **kwargs: [])
    monkeypatch.setattr(output_builder, "get_order_menu_snapshot", lambda order_id: None)
    monkeypatch.setattr(output_builder.daily_output_override_service, "apply_overrides_to_lines", lambda lines, facility_id: lines)
    monkeypatch.setattr(
        output_builder,
        "_workflow_v2_lines_for_outputs",
        lambda *args, **kwargs: (_ for _ in ()).throw(AssertionError("stale audit path must not materialize")),
    )

    lines = output_builder.build_order_lines_for_outputs(
        order,
        include_expanded_copy=False,
        allow_stale_draft_lines=True,
    )

    assert [(line["menu_name"], line["quantity_original"]) for line in lines] == [("current", 4)]


def test_build_daily_output_bundle_delivery_groups_and_merges_quantities(tmp_path, monkeypatch):
    monkeypatch.setattr(output_builder, "OUTPUT_DIR", tmp_path)
    monkeypatch.setattr(
        output_builder.order_service,
        "list_orders_by_line_date",
        lambda target_date, status=None: [
            {"id": "ORD-1", "facility": "FAC001"},
            {"id": "ORD-2", "facility": "FAC001"},
        ],
    )
    monkeypatch.setattr(
        output_builder.config_service,
        "get_facility_config",
        lambda facility_code: {"facility_name": "そよかぜ"},
    )
    monkeypatch.setattr(
        output_builder,
        "_prepare_output_context",
        lambda order_id: _make_context(order_id, "FAC001", "そよかぜ", "献立A"),
    )
    monkeypatch.setattr(
        output_builder,
        "_build_delivery_rows",
        lambda order, template, quantity_rules, facility_config, menu_meta: [
            {
                "date": TARGET_DATE,
                "daypart": "朝",
                "menu_category": "主菜",
                "menu_name": "献立A",
                "menu_display": "主菜 献立A",
                "_order_index": 1,
                "常食": 2 if order.get("id") == "ORD-1" else 3,
            }
        ],
    )

    def _fake_write_delivery_note(path, rows, columns, template_uri, include_menu_name, sheet_name=None, facility_name=None):
        workbook = Workbook()
        ws = workbook.active
        ws.title = "納品書"
        ws["A1"] = "メニュー"
        ws["B1"] = "常食"
        for index, row in enumerate(rows, start=2):
            ws.cell(row=index, column=1).value = row.get("menu_name")
            ws.cell(row=index, column=2).value = row.get("常食")
        workbook.save(path)

    monkeypatch.setattr(output_builder, "_write_delivery_note", _fake_write_delivery_note)

    bundle_path, summary = output_builder.build_daily_output_bundle(
        TARGET_DATE,
        bundle_type="delivery",
    )

    assert bundle_path.suffix == ".xlsx"
    assert summary["success_orders"] == 1
    workbook = load_workbook(bundle_path)
    assert workbook.sheetnames == ["そよかぜ"]
    assert workbook["そよかぜ"]["A2"].value == "献立A"
    assert workbook["そよかぜ"]["B2"].value == 5


def test_build_daily_output_bundle_empty_orders_are_not_errors(tmp_path, monkeypatch):
    monkeypatch.setattr(output_builder, "OUTPUT_DIR", tmp_path)
    monkeypatch.setattr(
        output_builder.order_service,
        "list_orders_by_line_date",
        lambda target_date, status=None: [
            {"id": "ORD-1", "facility": "FAC001"},
            {"id": "ORD-2", "facility": "FAC002"},
        ],
    )
    monkeypatch.setattr(
        output_builder.config_service,
        "get_facility_config",
        lambda facility_code: {
            "FAC001": {"facility_name": "そよかぜ"},
            "FAC002": {"facility_name": "池袋"},
        }.get(facility_code, {}),
    )
    monkeypatch.setattr(
        output_builder,
        "_prepare_output_context",
        lambda order_id: _make_context(
            order_id,
            "FAC001" if order_id == "ORD-1" else "FAC002",
            "そよかぜ" if order_id == "ORD-1" else "池袋",
            "献立A",
        ),
    )
    monkeypatch.setattr(
        output_builder,
        "_build_delivery_rows",
        lambda order, template, quantity_rules, facility_config, menu_meta: [
            {
                "date": TARGET_DATE,
                "daypart": "朝",
                "menu_category": "主菜",
                "menu_name": "献立A",
                "menu_display": "主菜 献立A",
                "_order_index": 1,
                "常食": 2,
            }
        ]
        if order.get("id") == "ORD-1"
        else [],
    )

    def _fake_write_delivery_note(path, rows, columns, template_uri, include_menu_name, sheet_name=None, facility_name=None):
        workbook = Workbook()
        ws = workbook.active
        ws.title = "納品書"
        ws["A1"] = "メニュー"
        ws["A2"] = rows[0].get("menu_name") if rows else ""
        workbook.save(path)

    monkeypatch.setattr(output_builder, "_write_delivery_note", _fake_write_delivery_note)

    bundle_path, summary = output_builder.build_daily_output_bundle(
        TARGET_DATE,
        bundle_type="delivery",
    )

    assert bundle_path.suffix == ".xlsx"
    assert summary["success_orders"] == 1
    assert summary["empty_orders"] == 1
    assert summary["error_orders"] == 0
    assert [item["status"] for item in summary["items"]] == ["ok", "empty"]


def test_build_daily_output_bundle_both_uses_prefixed_sheet_titles(tmp_path, monkeypatch):
    monkeypatch.setattr(output_builder, "OUTPUT_DIR", tmp_path)
    monkeypatch.setattr(
        output_builder.order_service,
        "list_orders_by_line_date",
        lambda target_date, status=None: [{"id": "ORD-1", "facility": "FAC001"}],
    )
    monkeypatch.setattr(
        output_builder.config_service,
        "get_facility_config",
        lambda facility_code: {"facility_name": "そよかぜ"},
    )
    monkeypatch.setattr(
        output_builder,
        "_prepare_output_context",
        lambda order_id: _make_context(order_id, "FAC001", "そよかぜ", "献立A"),
    )
    monkeypatch.setattr(
        output_builder,
        "_build_label_rows",
        lambda bags, label_profile, facility_name: (
            [{"メニュー": bag.get("menu_name") or ""} for bag in bags],
            ["メニュー"],
            "jp",
        ),
    )
    monkeypatch.setattr(
        output_builder,
        "_build_delivery_rows",
        lambda order, template, quantity_rules, facility_config, menu_meta: [
            {
                "date": TARGET_DATE,
                "daypart": "朝",
                "menu_category": "主菜",
                "menu_name": "献立A",
                "menu_display": "主菜 献立A",
                "_order_index": 1,
                "常食": 2,
            }
        ],
    )

    def _fake_write_delivery_note(path, rows, columns, template_uri, include_menu_name, sheet_name=None, facility_name=None):
        workbook = Workbook()
        ws = workbook.active
        ws["A1"] = "メニュー"
        ws["A2"] = rows[0].get("menu_name") if rows else ""
        workbook.save(path)

    monkeypatch.setattr(output_builder, "_write_delivery_note", _fake_write_delivery_note)

    bundle_path, _ = output_builder.build_daily_output_bundle(
        TARGET_DATE,
        bundle_type="both",
    )

    workbook = load_workbook(bundle_path)
    assert workbook.sheetnames == ["ラベル_そよかぜ", "納品書_そよかぜ"]


def test_weekly_weight_collect_rows_counts_diabetes_as_regular_and_excludes_forbidden(monkeypatch):
    target_date = dt_date(2026, 5, 12)
    source_lines = [
        {
            "date": target_date,
            "daypart": "昼",
            "menu_category": "主Ａ",
            "menu_name": "麻婆茄子",
            "diet_type": "regular",
            "quantity_corrected": 402,
            "menu_unit_type": "g",
            "menu_qty_per_serving": 100,
        },
        {
            "date": target_date,
            "daypart": "昼",
            "menu_category": "主Ａ",
            "menu_name": "麻婆茄子",
            "diet_type": "daycare",
            "quantity_corrected": 36,
            "menu_unit_type": "g",
            "menu_qty_per_serving": 100,
        },
        {
            "date": target_date,
            "daypart": "昼",
            "menu_category": "主Ａ",
            "menu_name": "麻婆茄子",
            "diet_type": "staff",
            "quantity_corrected": 2,
            "menu_unit_type": "g",
            "menu_qty_per_serving": 100,
        },
        {
            "date": target_date,
            "daypart": "昼",
            "menu_category": "主Ａ",
            "menu_name": "麻婆茄子",
            "diet_type": "糖尿",
            "quantity_corrected": 5,
            "menu_unit_type": "g",
            "menu_qty_per_serving": 100,
        },
        {
            "date": target_date,
            "daypart": "昼",
            "menu_category": "主Ａ",
            "menu_name": "麻婆茄子",
            "diet_type": "no_meat",
            "quantity_corrected": 6,
            "menu_unit_type": "g",
            "menu_qty_per_serving": 100,
        },
        {
            "date": target_date,
            "daypart": "昼",
            "menu_category": "主Ａ",
            "menu_name": "麻婆茄子",
            "diet_type": "soft",
            "quantity_corrected": 28,
            "menu_unit_type": "g",
            "menu_qty_per_serving": 100,
        },
        {
            "date": target_date,
            "daypart": "昼",
            "menu_category": "主Ａ",
            "menu_name": "麻婆茄子",
            "diet_type": "mixer",
            "quantity_corrected": 21,
            "menu_unit_type": "g",
            "menu_qty_per_serving": 100,
        },
    ]

    class RaisingSession:
        def __enter__(self):
            raise RuntimeError("force fallback")

        def __exit__(self, exc_type, exc, tb):
            return False

    monkeypatch.setattr(output_builder, "session_scope", lambda: RaisingSession())
    monkeypatch.setattr(
        output_builder.order_service,
        "list_orders_by_line_date",
        lambda current_date, status=None: [{"id": "ORD-WEIGHT"}] if current_date == target_date else [],
    )
    monkeypatch.setattr(
        output_builder,
        "_prepare_output_context_for_bundle",
        lambda *args, **kwargs: {"order_lines": source_lines},
    )
    monkeypatch.setattr(output_builder.menu_service, "get_menu_entries_for_facility", lambda *args: [])

    rows = output_builder._weekly_weight_collect_rows(target_date)  # noqa: SLF001
    payload = rows[(target_date, "昼", "主Ａ")]

    assert payload["regular_quantity"] == 445
    assert payload["regular_amounts"] == {"g": 44500.0}
    assert payload["soft_mixer_quantity"] == 49
    assert payload["soft_mixer_amounts"] == {"g": 4900.0}


def test_weekly_weight_workbook_uses_reference_layout_and_calculated_values(tmp_path, monkeypatch):
    sample_path = _sawa_root() / "input_example" / "2026.0512" / "May 10-16 2026 Weight.xlsx"
    target_date = dt_date(2026, 5, 12)
    rows_by_key = {
        (dt_date(2026, 5, 10), "朝", "副①"): {
            "regular_menu": "計算生成メニューA",
            "regular_quantity": 10,
            "regular_amounts": {"g": 1200},
            "soft_mixer_menu": "計算生成メニューA軟菜",
            "soft_mixer_quantity": 3,
            "soft_mixer_amounts": {"g": 240},
        },
        (dt_date(2026, 5, 10), "昼", "主Ａ"): {
            "regular_menu": "個数生成メニュー",
            "regular_quantity": 4,
            "regular_amounts": {"__literal__": "4個＋ソース0.2"},
            "soft_mixer_menu": "",
            "soft_mixer_quantity": 0,
            "soft_mixer_amounts": {},
        },
        (dt_date(2026, 5, 10), "夕", "主"): {
            "regular_menu": "煮込みハンバーグ",
            "regular_quantity": 430,
            "regular_amounts": {
                "__main_unit__": "個",
                "__main_count__": 430,
                "__garnish_unit__": "g",
                "__garnish_amount__": 17200,
                "__garnish_label__": "ソース",
                "__garnish_separator__": "＋",
            },
            "soft_mixer_menu": "煮込みハンバーグ",
            "soft_mixer_quantity": 49,
            "soft_mixer_amounts": {
                "__main_unit__": "個",
                "__main_count__": 49,
                "__garnish_unit__": "g",
                "__garnish_amount__": 1960,
                "__garnish_label__": "",
                "__garnish_separator__": "、",
            },
        },
    }

    monkeypatch.setattr(output_builder, "OUTPUT_DIR", tmp_path)
    monkeypatch.setattr(
        output_builder,
        "_weekly_weight_collect_rows",
        lambda target_date, status=None: rows_by_key,
    )

    generated_path = output_builder.build_weekly_weight_summary_workbook(target_date)
    assert generated_path is not None
    generated_book = load_workbook(generated_path, data_only=False)
    sample_book_for_style = load_workbook(sample_path, data_only=False)
    generated_ws = generated_book.worksheets[0]
    sample_ws = sample_book_for_style.worksheets[0]

    assert generated_ws.title == sample_ws.title
    assert generated_ws.max_row == sample_ws.max_row
    assert generated_ws.max_column == sample_ws.max_column
    assert {str(item) for item in generated_ws.merged_cells.ranges} == {
        str(item) for item in sample_ws.merged_cells.ranges
    }
    assert _dimension_signature(generated_ws) == _dimension_signature(sample_ws)
    assert _page_signature(generated_ws) == _page_signature(sample_ws)
    assert generated_ws["D11"].value == "計算生成メニューA"
    assert generated_ws["E11"].value == 10
    assert generated_ws["F11"].value == 1.2
    assert generated_ws["G11"].value == 3
    assert generated_ws["H11"].value == 0.2
    assert generated_ws["I11"].value == "計算生成メニューA軟菜"
    assert generated_ws["D13"].value == "個数生成メニュー"
    assert generated_ws["E13"].value == 4
    assert generated_ws["F13"].value == "4個＋ソース0.2"
    assert generated_ws["D16"].value == "煮込みハンバーグ"
    assert generated_ws["E16"].value == 430
    assert generated_ws["F16"].value == "430個＋ソース17.2"
    assert generated_ws["G16"].value == 49
    assert generated_ws["H16"].value == "49個、2"


def test_weekly_weight_amount_rules_convert_piece_and_hidden_garnish_units():
    cases = [
        ("さつま芋の天ぷら", "regular", 422, "844個"),
        ("アジのちゃんちゃん焼き", "regular", 417, "1042.5切＋野菜16.7"),
        ("アジのちゃんちゃん焼き", "soft", 49, "98切、野菜2"),
        ("鶏唐揚げ", "regular", 410, "1230個、添12.3"),
        ("鶏唐揚げ", "soft", 49, "98個、添1.5"),
        ("サバの塩焼き", "regular", 417, "1042.5切、添12.5"),
        ("サバの塩焼き", "soft", 51, "102切、添1.5"),
        ("チキンカツ", "regular", 437, "437個、添13.1"),
        ("チキンカツ", "soft", 50, "50個、添1.5"),
        ("煮込みハンバーグ", "regular", 430, "430個＋ソース17.2"),
        ("煮込みハンバーグ", "soft", 49, "49個、2"),
    ]
    for menu_name, diet_type, quantity, expected in cases:
        amounts = {}
        output_builder._weekly_weight_add_amount(  # noqa: SLF001
            amounts,
            {
                "menu_name": menu_name,
                "diet_type": diet_type,
                "menu_unit_type": "g",
                "menu_qty_per_serving": 100,
            },
            quantity,
        )
        assert output_builder._weekly_weight_format_amount(amounts) == expected  # noqa: SLF001


def test_daily_bag_amount_rules_convert_piece_and_hidden_garnish_units():
    lines = [
        {
            "date": dt_date(2026, 5, 10),
            "daypart": "夕",
            "menu_name": "煮込みハンバーグ",
            "menu_category": "主菜",
            "diet_type": "regular",
            "area_id": "X",
            "quantity_corrected": 430,
            "menu_unit_type": "g",
            "menu_qty_per_serving": 100,
        },
        {
            "date": dt_date(2026, 5, 10),
            "daypart": "昼",
            "menu_name": "アジのちゃんちゃん焼き",
            "menu_category": "主菜",
            "diet_type": "soft",
            "area_id": "X",
            "quantity_corrected": 49,
            "menu_unit_type": "g",
            "menu_qty_per_serving": 100,
        },
        {
            "date": dt_date(2026, 5, 10),
            "daypart": "夕",
            "menu_name": "チキンカツ",
            "menu_category": "主菜",
            "diet_type": "regular",
            "area_id": "X",
            "quantity_corrected": 437,
            "menu_unit_type": "g",
            "menu_qty_per_serving": 100,
        },
    ]
    stats = order_service._build_daily_bag_amount_stats(lines)  # noqa: SLF001

    def per_serving(menu_name: str, diet_type: str) -> dict:
        key = order_service._build_non_condiment_amount_key(  # noqa: SLF001
            dt_date(2026, 5, 10),
            "夕" if menu_name != "アジのちゃんちゃん焼き" else "昼",
            menu_name,
            diet_type,
            "X",
        )
        return stats["per_serving_by_group"][key]

    assert per_serving("煮込みハンバーグ", "regular") == {"個": 1.0, "g": 40.0}
    assert per_serving("アジのちゃんちゃん焼き", "soft") == {"切": 2.0, "g": 40.0}
    assert per_serving("チキンカツ", "regular") == {"個": 1.0, "g": 30.0}


def test_build_daily_output_bundle_raises_when_no_rows_for_target_date(tmp_path, monkeypatch):
    monkeypatch.setattr(output_builder, "OUTPUT_DIR", tmp_path)
    monkeypatch.setattr(
        output_builder.order_service,
        "list_orders_by_line_date",
        lambda target_date, status=None: [{"id": "ORD-1", "facility": "FAC001"}],
    )
    monkeypatch.setattr(
        output_builder.config_service,
        "get_facility_config",
        lambda facility_code: {"facility_name": "そよかぜ"},
    )
    monkeypatch.setattr(
        output_builder,
        "_prepare_output_context",
        lambda order_id: _make_context(order_id, "FAC001", "そよかぜ", "献立A"),
    )
    monkeypatch.setattr(
        output_builder,
        "_build_label_rows",
        lambda bags, label_profile, facility_name: ([], ["メニュー"], "jp"),
    )

    try:
        output_builder.build_daily_output_bundle(TARGET_DATE, bundle_type="labels")
    except ValueError as exc:
        assert str(exc).startswith("対象日の出力対象がありません")
    else:
        raise AssertionError("ValueError was not raised")


def test_reference_daily_delivery_materializes_static_formula_labels(tmp_path):
    grouped_outputs = {
        "FAC00013": {
            "facility_code": "FAC00013",
            "facility_name": "いこいの森",
            "invoice_template": {},
            "contexts": [],
        }
    }
    workbook = output_builder._create_reference_daily_delivery_workbook(  # noqa: SLF001
        target_date=dt_date(2026, 5, 10),
        grouped_outputs=grouped_outputs,
    )
    output_path = tmp_path / "delivery.xlsx"
    workbook.save(output_path)

    saved = load_workbook(output_path, data_only=False)
    ws = saved["いこいの森"]

    assert ws["A17"].value == "(日)"
    assert ws["D17"].value == "煮込みハンバーグ"
    assert ws["D18"].value == "ジャーマンポテト"
    assert ws["D19"].value == "ほうれん草の和え物"


def test_reference_daily_delivery_rewrites_static_menu_cells_for_target_date(tmp_path):
    grouped_outputs = {
        "FAC00007": {
            "facility_code": "FAC00007",
            "facility_name": "アイテラス",
            "invoice_template": {},
            "contexts": [
                {
                    "order_for_outputs": {
                        "id": "ORD-menu-static",
                        "facility": "FAC00007",
                        "lines": [
                            {
                                "date": dt_date(2026, 5, 26),
                                "daypart": "朝",
                                "menu_category": "副①",
                                "menu_name": "厚揚げと竹輪の煮物",
                                "diet_type": "regular",
                                "quantity_original": 15,
                            }
                        ],
                    },
                    "quantity_rules": {"zero_as_empty": True},
                    "facility_config": {"facility_name": "アイテラス"},
                    "ocr_menu_meta": {},
                }
            ],
        }
    }
    workbook = output_builder._create_reference_daily_delivery_workbook(  # noqa: SLF001
        target_date=dt_date(2026, 5, 26),
        grouped_outputs=grouped_outputs,
    )
    output_path = tmp_path / "delivery.xlsx"
    workbook.save(output_path)

    saved = load_workbook(output_path, data_only=False)
    assert saved["山城"]["A12"].value.date() == dt_date(2026, 5, 26)
    assert saved["山城"]["D12"].value == "副① 厚揚げと竹輪の煮物"
    assert saved["山城"]["A17"].value == "(火)"


def test_menu_master_defaults_fill_missing_temperature(monkeypatch):
    monkeypatch.setattr(
        output_builder.menu_service,
        "resolve_menu_defaults",
        lambda menu_names, facility_id: {
            "ポークチャップ": {
                "unit_type": "g",
                "qty_per_serving": 100,
                "temp_type": "温菜",
                "category": "主菜",
            }
        },
    )

    rows = output_builder._apply_menu_master_defaults(  # noqa: SLF001
        [{"date": TARGET_DATE, "daypart": "昼", "menu_name": "ポークチャップ", "menu_category": "主菜"}],
        "FAC00009",
    )

    assert rows[0]["menu_temp_type"] == "温菜"
    assert rows[0]["menu_qty_per_serving"] == 100
    assert rows[0]["menu_unit_type"] == "g"


def test_reference_daily_delivery_preserves_table_borders(tmp_path):
    grouped_outputs = {
        "FAC00012": {
            "facility_code": "FAC00012",
            "facility_name": "ふれあいの丘",
            "invoice_template": {},
            "contexts": [],
        }
    }
    workbook = output_builder._create_reference_daily_delivery_workbook(  # noqa: SLF001
        target_date=dt_date(2026, 5, 10),
        grouped_outputs=grouped_outputs,
    )
    output_path = tmp_path / "delivery.xlsx"
    output_builder._save_reference_daily_delivery_workbook_preserving_template_package(  # noqa: SLF001
        workbook,
        output_path,
    )

    actual = load_workbook(output_path, data_only=False)
    for sheet_name in actual.sheetnames:
        assert sheet_name in actual.sheetnames
        actual_ws = actual[sheet_name]
        dotted_cells = _dotted_table_border_cells(actual_ws)
        assert not dotted_cells, (
            f"{sheet_name} daily delivery table must not keep dotted borders: {dotted_cells}"
        )


def test_reference_daily_delivery_template_is_single_blank_sheet():
    workbook = load_workbook(output_builder.DAILY_DELIVERY_REFERENCE_TEMPLATE, data_only=False)

    assert workbook.sheetnames == ["テンプレート"]
    ws = workbook["テンプレート"]
    formulas = []
    populated_dynamic_cells = []
    junk_cells = []
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            if isinstance(cell.value, str) and cell.value.startswith("="):
                formulas.append(cell.coordinate)
            if cell.value in {"v", "V", "√", "✓", "./", " ", "　"}:
                junk_cells.append(cell.coordinate)
            if 12 <= cell.row <= 19 and (
                cell.column == 1 or cell.column == 4 or cell.column >= 5
            ) and cell.value not in (None, ""):
                populated_dynamic_cells.append(cell.coordinate)

    assert formulas == []
    assert junk_cells == []
    assert populated_dynamic_cells == []
    assert ws["A4"].value in (None, "")
    assert {
        ws.cell(row=19, column=col_idx).border.bottom.style
        for col_idx in range(1, ws.max_column + 1)
        if not isinstance(ws.cell(row=19, column=col_idx), MergedCell)
    } == {"medium"}
    assert ws["A17"].border.bottom.style == "medium"
    assert ws["B17"].border.bottom.style == "medium"
    assert {
        ws.cell(row=20, column=col_idx).border.top.style
        for col_idx in range(1, 12)
    } == {None}


def test_reference_daily_delivery_writing_values_does_not_change_template_borders(tmp_path):
    template = load_workbook(output_builder.DAILY_DELIVERY_REFERENCE_TEMPLATE, data_only=False)
    before_ws = template["テンプレート"]
    before = {
        cell.coordinate: _border_signature(cell.border)
        for row in before_ws.iter_rows(min_row=12, max_row=19)
        for cell in row
        if not isinstance(cell, MergedCell)
    }

    workbook = output_builder._create_reference_daily_delivery_workbook(  # noqa: SLF001
        target_date=TARGET_DATE,
        grouped_outputs={
            "FAC00009": {
                "facility_code": "FAC00009",
                "facility_name": "グループホームそよかぜ",
                "invoice_template": {},
                "contexts": [
                    {
                        "order_for_outputs": {
                            "id": "ORD-border",
                            "facility": "FAC00009",
                            "lines": [
                                {
                                    "date": TARGET_DATE,
                                    "daypart": "朝",
                                    "menu_category": "副①",
                                    "menu_name": "献立A",
                                    "diet_type": "regular",
                                    "quantity_original": 3,
                                }
                            ],
                        },
                        "quantity_rules": {"zero_as_empty": True},
                        "facility_config": {"facility_name": "グループホームそよかぜ"},
                        "ocr_menu_meta": {},
                    }
                ],
            }
        },
    )
    output_path = tmp_path / "delivery.xlsx"
    workbook.save(output_path)
    saved = load_workbook(output_path, data_only=False)
    ws = saved["そよかぜ"]
    after = {
        cell.coordinate: _border_signature(cell.border)
        for row in ws.iter_rows(min_row=12, max_row=19)
        for cell in row
        if not isinstance(cell, MergedCell)
    }

    assert after == before
    assert ws["E12"].value == 3


def test_delivery_slot_assignment_does_not_cross_daypart_when_template_section_is_full():
    assignments = output_builder._assign_delivery_rows_to_slots(  # noqa: SLF001
        [
            {"daypart": "朝", "menu_category": "副①", "menu_name": "朝副1", "_order_index": 1},
            {"daypart": "朝", "menu_category": "副②", "menu_name": "朝副2", "_order_index": 2},
            {"daypart": "朝", "menu_category": "主", "menu_name": "朝主", "_order_index": 3},
        ],
        [12, 13, 14, 15, 16, 17, 18, 19],
        {
            ("朝", "副1"): 12,
            ("朝", "副2"): 13,
            ("昼", "主a"): 14,
            ("昼", "副1"): 15,
            ("昼", "副2"): 16,
            ("夕", "主"): 17,
            ("夕", "副1"): 18,
            ("夕", "副2"): 19,
        },
        {"副1": [12, 15, 18], "副2": [13, 16, 19], "主": [17], "主a": [14]},
        {
            "朝": {"副1": [12], "副2": [13]},
            "昼": {"主a": [14], "副1": [15], "副2": [16]},
            "夕": {"主": [17], "副1": [18], "副2": [19]},
        },
        {"朝": [12, 13], "昼": [14, 15, 16], "夕": [17, 18, 19]},
    )

    assert assignments[12]["menu_name"] == "朝副1"
    assert assignments[13]["menu_name"] == "朝副2"
    assert 17 not in assignments


def test_reference_daily_delivery_writes_excel_readable_workbook(tmp_path):
    grouped_outputs = {
        "FAC00012": {
            "facility_code": "FAC00012",
            "facility_name": "ふれあいの丘",
            "invoice_template": {},
            "contexts": [],
        }
    }
    workbook = output_builder._create_reference_daily_delivery_workbook(  # noqa: SLF001
        target_date=dt_date(2026, 5, 10),
        grouped_outputs=grouped_outputs,
    )
    output_path = tmp_path / "delivery.xlsx"
    output_builder._save_reference_daily_delivery_workbook_preserving_template_package(  # noqa: SLF001
        workbook,
        output_path,
    )

    saved = load_workbook(output_path, data_only=False)
    assert "ふれあいの丘" in saved.sheetnames


def test_daily_label_jp_keeps_units_servings_and_facility_floor_suffix():
    rows, fields, label_format = output_builder._build_label_rows(  # noqa: SLF001
        [
            {
                "facility": "FAC00009",
                "date": TARGET_DATE,
                "daypart": "朝",
                "menu_name": "ごぼうと竹輪の煮物",
                "menu_category": "副菜①",
                "diet_type": "regular",
                "area_id": "2F",
                "menu_unit_type": "g",
                "menu_qty_per_serving": 70,
                "menu_temp_type": "温菜",
                "quantity": 3,
            }
        ],
        {},
        "グループホームそよかぜ",
    )

    assert label_format == "jp"
    assert fields[-1] == ""
    assert rows[0]["時間"] == "朝　2階"
    assert rows[0]["内容量"] == "210g"
    assert rows[0]["内容詳細"] == "70g"
    assert rows[0][""] == "3人前"


def test_daily_label_jp_sorts_floor_rows_by_menu_before_floor():
    rows, fields, label_format = output_builder._build_label_rows(  # noqa: SLF001
        [
            {
                "facility": "FAC00008",
                "date": TARGET_DATE,
                "daypart": "夕",
                "menu_name": "れんこんの甘辛煮",
                "menu_category": "副菜",
                "diet_type": "soft",
                "area_id": "2F",
                "menu_unit_type": "g",
                "menu_qty_per_serving": 40,
                "menu_temp_type": "温菜",
                "quantity": 9,
            },
            {
                "facility": "FAC00008",
                "date": TARGET_DATE,
                "daypart": "夕",
                "menu_name": "鶏すき焼き風",
                "menu_category": "主菜",
                "diet_type": "soft",
                "area_id": "2F",
                "menu_unit_type": "g",
                "menu_qty_per_serving": 100,
                "menu_temp_type": "温菜",
                "quantity": 9,
            },
            {
                "facility": "FAC00008",
                "date": TARGET_DATE,
                "daypart": "夕",
                "menu_name": "鶏すき焼き風",
                "menu_category": "主菜",
                "diet_type": "soft",
                "area_id": "3F",
                "menu_unit_type": "g",
                "menu_qty_per_serving": 100,
                "menu_temp_type": "温菜",
                "quantity": 6,
            },
            {
                "facility": "FAC00008",
                "date": TARGET_DATE,
                "daypart": "夕",
                "menu_name": "れんこんの甘辛煮",
                "menu_category": "副菜",
                "diet_type": "soft",
                "area_id": "3F",
                "menu_unit_type": "g",
                "menu_qty_per_serving": 40,
                "menu_temp_type": "温菜",
                "quantity": 6,
            },
        ],
        {},
        "佐古グループホーム",
    )

    assert label_format == "jp"
    assert fields[-1] == ""
    assert [(row["商品名１"], row["時間"]) for row in rows] == [
        ("鶏すき焼き風", "夕　2階"),
        ("鶏すき焼き風", "夕　3階"),
        ("れんこんの甘辛煮", "夕　2階"),
        ("れんこんの甘辛煮", "夕　3階"),
    ]


def test_daily_label_jp_excludes_forbidden_diets():
    rows, _fields, _label_format = output_builder._build_label_rows(  # noqa: SLF001
        [
            {
                "facility": "FAC00009",
                "date": TARGET_DATE,
                "daypart": "昼",
                "menu_name": "豚肉と白菜のすき煮",
                "menu_category": "主菜",
                "diet_type": "no_fish",
                "area_id": "2F",
                "menu_unit_type": "g",
                "menu_qty_per_serving": 100,
                "menu_temp_type": "温菜",
                "quantity": 1,
            }
        ],
        {},
        "グループホームそよかぜ",
    )

    assert rows == []


def test_delivery_rows_use_reference_slot_labels_and_order_for_html():
    rows = output_builder._build_delivery_rows(  # noqa: SLF001
        {
            "id": "ORD-slots",
            "facility": "FAC00009",
            "lines": [
                {
                    "date": TARGET_DATE,
                    "daypart": "夕",
                    "menu_category": "主菜",
                    "menu_name": "ジャーマンポテト",
                    "diet_type": "regular",
                    "area_id": "2F",
                    "quantity_original": 3,
                },
                {
                    "date": TARGET_DATE,
                    "daypart": "夕",
                    "menu_category": "主菜",
                    "menu_name": "煮込みハンバーグ",
                    "diet_type": "regular",
                    "area_id": "2F",
                    "quantity_original": 3,
                },
                {
                    "date": TARGET_DATE,
                    "daypart": "夕",
                    "menu_category": "副菜",
                    "menu_name": "ほうれん草の和え物",
                    "diet_type": "regular",
                    "area_id": "2F",
                    "quantity_original": 3,
                },
                {
                    "date": TARGET_DATE,
                    "daypart": "昼",
                    "menu_category": "副菜",
                    "menu_name": "豚肉と白菜のすき煮",
                    "diet_type": "regular",
                    "area_id": "2F",
                    "quantity_original": 3,
                },
                {
                    "date": TARGET_DATE,
                    "daypart": "昼",
                    "menu_category": "副菜",
                    "menu_name": "さつま芋の天ぷら",
                    "diet_type": "regular",
                    "area_id": "2F",
                    "quantity_original": 3,
                },
            ],
        },
        {
            "columns": [
                {"name": "日付", "source": "date"},
                {"name": "区分", "source": "daypart"},
                {"name": "献立区分", "source": "menu_category"},
                {"name": "メニュー名", "source": "menu_name"},
                {"name": "常食2F", "source": "quantity", "diet_type": "regular", "area_id": "2F"},
            ]
        },
        {"zero_as_empty": True},
        {"facility_name": "グループホームそよかぜ"},
        {},
        allow_ocr_menu_meta=False,
    )

    assert [(row["daypart"], row["menu_category"], row["menu_name"]) for row in rows] == [
        ("昼", "主Ａ", "豚肉と白菜のすき煮"),
        ("昼", "副①", "さつま芋の天ぷら"),
        ("夕", "主", "煮込みハンバーグ"),
        ("夕", "副①", "ジャーマンポテト"),
        ("夕", "副②", "ほうれん草の和え物"),
    ]


def test_delivery_rows_use_reference_slots_when_source_categories_are_raw():
    lines = []
    for index, (daypart, category, menu_name) in enumerate(
        [
            ("朝", "主", "厚揚げと竹輪の煮物"),
            ("朝", "副菜", "ほうれん草のお浸し"),
            ("昼", "主Ａ", "ポークチャップ"),
            ("昼", "副菜", "マカロニソテー"),
            ("昼", "副菜", "胡瓜のサラダ"),
            ("夕", "主", "オムレツのカニ玉風"),
            ("夕", "主", "さつま芋の煮物"),
            ("夕", "副菜", "三色ナムル"),
        ]
    ):
        lines.append(
            {
                "date": TARGET_DATE,
                "daypart": daypart,
                "menu_category": category,
                "menu_name": menu_name,
                "diet_type": "regular",
                "area_id": "2F",
                "quantity_original": 3,
            }
        )

    rows = output_builder._build_delivery_rows(  # noqa: SLF001
        {"id": "ORD-raw-slots", "facility": "FAC00008", "lines": lines},
        {
            "columns": [
                {"name": "日付", "source": "date"},
                {"name": "区分", "source": "daypart"},
                {"name": "献立区分", "source": "menu_category"},
                {"name": "メニュー名", "source": "menu_name"},
                {"name": "常食2F", "source": "quantity", "diet_type": "regular", "area_id": "2F"},
            ]
        },
        {"zero_as_empty": True},
        {"facility_name": "佐古"},
        {},
        allow_ocr_menu_meta=False,
    )

    assert [(row["daypart"], row["menu_category"], row["menu_name"]) for row in rows] == [
        ("朝", "副①", "厚揚げと竹輪の煮物"),
        ("朝", "副②", "ほうれん草のお浸し"),
        ("昼", "主Ａ", "ポークチャップ"),
        ("昼", "副①", "マカロニソテー"),
        ("昼", "副②", "胡瓜のサラダ"),
        ("夕", "主", "オムレツのカニ玉風"),
        ("夕", "副①", "さつま芋の煮物"),
        ("夕", "副②", "三色ナムル"),
    ]


def test_delivery_rows_show_condiment_next_to_main_menu(monkeypatch):
    monkeypatch.setattr(
        output_builder.menu_service,
        "resolve_menu_defaults",
        lambda menu_names, facility_id: {"主菜A": {"condiments": ["キャベツ"]}},
    )
    output_builder._cached_menu_defaults.cache_clear()  # noqa: SLF001

    rows = output_builder._build_delivery_rows(  # noqa: SLF001
        {
            "id": "ORD-condiment",
            "facility": "FAC00009",
            "lines": [
                {
                    "date": TARGET_DATE,
                    "daypart": "昼",
                    "menu_category": "主Ａ",
                    "menu_name": "主菜A",
                    "diet_type": "regular",
                    "area_id": "2F",
                    "quantity_original": 1,
                }
            ],
        },
        {
            "columns": [
                {"name": "日付", "source": "date"},
                {"name": "区分", "source": "daypart"},
                {"name": "メニュー名", "source": "menu_display"},
                {"name": "常食2F", "source": "quantity", "diet_type": "regular", "area_id": "2F"},
            ]
        },
        {"zero_as_empty": True},
        {"facility_name": "グループホームそよかぜ"},
        {},
        allow_ocr_menu_meta=False,
    )

    assert rows[0]["menu_display"] == "主Ａ 主菜A 添）キャベツ"


def test_reference_daily_delivery_removes_static_artifacts(tmp_path):
    workbook = output_builder._create_reference_daily_delivery_workbook(  # noqa: SLF001
        target_date=dt_date(2026, 5, 10),
        grouped_outputs={},
    )
    output_path = tmp_path / "delivery.xlsx"
    output_builder._save_reference_daily_delivery_workbook_preserving_template_package(  # noqa: SLF001
        workbook,
        output_path,
    )

    saved = load_workbook(output_path, data_only=False)
    assert saved["山城"]["C27"].value is None
    for ws in saved.worksheets:
        for col_idx in range(1, ws.max_column + 1):
            assert ws.cell(row=20, column=col_idx).value in (None, ""), (
                f"{ws.title}!{ws.cell(row=20, column=col_idx).coordinate} must not contain static artifacts"
            )
    for row_idx in range(12, 20):
        assert saved["池袋病院"].cell(row=row_idx, column=5).value is None
        assert saved["池袋病院"].cell(row=row_idx, column=6).value is None


def test_daily_output_both_bundle_uses_reference_delivery_and_label_rows(tmp_path, monkeypatch):
    monkeypatch.setattr(output_builder, "OUTPUT_DIR", tmp_path)
    order_id = "ORD-both"
    facility_code = "FAC00009"
    facility_name = "グループホームそよかぜ"
    order_lines = [
        {
            "date": TARGET_DATE,
            "daypart": "朝",
            "menu_category": "副①",
            "menu_name": "ごぼうと竹輪の煮物",
            "diet_type": "regular",
            "area_id": "2F",
            "quantity_original": 3,
            "menu_unit_type": "g",
            "menu_qty_per_serving": 70,
            "menu_temp_type": "温菜",
        },
        {
            "date": TARGET_DATE,
            "daypart": "朝",
            "menu_category": "副①",
            "menu_name": "ごぼうと竹輪の煮物",
            "diet_type": "no_fish",
            "area_id": "2F",
            "quantity_original": 1,
            "menu_unit_type": "g",
            "menu_qty_per_serving": 70,
            "menu_temp_type": "温菜",
        },
    ]
    context = {
        "order": {"id": order_id, "facility": facility_code},
        "order_lines": order_lines,
        "order_for_outputs": {"id": order_id, "facility": facility_code, "lines": order_lines},
        "bags": output_builder.build_bag_rows_for_outputs(
            {"id": order_id, "facility": facility_code, "lines": order_lines},
            order_lines=order_lines,
            facility_config={"facility_name": facility_name},
        ),
        "label_profile": {},
        "facility_config": {"facility_name": facility_name},
        "invoice_template": {},
        "quantity_rules": {"zero_as_empty": True},
        "ocr_menu_meta": {},
    }
    monkeypatch.setattr(
        output_builder.order_service,
        "list_orders_by_line_date",
        lambda target_date, status=None: [{"id": order_id, "facility": facility_code}],
    )
    monkeypatch.setattr(
        output_builder.config_service,
        "get_facility_config",
        lambda code: {"facility_name": facility_name},
    )
    monkeypatch.setattr(
        output_builder,
        "_prepare_output_context_for_bundle",
        lambda order_id, **kwargs: context,
    )

    output_path, manifest = output_builder.build_daily_output_bundle(TARGET_DATE, bundle_type="both", status="confirmed")
    saved = load_workbook(output_path, data_only=False)

    assert manifest["success_orders"] == 1
    assert "そよかぜ" in saved.sheetnames
    label_sheets = [name for name in saved.sheetnames if name.startswith("ラベル_そよかぜ")]
    assert label_sheets
    label_ws = saved[label_sheets[0]]
    assert label_ws["D2"].value == "朝　2階"
    assert label_ws["E2"].value == "副菜①"
    assert label_ws["I2"].value == "210g"
    assert label_ws["J2"].value == "70g"
    assert label_ws["K2"].value == "3人前"
    assert saved["そよかぜ"]["E12"].value == 3
    assert saved["そよかぜ"]["L12"].value in (None, "")
    assert not _dotted_table_border_cells(saved["そよかぜ"])


def test_daily_bundle_blocks_embedding_templated_delivery_workbook(tmp_path):
    template_path = tmp_path / "delivery_template.xlsx"
    workbook = Workbook()
    workbook.active.title = "Template"
    workbook.save(template_path)

    try:
        output_builder._create_daily_delivery_sheet(  # noqa: SLF001
            Workbook(),
            set(),
            title_seed="delivery",
            rows=[{"date": dt_date(2026, 5, 10), "menu_name": "A"}],
            invoice_template={
                "template_uri": template_path.as_uri(),
                "columns": [{"name": "メニュー", "source": "menu_name", "column_index": 1}],
            },
            facility_name="施設",
            order_id="ORD-test",
        )
    except ValueError as exc:
        assert str(exc) == "templated delivery notes cannot be embedded into a rebuilt daily bundle workbook"
    else:
        raise AssertionError("templated delivery embedding should be blocked")
