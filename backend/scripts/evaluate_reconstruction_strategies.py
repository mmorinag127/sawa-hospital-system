#!/usr/bin/env python3

from __future__ import annotations

import argparse
from datetime import date, datetime
import json
from pathlib import Path
import re
import sys
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
OCR_PIPELINE_ROOT = ROOT.parent / "ocr_pipeline"
for candidate in (str(ROOT), str(OCR_PIPELINE_ROOT)):
    if candidate not in sys.path:
        sys.path.insert(0, candidate)

from scripts import compare_structure_guided_ocr as compare
from src.services import config_service, order_form_service
from src.services.quantity_subgrid_experiment import (
    crop_image_by_norm_box,
    infer_quantity_subgrid,
    reread_suspicious_quantity_cells,
)
from src.services.structure_guided_ocr import (
    assign_words_to_structure_table,
    assign_words_to_structure_table_by_overlap,
    build_sequence_guided_table,
    repair_menu_tail_quantity_shift,
    select_primary_table,
)
from src.services.workbook_pdf_renderer import render_workbook_path_to_pdf


_DEFAULT_BODY_START_ROW = 11
_DEFAULT_HEADER_ROW_COUNT = 2
_DATE_TOKEN_RE = re.compile(r"_(\d{4})")
_MENU_TAIL_DIGITS_RE = re.compile(r"\d{2,3}$")
_NUMERIC_RE = re.compile(r"^\d+$")
_CANONICAL_DAYPARTS = {"", "朝", "昼", "夕"}


def _write_json(path: Path, payload: Any) -> None:
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def _normalize_text(value: object) -> str:
    if isinstance(value, datetime):
        return f"{value.month}/{value.day}"
    if isinstance(value, date):
        return f"{value.month}/{value.day}"
    return str(value or "").strip()


def _table_rows(table: dict[str, Any] | None, *, row_count: int, col_count: int) -> list[list[str]]:
    normalized = [["" for _ in range(col_count)] for _ in range(row_count)]
    if not isinstance(table, dict):
        return normalized
    source_rows = table.get("rows") or []
    for row_index in range(min(row_count, len(source_rows))):
        raw_row = source_rows[row_index]
        if not isinstance(raw_row, list):
            continue
        for col_index in range(min(col_count, len(raw_row))):
            normalized[row_index][col_index] = _normalize_text(raw_row[col_index])
    return normalized


def _copy_table_with_rows(table: dict[str, Any], rows: list[list[str]], *, source: str) -> dict[str, Any]:
    copied = dict(table)
    copied["rows"] = rows
    copied["row_count"] = len(rows)
    copied["col_count"] = max((len(row) for row in rows), default=0)
    copied["source"] = source
    return copied


def _load_facilities() -> list[dict[str, Any]]:
    facility_master_path = ROOT / "src" / "data" / "facility_master.template.json"
    payload = json.loads(facility_master_path.read_text(encoding="utf-8"))
    facilities = payload.get("facilities") or []
    return [item for item in facilities if isinstance(item, dict)]


def _source_context(*, facility_id: str, week_sheet_name: str) -> tuple[Path, Any]:
    facility = config_service.get_facility_config(facility_id)
    if not facility:
        raise ValueError(f"facility not found: {facility_id}")
    fax_template_id = str(order_form_service._infer_fax_template_id_from_facility(facility) or "").strip()  # noqa: SLF001
    source_workbook_name = order_form_service._resolve_source_workbook_name_for_week_sheet(  # noqa: SLF001
        fax_template_id,
        week_sheet_name,
    )
    source_workbook_path = order_form_service._resolve_source_workbook_path(source_workbook_name)  # noqa: SLF001
    workbook = load_workbook(source_workbook_path, data_only=True)
    return source_workbook_path, workbook[week_sheet_name]


def _canonical_rows_from_worksheet(
    worksheet,
    *,
    row_count: int,
    col_count: int,
) -> list[dict[str, Any]]:
    canonical_rows: list[dict[str, Any]] = []
    data_rows = max(0, row_count - _DEFAULT_HEADER_ROW_COUNT)
    for offset in range(data_rows):
        worksheet_row = _DEFAULT_BODY_START_ROW + offset
        cells = [
            _normalize_text(worksheet.cell(row=worksheet_row, column=col_index + 1).value)
            for col_index in range(col_count)
        ]
        canonical_rows.append(
            {
                "row_index": _DEFAULT_HEADER_ROW_COUNT + offset,
                "worksheet_row": worksheet_row,
                "date": cells[0] if col_count > 0 else "",
                "daypart": cells[1] if col_count > 1 else "",
                "aux": cells[2] if col_count > 2 else "",
                "menu_name": cells[3] if col_count > 3 else "",
                "cells": cells,
            }
        )
    return canonical_rows


def _expected_rows(
    *,
    structure_table: dict[str, Any],
    worksheet,
) -> list[list[str]]:
    row_count = int(structure_table.get("row_count") or 0)
    col_count = int(structure_table.get("col_count") or 0)
    rows = [["" for _ in range(col_count)] for _ in range(row_count)]
    structure_rows = structure_table.get("rows") or []
    for row_index in range(min(_DEFAULT_HEADER_ROW_COUNT, row_count, len(structure_rows))):
        raw_row = structure_rows[row_index]
        if not isinstance(raw_row, list):
            continue
        for col_index in range(min(col_count, len(raw_row))):
            rows[row_index][col_index] = _normalize_text(raw_row[col_index])

    canonical_rows = _canonical_rows_from_worksheet(
        worksheet,
        row_count=row_count,
        col_count=col_count,
    )
    for canonical in canonical_rows:
        row_index = int(canonical["row_index"])
        cells = canonical.get("cells") or []
        for col_index in range(min(col_count, len(cells))):
            rows[row_index][col_index] = _normalize_text(cells[col_index])
    return rows


def _metric_bundle(
    *,
    rows: list[list[str]],
    expected_rows: list[list[str]] | None,
    menu_col_index: int = 3,
    quantity_start_col_index: int = 4,
) -> dict[str, Any]:
    row_count = len(rows)
    col_count = max((len(row) for row in rows), default=0)
    exact_total = 0
    exact_meta = 0
    exact_qty = 0
    expected_non_empty = 0
    expected_qty_non_empty = 0
    row_exact = 0

    if expected_rows is not None:
        compare_row_count = min(len(expected_rows), row_count)
        compare_col_count = min(max((len(row) for row in expected_rows), default=0), col_count)
        for row_index in range(compare_row_count):
            row_matches = True
            for col_index in range(compare_col_count):
                actual = _normalize_text(rows[row_index][col_index] if col_index < len(rows[row_index]) else "")
                expected = _normalize_text(expected_rows[row_index][col_index] if col_index < len(expected_rows[row_index]) else "")
                if expected:
                    expected_non_empty += 1
                    if col_index >= quantity_start_col_index:
                        expected_qty_non_empty += 1
                    if actual == expected:
                        exact_total += 1
                        if col_index < quantity_start_col_index:
                            exact_meta += 1
                        else:
                            exact_qty += 1
                    else:
                        row_matches = False
            if row_matches:
                row_exact += 1

    menu_tail_rows = 0
    quantity_numeric_cells = 0
    quantity_non_numeric_cells = 0
    valid_daypart_rows = 0
    for row in rows[_DEFAULT_HEADER_ROW_COUNT:]:
        menu_text = _normalize_text(row[menu_col_index] if menu_col_index < len(row) else "")
        if menu_text and _MENU_TAIL_DIGITS_RE.search(menu_text):
            menu_tail_rows += 1
        daypart = _normalize_text(row[1] if len(row) > 1 else "")
        if daypart in _CANONICAL_DAYPARTS:
            valid_daypart_rows += 1
        for cell in row[quantity_start_col_index:]:
            text = _normalize_text(cell)
            if not text:
                continue
            if _NUMERIC_RE.fullmatch(text):
                quantity_numeric_cells += 1
            else:
                quantity_non_numeric_cells += 1

    return {
        "row_count": row_count,
        "col_count": col_count,
        "exact_total": exact_total,
        "exact_meta": exact_meta,
        "exact_qty": exact_qty,
        "expected_non_empty": expected_non_empty,
        "expected_qty_non_empty": expected_qty_non_empty,
        "row_exact": row_exact,
        "menu_tail_rows": menu_tail_rows,
        "quantity_numeric_cells": quantity_numeric_cells,
        "quantity_non_numeric_cells": quantity_non_numeric_cells,
        "valid_daypart_rows": valid_daypart_rows,
        "head_rows": rows[:5],
    }


def _week_from_pdf_name(pdf_name: str) -> str | None:
    match = _DATE_TOKEN_RE.search(pdf_name)
    if not match:
        return None
    token = match.group(1)
    mapping = {
        "0208": "2月8日～2月14日",
        "0215": "2月15日～2月21日",
        "0322": "3月22日～3月28日",
        "0426": "4月26日～4月30日",
    }
    return mapping.get(token)


def _find_pdf_from_inventory_name(pdf_name: str) -> Path | None:
    search_roots = [
        ROOT.parent.parent / "input_example",
        ROOT.parent.parent / "tmp" / "reupload_legacy_split",
    ]
    normalized_name = pdf_name.replace("_", "").replace(" ", "")
    for root in search_roots:
        if not root.exists():
            continue
        for candidate in root.rglob("*.pdf"):
            normalized_candidate = candidate.name.replace("_", "").replace(" ", "")
            if normalized_candidate == normalized_name:
                return candidate
    return None


def _prepare_structure_bundle(
    *,
    facility_id: str,
    week_sheet_name: str,
    output_dir: Path,
    dpi: int,
    device: str,
) -> dict[str, Any]:
    structure_xlsx = order_form_service.build_fax_structure_only_excel(
        facility_id=facility_id,
        week_sheet_name=week_sheet_name,
        output_dir=output_dir,
    )
    structure_pdf_path = output_dir / f"{structure_xlsx.stem}.pdf"
    render_workbook_path_to_pdf(
        structure_xlsx,
        output_path=structure_pdf_path,
        sheet_name=week_sheet_name,
        dpi=dpi,
    )
    structure_pdf_bytes = structure_pdf_path.read_bytes()
    structure_result = compare._run_structure_template(  # noqa: SLF001
        structure_pdf_bytes=structure_pdf_bytes,
        dpi=dpi,
        device=device,
    )
    structure_table = structure_result["primary_table"]
    if not isinstance(structure_table, dict):
        raise RuntimeError("structure-only template PDF did not produce a primary table")

    template_png_bytes = compare.render_pdf_to_png_bytes(structure_pdf_bytes, dpi=dpi, page=1)
    template_png_path = output_dir / "structure_template_page1.png"
    template_png_path.write_bytes(template_png_bytes)
    template_width, template_height = compare._decode_png_size(template_png_bytes)  # noqa: SLF001
    template_bgr = compare._decode_png_bgr(template_png_bytes)  # noqa: SLF001
    return {
        "structure_xlsx": structure_xlsx,
        "structure_pdf": structure_pdf_path,
        "structure_table": structure_table,
        "template_png_path": template_png_path,
        "template_width": template_width,
        "template_height": template_height,
        "template_bgr": template_bgr,
    }


def _page_image(
    *,
    pdf_bytes: bytes,
    page_index: int,
    dpi: int,
    apply_correction: bool,
) -> tuple[Any, dict[str, Any]]:
    correction_summary: dict[str, Any] = {"applied": False}
    corrected_pdf_bytes = pdf_bytes
    corrected_pages = None
    if apply_correction:
        try:
            corrected_pdf_bytes, correction_summary, corrected_pages = compare.correct_pdf_for_yomitoku(
                pdf_bytes=pdf_bytes,
                dpi=dpi,
                db=None,
            )
        except Exception as exc:  # noqa: BLE001
            correction_summary = {"applied": False, "error": str(exc)}
            corrected_pdf_bytes = pdf_bytes
            corrected_pages = None
    corrected_page = compare._normalize_correction_page(  # noqa: SLF001
        corrected_pages,
        pdf_bytes=corrected_pdf_bytes,
        page_index=page_index,
        dpi=dpi,
    )
    return corrected_page, correction_summary


def _position_guided_table(
    *,
    default_table: dict[str, Any],
    corrected_page,
    structure_table: dict[str, Any],
    template_bgr,
    template_png_path: Path,
    template_width: int,
    template_height: int,
    device: str,
    assignment_mode: str = "center",
) -> dict[str, Any]:
    match_bgr, ocr_bgr, _ocr_keep_lines_bgr = compare.build_images_for_match_and_ocr_from_bgr(corrected_page)
    _alignment_method, _warped_match_bgr, warped_ocr_bgr, matched_template = compare._align_to_structure_template(  # noqa: SLF001
        source_table_bbox=default_table.get("bbox") if isinstance(default_table.get("bbox"), list) else None,
        template_table_bbox=structure_table.get("bbox") if isinstance(structure_table.get("bbox"), list) else None,
        template_bgr=template_bgr,
        template_png_path=template_png_path,
        template_width=template_width,
        template_height=template_height,
        match_bgr=match_bgr,
        ocr_bgr=ocr_bgr,
    )
    if not isinstance(matched_template, dict):
        raise RuntimeError("template alignment failed")
    words = compare.ocr_image_words(warped_ocr_bgr, device=device)
    if assignment_mode == "overlap":
        return assign_words_to_structure_table_by_overlap(
            structure_table=structure_table,
            words=words,
        )
    return assign_words_to_structure_table(
        structure_table=structure_table,
        words=words,
    )


def _subgrid_patched_default_table(
    *,
    default_table: dict[str, Any],
    corrected_page,
    dpi: int,
) -> tuple[dict[str, Any], dict[str, Any]]:
    spec = infer_quantity_subgrid(default_table)
    if spec is None:
        return default_table, {"applied": False, "reason": "spec_not_found"}
    crop = crop_image_by_norm_box(corrected_page, spec.crop_box_norm, padding_px=4)
    sub_results, _ocr_pdf, _layout_pdf = compare.run_yomitoku(
        pdf_bytes=None,
        dpi=dpi,
        device="cpu",
        visualize=True,
        ignore_line_break=True,
        no_figure=True,
        figure_width=800,
        figure_dir="figures",
        page_images=[(1, crop)],
    )
    sub_table = select_primary_table(sub_results[0].tables if sub_results else [])
    if not isinstance(sub_table, dict):
        return default_table, {"applied": False, "reason": "subgrid_table_not_found"}
    improved_rows, patches = reread_suspicious_quantity_cells(
        quantity_crop_bgr=crop,
        quantity_table=sub_table,
        dpi=dpi,
    )
    full_rows = _table_rows(
        default_table,
        row_count=int(default_table.get("row_count") or 0),
        col_count=int(default_table.get("col_count") or 0),
    )
    for sub_row_index, sub_row in enumerate(improved_rows):
        target_row_index = spec.body_start_row + sub_row_index
        if target_row_index >= len(full_rows):
            break
        for sub_col_index, value in enumerate(sub_row):
            target_col_index = spec.quantity_start_col_index + sub_col_index
            if target_col_index >= len(full_rows[target_row_index]):
                break
            full_rows[target_row_index][target_col_index] = _normalize_text(value)
    return (
        _copy_table_with_rows(default_table, full_rows, source="default_subgrid_reread"),
        {
            "applied": True,
            "patch_count": len(patches),
            "body_start_row": spec.body_start_row,
            "quantity_start_col_index": spec.quantity_start_col_index,
        },
    )


def _generated_pdf_case(
    *,
    facility_id: str,
    week_sheet_name: str,
    output_dir: Path,
    dpi: int,
    device: str,
) -> dict[str, Any]:
    source_workbook_path, worksheet = _source_context(
        facility_id=facility_id,
        week_sheet_name=week_sheet_name,
    )
    structure_bundle = _prepare_structure_bundle(
        facility_id=facility_id,
        week_sheet_name=week_sheet_name,
        output_dir=output_dir,
        dpi=dpi,
        device=device,
    )
    generated_xlsx = order_form_service.build_fax_order_form_excel(
        facility_id=facility_id,
        week_sheet_name=week_sheet_name,
        output_dir=output_dir,
    )
    generated_pdf = output_dir / f"{generated_xlsx.stem}.pdf"
    render_workbook_path_to_pdf(
        generated_xlsx,
        output_path=generated_pdf,
        sheet_name=week_sheet_name,
        dpi=dpi,
    )
    pdf_bytes = generated_pdf.read_bytes()
    corrected_page, correction_summary = _page_image(
        pdf_bytes=pdf_bytes,
        page_index=1,
        dpi=dpi,
        apply_correction=False,
    )
    default_result = compare._run_default_table(  # noqa: SLF001
        page_image=corrected_page,
        page_index=1,
        dpi=dpi,
        device=device,
    )
    default_table = default_result["primary_table"]
    if not isinstance(default_table, dict):
        raise RuntimeError("default yomitoku path did not produce a primary table")

    structure_table = structure_bundle["structure_table"]
    expected_rows = _expected_rows(
        structure_table=structure_table,
        worksheet=worksheet,
    )
    canonical_rows = _canonical_rows_from_worksheet(
        worksheet,
        row_count=int(structure_table.get("row_count") or 0),
        col_count=int(structure_table.get("col_count") or 0),
    )
    position_guided_table = _position_guided_table(
        default_table=default_table,
        corrected_page=corrected_page,
        structure_table=structure_table,
        template_bgr=structure_bundle["template_bgr"],
        template_png_path=structure_bundle["template_png_path"],
        template_width=structure_bundle["template_width"],
        template_height=structure_bundle["template_height"],
        device=device,
    )
    position_guided_overlap_table = _position_guided_table(
        default_table=default_table,
        corrected_page=corrected_page,
        structure_table=structure_table,
        template_bgr=structure_bundle["template_bgr"],
        template_png_path=structure_bundle["template_png_path"],
        template_width=structure_bundle["template_width"],
        template_height=structure_bundle["template_height"],
        device=device,
        assignment_mode="overlap",
    )
    position_tail_rows = repair_menu_tail_quantity_shift(
        rows=_table_rows(
            position_guided_table,
            row_count=int(structure_table.get("row_count") or 0),
            col_count=int(structure_table.get("col_count") or 0),
        )
    )
    position_tail_table = _copy_table_with_rows(
        position_guided_table,
        position_tail_rows,
        source="position_guided_tail_shift",
    )
    position_overlap_tail_rows = repair_menu_tail_quantity_shift(
        rows=_table_rows(
            position_guided_overlap_table,
            row_count=int(structure_table.get("row_count") or 0),
            col_count=int(structure_table.get("col_count") or 0),
        )
    )
    position_overlap_tail_table = _copy_table_with_rows(
        position_guided_overlap_table,
        position_overlap_tail_rows,
        source="position_guided_overlap_tail_shift",
    )
    sequence_guided_table = build_sequence_guided_table(
        structure_table=structure_table,
        observed_table=default_table,
        canonical_rows=canonical_rows,
        header_row_count=_DEFAULT_HEADER_ROW_COUNT,
    )
    default_tail_rows = repair_menu_tail_quantity_shift(
        rows=_table_rows(
            default_table,
            row_count=int(default_table.get("row_count") or 0),
            col_count=int(default_table.get("col_count") or 0),
        )
    )
    sequence_tail_table = build_sequence_guided_table(
        structure_table=structure_table,
        observed_table=_copy_table_with_rows(default_table, default_tail_rows, source="default_tail_shift"),
        canonical_rows=canonical_rows,
        header_row_count=_DEFAULT_HEADER_ROW_COUNT,
    )
    default_subgrid_table, subgrid_meta = _subgrid_patched_default_table(
        default_table=default_table,
        corrected_page=corrected_page,
        dpi=dpi,
    )
    sequence_subgrid_table = build_sequence_guided_table(
        structure_table=structure_table,
        observed_table=default_subgrid_table,
        canonical_rows=canonical_rows,
        header_row_count=_DEFAULT_HEADER_ROW_COUNT,
    )

    row_count = int(structure_table.get("row_count") or 0)
    col_count = int(structure_table.get("col_count") or 0)
    methods = {
        "default": default_table,
        "position_guided": position_guided_table,
        "position_guided_tail_shift": position_tail_table,
        "position_guided_overlap": position_guided_overlap_table,
        "position_guided_overlap_tail_shift": position_overlap_tail_table,
        "sequence_guided": sequence_guided_table,
        "sequence_guided_tail_shift": sequence_tail_table,
        "sequence_guided_subgrid": sequence_subgrid_table,
    }
    method_metrics: dict[str, Any] = {}
    for method_name, table in methods.items():
        rows = _table_rows(table, row_count=row_count, col_count=col_count)
        method_metrics[method_name] = _metric_bundle(
            rows=rows,
            expected_rows=expected_rows,
        )
        _write_json(output_dir / f"{method_name}.json", table)
    return {
        "facility_id": facility_id,
        "week": week_sheet_name,
        "source_workbook_path": str(source_workbook_path),
        "generated_pdf": str(generated_pdf),
        "page_correction": correction_summary,
        "subgrid_meta": subgrid_meta,
        "methods": method_metrics,
    }


def _inventory_cases() -> list[dict[str, Any]]:
    inventory_path = ROOT.parent / "docs" / "fax_pdf_inventory_system_2026_02.md"
    if not inventory_path.exists():
        return []
    cases: list[dict[str, Any]] = []
    pattern = re.compile(r"^\|\s*([^|]+)\|\s*`([^`]+)`\s*\|\s*([^|]*)\|")
    for line in inventory_path.read_text(encoding="utf-8").splitlines():
        match = pattern.match(line)
        if not match:
            continue
        _order_id, pdf_name, facility_id = [token.strip() for token in match.groups()]
        if not pdf_name.endswith(".pdf") or not facility_id or facility_id == "-":
            continue
        week = _week_from_pdf_name(pdf_name)
        if not week:
            continue
        pdf_path = _find_pdf_from_inventory_name(pdf_name)
        if pdf_path is None:
            continue
        cases.append(
            {
                "facility_id": facility_id,
                "pdf_path": str(pdf_path),
                "pdf_name": pdf_name,
                "week": week,
                "page_index": 1,
            }
        )
    extras = [
        {
            "facility_id": "FAC00002",
            "pdf_path": "/Users/mmorinag/Sawa/2025.12/tmp/reupload_legacy_split/5.fax000362715_0426_0501_.pdf",
            "pdf_name": "5.fax000362715_0426_0501_.pdf",
            "week": "4月26日～4月30日",
            "page_index": 1,
        },
        {
            "facility_id": "FAC00004",
            "pdf_path": "/Users/mmorinag/Sawa/2025.12/tmp/reupload_legacy_split/18.fax000362705_0426_0501_.pdf",
            "pdf_name": "18.fax000362705_0426_0501_.pdf",
            "week": "4月26日～4月30日",
            "page_index": 1,
        },
        {
            "facility_id": "FAC00005",
            "pdf_path": "/Users/mmorinag/Sawa/2025.12/tmp/reupload_legacy_split/16.fax000364455_0426_0501_.pdf",
            "pdf_name": "16.fax000364455_0426_0501_.pdf",
            "week": "4月26日～4月30日",
            "page_index": 1,
        },
        {
            "facility_id": "FAC00006",
            "pdf_path": "/Users/mmorinag/Sawa/2025.12/tmp/reupload_legacy_split/14.fax000363668_0426-1_0501-1.pdf",
            "pdf_name": "14.fax000363668_0426-1_0501-1.pdf",
            "week": "4月26日～4月30日",
            "page_index": 1,
        },
        {
            "facility_id": "FAC00007",
            "pdf_path": "/Users/mmorinag/Sawa/2025.12/tmp/reupload_legacy_split/1.fax000363310_0426_0501_.pdf",
            "pdf_name": "1.fax000363310_0426_0501_.pdf",
            "week": "4月26日～4月30日",
            "page_index": 2,
        },
        {
            "facility_id": "FAC00012",
            "pdf_path": "/Users/mmorinag/Sawa/2025.12/tmp/reupload_legacy_split/13.fax000363999_0426_0501_.pdf",
            "pdf_name": "13.fax000363999_0426_0501_.pdf",
            "week": "4月26日～4月30日",
            "page_index": 1,
        },
        {
            "facility_id": "FAC00014",
            "pdf_path": "/Users/mmorinag/Sawa/2025.12/tmp/reupload_legacy_split/19.fax000364233_0426_0501_.pdf",
            "pdf_name": "19.fax000364233_0426_0501_.pdf",
            "week": "4月26日～4月30日",
            "page_index": 1,
        },
    ]
    deduped: dict[str, dict[str, Any]] = {}
    for case in cases + extras:
        if not Path(case["pdf_path"]).exists():
            continue
        facility_id = str(case["facility_id"])
        existing = deduped.get(facility_id)
        if existing is None or str(case["week"]).startswith("4月26日"):
            deduped[facility_id] = case
    return [deduped[key] for key in sorted(deduped)]


def _actual_case(
    *,
    facility_id: str,
    pdf_path: Path,
    week_sheet_name: str,
    page_index: int,
    output_dir: Path,
    dpi: int,
    device: str,
) -> dict[str, Any]:
    _source_workbook_path, worksheet = _source_context(
        facility_id=facility_id,
        week_sheet_name=week_sheet_name,
    )
    structure_bundle = _prepare_structure_bundle(
        facility_id=facility_id,
        week_sheet_name=week_sheet_name,
        output_dir=output_dir,
        dpi=dpi,
        device=device,
    )
    pdf_bytes = pdf_path.read_bytes()
    corrected_page, correction_summary = _page_image(
        pdf_bytes=pdf_bytes,
        page_index=page_index,
        dpi=dpi,
        apply_correction=True,
    )
    default_result = compare._run_default_table(  # noqa: SLF001
        page_image=corrected_page,
        page_index=page_index,
        dpi=dpi,
        device=device,
    )
    default_table = default_result["primary_table"]
    if not isinstance(default_table, dict):
        raise RuntimeError("default yomitoku path did not produce a primary table")
    structure_table = structure_bundle["structure_table"]
    canonical_rows = _canonical_rows_from_worksheet(
        worksheet,
        row_count=int(structure_table.get("row_count") or 0),
        col_count=int(structure_table.get("col_count") or 0),
    )
    position_guided_table = _position_guided_table(
        default_table=default_table,
        corrected_page=corrected_page,
        structure_table=structure_table,
        template_bgr=structure_bundle["template_bgr"],
        template_png_path=structure_bundle["template_png_path"],
        template_width=structure_bundle["template_width"],
        template_height=structure_bundle["template_height"],
        device=device,
    )
    position_guided_overlap_table = _position_guided_table(
        default_table=default_table,
        corrected_page=corrected_page,
        structure_table=structure_table,
        template_bgr=structure_bundle["template_bgr"],
        template_png_path=structure_bundle["template_png_path"],
        template_width=structure_bundle["template_width"],
        template_height=structure_bundle["template_height"],
        device=device,
        assignment_mode="overlap",
    )
    position_tail_table = _copy_table_with_rows(
        position_guided_table,
        repair_menu_tail_quantity_shift(
            rows=_table_rows(
                position_guided_table,
                row_count=int(structure_table.get("row_count") or 0),
                col_count=int(structure_table.get("col_count") or 0),
            )
        ),
        source="position_guided_tail_shift",
    )
    position_overlap_tail_table = _copy_table_with_rows(
        position_guided_overlap_table,
        repair_menu_tail_quantity_shift(
            rows=_table_rows(
                position_guided_overlap_table,
                row_count=int(structure_table.get("row_count") or 0),
                col_count=int(structure_table.get("col_count") or 0),
            )
        ),
        source="position_guided_overlap_tail_shift",
    )
    sequence_guided_table = build_sequence_guided_table(
        structure_table=structure_table,
        observed_table=default_table,
        canonical_rows=canonical_rows,
        header_row_count=_DEFAULT_HEADER_ROW_COUNT,
    )
    default_tail_rows = repair_menu_tail_quantity_shift(
        rows=_table_rows(
            default_table,
            row_count=int(default_table.get("row_count") or 0),
            col_count=int(default_table.get("col_count") or 0),
        )
    )
    sequence_tail_table = build_sequence_guided_table(
        structure_table=structure_table,
        observed_table=_copy_table_with_rows(default_table, default_tail_rows, source="default_tail_shift"),
        canonical_rows=canonical_rows,
        header_row_count=_DEFAULT_HEADER_ROW_COUNT,
    )
    default_subgrid_table, subgrid_meta = _subgrid_patched_default_table(
        default_table=default_table,
        corrected_page=corrected_page,
        dpi=dpi,
    )
    sequence_subgrid_table = build_sequence_guided_table(
        structure_table=structure_table,
        observed_table=default_subgrid_table,
        canonical_rows=canonical_rows,
        header_row_count=_DEFAULT_HEADER_ROW_COUNT,
    )
    row_count = int(structure_table.get("row_count") or 0)
    col_count = int(structure_table.get("col_count") or 0)
    methods = {
        "default": default_table,
        "position_guided": position_guided_table,
        "position_guided_tail_shift": position_tail_table,
        "position_guided_overlap": position_guided_overlap_table,
        "position_guided_overlap_tail_shift": position_overlap_tail_table,
        "sequence_guided": sequence_guided_table,
        "sequence_guided_tail_shift": sequence_tail_table,
        "sequence_guided_subgrid": sequence_subgrid_table,
    }
    method_metrics: dict[str, Any] = {}
    for method_name, table in methods.items():
        rows = _table_rows(table, row_count=row_count, col_count=col_count)
        method_metrics[method_name] = _metric_bundle(rows=rows, expected_rows=None)
        _write_json(output_dir / f"{method_name}.json", table)
    return {
        "facility_id": facility_id,
        "week": week_sheet_name,
        "pdf_path": str(pdf_path),
        "page_index": page_index,
        "page_correction": correction_summary,
        "subgrid_meta": subgrid_meta,
        "methods": method_metrics,
    }


def _aggregate(results: list[dict[str, Any]]) -> dict[str, Any]:
    aggregated: dict[str, dict[str, float]] = {}
    counts: dict[str, int] = {}
    for case in results:
        methods = case.get("methods") or {}
        for method_name, metrics in methods.items():
            if not isinstance(metrics, dict):
                continue
            bucket = aggregated.setdefault(method_name, {})
            counts[method_name] = counts.get(method_name, 0) + 1
            for key, value in metrics.items():
                if isinstance(value, (int, float)):
                    bucket[key] = bucket.get(key, 0.0) + float(value)
    summary: dict[str, Any] = {}
    for method_name, bucket in aggregated.items():
        total_cases = counts.get(method_name, 1)
        summary[method_name] = {
            **bucket,
            "case_count": total_cases,
        }
        for key, value in list(bucket.items()):
            summary[method_name][f"avg_{key}"] = value / total_cases
    return summary


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Evaluate reconstruction strategies across all facilities on generated forms and mapped actual faxes.",
    )
    parser.add_argument("--week", default="4月26日～4月30日")
    parser.add_argument("--dpi", type=int, default=200)
    parser.add_argument("--device", default="cpu")
    parser.add_argument("--facility-id", action="append", dest="facility_ids")
    parser.add_argument("--skip-actual", action="store_true")
    parser.add_argument("--skip-generated", action="store_true")
    parser.add_argument("--output-dir", default=None)
    args = parser.parse_args()

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_root = Path(args.output_dir or (ROOT / "tmp" / f"reconstruction_strategy_eval_{stamp}"))
    output_root.mkdir(parents=True, exist_ok=True)

    generated_results: list[dict[str, Any]] = []
    facilities = _load_facilities()
    requested_facilities = {str(item).strip() for item in (args.facility_ids or []) if str(item).strip()}
    for facility in facilities:
        facility_id = str(facility.get("facility_id") or "").strip()
        if not facility_id:
            continue
        if requested_facilities and facility_id not in requested_facilities:
            continue
        if args.skip_generated:
            continue
        case_dir = output_root / "generated" / facility_id
        case_dir.mkdir(parents=True, exist_ok=True)
        try:
            result = _generated_pdf_case(
                facility_id=facility_id,
                week_sheet_name=args.week,
                output_dir=case_dir,
                dpi=args.dpi,
                device=args.device,
            )
        except Exception as exc:  # noqa: BLE001
            result = {
                "facility_id": facility_id,
                "week": args.week,
                "error": str(exc),
            }
        generated_results.append(result)
        if "methods" in result:
            print("generated", facility_id, result["methods"]["sequence_guided"]["exact_total"], "/", result["methods"]["default"]["exact_total"])
        else:
            print("generated", facility_id, "error=", result["error"])

    actual_results: list[dict[str, Any]] = []
    for case in _inventory_cases():
        if args.skip_actual:
            continue
        if requested_facilities and str(case["facility_id"]) not in requested_facilities:
            continue
        case_dir = output_root / "actual" / str(case["facility_id"])
        case_dir.mkdir(parents=True, exist_ok=True)
        try:
            result = _actual_case(
                facility_id=str(case["facility_id"]),
                pdf_path=Path(str(case["pdf_path"])),
                week_sheet_name=str(case["week"]),
                page_index=int(case["page_index"]),
                output_dir=case_dir,
                dpi=args.dpi,
                device=args.device,
            )
        except Exception as exc:  # noqa: BLE001
            result = {
                "facility_id": str(case["facility_id"]),
                "week": str(case["week"]),
                "pdf_path": str(case["pdf_path"]),
                "error": str(exc),
            }
        actual_results.append(result)
        if "methods" in result:
            print("actual", case["facility_id"], case["pdf_name"], result["methods"]["sequence_guided"]["menu_tail_rows"])
        else:
            print("actual", case["facility_id"], case["pdf_name"], "error=", result["error"])

    generated_summary = {
        "week": args.week,
        "facility_count": len(generated_results),
        "cases": generated_results,
        "aggregate": _aggregate(generated_results),
    }
    actual_summary = {
        "case_count": len(actual_results),
        "facility_count": len({case["facility_id"] for case in actual_results}),
        "cases": actual_results,
        "aggregate": _aggregate(actual_results),
    }
    combined = {
        "generated_summary_path": str(output_root / "generated_summary.json"),
        "actual_summary_path": str(output_root / "actual_summary.json"),
        "generated_facility_count": len(generated_results),
        "actual_facility_count": len({case["facility_id"] for case in actual_results}),
        "actual_missing_facilities": sorted(
            {
                str(facility.get("facility_id"))
                for facility in facilities
                if str(facility.get("facility_id"))
                not in {case["facility_id"] for case in actual_results}
            }
        ),
    }

    _write_json(output_root / "generated_summary.json", generated_summary)
    _write_json(output_root / "actual_summary.json", actual_summary)
    _write_json(output_root / "summary.json", combined)
    print("summary=", output_root / "summary.json")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
