#!/usr/bin/env python3

from __future__ import annotations

import argparse
import json
from datetime import date, datetime
from pathlib import Path
import sys
from typing import Any

import cv2
import numpy as np
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
OCR_PIPELINE_ROOT = ROOT.parent / "ocr_pipeline"
for candidate in (str(ROOT), str(OCR_PIPELINE_ROOT)):
    if candidate not in sys.path:
        sys.path.insert(0, candidate)

from app.page_correction import correct_pdf_for_yomitoku
from app.pdf_render import render_pdf_to_page_images
from app.preprocess import build_images_for_match_and_ocr_from_bgr
from app.yomitoku_runner import ocr_image_words, run_yomitoku
from src.services import config_service, order_form_service
from src.services.fax_template_matcher import choose_template_and_warp
from src.services.pdf_render import render_pdf_to_png_bytes
from src.services.structure_guided_ocr import (
    assign_words_to_structure_table,
    build_sequence_guided_table,
    select_primary_table,
    table_rows_to_markdown,
)
from src.services.workbook_pdf_renderer import render_workbook_path_to_pdf


_DEFAULT_BODY_START_ROW = 11
_DEFAULT_BODY_END_ROW = 67
_DEFAULT_HEADER_ROW_COUNT = 2


def _write_json(path: Path, payload: Any) -> None:
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def _head_rows(rows: list[list[str]], *, limit: int = 12) -> list[list[str]]:
    return [list(row) for row in rows[:limit] if isinstance(row, list)]


def _normalize_text(value: object) -> str:
    if isinstance(value, datetime):
        return f"{value.month}/{value.day}"
    if isinstance(value, date):
        return f"{value.month}/{value.day}"
    return str(value or "").strip()


def _load_source_worksheet(*, facility_id: str, week_sheet_name: str):
    facility = config_service.get_facility_config(facility_id)
    if not facility:
        raise ValueError("facility not found")
    fax_template_id = str(order_form_service._infer_fax_template_id_from_facility(facility) or "").strip()  # noqa: SLF001
    if not fax_template_id:
        raise ValueError("facility fax_template_id not found")
    source_workbook_name = order_form_service._resolve_source_workbook_name_for_week_sheet(  # noqa: SLF001
        fax_template_id,
        week_sheet_name,
    )
    source_workbook_path = order_form_service._resolve_source_workbook_path(source_workbook_name)  # noqa: SLF001
    workbook = load_workbook(source_workbook_path, data_only=True)
    if week_sheet_name not in workbook.sheetnames:
        raise ValueError(f"week sheet not found in source workbook: {week_sheet_name}")
    return source_workbook_path, workbook[week_sheet_name]


def _canonical_rows_from_worksheet(
    worksheet,
    *,
    row_count: int,
    header_row_count: int = _DEFAULT_HEADER_ROW_COUNT,
) -> list[dict[str, Any]]:
    canonical_rows: list[dict[str, Any]] = []
    max_data_rows = max(0, row_count - header_row_count)
    available_rows = _DEFAULT_BODY_END_ROW - _DEFAULT_BODY_START_ROW + 1
    data_rows = min(max_data_rows, available_rows)
    for offset in range(data_rows):
        worksheet_row = _DEFAULT_BODY_START_ROW + offset
        canonical_rows.append(
            {
                "row_index": header_row_count + offset,
                "worksheet_row": worksheet_row,
                "date": _normalize_text(worksheet.cell(row=worksheet_row, column=1).value),
                "daypart": _normalize_text(worksheet.cell(row=worksheet_row, column=2).value),
                "aux": _normalize_text(worksheet.cell(row=worksheet_row, column=3).value),
                "menu_name": _normalize_text(worksheet.cell(row=worksheet_row, column=4).value),
            }
        )
    return canonical_rows


def _non_empty_cells(rows: list[list[str]]) -> int:
    count = 0
    for row in rows:
        if not isinstance(row, list):
            continue
        for cell in row:
            if str(cell or "").strip():
                count += 1
    return count


def _decode_png_size(png_bytes: bytes) -> tuple[int, int]:
    image = cv2.imdecode(np.frombuffer(png_bytes, np.uint8), cv2.IMREAD_COLOR)
    if image is None:
        raise RuntimeError("failed to decode rendered template PNG")
    height, width = image.shape[:2]
    return width, height


def _decode_png_bgr(png_bytes: bytes) -> np.ndarray:
    image = cv2.imdecode(np.frombuffer(png_bytes, np.uint8), cv2.IMREAD_COLOR)
    if image is None:
        raise RuntimeError("failed to decode rendered template PNG")
    return image


def _prepare_alignment_mask(image_bgr: np.ndarray, *, size: tuple[int, int]) -> np.ndarray:
    gray = cv2.cvtColor(image_bgr, cv2.COLOR_BGR2GRAY)
    if gray.shape[1] != size[0] or gray.shape[0] != size[1]:
        gray = cv2.resize(gray, size, interpolation=cv2.INTER_AREA)
    gray = cv2.GaussianBlur(gray, (5, 5), 0)
    _threshold, binary = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY | cv2.THRESH_OTSU)
    return (255 - binary).astype(np.float32) / 255.0


def _normalized_bbox_to_pixels(
    bbox: list[float] | tuple[float, float, float, float],
    *,
    width: int,
    height: int,
) -> tuple[float, float, float, float]:
    if len(bbox) != 4:
        raise ValueError("bbox must have 4 items")
    return (
        float(bbox[0]) * width,
        float(bbox[1]) * height,
        float(bbox[2]) * width,
        float(bbox[3]) * height,
    )


def _warp_by_table_bbox(
    *,
    source_table_bbox: list[float],
    template_table_bbox: list[float],
    match_bgr: np.ndarray,
    ocr_bgr: np.ndarray,
    output_size: tuple[int, int],
) -> tuple[str, np.ndarray, np.ndarray, dict[str, Any]]:
    source_height, source_width = match_bgr.shape[:2]
    target_width, target_height = output_size
    sx0, sy0, sx1, sy1 = _normalized_bbox_to_pixels(
        source_table_bbox,
        width=source_width,
        height=source_height,
    )
    dx0, dy0, dx1, dy1 = _normalized_bbox_to_pixels(
        template_table_bbox,
        width=target_width,
        height=target_height,
    )
    source_width_px = max(1.0, sx1 - sx0)
    source_height_px = max(1.0, sy1 - sy0)
    scale_x = (dx1 - dx0) / source_width_px
    scale_y = (dy1 - dy0) / source_height_px
    warp_matrix = np.array(
        [
            [scale_x, 0.0, dx0 - sx0 * scale_x],
            [0.0, scale_y, dy0 - sy0 * scale_y],
        ],
        dtype=np.float32,
    )
    warped_match = cv2.warpAffine(
        match_bgr,
        warp_matrix,
        output_size,
        flags=cv2.INTER_LINEAR,
        borderMode=cv2.BORDER_CONSTANT,
        borderValue=(255, 255, 255),
    )
    warped_ocr = cv2.warpAffine(
        ocr_bgr,
        warp_matrix,
        output_size,
        flags=cv2.INTER_LINEAR,
        borderMode=cv2.BORDER_CONSTANT,
        borderValue=(255, 255, 255),
    )
    return "table_bbox", warped_match, warped_ocr, {"warp_matrix": warp_matrix.tolist()}


def _warp_by_ecc(
    *,
    template_bgr: np.ndarray,
    match_bgr: np.ndarray,
    ocr_bgr: np.ndarray,
) -> tuple[str, np.ndarray, np.ndarray, dict[str, Any]]:
    template_height, template_width = template_bgr.shape[:2]
    output_size = (template_width, template_height)
    source_match = cv2.resize(match_bgr, output_size, interpolation=cv2.INTER_LINEAR)
    source_ocr = cv2.resize(ocr_bgr, output_size, interpolation=cv2.INTER_LINEAR)
    template_mask = _prepare_alignment_mask(template_bgr, size=output_size)
    source_mask = _prepare_alignment_mask(source_match, size=output_size)
    criteria = (
        cv2.TERM_CRITERIA_EPS | cv2.TERM_CRITERIA_COUNT,
        500,
        1e-6,
    )
    attempts: list[tuple[str, int, np.ndarray]] = [
        ("ecc_homography", cv2.MOTION_HOMOGRAPHY, np.eye(3, 3, dtype=np.float32)),
        ("ecc_affine", cv2.MOTION_AFFINE, np.eye(2, 3, dtype=np.float32)),
    ]
    errors: list[str] = []
    for method_name, motion_model, warp_seed in attempts:
        try:
            correlation, warp_matrix = cv2.findTransformECC(
                template_mask,
                source_mask,
                warp_seed,
                motion_model,
                criteria,
                None,
                5,
            )
            if motion_model == cv2.MOTION_HOMOGRAPHY:
                warped_match = cv2.warpPerspective(
                    source_match,
                    warp_matrix,
                    output_size,
                    flags=cv2.INTER_LINEAR | cv2.WARP_INVERSE_MAP,
                    borderMode=cv2.BORDER_CONSTANT,
                    borderValue=(255, 255, 255),
                )
                warped_ocr = cv2.warpPerspective(
                    source_ocr,
                    warp_matrix,
                    output_size,
                    flags=cv2.INTER_LINEAR | cv2.WARP_INVERSE_MAP,
                    borderMode=cv2.BORDER_CONSTANT,
                    borderValue=(255, 255, 255),
                )
            else:
                warped_match = cv2.warpAffine(
                    source_match,
                    warp_matrix,
                    output_size,
                    flags=cv2.INTER_LINEAR | cv2.WARP_INVERSE_MAP,
                    borderMode=cv2.BORDER_CONSTANT,
                    borderValue=(255, 255, 255),
                )
                warped_ocr = cv2.warpAffine(
                    source_ocr,
                    warp_matrix,
                    output_size,
                    flags=cv2.INTER_LINEAR | cv2.WARP_INVERSE_MAP,
                    borderMode=cv2.BORDER_CONSTANT,
                    borderValue=(255, 255, 255),
                )
            return method_name, warped_match, warped_ocr, {"correlation": float(correlation)}
        except cv2.error as exc:
            errors.append(f"{method_name}: {exc}")
    raise RuntimeError("template alignment failed: " + " | ".join(errors))


def _align_to_structure_template(
    *,
    source_table_bbox: list[float] | None,
    template_table_bbox: list[float] | None,
    template_bgr: np.ndarray,
    template_png_path: Path,
    template_width: int,
    template_height: int,
    match_bgr: np.ndarray,
    ocr_bgr: np.ndarray,
) -> tuple[str, np.ndarray, np.ndarray, dict[str, Any]]:
    if isinstance(source_table_bbox, list) and len(source_table_bbox) == 4 and isinstance(template_table_bbox, list) and len(template_table_bbox) == 4:
        return _warp_by_table_bbox(
            source_table_bbox=source_table_bbox,
            template_table_bbox=template_table_bbox,
            match_bgr=match_bgr,
            ocr_bgr=ocr_bgr,
            output_size=(template_width, template_height),
        )
    orb_attempts = [
        {"orb_nfeatures": 2000, "min_matches": 25, "min_inlier_ratio": 0.15, "label": "orb_strict"},
        {"orb_nfeatures": 4000, "min_matches": 8, "min_inlier_ratio": 0.05, "label": "orb_relaxed"},
    ]
    for attempt in orb_attempts:
        try:
            matched_template_id, warped_match_bgr, warped_ocr_bgr, matched_template = choose_template_and_warp(
                {
                    "structure_only_template": {
                        "template_image_gcs_uri": f"file://{template_png_path}",
                        "warp": {"output_size": [template_width, template_height]},
                        "match": {
                            "orb_nfeatures": attempt["orb_nfeatures"],
                            "min_matches": attempt["min_matches"],
                            "min_inlier_ratio": attempt["min_inlier_ratio"],
                        },
                    }
                },
                match_bgr,
                ocr_bgr,
            )
            if matched_template_id and isinstance(matched_template, dict):
                return attempt["label"], warped_match_bgr, warped_ocr_bgr, matched_template
        except Exception:  # noqa: BLE001
            continue
    return _warp_by_ecc(template_bgr=template_bgr, match_bgr=match_bgr, ocr_bgr=ocr_bgr)


def _page_image_from_pdf(
    pdf_bytes: bytes,
    *,
    page_index: int,
    dpi: int,
) -> np.ndarray:
    pages = render_pdf_to_page_images(pdf_bytes, dpi)
    for current_page_index, image in pages:
        if int(current_page_index) == int(page_index):
            return image
    raise ValueError(f"page {page_index} not found in PDF")


def _normalize_correction_page(
    corrected_pages: list[tuple[int, np.ndarray]] | None,
    *,
    pdf_bytes: bytes,
    page_index: int,
    dpi: int,
) -> np.ndarray:
    if corrected_pages:
        for current_page_index, image in corrected_pages:
            if int(current_page_index) == int(page_index):
                return image
    return _page_image_from_pdf(pdf_bytes, page_index=page_index, dpi=dpi)


def _run_default_table(
    *,
    page_image: np.ndarray,
    page_index: int,
    dpi: int,
    device: str,
) -> dict[str, Any]:
    page_results, ocr_pdf, layout_pdf = run_yomitoku(
        pdf_bytes=None,
        dpi=dpi,
        device=device,
        visualize=True,
        ignore_line_break=True,
        no_figure=True,
        figure_width=800,
        figure_dir="figures",
        page_images=[(page_index, page_image)],
    )
    table = select_primary_table(page_results[0].tables if page_results else [])
    return {
        "page_results": page_results,
        "primary_table": table,
        "ocr_pdf": ocr_pdf,
        "layout_pdf": layout_pdf,
    }


def _run_structure_template(
    *,
    structure_pdf_bytes: bytes,
    dpi: int,
    device: str,
) -> dict[str, Any]:
    page_results, ocr_pdf, layout_pdf = run_yomitoku(
        pdf_bytes=structure_pdf_bytes,
        dpi=dpi,
        device=device,
        visualize=True,
        ignore_line_break=True,
        no_figure=True,
        figure_width=800,
        figure_dir="figures",
    )
    table = select_primary_table(page_results[0].tables if page_results else [])
    return {
        "page_results": page_results,
        "primary_table": table,
        "ocr_pdf": ocr_pdf,
        "layout_pdf": layout_pdf,
    }


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Compare default yomitoku output vs structure-guided assignment using a structure-only generated order-form PDF.",
    )
    parser.add_argument("--facility-id", required=True)
    parser.add_argument("--fax-pdf", required=True)
    parser.add_argument("--week", default=order_form_service._DEFAULT_WEEK_SHEET)  # noqa: SLF001
    parser.add_argument("--page-index", type=int, default=1)
    parser.add_argument("--dpi", type=int, default=200)
    parser.add_argument("--device", default="cpu")
    parser.add_argument("--output-dir", default=None)
    args = parser.parse_args()

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_dir = Path(args.output_dir or (ROOT / "tmp" / f"structure_guided_compare_{stamp}"))
    output_dir.mkdir(parents=True, exist_ok=True)

    fax_pdf_path = Path(args.fax_pdf)
    fax_pdf_bytes = fax_pdf_path.read_bytes()

    structure_xlsx = order_form_service.build_fax_structure_only_excel(
        facility_id=args.facility_id,
        week_sheet_name=args.week,
        output_dir=output_dir,
    )
    structure_pdf_path = output_dir / f"{structure_xlsx.stem}.pdf"
    render_workbook_path_to_pdf(
        structure_xlsx,
        output_path=structure_pdf_path,
        sheet_name=args.week,
        dpi=args.dpi,
    )
    structure_pdf_bytes = structure_pdf_path.read_bytes()

    correction_summary: dict[str, Any] = {"applied": False}
    corrected_pdf_bytes = fax_pdf_bytes
    corrected_pages = None
    try:
        corrected_pdf_bytes, correction_summary, corrected_pages = correct_pdf_for_yomitoku(
            pdf_bytes=fax_pdf_bytes,
            dpi=args.dpi,
            db=None,
        )
    except Exception as exc:  # noqa: BLE001
        correction_summary = {"applied": False, "error": str(exc)}
        corrected_pdf_bytes = fax_pdf_bytes
        corrected_pages = None

    corrected_page = _normalize_correction_page(
        corrected_pages,
        pdf_bytes=corrected_pdf_bytes,
        page_index=args.page_index,
        dpi=args.dpi,
    )

    default_result = _run_default_table(
        page_image=corrected_page,
        page_index=args.page_index,
        dpi=args.dpi,
        device=args.device,
    )
    default_table = default_result["primary_table"]
    if not isinstance(default_table, dict):
        raise RuntimeError("default yomitoku path did not produce a primary table")

    structure_result = _run_structure_template(
        structure_pdf_bytes=structure_pdf_bytes,
        dpi=args.dpi,
        device=args.device,
    )
    structure_table = structure_result["primary_table"]
    if not isinstance(structure_table, dict):
        raise RuntimeError("structure-only template PDF did not produce a primary table")

    template_png_bytes = render_pdf_to_png_bytes(structure_pdf_bytes, dpi=args.dpi, page=1)
    template_png_path = output_dir / "structure_template_page1.png"
    template_png_path.write_bytes(template_png_bytes)
    template_width, template_height = _decode_png_size(template_png_bytes)
    template_bgr = _decode_png_bgr(template_png_bytes)

    match_bgr, ocr_bgr, _ocr_keep_lines_bgr = build_images_for_match_and_ocr_from_bgr(corrected_page)
    alignment_method, warped_match_bgr, warped_ocr_bgr, matched_template = _align_to_structure_template(
        source_table_bbox=default_table.get("bbox") if isinstance(default_table.get("bbox"), list) else None,
        template_table_bbox=structure_table.get("bbox") if isinstance(structure_table.get("bbox"), list) else None,
        template_bgr=template_bgr,
        template_png_path=template_png_path,
        template_width=template_width,
        template_height=template_height,
        match_bgr=match_bgr,
        ocr_bgr=ocr_bgr,
    )
    if not alignment_method or not isinstance(matched_template, dict):
        raise RuntimeError("structure-only template match failed")

    warped_match_path = output_dir / "warped_match.png"
    warped_ocr_path = output_dir / "warped_ocr.png"
    cv2.imwrite(str(warped_match_path), warped_match_bgr)
    cv2.imwrite(str(warped_ocr_path), warped_ocr_bgr)

    structure_words = ocr_image_words(warped_ocr_bgr, device=args.device)
    structure_guided_table = assign_words_to_structure_table(
        structure_table=structure_table,
        words=structure_words,
    )

    if default_result["ocr_pdf"]:
        (output_dir / "default_ocr_overlay.pdf").write_bytes(default_result["ocr_pdf"])
    if default_result["layout_pdf"]:
        (output_dir / "default_layout_overlay.pdf").write_bytes(default_result["layout_pdf"])
    if structure_result["ocr_pdf"]:
        (output_dir / "structure_template_ocr_overlay.pdf").write_bytes(structure_result["ocr_pdf"])
    if structure_result["layout_pdf"]:
        (output_dir / "structure_template_layout_overlay.pdf").write_bytes(structure_result["layout_pdf"])

    _source_workbook_path, worksheet = _load_source_worksheet(
        facility_id=args.facility_id,
        week_sheet_name=args.week,
    )
    canonical_rows = _canonical_rows_from_worksheet(
        worksheet,
        row_count=int(structure_table.get("row_count") or 0),
    )

    default_rows = list(default_table.get("rows") or [])
    position_guided_table = structure_guided_table
    structure_guided_table = build_sequence_guided_table(
        structure_table=structure_table,
        observed_table=default_table,
        canonical_rows=canonical_rows,
        header_row_count=_DEFAULT_HEADER_ROW_COUNT,
    )
    guided_rows = list(structure_guided_table.get("rows") or [])
    position_guided_rows = list(position_guided_table.get("rows") or [])
    summary = {
        "facility_id": args.facility_id,
        "fax_pdf": str(fax_pdf_path),
        "week": args.week,
        "page_index": args.page_index,
        "dpi": args.dpi,
        "device": args.device,
        "page_correction": correction_summary,
        "alignment_method": alignment_method,
        "artifacts": {
            "structure_xlsx": str(structure_xlsx),
            "structure_pdf": str(structure_pdf_path),
            "template_png": str(template_png_path),
            "warped_match_png": str(warped_match_path),
            "warped_ocr_png": str(warped_ocr_path),
        },
        "default": {
            "row_count": int(default_table.get("row_count") or 0),
            "col_count": int(default_table.get("col_count") or 0),
            "non_empty_cells": _non_empty_cells(default_rows),
            "head_rows": _head_rows(default_rows),
            "markdown": table_rows_to_markdown(default_rows),
        },
        "position_guided": {
            "row_count": int(position_guided_table.get("row_count") or 0),
            "col_count": int(position_guided_table.get("col_count") or 0),
            "non_empty_cells": _non_empty_cells(position_guided_rows),
            "head_rows": _head_rows(position_guided_rows),
            "markdown": table_rows_to_markdown(position_guided_rows),
        },
        "structure_template": {
            "row_count": int(structure_table.get("row_count") or 0),
            "col_count": int(structure_table.get("col_count") or 0),
            "head_rows": _head_rows(list(structure_table.get("rows") or [])),
        },
        "structure_guided": {
            "row_count": int(structure_guided_table.get("row_count") or 0),
            "col_count": int(structure_guided_table.get("col_count") or 0),
            "non_empty_cells": _non_empty_cells(guided_rows),
            "head_rows": _head_rows(guided_rows),
            "markdown": table_rows_to_markdown(guided_rows),
        },
    }

    _write_json(output_dir / "summary.json", summary)
    _write_json(output_dir / "default_table.json", default_table)
    _write_json(output_dir / "position_guided_table.json", position_guided_table)
    _write_json(output_dir / "structure_template_table.json", structure_table)
    _write_json(output_dir / "structure_guided_table.json", structure_guided_table)

    print(f"output_dir={output_dir}")
    print(f"summary={output_dir / 'summary.json'}")
    print(f"default_non_empty_cells={summary['default']['non_empty_cells']}")
    print(f"structure_guided_non_empty_cells={summary['structure_guided']['non_empty_cells']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
