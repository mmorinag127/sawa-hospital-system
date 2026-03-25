#!/usr/bin/env python3

from __future__ import annotations

from datetime import datetime
from pathlib import Path
import os
import re
import sys

from openpyxl.cell.cell import MergedCell
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))
os.chdir(ROOT)

from src.services import config_service, order_form_service  # noqa: E402


DEFAULT_FACILITY_ID = "FAC00002"
DEFAULT_WEEK = "3月22日～3月28日"
_BODY_START_ROW = 11
_BODY_END_ROW = 67


def _sanitize_fragment(value: str) -> str:
    safe = re.sub(r"[^0-9A-Za-z_-]+", "_", str(value or "").strip())
    return safe.strip("_") or "value"


def _build_output_dir() -> Path:
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = Path("/tmp/order-form-layered-example") / stamp
    path.mkdir(parents=True, exist_ok=True)
    return path


def _load_week_workbook(*, facility_id: str, week_sheet_name: str):
    facility = config_service.get_facility_config(facility_id)
    if not facility:
        raise ValueError(f"facility not found: {facility_id}")
    fax_template_id = str(facility.get("fax_template_id") or "").strip()
    spec = order_form_service._resolve_fax_family_spec(fax_template_id)
    source_path = order_form_service._FAX_SOURCE_TEMPLATE_DIR / str(spec["source_workbook"])
    workbook = load_workbook(source_path)
    order_form_service._keep_only_target_sheet(workbook, week_sheet_name)
    worksheet = workbook[week_sheet_name]
    return workbook, worksheet, facility, fax_template_id, spec


def _clear_menu_body(worksheet) -> None:
    for row in range(_BODY_START_ROW, _BODY_END_ROW + 1):
        for col in range(1, 5):
            _set_cell_value(worksheet, row=row, col=col, value=None)


def _set_deadline_placeholder(worksheet) -> None:
    worksheet["G3"] = "締切日記入欄"


def _apply_baseline_header_placeholders(worksheet) -> None:
    worksheet["A3"] = "施設名記入欄"
    worksheet["G4"] = "※週次メニューが入る前の共通土台です。"
    worksheet["G5"] = "※食種欄は施設設定で差し替えます。"
    worksheet["G6"] = "※訂正は右側欄へ追記してください。"
    worksheet["E7"] = "食種"
    worksheet["G7"] = "禁食"
    worksheet["I7"] = "訂正①"
    worksheet["J7"] = "訂正②"
    worksheet["G9"] = "禁食①"
    worksheet["H9"] = "禁食②"


def _apply_facility_header_values(target_ws, source_ws, *, facility_name: str) -> None:
    target_ws["A3"] = facility_name
    for row in range(4, 11):
        for col in range(1, 13):
            _set_cell_value(
                target_ws,
                row=row,
                col=col,
                value=source_ws.cell(row=row, column=col).value,
            )
    target_ws["A3"] = facility_name
    _set_deadline_placeholder(target_ws)


def _apply_completed_order_values(target_ws, source_ws, *, facility_name: str) -> None:
    target_ws["A3"] = facility_name
    target_ws["G3"] = source_ws["G3"].value
    for row in range(4, _BODY_END_ROW + 1):
        for col in range(1, 13):
            _set_cell_value(
                target_ws,
                row=row,
                col=col,
                value=source_ws.cell(row=row, column=col).value,
            )
    target_ws["A3"] = facility_name


def _set_cell_value(worksheet, *, row: int, col: int, value) -> None:
    cell = worksheet.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        return
    cell.value = value


def _decorate_workbook(
    workbook,
    worksheet,
    *,
    title: str,
    facility_id: str,
    facility_name: str,
    fax_template_id: str,
    family_label: str,
    mode: str,
    source_workbook_name: str,
    week_sheet_name: str,
) -> None:
    worksheet.title = title
    order_form_service._write_facility_name(worksheet, facility_name)
    order_form_service._apply_fax_metadata_header(
        worksheet,
        fax_template_id=fax_template_id,
        facility_id=facility_id,
        facility_name=facility_name,
        week_sheet_name=week_sheet_name,
        family_label=family_label,
    )
    order_form_service._apply_fax_markers(worksheet)
    order_form_service._apply_bottom_instruction_strip(
        worksheet,
        fax_template_id=fax_template_id,
        base_label=mode,
    )
    order_form_service._extend_print_area(worksheet, bottom_row=69)
    order_form_service._append_hidden_metadata_sheet(
        workbook,
        source_workbook_name=source_workbook_name,
        facility_id=facility_id,
        facility_name=facility_name,
        fax_template_id=fax_template_id,
        family_label=family_label,
        week_sheet_name=week_sheet_name,
        base_label=mode,
    )


def main() -> int:
    output_dir = _build_output_dir()
    source_wb, source_ws, facility, fax_template_id, spec = _load_week_workbook(
        facility_id=DEFAULT_FACILITY_ID,
        week_sheet_name=DEFAULT_WEEK,
    )
    facility_name = str(facility.get("facility_name") or facility.get("name") or DEFAULT_FACILITY_ID)
    family_label = str(spec["family_label"])
    source_workbook_name = str(spec["source_workbook"])

    baseline_wb, baseline_ws, _, _, _ = _load_week_workbook(
        facility_id=DEFAULT_FACILITY_ID,
        week_sheet_name=DEFAULT_WEEK,
    )
    _clear_menu_body(baseline_ws)
    _set_deadline_placeholder(baseline_ws)
    _apply_baseline_header_placeholders(baseline_ws)
    _decorate_workbook(
        baseline_wb,
        baseline_ws,
        title="ベースライン",
        facility_id="BASE",
        facility_name="施設名記入欄",
        fax_template_id=fax_template_id,
        family_label=family_label,
        mode="baseline_template",
        source_workbook_name=source_workbook_name,
        week_sheet_name=DEFAULT_WEEK,
    )

    facility_wb, facility_ws, _, _, _ = _load_week_workbook(
        facility_id=DEFAULT_FACILITY_ID,
        week_sheet_name=DEFAULT_WEEK,
    )
    _clear_menu_body(facility_ws)
    _set_deadline_placeholder(facility_ws)
    _apply_facility_header_values(facility_ws, source_ws, facility_name=facility_name)
    _decorate_workbook(
        facility_wb,
        facility_ws,
        title="施設テンプレート",
        facility_id=DEFAULT_FACILITY_ID,
        facility_name=facility_name,
        fax_template_id=fax_template_id,
        family_label=family_label,
        mode="facility_template",
        source_workbook_name=source_workbook_name,
        week_sheet_name=DEFAULT_WEEK,
    )

    completed_wb, completed_ws, _, _, _ = _load_week_workbook(
        facility_id=DEFAULT_FACILITY_ID,
        week_sheet_name=DEFAULT_WEEK,
    )
    _apply_completed_order_values(completed_ws, source_ws, facility_name=facility_name)
    _decorate_workbook(
        completed_wb,
        completed_ws,
        title=DEFAULT_WEEK,
        facility_id=DEFAULT_FACILITY_ID,
        facility_name=facility_name,
        fax_template_id=fax_template_id,
        family_label=family_label,
        mode="completed_order_form",
        source_workbook_name=source_workbook_name,
        week_sheet_name=DEFAULT_WEEK,
    )

    baseline_path = output_dir / (
        f"1_baseline_{_sanitize_fragment(fax_template_id)}_{_sanitize_fragment(DEFAULT_WEEK)}.xlsx"
    )
    facility_path = output_dir / (
        f"2_facility_template_{DEFAULT_FACILITY_ID}_{_sanitize_fragment(facility_name)}.xlsx"
    )
    completed_path = output_dir / (
        f"3_completed_order_form_{DEFAULT_FACILITY_ID}_{_sanitize_fragment(DEFAULT_WEEK)}.xlsx"
    )
    baseline_wb.save(baseline_path)
    facility_wb.save(facility_path)
    completed_wb.save(completed_path)

    print(f"output_dir={output_dir}")
    print(f"baseline={baseline_path}")
    print(f"facility_template={facility_path}")
    print(f"completed={completed_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
