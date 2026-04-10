#!/usr/bin/env python3

from __future__ import annotations

import argparse
from datetime import date, datetime, timedelta
from pathlib import Path
import os
import sys

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter, range_boundaries


ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))
os.chdir(ROOT)

from src.services import config_service, order_form_service  # noqa: E402


DEFAULT_FACILITY_ID = "FAC00016"
DEFAULT_WEEK = "3月22日～3月28日"
BODY_START_ROW = 11
BODY_END_ROW = 67


def _build_output_dir(output_dir: str | None) -> Path:
    if output_dir:
        path = Path(output_dir)
    else:
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        path = Path("/tmp/order-form-layer-demo") / stamp
    path.mkdir(parents=True, exist_ok=True)
    return path


def _resolve_demo_context(facility_id: str) -> tuple[dict, dict]:
    facility = config_service.get_facility_config(facility_id)
    if not facility:
        raise ValueError(f"facility not found: {facility_id}")
    fax_template_id = str(facility.get("fax_template_id") or "").strip()
    if not fax_template_id:
        raise ValueError(f"facility fax_template_id not found: {facility_id}")
    spec = order_form_service._resolve_fax_family_spec(fax_template_id)
    return facility, spec


def _load_source_week(source_workbook_name: str, week_sheet_name: str):
    source_path = order_form_service._FAX_SOURCE_TEMPLATE_DIR / source_workbook_name
    workbook = load_workbook(source_path)
    order_form_service._keep_only_target_sheet(workbook, week_sheet_name)
    worksheet = workbook[week_sheet_name]
    return workbook, worksheet


def _insert_rows_preserving_merges(worksheet, idx: int, amount: int = 1) -> None:
    original_ranges = [str(rng) for rng in worksheet.merged_cells.ranges]
    for rng in original_ranges:
        worksheet.unmerge_cells(rng)
    worksheet.insert_rows(idx, amount)
    for rng in original_ranges:
        min_col, min_row, max_col, max_row = range_boundaries(rng)
        if min_row >= idx:
            min_row += amount
            max_row += amount
        elif min_row < idx <= max_row:
            max_row += amount
        shifted = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
        worksheet.merge_cells(shifted)


def _insert_cols_preserving_merges(worksheet, idx: int, amount: int = 1) -> None:
    original_ranges = [str(rng) for rng in worksheet.merged_cells.ranges]
    for rng in original_ranges:
        worksheet.unmerge_cells(rng)
    worksheet.insert_cols(idx, amount)
    for rng in original_ranges:
        min_col, min_row, max_col, max_row = range_boundaries(rng)
        if min_col >= idx:
            min_col += amount
            max_col += amount
        elif min_col < idx <= max_col:
            max_col += amount
        shifted = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
        worksheet.merge_cells(shifted)


def _shift_image_anchors(worksheet, *, delta_cols: int = 0, delta_rows: int = 0) -> None:
    for image in getattr(worksheet, "_images", []):
        anchor = getattr(image, "anchor", None)
        start = getattr(anchor, "_from", None)
        end = getattr(anchor, "_to", None) or getattr(anchor, "to", None)
        if start is not None:
            start.col += delta_cols
            start.row += delta_rows
        if end is not None:
            end.col += delta_cols
            end.row += delta_rows


def _fill_marker_cells(worksheet, *cell_refs: str) -> None:
    for cell_ref in cell_refs:
        cell = worksheet[cell_ref]
        cell.fill = order_form_service._MARKER_FILL
        cell.border = order_form_service._THIN_BORDER


def _append_facility_profile_sheet(workbook, *, facility: dict, family_label: str) -> None:
    if "施設情報" in workbook.sheetnames:
        del workbook["施設情報"]
    worksheet = workbook.create_sheet("施設情報")
    worksheet["A1"] = "施設設定サマリー"
    worksheet["A1"].font = Font(name="Meiryo", size=14, bold=True)

    rows = [
        ("facility_id", str(facility.get("facility_id") or "")),
        ("facility_name", str(facility.get("facility_name") or "")),
        ("fax_template_id", str(facility.get("fax_template_id") or "")),
        ("fax_template_ids", ", ".join(str(item) for item in (facility.get("fax_template_ids") or []))),
        ("family_label", family_label),
        ("aliases", ", ".join(str(item) for item in (facility.get("aliases") or []))),
    ]
    row_index = 3
    for key, value in rows:
        worksheet[f"A{row_index}"] = key
        worksheet[f"A{row_index}"].font = Font(name="Meiryo", size=10, bold=True)
        worksheet[f"B{row_index}"] = value
        worksheet[f"B{row_index}"].alignment = Alignment(vertical="top", wrap_text=True)
        row_index += 1

    row_index += 1
    worksheet[f"A{row_index}"] = "fax_template_override.columns"
    worksheet[f"A{row_index}"].font = Font(name="Meiryo", size=11, bold=True)
    row_index += 1
    headers = ["index", "role", "header", "diet_type", "area_id"]
    for col_index, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=row_index, column=col_index, value=header)
        cell.font = Font(name="Meiryo", size=10, bold=True)
        cell.fill = order_form_service._META_FILL
        cell.border = order_form_service._THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")
    row_index += 1

    columns = (facility.get("fax_template_override") or {}).get("columns") or []
    for column in columns:
        worksheet.cell(row=row_index, column=1, value=column.get("index"))
        worksheet.cell(row=row_index, column=2, value=column.get("role"))
        worksheet.cell(row=row_index, column=3, value=column.get("header"))
        worksheet.cell(row=row_index, column=4, value=column.get("diet_type"))
        worksheet.cell(row=row_index, column=5, value=column.get("area_id"))
        row_index += 1

    worksheet.column_dimensions["A"].width = 20
    worksheet.column_dimensions["B"].width = 42
    worksheet.column_dimensions["C"].width = 18
    worksheet.column_dimensions["D"].width = 18
    worksheet.column_dimensions["E"].width = 14
    worksheet.freeze_panes = "A10"


def _set_facility_name_box(worksheet, *, facility_name: str) -> None:
    name_length = len(facility_name.strip())
    if name_length <= 10:
        font_size = 32
    elif name_length <= 16:
        font_size = 28
    elif name_length <= 24:
        font_size = 24
    elif name_length <= 32:
        font_size = 20
    else:
        font_size = 16
    worksheet["B4"] = facility_name
    worksheet["B4"].font = Font(name="Meiryo", size=font_size, bold=True)
    worksheet["B4"].alignment = Alignment(
        horizontal="left",
        vertical="center",
        shrink_to_fit=True,
        indent=1,
    )


def _clear_body_values(worksheet) -> None:
    for row in worksheet.iter_rows(min_row=BODY_START_ROW, max_row=BODY_END_ROW, min_col=1, max_col=4):
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            cell.value = None


def _set_deadline_placeholder(worksheet, text: str) -> None:
    worksheet["G3"] = text


def _extract_first_menu_date(worksheet) -> date | None:
    for row in range(BODY_START_ROW, worksheet.max_row + 1):
        value = worksheet[f"A{row}"].value
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
    return None


def _compute_deadline_text_for_week(worksheet) -> str:
    first_menu_date = _extract_first_menu_date(worksheet)
    if first_menu_date is None:
        return "締切日 [週次生成時に自動計算]"
    # Weekly order deadlines in the sample books are keyed to the containing
    # Sunday-start block, even for partial weeks at month boundaries.
    sunday_anchor = first_menu_date - timedelta(days=(first_menu_date.weekday() + 1) % 7)
    deadline = sunday_anchor - timedelta(days=16)
    return f"締切日{deadline.month}月{deadline.day}日まで"


def _set_generic_notes(worksheet) -> None:
    worksheet["G4"] = "※ご希望メニューに食数をご記入ください。"
    worksheet["G5"] = "※制約食は対象欄に食数をご記入ください。"
    worksheet["G6"] = "※訂正は右欄へ追記してください。"


def _set_generic_headers(worksheet) -> None:
    worksheet["E7"] = "食種A"
    worksheet["F7"] = "食種B"
    worksheet["G7"] = "制約食"
    worksheet["G9"] = "制約A"
    worksheet["H9"] = "制約B"
    worksheet["I7"] = "訂正①"
    worksheet["J7"] = "訂正②"
    worksheet["K7"] = "備考欄"


def _apply_marker_frame(
    workbook,
    worksheet,
    *,
    facility_name: str,
    facility_id: str,
    fax_template_id: str,
    family_label: str,
    week_label: str,
    mode_label: str,
) -> None:
    # Create dedicated blank gutters around the printed table so OCR markers stay
    # out of user-write areas while still surviving fax distortion.
    original_widths = {
        col_idx: worksheet.column_dimensions[get_column_letter(col_idx)].width
        for col_idx in range(1, worksheet.max_column + 1)
    }
    _insert_rows_preserving_merges(worksheet, 7, 1)
    _insert_cols_preserving_merges(worksheet, 1, 1)
    _insert_cols_preserving_merges(worksheet, worksheet.max_column + 1, 1)
    _shift_image_anchors(worksheet, delta_cols=1)

    worksheet.column_dimensions["A"].width = 2.8
    worksheet.column_dimensions["N"].width = 2.8
    for original_col_idx, width in original_widths.items():
        worksheet.column_dimensions[get_column_letter(original_col_idx + 1)].width = width
    worksheet.row_dimensions[1].height = 18
    worksheet.row_dimensions[7].height = 8
    worksheet.row_dimensions[69].height = 8
    worksheet.row_dimensions[70].height = 18

    # Give each corner a distinct signature so 180-degree rotation and flips are
    # detectable before finer gutter matching runs.
    _fill_marker_cells(worksheet, "A1", "A2")
    _fill_marker_cells(worksheet, "N1", "N2", "N3")
    _fill_marker_cells(worksheet, "A69", "B69", "A70")
    _fill_marker_cells(worksheet, "M69", "N68", "N69", "N70")

    # Top/bottom marker gutters capture horizontal stretch and curvature. The
    # patterns intentionally differ so upside-down scans remain detectable.
    _fill_marker_cells(worksheet, "B7", "E7", "H7", "K7", "M7")
    _fill_marker_cells(worksheet, "C69", "F69", "I69", "L69")

    # Side gutters add anchors for vertical drift and center wobble.
    _fill_marker_cells(worksheet, "A10", "A24", "A38", "A52", "A66")
    _fill_marker_cells(worksheet, "N14", "N30", "N46", "N62")

    # Top metadata remains above the marker gutter.
    worksheet.merge_cells("B1:M1")
    header_cell = worksheet["B1"]
    header_cell.value = (
        f"TEMPLATE={fax_template_id} | FAMILY={family_label} | "
        f"FACILITY={facility_id}:{facility_name} | WEEK={week_label} | PAGE=1/1"
    )
    header_cell.font = Font(name="Meiryo", size=8, bold=True)
    header_cell.alignment = Alignment(horizontal="center", vertical="center")
    header_cell.fill = order_form_service._META_FILL
    header_cell.border = order_form_service._THIN_BORDER

    worksheet.merge_cells("B70:M70")
    info_cell = worksheet["B70"]
    info_cell.value = (
        "OCR補助: 枠内に濃く記入 / 訂正は右側へ追記 / "
        f"template={fax_template_id} / mode={mode_label}"
    )
    info_cell.font = Font(name="Meiryo", size=8, bold=True)
    info_cell.alignment = Alignment(horizontal="center", vertical="center")
    info_cell.fill = order_form_service._META_FILL
    info_cell.border = order_form_service._THIN_BORDER

    worksheet.print_area = "A1:N70"
    order_form_service._append_hidden_metadata_sheet(
        workbook,
        source_workbook_name="layer_demo",
        facility_id=facility_id,
        facility_name=facility_name,
        fax_template_id=fax_template_id,
        family_label=family_label,
        week_sheet_name=week_label,
        base_label=mode_label,
    )


def _build_baseline_template(
    *,
    source_workbook_name: str,
    week_sheet_name: str,
    fax_template_id: str,
    family_label: str,
    output_dir: Path,
) -> Path:
    workbook, worksheet = _load_source_week(source_workbook_name, week_sheet_name)
    _clear_body_values(worksheet)
    _set_deadline_placeholder(worksheet, "締切日 [週次生成時に自動計算]")
    _set_generic_notes(worksheet)
    _set_generic_headers(worksheet)
    _apply_marker_frame(
        workbook,
        worksheet,
        facility_name="施設名",
        facility_id="BASELINE",
        fax_template_id=fax_template_id,
        family_label=family_label,
        week_label="UNSET",
        mode_label="baseline",
    )
    _set_facility_name_box(worksheet, facility_name="施設名")
    worksheet.title = "ベースライン"
    output_path = output_dir / "1_baseline_template.xlsx"
    workbook.save(output_path)
    return output_path


def _build_facility_template(
    *,
    facility: dict,
    source_workbook_name: str,
    week_sheet_name: str,
    fax_template_id: str,
    family_label: str,
    output_dir: Path,
) -> Path:
    workbook, worksheet = _load_source_week(source_workbook_name, week_sheet_name)
    _clear_body_values(worksheet)
    _set_deadline_placeholder(worksheet, "締切日 [週次生成時に自動計算]")
    _apply_marker_frame(
        workbook,
        worksheet,
        facility_name=str(facility.get("facility_name") or ""),
        facility_id=str(facility.get("facility_id") or ""),
        fax_template_id=fax_template_id,
        family_label=family_label,
        week_label="UNSET",
        mode_label="facility_template",
    )
    _set_facility_name_box(worksheet, facility_name=str(facility.get("facility_name") or ""))
    worksheet.title = "施設テンプレート"
    _append_facility_profile_sheet(workbook, facility=facility, family_label=family_label)
    output_path = output_dir / f"2_facility_template_{facility['facility_id']}.xlsx"
    workbook.save(output_path)
    return output_path


def _build_completed_order_form(
    *,
    facility: dict,
    source_workbook_name: str,
    week_sheet_name: str,
    fax_template_id: str,
    family_label: str,
    output_dir: Path,
) -> Path:
    workbook, worksheet = _load_source_week(source_workbook_name, week_sheet_name)
    worksheet["G3"] = _compute_deadline_text_for_week(worksheet)
    _apply_marker_frame(
        workbook,
        worksheet,
        facility_name=str(facility.get("facility_name") or ""),
        facility_id=str(facility.get("facility_id") or ""),
        fax_template_id=fax_template_id,
        family_label=family_label,
        week_label=week_sheet_name,
        mode_label="completed",
    )
    _set_facility_name_box(worksheet, facility_name=str(facility.get("facility_name") or ""))
    output_path = output_dir / f"3_completed_order_form_{facility['facility_id']}_{week_sheet_name}.xlsx"
    workbook.save(output_path)
    return output_path


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Generate a three-layer order-form example: baseline, facility template, completed order form.",
    )
    parser.add_argument("--facility-id", default=DEFAULT_FACILITY_ID)
    parser.add_argument("--week", default=DEFAULT_WEEK)
    parser.add_argument("--output-dir", default=None)
    args = parser.parse_args()

    output_dir = _build_output_dir(args.output_dir)
    facility, spec = _resolve_demo_context(args.facility_id)
    fax_template_id = str(facility.get("fax_template_id") or "")
    source_workbook_name = str(spec["source_workbook"])
    family_label = str(spec["family_label"])

    baseline_path = _build_baseline_template(
        source_workbook_name=source_workbook_name,
        week_sheet_name=args.week,
        fax_template_id=fax_template_id,
        family_label=family_label,
        output_dir=output_dir,
    )
    facility_template_path = _build_facility_template(
        facility=facility,
        source_workbook_name=source_workbook_name,
        week_sheet_name=args.week,
        fax_template_id=fax_template_id,
        family_label=family_label,
        output_dir=output_dir,
    )
    completed_path = _build_completed_order_form(
        facility=facility,
        source_workbook_name=source_workbook_name,
        week_sheet_name=args.week,
        fax_template_id=fax_template_id,
        family_label=family_label,
        output_dir=output_dir,
    )

    print(f"output_dir={output_dir}")
    print(f"baseline={baseline_path}")
    print(f"facility_template={facility_template_path}")
    print(f"completed={completed_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
