#!/usr/bin/env python3

from __future__ import annotations

import argparse
from datetime import date, datetime
import json
from pathlib import Path
import re
import sys
import unicodedata
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import range_boundaries


ROOT = Path(__file__).resolve().parents[1]
OCR_PIPELINE_ROOT = ROOT.parent / "ocr_pipeline"
for candidate in (str(ROOT), str(OCR_PIPELINE_ROOT)):
    if candidate not in sys.path:
        sys.path.insert(0, candidate)

from scripts import compare_structure_guided_ocr as compare
from src.services import config_service, order_form_service
from src.services.workbook_pdf_renderer import render_workbook_path_to_pdf


_DIGIT_RE = re.compile(r"^\d{1,2}$")
_DEFAULT_HEADER_ROWS = (7, 8, 9)
_DEFAULT_BODY_START_ROW = 11
_DEFAULT_BODY_END_ROW = 67


def _write_json(path: Path, payload: Any) -> None:
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def _normalize_text(value: object) -> str:
    if isinstance(value, datetime):
        return f"{value.month}/{value.day}"
    if isinstance(value, date):
        return f"{value.month}/{value.day}"
    return unicodedata.normalize("NFKC", str(value or "")).strip()


def _normalize_menu_key(value: object) -> str:
    text = _normalize_text(value)
    text = re.sub(r"\s+", "", text)
    text = text.replace("（", "(").replace("）", ")")
    text = text.replace("①", "1").replace("②", "2")
    return text


def _normalize_digit_text(value: object) -> str:
    text = _normalize_text(value)
    text = re.sub(r"\s+", "", text)
    return text


def _is_numeric_text(value: object) -> bool:
    return bool(_DIGIT_RE.fullmatch(_normalize_digit_text(value)))


def _load_source_context(
    *,
    facility_id: str,
    week_sheet_name: str,
) -> tuple[Path, Any]:
    facility = config_service.get_facility_config(facility_id)
    if not facility:
        raise ValueError("facility not found")
    fax_template_id = str(order_form_service._infer_fax_template_id_from_facility(facility) or "").strip()  # noqa: SLF001
    if not fax_template_id:
        raise ValueError("facility fax_template_id not found")
    source_workbook_name = order_form_service._resolve_source_workbook_name_for_week_sheet(  # noqa: SLF001
        fax_template_id,
        week_sheet_name,
    )
    source_workbook_path = order_form_service._resolve_source_workbook_path(source_workbook_name)  # noqa: SLF001
    workbook = load_workbook(source_workbook_path, data_only=True)
    if week_sheet_name not in workbook.sheetnames:
        raise ValueError(f"week sheet not found in source workbook: {week_sheet_name}")
    return source_workbook_path, workbook[week_sheet_name]


def _is_weekday_only(value: object) -> bool:
    text = _normalize_text(value)
    if not text:
        return False
    return bool(re.fullmatch(r"[()（）月火水木金土日]+", text))


def _merged_value_map(worksheet) -> dict[tuple[int, int], object]:
    merged: dict[tuple[int, int], object] = {}
    for merged_range in worksheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
        anchor_value = worksheet.cell(min_row, min_col).value
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                merged[(row, col)] = anchor_value
    return merged


def _effective_cell_value(
    worksheet,
    merged_values: dict[tuple[int, int], object],
    *,
    row: int,
    col: int,
) -> object:
    value = worksheet.cell(row=row, column=col).value
    if value is not None:
        return value
    return merged_values.get((row, col))


def _header_values_for_column(
    worksheet,
    merged_values: dict[tuple[int, int], object],
    *,
    col: int,
) -> list[str]:
    values: list[str] = []
    for row in _DEFAULT_HEADER_ROWS:
        text = _normalize_text(_effective_cell_value(worksheet, merged_values, row=row, col=col))
        values.append(text)
    return values


def _body_samples_for_column(
    worksheet,
    *,
    col: int,
    limit: int = 8,
) -> list[str]:
    samples: list[str] = []
    for row in range(_DEFAULT_BODY_START_ROW, _DEFAULT_BODY_END_ROW + 1):
        text = _normalize_text(worksheet.cell(row=row, column=col).value)
        if not text:
            continue
        samples.append(text)
        if len(samples) >= limit:
            break
    return samples


def _infer_column_slot(
    *,
    raw_col_index: int,
    header_values: list[str],
    body_samples: list[str],
) -> dict[str, Any]:
    header_blob = " ".join(value for value in header_values if value).replace(" ", "")
    sample_blob = " ".join(body_samples)

    role = "unknown"
    slot_name = f"col_{raw_col_index}"
    label = header_blob or slot_name

    if "日付" in header_blob:
        role = "date"
        slot_name = "date"
        label = "日付"
    elif "区分" in header_blob:
        normalized_samples = {sample for sample in body_samples if sample}
        if normalized_samples and normalized_samples <= {"朝", "昼", "夕", "タ", "粉", "塩", "師"}:
            role = "daypart"
            slot_name = "daypart"
            label = "区分"
        else:
            role = "aux"
            slot_name = "aux"
            label = "補助区分"
    elif "献立" in header_blob:
        role = "menu_name"
        slot_name = "menu_name"
        label = "献立"
    elif "肉禁" in header_blob:
        role = "quantity"
        slot_name = "qty.no_meat_x"
        label = "肉禁"
    elif "魚禁" in header_blob:
        role = "quantity"
        slot_name = "qty.no_fish_x"
        label = "魚禁"
    elif "常食" in header_blob:
        role = "quantity"
        slot_name = "qty.regular_x"
        label = "常食"
    elif "変更1" in header_blob or "変更①" in header_blob:
        role = "quantity"
        slot_name = "qty.change_1_x"
        label = "変更1"
    elif "変更2" in header_blob or "変更②" in header_blob:
        role = "quantity"
        slot_name = "qty.change_2_x"
        label = "変更2"
    elif "備考" in header_blob:
        role = "note"
        slot_name = "note"
        label = "備考欄"
    elif not header_blob and not sample_blob:
        role = "spacer"
        slot_name = "spacer"
        label = "spacer"

    return {
        "raw_col_index": raw_col_index,
        "worksheet_col_index": raw_col_index + 1,
        "role": role,
        "slot_name": slot_name,
        "label": label,
        "header_values": header_values,
        "body_samples": body_samples,
    }


def _column_slots(
    worksheet,
    *,
    structure_col_count: int,
) -> list[dict[str, Any]]:
    merged_values = _merged_value_map(worksheet)
    slots: list[dict[str, Any]] = []
    for raw_col_index in range(structure_col_count):
        worksheet_col_index = raw_col_index + 1
        slots.append(
            _infer_column_slot(
                raw_col_index=raw_col_index,
                header_values=_header_values_for_column(
                    worksheet,
                    merged_values,
                    col=worksheet_col_index,
                ),
                body_samples=_body_samples_for_column(worksheet, col=worksheet_col_index),
            )
        )
    return slots


def _canonical_workbook_rows(worksheet) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    last_date = ""
    for worksheet_row in range(_DEFAULT_BODY_START_ROW, _DEFAULT_BODY_END_ROW + 1):
        date_text = _normalize_text(worksheet.cell(row=worksheet_row, column=1).value)
        if date_text and not _is_weekday_only(date_text):
            last_date = date_text
        effective_date = last_date
        daypart = _normalize_text(worksheet.cell(row=worksheet_row, column=2).value)
        aux = _normalize_text(worksheet.cell(row=worksheet_row, column=3).value)
        menu_name = _normalize_text(worksheet.cell(row=worksheet_row, column=4).value)
        rows.append(
            {
                "worksheet_row": worksheet_row,
                "date": date_text,
                "effective_date": effective_date,
                "daypart": daypart,
                "aux": aux,
                "menu_name": menu_name,
                "menu_key": _normalize_menu_key(menu_name),
            }
        )
    return rows


def _canonical_table_rows(canonical_table: dict[str, Any]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for row_index, row in enumerate(canonical_table.get("rows") or []):
        if not isinstance(row, list):
            continue
        date_text = _normalize_text(row[0] if len(row) > 0 else "")
        daypart = _normalize_text(row[1] if len(row) > 1 else "")
        aux = _normalize_text(row[2] if len(row) > 2 else "")
        menu_name = _normalize_text(row[3] if len(row) > 3 else "")
        rows.append(
            {
                "row_index": row_index,
                "date": date_text,
                "daypart": daypart,
                "aux": aux,
                "menu_name": menu_name,
                "menu_key": _normalize_menu_key(menu_name),
            }
        )
    return rows


def _map_canonical_rows_to_workbook(
    *,
    canonical_table: dict[str, Any],
    worksheet,
) -> dict[int, dict[str, Any]]:
    canonical_rows = _canonical_table_rows(canonical_table)
    workbook_rows = _canonical_workbook_rows(worksheet)
    workbook_pointer = 0
    row_map: dict[int, dict[str, Any]] = {}
    for canonical in canonical_rows:
        if int(canonical.get("row_index") or 0) < 2:
            continue
        menu_key = str(canonical.get("menu_key") or "").strip()
        if not menu_key or menu_key == "献立":
            continue
        while workbook_pointer < len(workbook_rows):
            workbook_entry = workbook_rows[workbook_pointer]
            workbook_key = str(workbook_entry.get("menu_key") or "").strip()
            if workbook_key == menu_key:
                row_map[int(canonical["row_index"])] = {
                    **workbook_entry,
                    "row_index": int(canonical["row_index"]),
                }
                workbook_pointer += 1
                break
            workbook_pointer += 1
    return row_map


def _fill_row_map_from_workbook_positions(
    *,
    row_map: dict[int, dict[str, Any]],
    worksheet,
    row_count: int,
) -> dict[int, dict[str, Any]]:
    workbook_rows = _canonical_workbook_rows(worksheet)
    if not workbook_rows:
        return row_map
    header_rows = min(row_map) if row_map else 2
    row_offset = _DEFAULT_BODY_START_ROW - header_rows
    filled = dict(row_map)
    for row_index in range(header_rows, row_count):
        if row_index in filled:
            continue
        worksheet_row = row_index + row_offset
        if worksheet_row < _DEFAULT_BODY_START_ROW or worksheet_row > _DEFAULT_BODY_END_ROW:
            continue
        workbook_index = worksheet_row - _DEFAULT_BODY_START_ROW
        if workbook_index < 0 or workbook_index >= len(workbook_rows):
            continue
        filled[row_index] = {
            **workbook_rows[workbook_index],
            "row_index": row_index,
        }
    return filled


def _anchor_row_map(row_map: dict[int, dict[str, Any]], *, row_count: int) -> dict[int, dict[str, Any] | None]:
    anchor_map: dict[int, dict[str, Any] | None] = {}
    current_anchor: dict[str, Any] | None = None
    for row_index in range(row_count):
        current = row_map.get(row_index)
        if current and str(current.get("daypart") or "").strip():
            current_anchor = current
        anchor_map[row_index] = current_anchor
    return anchor_map


def _prepare_structure_template(
    *,
    facility_id: str,
    week_sheet_name: str,
    output_dir: Path,
    dpi: int,
    device: str,
) -> dict[str, Any]:
    structure_xlsx = order_form_service.build_fax_structure_only_excel(
        facility_id=facility_id,
        week_sheet_name=week_sheet_name,
        output_dir=output_dir,
    )
    structure_pdf_path = output_dir / f"{structure_xlsx.stem}.pdf"
    render_workbook_path_to_pdf(
        structure_xlsx,
        output_path=structure_pdf_path,
        sheet_name=week_sheet_name,
        dpi=dpi,
    )
    structure_pdf_bytes = structure_pdf_path.read_bytes()
    structure_result = compare._run_structure_template(  # noqa: SLF001
        structure_pdf_bytes=structure_pdf_bytes,
        dpi=dpi,
        device=device,
    )
    structure_table = structure_result["primary_table"]
    if not isinstance(structure_table, dict):
        raise RuntimeError("structure-only template PDF did not produce a primary table")

    template_png_bytes = compare.render_pdf_to_png_bytes(structure_pdf_bytes, dpi=dpi, page=1)
    template_png_path = output_dir / "structure_template_page1.png"
    template_png_path.write_bytes(template_png_bytes)
    template_width, template_height = compare._decode_png_size(template_png_bytes)  # noqa: SLF001
    template_bgr = compare._decode_png_bgr(template_png_bytes)  # noqa: SLF001
    return {
        "structure_xlsx": structure_xlsx,
        "structure_pdf": structure_pdf_path,
        "structure_table": structure_table,
        "template_png_path": template_png_path,
        "template_width": template_width,
        "template_height": template_height,
        "template_bgr": template_bgr,
    }


def _structure_guided_table_for_pdf_page(
    *,
    pdf_bytes: bytes,
    page_index: int,
    dpi: int,
    device: str,
    structure_table: dict[str, Any],
    template_png_path: Path,
    template_width: int,
    template_height: int,
    template_bgr,
    apply_correction: bool,
    output_dir: Path,
    prefix: str,
) -> dict[str, Any]:
    correction_summary: dict[str, Any] = {"applied": False}
    corrected_pdf_bytes = pdf_bytes
    corrected_pages = None
    if apply_correction:
        try:
            corrected_pdf_bytes, correction_summary, corrected_pages = compare.correct_pdf_for_yomitoku(
                pdf_bytes=pdf_bytes,
                dpi=dpi,
                db=None,
            )
        except Exception as exc:  # noqa: BLE001
            correction_summary = {"applied": False, "error": str(exc)}
            corrected_pdf_bytes = pdf_bytes
            corrected_pages = None

    corrected_page = compare._normalize_correction_page(  # noqa: SLF001
        corrected_pages,
        pdf_bytes=corrected_pdf_bytes,
        page_index=page_index,
        dpi=dpi,
    )

    default_result = compare._run_default_table(  # noqa: SLF001
        page_image=corrected_page,
        page_index=page_index,
        dpi=dpi,
        device=device,
    )
    default_table = default_result["primary_table"]
    if not isinstance(default_table, dict):
        raise RuntimeError(f"default yomitoku path did not produce a primary table: {prefix}")

    match_bgr, ocr_bgr, _ocr_keep_lines_bgr = compare.build_images_for_match_and_ocr_from_bgr(corrected_page)
    alignment_method, warped_match_bgr, warped_ocr_bgr, matched_template = compare._align_to_structure_template(  # noqa: SLF001
        source_table_bbox=default_table.get("bbox") if isinstance(default_table.get("bbox"), list) else None,
        template_table_bbox=structure_table.get("bbox") if isinstance(structure_table.get("bbox"), list) else None,
        template_bgr=template_bgr,
        template_png_path=template_png_path,
        template_width=template_width,
        template_height=template_height,
        match_bgr=match_bgr,
        ocr_bgr=ocr_bgr,
    )
    if not alignment_method or not isinstance(matched_template, dict):
        raise RuntimeError(f"template alignment failed: {prefix}")

    warped_match_path = output_dir / f"{prefix}_warped_match.png"
    warped_ocr_path = output_dir / f"{prefix}_warped_ocr.png"
    compare.cv2.imwrite(str(warped_match_path), warped_match_bgr)
    compare.cv2.imwrite(str(warped_ocr_path), warped_ocr_bgr)

    words = compare.ocr_image_words(warped_ocr_bgr, device=device)
    guided_table = compare.assign_words_to_structure_table(
        structure_table=structure_table,
        words=words,
    )
    return {
        "alignment_method": alignment_method,
        "page_correction": correction_summary,
        "default_table": default_table,
        "guided_table": guided_table,
        "warped_match_png": str(warped_match_path),
        "warped_ocr_png": str(warped_ocr_path),
    }


def _collect_slot_hits(
    *,
    guided_table: dict[str, Any],
    column_slots: list[dict[str, Any]],
    row_map: dict[int, dict[str, Any]],
) -> dict[str, Any]:
    row_count = int(guided_table.get("row_count") or 0)
    anchor_map = _anchor_row_map(row_map, row_count=row_count)
    all_numeric_hits: list[dict[str, Any]] = []
    quantity_slot_hits: list[dict[str, Any]] = []

    for cell in guided_table.get("cells") or []:
        if not isinstance(cell, dict):
            continue
        text = _normalize_text(cell.get("text"))
        if not text:
            continue
        row_index = int(cell.get("row_index") or 0)
        if row_index < 2:
            continue
        col_index = int(cell.get("col_index") or 0)
        slot = column_slots[col_index] if 0 <= col_index < len(column_slots) else None
        normalized_text = _normalize_digit_text(text)
        if _is_numeric_text(text):
            all_numeric_hits.append(
                {
                    "row_index": row_index,
                    "col_index": col_index,
                    "text": text,
                    "slot": slot,
                }
            )
        if not isinstance(slot, dict) or slot.get("role") != "quantity":
            continue
        if not text:
            continue
        anchor_row = anchor_map.get(row_index)
        quantity_slot_hits.append(
            {
                "row_index": row_index,
                "col_index": col_index,
                "text": text,
                "normalized_text": normalized_text,
                "is_numeric": _is_numeric_text(text),
                "slot": slot,
                "canonical_anchor_row": anchor_row,
            }
        )

    return {
        "all_numeric_hits": all_numeric_hits,
        "quantity_slot_hits": quantity_slot_hits,
    }


def _business_preview(quantity_slot_hits: list[dict[str, Any]]) -> list[dict[str, Any]]:
    preview: list[dict[str, Any]] = []
    for hit in quantity_slot_hits:
        anchor = hit.get("canonical_anchor_row") or {}
        preview.append(
            {
                "slot": (hit.get("slot") or {}).get("slot_name"),
                "label": (hit.get("slot") or {}).get("label"),
                "row_index": hit.get("row_index"),
                "anchor_row_index": (anchor or {}).get("row_index"),
                "date": (anchor or {}).get("effective_date") or (anchor or {}).get("date"),
                "daypart": (anchor or {}).get("daypart"),
                "aux": (anchor or {}).get("aux"),
                "menu_name": (anchor or {}).get("menu_name"),
                "text": hit.get("text"),
                "is_numeric": hit.get("is_numeric"),
            }
        )
    return preview


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Try canonical structure-slot assignment on an actual fax PDF and emit semantic quantity hits.",
    )
    parser.add_argument("--facility-id", required=True)
    parser.add_argument("--fax-pdf", required=True)
    parser.add_argument("--week", default=order_form_service._DEFAULT_WEEK_SHEET)  # noqa: SLF001
    parser.add_argument("--page-index", type=int, required=True)
    parser.add_argument("--dpi", type=int, default=200)
    parser.add_argument("--device", default="cpu")
    parser.add_argument("--output-dir", default=None)
    args = parser.parse_args()

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_dir = Path(args.output_dir or (ROOT / "tmp" / f"structure_slot_hits_{stamp}"))
    output_dir.mkdir(parents=True, exist_ok=True)

    source_workbook_path, worksheet = _load_source_context(
        facility_id=args.facility_id,
        week_sheet_name=args.week,
    )
    structure_template = _prepare_structure_template(
        facility_id=args.facility_id,
        week_sheet_name=args.week,
        output_dir=output_dir,
        dpi=args.dpi,
        device=args.device,
    )

    canonical_xlsx = order_form_service.build_fax_order_form_excel(
        facility_id=args.facility_id,
        week_sheet_name=args.week,
        output_dir=output_dir,
    )
    canonical_pdf = output_dir / f"{canonical_xlsx.stem}.pdf"
    render_workbook_path_to_pdf(
        canonical_xlsx,
        output_path=canonical_pdf,
        sheet_name=args.week,
        dpi=args.dpi,
    )

    canonical_result = _structure_guided_table_for_pdf_page(
        pdf_bytes=canonical_pdf.read_bytes(),
        page_index=1,
        dpi=args.dpi,
        device=args.device,
        structure_table=structure_template["structure_table"],
        template_png_path=structure_template["template_png_path"],
        template_width=structure_template["template_width"],
        template_height=structure_template["template_height"],
        template_bgr=structure_template["template_bgr"],
        apply_correction=False,
        output_dir=output_dir,
        prefix="canonical",
    )

    actual_result = _structure_guided_table_for_pdf_page(
        pdf_bytes=Path(args.fax_pdf).read_bytes(),
        page_index=args.page_index,
        dpi=args.dpi,
        device=args.device,
        structure_table=structure_template["structure_table"],
        template_png_path=structure_template["template_png_path"],
        template_width=structure_template["template_width"],
        template_height=structure_template["template_height"],
        template_bgr=structure_template["template_bgr"],
        apply_correction=True,
        output_dir=output_dir,
        prefix="actual",
    )

    canonical_row_map = _map_canonical_rows_to_workbook(
        canonical_table=canonical_result["guided_table"],
        worksheet=worksheet,
    )
    canonical_row_map = _fill_row_map_from_workbook_positions(
        row_map=canonical_row_map,
        worksheet=worksheet,
        row_count=int(actual_result["guided_table"].get("row_count") or 0),
    )
    column_slots = _column_slots(
        worksheet,
        structure_col_count=int(actual_result["guided_table"].get("col_count") or 0),
    )
    collected = _collect_slot_hits(
        guided_table=actual_result["guided_table"],
        column_slots=column_slots,
        row_map=canonical_row_map,
    )

    summary = {
        "facility_id": args.facility_id,
        "fax_pdf": str(Path(args.fax_pdf)),
        "week": args.week,
        "page_index": args.page_index,
        "dpi": args.dpi,
        "device": args.device,
        "source_workbook": str(source_workbook_path),
        "artifacts": {
            "structure_xlsx": str(structure_template["structure_xlsx"]),
            "structure_pdf": str(structure_template["structure_pdf"]),
            "canonical_xlsx": str(canonical_xlsx),
            "canonical_pdf": str(canonical_pdf),
            "template_png": str(structure_template["template_png_path"]),
            "canonical_warped_ocr_png": canonical_result["warped_ocr_png"],
            "actual_warped_ocr_png": actual_result["warped_ocr_png"],
            "actual_warped_match_png": actual_result["warped_match_png"],
        },
        "column_slots": column_slots,
        "actual_alignment_method": actual_result["alignment_method"],
        "canonical_alignment_method": canonical_result["alignment_method"],
        "actual_page_correction": actual_result["page_correction"],
        "all_numeric_hits": collected["all_numeric_hits"],
        "quantity_slot_hits": collected["quantity_slot_hits"],
        "business_preview": _business_preview(collected["quantity_slot_hits"]),
    }

    _write_json(output_dir / "summary.json", summary)
    _write_json(output_dir / "actual_guided_table.json", actual_result["guided_table"])
    _write_json(output_dir / "canonical_guided_table.json", canonical_result["guided_table"])
    _write_json(output_dir / "canonical_row_map.json", canonical_row_map)

    print(f"output_dir={output_dir}")
    print(f"summary={output_dir / 'summary.json'}")
    print(f"all_numeric_hits={len(summary['all_numeric_hits'])}")
    print(f"quantity_slot_hits={len(summary['quantity_slot_hits'])}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
